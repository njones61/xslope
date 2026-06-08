"""Generate sample DXF cross-section files from the polygon-sheet xlsx inputs.

This produces realistic CAD test fixtures for the planned DXF import utility
(see plans/plan_polygons.md, section 7). Each material zone is written as a
*closed* LWPOLYLINE, and polygons are organized onto DXF layers named after
their material -- exactly the convention the importer expects ("all polylines
on the CLAY layer are clay", etc.). This mirrors how commercial packages ask
users to organize a drawing before export.

For every xslope_*_poly.xlsx in this folder it reads the `polygon` and `mat`
sheets directly (load_slope_data does not parse the polygon sheet yet) and
writes a matching .dxf next to it. After writing, it round-trips each file
back through ezdxf as a smoke test for the future reader.

Run from the repo root:  python3 poly_test/_build_dxf_from_poly.py
"""
import glob
import os
import re

import ezdxf
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))

# A small palette of AutoCAD Color Index values so zones are visually distinct
# when the DXF is opened in CAD. Cycled by material id.
_ACI_PALETTE = [40, 30, 8, 5, 3, 1, 6, 2, 4, 7]


def _layer_name(material_name, mat_id):
    """Turn a material name into a valid, CAD-conventional DXF layer name.

    Uppercase, spaces/illegal chars -> underscores (e.g. 'Silty Clay' ->
    'SILTY_CLAY'). Falls back to MAT_<id> when the material has no name.
    DXF disallows: < > / \\ " : ; ? * | = ` and control chars.
    """
    name = (material_name or "").strip()
    if not name:
        return f"MAT_{mat_id}"
    name = name.upper()
    name = re.sub(r'[<>/\\":;?*|=`\s]+', "_", name)
    name = name.strip("_")
    return name or f"MAT_{mat_id}"


def _read_materials(wb):
    """mat sheet: row 8 headers, rows 9+ = materials (col1=id, col2=name)."""
    ms = wb["mat"]
    names = {}
    for r in range(9, ms.max_row + 1):
        mat_id = ms.cell(row=r, column=1).value
        if mat_id is None:
            continue
        names[int(mat_id)] = ms.cell(row=r, column=2).value
    return names


def _read_polygons(wb):
    """polygon sheet: block p has x_col=1+3*(p-1), y_col=x_col+1.

    Row 5 (y_col) = Mat ID; coords from row 8 down until a blank row.
    Returns list of dicts {'mat_id': int, 'coords': [(x, y), ...]}.
    """
    ps = wb["polygon"]
    polys = []
    p = 1
    while True:
        x_col = 1 + 3 * (p - 1)
        y_col = x_col + 1
        if x_col > ps.max_column:
            break
        mat_id = ps.cell(row=5, column=y_col).value
        if mat_id in (None, ""):
            # No more populated blocks.
            break
        coords = []
        r = 8
        while True:
            x = ps.cell(row=r, column=x_col).value
            y = ps.cell(row=r, column=y_col).value
            if x in (None, "") or y in (None, ""):
                break
            coords.append((float(x), float(y)))
            r += 1
        if coords:
            polys.append({"mat_id": int(mat_id), "coords": coords})
        p += 1
    return polys


def _dedupe_closing_vertex(coords):
    """Drop a trailing vertex that repeats the first -- a closed LWPOLYLINE
    closes implicitly, so an explicit duplicate would add a zero-length edge."""
    if len(coords) >= 2 and coords[0] == coords[-1]:
        return coords[:-1]
    return coords


def build_dxf(xlsx_path, dxf_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    mat_names = _read_materials(wb)
    polys = _read_polygons(wb)

    doc = ezdxf.new("R2010")
    msp = doc.modelspace()

    # Create a layer per material actually used.
    used_mat_ids = sorted({p["mat_id"] for p in polys})
    layer_for = {}
    for mid in used_mat_ids:
        lname = _layer_name(mat_names.get(mid), mid)
        layer_for[mid] = lname
        if lname not in doc.layers:
            color = _ACI_PALETTE[(mid - 1) % len(_ACI_PALETTE)]
            doc.layers.add(name=lname, color=color)

    for poly in polys:
        coords = _dedupe_closing_vertex(poly["coords"])
        msp.add_lwpolyline(
            coords,
            close=True,
            dxfattribs={"layer": layer_for[poly["mat_id"]]},
        )

    doc.saveas(dxf_path)
    return polys, mat_names, layer_for


def verify_dxf(dxf_path):
    """Round-trip: read the file back the way the importer will, return a
    summary of closed LWPOLYLINEs grouped by layer."""
    doc = ezdxf.readfile(dxf_path)
    summary = []
    for e in doc.modelspace().query("LWPOLYLINE"):
        pts = [(round(p[0], 3), round(p[1], 3)) for p in e.get_points("xy")]
        summary.append((e.dxf.layer, e.closed, len(pts)))
    return summary


if __name__ == "__main__":
    xlsx_files = sorted(glob.glob(os.path.join(HERE, "xslope_*_poly.xlsx")))
    for xlsx_path in xlsx_files:
        dxf_path = os.path.splitext(xlsx_path)[0] + ".dxf"
        polys, mat_names, layer_for = build_dxf(xlsx_path, dxf_path)
        summary = verify_dxf(dxf_path)

        base = os.path.basename(dxf_path)
        layers = sorted(set(layer_for.values()))
        print(f"{base}")
        print(f"   {len(polys)} polygons, {len(layers)} layers: {', '.join(layers)}")
        # Confirm every polyline read back as closed.
        n_closed = sum(1 for (_, closed, _) in summary if closed)
        print(f"   round-trip: {n_closed}/{len(summary)} LWPOLYLINEs closed OK")
