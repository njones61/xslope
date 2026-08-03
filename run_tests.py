#!/usr/bin/env python3
"""
Regression test suite for xslope.

Scans docs/{lem,fem,seep}/samples.md and docs/seep/seep_slope.md for test
tags of the form:

    <!-- test: file=files/foo.xlsx, type=circular_search, method=spencer, expected_fs=1.234, num_slices=30 -->
    <!-- test: file=files/foo.xlsx, type=fem_ssrm, expected_fs=1.38, element_type=quad8, target_size=3.5, tolerance=0.025 -->
    <!-- test: file=files/foo.xlsx, type=seep, expected_flowrate=40.062, tolerance=0.05 -->
    <!-- test: file=files/foo.xlsx, type=seep, expected_flowrate=28.6, element_type=tri6, target_size=2.0, tolerance=0.01 -->
    <!-- test: file=files/foo.xlsx, type=seep_head, points=2:2:4.05;4:2:4.15, tolerance=0.02 -->
    <!-- test: file=files/foo.xlsx, type=tseep_head, time=600, points=0:0:5.0;30:0:2.4, tolerance=0.05 -->
    <!-- test: file=files/foo.xlsx, type=seep_elements, expected_flowrate=40.062, target_size=1.5, tolerance=0.05 -->
    <!-- test: file=files/foo.xlsx, type=fem_elements, expected_fs=1.36, target_size=3.5, tolerance=0.04, f_min=1.0, f_max=1.8, max_iter=4000, benchmark=SSRM-elements -->
    <!-- test: file=files/foo.xlsx, type=mesh_elements, element_type=tri6, target_size=6.5, expected_elements=3166, expected_nodes=6555, benchmark=RS2-4-mesh -->

The mesh_elements type locks a published MESH SIZE — the element and node count
one model meshes to at one target size. It builds the mesh and counts it, and
solves nothing, so it costs seconds where the strength-reduction row on the same
file costs minutes. It exists because several verification rows quote XSLOPE's
own discretization beside the vendor's to say which comparison is the finer, and
nothing else checks those numbers: the SSRM locks beside them are tolerant to a
percent or two of mesh drift, so a mesher change moves every printed count and
leaves every page asserting a stale one.

The seep_elements / fem_elements types solve ONE problem with every supported
element type (seep: tri3/tri6/quad4/quad8/quad9; FEM: the quadratic tri6/quad8/
quad9) and check each converges and matches the expected flowrate/FS within
tolerance — coverage that a solver/mesh change hasn't broken an element type.
element_types= overrides the default set.

An optional benchmark=<ID> key (e.g. benchmark=SSRM-2) marks a verification
benchmark; these can be excluded with --skip-benchmarks for quicker routine
runs.

Runs each test and compares results against expected values.
As new samples are added with test tags, they automatically become part of the suite.

Usage:
    python run_tests.py              # run all tests
    python run_tests.py --lem        # run only LEM tests
    python run_tests.py --fem        # run only FEM tests
    python run_tests.py --seep       # run only (steady) seepage tests
    python run_tests.py --tseep      # run only transient-seepage tests (type=tseep_head)
    python run_tests.py --roundtrip  # run only the Excel save/load round-trip tests
    python run_tests.py --tolerance 0.02  # custom FS tolerance (default 0.01)
    python run_tests.py --skip-benchmarks # exclude verification benchmarks (faster)
    python run_tests.py --verbose    # print details for passing tests too
    python run_tests.py --list       # enumerate the discovered tests, run nothing

The suite may be invoked from any working directory: every path it reads is
resolved against this file's own location (see ``REPO_ROOT`` / ``_repo``), and
the ``cwd_invariant`` row checks that discovery is the same either way.
"""

import argparse
import glob
import os
import re
import sys
import time
from pathlib import Path

import matplotlib
matplotlib.use('Agg')  # non-interactive backend — no plot windows


# === Repo-relative paths are anchored to THIS FILE, never to the process CWD ===
# Every fixture, corpus file and docs page this suite reads lives at a fixed place
# inside the repo, so its location is a property of the script, not of wherever the
# script happens to be invoked from. Resolving them against the CWD made discovery
# fail SILENTLY from any other directory: the `.exists()` guards around the docs
# scan simply found nothing and the run reported a smaller, still-green suite.
# `--list` prints the discovered count so that invariance is checkable, and the
# `cwd_invariant` test checks it.
REPO_ROOT = Path(__file__).resolve().parent


def _repo(rel):
    """Absolute path to a repo-relative location, anchored to this script."""
    return str(REPO_ROOT / rel)


def _private_dir():
    """The private test repo, or None when this clone does not have one.

    $XSLOPE_PRIVATE_TESTS wins; otherwise the repo's own SIBLING directories are
    searched, which is what the name means — a checkout beside this one, not a
    directory beside wherever the suite was invoked from. Absent on public clones,
    where the private rows are simply not collected.
    """
    d = os.environ.get('XSLOPE_PRIVATE_TESTS')
    if d:
        return d
    # 'xslope_private_tests' is the pre-rename name, still used by old clones.
    for name in ('xslope_private', 'xslope_private_tests'):
        cand = REPO_ROOT.parent / name
        if cand.is_dir():
            return str(cand)
    return None



# === Excel round-trip regression (save_slope_data_to_xlsx) ===
# load -> save into a blank template -> reload must reproduce every input
# category. The file set spans circular & non-circular surfaces, profile &
# polygon geometry, reinforcement, piles, distributed loads (both sets), a
# second piezo line, reliability sigmas, and seepage BCs (both sets).
ROUNDTRIP_TEMPLATE = _repo('docs/inputs/input_template.xlsx')   # editable master (docs link)
BUNDLED_TEMPLATE = _repo('xslope/resources/input_template.xlsx')  # copy shipped in the wheel
# The /xslope skill body: docs master + the copy shipped in the wheel (used by the
# Studio assistant on pip installs, where docs/ is absent). Must stay byte-identical.
SKILL_MASTER = _repo('docs/usage/claude/xslope.md')
BUNDLED_SKILL = _repo('xslope/resources/xslope_skill.md')
# M-P with f(x)==1 must reproduce Spencer exactly, on both slope facings (the S3b gate).
AXIAL_MIRROR_LEFT = _repo('docs/inputs/slope/xslope_nail_axial.xlsx')
AXIAL_MIRROR_RIGHT = _repo('docs/inputs/slope/xslope_nail_axial_rface.xlsx')
MP_SPENCER_LEFT = _repo('docs/inputs/slope/xslope_simple1.xlsx')
MP_SPENCER_RIGHT = _repo('docs/inputs/slope/xslope_rface.xlsx')
# Stage-1 FS 1.91 unweakened; halving c/phi/d/psi drops it below 1, which must be refused.
DRAWDOWN_GUARD_FILE = _repo('docs/inputs/slope/xslope_rapid.xlsx')
# --- backward-compatibility fixtures: SYNTHESIZED, never sample files ------
#
# This guard's whole premise is that its fixtures are OLD: the check that matters
# is that saving a pre-v19 model into the current template does not let the
# template's own pre-filled cells (D17='YES', D22='rollers', D23='auto') leak in,
# and that only means something when the model really came from a file that had
# no such cell. Pointing it at the sample files made that premise a hostage to
# their format: the moment the download set moved to v22, the guard would have
# been silently testing v22 -> v22 and proving nothing. (The same erosion had
# already happened once, unnoticed, to the v16 fixture below.)
#
# So the fixtures are BUILT, from the archived templates in docs/inputs. Each
# entry names a real model and the template version to express it in; the file is
# written into a temp dir at test time through the archived vNN template, which is
# the writer's own version-faithful path. The archived templates are the
# historical record and are never migrated, so the backward-compatibility
# guarantee is now anchored to them rather than to files that move.
LEGACY_TEMPLATE_FMT = _repo('docs/inputs/input_template_v{}.xlsx')
# (model, template version to express it in). The version spread is deliberate:
# v12 predates the v18 main-sheet shift AND the v16/v17 mat-column inserts, v13
# and v15 sit between them, v17 is the last version before the v18 shift, and v20
# is post-shift but pre-v21/v22 — so the set spans every layout the loader must
# still read. Content-wise it spans circular & non-circular surfaces, profile &
# polygon geometry, reinforcement, piles, distributed loads (both sets), a second
# piezo line, reliability sigmas, and seepage BCs (both sets).
ROUNDTRIP_FILES = [
    (_repo('docs/inputs/slope/xslope_simple1.xlsx'), 12),
    (_repo('docs/inputs/slope/xslope_dam.xlsx'), 12),
    (_repo('docs/inputs/slope/xslope_rapid.xlsx'), 12),
    (_repo('docs/inputs/slope/xslope_reliability.xlsx'), 12),
    (_repo('docs/inputs/slope/xslope_rface.xlsx'), 12),
    (_repo('docs/fem/files/xslope_reinforce_fem.xlsx'), 12),
    (_repo('docs/fem/files/xslope_piles_fem.xlsx'), 15),
    (_repo('docs/fem/files/xslope_noncircular_fem.xlsx'), 12),
    (_repo('docs/fem/files/xslope_griffiths1_load.xlsx'), 17),
    (_repo('docs/seep/files/xslope_earth_dam1.xlsx'), 12),
    (_repo('docs/inputs/seep/xslope_earth_dam_bc2.xlsx'), 12),
    (_repo('docs/seep/files/xslope_levee_poly.xlsx'), 13),
    (_repo('docs/inputs/seep/xslope_lost_lake.xlsx'), 12),
    # Rocscience GW#5: the only corpus model carrying a conductivity below 1e-10
    # (a 1e-13 m/s lens). It is here as the small-magnitude guard — the cell
    # writer used to round every float to ten DECIMAL places, which wrote that
    # lens as 0, and the nonzero-to-zero check in _roundtrip_eq catches it.
    (_repo('docs/verification/files/rocscience_gw/gw005.xlsx'), 20),
]
# Structured-DXF round-trip files (export_dxf -> read_dxf_layers -> default wizard
# mapping -> build_from_dxf_mapping). Each entry is (file, kind) where kind drives
# the per-feature assertion: 'profile', 'polygon', or 'reinforce'. Needs ezdxf +
# PySide6 (build_from_dxf_mapping lives in studio); skipped when either is absent.
DXF_FILES = [
    (_repo('docs/inputs/slope/xslope_simple1.xlsx'), 'profile'),
    (_repo('docs/inputs/slope/xslope_reinf.xlsx'), 'reinforce'),
    (_repo('docs/seep/files/xslope_levee_poly.xlsx'), 'polygon'),
]
# Source (non-derived) keys that must survive a round-trip. Derived geometry
# (ground_surface, domain_polygon, and polygons-from-profile) is recomputed by
# the loader and so is compared separately (polygon geometry only).
ROUNDTRIP_KEYS = [
    'gamma_water', 'tcrack_depth', 'tcrack_water', 'k_seismic', 'max_depth',
    'profile_lines', 'materials', 'piezo_line', 'piezo_line2', 'circles',
    'non_circ', 'dloads', 'dloads2', 'dload_dirs', 'dload2_dirs',
    'reinforcement_lines', 'pile_lines',
    'line_loads', 'seepage_bc', 'seepage_bc2',
    # v19 file-carried run options + the circles-sheet search window. On the
    # pre-v19 fixtures above (all but gw005) every one of these is None on both
    # sides, which is itself the check that matters there: saving a v18 model into
    # the v19 template must NOT let the template's own pre-filled D17='YES' leak in.
    'lem_method', 'num_slices', 'k0', 'tension_srf', 'element_type',
    'target_size', 'ssrm_f_min', 'ssrm_f_max', 'search_window',
    # v22 water-load mode. Unlike the options above it is never None: these
    # pre-v22 fixtures resolve to 'manual' on load, and the check that matters
    # here is that saving them into the v22 template — which ships D23 pre-filled
    # 'auto' — brings them back MANUAL. A leak in this cell would not merely lose
    # a setting; it would double the reservoir on every file that carries one.
    'water_loads',
    # v22 surface family (main D24). None on every corpus file, which is the check
    # that matters here: these decks define one family, nothing had to choose, and a
    # save must not write an answer into a cell whose blank is the normal state. The
    # SURVIVAL of a stated family is the surface_family_roundtrip row below.
    'surface_family',
    # v20 SSR zone overlays. Empty on every corpus file, which is the check that
    # matters here: saving a model that carries none must not invent one (the
    # template's own pre-labelled polygon blocks are the way that could happen).
    # The SURVIVAL of a populated set is the ssr_zone_roundtrip row below.
    'ssr_zones',
]


def _legacy_manual_water(sd):
    """A model expressed the way every template before v22 had to express it.

    Automatic mode states the reservoir once, as the water definition; before v22
    there was no cell to say so, so the same reservoir had to be typed onto the
    dloads sheet. This runs that conversion backwards — the engine's own derived
    blocks are materialized onto the sheets they used to occupy and the mode goes
    back to manual — which is what lets a current model be written into an
    archived template without losing its standing water.

    A manual-mode model is returned as-is: it is already in the old idiom.
    """
    from xslope.water import (with_water_loads, derived_blocks, water_loads_mode,
                              DERIVED_KEYS, DERIVED_META_KEY)
    if water_loads_mode(sd) != 'auto':
        return dict(sd)
    full = with_water_loads(sd)
    out = dict(full)
    for stage, key, dirs_key in ((1, 'dloads', 'dload_dirs'),
                                 (2, 'dloads2', 'dload2_dirs')):
        blocks = derived_blocks(full, stage)
        if not blocks:
            continue
        dirs = list(out.get(dirs_key) or [])
        while len(dirs) < len(out.get(key) or []):
            dirs.append('normal')
        out[key] = list(out.get(key) or []) + list(blocks)
        out[dirs_key] = dirs + ['normal'] * len(blocks)
    out['water_loads'] = 'manual'
    for k in list(DERIVED_KEYS.values()) + [DERIVED_META_KEY]:
        out.pop(k, None)
    return out


def provision_legacy_file(source, version, dest):
    """Write ``source``'s model into the archived version-``version`` template.

    The fixture builder for every backward-compatibility guard here. The writer is
    already version-faithful to its destination template (fileio's ``_dest_version``
    branches), so writing through ``input_template_vNN.xlsx`` produces a genuine
    vNN file — the newer cells simply do not exist to be written, which is exactly
    the condition the guards need. Returns ``dest``.
    """
    import shutil
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx
    template = LEGACY_TEMPLATE_FMT.format(version)
    if not os.path.exists(template):
        raise FileNotFoundError(f"archived template missing: {template}")
    save_slope_data_to_xlsx(_legacy_manual_water(load_slope_data(source)),
                            dest, template=template)
    # Companion files travel with the fixture. They are found by naming convention
    # beside the .xlsx, so a model whose pore pressures come from a solved field
    # would otherwise load in the temp dir as a model with no field at all.
    src_base = os.path.splitext(source)[0]
    dst_base = os.path.splitext(dest)[0]
    for suffix in ('_mesh.json', '_seep.csv', '_seep2.csv'):
        if os.path.exists(src_base + suffix):
            shutil.copy(src_base + suffix, dst_base + suffix)
    return dest


# --- v19 run-option round-trip ---
# The ROUNDTRIP_FILES fixtures are nearly all pre-v19, so they mostly prove the new
# fields stay absent. This synthetic case proves they SURVIVE: a real model is
# loaded, every v19 field is set to a distinct non-default value, saved through the
# current template, reloaded, and compared field by field. It is built in a temp
# dir rather than checked in so there is no corpus file to keep in sync.
V19_ROUNDTRIP_BASE = _repo('docs/inputs/slope/xslope_simple1.xlsx')
V19_ROUNDTRIP_VALUES = {
    'lem_method': 'mprice',
    'num_slices': 37,
    'k0': 0.65,
    'tension_srf': False,          # NO — the non-default, so a leak would show
    'element_type': 'quad9',
    'target_size': 2.75,
    'ssrm_f_min': 0.8,
    'ssrm_f_max': 2.4,
}
# --- v20 SSR-zone overlay round-trip ---
# A polygon-geometry model, so the zone rows are written AFTER real material-zone
# rows (the ordering the writer has to get right) and the "zones never enter
# 'polygons'" check has material zones to distinguish them from.
SSR_ZONE_ROUNDTRIP_BASE = _repo('docs/seep/files/xslope_levee_poly.xlsx')
# --- v21 round-trip ---
# The v21 additions (polygon Type/Size, profile Size, dload Direction, main Side BC)
# exist on NO corpus file, so like the v19 block above they are proved on synthetic
# values written onto real models. Two bases, because the polygon and profile
# geometry forms take different writer branches and the Size row moved in both.
V21_ROUNDTRIP_BASE = _repo('docs/seep/files/xslope_levee_poly.xlsx')
V21_ROUNDTRIP_PROFILE_BASE = _repo('docs/fem/files/xslope_griffiths1_load.xlsx')
#: A tseep-bearing model, for the save_times column — the one tseep anchor whose row
#: is version-dependent, and the only one no corpus round-trip would notice moving.
V21_ROUNDTRIP_TSEEP_BASE = _repo('docs/seep/files/xslope_earth_dam_tseep.xlsx')
# --- v22 surface-family round-trip ---
# A one-circle model, so the test can make a BOTH-family deck out of it (the only
# case the cell means anything for) and still check that the single-family original
# is unaffected. The non-circular surface is synthetic: what is under test is the
# cell, not the geometry, and no corpus file carries both families.
SURFACE_FAMILY_BASE = _repo('docs/inputs/slope/xslope_simple1.xlsx')
SURFACE_FAMILY_NONCIRC = [
    {'X': 20.0, 'Y': 40.0, 'Movement': 'Free'},
    {'X': 60.0, 'Y': 10.0, 'Movement': 'Horiz'},
    {'X': 110.0, 'Y': 40.0, 'Movement': 'Free'},
]
V19_SEARCH_WINDOW = {
    'entry_x_min': 41.0, 'entry_x_max': 54.5,
    'exit_x_min': 23.25, 'exit_x_max': 32.0,
    'center_box_x_min': 30.0, 'center_box_x_max': 70.0,
    'center_box_y_min': 40.0, 'center_box_y_max': 90.0,
    'max_tangent_depth': 16.5, 'min_slip_depth': 1.25,
}


def _roundtrip_eq(a, b):
    """Scalar equality with float tolerance; strings compared verbatim (so the
    loader's empty-cell 'nan' sentinel matches itself)."""
    import math
    if str(a) == str(b):
        return True
    try:
        fa, fb = float(a), float(b)
    except (TypeError, ValueError):
        return False
    # A value that goes in nonzero and comes back EXACTLY zero is never a
    # round-trip, whatever its magnitude: the abs_tol below would otherwise wave
    # through a 1e-13 m/s hydraulic conductivity written as 0, which reloads as a
    # silent hole in the model rather than an error.
    if (fa == 0.0) != (fb == 0.0):
        return False
    return math.isclose(fa, fb, rel_tol=1e-6, abs_tol=1e-6)


def _roundtrip_diff(a, b, path=''):
    """Recursively compare two slope_data values; return a list of mismatch paths."""
    out = []
    if isinstance(a, dict):
        if not isinstance(b, dict):
            return [f"{path}: dict vs {type(b).__name__}"]
        for k in a:
            if k not in b:
                out.append(f"{path}.{k}: missing")
            else:
                out += _roundtrip_diff(a[k], b[k], f"{path}.{k}")
        return out
    if isinstance(a, (list, tuple)):
        if not isinstance(b, (list, tuple)) or len(a) != len(b):
            blen = len(b) if hasattr(b, '__len__') else '?'
            return [f"{path}: len {len(a)} vs {blen}"]
        for i, (x, y) in enumerate(zip(a, b)):
            out += _roundtrip_diff(x, y, f"{path}[{i}]")
        return out
    if not _roundtrip_eq(a, b):
        out.append(f"{path}: {a!r} != {b!r}")
    return out


def parse_test_tags(md_path):
    """Parse <!-- test: ... --> tags from a markdown file.

    A tag may name a single method the classic way (method=spencer,
    expected_fs=1.23) or carry per-method values (fs_oms=0.94, fs_bishop=0.99,
    ...). The latter form is expanded into one test per method, so every method
    becomes an independently-checked regression.
    """
    tests = []
    md_dir = Path(md_path).parent

    with open(md_path, 'r') as f:
        content = f.read()

    pattern = r'<!--\s*test:\s*(.*?)\s*-->'
    for match in re.finditer(pattern, content):
        tag_str = match.group(1)
        params = {}
        for pair in tag_str.split(','):
            key, _, value = pair.partition('=')
            params[key.strip()] = value.strip()

        # Resolve file path relative to the markdown file's directory
        if 'file' in params:
            params['file'] = str(md_dir / params['file'])
        if 'file2' in params:
            params['file2'] = str(md_dir / params['file2'])

        # Convert numeric fields
        for key in ['expected_fs', 'expected_flowrate', 'expected_beta', 'tolerance', 'target_size', 'f_min', 'f_max', 'beta', 'k0',
                    'expected_kc', 'k_min', 'k_max', 'fs_tol', 'kc_tol', 'refine_factor',
                    'expected_pf', 'pf_tol']:
            if key in params:
                params[key] = float(params[key])
        for key in ['num_slices', 'n_samples', 'rng_seed', 'circle_index',
                    'expected_elements', 'expected_nodes']:
            if key in params:
                params[key] = int(params[key])

        params['source'] = str(md_path)

        # Expand a compact multi-method tag (fs_oms=..., fs_bishop=..., ...) into
        # one test per method; otherwise keep the single tag as-is.
        fs_keys = [k for k in params if k.startswith('fs_')]
        if fs_keys:
            shared = {k: v for k, v in params.items() if not k.startswith('fs_')}
            for k in fs_keys:
                t = dict(shared)
                t['method'] = k[3:]      # fs_<method> tag -> solver function name
                t['expected_fs'] = float(params[k])
                tests.append(t)
        else:
            tests.append(params)

    return tests


def _refine_kwargs(test):
    """Optional feature-aware mesh-refinement knobs from a test tag, as build-mesh
    kwargs. ``refine_factor=N`` turns it on (None/absent = OFF = byte-identical mesh);
    ``refine_features=a;b`` (semicolon-delimited, since the tag splits on commas)
    selects feature classes, defaulting to all. Returns {} when no refinement."""
    rf = test.get('refine_factor')
    if rf is None:
        return {}
    kw = {'refine_factor': float(rf)}
    feats = test.get('refine_features')
    if feats:
        kw['refine_features'] = [s.strip() for s in str(feats).split(';') if s.strip()]
    return kw


def run_lem_test(test):
    """Run a single LEM test (single_circle, circular_search, or noncircular_search)."""
    from xslope.fileio import load_slope_data
    from xslope.slice import generate_slices
    from xslope.solve import solve_selected
    from xslope.search import circular_search, noncircular_search

    file_path = test['file']
    test_type = test['type']
    method = test['method']
    num_slices = test.get('num_slices', 30)
    # `rapid=true` runs the 3-stage rapid-drawdown analysis (needs the seep solution
    # + dloads/seep-bc stage-2 data in the input). Works with any test type.
    rapid = str(test.get('rapid', 'false')).strip().lower() in ('true', '1', 'yes')
    # `composite=true` lets a circle be truncated at the bottom of the model and run
    # along it (slice.CompositeSurface). Only for problems whose base is a real
    # impenetrable boundary; off by default.
    composite = str(test.get('composite', 'false')).strip().lower() in ('true', '1', 'yes')
    # `seed=grid` runs the circular search from an automatic grid-and-tangent sweep
    # instead of (only) the circles sheet — the global-search mode.
    seed = str(test.get('seed', 'circles')).strip().lower()
    # `right_facing=true/false` overrides the automatic facing detection. Needed for
    # level-ground bearing surfaces (flat arc), where the surface is symmetric and the
    # facing is set by the load asymmetry the geometry can't see (VP26). Absent → the
    # engine's automatic rule (unchanged for every normal slope).
    _rf = str(test.get('right_facing', '')).strip().lower()
    right_facing = True if _rf in ('true', '1', 'yes') else (
        False if _rf in ('false', '0', 'no') else None)

    # Optional circular_search window constraints (Slide2 "search limits" semantics).
    # Values are semicolon-separated (commas already split the tag): center_box has 4
    # numbers, entry_range/exit_range/tangent_depth have 2. Absent -> None (the
    # unconstrained global search, unchanged). See xslope.search.circular_search.
    def _bounds(key, n):
        raw = test.get(key)
        if raw is None or str(raw).strip() == '':
            return None
        parts = [float(v) for v in str(raw).split(';') if v.strip() != '']
        if len(parts) != n:
            raise ValueError(f"{key} needs {n} ';'-separated numbers, got {raw!r}")
        return tuple(parts)
    center_box = _bounds('center_box', 4)
    entry_range = _bounds('entry_range', 2)
    exit_range = _bounds('exit_range', 2)
    tangent_depth = _bounds('tangent_depth', 2)

    # Optional matric-suction apparent-cohesion run options (opt-in Fredlund
    # extended Mohr-Coulomb). `suction_phi_b` is "Name:deg;Name2:deg" — a per-
    # material suction-strength angle φᵇ; `suction_cap` is one number (stress
    # units) bounding the suction before it becomes apparent cohesion. Absent →
    # None (the clamped baseline, unchanged). Threaded through generate_slices on
    # the direct-surface test types; the search types would need the angle carried
    # through the search itself, so a suction tag on them is rejected rather than
    # silently dropped.
    def _suction_phi_b():
        raw = test.get('suction_phi_b')
        if raw is None or str(raw).strip() == '':
            return None
        out = {}
        for tok in str(raw).split(';'):
            tok = tok.strip()
            if not tok:
                continue
            name, sep, deg = tok.rpartition(':')
            if not sep or not name.strip() or deg.strip() == '':
                raise ValueError(f"suction_phi_b entry {tok!r} must be 'Name:degrees'")
            out[name.strip()] = float(deg)
        return out or None
    suction_phi_b = _suction_phi_b()
    _sc = test.get('suction_cap')
    suction_cap = float(_sc) if _sc is not None and str(_sc).strip() != '' else None
    if (suction_phi_b or suction_cap is not None) and test_type in (
            'circular_search', 'noncircular_search'):
        return None, ("suction_phi_b/suction_cap are supported on single_circle and "
                      "single_noncirc tests only (the search path does not thread "
                      "the suction-strength angle)")

    slope_data = load_slope_data(file_path)

    if test_type == 'single_circle':
        # circle_index picks one of several specified surfaces stored in the same
        # file (default 0, the historical behaviour). A model whose published
        # answers are per-surface — SLOPE/W's Cannon Dam #2 stores nine of
        # Hassan & Wolff's fixed circles on one geometry — is one input file with
        # one tag per circle, not one duplicated file per circle.
        ci = int(test.get('circle_index', 0))
        circles = slope_data['circles']
        if not 0 <= ci < len(circles):
            return None, (f"circle_index={ci} out of range: the file has "
                          f"{len(circles)} circle(s)")
        circle = circles[ci]
        success, result = generate_slices(slope_data, circle=circle, num_slices=num_slices,
                                          composite=composite, right_facing=right_facing,
                                          suction_phi_b=suction_phi_b, suction_cap=suction_cap)
        if not success:
            return None, f"generate_slices failed: {result}"
        slice_df, failure_surface = result
        solver_result = solve_selected(method, slice_df, rapid=rapid)
        if isinstance(solver_result, str):
            return None, f"solve failed: {solver_result}"
        # Regression guard: rapid drawdown must write the winning stage's stresses
        # back to the caller's slice_df (else plots get n_eff=0 — see
        # advanced.rapid_drawdown). The FS can be correct while n_eff is stale.
        if rapid and not bool((slice_df['n_eff'].abs() > 0).any()):
            return None, "rapid drawdown left n_eff all-zero in the slice_df (not written back)"
        return solver_result['FS'], None

    elif test_type == 'circular_search':
        fs_cache, converged, search_path, circle_cache = circular_search(
            slope_data, method, num_slices=num_slices, rapid=rapid, composite=composite,
            seed=seed, center_box=center_box, entry_range=entry_range,
            exit_range=exit_range, tangent_depth=tangent_depth
        )
        if not fs_cache or fs_cache[0]['FS'] >= 9999:
            return None, "circular_search found no valid surface"
        if rapid:
            crit = fs_cache[0].get('slices')
            if crit is not None and not bool((crit['n_eff'].abs() > 0).any()):
                return None, "rapid search: critical-surface n_eff all-zero (not written back)"
        return fs_cache[0]['FS'], None

    elif test_type == 'single_noncirc':
        # Evaluate the file's specified non-circular surface as-is (no search) —
        # the "predefined slip surface" form of the verification problems.
        success, result = generate_slices(slope_data, non_circ=slope_data['non_circ'],
                                          num_slices=num_slices, right_facing=right_facing,
                                          suction_phi_b=suction_phi_b, suction_cap=suction_cap)
        if not success:
            return None, f"generate_slices failed: {result}"
        slice_df, failure_surface = result
        solver_result = solve_selected(method, slice_df, rapid=rapid)
        if isinstance(solver_result, str):
            return None, f"solve failed: {solver_result}"
        return solver_result['FS'], None

    elif test_type == 'noncircular_search':
        fs_cache, converged, search_path = noncircular_search(
            slope_data, method, num_slices=num_slices, rapid=rapid
        )
        if not fs_cache or fs_cache[0]['FS'] >= 9999:
            return None, "noncircular_search found no valid surface"
        return fs_cache[0]['FS'], None

    else:
        return None, f"Unknown LEM test type: {test_type}"


def run_critical_kc_test(test):
    """Find the CRITICAL seismic coefficient k_c — the horizontal pseudo-static
    coefficient at which the searched minimum factor of safety equals 1.0 — and
    compare it to the published k_c. This is the verification target for RS2 #68
    (Loukidis, Bandini & Salgado 2003), which publishes k_c, not an FS.

    Tag keys: file, method, expected_kc, k_min, k_max, num_slices (default 40),
    fs_tol (confirming-search FS band, default 0.005), kc_tol (the comparison
    tolerance, via _expected_and_tol), max_outer (default 3).

    Algorithm (cheap and honest): FS decreases monotonically as k rises, so k_c is
    a single crossing. A full circular search at the bracket midpoint fixes a near-
    critical circle; k is then bisected on that FIXED circle with fast single-circle
    solves (~0.1 s each) until FS = 1; a confirming FULL search at that k re-checks
    that the true minimum surface there is also FS ≈ 1. If a more critical surface
    has emerged (the critical circle migrates with k), it is adopted and the inner
    bisection repeats — so the result converges to the search-minimised k_c from
    above, at the cost of ~2 full searches rather than one per bisection step.
    Returns the k_c (compared to expected_kc within kc_tol by the framework)."""
    from xslope.fileio import load_slope_data
    from xslope.slice import generate_slices
    from xslope.solve import solve_selected
    from xslope.search import circular_search

    method = test['method']
    num_slices = int(test.get('num_slices', 40))
    k_min = float(test['k_min'])
    k_max = float(test['k_max'])
    fs_tol = float(test.get('fs_tol', 0.005))
    max_outer = int(test.get('max_outer', 3))

    slope_data = load_slope_data(test['file'])

    def search_min(k):
        """Full circular search at coefficient k → (min_FS, critical circle)."""
        slope_data['k_seismic'] = k
        fs_cache, _conv, _path, _cc = circular_search(
            slope_data, method, num_slices=num_slices)
        if not fs_cache or fs_cache[0]['FS'] >= 9999:
            return None, None
        b = fs_cache[0]
        return b['FS'], {'Xo': b['Xo'], 'Yo': b['Yo'], 'Depth': b['Depth'],
                         'R': b['Yo'] - b['Depth']}

    def fs_on(circle, k):
        """FS on a FIXED circle at coefficient k (one single_circle solve)."""
        slope_data['k_seismic'] = k
        ok, res = generate_slices(slope_data, circle=circle, num_slices=num_slices)
        if not ok:
            return None
        slice_df, _fs = res
        out = solve_selected(method, slice_df)
        if isinstance(out, str):
            return None
        return out['FS']

    _fs_mid, circle = search_min(0.5 * (k_min + k_max))
    if circle is None:
        return None, f"circular search found no surface at k={0.5*(k_min+k_max):.4f}"

    kc = 0.5 * (k_min + k_max)
    for _outer in range(max_outer):
        lo, hi = k_min, k_max
        fs_lo, fs_hi = fs_on(circle, lo), fs_on(circle, hi)
        if fs_lo is None or fs_hi is None:
            return None, "fixed-circle solve failed at a bracket endpoint"
        if not (fs_hi <= 1.0 <= fs_lo):
            return None, (f"k_c is not bracketed on the critical circle: "
                          f"FS({lo})={fs_lo:.3f}, FS({hi})={fs_hi:.3f} "
                          f"(need FS(k_min)>=1>=FS(k_max)) — widen k_min/k_max")
        fs = fs_hi
        for _inner in range(50):
            kc = 0.5 * (lo + hi)
            fs = fs_on(circle, kc)
            if fs > 1.0:
                lo = kc            # too stable — raise k
            else:
                hi = kc            # unstable — lower k
            if abs(fs - 1.0) <= 1e-4:
                break
        fs_chk, circle_chk = search_min(kc)
        if circle_chk is None:
            return None, f"confirming search found no surface at k={kc:.4f}"
        if abs(fs_chk - 1.0) <= fs_tol:
            return kc, None        # the true minimum surface at k_c is also FS≈1
        circle = circle_chk        # critical circle migrated — adopt and re-bisect
    return kc, None                # best estimate after max_outer iterations


def run_design_test(test):
    """Run a design-sweep regression: edit a profile point to set the slope-face
    angle, rebuild the polygon geometry, then run a circular search and return FS.

    This guards the bug found in the design driver (main_design.py), where editing
    profile_lines and rebuilding only the ground surface left the slice weights —
    and therefore the factor of safety — pinned to the ORIGINAL geometry. Slice
    weights are computed from slope_data['polygons'], so the polygons must be
    regenerated from the edited profile. The expected FS reflects the NEW slope
    angle; if the polygon resync regresses, the search returns the stale
    base-geometry FS and this test fails.
    """
    import math
    from shapely.geometry import Polygon
    from xslope.fileio import load_slope_data, build_ground_surface_from_polygons
    from xslope.mesh import build_polygons
    from xslope.search import circular_search

    file_path = test['file']
    method = test.get('method', 'bishop')
    beta = float(test['beta'])
    toe_index = int(test.get('toe_index', 1))
    slope_index = int(test.get('slope_index', 2))
    num_slices = test.get('num_slices', 30)

    slope_data = load_slope_data(file_path)

    # Move the slope-top point so the slope face makes angle `beta`:
    #   tan(beta) = (y_top - y_toe) / (x_top - x_toe)
    profile = slope_data['profile_lines'][0]['coords']
    x_toe, y_toe = profile[toe_index]
    _x_top, y_top = profile[slope_index]
    x_top_new = x_toe + (y_top - y_toe) / math.tan(math.radians(beta))
    slope_data['profile_lines'][0]['coords'][slope_index] = (x_top_new, y_top)

    # Rebuild material polygons from the edited profile, then re-derive the ground
    # surface / domain polygon from them — the resync that the original bug omitted.
    polys = [
        {'polygon': Polygon(p['coords']), 'mat_id': p['mat_id']}
        for p in build_polygons(slope_data={'profile_lines': slope_data['profile_lines'],
                                             'max_depth': slope_data.get('max_depth')})
    ]
    slope_data['polygons'] = polys
    ground_surface, domain_polygon = build_ground_surface_from_polygons(polys)
    slope_data['ground_surface'] = ground_surface
    slope_data['domain_polygon'] = domain_polygon

    fs_cache, converged, search_path, circle_cache = circular_search(
        slope_data, method, num_slices=num_slices)
    if not fs_cache or fs_cache[0]['FS'] >= 9999:
        return None, "design_search found no valid surface"
    return fs_cache[0]['FS'], None


def build_fem_ssrm_case(test):
    """Turn a ``fem_ssrm`` test tag into everything ``solve_ssrm`` needs.

    Returns ``(fem_data, kwargs, f_min, f_max, ssrm_tolerance)``. Split out of
    ``run_fem_test`` so that the tag -> (mesh, fem_data, solver options) mapping has
    exactly ONE implementation: the suite calls it, and so does any driver that has
    to run the same benchmark under different solver settings (e.g.
    ``benchmarks/hybrid_criterion_ab.py``, which solves each case under two failure
    criteria). Keeping it here rather than copying the mapping is what stops a
    driver from silently benchmarking a different mesh than the lock."""
    from xslope.fileio import load_slope_data
    from xslope.fem import build_fem_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_constraint_line_geometry, extract_point_constraints,
                             extract_size_regions)

    file_path = test['file']
    element_type = test.get('element_type', 'tri6')
    target_size = test.get('target_size')  # default computed from domain extent below
    ssrm_tolerance = test.get('tolerance', 0.05)

    slope_data = load_slope_data(file_path)

    # Seepage-coupled models (material u option = 'seep') must run on the SAME
    # mesh as the stored seepage solution: seep_u is nodal, and build_fem_data
    # silently zeroes the pore pressures if the node count differs.
    uses_seep = any(str(m.get('u', '')).strip().lower() == 'seep'
                    for m in slope_data.get('materials', []))
    if uses_seep and slope_data.get('mesh') is not None:
        mesh = slope_data['mesh']
    else:
        # Default mesh size scales with the domain (like the seepage path) rather
        # than a fixed value, so a wide domain does not blow up to a huge mesh.
        if target_size is None:
            x_coords = [x for x, _ in slope_data['ground_surface'].coords]
            target_size = (max(x_coords) - min(x_coords)) / 100
        # constraint lines include BOTH reinforcement and pile axes — pile
        # beam elements require their axis lines in the mesh (a
        # reinforcement-only extraction silently drops the piles)
        constraint_lines, _n_reinf, _n_pile = extract_constraint_line_geometry(slope_data)
        polygons = get_material_polygons(slope_data, reinf_lines=constraint_lines)
        mesh = build_mesh_from_polygons(
            polygons, target_size=target_size, element_type=element_type,
            lines=constraint_lines,
            point_constraints=extract_point_constraints(slope_data),
            size_regions=extract_size_regions(slope_data),
            **_refine_kwargs(test)
        )

    fem_data = build_fem_data(slope_data, mesh)
    f_min = test.get('f_min', 0.5)
    f_max = test.get('f_max', 3.0)
    kwargs = {}
    if 'criterion' in test:
        kwargs['failure_criterion'] = test['criterion']
    if 'max_iter' in test:
        kwargs['max_iterations'] = int(test['max_iter'])
    # TEMP (#50): the Dawson per-node criterion needs ~3x the iterations the old rate-based
    # test did (displacement settles ~5k, the per-node max only ~11k). Let the experiment
    # raise the floor without editing 64 tags.
    import os as _os
    _floor = int(_os.environ.get('XSLOPE_MIN_MAX_ITER', '0'))
    if _floor:
        kwargs['max_iterations'] = max(kwargs.get('max_iterations', 0), _floor)
    if test.get('cutoff', '').lower() in ('true', '1', 'yes'):
        kwargs['tension_cutoff'] = True
    # tension_srf divides each tension cap by the trial F alongside c and tan(phi)
    # (RS2's tensilestrength_SRF=1). Per-material caps themselves come from the
    # file's mat t_cut column via build_fem_data. The solver default is ON, so an
    # absent tag rides that default (a no-op on the cap-less majority); the tag is
    # honored BOTH ways so a future tension_srf=false pins the static-cap run
    # instead of being silently overridden by the default.
    if 'tension_srf' in test:
        kwargs['tension_srf'] = str(test['tension_srf']).lower() in ('true', '1', 'yes')
    # K0 initial stress (v19). Absent -> the solver's None = gravity turn-on, and
    # the file's own main!D16 if it declares one. Present -> wins over the file.
    if 'k0' in test:
        kwargs['k0'] = float(test['k0'])
    # Surficial-skin filter: exclude any mechanism shallower than this depth below
    # the ground surface, so the row reports the deep-seated factor of safety
    # instead of a c = 0 face skin (docs/fem/overview.md, "Surficial (Skin)
    # Failures"). Absent -> the solver's None = filter off, the true global minimum.
    if 'min_slip_depth' in test:
        kwargs['min_slip_depth'] = float(test['min_slip_depth'])
    if 'char_x' in test and 'char_y' in test:
        kwargs['char_point'] = (float(test['char_x']), float(test['char_y']))
    # SSR-exclusion material names. Tags split on commas, so the material names
    # within ssr_exclude are separated by SEMICOLONS: ssr_exclude=Name1;Name2.
    if 'ssr_exclude' in test:
        kwargs['ssr_exclude'] = [s.strip() for s in str(test['ssr_exclude']).split(';')
                                 if s.strip()]
    # SSR search-area polygon (RS2's "SSR Search Area"): a flat x1;y1;x2;y2;... list
    # of the constraint polygon's vertices (semicolon-separated, since tags split on
    # commas). Strength reduction is confined to elements inside the polygon.
    if 'ssr_zone' in test:
        _z = [float(v) for v in str(test['ssr_zone']).split(';') if v.strip() != '']
        kwargs['ssr_zone'] = list(zip(_z[0::2], _z[1::2]))
    # Elastic-materials: names of materials held PURE LINEAR ELASTIC (skip plasticity
    # entirely, cannot yield) — RS2's "Plasticity Specifications: None". Distinct from
    # ssr_exclude (full strength but still yields). Tags split on commas, so the names
    # are SEMICOLON-separated: elastic_materials=Name1;Name2.
    if 'elastic_materials' in test:
        kwargs['elastic_materials'] = [s.strip() for s in
                                       str(test['elastic_materials']).split(';')
                                       if s.strip()]
    # Bond-slip load transfer for 1D reinforcement (opt-in). Tags split on commas,
    # so line entries are SEMICOLON-separated and their fields COLON-separated:
    # bond_slip=<line>:<bond_c>:<bond_phi_deg>:<perimeter>[;<line>:...]. <line> is a
    # reinforcement line label, a 1-based id, or '*' (all lines). Off by default.
    if 'bond_slip' in test:
        bs = {}
        for entry in str(test['bond_slip']).split(';'):
            entry = entry.strip()
            if not entry:
                continue
            key, c, phi, perim = entry.split(':')
            key = key.strip()
            # numeric key => 1-based line id; otherwise a label (or '*')
            try:
                key = int(key)
            except ValueError:
                pass
            bs[key] = (float(c), float(phi), float(perim))
        kwargs['bond_slip'] = bs
    # Matric-suction strength (Fredlund extended MC), opt-in. suction_phi_b is a
    # per-material angle list "Name:deg;Name2:deg" (semicolon-separated, since tags
    # split on commas); suction_cap is one number (stress units) bounding the credited
    # suction. Absent => auto-wired from the file's v17 phi_b/s_cap (off if blank). The
    # apparent cohesion s*tan(phi_b) above the water table is reduced by the trial F.
    if 'suction_phi_b' in test and str(test['suction_phi_b']).strip():
        sp = {}
        for tok in str(test['suction_phi_b']).split(';'):
            tok = tok.strip()
            if not tok:
                continue
            name, sep, deg = tok.rpartition(':')
            if not sep or not name.strip() or deg.strip() == '':
                raise ValueError(f"suction_phi_b entry {tok!r} must be 'Name:degrees'")
            sp[name.strip()] = float(deg)
        kwargs['suction_phi_b'] = sp or None
    if 'suction_cap' in test and str(test['suction_cap']).strip():
        kwargs['suction_cap'] = float(test['suction_cap'])
    return fem_data, kwargs, f_min, f_max, ssrm_tolerance


def run_fem_test(test, fast_kernel=None):
    """Run a single FEM SSRM test.

    ``fast_kernel`` selects the Mohr-Coulomb kernel used for this solve:

      * ``True``  — force the compiled fast kernel for the solve_ssrm trials.
      * ``False`` — force the pure-NumPy reference kernel.
      * ``None``  — inherit ``fem.solve_fem``'s own default, which is ``'auto'``
        (compiled kernel whenever it is built). NO SUITE PATH USES THIS: every
        caller in this file passes ``True`` or ``False`` explicitly, because a
        lock is defined by the reference path and an inherited default would let
        the machine's build state decide which kernel rendered a verdict. It is
        retained only for ad-hoc interactive use.

    ``solve_ssrm`` exposes no ``fast_kernel`` parameter (it calls ``solve_fem`` by
    bare name), so ``True``/``False`` are threaded by temporarily wrapping
    ``fem.solve_fem`` — see ``_force_fast_kernel``. Everything else — the mesh
    build, ``fem_data``, the assembled ``kwargs``, and (unchanged)
    ``capture_failure_state`` — is identical across all three modes, so the fast
    trial and the reference fallback each cost exactly what today's single solve
    costs and the suite pays no new capture overhead."""
    from xslope.fem import solve_ssrm

    try:
        fem_data, kwargs, f_min, f_max, ssrm_tolerance = build_fem_ssrm_case(test)
    except ValueError as exc:
        return None, str(exc)

    # capture_failure_state is deliberately NOT set here — the single-kernel runner
    # never has, so solve_ssrm's own default rides through unchanged. Both the fast
    # trial and the reference fallback reuse this exact call, so neither tier of the
    # two-tier fem_ssrm path introduces capture cost the suite did not already pay.
    if fast_kernel is None:
        result = solve_ssrm(fem_data, F_min=f_min, F_max=f_max, tolerance=ssrm_tolerance,
                            debug_level=0, **kwargs)
    else:
        import xslope.fem as _fem
        with _force_fast_kernel(_fem, fast_kernel):
            result = solve_ssrm(fem_data, F_min=f_min, F_max=f_max, tolerance=ssrm_tolerance,
                                debug_level=0, **kwargs)

    if result.get('converged', False):
        return result['FS'], None
    else:
        return None, f"SSRM failed: {result.get('error', 'Unknown error')}"


class _force_fast_kernel:
    """Context manager forcing solve_ssrm's internal solve_fem trials onto the fast
    (``on=True``) or reference (``on=False``) Mohr-Coulomb kernel.

    ``solve_ssrm`` has no ``fast_kernel`` parameter — it calls ``solve_fem`` by bare
    name — so the flag is threaded by temporarily wrapping ``fem.solve_fem``. This is
    the exact mechanism ``benchmarks/kernel_xcheck.py`` used in the 84-case soak,
    productionized here for the suite. (kernel_xcheck keeps its own copy so it stays
    runnable standalone; this one duplicate is intentional layering — the suite does
    not depend on a benchmark script importing.) Process-safe, not thread-safe: each
    parallel worker owns its interpreter and runs its rows serially, and the wrap is
    always restored in ``__exit__``."""

    def __init__(self, fem_mod, on):
        self._fem = fem_mod
        self._on = on
        self._orig = fem_mod.solve_fem

    def __enter__(self):
        orig, on = self._orig, self._on

        def _wrap(*a, **k):
            k['fast_kernel'] = on
            return orig(*a, **k)
        self._fem.solve_fem = _wrap
        return self

    def __exit__(self, *exc):
        self._fem.solve_fem = self._orig
        return False


def _fast_kernel_available():
    """True when the compiled Mohr-Coulomb kernel (built by setup_kernel.py) imports."""
    try:
        from xslope import _fem_kernel  # noqa: F401
        return True
    except ImportError:
        return False


def _ssrm_mode_notice(n_ssrm, reference_only):
    """One-line startup notice describing how the run's ``n_ssrm`` fem_ssrm rows
    will be verified. ``--reference-only`` wins; otherwise the compiled fast kernel
    being built (or not) decides between fast-first and reference-only. Returns the
    message string (the caller prints it)."""
    if reference_only:
        return (f"FEM SSRM: {n_ssrm} row(s) verified reference-only "
                f"(--reference-only; fast-first disabled)")
    if _fast_kernel_available():
        return (f"FEM SSRM: {n_ssrm} row(s) run fast-first with reference "
                f"fallback (fast kernel available)")
    return (f"FEM SSRM: {n_ssrm} row(s) verified reference-only "
            f"(compiled fast kernel not built)")


def _run_fem_ssrm(test):
    """Two-tier *fast-first-with-fallback* runner for a single ``fem_ssrm`` row.
    Returns ``(computed_FS, error_msg, annotation)`` where ``annotation`` is a
    ``(bucket, text)`` routing note the summary tallies (``bucket`` in
    ``{'fast', 'fallback', 'direct'}``).

    ================================ WHY THIS EXISTS ================================
    Read this before "simplifying" the two-tier scheme away — it is load-bearing.

    (1) THE LOCKS ARE PROPERTIES OF THE REFERENCE PATH. Every locked factor of
        safety in this suite is *defined* by the pure-NumPy Step-6 solver in
        ``fem.solve_fem`` (the oracle). The suite's job is to guard that oracle. The
        compiled fast kernel is an optimization that must reproduce the oracle
        bit-for-bit; it is never itself the definition of a lock.

    (2) FAST-FIRST IS SOUND because ~95% of the SSRM pipeline (mesh, assembly, BCs,
        the bisection driver, the plasticity return-mapping structure) is SHARED by
        both kernels. A regression in that shared code fails BOTH paths, so the fast
        solve misses the lock, we fall back to the reference solve, and it *also*
        misses — a TRUE alarm, correctly raised. Meanwhile the harmless case — a
        knife-edge fast miss where the fast kernel lands a hair off the lock (e.g.
        RS2-62c's soft band, fast FS 0.773 vs lock 0.801) — auto-resolves: the
        reference re-solve lands on the lock and the row passes with a "fell back"
        annotation, no false alarm and no human in the loop. So fast-first turns the
        common healthy run cheap without ever weakening a verdict: the reference
        verdict is always FINAL, whether it passes or fails.

    (3) THE ONE GAP is a change to REFERENCE-ONLY constitutive physics (a kernel the
        fast path does not share) that happens to keep the fast kernel passing the
        locks — fast-first would never trigger the fallback and the reference
        regression would hide. That gap is closed by the ``kernel_xcheck`` gate,
        which solves small cases BOTH ways and fails on any FS or field divergence
        between them. That is why ``kernel_xcheck`` is a REQUIRED companion to this
        scheme and MUST NOT be removed while fast-first is the default: it is the
        guard that makes fast-first safe against reference-only drift.

    (4) ``--reference-only`` FORCES the pure-reference verdict for every row (no fast
        first pass). Use it for strict verification runs: pre-release, or right after
        any constitutive-physics edit, when you want the oracle to speak directly and
        do not want a fast pass masking a reference change before kernel_xcheck runs.
    ================================================================================

    When the compiled kernel is absent, or ``--reference-only`` is set, this is a
    single pure-reference solve — exactly today's behavior."""
    default_tol = float(test.get('_default_tol', 0.01))
    reference_only = bool(test.get('_reference_only', False))
    expected, tol = _expected_and_tol(test, default_tol)

    # Reference-only, or the compiled kernel isn't built: today's behavior exactly —
    # a single pure-reference solve, no fast first pass. Direct bucket (not a
    # fallback: the fast path was never attempted).
    if reference_only or not _fast_kernel_available():
        # fast_kernel=False is passed EXPLICITLY, never left to solve_fem's own
        # default. That default is 'auto' (compiled kernel when it is built), so
        # inheriting it here would silently run --reference-only rows on the fast
        # kernel — the exact opposite of what the flag promises, and it would
        # dissolve the drift fence. Every tier of this runner pins its kernel.
        computed, err = run_fem_test(test, fast_kernel=False)
        why = '--reference-only' if reference_only else 'fast kernel not built'
        return computed, err, ('direct', f'via reference ({why})')

    # Tier 1 — fast kernel. A hit (within the same expected/tolerance the framework
    # uses) is a PASS by construction, annotated "via fast kernel".
    fast_fs, fast_err = run_fem_test(test, fast_kernel=True)
    if fast_err is None and expected is not None and abs(fast_fs - expected) <= tol:
        return fast_fs, None, ('fast', 'via fast kernel')

    # Tier 2 — reference fallback. The reference verdict is FINAL (PASS or FAIL); the
    # framework re-checks the returned reference FS against the same expected/tol.
    ref_fs, ref_err = run_fem_test(test, fast_kernel=False)
    if ref_err is not None:
        return None, ref_err, ('fallback', 'via reference (fast missed; reference errored)')
    if fast_err is not None:
        text = f'via reference (fast errored: {fast_err})'
    elif expected is not None:
        text = f'via reference (fast missed by d={abs(fast_fs - expected):.4f})'
    else:
        text = 'via reference (fast miss)'
    return ref_fs, None, ('fallback', text)


def run_seep_test(test):
    """Run a single seepage test."""
    from xslope.fileio import load_slope_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_size_regions)
    from xslope.seep import build_seep_data, run_seepage_analysis

    file_path = test['file']
    slope_data = load_slope_data(file_path)

    polygons = get_material_polygons(slope_data)
    element_type = test.get('element_type', 'tri3')
    target_size = test.get('target_size')
    if target_size is None:
        x_coords = [x for x, _ in slope_data['ground_surface'].coords]
        target_size = (max(x_coords) - min(x_coords)) / 120
    mesh = build_mesh_from_polygons(polygons, target_size, element_type,
                                    size_regions=extract_size_regions(slope_data),
                                    **_refine_kwargs(test))

    seep_data = build_seep_data(mesh, slope_data)
    solution = run_seepage_analysis(seep_data, tol=1e-4,
                                    max_iter=int(test.get('max_iter', 400)))

    if not solution.get('converged', True):
        return None, "seepage solution did not converge (flowrate unreliable)"
    return solution['flowrate'], None


def run_seep_head_test(test):
    """Check solved total head at named points (groundwater-corpus tags).

    Tag keys: file, points="x:y:h;x:y:h;...", tolerance (m, absolute),
    optional target_size / element_type. Meshes and solves live like
    run_seep_test, then interpolates head at each point by inverse-distance
    over the four nearest nodes. Pass/fail: returns 0.0 on success."""
    import numpy as np
    from xslope.fileio import load_slope_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_size_regions)
    from xslope.seep import build_seep_data, run_seepage_analysis

    slope_data = load_slope_data(test['file'])
    polygons = get_material_polygons(slope_data)
    target_size = test.get('target_size')
    if target_size is None:
        xs = [x for x, _ in slope_data['ground_surface'].coords]
        target_size = (max(xs) - min(xs)) / 120
    mesh = build_mesh_from_polygons(polygons, float(target_size),
                                    test.get('element_type', 'tri3'),
                                    size_regions=extract_size_regions(slope_data),
                                    **_refine_kwargs(test))
    seep_data = build_seep_data(mesh, slope_data)
    solution = run_seepage_analysis(seep_data, tol=1e-5,
                                    max_iter=int(test.get('max_iter', 400)))
    if not solution.get('converged', True):
        return None, "seepage solution did not converge"

    nodes = seep_data['nodes']
    h = np.asarray(solution['head'])

    def head_at(xq, yq):
        d2 = (nodes[:, 0] - xq) ** 2 + (nodes[:, 1] - yq) ** 2
        idx = np.argsort(d2)[:4]
        w = 1.0 / np.maximum(d2[idx], 1e-12)
        return float(np.sum(w * h[idx]) / np.sum(w))

    tol = float(test.get('tolerance', 0.01))
    errs = []
    for triplet in str(test['points']).split(';'):
        xs_, ys_, hs_ = (float(v) for v in triplet.split(':'))
        got = head_at(xs_, ys_)
        if abs(got - hs_) > tol:
            errs.append(f"({xs_:g},{ys_:g}): expected {hs_:.3f}, got {got:.3f}")
    if errs:
        return None, "head mismatch: " + "; ".join(errs)
    return 0.0, None


def run_tseep_head_test(test):
    """Check solved TRANSIENT total head at named points at a save time (the
    transient-seepage groundwater/GeoStudio corpus tags — the Pattern-B locks).

    This is the transient sibling of run_seep_head_test: it meshes and samples
    head at named (x,y) points the identical way (inverse-distance over the four
    nearest nodes, absolute head tolerance), but the field comes from a frame of
    a transient solve rather than a single steady solve.

    Tag keys: file (an .xlsx carrying a v18 ``tseep`` sheet), time (the save time
    t whose frame is sampled — it must land on the solver's save schedule, which
    the stepper guarantees for every ``save_times`` entry), points="x:y:h;...",
    tolerance (head units, absolute; default 0.01), optional target_size /
    element_type (mesh) and dt_max / max_head_change_frac / theta (stepper). The
    expected heads are LITERALS in the tag (locked-values law) — the runner never
    computes a reference. Pass/fail: returns 0.0 on success.

    Each row is one serial transient time-march (the stepper carries no internal
    parallelism), dispatched in the suite's per-row worker exactly like a
    seep_head row."""
    import numpy as np
    from xslope.fileio import load_slope_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_size_regions)
    from xslope.seep import (build_seep_data, build_tseep_data,
                             run_transient_seepage, transient_frame_index)

    slope_data = load_slope_data(test['file'])
    tseep_data = build_tseep_data(slope_data)
    if tseep_data is None:
        return None, "file carries no tseep sheet (not a transient model)"

    polygons = get_material_polygons(slope_data)
    target_size = test.get('target_size')
    if target_size is None:
        xs = [x for x, _ in slope_data['ground_surface'].coords]
        target_size = (max(xs) - min(xs)) / 120
    mesh = build_mesh_from_polygons(polygons, float(target_size),
                                    test.get('element_type', 'tri3'),
                                    size_regions=extract_size_regions(slope_data),
                                    **_refine_kwargs(test))
    seep_data = build_seep_data(mesh, slope_data)

    # Optional stepper knobs (analogous to seep_head's max_iter); default = the
    # solver's own defaults. verbose off so the suite stays quiet.
    kw = {'verbose': False}
    for key in ('dt_max', 'max_head_change_frac', 'theta'):
        if key in test and str(test[key]).strip() != '':
            kw[key] = float(test[key])
    solution = run_transient_seepage(seep_data, tseep_data, **kw)
    if not solution.get('converged', True):
        return None, "transient seepage solution did not converge"

    # Pull the frame at the requested save time. Saved frames are computed states
    # clamped to the schedule, so a save_times entry matches exactly;
    # transient_frame_index raises if the tag names a time off the schedule.
    t = float(test['time'])
    try:
        fi = transient_frame_index(solution, t)
    except ValueError as e:
        return None, str(e)
    h = np.asarray(solution['frames'][fi]['head'], dtype=float)

    nodes = seep_data['nodes']

    def head_at(xq, yq):
        d2 = (nodes[:, 0] - xq) ** 2 + (nodes[:, 1] - yq) ** 2
        idx = np.argsort(d2)[:4]
        w = 1.0 / np.maximum(d2[idx], 1e-12)
        return float(np.sum(w * h[idx]) / np.sum(w))

    tol = float(test.get('tolerance', 0.01))
    errs = []
    for triplet in str(test['points']).split(';'):
        xs_, ys_, hs_ = (float(v) for v in triplet.split(':'))
        got = head_at(xs_, ys_)
        if abs(got - hs_) > tol:
            errs.append(f"({xs_:g},{ys_:g}) @t={t:g}: expected {hs_:.3f}, got {got:.3f}")
    if errs:
        return None, "head mismatch: " + "; ".join(errs)
    return 0.0, None


def run_fs_vs_time_test(test):
    """Check the factor of safety at every saved instant of a transient march
    (the coupled FS-versus-time locks).

    The coupled question the seepage locks do not reach: the march produces a
    sequence of pore-pressure fields, and this runs the stability analysis against
    each in turn. It rebuilds the mesh and re-marches exactly as
    ``run_tseep_head_test`` does, so the field under the curve is the one that
    row locks, then hands the solution to ``xslope.sensitivity.fs_vs_time``.

    Tag keys: file, method (one LEM method), times="t1;t2;..." (the instants, in
    the file's time unit), expected="fs1;fs2;..." (one per time, LITERALS in the
    tag -- the runner never computes a reference), tolerance (default 0.005),
    optional critical_time / min_fs (the curve's own minimum, locked as such),
    num_slices, and the mesh/stepper knobs run_tseep_head_test takes
    (target_size, element_type, dt_max, max_head_change_frac, theta).

    A frame that produced no result is a failure of this row and is reported with
    the reason the mode recorded, never skipped: a curve missing its critical
    instant reads as a healthier slope than the model describes.

    Pass/fail: returns 0.0 on success. One row is one march plus one search per
    instant -- around a minute -- so it is dispatched like a tseep_head row."""
    from xslope.fileio import load_slope_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_size_regions)
    from xslope.seep import build_seep_data, build_tseep_data, run_transient_seepage
    from xslope.sensitivity import fs_vs_time

    slope_data = load_slope_data(test['file'])
    tseep_data = build_tseep_data(slope_data)
    if tseep_data is None:
        return None, "file carries no tseep sheet (not a transient model)"

    polygons = get_material_polygons(slope_data)
    target_size = test.get('target_size')
    if target_size is None:
        xs = [x for x, _ in slope_data['ground_surface'].coords]
        target_size = (max(xs) - min(xs)) / 120
    mesh = build_mesh_from_polygons(polygons, float(target_size),
                                    test.get('element_type', 'tri3'),
                                    size_regions=extract_size_regions(slope_data),
                                    **_refine_kwargs(test))
    slope_data['mesh'] = mesh                  # the u = 'seep' interpolation reads it
    seep_data = build_seep_data(mesh, slope_data)

    kw = {'verbose': False}
    for key in ('dt_max', 'max_head_change_frac', 'theta'):
        if key in test and str(test[key]).strip() != '':
            kw[key] = float(test[key])
    solution = run_transient_seepage(seep_data, tseep_data, **kw)
    if not solution.get('converged', True):
        return None, "transient seepage solution did not converge"

    times = [float(v) for v in str(test['times']).split(';')]
    expected = [float(v) for v in str(test['expected']).split(';')]
    if len(times) != len(expected):
        return None, (f"tag lists {len(times)} times but {len(expected)} expected "
                      f"factors of safety")

    ok, res = fs_vs_time(slope_data, solution, times=times,
                         methods=(test.get('method', 'spencer'),),
                         num_slices=int(test.get('num_slices', 40)))
    if not ok:
        return None, f"fs_vs_time refused the run: {res}"
    df = res['df']

    errs = []
    tol = float(test.get('tolerance', 0.005))
    for t, want in zip(times, expected):
        row = df.loc[(df['value'] - t).abs() < 1e-6]
        if row.empty:
            errs.append(f"t={t:g}: no row in the returned table")
            continue
        row = row.iloc[0]
        if not bool(row['success']):
            errs.append(f"t={t:g}: no result -- {row['msg']}")
            continue
        got = float(row['fs'])
        if abs(got - want) > tol:
            errs.append(f"t={t:g}: expected {want:.3f}, got {got:.3f}")
    if 'critical_time' in test:
        ct = res.get('critical_time')
        if ct is None or abs(float(ct) - float(test['critical_time'])) > 1e-6:
            errs.append(f"critical time: expected {float(test['critical_time']):g}, "
                        f"got {ct}")
    if 'min_fs' in test:
        mf = res.get('min_fs')
        if mf is None or abs(float(mf) - float(test['min_fs'])) > tol:
            errs.append(f"minimum FS: expected {float(test['min_fs']):.3f}, got {mf}")
    if errs:
        return None, "FS-vs-time mismatch: " + "; ".join(errs)
    return 0.0, None


def run_seep_elements_test(test):
    """Element-type coverage: solve ONE seepage problem with every supported
    element type (tri3, tri6, quad4, quad8, quad9) and check each converges and
    reproduces the expected flowrate within tolerance. The same physical problem
    must give the same flowrate on every element type, so a single expected value
    covers all — a change to the seepage solver or mesh that breaks (or shifts) an
    element type is caught. Returns (0.0, None) if all pass, else (None, message
    naming the failing types). element_types= overrides the default set."""
    import io
    import contextlib
    from xslope.fileio import load_slope_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_size_regions)
    from xslope.seep import build_seep_data, run_seepage_analysis

    slope_data = load_slope_data(test['file'])
    polygons = get_material_polygons(slope_data)
    element_types = [s.strip() for s in
                     test.get('element_types', 'tri3,tri6,quad4,quad8,quad9').split(',')]
    target_size = test.get('target_size')
    if target_size is None:
        x_coords = [x for x, _ in slope_data['ground_surface'].coords]
        target_size = (max(x_coords) - min(x_coords)) / 120
    expected = test['expected_flowrate']
    tol = test.get('tolerance', 0.05) * abs(expected)

    problems = []
    for et in element_types:
        try:
            with contextlib.redirect_stdout(io.StringIO()):
                mesh = build_mesh_from_polygons(polygons, target_size, et)
                seep_data = build_seep_data(mesh, slope_data)
                solution = run_seepage_analysis(seep_data, tol=1e-4)
        except Exception as e:
            problems.append(f"{et}: {type(e).__name__}: {e}")
            continue
        if not solution.get('converged', True):
            problems.append(f"{et}: did not converge")
        elif abs(solution['flowrate'] - expected) > tol:
            problems.append(f"{et}: flowrate {solution['flowrate']:.3f} vs {expected:.3f}")
    if problems:
        return None, "; ".join(problems)
    return 0.0, None


def run_fem_elements_test(test):
    """Element-type coverage: solve ONE FEM (SSRM) problem with every QUADRATIC
    element type (tri6, quad8, quad9) and check each converges to the expected
    factor of safety within tolerance. FEM uses quadratic elements only (linear
    tri3/quad4 lock and overestimate FS). Returns (0.0, None) if all pass, else
    (None, message naming the failing types). element_types= overrides the set."""
    import io
    import contextlib
    from xslope.fileio import load_slope_data
    from xslope.fem import build_fem_data, solve_ssrm
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_constraint_line_geometry, extract_size_regions)

    slope_data = load_slope_data(test['file'])
    element_types = [s.strip() for s in
                     test.get('element_types', 'tri6,quad8,quad9').split(',')]
    target_size = test.get('target_size')
    if target_size is None:
        x_coords = [x for x, _ in slope_data['ground_surface'].coords]
        target_size = (max(x_coords) - min(x_coords)) / 100
    expected = test['expected_fs']
    tol = test.get('tolerance', 0.03)
    kwargs = {}
    if 'max_iter' in test:
        kwargs['max_iterations'] = int(test['max_iter'])
    # ssrm_tol is the bisection precision (how tightly FS is bracketed); keep it
    # well below the cross-element comparison tolerance.
    ssrm_tol = float(test.get('ssrm_tol', 0.01))

    constraint_lines, _n_reinf, _n_pile = extract_constraint_line_geometry(slope_data)
    polygons = get_material_polygons(slope_data, reinf_lines=constraint_lines)
    problems = []
    for et in element_types:
        try:
            with contextlib.redirect_stdout(io.StringIO()):
                mesh = build_mesh_from_polygons(
                    polygons, target_size=target_size, element_type=et,
                    lines=constraint_lines,
                    size_regions=extract_size_regions(slope_data))
                fem_data = build_fem_data(slope_data, mesh)
                result = solve_ssrm(fem_data, F_min=test.get('f_min', 0.5),
                                    F_max=test.get('f_max', 3.0), tolerance=ssrm_tol,
                                    debug_level=0, **kwargs)
        except Exception as e:
            problems.append(f"{et}: {type(e).__name__}: {e}")
            continue
        if not result.get('converged', False):
            problems.append(f"{et}: SSRM failed ({result.get('error', '?')})")
        elif abs(result['FS'] - expected) > tol:
            problems.append(f"{et}: FS {result['FS']:.3f} vs {expected:.3f}")
    if problems:
        return None, "; ".join(problems)
    return 0.0, None


def run_sensitivity_test(test):
    """Run a sensitivity-sweep regression (docs/parametric/sensitivity.md tags).

    Tag keys: file, param, method, num_slices, n, rel_range, fs_base, fs_low,
    fs_high, tolerance. Runs sensitivity() with a searched sweep and checks
    the base row and the two range endpoints. Returns (fs_base, None) on pass
    so the table shows the base FS; a mismatch anywhere returns an error."""
    import io, contextlib
    from xslope.fileio import load_slope_data
    from xslope.sensitivity import sensitivity

    slope_data = load_slope_data(test['file'])
    tol = float(test.get('tolerance', 0.01))
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        ok, res = sensitivity(slope_data, param=test['param'],
                              rel_range=float(test.get('rel_range', 0.5)),
                              n=int(test.get('n', 3)),
                              methods=(test['method'],),
                              search=str(test.get('search', 'true')).lower()
                                     not in ('false', '0', 'no'),
                              num_slices=int(test.get('num_slices', 30)))
    if not ok:
        return None, f"sensitivity failed: {res}"
    df = res['df']
    swept = df.loc[~df['is_base']].sort_values('value')
    checks = [('expected_base', float(df.loc[df['is_base'], 'fs'].iloc[0])),
              ('expected_low', float(swept['fs'].iloc[0])),
              ('expected_high', float(swept['fs'].iloc[-1]))]
    if not bool(swept['success'].all()):
        return None, f"sweep had failed points: {swept.loc[~swept['success'], 'msg'].tolist()}"
    for key, got in checks:
        if key in test and abs(got - float(test[key])) > tol:
            return None, f"{key}: expected {float(test[key]):.3f}, got {got:.3f}"
    return checks[0][1], None


def run_design_callable_test(test):
    """Guard the design()/back_analysis() ``modify=`` callable path that was brought
    to the same exclusive contract as sensitivity() (docs/parametric/design.md).

    Two deterministic checks, file-less beyond loading the shipped ACADS sample:

    (a) Hand-computable crossing. On a FIXED circle (search=False) solved with the
        Ordinary Method of Slices, FS is EXACTLY linear in cohesion c — the OMS
        numerator's cohesion term is c·Σ(ΔL) and the denominator is c-independent —
        so a ``modify=`` callable that sets c and asks where FS = 1.35 has a crossing
        the linear closed form pins to machine precision. Asserts design() reports
        that locked crossing, that it equals the closed-form value, and bracketed.

    (b) param-vs-equivalent-callable equivalence. The identical sweep expressed as
        ``param='mat:Soil:c'`` and as a ``modify=`` callable setting the same field
        must be byte-identical — same crossing, same fs_range, same per-point FS —
        proving both travel the one shared ``_resolve_sweep_spec`` setter path. Also
        confirms back_analysis() inherits the callable path (it wraps design()).

    Returns (0.0, None) on success, else (None, message)."""
    import io
    import contextlib
    import numpy as np
    from xslope.fileio import load_slope_data
    from xslope.sensitivity import design, back_analysis

    sd = load_slope_data(_repo('docs/lem/files/xslope_acads_simple.xlsx'))

    def set_c(s, val):
        # modify= receives an already-copied slope_data (materials are deep-copied
        # by the engine before the setter runs), so mutate-and-return is correct and
        # matches exactly what param='mat:Soil:c' does.
        s['materials'][0]['c'] = val
        return s

    kw = dict(low=1.0, high=9.0, steps=5, target_fs=1.35, method='oms', search=False)
    with contextlib.redirect_stdout(io.StringIO()):
        okp, rp = design(sd, param='mat:Soil:c', **kw)
        okm, rm = design(sd, modify=set_c, label='cohesion c', **kw)
    if not okp:
        return None, f"design(param=) failed: {rp}"
    if not okm:
        return None, f"design(modify=) failed: {rm}"

    problems = []
    expected = 4.6351587374          # locked OMS/fixed-surface crossing of FS = 1.35

    # (a) hand-computable crossing on the callable path
    if not rm['bracketed']:
        problems.append("modify sweep did not bracket FS = 1.35")
    if rm['crossing'] is None or abs(rm['crossing'] - expected) > 1e-6:
        problems.append(f"modify crossing {rm['crossing']} != expected {expected}")
    sw = rm['df'].loc[~rm['df']['is_base']].sort_values('value')
    v = sw['value'].to_numpy(dtype=float)
    f = sw['fs'].to_numpy(dtype=float)
    B = (f[-1] - f[0]) / (v[-1] - v[0])
    A = f[0] - B * v[0]
    if np.max(np.abs(f - (A + B * v))) > 1e-9:
        problems.append("OMS fixed-surface FS is not linear in c (test premise broken)")
    c_closed = (1.35 - A) / B
    if rm['crossing'] is not None and abs(rm['crossing'] - c_closed) > 1e-9:
        problems.append(f"crossing {rm['crossing']} != closed-form {c_closed}")

    # (b) param == callable, byte-for-byte (the shared setter path)
    if rp['crossing'] is None or abs(rp['crossing'] - rm['crossing']) > 1e-9:
        problems.append(f"param crossing {rp['crossing']} != modify crossing {rm['crossing']}")
    if rp['bracketed'] != rm['bracketed'] or rp['fs_range'] != rm['fs_range']:
        problems.append("param/modify bracketed or fs_range differ")
    fp = rp['df'].loc[~rp['df']['is_base']].sort_values('value')['fs'].to_numpy(dtype=float)
    fm = rm['df'].loc[~rm['df']['is_base']].sort_values('value')['fs'].to_numpy(dtype=float)
    if fp.shape != fm.shape or np.max(np.abs(fp - fm)) > 1e-12:
        problems.append("param/modify per-point FS differ")

    # back_analysis() inherits the callable path (thin wrapper over design())
    with contextlib.redirect_stdout(io.StringIO()):
        okb, rb = back_analysis(sd, modify=set_c, label='cohesion c',
                                low=1.0, high=9.0, steps=5, target_fs=1.35,
                                method='oms', search=False)
    if not okb or rb.get('study') != 'back_analysis':
        problems.append("back_analysis did not inherit modify= / the study flag")
    elif rb['crossing'] is None or abs(rb['crossing'] - expected) > 1e-6:
        problems.append(f"back_analysis crossing {rb.get('crossing')} != {expected}")

    if problems:
        return None, "; ".join(problems[:6])
    return 0.0, None


def run_reliability_test(test):
    """Run a single reliability analysis, returning the lognormal reliability index beta."""
    from xslope.fileio import load_slope_data
    from xslope.advanced import reliability as reliability_analysis

    file_path = test['file']
    method = test.get('method', 'spencer')
    circular = str(test.get('circular', 'true')).lower() not in ('false', '0', 'no')

    do_search = str(test.get('search', 'true')).lower() not in ('false', '0', 'no')

    slope_data = load_slope_data(file_path)
    success, result = reliability_analysis(slope_data, method, circular=circular, debug_level=0,
                                           search=do_search)
    if not success:
        return None, f"reliability failed: {result}"
    return result['beta_ln'], None


def run_reliability_mc_test(test):
    """Monte Carlo reliability regression: sample the material sigmas, evaluate the
    factor of safety of each realization on the fixed surface, and return the
    lognormal reliability index beta_ln (compared to expected_beta). The seed is a
    fixed constant, so a given (file, n_samples, seed) is bit-reproducible. An
    optional expected_pf/pf_tol pair additionally checks the empirical probability
    of failure."""
    from xslope.fileio import load_slope_data
    from xslope.advanced import reliability_mc

    file_path = test['file']
    method = test.get('method', 'spencer')
    circular = str(test.get('circular', 'true')).lower() not in ('false', '0', 'no')
    do_search = str(test.get('search', 'true')).lower() not in ('false', '0', 'no')
    n_samples = int(test.get('n_samples', 10000))
    distribution = test.get('distribution', 'normal')
    num_slices = int(test.get('num_slices', 40))
    composite = str(test.get('composite', 'false')).lower() in ('true', '1', 'yes')
    kw = {}
    if 'rng_seed' in test:
        kw['rng_seed'] = int(test['rng_seed'])

    slope_data = load_slope_data(file_path)
    success, result = reliability_mc(slope_data, method, circular=circular, search=do_search,
                                     n_samples=n_samples, distribution=distribution,
                                     num_slices=num_slices, composite=composite,
                                     debug_level=-1, **kw)
    if not success:
        return None, f"reliability_mc failed: {result}"

    exp_pf = test.get('expected_pf')
    if exp_pf is not None:
        pf_tol = float(test.get('pf_tol', 0.02))
        if abs(result['pf_empirical'] - float(exp_pf)) > pf_tol:
            return None, (f"pf_empirical {result['pf_empirical']:.4f} vs expected "
                          f"{float(exp_pf):.4f} (tol {pf_tol})")
    return result['beta_ln'], None


def run_fem_reliability_test(test):
    """FEM reliability regression: run reliability_fem on a mesh and return the
    lognormal reliability index beta_ln (compared to expected_beta). Guards the
    whole TSPM-over-SSRM pipeline — F_MLV, the F+/F- perturbation solves, and the
    beta combination — on a coarse mesh so it stays affordable. ssrm_tol sets the
    bisection precision; the tag's `tolerance` is the beta comparison tolerance."""
    import io
    import contextlib
    from xslope.fileio import load_slope_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_constraint_line_geometry, extract_size_regions)
    from xslope.advanced import reliability_fem

    slope_data = load_slope_data(test['file'])
    element_type = test.get('element_type', 'tri6')
    target_size = test.get('target_size')
    if target_size is None:
        x_coords = [x for x, _ in slope_data['ground_surface'].coords]
        target_size = (max(x_coords) - min(x_coords)) / 100
    constraint_lines, _n_reinf, _n_pile = extract_constraint_line_geometry(slope_data)
    polygons = get_material_polygons(slope_data, reinf_lines=constraint_lines)
    mesh = build_mesh_from_polygons(polygons, target_size=target_size,
                                    element_type=element_type, lines=constraint_lines,
                                    size_regions=extract_size_regions(slope_data))
    with contextlib.redirect_stdout(io.StringIO()):
        success, result = reliability_fem(
            slope_data, mesh=mesh, F_min=test.get('f_min', 0.5),
            F_max=test.get('f_max', 3.0), tolerance=float(test.get('ssrm_tol', 0.001)),
            failure_criterion=test.get('criterion', 'non_convergence'))
    if not success:
        return None, f"reliability_fem failed: {result}"
    return result['beta_ln'], None


def run_gsat_pair_test(test):
    """The gamma_sat sidecar equivalence gate (plan §1.2 S1): a model hand-zoned
    into moist/saturated polygons at the water table must give the same FS as the
    single-material gamma/gamma_sat + piezo-sidecar formulation. Discretization
    differs (the two force different slice boundaries), so the comparison runs at
    300 slices where both converge to the same answer (verified 5e-8 at build
    time); tolerance 1e-5."""
    from xslope.fileio import load_slope_data
    from xslope.slice import generate_slices
    from xslope import solve

    vals = []
    for f in (test['file'], test['file2']):
        sd = load_slope_data(f)
        ok, result = generate_slices(sd, circle=sd['circles'][0], num_slices=300)
        if not ok:
            return None, f"slice generation failed for {f}: {result}"
        df, _ = result
        s_ok, r = solve.spencer(df)
        if not s_ok:
            return None, f"spencer failed for {f}: {r}"
        vals.append(r['FS'])
    d = abs(vals[0] - vals[1])
    if d > 1e-5:
        return None, (f"zoned vs sidecar FS differ by {d:.2e} "
                      f"({vals[0]:.6f} vs {vals[1]:.6f})")
    return 0.0, None    # pass/fail test: 0.0 = the two formulations agree


def run_roundtrip_test(test):
    """Verify save_slope_data_to_xlsx round-trips a file: load -> save into a
    blank template -> reload must reproduce every input category.

    The source is a fixture SYNTHESIZED at ``test['legacy_version']`` from the
    archived template of that version (see ``provision_legacy_file``), not the
    checked-in file — so what is being proved is that a genuinely old file
    survives the current writer, and that none of the current template's
    pre-filled cells leak into it on the way. When ``legacy_version`` is absent
    the file itself is used, which is how the DXF and editor paths call this.

    Returns (0.0, None) on success, or (None, message) listing the mismatches.
    """
    import tempfile
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx

    with tempfile.TemporaryDirectory() as td:
        src = test['file']
        version = test.get('legacy_version')
        if version is not None:
            src = provision_legacy_file(
                src, version,
                os.path.join(td, f"v{version}_{os.path.basename(test['file'])}"))
            got = load_slope_data(src).get('template_version')
            if got is not None and int(float(got)) != int(version):
                return None, (f"fixture provisioning wrote template version {got}, "
                              f"expected v{version}")
        d1 = load_slope_data(src)
        tmp = os.path.join(td, 'roundtrip.xlsx')
        save_slope_data_to_xlsx(d1, tmp,
                                template=test.get('template', ROUNDTRIP_TEMPLATE))
        d2 = load_slope_data(tmp)

    mismatches = []
    for k in ROUNDTRIP_KEYS:
        mismatches += _roundtrip_diff(d1.get(k), d2.get(k), k)

    # Polygon-sourced geometry has no profile_lines to compare, so check the
    # reconstructed polygon zones directly (geometry + material assignment).
    if not d1.get('profile_lines'):
        p1 = d1.get('polygons') or []
        p2 = d2.get('polygons') or []
        if len(p1) != len(p2):
            mismatches.append(f"polygons: len {len(p1)} vs {len(p2)}")
        else:
            for i, (a, b) in enumerate(zip(p1, p2)):
                if a.get('mat_id') != b.get('mat_id'):
                    mismatches.append(f"polygons[{i}].mat_id")
                if not a['polygon'].equals(b['polygon']):
                    mismatches.append(f"polygons[{i}].geom")

    if mismatches:
        return None, "round-trip mismatch: " + "; ".join(mismatches[:5])
    return 0.0, None


def run_v19_roundtrip_test(test):
    """Verify every v19 file-carried run option survives save -> load.

    Sets all eight main-sheet run options (D14:D21), the ten circles-sheet search
    window limits (K8:K17) to distinct non-default values, writes the model through
    the current template, reloads it, and compares.
    A field the writer forgets, or one the loader reads from the wrong cell, fails
    here — the corpus round-trips can only see that the fields stay ABSENT.
    """
    import tempfile
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx

    template = test.get('template', ROUNDTRIP_TEMPLATE)
    d1 = load_slope_data(test['file'])
    d1.update(V19_ROUNDTRIP_VALUES)
    d1['search_window'] = dict(V19_SEARCH_WINDOW)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    try:
        save_slope_data_to_xlsx(d1, tmp, template=template)
        d2 = load_slope_data(tmp)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)

    problems = []
    for k, v in V19_ROUNDTRIP_VALUES.items():
        got = d2.get(k)
        # tension_srf is a tri-state: False must come back False, never None.
        if k == 'tension_srf':
            if got is not v:
                problems.append(f"{k}: {got!r} != {v!r}")
        elif not _roundtrip_eq(v, got):
            problems.append(f"{k}: {got!r} != {v!r}")
    problems += _roundtrip_diff(V19_SEARCH_WINDOW, d2.get('search_window') or {},
                                'search_window')
    if problems:
        return None, "v19 round-trip mismatch: " + "; ".join(problems[:6])
    return 0.0, None


def run_surface_family_roundtrip_test(test):
    """Verify the v22 surface-family cell (main D24) survives save -> load.

    The rare deck defines BOTH a circular and a non-circular surface, and nothing in
    its geometry says which one it means — the circles simply win. The cell is where
    that answer is stored, and three things have to hold for storing it to be worth
    anything:

      1. **blank stays blank.** It is the normal state (every single-family model in
         the corpus), and a writer that invented a family here would make every file
         it touched claim an answer nobody gave.
      2. **both values come back**, in the model's own vocabulary, and the stated
         family drives the ``circular`` flag the runners and the plots read — which
         is what makes the selection outlive the session that made it. Drop the
         writer's emit and this is the check that fails.
      3. **a single-family deck is unaffected**: the cell stays blank through a
         save, and a value left in it never claims a surface the deck does not
         define (the flag still follows presence).
    """
    import tempfile
    import openpyxl
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx

    template = test.get('template', ROUNDTRIP_TEMPLATE)
    problems = []

    def _cycle(mutate):
        d1 = load_slope_data(test['file'])
        mutate(d1)
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        try:
            save_slope_data_to_xlsx(d1, tmp, template=template)
            wb = openpyxl.load_workbook(tmp)
            cell = wb['main']['D24'].value
            wb.close()
            return load_slope_data(tmp), cell
        finally:
            if os.path.exists(tmp):
                os.unlink(tmp)

    def _both(d, family):
        """Make a both-family deck out of a circles model, and state ``family``."""
        d['non_circ'] = list(SURFACE_FAMILY_NONCIRC)
        d['surface_family'] = family

    # (1) A model that states nothing writes a blank cell and reads back None.
    d2, cell = _cycle(lambda d: _both(d, None))
    if cell is not None:
        problems.append(f"blank: the writer put {cell!r} in D24 instead of a blank cell")
    if d2.get('surface_family') is not None:
        problems.append(f"blank: came back as {d2.get('surface_family')!r}, not None")
    if not d2.get('circular'):
        problems.append("blank: a both-family deck stating nothing lost the circles "
                        "(the automatic answer is the circular family)")

    # (2) Both values survive, and the stated family drives the flag every consumer
    #     outside a dialog reads.
    for family, want_cell, want_flag in (('circular', 'circular', True),
                                         ('noncircular', 'non-circular', False)):
        d2, cell = _cycle(lambda d, f=family: _both(d, f))
        if cell != want_cell:
            problems.append(f"{family}: D24 holds {cell!r}, expected {want_cell!r}")
        if d2.get('surface_family') != family:
            problems.append(f"{family}: came back as {d2.get('surface_family')!r}")
        if bool(d2.get('circular')) != want_flag:
            problems.append(f"{family}: the circular flag came back "
                            f"{d2.get('circular')!r}, expected {want_flag!r}")
        if len(d2.get('circles') or []) != 1 or len(d2.get('non_circ') or []) != 3:
            problems.append(f"{family}: the deck lost a surface "
                            f"({len(d2.get('circles') or [])} circle(s), "
                            f"{len(d2.get('non_circ') or [])} non-circ point(s))")

    # (3) A single-family deck is untouched: nothing states a family, the cell stays
    #     blank, and a value that somehow got in cannot claim the absent surface.
    d2, cell = _cycle(lambda d: None)
    if cell is not None or d2.get('surface_family') is not None:
        problems.append(f"single-family: D24 = {cell!r}, model = "
                        f"{d2.get('surface_family')!r} — both should be blank/None")
    d2, cell = _cycle(lambda d: d.update(surface_family='noncircular'))
    if not d2.get('circular'):
        problems.append("single-family: a stated non-circular family turned off the "
                        "circles on a deck that defines no non-circular surface")

    # (4) A word the cell cannot hold is REFUSED on the way out rather than coerced,
    #     the same contract as the water-load mode beside it.
    try:
        _cycle(lambda d: d.update(surface_family='spiral'))
    except ValueError as exc:
        if 'D24' not in str(exc):
            problems.append(f"an unknown family raised without naming the cell: {exc}")
    else:
        problems.append("an unknown surface family was written without an error")

    if problems:
        return None, "surface-family round-trip: " + "; ".join(problems[:6])
    return 0.0, None


def run_ssr_zone_roundtrip_test(test):
    """Verify the v20 SSR zone overlays survive save -> load, all three sentinels.

    Writes one polygon row of EACH kind (-1 "SSR reduce", -2 "SSR hold", -3 "SSR
    elastic") onto a real polygon-geometry model, saves it through the current
    template, reloads and compares. Three things have to hold, and each has its own
    failure mode:

      1. every zone comes back with its kind and its ring intact (a writer that drops
         the sentinel would silently turn a zone into material 0 or lose it);
      2. the zones stay OUT of slope_data['polygons'] — they are analysis overlays,
         and a zone that leaked into the geometry would be meshed and sliced;
      3. the material zones are untouched in count and area.

    Also checks that an unrecognized negative Mat ID is REFUSED rather than dropped —
    a mis-typed -4 that vanished would run the SSRM unconstrained.
    """
    import tempfile
    from xslope.fileio import (SSR_ZONE_LABELS as _SSR_LABELS, load_slope_data,
                              save_slope_data_to_xlsx)

    template = test.get('template', ROUNDTRIP_TEMPLATE)
    d1 = load_slope_data(test['file'])
    n_polys = len(d1.get('polygons') or [])
    areas = [round(p['polygon'].area, 6) for p in (d1.get('polygons') or [])]
    x0, y0, x1, y1 = d1['domain_polygon'].bounds
    dx, dy = (x1 - x0), (y1 - y0)

    def box(fx0, fy0, fx1, fy1):
        return [(x0 + fx0 * dx, y0 + fy0 * dy), (x0 + fx1 * dx, y0 + fy0 * dy),
                (x0 + fx1 * dx, y0 + fy1 * dy), (x0 + fx0 * dx, y0 + fy1 * dy)]

    zones = [
        {'kind': 'reduce', 'polygon': box(0.10, 0.05, 0.90, 0.95)},
        {'kind': 'hold', 'polygon': box(0.15, 0.10, 0.35, 0.50)},
        {'kind': 'hold_elastic', 'polygon': box(0.60, 0.10, 0.80, 0.50)},
    ]
    d1['ssr_zones'] = [dict(z) for z in zones]

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    try:
        save_slope_data_to_xlsx(d1, tmp, template=template)
        d2 = load_slope_data(tmp)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)

    problems = []
    got = d2.get('ssr_zones') or []
    if len(got) != len(zones):
        problems.append(f"ssr_zones: {len(got)} zone(s) back, expected {len(zones)}")
    else:
        for i, (want, have) in enumerate(zip(zones, got)):
            if have.get('kind') != want['kind']:
                problems.append(f"ssr_zones[{i}].kind: {have.get('kind')!r} != "
                                f"{want['kind']!r}")
            ring_w = [(round(x, 6), round(y, 6)) for x, y in want['polygon']]
            ring_h = [(round(x, 6), round(y, 6)) for x, y in have.get('polygon') or []]
            if ring_w != ring_h:
                problems.append(f"ssr_zones[{i}].polygon: {ring_h} != {ring_w}")
            if have.get('label') != _SSR_LABELS[want['kind']]:
                problems.append(f"ssr_zones[{i}].label: {have.get('label')!r}")
    got_polys = d2.get('polygons') or []
    if len(got_polys) != n_polys:
        problems.append(f"polygons: {len(got_polys)} back, expected {n_polys} — a "
                        "zone leaked into the material geometry")
    elif [round(p['polygon'].area, 6) for p in got_polys] != areas:
        problems.append("polygons: material-zone areas changed")

    # A zone code the reader does not recognize must RAISE, not be silently dropped.
    # Both vocabularies are checked on the SAME v21 file, because both readers stay
    # live: an unknown v21 Type WORD in the Type row (5), and an unknown v20 SENTINEL
    # in the Mat ID row (6) of a block whose Type is blank.
    from benchmarks._xlsx_writer import cell_ref, write_cells_to_xlsx
    from xslope.fileio import _read_template_info

    _dest_ver = _read_template_info(template)[0] if os.path.exists(template) else 0
    _type_row, _matid_row = (5, 6) if _dest_ver >= 21 else (None, 5)
    for _row, _bad, _want in ((_type_row, 'ssr reduc', 'not a recognized polygon type'),
                              (_matid_row, -4, 'SSR zone overlay')):
        if _row is None:
            continue
        d3 = load_slope_data(test['file'])
        d3['ssr_zones'] = [{'kind': 'reduce', 'polygon': box(0.1, 0.1, 0.9, 0.9)}]
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        try:
            save_slope_data_to_xlsx(d3, tmp, template=template)
            blk = len(d3.get('polygons') or []) + 1
            _cells = {cell_ref(_row, 2 + (blk - 1) * 3): _bad}
            if _row == _matid_row and _type_row is not None:
                _cells[cell_ref(_type_row, 2 + (blk - 1) * 3)] = None  # blank the Type
            write_cells_to_xlsx(tmp, {'polygon': _cells})
            try:
                load_slope_data(tmp)
                problems.append(f"an unknown zone code ({_bad!r}) loaded without error")
            except ValueError as exc:
                if _want not in str(exc):
                    problems.append(
                        f"{_bad!r} raised, but without naming the vocabulary: {exc}")
        finally:
            if os.path.exists(tmp):
                os.unlink(tmp)

    if problems:
        return None, "SSR zone round-trip mismatch: " + "; ".join(problems[:6])
    return 0.0, None


def run_v21_roundtrip_test(test):
    """Verify every v21 file-carried input survives save -> load.

    The v21 additions are spread over four sheets and none of them exists on any
    corpus file, so this is the only place they are exercised end to end:

      * polygon sheet — a Size on a material zone, a Size on an SSR overlay, and a
        Type='refine' overlay (which is neither geometry nor an SSR constraint and
        must come back on its own list);
      * profile sheet — a Size on a profile line (checked on a profile-geometry
        model, since the two geometry forms take different writer branches);
      * dloads / dloads (2) — a per-block Direction, with a 'normal' block beside a
        'vertical' one so a writer that wrote one direction for the whole sheet fails;
      * main sheet — the Side BC selector.

    It also carries the tseep save_times column, whose header row moved from J10 to
    J11 at v22. Every corpus tseep file is v18-v21, so only a write into the CURRENT
    master exercises the newer anchor — and a save_times list written one row off its
    header reads back short (or empty) with nothing else to show for it.
    """
    import tempfile
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx

    template = test.get('template', ROUNDTRIP_TEMPLATE)
    problems = []

    def _roundtrip(path, mutate):
        d1 = load_slope_data(path)
        mutate(d1)
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        try:
            save_slope_data_to_xlsx(d1, tmp, template=template)
            return d1, load_slope_data(tmp)
        finally:
            if os.path.exists(tmp):
                os.unlink(tmp)

    # --- polygon geometry: zone Size, overlay Size, refine overlay, Side BC ---
    base = test['file']
    x0, y0, x1, y1 = load_slope_data(base)['domain_polygon'].bounds
    dx, dy = (x1 - x0), (y1 - y0)

    def box(fx0, fy0, fx1, fy1):
        return [(x0 + fx0 * dx, y0 + fy0 * dy), (x0 + fx1 * dx, y0 + fy0 * dy),
                (x0 + fx1 * dx, y0 + fy1 * dy), (x0 + fx0 * dx, y0 + fy1 * dy)]

    refine_ring = box(0.30, 0.20, 0.55, 0.60)

    def _mut_poly(d):
        d['polygons'][0]['size'] = 1.75
        d['ssr_zones'] = [{'kind': 'reduce', 'polygon': box(0.1, 0.05, 0.9, 0.95),
                           'size': 0.9}]
        d['refine_zones'] = [{'polygon': list(refine_ring), 'size': 0.45}]
        d['side_bc'] = 'fixed'

    d1, d2 = _roundtrip(base, _mut_poly)
    if not _roundtrip_eq(1.75, (d2['polygons'][0] or {}).get('size')):
        problems.append(f"polygon size: {(d2['polygons'][0] or {}).get('size')!r} != 1.75")
    _z = (d2.get('ssr_zones') or [{}])[0]
    if not _roundtrip_eq(0.9, _z.get('size')):
        problems.append(f"ssr zone size: {_z.get('size')!r} != 0.9")
    if _z.get('kind') != 'reduce':
        problems.append(f"ssr zone kind: {_z.get('kind')!r} != 'reduce'")
    _r = d2.get('refine_zones') or []
    if len(_r) != 1:
        problems.append(f"refine_zones: {len(_r)} back, expected 1")
    else:
        if not _roundtrip_eq(0.45, _r[0].get('size')):
            problems.append(f"refine size: {_r[0].get('size')!r} != 0.45")
        problems += _roundtrip_diff([tuple(p) for p in refine_ring],
                                    [tuple(p) for p in _r[0]['polygon']], 'refine ring')
    if len(d2.get('polygons') or []) != len(d1.get('polygons') or []):
        problems.append("a refine overlay leaked into the material geometry")
    if d2.get('side_bc') != 'fixed':
        problems.append(f"side_bc: {d2.get('side_bc')!r} != 'fixed'")

    # A refine polygon with no Size is a no-op the user could not see — refused.
    try:
        _roundtrip(base, lambda d: d.update(
            refine_zones=[{'polygon': list(refine_ring), 'size': None}]))
        problems.append("a refine polygon with no Size was accepted")
    except ValueError:
        pass

    # --- profile geometry: a per-line Size, and dload Directions ---
    prof_file = test.get('profile_file')
    if not prof_file or not Path(prof_file).exists():
        # A missing base must FAIL, not silently halve the test's coverage.
        problems.append(f"profile base not found: {prof_file!r}")
    else:
        def _mut_prof(d):
            d['profile_lines'][0]['size'] = 2.25
            for i, _dir in enumerate(('vertical', 'normal')):
                if i < len(d.get('dload_dirs') or []):
                    d['dload_dirs'][i] = _dir
            if d.get('dload2_dirs'):
                d['dload2_dirs'][0] = 'vertical'
        d1, d2 = _roundtrip(prof_file, _mut_prof)
        if not _roundtrip_eq(2.25, d2['profile_lines'][0].get('size')):
            problems.append(f"profile size: {d2['profile_lines'][0].get('size')!r} != 2.25")
        for key in ('dload_dirs', 'dload2_dirs'):
            problems += _roundtrip_diff(d1.get(key) or [], d2.get(key) or [], key)
        # Geometry must be untouched by the row shift the Size inserted.
        problems += _roundtrip_diff(
            [ln['coords'] for ln in d1['profile_lines']],
            [ln['coords'] for ln in d2['profile_lines']], 'profile coords')
        problems += _roundtrip_diff(d1.get('dloads') or [], d2.get('dloads') or [],
                                    'dloads')

    # --- tseep: the whole block, and save_times in particular ---
    ts_file = test.get('tseep_file')
    if not ts_file or not Path(ts_file).exists():
        problems.append(f"tseep base not found: {ts_file!r}")
    else:
        d1, d2 = _roundtrip(ts_file, lambda d: None)
        t1, t2 = d1.get('tseep') or {}, d2.get('tseep') or {}
        if not t1.get('save_times'):
            problems.append("tseep base carries no save_times to round-trip")
        problems += _roundtrip_diff(t1, t2, 'tseep')

    if problems:
        return None, "v21 round-trip mismatch: " + "; ".join(problems[:6])
    return 0.0, None


# === Editor round-trip guard (studio.editors) ===
# Opening each CategoryEditor on a fully-populated record and applying with NO
# changes must leave its inputs deep-equal — the editor module's contract that "a
# round-trip through an editor never drops data". This is the guard that would
# have caught the materials-combo corruption: a Field whose `choices` lagged the
# loader's accepted set left the combo unselectable for a valid value, so OK
# silently rewrote option/u/unsat to the combo's first item. The Excel --roundtrip
# test never routes through the editors, so it can't see this. Needs PySide6 (the
# studio layer), so it's registered only when that imports — engine-only installs
# skip it cleanly, like the DXF tests.

# slope_data keys each editor round-trips (its "managed" inputs). Keys the editors
# recompute from these (ground_surface, domain_polygon, tcrack_surface,
# reinforce_lines, circular, has_seepage_bc2, ...) are derived, not source, so are
# excluded — only the source inputs must survive a no-op round-trip.
_EDITOR_MANAGED_KEYS = {
    "global": ["gamma_water", "tcrack_depth", "tcrack_water", "k_seismic",
               "unit_system", "time_unit",
               # v19 main-sheet globals surfaced in the Global parameters dialog.
               "k0", "tension_srf"],
    "materials": ["materials"],
    "circles": ["circles"],
    "non_circ": ["non_circ"],
    "piezo": ["piezo_line", "piezo_line2"],
    # The dloads editor also owns the v21 per-block Directions. They are stored as
    # lists PARALLEL to the block lists, which is exactly the shape that goes wrong
    # when a block is deleted, so they are managed keys and the round-trip covers
    # them.
    "dloads": ["dloads", "dloads2", "dload_dirs", "dload2_dirs"],
    "seep_bc": ["seepage_bc", "seepage_bc2"],
    "piles": ["pile_lines"],
    "reinforce": ["reinforcement_lines"],
    "line_loads": ["line_loads"],
    "profile": ["profile_lines"],
    # The polygon editor also owns the polygon sheet's overlay rows — SSR zones and
    # v21 refine regions. They are edited in the same dialog as the material zones
    # and split back out by Type on apply, so a dropped or mis-typed overlay has to
    # fail here.
    "polygons": ["polygons", "ssr_zones", "refine_zones"],
    "transient": ["tseep"],
}


def _editor_full_material(name, option, u, unsat, t_cut=4.0, phi_b=15.0, s_cap=30.0):
    """A material with EVERY loader-produced key set to a distinct non-default
    value, so a dropped key is caught; option/u/unsat carry the enum value under
    test. Together the fixture's rows exercise every accepted option (mc/cp/pow/hb/
    elastic), u (none/piezo/seep/ru) and unsat (lf/vg/gard) value, the v16 t_cut
    column, and the v17 matric-suction pair phi_b/s_cap (distinct non-None values,
    so a dropped key is caught; the elastic row carries None for all three)."""
    return {
        "name": name, "gamma": 120.0, "gamma_sat": 125.0, "option": option,
        "c": 100.0, "phi": 30.0, "cp": 0.5, "r_elev": 12.0, "d": 3.0, "psi": 5.0,
        "t_cut": t_cut, "phi_b": phi_b, "s_cap": s_cap,
        "pow_a": 1.1, "pow_b": 0.9, "pow_c": 2.0, "pow_d": 4.0,
        "u": u, "ru": 0.35,
        "sigma_gamma": 1.0, "sigma_c": 2.0, "sigma_phi": 3.0, "sigma_cp": 0.1,
        "sigma_d": 0.2, "sigma_psi": 0.3,
        "k1": 1.0, "k2": 2.0, "alpha": 0.4, "unsat": unsat,
        "kr0": 0.6, "h0": 7.0, "vg_a": 0.05, "vg_n": 1.4, "E": 5000.0, "nu": 0.3,
        "hb_sci": 50.0, "hb_gsi": 60.0, "hb_mi": 10.0, "hb_d": 0.0,
    }


def _editor_fixture():
    """A fully-populated slope_data spanning every editable category, canonical so
    each editor's apply-time transform (circle R/Depth, pile theta, geometry
    resync, ...) is idempotent on its managed keys."""
    import math
    from shapely.geometry import Polygon

    materials = [
        _editor_full_material("m-mc-none-lf",    "mc",  "none",  "lf"),
        _editor_full_material("m-cp-piezo-vg",   "cp",  "piezo", "vg"),
        _editor_full_material("m-pow-seep-gard", "pow", "seep",  "gard"),
        _editor_full_material("m-hb-ru-lf",      "hb",  "ru",    "lf"),
        # A seep-only material with a BLANK strength option ('' — valid per the
        # loader). Locks the MaterialsEditor combo's empty entry: without it the
        # combo would normalize blank -> 'mc' and drop the seep-only classification.
        _editor_full_material("m-blank-seep-vg", "",    "seep",  "vg"),
        # v16: an elastic material (with a BLANK t_cut, the common case), so the
        # editor round-trip locks both the new 'elastic' option-combo entry and a
        # None t_cut surviving unchanged (the combo would corrupt 'elastic' if it
        # weren't a choice; a dropped/zeroed t_cut would be caught by the mc rows).
        _editor_full_material("m-elastic-none-lf", "elastic", "none", "lf",
                              t_cut=None, phi_b=None, s_cap=None),
    ]

    def pile(x1, y1, x2, y2, appl="active"):
        p = {"label": "P", "x1": x1, "y1": y1, "x2": x2, "y2": y2, "H": 10.0,
             "theta_p": math.degrees(math.atan2(x2 - x1, -(y2 - y1))),
             "D_pile": 2.0, "S": 6.0, "E": 3000.0, "I": 1.5, "area": 4.0,
             "V_cap": 50.0, "M_cap": 200.0, "fixity": "fixed", "appl": appl}
        return p

    return {
        "gamma_water": 62.4, "tcrack_depth": 0.0, "tcrack_water": 0.0,
        "k_seismic": 0.15, "max_depth": 0.0,
        # v18 unit declaration — the GlobalEditor round-trips these two through its
        # Units/Time selectors; locking them here catches a future drop (matches the
        # fixture's Imperial gamma_water 62.4 and its ~120 pcf material gammas).
        "unit_system": "imperial", "time_unit": "day",
        # v19 main-sheet globals the Global parameters dialog now edits. Distinct
        # non-default values (tension_srf False is the interesting one — the tri-state
        # must not collapse to None or to the shipped YES).
        "k0": 0.65, "tension_srf": False,
        # 'size' (v21) is carried but not editable in either geometry dialog, so it
        # is exactly the kind of key a rebuild-from-fields apply() silently drops.
        "profile_lines": [{"coords": [(0.0, 0.0), (20.0, 20.0), (100.0, 20.0)],
                           "mat_id": 0, "size": 1.75}],
        "polygons": [{"polygon": Polygon([(0.0, 0.0), (20.0, 20.0), (100.0, 20.0),
                                          (100.0, 0.0)]), "mat_id": 0, "size": 1.25}],
        # SSR zone overlays — one row of EACH kind, so the polygon editor has to
        # round-trip all three (and keep them out of 'polygons', which the same check
        # compares). One carries a Size, which is legal on any polygon.
        "ssr_zones": [
            {"kind": "reduce", "label": "SSR reduce", "size": 0.9,
             "polygon": [(10.0, 0.0), (90.0, 0.0), (90.0, 18.0), (10.0, 18.0)]},
            {"kind": "hold", "label": "SSR hold", "size": None,
             "polygon": [(20.0, 1.0), (35.0, 1.0), (35.0, 9.0), (20.0, 9.0)]},
            {"kind": "hold_elastic", "label": "SSR elastic", "size": None,
             "polygon": [(60.0, 1.0), (75.0, 1.0), (75.0, 9.0), (60.0, 9.0)]}],
        # v21 refine region: a polygon that is NOTHING but a local mesh size. It is
        # edited in the polygon dialog alongside the material zones and the SSR
        # overlays, so it must come back on its own list, with its Size, and must
        # never leak into 'polygons'.
        "refine_zones": [{"polygon": [(40.0, 1.0), (55.0, 1.0), (55.0, 8.0),
                                      (40.0, 8.0)], "size": 0.45}],
        "ground_surface": None, "domain_polygon": None, "tcrack_surface": None,
        "materials": materials,
        "piezo_line": [(0.0, 5.0), (100.0, 5.0)],
        "piezo_line2": [(0.0, 3.0), (100.0, 3.0)],
        "circular": True,
        "circles": [{"Xo": 10.0, "Yo": 40.0, "Option": "Depth", "Depth": 5.0,
                     "Xi": 0.0, "Yi": 0.0, "R": 35.0}],
        "non_circ": [{"X": 0.0, "Y": 10.0, "Movement": "Free"},
                     {"X": 50.0, "Y": 2.0, "Movement": "Horiz"},
                     {"X": 100.0, "Y": 10.0, "Movement": "Fixed"}],
        # Three loads with MIXED v21 Directions (and a distinct Normal each, so a
        # block can be identified by its data). Mixed is the point: a writer that
        # wrote one direction for the whole set, or a list that slipped by one,
        # cannot survive this.
        "dloads": [[{"X": 0.0, "Y": 20.0, "Normal": 100.0},
                    {"X": 30.0, "Y": 20.0, "Normal": 100.0}],
                   [{"X": 40.0, "Y": 20.0, "Normal": 200.0},
                    {"X": 60.0, "Y": 20.0, "Normal": 200.0}],
                   [{"X": 70.0, "Y": 20.0, "Normal": 300.0},
                    {"X": 90.0, "Y": 20.0, "Normal": 300.0}]],
        "dload_dirs": ["vertical", "normal", "vertical"],
        "dloads2": [[{"X": 0.0, "Y": 20.0, "Normal": 50.0},
                     {"X": 30.0, "Y": 20.0, "Normal": 50.0}]],
        "dload2_dirs": ["vertical"],
        # One row per support type (blank/geosynthetic/nail/tieback/anchor) so the
        # editor exercises every Type value plus both Dir (tangent/axial) and Appl
        # (active/passive) enums and the tend1/tend2/spacing numerics.
        "reinforcement_lines": [
            {"x1": 0.0, "y1": 5.0, "x2": 40.0, "y2": 5.0, "t_max": 1000.0,
             "t_res": 800.0, "lp1": 2.0, "lp2": 3.0, "E": 2000.0, "area": 1.2,
             "type": "", "dir": "tangent", "appl": "active",
             "tend1": 0.0, "tend2": 0.0, "spacing": 1.0},
            {"x1": 0.0, "y1": 6.0, "x2": 38.0, "y2": 6.0, "t_max": 900.0,
             "t_res": 700.0, "lp1": 1.5, "lp2": 2.5, "E": 1800.0, "area": 1.1,
             "type": "geosynthetic", "dir": "tangent", "appl": "active",
             "tend1": 5.0, "tend2": 6.0, "spacing": 1.0},
            {"x1": 0.0, "y1": 7.0, "x2": 36.0, "y2": 7.0, "t_max": 800.0,
             "t_res": 600.0, "lp1": 1.0, "lp2": 2.0, "E": 1600.0, "area": 1.0,
             "type": "nail", "dir": "axial", "appl": "passive",
             "tend1": 10.0, "tend2": 12.0, "spacing": 1.5},
            {"x1": 0.0, "y1": 8.0, "x2": 34.0, "y2": 8.0, "t_max": 700.0,
             "t_res": 500.0, "lp1": 0.5, "lp2": 1.5, "E": 1400.0, "area": 0.9,
             "type": "tieback", "dir": "axial", "appl": "active",
             "tend1": 15.0, "tend2": 18.0, "spacing": 2.0},
            {"x1": 0.0, "y1": 9.0, "x2": 32.0, "y2": 9.0, "t_max": 600.0,
             "t_res": 400.0, "lp1": 0.5, "lp2": 1.0, "E": 1200.0, "area": 0.8,
             "type": "anchor", "dir": "axial", "appl": "passive",
             "tend1": 20.0, "tend2": 22.0, "spacing": 2.5},
        ],
        "pile_lines": [pile(20.0, 20.0, 20.0, 0.0, "passive"),
                       pile(35.0, 20.0, 35.0, 2.0, "active")],
        # ground_surface is None in this fixture, so LineLoadsEditor.apply performs
        # no snapping and the record must round-trip byte-for-byte.
        "line_loads": [{"x": 30.0, "y": 20.0, "P": 500.0, "angle": -90.0,
                        "label": "Facing plate"}],
        "seepage_bc": {
            "specified_heads": [
                {"head": 18.0, "kind": "reservoir",
                 "coords": [(0.0, 0.0), (0.0, 18.0)]},
                {"head": 5.0, "kind": "head", "coords": [(100.0, 0.0), (100.0, 5.0)]}],
            # Two flux BCs (v15) — the SeepBcEditor now edits fluxes in the master
            # list alongside heads/exit; multiple fluxes lock their list ordering.
            "specified_fluxes": [
                {"flux": 1.5, "coords": [(40.0, 20.0), (60.0, 20.0)]},
                {"flux": -0.75, "coords": [(70.0, 20.0), (85.0, 20.0)]}],
            "exit_face": [(60.0, 20.0), (100.0, 5.0)],
        },
        "seepage_bc2": {
            # Set 2 is the constant-steady rapid-drawdown set: plain heads only (the
            # SeepBcEditor hides the reservoir type for it, and fileio rejects a
            # reservoir/series there). Set 1 above keeps a reservoir to exercise that
            # type's round-trip.
            "specified_heads": [{"head": 10.0, "kind": "head",
                                 "coords": [(0.0, 0.0), (0.0, 10.0)]}],
            "specified_fluxes": [{"flux": 2.25, "coords": [(30.0, 20.0), (45.0, 20.0)]}],
            "exit_face": [(60.0, 20.0), (100.0, 5.0)],
        },
        "has_seepage_bc2": True,
        "mesh": None,
    }


def _editor_norm(key, val):
    """Normalize a managed value for deep comparison: material-zone polygons hold
    shapely objects rebuilt fresh on apply, so compare their exterior coordinate
    rings + mat_id rather than object identity."""
    if key == "polygons":
        out = []
        for p in (val or []):
            poly = p.get("polygon")
            coords = list(poly.exterior.coords) if poly is not None else []
            out.append({"coords": coords, "mat_id": p.get("mat_id"),
                        "size": p.get("size")})
        return out
    if key == "ssr_zones":
        # Vertex containers vary (list vs tuple) across the editor round-trip; the
        # kind, the ring and the local mesh size are what must survive.
        return [{"kind": z.get("kind"), "size": z.get("size"),
                 "coords": [tuple(c) for c in (z.get("polygon") or [])]}
                for z in (val or [])]
    if key == "refine_zones":
        return [{"size": z.get("size"),
                 "coords": [tuple(c) for c in (z.get("polygon") or [])]}
                for z in (val or [])]
    return val


def run_editor_roundtrip_test(test):
    """Open each studio CategoryEditor on a fully-populated record, apply with NO
    changes, and assert its managed keys are deep-equal in -> out.

    Returns (0.0, None) on success, or (None, message) naming the dropped/corrupted
    fields. Fast (one offscreen QApplication, no file or solver work).
    """
    import copy
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    from PySide6.QtWidgets import QApplication
    from studio.editors import CATEGORY_EDITORS

    app = QApplication.instance() or QApplication([])
    problems = []
    for key, editor in CATEGORY_EDITORS.items():
        keys = _EDITOR_MANAGED_KEYS.get(key, [])
        sd = _editor_fixture()
        before = {k: _editor_norm(k, copy.deepcopy(sd.get(k))) for k in keys}
        dlg = editor.build(sd, None)
        editor.apply(sd, dlg)
        for k in keys:
            problems += _roundtrip_diff(before[k], _editor_norm(k, sd.get(k)),
                                        f"{key}:{k}")

    # Every editor's live PREVIEW must actually draw. MplCanvas deliberately keeps
    # the last good pixmap when a draw hook raises, so that a half-typed row doesn't
    # blank the pane — which means a preview closure that is simply broken shows as
    # an empty panel and nothing else. (It hid a real one: the dloads preview called
    # a widget method that didn't exist, and the pane just stayed white.) Here the
    # hook is called directly, with no canvas to swallow it.
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as _plt
    from studio.canvas import PreviewPane
    for key, editor in CATEGORY_EDITORS.items():
        sd = _editor_fixture()
        dlg = editor.build(sd, None)
        panes = dlg.findChildren(PreviewPane)
        for n, pane in enumerate(panes):
            fig, ax = _plt.subplots()
            try:
                pane._draw_fn(ax)
            except Exception as exc:
                problems.append(f"{key}: preview {n} raised {type(exc).__name__}: {exc}")
            finally:
                _plt.close(fig)
        dlg.deleteLater()
    app.processEvents()

    # The materials editor has a second view (List) that binds to the SAME rows.
    # Both the list-view round-trip and a table->list->table switch mid-edit must
    # preserve every key — the lossless-switch invariant. The generic loop above
    # only exercises the default (Table) view, so cover the list path explicitly.
    mat_editor = CATEGORY_EDITORS["materials"]

    # (1) Open, switch to LIST view, apply unchanged — deep-equal.
    sd = _editor_fixture()
    before_mats = copy.deepcopy(sd["materials"])
    dlg = mat_editor.build(sd, None)
    dlg.set_view_mode("list")
    mat_editor.apply(sd, dlg)
    problems += _roundtrip_diff(before_mats, sd["materials"], "materials(list)")
    dlg.deleteLater()
    app.processEvents()

    # (2) Edit a value in the TABLE view, switch table->list->table, apply — the
    #     edit must survive both switches and nothing else may change.
    sd = _editor_fixture()
    expected_mats = copy.deepcopy(sd["materials"])
    dlg = mat_editor.build(sd, None)
    dlg.set_view_mode("table")                       # explicit (session may remember List)
    c_col = next(j for j, f in enumerate(mat_editor.FIELDS) if f.key == "c")
    dlg._table.table.item(0, c_col).setText("137.5")
    expected_mats[0]["c"] = 137.5
    dlg.set_view_mode("list")
    dlg.set_view_mode("table")
    mat_editor.apply(sd, dlg)
    problems += _roundtrip_diff(expected_mats, sd["materials"], "materials(switch)")
    dlg.deleteLater()
    app.processEvents()

    # (3) Display colors live in the STYLE delta, not the material dicts. Opening the
    #     editor on a project that carries per-material color overrides and applying
    #     with NO color change must leave the sparse style delta byte-for-byte equal —
    #     in both views — so colors stay out of the material data path.
    # (4) A distributed load's Direction must travel WITH its block through a
    #     delete. The dloads editor used to rewrite the block list without touching
    #     dload_dirs, which is a list parallel to it: removing a load left the
    #     directions at their old positions, so every later load silently inherited
    #     the direction of the one before it — a vertical surcharge applied as a
    #     surface-normal thrust, or the reverse, with nothing on screen to show it.
    dl_editor = CATEGORY_EDITORS["dloads"]

    def _dl_state(d):
        """(identifying Normal, direction) per load — the pairing under test."""
        dirs = d.get("dload_dirs") or []
        return [(b[0]["Normal"], dirs[i] if i < len(dirs) else None)
                for i, b in enumerate(d.get("dloads") or [])]

    sd = _editor_fixture()
    dlg = dl_editor.build(sd, None)
    w1 = dlg._sets[0]
    w1.list.setCurrentRow(1)
    w1._remove_block()                       # delete the MIDDLE load
    dl_editor.apply(sd, dlg)
    got, want = _dl_state(sd), [(100.0, "vertical"), (300.0, "vertical")]
    if got != want:
        problems.append(f"dloads(delete): {got} != {want} — a load's Direction did "
                        "not travel with its block")
    if len(sd.get("dload_dirs") or []) != len(sd.get("dloads") or []):
        problems.append(f"dloads(delete): {len(sd.get('dload_dirs') or [])} "
                        f"direction(s) for {len(sd.get('dloads') or [])} load(s)")
    dlg.deleteLater()
    app.processEvents()

    #     …and through an ADD: the new load takes the default (normal) without
    #     disturbing the directions already on the loads around it.
    sd = _editor_fixture()
    dlg = dl_editor.build(sd, None)
    w1 = dlg._sets[0]
    w1.list.setCurrentRow(0)
    w1._remove_block()                       # delete the FIRST load
    w1._add_block()                          # then append a fresh one
    for _x, _n in ((5.0, 400.0), (15.0, 400.0)):
        w1.table._add_row()
        _r = w1.table.table.rowCount() - 1
        w1.table.table.item(_r, 0).setText(str(_x))      # X
        w1.table.table.item(_r, 1).setText("20.0")       # Y
        w1.table.table.item(_r, 2).setText(str(_n))      # Normal
    dl_editor.apply(sd, dlg)
    got = _dl_state(sd)
    want = [(200.0, "normal"), (300.0, "vertical"), (400.0, "normal")]
    if got != want:
        problems.append(f"dloads(add): {got} != {want} — adding a load disturbed the "
                        "existing Directions")
    dlg.deleteLater()
    app.processEvents()

    # (5) The Run FEM dialog's v21 Side BC. It is a RUN option, not an editor-managed
    #     input, so it has no CategoryEditor to round-trip — but it is the one v21
    #     addition whose only Studio surface is a dialog, and a seeding bug there
    #     would silently run every model on the default restraint. Checked here
    #     because this is the suite's Qt-layer slot.
    from studio.dialogs import RunFemDialog
    for _seed, _want in (({}, "rollers"),                    # dialog's own default
                         ({"side_bc": "rollers"}, "rollers"),
                         ({"side_bc": "fixed"}, "fixed")):   # seeded from the file
        _fd = RunFemDialog(defaults=_seed)
        _got = _fd.options().get("side_bc")
        if _got != _want:
            problems.append(f"RunFemDialog side_bc: seeded {_seed!r} -> {_got!r}, "
                            f"expected {_want!r}")
        _fd.deleteLater()
    app.processEvents()

    from studio.editors import MaterialsDialog, _new_material
    sd = _editor_fixture()
    style = {"materials": {"0": {"color": "#123456"},
                           "1": {"color": "#abcdef", "hatch": "....", "alpha": 0.5}}}
    before_style = copy.deepcopy(style)
    dlg = MaterialsDialog("Materials", mat_editor.FIELDS, sd["materials"],
                          _new_material, None, style=style)
    dlg.set_view_mode("table")
    problems += _roundtrip_diff(before_style, dlg.result_style(), "materials(style)")
    dlg.set_view_mode("list")
    problems += _roundtrip_diff(before_style, dlg.result_style(), "materials(style-list)")
    dlg.deleteLater()
    app.processEvents()

    # The reinforcement and pile editors likewise have a List view bound to the SAME
    # rows (default view; the table is the bulk-entry path). Cover both invariants for
    # each — a list-view no-op round-trip and a table->list->table mid-edit survival —
    # mirroring the materials cases. The switch case edits a NON-geometry field so the
    # pile's axis-derived θ stays idempotent (editing x/y would legitimately recompute
    # it, which isn't what this guard is measuring).
    for cat, mkey, edit_key, edit_val in (
            ("reinforce", "reinforcement_lines", "t_max", 1234.5),
            ("piles",     "pile_lines",          "V_cap", 77.5)):
        editor = CATEGORY_EDITORS[cat]

        # (1) Open, switch to LIST view, apply unchanged — deep-equal.
        sd = _editor_fixture()
        before = copy.deepcopy(sd[mkey])
        dlg = editor.build(sd, None)
        dlg.set_view_mode("list")
        editor.apply(sd, dlg)
        problems += _roundtrip_diff(before, sd[mkey], f"{cat}(list)")
        dlg.deleteLater()
        app.processEvents()

        # (2) Edit a value in the TABLE view, switch table->list->table, apply — the
        #     edit must survive both switches and nothing else may change.
        sd = _editor_fixture()
        expected = copy.deepcopy(sd[mkey])
        dlg = editor.build(sd, None)
        dlg.set_view_mode("table")                   # explicit (default/session is List)
        col = next(j for j, f in enumerate(editor.FIELDS) if f.key == edit_key)
        dlg._table.table.item(0, col).setText(str(edit_val))
        expected[0][edit_key] = edit_val
        dlg.set_view_mode("list")
        dlg.set_view_mode("table")
        editor.apply(sd, dlg)
        problems += _roundtrip_diff(expected, sd[mkey], f"{cat}(switch)")
        dlg.deleteLater()
        app.processEvents()

    # The reinforcement and pile editors carry per-field help and — for the
    # spacing-scaled fields — a live "per element / per unit width" label that flips
    # with the row's Spacing/S. Guard both, in both views, on the (Imperial) fixture:
    #   (a) every field of each editor has non-empty help on the table header, the
    #       list-view label AND edit, and the context-sensitive help strip;
    #   (b) the dynamic label re-words set -> "per element" / blank -> "per unit width",
    #       joining the declared unit string, and flips live as Spacing/S changes.
    for cat, driver, dyn_key, per_elem_unit, per_width_unit in (
            ("reinforce", "spacing", "t_max", "lb", "lb/ft"),
            ("piles",     "S",       "V_cap", "lb", "lb/ft")):
        editor = CATEGORY_EDITORS[cat]
        # Every field carries help text.
        for f in editor.FIELDS:
            if not (getattr(f, "tooltip", "") or "").strip():
                problems.append(f"{cat}:tooltip:{f.key} is empty")

        sd = _editor_fixture()
        dlg = editor.build(sd, None)

        # Table view: every header item has a non-empty tooltip.
        dlg.set_view_mode("table")
        app.processEvents()
        tbl = dlg._table.table
        for j, f in enumerate(editor.FIELDS):
            it = tbl.horizontalHeaderItem(j)
            if it is None or not it.toolTip():
                problems.append(f"{cat}:header-tooltip:{f.key} missing")

        # List view: every edit widget has a non-empty tooltip; the help strip is
        # wired (field_help maps every field key).
        dlg.set_view_mode("list")
        lv = dlg._list_view
        lv.list.setCurrentRow(0)
        app.processEvents()
        for f in editor.FIELDS:
            w = lv._edits.get(f.key)
            if w is None or not w.toolTip():
                problems.append(f"{cat}:list-tooltip:{f.key} missing")

        # Dynamic suffix flips live with the driver (Spacing/S). Blank -> per unit
        # width (+ per-width unit); a real spacing -> per element (+ per-element unit).
        drv = lv._edits[driver]
        drv.setText("")
        app.processEvents()
        lab_blank = lv._dyn_labels[dyn_key][0].text()
        if "per unit width" not in lab_blank or per_width_unit not in lab_blank:
            problems.append(f"{cat}:dyn blank label = {lab_blank!r}")
        drv.setText("2.0")
        app.processEvents()
        lab_set = lv._dyn_labels[dyn_key][0].text()
        if "per element" not in lab_set or per_elem_unit not in lab_set:
            problems.append(f"{cat}:dyn set label = {lab_set!r}")
        dlg.deleteLater()
        app.processEvents()

    # The Transient editor round-trips the whole tseep contract — the shared time axis,
    # named series aligned to it (with None gaps), the run controls, save_times and the
    # rapid-drawdown stage times. The generic loop above only sees the fixture (which
    # carries no tseep, so the editor opens disabled and applies None), so round-trip a
    # REAL tseep-bearing file explicitly. The fixture is used read-only (copied to a
    # scratch temp first, then loaded from there).
    tseep_master = _repo('docs/seep/files/xslope_earth_dam_tseep.xlsx')
    if os.path.exists(tseep_master):
        import shutil
        import tempfile
        from xslope.fileio import load_slope_data as _load_ts
        with tempfile.TemporaryDirectory() as _td:
            _fx = os.path.join(_td, "tseep_fixture.xlsx")
            shutil.copy(tseep_master, _fx)
            sd = _load_ts(_fx)
        before_ts = copy.deepcopy(sd.get("tseep"))
        if not before_ts:
            problems.append("transient(tseep-file): fixture carries no tseep data")
        else:
            ed = CATEGORY_EDITORS["transient"]
            dlg = ed.build(sd, None)
            ed.apply(sd, dlg)
            problems += _roundtrip_diff(before_ts, sd.get("tseep"), "transient(tseep-file)")
            dlg.deleteLater()
            app.processEvents()

    problems += _run_dialog_preflight_checks(app)

    if problems:
        return None, "editor round-trip dropped/corrupted data: " + "; ".join(problems[:6])
    return 0.0, None


# === Run-dialog preflight guard (studio.dialogs + studio.preflight_panel) ===
# The dialogs are the run gate now: they render the registry's findings, refuse on an
# ERROR, never refuse on a WARNING, dim a method with the SAME sentence the gate would
# have printed, and offer a remedy that is applied only on request. Each of those is a
# claim about behaviour that a wording change or a re-wired signal could break
# silently, so each is checked here — in the suite's Qt slot, beside the editor
# round-trip, since it needs the same offscreen QApplication and no solver work.

def _preflight_dialog_model():
    """A small LEM model the dialog checks have something to say about."""
    from xslope.fileio import load_slope_data
    return load_slope_data(_repo('docs/inputs/slope/xslope_dam.xlsx'))


def _run_dialog_preflight_checks(app):
    import copy

    from PySide6.QtWidgets import QLabel, QPushButton
    from studio.dialogs import RunLemDialog
    from studio.document import new_slope_data
    from xslope import preflight as pf

    problems = []

    def _findings(dlg):
        return {lbl.objectName()[len("finding_"):]
                for lbl in dlg.preflight.findChildren(QLabel)
                if lbl.objectName().startswith("finding_")}

    # (1) A clean model: no error, Run live.
    sd = _preflight_dialog_model()
    dlg = RunLemDialog(slope_data=sd)
    if dlg.preflight.blocked or not dlg._ok.isEnabled():
        problems.append("preflight(clean): Run is refused on a model with no error")
    dlg.deleteLater()

    # (2) An ERROR blocks Run, is rendered, and carries the registry's own message.
    sd = _preflight_dialog_model()
    sd["k_seismic"] = None                      # main!D13 blank -> every kw is NaN
    dlg = RunLemDialog(slope_data=sd)
    if not dlg.preflight.blocked or dlg._ok.isEnabled():
        problems.append("preflight(error): a blank main!D13 did not refuse the run")
    if "main.seismic_missing" not in _findings(dlg):
        problems.append(f"preflight(error): finding not rendered ({_findings(dlg)})")
    gate = pf.preflight(sd, "lem", dlg._selection()).errors
    if gate and dlg.preflight.block_reason() != gate[0].message:
        problems.append("preflight(error): the dialog's reason is not the gate's message")
    dlg.deleteLater()

    # (3) A WARNING is rendered and does NOT block — the severity contract.
    sd = _preflight_dialog_model()
    sd["piezo_line"] = list(reversed(list(sd["piezo_line"])))
    dlg = RunLemDialog(slope_data=sd)
    if "order.piezo_reversed" not in _findings(dlg):
        problems.append("preflight(warning): a reversed piezo line was not reported")
    if dlg.preflight.blocked or not dlg._ok.isEnabled():
        problems.append("preflight(warning): a WARNING blocked the run")

    # (4) …and it carries its remedy as a live button, whose click applies the
    #     change to the model and clears the finding.
    btns = [b for b in dlg.preflight.findChildren(QPushButton)
            if b.objectName().startswith("remedy_")]
    keys = [b.objectName()[len("remedy_"):] for b in btns]
    if "reverse_polyline:piezo1" not in keys:
        problems.append(f"preflight(remedy): no reversal button offered ({keys})")
    else:
        from xslope import remedies as rem
        before = list(sd["piezo_line"])
        # The confirmation is the dialog's; the change is what is under test.
        dlg.preflight.apply_proposal(rem.propose(sd, "reverse_polyline", "piezo1"))
        after = list(dlg.preflight.model["piezo_line"])
        if after != list(reversed(before)):
            problems.append("preflight(remedy): the line was not reversed in the model")
        if "order.piezo_reversed" in _findings(dlg):
            problems.append("preflight(remedy): the finding survived its own remedy")
    dlg.deleteLater()
    app.processEvents()

    # (5) A remedy that cannot fully succeed DIMS with its reason instead of failing
    #     when pressed, and the gate is evaluated PER TARGET: line 1 simply runs
    #     backwards and can be reversed, line 2 rises and then falls and cannot, so
    #     the two buttons must not share one answer.
    sd = _preflight_dialog_model()
    pts = [(float(x), float(y)) for x, y in sd["piezo_line"]]
    sd["piezo_line"] = list(reversed(pts))                     # reversible
    sd["piezo_line2"] = [pts[0], pts[-1], pts[len(pts) // 2]]  # rises, then falls
    dlg = RunLemDialog(slope_data=sd)
    state = {b.objectName()[len("remedy_"):]: (b.isEnabled(), b.toolTip())
             for b in dlg.preflight.findChildren(QPushButton)
             if b.objectName().startswith("remedy_")}
    live = state.get("reverse_polyline:piezo1")
    dimmed = state.get("reverse_polyline:piezo2")
    if not (live and live[0]):
        problems.append("preflight(remedy): the reversible line offered no live button")
    if dimmed is None or dimmed[0]:
        problems.append("preflight(remedy): a non-reversible line offered a live button")
    elif not dimmed[1].strip():
        problems.append("preflight(remedy): a dimmed button carries no reason")
    dlg.deleteLater()

    # (6) Method dimming: the reason on a dimmed item IS the gate's message, and a
    #     surface change re-filters live.
    sd = _preflight_dialog_model()
    sd["non_circ"] = [{"X": -50.0, "Y": 40.0, "Movement": "Free"},
                      {"X": 100.0, "Y": 10.0, "Movement": "Horiz"},
                      {"X": 300.0, "Y": 40.0, "Movement": "Free"}]
    dlg = RunLemDialog(slope_data=sd)
    if dlg.surface is None:
        problems.append("preflight(surface): no family selector on a both-family deck")
    else:
        model = dlg.method.model()
        enabled_circ = {dlg.method.itemData(i): model.item(i).isEnabled()
                        for i in range(dlg.method.count())}
        if not all(enabled_circ.values()):
            problems.append("preflight(methods): a method was dimmed on a circle")
        dlg.surface.setCurrentIndex(1)                     # -> non-circular
        app.processEvents()
        for i in range(dlg.method.count()):
            key = dlg.method.itemData(i)
            item = model.item(i)
            want = pf.method_surface_reason(key, "noncircular")
            if bool(want) == item.isEnabled():
                problems.append(f"preflight(methods): {key} enabled={item.isEnabled()} "
                                f"with reason {want!r}")
            if want and not item.toolTip():
                problems.append(f"preflight(methods): {key} dimmed with no reason")
        if dlg.method.currentData() in pf.CIRCULAR_ONLY_METHODS:
            problems.append("preflight(methods): the dialog kept a refused method selected")
    dlg.deleteLater()
    app.processEvents()

    # (7) A new document takes its water loads automatically (main!D23), like the
    #     template it will be saved into.
    if str(new_slope_data().get("water_loads") or "").lower() != "auto":
        problems.append("new_slope_data(): a new project does not default to auto "
                        "water loads")

    # (8) Blank preservation, end to end: a cleared cell reaches the model as None
    #     (not 0.0), the checks can see the difference, and the writer puts a blank
    #     cell in the sheet rather than inventing a zero on the way out.
    problems += _blank_preservation_checks(app)

    # (9) The surface family the dialog chose survives the session: the write-back
    #     lands in the cell-backed key, a save carries it, and the next dialog opens
    #     on it instead of on the circles again.
    problems += _surface_family_persistence_checks(app)
    return problems


def _surface_family_persistence_checks(app):
    """The dialog's surface choice reaches the file and comes back.

    The file stores and the dialog edits. Wave 5 gave the both-family deck a choice
    but kept it in memory, so it lasted exactly as long as the session -- reopen the
    file and the circles won again, silently. The amendment is the ``main`` sheet's
    Surface family cell, and this is the chain that has to hold end to end: the
    dialog's choice -> the model's ``surface_family`` -> the saved cell -> the next
    dialog's opening state. A single-family deck is the control: nothing asks, and
    nothing is written.
    """
    import tempfile

    from studio.dialogs import RunLemDialog
    from studio.main_window import MainWindow
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx

    problems = []

    class _Doc:
        """Just enough document for the write-back: a model and an undo bracket."""
        def __init__(self, sd):
            self.slope_data = sd
            self.edits = []

        def begin_edit(self, label):
            self.edits.append(label)

        def commit_edit(self):
            pass

    class _Win:
        def __init__(self, sd):
            self.doc = _Doc(sd)

    def _both_family_model():
        sd = load_slope_data(_repo('docs/inputs/slope/xslope_simple1.xlsx'))
        sd["non_circ"] = [dict(p) for p in SURFACE_FAMILY_NONCIRC]
        return sd

    # A both-family deck opens on the circular family (nothing stated yet) and the
    # selector is there to be asked.
    sd = _both_family_model()
    dlg = RunLemDialog(slope_data=sd)
    if dlg.surface is None:
        problems.append("surface_family: no selector on a both-family deck")
        dlg.deleteLater()
        return problems
    if dlg.surface.currentData() != "circular":
        problems.append(f"surface_family: a deck stating nothing opened on "
                        f"{dlg.surface.currentData()!r}, not the automatic answer")
    dlg.surface.setCurrentIndex(1)                            # -> non-circular
    app.processEvents()
    opts = dlg.options()
    dlg.deleteLater()

    # The write-back is the real one, driven through a stand-in document (building a
    # whole MainWindow offscreen would test the window, not the write-back).
    win = _Win(sd)
    MainWindow._store_surface_family(win, opts["surface"])
    if sd.get("surface_family") != "noncircular":
        problems.append(f"surface_family: the run's choice was stored as "
                        f"{sd.get('surface_family')!r}, not 'noncircular'")
    if sd.get("circular"):
        problems.append("surface_family: the circular flag stayed on after a "
                        "non-circular run (the runners and the plots read it)")
    if not win.doc.edits:
        problems.append("surface_family: the choice was written outside the undo stack")

    # It survives the file: save, reload, and the next dialog opens on it.
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    try:
        save_slope_data_to_xlsx(sd, tmp)
        reloaded = load_slope_data(tmp)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)
    if reloaded.get("surface_family") != "noncircular":
        problems.append(f"surface_family: save/reload lost the choice "
                        f"({reloaded.get('surface_family')!r})")
    dlg = RunLemDialog(slope_data=reloaded)
    if dlg.surface is None or dlg.surface.currentData() != "noncircular":
        got = None if dlg.surface is None else dlg.surface.currentData()
        problems.append(f"surface_family: a reopened file's dialog opened on {got!r}")
    dlg.deleteLater()
    app.processEvents()

    # The control: a single-family deck asks nothing and writes nothing.
    sd = load_slope_data(_repo('docs/inputs/slope/xslope_simple1.xlsx'))
    dlg = RunLemDialog(slope_data=sd)
    if dlg.surface is not None:
        problems.append("surface_family: a single-family deck was asked to choose")
    chosen = dlg.options()["surface"]
    dlg.deleteLater()
    app.processEvents()
    win = _Win(sd)
    MainWindow._store_surface_family(win, chosen)
    if sd.get("surface_family") is not None or win.doc.edits:
        problems.append(f"surface_family: a single-family deck was given a stored "
                        f"family ({sd.get('surface_family')!r})")
    return problems


def _blank_preservation_checks(app):
    """A blank material cell must stay blank through the editor and the writer.

    The mutation is the point: clear one cell that a plain float field would have
    stamped 0.0 onto, and require (a) None in the model, (b) a finding that names
    it, and (c) an empty cell in the saved file. Without all three, the model checks
    cannot tell a cohesionless material from an unfilled one on any model Studio has
    touched -- which is every model a Studio user has.
    """
    import copy
    import shutil
    import tempfile

    from studio.editors import CATEGORY_EDITORS
    from xslope import preflight as pf
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx

    problems = []
    editor = CATEGORY_EDITORS["materials"]
    blank_keys = ("gamma", "c", "phi", "d", "psi", "E", "nu", "ru", "k1")

    # (a) Opening the editor and applying with NOTHING typed must not stamp zeros
    #     over a property that was already unset.
    sd = _editor_fixture()
    for m in sd["materials"]:
        m["nu"] = None
    dlg = editor.build(sd, None)
    editor.apply(sd, dlg)
    if any(m.get("nu") is not None for m in sd["materials"]):
        problems.append("blank(nu): a round-trip through the editor filled in a "
                        "blank Poisson's ratio")
    dlg.deleteLater()
    app.processEvents()

    # (b) Clearing a cell in the table gives None, not 0.0 -- for every field whose
    #     zero is physically meaningful.
    sd = _editor_fixture()
    dlg = editor.build(sd, None)
    dlg.set_view_mode("table")
    for key in blank_keys:
        col = next(j for j, f in enumerate(editor.FIELDS) if f.key == key)
        dlg._table.table.item(0, col).setText("")
    editor.apply(sd, dlg)
    got = {k: sd["materials"][0].get(k) for k in blank_keys}
    bad = {k: v for k, v in got.items() if v is not None}
    if bad:
        problems.append(f"blank(clear): cleared cells came back as {bad}, not None")
    dlg.deleteLater()
    app.processEvents()

    # (c) The checks see the blank -- an unset unit weight is an error naming the row.
    ids = {f.rule_id for f in pf.preflight(sd, "lem")}
    if "mat.gamma_nonpositive" not in ids:
        problems.append(f"blank(check): a blank unit weight raised no finding ({ids})")

    # (d) The writer blanks the cell instead of writing 0.0. Round-tripped through a
    #     REAL model so the whole save path runs, not just the mat block.
    src = _repo('docs/inputs/slope/xslope_dam.xlsx')
    if os.path.exists(src):
        with tempfile.TemporaryDirectory() as td:
            model = load_slope_data(src)
            model["materials"][0]["nu"] = None
            out = os.path.join(td, "blank.xlsx")
            try:
                save_slope_data_to_xlsx(model, out)
            except Exception as exc:
                problems.append(f"blank(save): saving a None property raised "
                                f"{type(exc).__name__}: {exc}")
            else:
                import openpyxl
                wb = openpyxl.load_workbook(out)
                ws = wb["mat"]
                header = _mat_header_columns(ws)
                col = header.get("nu")
                if col is None:
                    problems.append("blank(save): no nu column found in the saved sheet")
                else:
                    val = ws.cell(row=header["_row"] + 1, column=col).value
                    if val is not None:
                        problems.append(f"blank(save): a None Poisson's ratio was "
                                        f"written as {val!r}, not a blank cell")
                wb.close()
    return problems


def _mat_header_columns(ws):
    """``{header: column}`` for the mat sheet, plus ``_row`` — its header row."""
    out = {}
    for row in range(1, 25):
        if str(ws.cell(row=row, column=1).value or "").strip().lower() == "mat":
            out["_row"] = row
            for col in range(1, ws.max_column + 1):
                name = str(ws.cell(row=row, column=col).value or "").strip().lower()
                if name:
                    out[name.replace("_", "")] = col
            break
    return out


def run_template_sync_test(test):
    """Verify a copy shipped in the wheel (xslope/resources) is byte-identical to
    its editable docs master, so the two can't silently drift when the master is
    tweaked. test['master']/test['copy'] name the pair (defaults: input template).

    Returns (0.0, None) if identical, else (None, message).
    """
    import filecmp
    master = test.get('master', ROUNDTRIP_TEMPLATE)
    copy = test.get('copy', BUNDLED_TEMPLATE)
    if not os.path.exists(master):
        return None, f"master file missing: {master}"
    if not os.path.exists(copy):
        return None, f"packaged copy missing: {copy}"
    if not filecmp.cmp(master, copy, shallow=False):
        return None, (f"packaged copy differs from master — run: "
                      f"cp {master} {copy}")
    return 0.0, None


# === v16/v17 backward-compatibility guard ================================ #
# The v16 mat-sheet reshuffle moved E/nu (to columns M/N) and inserted t_cut (L);
# v17 then inserted the matric-suction pair phi_b/s_cap (cols L/M, shifting t_cut to
# N). The loader keys every material column BY NAME, so an OLDER file — where E/nu
# sit wherever they did and the newer columns are simply absent — must still load to
# exactly the same slope_data, save for the new-key defaults (t_cut/phi_b/s_cap all
# None: no cutoff, no suction strength, uncapped). Real corpus MODELS are frozen
# here, materials key-by-key, captured from the loader BEFORE the change:
#   - two PRE-v16 files (no t_cut column): the v16 reshuffle proof.
#   - one V16-ERA file (t_cut column present, no phi_b/s_cap): the v17 insert proof,
#     that adding phi_b/s_cap (cols L/M) is transparent to a v16 file.
# A regression that hardcodes a column position, or breaks the by-name E/nu lookup,
# would read E/nu back as 0.0 and trip this.
#
# The mat-sheet VERSION each model is read at is declared below and the fixture is
# SYNTHESIZED at it (provision_legacy_file), for the reason given at
# ROUNDTRIP_FILES — and this guard is the worked example of why. Its third row was
# described here as "a v16-era file" long after the file itself had been rebuilt to
# v20: the premise had quietly evaporated and the guard was proving the v17 insert
# against a post-v17 file. Declaring the version instead of inheriting whatever the
# corpus happens to carry makes that impossible.
_V16_BACKCOMPAT_EXPECTED = {
    # docs/verification/files/rocscience/vp039d.xlsx — read at v13 (E/nu present; E at old
    # col AF). Same two-material Fill/Soft Clay model as its sibling vp039c, which used to
    # be the sample here: vp039c carries an RS2 vendor tensile cap now (t_cut is no longer
    # None on it), so it can no longer stand in for a pre-v16 file.
    # Retaken after vp039d's builder was registered: the file had been out of the rebuild
    # guard's reach and so carried two staleness marks this capture had frozen in place.
    # (1) The sigmas were 1.2/1.8/2.744 on BOTH materials — the ACADS donor's reliability
    # dispersions copied into a deterministic reinforced-slope problem, the same leak vp042
    # below records; the builder clears them. (2) E/nu were a flat 100000/0.3 placeholder
    # predating the soil-type elastic classifier, now 137000/0.3 for the compacted fill and
    # 8000/0.45 for the soft clay. The row stays probative either way: what it proves is
    # that a v13 file's E/nu are still found BY NAME after the v16/v17 column inserts, so
    # the values only have to be the two distinct ones the file actually holds — and two
    # different E's per file is a sharper probe of a positional-column regression than the
    # single repeated 100000 was.
    _repo('docs/verification/files/rocscience/vp039d.xlsx'): [
        {'name': 'Fill', 'gamma': 17.0, 'gamma_sat': 17.0, 'option': 'mc', 'c': 0.0, 'phi': 37.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 0.0, 'psi': 0.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'none', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 137000.0, 'nu': 0.3, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
        {'name': 'Soft Clay', 'gamma': 20.0, 'gamma_sat': 20.0, 'option': 'mc', 'c': 20.0, 'phi': 0.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 0.0, 'psi': 0.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'none', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 8000.0, 'nu': 0.45, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
    ],
    # docs/inputs/slope/xslope_dam.xlsx — read at v12 (piezo u; blank gsat -> None)
    _repo('docs/inputs/slope/xslope_dam.xlsx'): [
        {'name': 'Shell', 'gamma': 125.0, 'gamma_sat': None, 'option': 'mc', 'c': 0.0, 'phi': 34.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 0.0, 'psi': 0.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'piezo', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 700000.0, 'nu': 0.3, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
        {'name': 'Core', 'gamma': 122.0, 'gamma_sat': None, 'option': 'mc', 'c': 100.0, 'phi': 26.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 300.0, 'psi': 20.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'piezo', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 700000.0, 'nu': 0.3, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
        {'name': 'Clay', 'gamma': 123.0, 'gamma_sat': None, 'option': 'mc', 'c': 0.0, 'phi': 24.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 100.0, 'psi': 19.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'piezo', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 700000.0, 'nu': 0.3, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
        {'name': 'Sand', 'gamma': 127.0, 'gamma_sat': None, 'option': 'mc', 'c': 0.0, 'phi': 32.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 0.0, 'psi': 0.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'piezo', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 0.0, 'nu': 0.0, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
    ],
    # docs/verification/files/rocscience/vp042.xlsx — read at v16 (t_cut column
    # present, all blank; no phi_b/s_cap columns). Captured pre-v17. Proves inserting
    # phi_b/s_cap at cols L/M is transparent: E/nu (now at N.. after the v17 shift)
    # still read by name, and phi_b/s_cap default to None.
    # The sigmas here were 1.2/1.8/2.744 on all four materials — one dispersion for a
    # granular fill, a clay core and a hard base, which is the signature of the ACADS
    # geometry donor's reliability numbers being copied into a derived problem rather
    # than of anything Slide #42 publishes (it is a deterministic safety-map dam). The
    # builder now clears the donor's sigmas, so the capture is retaken at zero; the old
    # expectation had frozen the leak in place.
    _repo('docs/verification/files/rocscience/vp042.xlsx'): [
        {'name': 'Granular fill', 'gamma': 21.5, 'gamma_sat': 21.5, 'option': 'mc', 'c': 0.0, 'phi': 40.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 0.0, 'psi': 0.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'piezo', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 175000.0, 'nu': 0.28, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
        {'name': 'Clay core', 'gamma': 20.0, 'gamma_sat': 20.0, 'option': 'mc', 'c': 20.0, 'phi': 20.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 0.0, 'psi': 0.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'piezo', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 32000.0, 'nu': 0.4, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
        {'name': 'Granular fill (below core)', 'gamma': 21.5, 'gamma_sat': 21.5, 'option': 'mc', 'c': 0.0, 'phi': 40.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 0.0, 'psi': 0.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'piezo', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 175000.0, 'nu': 0.28, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
        {'name': 'Hard base', 'gamma': 24.0, 'gamma_sat': 24.0, 'option': 'mc', 'c': 200.0, 'phi': 45.0, 'cp': 0.0, 'r_elev': 0.0, 'd': 0.0, 'psi': 0.0, 'pow_a': 0.0, 'pow_b': 0.0, 'pow_c': 0.0, 'pow_d': 0.0, 'u': 'piezo', 'ru': 0.0, 'sigma_gamma': 0.0, 'sigma_c': 0.0, 'sigma_phi': 0.0, 'sigma_cp': 0.0, 'sigma_d': 0.0, 'sigma_psi': 0.0, 'k1': 0.0, 'k2': 0.0, 'alpha': 0.0, 'unsat': 'lf', 'kr0': 0.0, 'h0': 0.0, 'vg_a': 0.0, 'vg_n': 0.0, 'E': 125000.0, 'nu': 0.3, 'hb_sci': 0.0, 'hb_gsi': 0.0, 'hb_mi': 0.0, 'hb_d': 0.0},
    ],
}

#: The mat-sheet version each frozen model above is read at. Two pre-v16 (no t_cut
#: column) and one v16-era (t_cut, no phi_b/s_cap) — the two eras the guard exists
#: to compare. Synthesized, so they stay those eras.
_V16_BACKCOMPAT_VERSIONS = {
    _repo('docs/verification/files/rocscience/vp039d.xlsx'): 13,
    _repo('docs/inputs/slope/xslope_dam.xlsx'): 12,
    _repo('docs/verification/files/rocscience/vp042.xlsx'): 16,
}


def run_v16_backcompat_test(test):
    """An older file loads to the same materials as before, plus the new-key defaults.

    Each frozen model is written into its declared archived template
    (_V16_BACKCOMPAT_VERSIONS) and read back, then asserted key-by-key against its
    captured baseline once the newer keys are set aside — t_cut (v16), phi_b and
    s_cap (v17), which must all be None (no cutoff, no suction strength, uncapped)
    — and that no material is 'elastic'. This is the load-side proof that the
    column inserts/reshuffles are transparent to old files, and the fixture is
    built at the old version rather than assumed to still be at it.

    Returns (0.0, None) on success, or (None, message) naming the divergences.
    """
    import tempfile
    from xslope.fileio import load_slope_data
    problems = []
    _NEW_KEYS = ("t_cut", "phi_b", "s_cap")
    with tempfile.TemporaryDirectory() as td:
        for fp, expected in _V16_BACKCOMPAT_EXPECTED.items():
            version = _V16_BACKCOMPAT_VERSIONS[fp]
            if not os.path.exists(fp):
                problems.append(f"{fp}: file missing")
                continue
            if not os.path.exists(LEGACY_TEMPLATE_FMT.format(version)):
                problems.append(f"{fp}: archived v{version} template missing")
                continue
            try:
                fixture = provision_legacy_file(
                    fp, version,
                    os.path.join(td, f"v{version}_{os.path.basename(fp)}"))
            except Exception as exc:                      # noqa: BLE001
                problems.append(f"{fp}: could not provision a v{version} fixture: "
                                f"{exc}")
                continue
            sd = load_slope_data(fixture)
            got_v = sd.get("template_version")
            if got_v is not None and int(float(got_v)) != version:
                problems.append(f"{fp}: fixture is v{got_v}, expected v{version}")
            mats = sd.get("materials", [])
            if len(mats) != len(expected):
                problems.append(f"{fp}: {len(mats)} materials, expected {len(expected)}")
                continue
            for i, (m, exp) in enumerate(zip(mats, expected)):
                for nk in _NEW_KEYS:
                    if nk not in m:
                        problems.append(f"{fp}[{i}]: missing new {nk} key")
                    elif m.get(nk) is not None:
                        problems.append(f"{fp}[{i}].{nk}: {m.get(nk)!r}, expected None")
                if str(m.get("option", "")).strip().lower() == "elastic":
                    problems.append(f"{fp}[{i}]: unexpectedly 'elastic'")
                stripped = {k: v for k, v in m.items() if k not in _NEW_KEYS}
                problems += _roundtrip_diff(exp, stripped,
                                            f"{os.path.basename(fp)}[{i}]")
    if problems:
        return None, "v16 back-compat mismatch: " + "; ".join(problems[:6])
    return 0.0, None


def run_deps_declared_test(test):
    """Every third-party module imported at module scope in xslope/ must be a declared
    runtime dependency in pyproject.toml.

    A module-scope import that isn't declared makes `pip install xslope` install a
    package that cannot be imported — while every test here still passes, because the
    suite runs against the source tree in a dev environment that happens to have the
    module. `twine check` doesn't catch it either; it only validates metadata. That is
    exactly how lxml (fileio, since 2026-06-08) and tabulate (solve, advanced) shipped
    undeclared through releases 0.1.50-0.1.54: six weeks of broken clean installs.

    Import-time is the criterion, not usage. A module imported lazily inside a function
    may legitimately be an optional extra (ezdxf in cad.py is, and __init__ imports only
    _version, so a plain `import xslope` never pulls it). One imported at module scope
    cannot be optional for any module a user imports.

    Returns (0.0, None) if clean, else (None, message).
    """
    import ast as _ast
    root = Path(__file__).parent
    pyproject = root / "pyproject.toml"
    if not pyproject.exists():
        return None, "pyproject.toml missing"

    # Declared runtime deps: the [project] dependencies array. Parsed by hand so the
    # check runs on 3.9/3.10 without tomllib.
    txt = pyproject.read_text()
    m = re.search(r"^dependencies\s*=\s*\[(.*?)\]", txt, re.S | re.M)
    if not m:
        return None, "could not find [project] dependencies in pyproject.toml"
    declared = {re.split(r"[<>=!\[; ]", s)[0].strip().lower()
                for s in re.findall(r'"([^"]+)"', m.group(1))}

    # Optional extras are allowed at module scope only in modules the package never
    # imports on its own; those are opt-in imports by the user.
    extras = set()
    for em in re.finditer(r'^\w[\w-]*\s*=\s*\[(.*?)\]', txt, re.S | re.M):
        for s in re.findall(r'"([^"]+)"', em.group(1)):
            extras.add(re.split(r"[<>=!\[; ]", s)[0].strip().lower())
    OPTIONAL_OK = {"ezdxf", "gmsh", "pyside6", "litellm", "keyring", "pyobjc-framework-cocoa"}

    # Distributions whose IMPORT name is not their package name. The declared list
    # holds distribution names (what pip installs) and the scan sees import names
    # (what Python loads); every other dependency here happens to spell them the
    # same, so only the exceptions are stated.
    IMPORT_TO_DIST = {"docx": "python-docx"}

    stdlib = getattr(sys, "stdlib_module_names", frozenset())
    offenders = {}
    for py in sorted((root / "xslope").glob("*.py")):
        try:
            tree = _ast.parse(py.read_text())
        except SyntaxError as e:
            return None, f"could not parse {py.name}: {e}"
        for node in tree.body:                       # module scope only
            mods = []
            if isinstance(node, _ast.Import):
                mods = [a.name.split('.')[0] for a in node.names]
            elif isinstance(node, _ast.ImportFrom):
                if node.level:                        # relative -> in-package
                    continue
                if node.module:
                    mods = [node.module.split('.')[0]]
            for mod in mods:
                low = IMPORT_TO_DIST.get(mod.lower(), mod.lower())
                if mod in stdlib or mod == "xslope" or low in declared:
                    continue
                if low in OPTIONAL_OK:
                    continue
                offenders.setdefault(mod, []).append(py.name)

    if offenders:
        parts = [f"{mod} ({', '.join(sorted(set(files)))})"
                 for mod, files in sorted(offenders.items())]
        return None, ("imported at module scope but not a declared runtime dependency: "
                      + "; ".join(parts)
                      + " — add to [project] dependencies in pyproject.toml, or make the "
                        "import lazy if it is genuinely optional")
    return 0.0, None


# ===========================================================================
# Preflight (xslope.preflight) -- rule registry regression family
#
# Three checks, in increasing scope:
#
#   preflight_contract  the registry's own invariants, and the two entry points
#   preflight_rules     one mutation per rule, through the real Excel path where
#                       the writer can express the mutation: break exactly that
#                       requirement, assert exactly that finding, and assert the
#                       unbroken model raises nothing (the negative control)
#   preflight_corpus    every tagged file must preflight CLEAN (zero ERROR) for
#                       the analysis its tag names -- a rule that refuses a
#                       locked-green corpus file is miscalibrated, not strict
# ===========================================================================

#: Sample files the mutation tests break copies of. Each is preflight-clean for
#: its analysis, which is what makes it usable as a negative control.
PREFLIGHT_BASE_LEM = _repo('docs/inputs/slope/xslope_dam.xlsx')
PREFLIGHT_BASE_SEEP = _repo('docs/inputs/seep/xslope_earth_dam1.xlsx')
PREFLIGHT_BASE_FEM = _repo('docs/fem/files/xslope_griffiths1.xlsx')
PREFLIGHT_BASE_NONCIRC = _repo('docs/verification/files/rocscience/vp047.xlsx')
PREFLIGHT_BASE_BOTH = _repo('docs/verification/files/rocscience/vp042.xlsx')
PREFLIGHT_BASE_TSEEP = _repo('docs/seep/files/xslope_earth_dam_tseep.xlsx')
PREFLIGHT_BASE_RAPID = _repo('docs/verification/files/rocscience/vp096.xlsx')
PREFLIGHT_BASE_RAPID_MULTI = _repo('docs/verification/files/rocscience/vp099.xlsx')
PREFLIGHT_BASE_PILES = _repo('docs/lem/files/xslope_piles.xlsx')
PREFLIGHT_BASE_PILES_FEM = _repo('docs/fem/files/xslope_piles_fem.xlsx')
PREFLIGHT_BASE_REINF = _repo('docs/inputs/slope/xslope_reinf.xlsx')
PREFLIGHT_BASE_REINF_FEM = _repo('docs/fem/files/xslope_reinforce_fem.xlsx')
PREFLIGHT_BASE_CRACK = _repo('docs/verification/files/rocscience/vp053.xlsx')
#: A profile-line model, so its domain is built from Max Depth (profile sheet B2)
#: and the writer can express the mutation that breaks it. This is also the file
#: that shipped the defect: its Max Depth sat at the elevation of the toe, so the
#: base of the domain retraced the ground surface and the ring closed on itself.
PREFLIGHT_BASE_PROFILE = _repo('docs/inputs/slope/xslope_reliability.xlsx')
#: A reliability model that is preflight-clean as one: six Mohr-Coulomb materials,
#: standard deviations on two of them, and nothing else to report. The reliability
#: mutations break copies of THIS, so a firing rule is the mutation and not the file.
PREFLIGHT_BASE_RELIABILITY = _repo('docs/verification/files/rocscience/vp035.xlsx')
#: The three models the weak-zone remedy has to tell apart, since what it may offer
#: depends entirely on which of these a file is. ACADS problem 3(b) is a 0.5-unit
#: weak layer under a 12-unit slope -- one zone weaker than the rest by a wide
#: margin, so the generator picks it and the remedy is live. Loukidis' three dipping
#: bands are the ambiguous case: two comparable candidates, which is a question for
#: the zone picker rather than a fault for a remedy. vp047 has a single material and
#: therefore no weak layer at all. The same three the generator's own standing check
#: uses (test/noncircular_generator_check.py), so the two cannot drift apart.
PREFLIGHT_NONCIRC_SEAM = _repo('docs/lem/files/xslope_acads_weak_layer.xlsx')
PREFLIGHT_NONCIRC_AMBIGUOUS = _repo('docs/verification/files/rocscience/vp063.xlsx')
PREFLIGHT_NONCIRC_ONE_ZONE = _repo('docs/verification/files/rocscience/vp047.xlsx')
#: The Griffiths soft-band section: one thin inclined seam through a uniform slope,
#: carrying the local Size that makes it meshable. The base for the thin-zone
#: advisory, because its mitigation is already in the file — the control can be the
#: model as shipped, and the mutation is taking the mitigation away.
PREFLIGHT_BASE_THIN = _repo('docs/fem/files/xslope_griffiths3_r0p2_thin.xlsx')


def _pf_set(d, **kw):
    d.update(kw)
    return d


def _pf_mats(sd, **kw):
    for m in sd['materials']:
        m.update(kw)
    return sd


def _pf_mat(sd, i, **kw):
    """Update ONE material row (0-based) and return the model, not the row."""
    sd['materials'][i].update(kw)
    return sd


def _pf_drop(sd, key):
    """Remove a top-level model key (a mesh, a stored field) and return the model."""
    sd.pop(key, None)
    return sd


def _pf_rows(sd, key, **kw):
    """Update every row of a structural sheet (``pile_lines``, ``reinforcement_lines``)."""
    for r in sd.get(key) or []:
        r.update(kw)
    return sd


def _pf_tseep(sd, **kw):
    sd['tseep'] = dict(sd['tseep'] or {}, **kw)
    return sd


def _pf_pts(obj):
    """(x, y) pairs from a shapely line or a plain coordinate list."""
    return [(float(x), float(y)) for x, y in getattr(obj, 'coords', obj)]


def _pf_circle_depth(sd, below):
    """Put every circle's nadir ``below`` units under the domain floor.

    ``R`` follows ``Depth`` (the loader collapses every circle to ``R = Yo -
    Depth``), so the mutated circle is the one a user would get back from the
    sheet, not an inconsistent pair.
    """
    floor = sd['domain_polygon'].bounds[1]
    for c in sd.get('circles') or []:
        c['Depth'] = floor - below
        c['R'] = c['Yo'] - c['Depth']
    return sd


def _pf_zone_sizes(sd, size=None):
    """Set (or clear, with None) the local element Size on every material zone."""
    for p in sd.get('polygons') or []:
        p['size'] = size
    return sd


def _pf_grid_mesh(sd, size):
    """Attach a uniform triangular grid of element size ``size`` over the model.

    Hand-built rather than meshed, because the thin-zone advisory reads exactly two
    things off a mesh — where each element is and how big it is — and a grid states
    both without a gmsh call inside a mutation spec. Two calls at two sizes are the
    whole mitigation test for the attached-mesh branch: a mesh whose elements do not
    fit across a thin zone, and one whose elements do (which is what the Build-mesh
    dialog's thin-zone refinement produces).
    """
    xmin, ymin, xmax, ymax = sd['domain_polygon'].bounds
    nx = max(int(round((xmax - xmin) / size)), 1)
    ny = max(int(round((ymax - ymin) / size)), 1)
    dx, dy = (xmax - xmin) / nx, (ymax - ymin) / ny
    nodes, index = [], {}
    for j in range(ny + 1):
        for i in range(nx + 1):
            index[(i, j)] = len(nodes)
            nodes.append([xmin + i * dx, ymin + j * dy])
    elements, types = [], []
    for j in range(ny):
        for i in range(nx):
            a, b = index[(i, j)], index[(i + 1, j)]
            c, d = index[(i + 1, j + 1)], index[(i, j + 1)]
            for tri in ((a, b, c), (a, c, d)):
                elements.append([tri[0], tri[1], tri[2], 0, 0, 0, 0, 0, 0])
                types.append(3)
    sd['mesh'] = {'nodes': nodes, 'elements': elements, 'element_types': types,
                  'element_materials': [1] * len(elements)}
    return sd


def _pf_move(sd, key, dx=0.0, dy=0.0):
    """Translate every structural line on a sheet, to take it off the surface."""
    for r in sd.get(key) or []:
        r['x1'] += dx; r['x2'] += dx
        r['y1'] += dy; r['y2'] += dy
    return sd


#: One entry per rule. Fields:
#:   rule      the rule id under test
#:   base      the sample file to break a copy of
#:   analysis  / selection   what the run would be
#:   mutation  callable(slope_data) breaking exactly that requirement
#:   control   callable(slope_data) applied instead, to prove the rule is silent
#:             on a model that satisfies it (default: leave the file alone)
#:   mode      'excel' round-trips the mutated model through the writer and the
#:             loader, so the rule is exercised on a file a user could actually
#:             save; 'dict' mutates in memory, which is the honest mode for the
#:             inputs the writer does not carry (a mesh, a stored seepage field)
#:             and for the vocabulary guards the LOADER already refuses -- those
#:             carry a paired 'excel' spec asserting the load-time refusal.
#:   expect    a substring the finding's message must contain
#:   load_error   for a paired spec: the substring the LOADER must refuse with
PREFLIGHT_RULE_SPECS = [
    # --- water and the unit weight of water --------------------------------
    dict(rule='water.gamma_water_missing', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, gamma_water=0.0),
         expect='Unit weight of water is'),
    dict(rule='piezo.line_missing', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, piezo_line=[]),
         expect='fewer than two points'),
    dict(rule='piezo.extent_short', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, piezo_line=sd['piezo_line'][:3]),
         expect='does not cover the whole section'),
    dict(rule='piezo.no_consumer', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, u='none'),
         expect='no material uses u = piezo'),
    dict(rule='water.ponded_no_dload', base=PREFLIGHT_BASE_LEM, mode='excel',
         manual_water=True,
         mutation=lambda sd: _pf_set(sd, dloads=[], dload_dirs=[]),
         expect='above the ground surface'),

    # --- automatic water loads (v22 main!D23) ------------------------------
    # The base is taken in the pre-v22 idiom (manual_water), so it carries its
    # reservoir as a transcribed block and flipping the mode alone IS the double
    # count: the engine derives the same load beside it.
    dict(rule='water.auto_dload_double_count', base=PREFLIGHT_BASE_LEM, mode='excel',
         manual_water=True,
         mutation=lambda sd: _pf_set(sd, water_loads='auto'),
         expect='counted twice'),
    # Automatic mode with a water definition the derivation cannot measure along:
    # the reservoir silently disappears unless something says so. The seepage head
    # boundaries are cleared first so the piezometric line is the source, and it is
    # the line that is broken.
    dict(rule='water.auto_derivation_empty', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, water_loads='auto', seepage_bc={},
                                     piezo_line=list(reversed(_pf_pts(sd['piezo_line'])))),
         expect='derived no water load at all'),
    # Two sheets stating where the same pool stands, disagreeing by 5 ft.
    dict(rule='water.sources_disagree', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(
             sd, piezo_line=[(x, y + 5.0) for x, y in _pf_pts(sd['piezo_line'])]),
         expect='describe different pools'),

    # --- material vocabulary and strength ----------------------------------
    dict(rule='mat.u_vocabulary', base=PREFLIGHT_BASE_LEM, mode='dict',
         mutation=lambda sd: _pf_mats(sd, u='pieso'),
         expect='is not a pore-pressure option'),
    dict(rule='mat.u_vocabulary', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, u='pieso'),
         load_error='unrecognized pore pressure option'),
    dict(rule='mat.unsat_vocabulary', base=PREFLIGHT_BASE_SEEP, mode='dict',
         analysis='seep',
         mutation=lambda sd: _pf_mats(sd, unsat='vgn'),
         expect='is not an unsaturated conductivity model'),
    dict(rule='mat.option_vocabulary', base=PREFLIGHT_BASE_LEM, mode='dict',
         mutation=lambda sd: _pf_mats(sd, option='mohr'),
         expect='is not a strength model'),
    dict(rule='mat.option_vocabulary', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, option='mohr'),
         load_error='unrecognized strength option'),
    dict(rule='mat.option_missing', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, option=''),
         expect='has no strength model'),
    dict(rule='mat.gamma_nonpositive', base=PREFLIGHT_BASE_LEM, mode='dict',
         mutation=lambda sd: _pf_mats(sd, gamma=0.0),
         expect='has no unit weight'),
    dict(rule='mat.gamma_nonpositive', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, gamma=0.0),
         load_error='non-positive unit weight'),
    dict(rule='mat.no_shear_strength', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, option='mc', c=0.0, phi=0.0),
         expect='no shear strength at all'),
    dict(rule='mat.cp_zero_strength', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, option='cp', c=0.0, cp=0.0),
         expect='undrained strength Su is zero'),
    dict(rule='mat.ru_zero', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, u='ru', ru=0.0),
         expect='selects u = ru'),

    # --- main-sheet scalars ------------------------------------------------
    dict(rule='main.seismic_missing', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, k_seismic=float('nan')),
         expect='Seismic coefficient k is blank'),
    dict(rule='main.seismic_magnitude', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, k_seismic=15.0),
         expect='fraction of gravity'),
    dict(rule='main.seismic_negative_lem', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, k_seismic=-0.2),
         expect='is negative'),
    dict(rule='main.crack_water_deeper_than_crack', base=PREFLIGHT_BASE_LEM,
         mode='excel',
         mutation=lambda sd: _pf_set(sd, tcrack_water=10.0, tcrack_depth=0.0),
         expect='no crack for the water to stand in'),
    dict(rule='main.lem_method_unknown', base=PREFLIGHT_BASE_LEM, mode='dict',
         selection={'surface': 'circular', 'method': 'wibble'},
         mutation=lambda sd: sd,
         control=lambda sd: sd,
         control_selection={'surface': 'circular', 'method': 'bishop'},
         expect='not a method this package implements'),
    dict(rule='main.lem_method_unknown', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, lem_method='wibble'),
         load_error='unrecognized LEM method'),

    # --- surface family and method compatibility ---------------------------
    dict(rule='surface.none_defined', base=PREFLIGHT_BASE_LEM, mode='dict',
         mutation=lambda sd: _pf_set(sd, circles=[], non_circ=[], circular=False),
         expect='defines no failure surface'),
    # Selection-awareness: the SAME empty-sheet model, and the only difference is
    # that the control's run brought its own surface (generate_slices(circle=...),
    # a search's trial circle, a sweep step). The sheet is then not the source, so
    # the rule must be silent — and this pair is what proves the flag is read
    # rather than merely accepted.
    dict(rule='surface.none_defined', base=PREFLIGHT_BASE_LEM, mode='dict',
         selection={'surface': 'circular'},
         control_selection={'surface': 'circular', 'surface_supplied': True},
         mutation=lambda sd: _pf_set(sd, circles=[], non_circ=[], circular=False),
         control=lambda sd: _pf_set(sd, circles=[], non_circ=[], circular=False),
         expect='defines no failure surface'),
    dict(rule='surface.method_requires_circle', base=PREFLIGHT_BASE_NONCIRC,
         mode='dict', selection={'surface': 'noncircular', 'method': 'oms'},
         control_selection={'surface': 'noncircular', 'method': 'spencer'},
         mutation=lambda sd: sd, control=lambda sd: sd,
         expect='takes moments about a circle centre'),
    dict(rule='surface.family_ambiguous', base=PREFLIGHT_BASE_BOTH, mode='dict',
         selection={}, control_selection={'surface': 'circular'},
         mutation=lambda sd: sd, control=lambda sd: sd,
         expect='did not state which to analyse'),

    # --- the model domain, and the circles that must fit inside it ---------
    # The mutation reconstructs the shipped defect exactly: Max Depth raised to the
    # elevation of the toe, so the base of the domain runs back along the ground
    # surface and the ring closes on the first edge it drew.
    dict(rule='domain.degenerate_ring', base=PREFLIGHT_BASE_PROFILE, mode='excel',
         mutation=lambda sd: _pf_set(sd, max_depth=0.0),
         expect='crosses or retraces itself at (0, 0)'),
    # The control is the boundary this rule has to get right, not an untouched
    # file: a circle 20 below the floor, which is a MODELLING CHOICE (a composite
    # surface truncated along the base) and slices perfectly well. Only the circle
    # that can be sliced NEITHER way is reported.
    dict(rule='surface.circle_below_domain_floor', base=PREFLIGHT_BASE_LEM,
         mode='excel',
         mutation=lambda sd: _pf_circle_depth(sd, 200.0),
         control=lambda sd: _pf_circle_depth(sd, 20.0),
         expect='no slices can be generated from it'),

    # --- polyline ordering -------------------------------------------------
    dict(rule='order.piezo_reversed', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, piezo_line=list(reversed(sd['piezo_line']))),
         expect='run right to left'),
    dict(rule='order.piezo_nonmonotonic', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(
             sd, piezo_line=[sd['piezo_line'][0], sd['piezo_line'][2],
                             sd['piezo_line'][1], sd['piezo_line'][3]]),
         expect='X values are not in order'),
    dict(rule='order.dload_reversed', base=PREFLIGHT_BASE_LEM, mode='dict',
         manual_water=True,
         mutation=lambda sd: _pf_set(sd, dloads=[list(reversed(sd['dloads'][0]))]),
         expect='run right to left'),
    dict(rule='order.dload_nonmonotonic', base=PREFLIGHT_BASE_LEM, mode='excel',
         manual_water=True,
         mutation=lambda sd: _pf_set(
             sd, dloads=[[sd['dloads'][0][0], sd['dloads'][0][2],
                          sd['dloads'][0][1]]]),
         expect='X values are not in order'),

    # --- unit-system plausibility ------------------------------------------
    dict(rule='units.gamma_water_off_band', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, gamma_water=9.81),
         expect='unit weight of water'),
    dict(rule='units.material_gamma_off_band', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, gamma=20.0),
         expect='material unit weights'),
    dict(rule='units.signals_disagree', base=PREFLIGHT_BASE_LEM, mode='dict',
         mutation=lambda sd: _pf_mats(_pf_set(sd, unit_system=None), gamma=20.0),
         control=lambda sd: _pf_set(sd, unit_system=None),
         expect='Two independent signals disagree'),
    dict(rule='mat.E_unusable', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_mats(sd, E=0.0),
         expect="Young's modulus"),

    # --- mesh and stored seepage fields ------------------------------------
    dict(rule='mesh.missing', base=PREFLIGHT_BASE_SEEP, mode='dict',
         analysis='seep', availability=True,
         mutation=lambda sd: _pf_set(sd, mesh=None),
         control=lambda sd: _pf_set(sd, mesh={'nodes': [], 'elements': [],
                                              'element_types': [],
                                              'element_materials': []}),
         expect='carries no finite element mesh'),
    dict(rule='mesh.element_type_unsupported', base=PREFLIGHT_BASE_FEM, mode='dict',
         analysis='seep',
         mutation=lambda sd: _pf_set(sd, mesh=dict(sd['mesh'],
                                                   element_types=[5, 5])),
         expect='element type code'),
    # Linear elements on a stress run: once from the mesh the run would solve,
    # once from the main!D18 declaration alone, which is what Studio's run gate
    # can see before any mesh has been built.
    dict(rule='mesh.element_type_linear', base=PREFLIGHT_BASE_FEM, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_set(sd, mesh=dict(sd['mesh'],
                                                   element_types=[3, 3])),
         expect='lock volumetrically'),
    dict(rule='mesh.element_type_linear', base=PREFLIGHT_BASE_FEM, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_set(sd, mesh=None, element_type='quad4'),
         control=lambda sd: _pf_set(sd, mesh=None, element_type='quad8'),
         expect='Element type is quad4'),
    dict(rule='mesh.material_id_out_of_range', base=PREFLIGHT_BASE_FEM, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_set(sd, mesh=dict(sd['mesh'],
                                                   element_materials=[99])),
         expect='references material ID'),
    dict(rule='seep_field.node_count_mismatch', base=PREFLIGHT_BASE_FEM, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_set(sd, seep_u=[0.0, 1.0, 2.0]),
         expect='computed on a different mesh'),
    dict(rule='seep_field.missing', base=PREFLIGHT_BASE_LEM, mode='dict',
         mutation=lambda sd: _pf_mats(sd, u='seep'),
         expect='carries no mesh'),
    # The transient counterpart: the same u = seep materials, but the RUN states
    # that it will stage one instant of a transient march into the model before the
    # solver starts. The selection is the mutation — that fact is API-only, which is
    # exactly why the rule exists to report which instant was read.
    dict(rule='seep_field.transient_frame', base=PREFLIGHT_BASE_LEM, mode='dict',
         selection={'surface': 'circular', 'seep_frame': {'times': [30.0]}},
         mutation=lambda sd: _pf_mats(sd, u='seep'),
         expect='transient seepage solution'),

    # --- steady seepage: material properties -------------------------------
    dict(rule='seep.k1_nonpositive', base=PREFLIGHT_BASE_SEEP, mode='excel',
         analysis='seep',
         mutation=lambda sd: _pf_mats(sd, k1=0.0),
         expect='has no hydraulic conductivity'),
    dict(rule='seep.k2_nonpositive', base=PREFLIGHT_BASE_SEEP, mode='excel',
         analysis='seep',
         mutation=lambda sd: _pf_mats(sd, k2=0.0),
         expect='has no minor hydraulic conductivity'),
    dict(rule='seep.k2_greater_than_k1', base=PREFLIGHT_BASE_SEEP, mode='excel',
         analysis='seep',
         mutation=lambda sd: _pf_mats(sd, k1=1.0, k2=10.0),
         expect='is greater than k1'),
    dict(rule='seep.anisotropy_angle_unset', base=PREFLIGHT_BASE_SEEP, mode='excel',
         analysis='seep',
         mutation=lambda sd: _pf_mats(sd, k1=10.0, k2=1.0, alpha=0.0),
         control=lambda sd: _pf_mats(sd, k1=10.0, k2=1.0, alpha=30.0),
         expect='is anisotropic'),
    dict(rule='seep.unsat_params_missing', base=PREFLIGHT_BASE_SEEP, mode='excel',
         analysis='seep',
         mutation=lambda sd: _pf_mats(sd, unsat='lf', kr0=0.0, h0=0.0),
         expect='needs kr0 > 0'),
    dict(rule='seep.confined_unsat_unused', base=PREFLIGHT_BASE_SEEP, mode='dict',
         analysis='seep',
         mutation=lambda sd: _pf_mats(
             _pf_set(sd, seepage_bc=dict(sd['seepage_bc'], exit_face=[])),
             unsat='vg', vg_a=1.0, vg_n=2.0),
         expect='No exit face is defined'),

    # --- steady seepage: boundary conditions -------------------------------
    dict(rule='seep.no_boundary_conditions', base=PREFLIGHT_BASE_SEEP, mode='dict',
         analysis='seep',
         mutation=lambda sd: _pf_set(sd, seepage_bc={'specified_heads': [],
                                                     'specified_fluxes': [],
                                                     'exit_face': []}),
         expect='no seepage boundary conditions at all'),
    dict(rule='seep.no_dirichlet', base=PREFLIGHT_BASE_SEEP, mode='dict',
         analysis='seep',
         mutation=lambda sd: _pf_set(sd, seepage_bc={
             'specified_heads': [], 'exit_face': [],
             'specified_fluxes': [{'flux': 1.0, 'coords': [(0.0, 0.0), (1.0, 0.0)],
                                   'kind': 'flux'}]}),
         expect='defined only up to an additive constant'),
    dict(rule='seep.exit_face_without_head', base=PREFLIGHT_BASE_SEEP, mode='dict',
         analysis='seep',
         mutation=lambda sd: _pf_set(sd, seepage_bc={
             'specified_heads': [], 'specified_fluxes': [],
             'exit_face': sd['seepage_bc']['exit_face']}),
         expect='no specified head and no flux'),
    dict(rule='seep.no_gradient', base=PREFLIGHT_BASE_SEEP, mode='dict',
         analysis='seep',
         mutation=lambda sd: _pf_set(sd, seepage_bc={
             'specified_fluxes': [], 'exit_face': [],
             'specified_heads': [dict(b, head=5.0)
                                 for b in sd['seepage_bc']['specified_heads']]}),
         expect='no gradient and no flow'),
    dict(rule='seep.bc_polyline_too_short', base=PREFLIGHT_BASE_SEEP, mode='dict',
         analysis='seep',
         mutation=lambda sd: _pf_set(sd, seepage_bc=dict(
             sd['seepage_bc'],
             specified_heads=[dict(b, coords=b['coords'][:1])
                              for b in sd['seepage_bc']['specified_heads']])),
         expect='fewer than two points'),

    # --- finite element elastic properties and the tensile cap -------------
    dict(rule='mat.nu_unusable', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_mats(sd, nu=0.0),
         expect="Poisson's ratio"),
    dict(rule='mat.tensile_cap_missing', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_mats(sd, t_cut=None),
         control=lambda sd: _pf_mats(sd, t_cut=50.0),
         expect='has no tensile strength'),
    dict(rule='main.tension_srf_unset', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_mats(_pf_set(sd, tension_srf=None), t_cut=50.0),
         control=lambda sd: _pf_mats(_pf_set(sd, tension_srf=True), t_cut=50.0),
         expect='Tension SRF'),

    # --- finite element initial stress and SSR zones -----------------------
    dict(rule='fem.k0_without_zone_geometry', base=PREFLIGHT_BASE_FEM, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_set(sd, k0=0.5, polygons=[], profile_lines=[]),
         control=lambda sd: _pf_set(sd, k0=0.5),
         expect='no material-zone geometry'),
    dict(rule='fem.k0_pre_equilibration', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='fem',
         mutation=lambda sd: _pf_set(sd, k0=0.5),
         expect='ground that is not level'),
    dict(rule='ssr.zone_contains_no_elements', base=PREFLIGHT_BASE_FEM, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_set(sd, ssr_zones=[
             {'kind': 'reduce', 'label': 'SSR reduce',
              'polygon': [(1e5, 1e5), (1e5 + 10, 1e5), (1e5 + 10, 1e5 + 10)]}]),
         control=lambda sd: _pf_set(sd, ssr_zones=[
             {'kind': 'reduce', 'label': 'SSR reduce',
              'polygon': [(0.0, 0.0), (80.0, 0.0), (80.0, 50.0), (0.0, 50.0)]}]),
         expect='contains no mesh elements'),
    dict(rule='mesh.zone_size_not_finer', base=PREFLIGHT_BASE_FEM, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_set(sd, target_size=2.0, refine_zones=[
             {'polygon': [(0.0, 0.0), (5.0, 0.0), (5.0, 5.0)], 'size': 4.0}]),
         control=lambda sd: _pf_set(sd, target_size=2.0, refine_zones=[
             {'polygon': [(0.0, 0.0), (5.0, 0.0), (5.0, 5.0)], 'size': 0.5}]),
         expect='not finer than the global target'),
    # A thin zone too coarse to mesh, twice over: once as the model DECLARES it and
    # once as the attached mesh actually resolved it. The base is the Griffiths
    # soft-band section, whose seam is 2.95 wide and whose polygon sheet carries the
    # Size that makes it meshable — so each control is the mitigation itself rather
    # than an unrelated model, which is the only way to test that the mitigation
    # stops the rule.
    dict(rule='mesh.thin_zone_unresolved', base=PREFLIGHT_BASE_THIN, mode='excel',
         analysis='fem',
         mutation=lambda sd: _pf_zone_sizes(_pf_set(sd, target_size=3.0), None),
         control=lambda sd: _pf_set(sd, target_size=3.0),
         expect='cannot develop a shear band'),
    dict(rule='mesh.thin_zone_unresolved', base=PREFLIGHT_BASE_THIN, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_grid_mesh(_pf_zone_sizes(sd, None), 3.0),
         control=lambda sd: _pf_grid_mesh(_pf_zone_sizes(sd, None), 0.7),
         expect='element rows across it'),
    # The elastic carve-out, pinned as the pair it is: the SAME unmeshable seam,
    # differing only in the strength model on its material row. An elastic zone
    # cannot yield at any element size, so there is no shear band for a coarse mesh
    # to lose; anything that can yield is warned about, which is why the mutation
    # sets an ordinary Mohr-Coulomb row rather than merely leaving one there.
    dict(rule='mesh.thin_zone_unresolved', base=PREFLIGHT_BASE_THIN, mode='dict',
         analysis='fem',
         mutation=lambda sd: _pf_mat(_pf_zone_sizes(_pf_set(sd, target_size=3.0),
                                                    None), 1, option='mc'),
         control=lambda sd: _pf_mat(_pf_zone_sizes(_pf_set(sd, target_size=3.0),
                                                   None), 1, option='elastic'),
         expect='cannot develop a shear band'),

    # --- transient seepage -------------------------------------------------
    dict(rule='tseep.time_unit_missing', base=PREFLIGHT_BASE_TSEEP, mode='dict',
         analysis='tseep',
         mutation=lambda sd: _pf_set(sd, time_unit=None),
         expect='declares no time unit'),
    dict(rule='tseep.storage_nonpositive', base=PREFLIGHT_BASE_TSEEP, mode='excel',
         analysis='tseep',
         mutation=lambda sd: _pf_mats(sd, Ss=0.0),
         expect='specific storage'),
    dict(rule='tseep.retention_band_substituted', base=PREFLIGHT_BASE_TSEEP,
         mode='excel', analysis='tseep',
         mutation=lambda sd: _pf_mats(sd, unsat='lf', h0=0.0),
         expect='invented rather than'),
    dict(rule='tseep.duration_invalid', base=PREFLIGHT_BASE_TSEEP, mode='dict',
         analysis='tseep',
         mutation=lambda sd: _pf_tseep(sd, duration=0.0),
         expect='needs a positive duration'),
    dict(rule='tseep.save_interval', base=PREFLIGHT_BASE_TSEEP, mode='excel',
         analysis='tseep',
         mutation=lambda sd: _pf_tseep(sd, save_interval=None),
         expect='Save interval is blank'),
    dict(rule='tseep.stage_times', base=PREFLIGHT_BASE_TSEEP, mode='excel',
         analysis='tseep',
         mutation=lambda sd: _pf_tseep(sd, stage_1=100.0, stage_2=50.0),
         expect='not earlier than Stage 2 time'),
    dict(rule='tseep.save_times_beyond_duration', base=PREFLIGHT_BASE_TSEEP,
         mode='excel', analysis='tseep',
         mutation=lambda sd: _pf_tseep(sd, save_times=[1.0e9]),
         expect='lie beyond the Duration'),
    dict(rule='tseep.initial_condition', base=PREFLIGHT_BASE_TSEEP, mode='dict',
         analysis='tseep',
         mutation=lambda sd: _pf_tseep(sd, series={
             k: [None] + list(v[1:]) for k, v in sd['tseep']['series'].items()}),
         expect='no value at the first time'),
    dict(rule='tseep.reservoir_face_above_level', base=PREFLIGHT_BASE_TSEEP,
         mode='dict', analysis='tseep',
         mutation=lambda sd: _pf_tseep(sd, series={
             k: [1.0] + list(v[1:]) for k, v in sd['tseep']['series'].items()}),
         expect='reservoir face rising'),

    # --- rapid drawdown ----------------------------------------------------
    dict(rule='rapid.stage2_water_missing', base=PREFLIGHT_BASE_RAPID, mode='excel',
         analysis='rapid',
         mutation=lambda sd: _pf_set(sd, piezo_line2=[]),
         expect='Piezometric Line 2'),
    dict(rule='rapid.ru_has_no_stage2', base=PREFLIGHT_BASE_RAPID, mode='excel',
         analysis='rapid',
         mutation=lambda sd: _pf_mats(sd, u='ru', ru=0.3),
         expect='no staged variant'),
    dict(rule='rapid.dloads2_missing', base=PREFLIGHT_BASE_RAPID, mode='dict',
         analysis='rapid', manual_water=True,
         mutation=lambda sd: _pf_set(sd, dloads2=[], dload2_dirs=[]),
         expect='stage-2 distributed loads are empty'),
    dict(rule='rapid.d_psi_incomplete', base=PREFLIGHT_BASE_RAPID, mode='excel',
         analysis='rapid',
         mutation=lambda sd: _pf_mats(sd, psi=0.0),
         expect='carries d but not psi'),
    dict(rule='rapid.no_low_k_material', base=PREFLIGHT_BASE_RAPID, mode='excel',
         analysis='rapid',
         mutation=lambda sd: _pf_mats(sd, d=0.0, psi=0.0),
         expect='no low-permeability material'),
    dict(rule='rapid.free_draining_materials', base=PREFLIGHT_BASE_RAPID_MULTI,
         mode='dict', analysis='rapid',
         mutation=lambda sd: sd,
         control=lambda sd: _pf_mats(sd, d=500.0, psi=15.0),
         expect='treated as free-draining'),
    dict(rule='rapid.cp_ignores_d_psi', base=PREFLIGHT_BASE_RAPID, mode='dict',
         analysis='rapid',
         mutation=lambda sd: _pf_mats(sd, option='cp', cp=0.3),
         expect='option = cp'),
    dict(rule='rapid.curved_envelope_unsupported', base=PREFLIGHT_BASE_RAPID,
         mode='dict', analysis='rapid',
         mutation=lambda sd: _pf_mats(sd, option='pow'),
         expect='rapid drawdown does not support'),
    dict(rule='rapid.pool_rises', base=PREFLIGHT_BASE_RAPID, mode='dict',
         analysis='rapid',
         mutation=lambda sd: _pf_set(sd, piezo_line2=[
             (x, y + 25.0) for x, y in _pf_pts(sd['piezo_line'])]),
         expect='ABOVE Line 1'),
    dict(rule='rapid.stage2_water_without_load', base=PREFLIGHT_BASE_RAPID,
         mode='dict', analysis='rapid', manual_water=True,
         mutation=lambda sd: _pf_set(sd, dloads2=[], dload2_dirs=[], piezo_line2=[
             (x, y + 25.0) for x, y in _pf_pts(sd['ground_surface'])]),
         expect='above the ground surface'),
    dict(rule='rapid.stage2_load_repeats_stage1', base=PREFLIGHT_BASE_RAPID,
         mode='dict', analysis='rapid', manual_water=True,
         mutation=lambda sd: _pf_set(sd, dloads2=[list(b) for b in sd['dloads']],
                                     dload2_dirs=list(sd['dload_dirs'])),
         expect='same resultant'),

    # --- seismic direction -------------------------------------------------
    dict(rule='main.seismic_direction_lem', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, k_seismic=0.15),
         expect='the sign is ignored'),
    dict(rule='main.seismic_direction_fem', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_set(sd, k_seismic=0.15),
         expect='pushes in +x'),

    # --- tension cracks ----------------------------------------------------
    dict(rule='crack.deeper_than_slope', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_set(sd, tcrack_depth=999.0),
         expect='base of a slope'),
    dict(rule='crack.exceeds_theoretical_depth', base=PREFLIGHT_BASE_CRACK,
         mode='dict', selection={'surface': 'noncircular'},
         control_selection={'surface': 'noncircular'},
         mutation=lambda sd: sd,
         control=lambda sd: _pf_set(sd, tcrack_depth=1.0),
         expect='against a theoretical depth'),
    dict(rule='crack.cohesionless_materials', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(_pf_set(sd, tcrack_depth=2.0), c=0.0),
         control=lambda sd: _pf_set(sd, tcrack_depth=2.0),
         expect='cohesionless'),
    dict(rule='crack.no_surface_intersection', base=PREFLIGHT_BASE_LEM, mode='dict',
         mutation=lambda sd: _pf_set(sd, tcrack_depth=80.0,
                                     circles=sd['circles'][:1]),
         control=lambda sd: _pf_set(sd, tcrack_depth=5.0,
                                    circles=sd['circles'][:1]),
         expect='intersects none of the failure surfaces'),
    dict(rule='crack.ignored_by_fem', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_set(sd, tcrack_depth=3.0),
         expect='does not include'),

    # --- piles -------------------------------------------------------------
    dict(rule='pile.spacing_invalid', base=PREFLIGHT_BASE_PILES, mode='excel',
         mutation=lambda sd: _pf_rows(sd, 'pile_lines', S=0.0),
         expect='has a spacing of 0'),
    dict(rule='pile.spacing_invalid', base=PREFLIGHT_BASE_PILES, mode='excel',
         mutation=lambda sd: _pf_rows(sd, 'pile_lines', S=-6.0),
         expect='silently unconservative'),
    dict(rule='pile.spacing_not_greater_than_diameter', base=PREFLIGHT_BASE_PILES,
         mode='dict',
         mutation=lambda sd: _pf_rows(sd, 'pile_lines', H=None, S=1.0, D_pile=2.0),
         control=lambda sd: _pf_rows(sd, 'pile_lines', H=None, S=6.0, D_pile=2.0),
         expect='not greater than D'),
    dict(rule='pile.fem_incomplete', base=PREFLIGHT_BASE_PILES_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_rows(sd, 'pile_lines', E=None),
         expect='E is blank'),
    dict(rule='pile.units_convention', base=PREFLIGHT_BASE_PILES, mode='excel',
         mutation=lambda sd: _pf_rows(sd, 'pile_lines', H=100.0),
         expect='different conventions'),
    dict(rule='pile.section_derived', base=PREFLIGHT_BASE_PILES_FEM, mode='dict',
         analysis='ssrm',
         mutation=lambda sd: sd,
         control=lambda sd: _pf_rows(sd, 'pile_lines', area=3.14, I=0.785),
         expect='derives the solid circular section'),
    dict(rule='pile.no_surface_engagement', base=PREFLIGHT_BASE_PILES, mode='dict',
         mutation=lambda sd: _pf_move(sd, 'pile_lines', dx=1.0e5),
         expect='crosses any failure surface'),

    # --- reinforcement -----------------------------------------------------
    dict(rule='reinforce.dir_vocabulary', base=PREFLIGHT_BASE_REINF_FEM,
         mode='dict', analysis='ssrm',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', dir='sideways'),
         expect='is not a direction'),
    dict(rule='reinforce.dir_vocabulary', base=PREFLIGHT_BASE_REINF_FEM,
         mode='excel', analysis='ssrm',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', dir='sideways'),
         load_error="unrecognized Dir"),
    dict(rule='reinforce.tmax_nonpositive', base=PREFLIGHT_BASE_REINF_FEM,
         mode='dict', analysis='ssrm',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', t_max=0.0),
         expect='has no tensile capacity'),
    dict(rule='reinforce.fem_incomplete', base=PREFLIGHT_BASE_REINF_FEM,
         mode='excel', analysis='ssrm',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', E=float('nan')),
         expect='no usable axial stiffness'),
    dict(rule='reinforce.fem_incomplete_on_lem', base=PREFLIGHT_BASE_REINF_FEM,
         mode='excel',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', E=float('nan')),
         expect='not approximating the same reinforcement'),
    dict(rule='reinforce.type_blank', base=PREFLIGHT_BASE_REINF_FEM, mode='excel',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', type=''),
         control=lambda sd: _pf_rows(sd, 'reinforcement_lines',
                                     type='geosynthetic'),
         expect='leave Type blank'),
    dict(rule='reinforce.pullout_negative', base=PREFLIGHT_BASE_REINF_FEM,
         mode='excel',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', lp1=-2.0),
         expect='read as FULLY ANCHORED'),
    dict(rule='reinforce.envelope_inconsistent', base=PREFLIGHT_BASE_REINF_FEM,
         mode='excel',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', lp1=500.0),
         expect='longer than the line itself'),
    dict(rule='reinforce.no_surface_engagement', base=PREFLIGHT_BASE_REINF_FEM,
         mode='dict',
         mutation=lambda sd: _pf_move(sd, 'reinforcement_lines', dx=1.0e5),
         expect='crosses any failure surface'),

    # --- magnitude plausibility (the sniff tests) --------------------------
    dict(rule='mat.E_off_soil_type_band', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_mats(sd, E=1.0e12),
         expect='times that'),
    dict(rule='mat.nu_implausible', base=PREFLIGHT_BASE_FEM, mode='excel',
         analysis='ssrm',
         mutation=lambda sd: _pf_mats(sd, nu=0.05),
         expect='almost no lateral coupling'),
    dict(rule='structural.modulus_off_band', base=PREFLIGHT_BASE_REINF_FEM,
         mode='excel', analysis='ssrm',
         mutation=lambda sd: _pf_rows(sd, 'reinforcement_lines', E=12.0),
         expect='outside that range'),

    # --- material field ranges (what a swept or perturbed value trips) ------
    dict(rule='mat.strength_negative', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, c=-500.0),
         expect='A strength cannot be negative'),
    dict(rule='mat.phi_out_of_range', base=PREFLIGHT_BASE_LEM, mode='excel',
         mutation=lambda sd: _pf_mats(sd, phi=95.0),
         expect='no finite value at 90 degrees'),

    # --- parametric sweeps -------------------------------------------------
    # These rules read the RUN, not the file: a sweep's inputs are API-only by
    # design. The mutation is therefore the selection, and the control is the same
    # model swept on something it does read.
    dict(rule='sensitivity.ru_without_consumer', base=PREFLIGHT_BASE_LEM,
         mode='dict', analysis='sensitivity',
         selection={'param': 'mat:Shell:ru'},
         control_selection={'param': 'mat:Shell:gamma'},
         mutation=lambda sd: sd,
         expect='ru is never read'),
    dict(rule='sensitivity.rapid_only_field', base=PREFLIGHT_BASE_LEM,
         mode='dict', analysis='sensitivity',
         selection={'param': 'mat:Shell:d'},
         control_selection={'param': 'mat:Shell:gamma'},
         mutation=lambda sd: sd,
         expect='read only by the rapid drawdown analysis'),
    dict(rule='sensitivity.range_crosses_zero', base=PREFLIGHT_BASE_LEM,
         mode='dict', analysis='sensitivity',
         selection={'param': 'mat:Shell:gamma', 'values': [-10.0, 0.0, 10.0]},
         control_selection={'param': 'mat:Shell:gamma', 'values': [10.0, 20.0]},
         mutation=lambda sd: sd,
         expect='will be skipped'),
    dict(rule='sensitivity.seismic_sign_discarded', base=PREFLIGHT_BASE_LEM,
         mode='dict', analysis='sensitivity',
         selection={'param': 'global:k_seismic', 'values': [-0.2, 0.2],
                    'mode': 'lem'},
         control_selection={'param': 'global:k_seismic', 'values': [0.1, 0.2],
                            'mode': 'lem'},
         mutation=lambda sd: sd,
         expect='discards the sign'),

    # --- reliability -------------------------------------------------------
    dict(rule='reliability.no_standard_deviations', base=PREFLIGHT_BASE_RELIABILITY,
         mode='excel', analysis='reliability',
         mutation=lambda sd: _pf_mats(sd, sigma_c=0.0, sigma_phi=0.0,
                                      sigma_gamma=0.0, sigma_cp=0.0),
         expect='Standard Deviations'),
    dict(rule='reliability.elastic_carries_sigma', base=PREFLIGHT_BASE_RELIABILITY,
         mode='excel', analysis='reliability',
         mutation=lambda sd: _pf_mat(sd, 0, option='elastic'),
         expect='it is elastic'),
    dict(rule='reliability.option_unsupported', base=PREFLIGHT_BASE_RELIABILITY,
         mode='excel', analysis='reliability',
         mutation=lambda sd: _pf_mat(sd, 0, option='pow', pow_a=2.0, pow_b=0.7),
         expect="does not support the strength option"),
    dict(rule='reliability.sigma_ignored_by_option', base=PREFLIGHT_BASE_RELIABILITY,
         mode='excel', analysis='reliability',
         mutation=lambda sd: _pf_mat(sd, 0, sigma_cp=2.0),
         expect='does not read that parameter'),
    dict(rule='reliability.dead_sigma_columns', base=PREFLIGHT_BASE_RELIABILITY,
         mode='excel', analysis='reliability',
         mutation=lambda sd: _pf_mat(sd, 0, sigma_d=1.0),
         expect='no reliability engine reads'),
    dict(rule='reliability.sigma_exceeds_mean', base=PREFLIGHT_BASE_RELIABILITY,
         mode='excel', analysis='reliability',
         mutation=lambda sd: _pf_mat(sd, 0, sigma_phi=999.0),
         expect='exceeds the mean'),
    dict(rule='reliability.rapid_has_no_effect', base=PREFLIGHT_BASE_RELIABILITY,
         mode='dict', analysis='reliability',
         selection={'engine': 'mc', 'rapid': True},
         control_selection={'engine': 'taylor', 'rapid': True, 'search': True},
         mutation=lambda sd: sd,
         expect='solved WITHOUT drawdown'),
    dict(rule='reliability.mc_samples_too_few', base=PREFLIGHT_BASE_RELIABILITY,
         mode='dict', analysis='reliability',
         selection={'engine': 'mc', 'n_samples': 1},
         control_selection={'engine': 'mc', 'n_samples': 1000},
         mutation=lambda sd: sd,
         expect='at least 2 samples'),
    dict(rule='reliability.fem_mesh_generated', base=PREFLIGHT_BASE_FEM,
         mode='dict', analysis='reliability',
         selection={'base': 'fem'},
         mutation=lambda sd: _pf_drop(_pf_mats(sd, sigma_c=1.0), 'mesh'),
         control=lambda sd: _pf_mats(sd, sigma_c=1.0),
         expect='generated automatically at a target element size'),
]


def _preflight_eval(sd, spec, selection_key='selection'):
    """Run preflight for a spec and return the findings from its rule."""
    from xslope.preflight import preflight
    analysis = spec.get('analysis', 'lem')
    sel = spec.get(selection_key)
    if sel is None:
        sel = spec.get('selection')
    if sel is None:
        sel = {'surface': 'circular'} if analysis in ('lem', 'rapid') else {}
    report = preflight(sd, analysis, sel,
                       include_availability=spec.get('availability', False))
    return [f for f in report.findings if f.rule_id == spec['rule']]


def _preflight_apply(spec, mutate, template):
    """Load the spec's base file, apply a mutation, and return the model.

    In ``excel`` mode the mutated model is written through the real writer and read
    back through the real loader, so the rule is exercised on a file a user could
    have saved. A loader refusal is returned rather than raised, because for the
    vocabulary guards the refusal IS the assertion.

    ``manual_water`` on the spec asks for the base in the pre-v22 idiom: mode
    manual, with the reservoir typed onto the dloads sheet. Rules about a load a
    USER entered need a load a user entered, and the corpus states its water as a
    definition now — so the fixture is built rather than borrowed, which is what
    keeps these rules testable whatever the sample files happen to carry.
    """
    import tempfile
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx
    sd = load_slope_data(spec['base'])
    if spec.get('manual_water'):
        sd = _legacy_manual_water(sd)
    sd = mutate(sd) or sd
    if spec.get('mode') != 'excel':
        return sd, None
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    try:
        save_slope_data_to_xlsx(sd, tmp, template=template)
        return load_slope_data(tmp), None
    except Exception as exc:
        return None, f"{type(exc).__name__}: {exc}"
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)


def run_preflight_rules_test(test):
    """Mutation test per preflight rule, plus each rule's negative control.

    For every rule: break exactly the requirement it exists to protect and assert
    that rule fires with the message it promises, then assert it stays silent on
    the unbroken model. A rule that cannot be made to fire is dead; a rule that
    fires on a valid model is miscalibrated. Returns (0.0, None) or (None, message).
    """
    import warnings as _warnings
    from xslope.preflight import rules as _rules

    template = test.get('template', ROUNDTRIP_TEMPLATE)
    problems = []
    covered = set()

    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        for spec in PREFLIGHT_RULE_SPECS:
            rid = spec['rule']
            covered.add(rid)

            # 1. the mutation must make exactly this rule fire (or, for a spec that
            #    tests a load-time guard, the LOADER must refuse it).
            sd, load_err = _preflight_apply(spec, spec['mutation'], template)
            if 'load_error' in spec:
                if load_err is None:
                    problems.append(f"{rid}: the loader accepted a model it should "
                                    f"refuse ({spec['load_error']!r})")
                elif spec['load_error'] not in load_err:
                    problems.append(f"{rid}: loader refused with {load_err[:90]!r}, "
                                    f"expected {spec['load_error']!r}")
                continue
            if load_err is not None:
                problems.append(f"{rid}: mutated model failed to round-trip: "
                                f"{load_err[:110]}")
                continue
            fired = _preflight_eval(sd, spec)
            if not fired:
                problems.append(f"{rid}: the mutation did not make the rule fire")
            elif not any(spec['expect'] in f.message for f in fired):
                problems.append(f"{rid}: fired, but no message contained "
                                f"{spec['expect']!r} (got {fired[0].message[:80]!r})")

            # 2. the negative control: the rule must be silent on a model that
            #    satisfies it.
            control = spec.get('control', lambda sd_: sd_)
            sd0, load_err0 = _preflight_apply(spec, control, template)
            if load_err0 is not None:
                problems.append(f"{rid}: control model failed to round-trip: "
                                f"{load_err0[:110]}")
                continue
            still = _preflight_eval(sd0, spec, selection_key='control_selection')
            if still:
                problems.append(f"{rid}: fired on the control model "
                                f"({still[0].message[:80]!r})")

    uncovered = sorted(r.id for r in _rules() if r.id not in covered)
    if uncovered:
        problems.append("rules with no mutation test: " + ", ".join(uncovered))

    if problems:
        return None, f"{len(problems)} problem(s): " + "; ".join(problems[:6])
    return 0.0, None


#: How a docs test tag's ``type=`` maps onto a preflight analysis and selection.
#: The tag vocabulary IS the surface-family selection, which is what makes the
#: corpus usable as the regression base: every row states what it runs.
PREFLIGHT_TAG_ANALYSIS = {
    'single_circle': ('lem', {'surface': 'circular'}),
    'single_noncirc': ('lem', {'surface': 'noncircular'}),
    'circular_search': ('lem', {'surface': 'circular', 'search': True}),
    'noncircular_search': ('lem', {'surface': 'noncircular', 'search': True}),
    'mp_spencer': ('lem', {'surface': 'circular'}),
    'gsat_pair': ('lem', {}),
    'design_search': ('lem', {'surface': 'circular', 'search': True}),
    'critical_kc': ('lem', {'surface': 'circular', 'search': True}),
    'sensitivity': ('sensitivity', {}),
    'reliability': ('reliability', {}),
    'reliability_mc': ('reliability', {}),
    'fem_reliability': ('reliability', {'base': 'fem'}),
    'seep': ('seep', {}),
    'seep_head': ('seep', {}),
    'seep_elements': ('seep', {}),
    'tseep_head': ('tseep', {}),
    'fem_ssrm': ('ssrm', {}),
    'fem_elements': ('fem', {}),
    # A mesh-size lock names no analysis of its own: the file's OTHER tags say
    # what the model is for, and this row only counts what the mesher produced.
    'mesh_elements': ('fem', {}),
}


def run_preflight_corpus_test(test):
    """Every tagged file in one markdown source must preflight CLEAN.

    The corpus is the regression base for the rule registry: these files carry
    standing locked answers, so a rule that reports an ERROR on one of them is
    miscalibrated -- the model is known-good and the rule is wrong about it. Each
    file is checked for the analysis its own test tag names, with the tag's surface
    family as the selection.

    Also fails on a rule that could not be EVALUATED: preflight downgrades a
    raising rule to an info so a broken rule can never block a valid model, and
    this is what stops that safety net from hiding the breakage.
    """
    import warnings as _warnings
    from xslope.fileio import load_slope_data
    from xslope.preflight import preflight, ERROR

    src = test['file']
    if not os.path.exists(src):
        return None, f"source markdown missing: {src}"

    cases = {}
    for t in parse_test_tags(src):
        mapped = PREFLIGHT_TAG_ANALYSIS.get(t.get('type', 'single_circle'))
        f = t.get('file')
        if mapped is None or not f or not f.endswith('.xlsx'):
            continue
        analysis, sel = mapped
        # A tag's own `rapid=true` is the surface-family selection's counterpart for
        # the analysis TYPE: the same circular run, checked against the rapid
        # drawdown requirements it actually carries. Without this the rapid family
        # would have no corpus regression at all.
        if analysis == 'lem' and str(t.get('rapid', '')).strip().lower() == 'true':
            analysis = 'rapid'
        cases.setdefault((f, analysis, tuple(sorted(sel.items()))), (analysis, sel))

    problems = []
    checked = 0
    loaded = {}
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        for (path, _a, _s), (analysis, sel) in sorted(cases.items()):
            if not os.path.exists(path):
                continue
            if path not in loaded:
                try:
                    loaded[path] = load_slope_data(path)
                except Exception as exc:
                    loaded[path] = exc
            sd = loaded[path]
            if isinstance(sd, Exception):
                problems.append(f"{os.path.basename(path)}: load failed "
                                f"({type(sd).__name__})")
                continue
            checked += 1
            report = preflight(sd, analysis, sel)
            for f in report.errors:
                problems.append(f"{os.path.basename(path)} [{analysis}]: "
                                f"{f.rule_id}: {f.message[:100]}")
            for f in report.infos:
                if 'could not be evaluated' in f.message:
                    problems.append(f"{os.path.basename(path)}: {f.message[:120]}")

    if problems:
        return None, (f"{len(problems)} corpus file(s) refused by preflight "
                      f"(a locked-green file that a rule rejects means the RULE is "
                      f"miscalibrated): " + "; ".join(problems[:5]))
    if checked == 0:
        return None, f"no tagged files found in {src}"
    return 0.0, None


#: A small, fast model for the sweep-gate guard: one Mohr-Coulomb material, a
#: circle on the sheet, and a factor of safety in a fraction of a second on a
#: fixed surface. The sweep contract is about WHICH rows come back and what they
#: carry, not about the numbers, so the cheapest real model is the right one.
SWEEP_GATE_BASE = _repo('docs/inputs/slope/xslope_simple1.xlsx')


def run_sweep_gate_test(test):
    """The two-stage sweep contract, and the three ways it could go silent.

    A parametric sweep validates in two stages -- one full preflight of the base
    model, then a per-step re-check of only the rules that read the substituted
    field -- and a step carrying an ERROR is SKIPPED WITH A STATED REASON rather
    than killing the sweep. Every clause of that is asserted here, and each of the
    three ways the machinery could stop working is written as a mutation that must
    trip a check:

    1. **A rule that stops firing.** If the field declaration is dropped from a
       rule, ``rules_for_field`` no longer selects it and the per-step gate goes
       silent on exactly the value it exists to catch. Asserted both ways: the
       filter must name the rule, and the same check with no field must be silent
       (which is what proves the filter is doing the work).
    2. **A skip that stops carrying its reason.** The skipped row's message must be
       the registry's own sentence for that model, character for character -- not a
       generic "point failed". A user reading the sweep sees why.
    3. **A sentinel screen that stops screening.** ``fs = 9999`` is the search's
       no-admissible-surface flag and a negative factor of safety is not a result;
       both are recorded with ``success=True`` by the solvers and must be turned
       into failed rows here. A discharge is exempt from the sign test, because for
       q the sign is a direction.

    Plus the two defects §5.3 of the plan measured: a failing step must not destroy
    the points before it, and a model whose geometry was ALREADY invalid must not
    have that blamed on every edit.

    Returns (0.0, None) on success, else (None, message).
    """
    import warnings as _warnings
    import numpy as _np
    from xslope.fileio import load_slope_data
    from xslope import preflight as pf
    from xslope import sensitivity as sens

    problems = []
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        sd = load_slope_data(SWEEP_GATE_BASE)
    mat = sd['materials'][0]['name']

    # -- 1. the per-step filter selects the rule that reads the field ---------
    picked = pf.rules_for_field('gamma', 'lem')
    if 'mat.gamma_nonpositive' not in picked:
        problems.append("rules_for_field('gamma') no longer selects "
                        "mat.gamma_nonpositive — the per-step gate would pass a "
                        "zero unit weight")
    if 'main.seismic_missing' in picked:
        problems.append("rules_for_field('gamma') selected a rule that does not "
                        "read gamma — the filter is not filtering")
    if not pf.rules_for_field('k_seismic', 'lem'):
        problems.append("rules_for_field('k_seismic') selects nothing")

    # the filtered preflight must be the same subset
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        broken = sens.set_param(sd, f'mat:{mat}:gamma', 0.0)
    ids = {f.rule_id for f in pf.preflight(broken, 'lem', {'surface': 'circular'},
                                           fields=['gamma'])}
    if ids - set(picked):
        problems.append(f"preflight(fields=['gamma']) evaluated rules outside the "
                        f"field's own subset: {sorted(ids - set(picked))}")

    # -- 2. _validate_model: geometry baseline, and the field check -----------
    baseline = sens._geometry_baseline(sd)
    reason = sens._validate_model(broken, baseline, field='gamma', mode='lem',
                                  selection={'surface': 'circular'})
    if not reason:
        problems.append("_validate_model passed a zero unit weight")
    if sens._validate_model(broken, baseline, field=None) is not None:
        problems.append("_validate_model reported a field problem with no field "
                        "named — the subset filter is not what is doing the work")
    expected = [f.message for f in pf.preflight(broken, 'lem',
                                                {'surface': 'circular'},
                                                fields=['gamma']).errors]
    if reason not in expected:
        problems.append("the per-step reason is not the registry's own message")

    # an inherited invalid polygon is the MODEL's condition, never the edit's
    from shapely.geometry import Polygon
    inherited = dict(sd)
    inherited['polygons'] = [dict(p) for p in sd['polygons']]
    inherited['polygons'][0] = dict(inherited['polygons'][0],
                                    polygon=Polygon([(0, 0), (2, 0), (2, 2),
                                                     (0, 2), (1, 0), (0, 0)]))
    if sens._validate_model(inherited, sens._geometry_baseline(inherited)) is not None:
        problems.append("a polygon that was invalid BEFORE the edit was blamed on "
                        "the edit")
    if sens._validate_model(inherited, frozenset()) is None:
        problems.append("a polygon the edit invalidated was not reported — the "
                        "geometry check has stopped checking")

    # -- 3. the post-step sentinel screen ------------------------------------
    rows = [{'method': 'spencer', 'fs': 9999.0, 'success': True, 'msg': ''},
            {'method': 'spencer', 'fs': -1.2, 'success': True, 'msg': ''},
            {'method': 'spencer', 'fs': _np.nan, 'success': True, 'msg': ''},
            {'method': 'spencer', 'fs': 1.35, 'success': True, 'msg': ''}]
    screened = sens._screen_point([dict(r) for r in rows], 'lem')
    if screened[0]['success'] or 'sentinel' not in screened[0]['msg']:
        problems.append("the sentinel screen recorded fs = 9999 as a success")
    if screened[1]['success'] or not screened[1]['msg']:
        problems.append("the sentinel screen recorded a negative factor of safety "
                        "as a success")
    if screened[2]['success']:
        problems.append("the sentinel screen recorded a NaN as a success")
    if not screened[3]['success'] or screened[3]['msg']:
        problems.append("the sentinel screen failed a valid factor of safety")
    q = sens._screen_point([{'method': 'seep', 'fs': -3.2, 'success': True,
                             'msg': ''},
                            {'method': 'seep', 'fs': _np.nan, 'success': True,
                             'msg': ''}], 'seep')
    if not q[0]['success']:
        problems.append("the sentinel screen failed a negative discharge, whose "
                        "sign is a direction rather than an error")
    if q[1]['success']:
        problems.append("the sentinel screen passed a non-finite discharge")

    # -- 4. the sweep end to end: skipped with a reason, never killed ---------
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        ok, res = sens.sensitivity(sd, param=f'mat:{mat}:gamma',
                                   values=[125.0, 60.0, 0.0, -60.0],
                                   methods=('spencer',), search=False,
                                   num_slices=20)
    if not ok:
        problems.append(f"a sweep with two inadmissible values was refused "
                        f"outright: {res}")
    else:
        df = res['df']
        swept = df.loc[~df['is_base']].set_index('value')
        for v in (125.0, 60.0):
            if not bool(swept.loc[v, 'success']):
                problems.append(f"the admissible point {v:g} was lost")
        for v in (0.0, -60.0):
            if bool(swept.loc[v, 'success']):
                problems.append(f"the inadmissible point {v:g} was recorded as a "
                                f"success")
            elif 'unit weight' not in str(swept.loc[v, 'msg']):
                problems.append(f"the skipped point {v:g} does not carry the "
                                f"rule's reason: {swept.loc[v, 'msg']!r}")

    # a setter that leaves the model unsolvable must cost ONE point, not the sweep
    def _break_at_two(sd_, value):
        sd_ = sens._copy_for_edit(sd_)
        if value == 2.0:
            sd_['materials'][0]['gamma'] = None
        return sd_

    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        ok2, res2 = sens.sensitivity(sd, modify=_break_at_two, label='probe',
                                     values=[1.0, 2.0, 3.0], methods=('spencer',),
                                     search=False, num_slices=20)
    if not ok2:
        problems.append(f"a point that raised destroyed the whole sweep: {res2}")
    else:
        got = res2['df'].loc[~res2['df']['is_base']]
        if len(got) != 3:
            problems.append(f"the raising point cost the sweep rows: {len(got)} of 3")
        elif bool(got.set_index('value').loc[2.0, 'success']):
            problems.append("the raising point was recorded as a success")

    # -- 5. the base-model preflight, once, at the door ----------------------
    blank_k = dict(sd)
    blank_k['k_seismic'] = None
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        ok3, res3 = sens.sensitivity(blank_k, param=f'mat:{mat}:gamma',
                                     values=[100.0, 120.0], methods=('spencer',),
                                     search=False, num_slices=20)
    if ok3:
        problems.append("a sweep started on a base model preflight refuses")
    elif 'Seismic coefficient k' not in str(res3):
        problems.append(f"the base-model refusal does not name the field: {res3}")
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        ok4, _res4 = sens.sensitivity(blank_k, param=f'mat:{mat}:gamma',
                                      values=[100.0], methods=('spencer',),
                                      search=False, num_slices=20,
                                      check_inputs=False)
    if not ok4:
        problems.append("check_inputs=False did not bypass the base-model gate")

    # -- 6. the reliability gate: refuses, never raises, in one voice --------
    from xslope.reliability import reliability
    no_sigma = dict(sd)
    no_sigma['materials'] = [dict(m, sigma_gamma=0.0, sigma_c=0.0, sigma_phi=0.0,
                                  sigma_cp=0.0) for m in sd['materials']]
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        okr, resr = reliability(no_sigma, 'bishop')
    if okr or resr != pf.no_sigmas_message():
        problems.append("a reliability run with no standard deviations did not "
                        "refuse in the registry's own words")

    elastic = dict(sd)
    elastic['materials'] = [dict(m, option='elastic', sigma_c=2.0)
                            for m in sd['materials']]
    try:
        with _warnings.catch_warnings():
            _warnings.simplefilter('ignore')
            oke, rese = reliability(elastic, 'bishop')
    except Exception as exc:
        oke, rese = True, f'raised {exc!r}'
    want = pf.elastic_sigma_message(elastic['materials'][0]['name'], ['sigma_c'])
    if oke or rese != want:
        problems.append(f"a probabilistic elastic material was not refused before "
                        f"the run, in the same words the engine uses: {rese!r}")

    # ...and the query a dialog dims from must say the same sentence.
    caps = pf.capabilities(no_sigma, {'analysis': 'reliability'})
    cap = caps['analysis']['reliability']
    if cap.available:
        problems.append("capabilities() offers a reliability run on a model with "
                        "no standard deviations")
    elif cap.reason != pf.no_sigmas_message():
        problems.append("the dimming reason and the run's refusal differ")
    if not caps['analysis']['lem'].available:
        problems.append("capabilities() dimmed the plain LEM analysis on a "
                        "runnable model")

    if problems:
        return None, f"{len(problems)} problem(s): " + "; ".join(problems[:6])
    return 0.0, None


def run_preflight_contract_test(test):
    """The registry's own invariants and the two entry points' contracts."""
    import warnings as _warnings
    from xslope.fileio import load_slope_data
    from xslope import preflight as pf

    problems = []
    seen = set()
    for r in pf.rules():
        if r.id in seen:
            problems.append(f"duplicate rule id {r.id}")
        seen.add(r.id)
        if r.severity not in pf.SEVERITY_ORDER:
            problems.append(f"{r.id}: unknown severity {r.severity!r}")
        if not r.summary or not r.summary.strip():
            problems.append(f"{r.id}: no summary")
        for a in r.analyses:
            if a != '*' and a not in pf.ANALYSES:
                problems.append(f"{r.id}: unknown analysis {a!r}")
        for name in r.remedies:
            if name not in pf.REMEDIES:
                problems.append(f"{r.id}: unknown remedy {name!r}")
        # A rule offering several repairs still has exactly one primary, and it is
        # the first: that is the one a Finding carries and an interface reaches for.
        if r.remedies[:1] != ((r.remedy,) if r.remedy is not None else ()):
            problems.append(f"{r.id}: remedy {r.remedy!r} is not the first of "
                            f"{r.remedies!r}")
        if r.capability is not None and r.capability not in pf.CAPABILITY_GROUPS:
            problems.append(f"{r.id}: unknown capability group {r.capability!r}")

    # composition: a composite analysis inherits its base's rules
    if 'lem' not in pf.expand_analysis('rapid'):
        problems.append("rapid does not inherit lem")
    if 'seep' not in pf.expand_analysis('tseep'):
        problems.append("tseep does not inherit seep")
    if 'fem' not in pf.expand_analysis('ssrm'):
        problems.append("ssrm does not inherit fem")
    if 'fem' not in pf.expand_analysis('reliability', base='fem'):
        problems.append("reliability-on-fem does not inherit fem")
    try:
        pf.expand_analysis('nonsense')
        problems.append("expand_analysis accepted an unknown analysis")
    except ValueError:
        pass

    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        sd = load_slope_data(PREFLIGHT_BASE_LEM)
        nc = load_slope_data(PREFLIGHT_BASE_NONCIRC)

    # the gate and the query must agree: a method the query dims is a method the
    # gate refuses, in the same words.
    caps = pf.capabilities(nc)
    for group in pf.CAPABILITY_GROUPS:
        if group not in caps:
            problems.append(f"capabilities() omitted the {group!r} group")
    for m in ('oms', 'bishop'):
        cap = caps['lem_method'][m]
        if cap.available:
            problems.append(f"capabilities(): {m} should be unavailable on a "
                            f"non-circular-only model")
        elif not cap.reason:
            problems.append(f"capabilities(): {m} dimmed with no reason string")
        else:
            gate = pf.preflight(nc, 'lem',
                                {'surface': 'noncircular', 'method': m}).errors
            if not gate:
                problems.append(f"{m}: dimmed by capabilities() but not refused by "
                                f"the gate")
            elif gate[0].message != cap.reason:
                problems.append(f"{m}: the dimming reason and the refusal differ")
    for m in ('janbu', 'spencer', 'mprice', 'corps', 'lowe'):
        if not caps['lem_method'][m].available:
            problems.append(f"capabilities(): {m} should be available on a "
                            f"non-circular model")

    # report contract
    clean = pf.preflight(sd, 'lem', {'surface': 'circular'})
    if not clean.ok or clean.errors:
        problems.append(f"the sample model {PREFLIGHT_BASE_LEM} is not preflight "
                        f"clean: {[f.rule_id for f in clean.errors]}")
    if not bool(clean):
        problems.append("a clean report is falsy")
    clean.raise_for_errors()          # must not raise

    broken = dict(sd)
    broken['k_seismic'] = None
    bad = pf.preflight(broken, 'lem', {'surface': 'circular'})
    if bad.ok:
        problems.append("a model with a blank seismic coefficient passed")
    try:
        bad.raise_for_errors()
        problems.append("raise_for_errors() did not raise on an error report")
    except pf.PreflightError as exc:
        if not isinstance(exc, ValueError):
            problems.append("PreflightError is not a ValueError")
        if getattr(exc, 'report', None) is None:
            problems.append("PreflightError carries no report")
    if 'Seismic coefficient k' not in bad.format():
        problems.append("format() lost the finding text")
    if bad.format(min_severity=pf.ERROR).count('WARNING') != 0:
        problems.append("format(min_severity=ERROR) leaked a warning")

    # the escape hatch must actually bypass the gate
    from xslope.slice import generate_slices
    try:
        generate_slices(broken, circle=sd['circles'][0], num_slices=10, debug=False)
        problems.append("generate_slices ran a model preflight refuses")
    except pf.PreflightError:
        pass
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        ok, _res = generate_slices(dict(sd), circle=sd['circles'][0], num_slices=10,
                                   debug=False, check_inputs=False)
    if not ok:
        problems.append("check_inputs=False did not bypass the gate")

    # A caller that hands generate_slices its own surface must not be refused for
    # an empty circles sheet. This is the whole-path version of the
    # surface.none_defined selection pair above: the sheet is emptied on a real
    # model and the same circle is passed in as an argument, which is what every
    # search, sweep and plot does.
    sheetless = dict(sd)
    sheetless['circles'] = []
    sheetless['non_circ'] = []
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        try:
            ok2, _res2 = generate_slices(sheetless, circle=sd['circles'][0],
                                         num_slices=10, debug=False)
            if not ok2:
                problems.append("generate_slices refused a caller-supplied circle "
                                "on a model with an empty circles sheet")
        except pf.PreflightError as exc:
            problems.append(f"generate_slices refused a caller-supplied surface: "
                            f"{exc}")
    # ...and the sheet-emptiness rule still fires when NOTHING supplied a surface.
    if not any(f.rule_id == 'surface.none_defined'
               for f in pf.preflight(sheetless, 'lem',
                                     {'surface': 'circular'}).errors):
        problems.append("surface.none_defined no longer fires on a model that "
                        "really defines no surface")

    # ids= isolates one rule, skip= suppresses it
    one = pf.preflight(broken, 'lem', {'surface': 'circular'},
                       ids=['main.seismic_missing'])
    if len(one.findings) != 1 or one.findings[0].rule_id != 'main.seismic_missing':
        problems.append("ids= did not isolate a single rule")
    none_ = pf.preflight(broken, 'lem', {'surface': 'circular'},
                         skip=['main.seismic_missing'])
    if any(f.rule_id == 'main.seismic_missing' for f in none_.findings):
        problems.append("skip= did not suppress the rule")

    if problems:
        return None, f"{len(problems)} problem(s): " + "; ".join(problems[:6])
    return 0.0, None


# ===========================================================================
# Remedies (xslope.remedies) -- the offered-never-silent contract
#
# A remedy changes the model, so every clause of its contract is a claim that
# has to be falsifiable: it resolves the finding it is offered for, it says
# beforehand exactly what it will do, it leaves the caller's model untouched, and
# it declines rather than half-applying. Each is asserted below on a model broken
# through the real Excel writer and loader, so the remedy is exercised on a file a
# user could have saved.
#
# The strongest assertion here is the water audit: on a model whose transcribed
# reservoir load is removed, the DERIVED load must reproduce the removed one --
# same resultant, to a fraction of a percent. That is the equivalence the
# automatic water-load mode will rest on, measured now on the file that carries
# both halves.
# ===========================================================================

def _remedy_excel(path, mutate, template):
    """Load a sample file, mutate it, and round-trip through the writer/loader."""
    import tempfile
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx
    sd = mutate(load_slope_data(path))
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    try:
        save_slope_data_to_xlsx(sd, tmp, template=template)
        return load_slope_data(tmp)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)


def _dload_resultant(blocks):
    """Total distributed load per unit width -- the quantity the slicer applies."""
    import math
    total = 0.0
    for blk in blocks or []:
        for a, b in zip(blk, blk[1:]):
            total += 0.5 * (a['Normal'] + b['Normal']) * math.hypot(b['X'] - a['X'],
                                                                    b['Y'] - a['Y'])
    return total


def run_preflight_remedies_test(test):
    """The remedy contract: offered, described first, pure, and declining cleanly."""
    import warnings as _warnings
    from xslope.fileio import load_slope_data
    from xslope import preflight as pf
    from xslope import remedies as rm

    template = test.get('template', ROUNDTRIP_TEMPLATE)
    problems = []

    # -- the registry ----------------------------------------------------
    for name in rm.implemented():
        if name not in pf.REMEDIES:
            problems.append(f"{name}: implemented but not declared in REMEDIES")
    for r in pf.rules():
        for name in r.remedies:
            if name not in pf.REMEDIES:
                problems.append(f"{r.id}: offers unknown remedy {name!r}")
    try:
        rm.remedy_proposals({}, ['set_seismic_zero'])
        problems.append("a declared-but-unimplemented remedy silently did nothing")
    except ValueError:
        pass
    try:
        rm.remedy_proposals({}, ['no_such_remedy'])
        problems.append("an unknown remedy name was accepted")
    except ValueError:
        pass

    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        # Both water remedies are about a load a USER entered — one offers to add
        # it, the other to hand it back to the engine — so the fixture is the base
        # model in the pre-v22 idiom: manual mode, reservoir typed on the sheet.
        # Built rather than borrowed, since the corpus states its water as a
        # definition now (see _preflight_apply's manual_water).
        base = _legacy_manual_water(load_slope_data(PREFLIGHT_BASE_LEM))
        _manual = lambda sd: _legacy_manual_water(sd)          # noqa: E731

        # -- add_ponded_water_load, through the Excel path ----------------
        transcribed = _dload_resultant(base['dloads'])
        stripped = _remedy_excel(PREFLIGHT_BASE_LEM,
                                 lambda sd: dict(_legacy_manual_water(sd),
                                                 dloads=[], dload_dirs=[]),
                                 template)
        before = pf.preflight(stripped, 'lem', {'surface': 'circular'})
        if not any(f.rule_id == 'water.ponded_no_dload' for f in before.findings):
            problems.append("the stripped model did not report its missing water load")
        proposal = rm.propose(stripped, 'add_ponded_water_load', 'stage1')
        if not proposal.available:
            problems.append(f"the water remedy declined on a model that needs it: "
                            f"{proposal.reason}")
        else:
            after = pf.preflight(stripped, 'lem', {'surface': 'circular'},
                                 remedies=['add_ponded_water_load'])
            if any(f.rule_id == 'water.ponded_no_dload' for f in after.findings):
                problems.append("the applied water remedy did not resolve its finding")
            info = [f for f in after.findings
                    if f.rule_id == 'remedy.add_ponded_water_load'
                    and f.severity == pf.INFO]
            if not info:
                problems.append("applying a remedy produced no INFO finding")
            elif proposal.description not in info[0].message:
                problems.append("the applied INFO does not carry the description the "
                                "proposal computed beforehand")
            if stripped['dloads']:
                problems.append("the remedy mutated the caller's model")
            derived = _dload_resultant(after.model['dloads'])
            if transcribed <= 0:
                problems.append("the sample model carries no transcribed water load to "
                                "audit the derivation against")
            elif abs(derived - transcribed) > 1e-3 * transcribed:
                problems.append(
                    f"the derived water load does not reproduce the transcribed one: "
                    f"{derived:.6g} against {transcribed:.6g}")

        # a model already carrying the load must not be offered a second one
        again = rm.propose(base, 'add_ponded_water_load', 'stage1')
        if again.available:
            problems.append("the water remedy offered to add a load over a reach that "
                            "already carries one (the double count)")
        # nor may it in automatic mode
        auto = rm.propose(dict(stripped, water_loads='auto'),
                          'add_ponded_water_load', 'stage1')
        if auto.available or 'auto' not in auto.reason:
            problems.append("the water remedy did not stand down in automatic mode")

        # -- reverse_polyline --------------------------------------------
        pts = [(float(x), float(y)) for x, y in
               getattr(base['piezo_line'], 'coords', base['piezo_line'])]
        rev = _remedy_excel(PREFLIGHT_BASE_LEM,
                            lambda sd: dict(sd, piezo_line=list(reversed(pts))),
                            template)
        if not any(f.rule_id == 'order.piezo_reversed'
                   for f in pf.preflight(rev, 'lem', {'surface': 'circular'}).findings):
            problems.append("a right-to-left piezometric line was not reported")
        fixed = pf.preflight(rev, 'lem', {'surface': 'circular'},
                             remedies=['reverse_polyline'])
        if any(f.rule_id == 'order.piezo_reversed' for f in fixed.findings):
            problems.append("the applied reversal did not resolve its finding")
        got = [(float(x), float(y)) for x, y in
               getattr(fixed.model['piezo_line'], 'coords', fixed.model['piezo_line'])]
        if got != pts:
            problems.append("the reversed line did not come back in its original order")

        # a line that rises then falls is NOT a reversed line: decline, do not sort
        mixed = dict(base, piezo_line=[(pts[0][0], pts[0][1]),
                                       (pts[-1][0], pts[-1][1]),
                                       (0.5 * (pts[0][0] + pts[-1][0]), pts[0][1]),
                                       (pts[-1][0] + 1.0, pts[-1][1])])
        m_prop = rm.propose(mixed, 'reverse_polyline', 'piezo1')
        if m_prop.available:
            problems.append("the reversal remedy offered to reverse a non-monotonic line")
        try:
            rm.apply_remedy(mixed, 'reverse_polyline', 'piezo1')
            problems.append("apply_remedy applied a remedy it had declined")
        except rm.RemedyDeclined as exc:
            if str(exc) != m_prop.reason:
                problems.append("the decline and the dimming reason differ")

        # -- generate_starting_circles ------------------------------------
        no_surface = _remedy_excel(PREFLIGHT_BASE_LEM,
                                   lambda sd: dict(sd, circles=[], circular=False),
                                   template)
        gate = pf.preflight(no_surface, 'lem', {'surface': 'circular'})
        if not any(f.rule_id == 'surface.none_defined' for f in gate.errors):
            problems.append("a model with no failure surface was not refused")
        elif gate.errors[0].remedy != 'generate_starting_circles':
            problems.append("the no-surface error does not offer the generator")
        made = pf.preflight(no_surface, 'lem', {'surface': 'circular'},
                            remedies=['generate_starting_circles'])
        if any(f.rule_id == 'surface.none_defined' for f in made.errors):
            problems.append("the generated circles did not satisfy the surface rule")
        if no_surface['circles']:
            problems.append("the generator remedy mutated the caller's model")
        if not made.model['circles']:
            problems.append("the generator remedy produced no circles")
        # The same rule offers a SECOND repair. A Finding carries one remedy name --
        # the primary -- so the registry is what ties the other one back to the rule
        # it belongs beside, and the INFO an applied remedy writes names that rule
        # from here. Assert the link, not the finding field.
        if 'surface.none_defined' not in rm.rule_ids_for(
                'generate_noncircular_surface'):
            problems.append("the no-surface rule does not offer the weak-zone "
                            "generator alongside the circles one")

        # -- generate_noncircular_surface ---------------------------------
        # Same fault, second repair: where a weak layer controls the mechanism, no
        # circle can follow the seam. Offered only where the generator picks a zone
        # ITSELF -- the ambiguous model below is the gate, and it is the clause that
        # would otherwise rot, because a remedy that offered on an ambiguous model
        # would look right in every way except that its "exact change" was a guess.
        seam = dict(load_slope_data(PREFLIGHT_NONCIRC_SEAM),
                    circles=[], non_circ=[], circular=False)
        seam_gate = pf.preflight(seam, 'lem')
        if not any(f.rule_id == 'surface.none_defined' for f in seam_gate.errors):
            problems.append("the weak-seam model with both sheets emptied was not "
                            "refused")
        seam_prop = rm.propose(seam, 'generate_noncircular_surface')
        if not seam_prop.available:
            problems.append(f"the weak-zone remedy declined on a model with a clearly "
                            f"weakest zone: {seam_prop.reason}")
        else:
            tracked = pf.preflight(seam, 'lem',
                                   remedies=['generate_noncircular_surface'])
            if any(f.rule_id == 'surface.none_defined' for f in tracked.errors):
                problems.append("the generated surface did not satisfy the surface rule")
            if seam['non_circ']:
                problems.append("the weak-zone remedy mutated the caller's model")
            pts = tracked.model['non_circ']
            if not pts:
                problems.append("the weak-zone remedy produced no surface")
            elif any(p.get('Y') is None or p.get('Y') != p.get('Y')
                     or not str(p.get('Movement') or '').strip() for p in pts):
                problems.append("a generated surface point carries a blank Y or "
                                "Movement (the two blanks the slicer cannot read)")
            info = [f for f in tracked.findings
                    if f.rule_id == 'remedy.generate_noncircular_surface']
            if not info:
                problems.append("applying the weak-zone remedy produced no INFO")
            elif seam_prop.description not in info[0].message:
                problems.append("the weak-zone INFO does not carry the description "
                                "the proposal computed beforehand")

        # THE GATE. Two comparable candidate zones is a question, not a fault, and a
        # remedy cannot ask one -- it has to state its exact change before applying
        # it. So the remedy stands down entirely here and Studio's zone picker takes
        # the case. Offering it (available OR dimmed) fails this.
        ambiguous = dict(load_slope_data(PREFLIGHT_NONCIRC_AMBIGUOUS),
                         circles=[], non_circ=[], circular=False)
        from xslope.generators import generate_noncircular_surface as _gen
        amb = _gen(ambiguous, report=True)
        if amb['zone'] is not None or len(amb['candidates']) < 2:
            problems.append("the ambiguity fixture is no longer ambiguous: the "
                            "generator picked a zone, so the gate below proves nothing")
        offered = rm.remedy_proposals(ambiguous, ['generate_noncircular_surface'])
        if offered:
            problems.append(
                f"the weak-zone remedy was offered on a model with no clearly weakest "
                f"zone ({offered[0].key}, available={offered[0].available}); that "
                f"case belongs to the zone picker, not to a remedy")
        try:
            rm.apply_remedy(ambiguous, 'generate_noncircular_surface')
            problems.append("the weak-zone remedy applied itself on an ambiguous model")
        except rm.RemedyDeclined:
            pass
        # ...and a one-material model has no weak layer to track at all.
        lone = dict(load_slope_data(PREFLIGHT_NONCIRC_ONE_ZONE),
                    circles=[], non_circ=[], circular=False)
        if rm.remedy_proposals(lone, ['generate_noncircular_surface']):
            problems.append("the weak-zone remedy was offered on a single-material "
                            "model, which has no weak layer for a surface to track")

        # -- capability gating: the same answer, in the same words --------
        caps = rm.remedy_capabilities(stripped)
        for p in rm.remedy_proposals(stripped):
            cap = caps.get(p.key)
            if cap is None:
                problems.append(f"remedy_capabilities() omitted {p.key}")
            elif bool(cap.available) != p.available or cap.reason != p.reason:
                problems.append(f"{p.key}: the capability and the proposal disagree")
            if not p.available and not p.reason:
                problems.append(f"{p.key}: dimmed with no reason string")
            if p.available and not p.description:
                problems.append(f"{p.key}: offered with nothing to show the user")

    if problems:
        return None, f"{len(problems)} problem(s): " + "; ".join(problems[:6])
    return 0.0, None


# ===========================================================================
# Automatic water loads (v22 main!D23, xslope.water)
#
# Two properties, and the second is the one that would be expensive to discover
# late:
#
#   MANUAL MODE CHANGES NOTHING. Every model that does not say otherwise -- which
#   is every file at template v21 and earlier -- must solve exactly as it did
#   before the engine could derive water at all. The check is identity of the
#   model object plus absence of the derived keys, because that is what makes the
#   downstream arithmetic bit-identical rather than merely close.
#
#   AUTO MODE IS THE SAME ANALYSIS, EXPRESSED DIFFERENTLY. A file whose water is
#   transcribed and the same file with those blocks removed and D23 = auto must
#   produce the same factor of safety, the same slice-level load resultants and
#   the same finite-element tractions. That is the equivalence the corpus
#   migration rests on, checked here on one file per class so a regression in the
#   derivation is caught by the routine suite rather than by the migration.
# ===========================================================================

#: ``(file, what it exercises)``. One per water source and per consumer: a
#: seep-BC pool, a piezometric pool, a two-stage rapid-drawdown deck, and a FEM
#: model whose reservoir becomes tractions.
AUTO_WATER_CASES = [
    (_repo('docs/inputs/slope/xslope_dam.xlsx'), 'seep bc pool, LEM'),
    (_repo('docs/verification/files/rocscience/vp042.xlsx'), 'piezometric pool, LEM'),
    (_repo('docs/verification/files/rocscience/vp096.xlsx'), 'rapid drawdown, both stages'),
    (_repo('docs/fem/files/xslope_griffiths5_0p5.xlsx'), 'reservoir as FEM tractions'),
]

#: How far an automatic model may differ from the manual one it replaces. The
#: derivation samples at the union of the water line's and the ground's vertices
#: and tapers at the crossing, so a transcribed block need not agree to the last
#: bit; the measured spread across the pilot classes was 0 to 1.3e-7 relative,
#: which is what makes 1e-6 a fence rather than a fitted number.
AUTO_WATER_TOL = 1e-6


def _auto_water_lem(sd):
    """``(FS, total |load| on the slices)`` for a model, on its own first surface."""
    from xslope.slice import generate_slices
    from xslope import solve as _solve
    circ = (sd.get('circles') or [None])[0]
    if circ is not None:
        ok, res = generate_slices(sd, circle=circ, num_slices=40, debug=False,
                                  check_inputs=False)
    else:
        ok, res = generate_slices(sd, non_circ=sd.get('non_circ'), num_slices=40,
                                  debug=False, check_inputs=False)
    if not ok:
        return None, None, f"slices failed: {str(res)[:60]}"
    df = res[0] if isinstance(res, tuple) else res
    load = float(df['dload'].abs().sum())
    if 'dload2' in df.columns:
        load += float(df['dload2'].abs().sum())
    for method in ('spencer', 'bishop', 'oms'):
        try:
            good, out = getattr(_solve, method)(df)
        except Exception:
            continue
        if good:
            return float(out['FS']), load, None
    return None, load, 'no method converged'


def _auto_water_mesh(sd):
    """A mesh for a model, built once and shared by both variants of it.

    Shared deliberately: the mesher is free to produce a different (equally valid)
    triangulation for the same geometry, and comparing tractions node by node only
    means something on ONE mesh.
    """
    from xslope.mesh import (build_mesh_from_polygons, get_material_polygons,
                             extract_constraint_line_geometry,
                             extract_point_constraints, extract_size_regions)
    xs = [x for x, _ in sd['ground_surface'].coords]
    lines, _r, _p = extract_constraint_line_geometry(sd)
    return build_mesh_from_polygons(get_material_polygons(sd, reinf_lines=lines),
                                    target_size=(max(xs) - min(xs)) / 40,
                                    element_type='tri3', lines=lines,
                                    point_constraints=extract_point_constraints(sd),
                                    size_regions=extract_size_regions(sd))


def _auto_water_fem(sd, mesh):
    """The assembled nodal force vector for a model on a given mesh."""
    import numpy as np
    from xslope.fem import build_fem_data
    fd = build_fem_data(sd, mesh)
    return np.asarray(fd['bc_type']), np.asarray(fd['bc_values'], dtype=float)


def _auto_water_geometry(problems):
    """The derivation's geometry, on shapes whose answer is closed-form.

    Four properties the Wave 4b corpus audit turned up as defects, each checked
    against arithmetic rather than a stored number, and each on the smallest shape
    that shows it:

    * **A vertical face carries its water.** A wall standing in a pool is loaded
      hydrostatically over its wetted height. Sampling the ground by its x values
      instead of its vertices loses the wall and draws a fictitious diagonal to the
      next vertex -- worth 13.8% of the total on the shape below, and worth vp092's
      whole disagreement with its own transcription.
    * **The pool ends at the shoreline.** Not at the next ground vertex above it:
      that runs the load up the dry part of the face, and on a boundary block drawn
      up a whole face (gw017, gw018, vp038) it over-measures the reservoir threefold.
    * **A hairline is not a pool.** Two millimetres between a water line and the
      ground it was drawn to meet is the coordinates' precision, and is dropped --
      but a real shallow pool an order of magnitude deeper is kept, so the fence is
      checked from both sides.
    * **A vertical face between two pools is reported.** The water surface is a
      function of x and cannot hold two elevations at one station, so the derivation
      says so instead of quietly loading the face to the higher one.
    """
    from shapely.geometry import LineString
    from xslope import water

    def model(ground, **kw):
        sd = {'ground_surface': LineString(ground), 'gamma_water': 10.0,
              'piezo_line': [], 'seepage_bc': {}, 'seepage_bc2': {}, 'dloads': []}
        sd.update(kw)
        return sd

    def res(sd, stage=1):
        d = water.derive_water_loads(sd, stage)
        return d, sum(water.block_resultant(b) for b in d['blocks'])

    # -- a 10 m wall standing in 10 m of water, on a 10 m bed --------------
    # bed: 10 m of 100 kPa = 1000; wall: hydrostatic triangle 0.5 x 100 x 10 = 500
    _d, got = res(model([(0, 0), (10, 0), (10, 10), (20, 10)],
                        piezo_line=[(-1, 10), (21, 10)]))
    if abs(got - 1500.0) > 1e-6:
        problems.append(f"a vertical face in a pool: expected 1500 (1000 on the bed "
                        f"+ 500 on the wall), derived {got:.6g}")

    # -- 5 m of water against a 45 deg face, the boundary drawn up all of it
    # 0.5 x gamma x h x (h / sin 45) = 0.5 x 10 x 5 x 7.0711 = 176.78
    _d, got = res(model([(0, 0), (20, 20)], seepage_bc={'specified_heads': [
        {'head': 5.0, 'coords': [(0, 0), (20, 20)]}]}))
    if abs(got - 176.7767) > 1e-3:
        problems.append(f"the pool must end at the shoreline: expected 176.78 on the "
                        f"wetted 5 m of a 45 deg face, derived {got:.6g}")

    # -- 2 mm of water on a 10 m section is round-off; 100 mm is a pool ----
    d, got = res(model([(0, 10), (50, 9.998), (60, 20), (100, 20)],
                       piezo_line=[(-1, 10), (101, 10)]))
    if d['blocks'] or got:
        problems.append(f"a 2 mm hairline was carried as a pool ({got:.6g})")
    if not d['negligible']:
        problems.append("a hairline pool was dropped without saying so")
    d, got = res(model([(0, 10), (50, 9.9), (60, 20), (100, 20)],
                       piezo_line=[(-1, 10), (101, 10)]))
    if not d['blocks'] or abs(got - 25.0) > 0.5:
        problems.append(f"a real 100 mm pool was dropped as round-off: derived "
                        f"{got:.6g} against 25")

    # -- two pools, a wall between them: reported, not guessed -------------
    d, _got = res(model([(0, 10), (10, 10), (10, 2), (20, 2)],
                        seepage_bc={'specified_heads': [
                            {'head': 10.0, 'coords': [(0, 10), (10, 10)]},
                            {'head': 4.0, 'coords': [(20, 2)]}]}))
    if not d['ambiguous']:
        problems.append("a vertical face between two pools at different elevations "
                        "was loaded without saying which pool was used")
    d, _got = res(model([(0, 0), (10, 0), (10, 10), (20, 10)],
                        piezo_line=[(-1, 10), (21, 10)]))
    if d['ambiguous']:
        problems.append("a section with one pool reported a two-pool ambiguity")

    # -- the block carries its shape, not its sampling ---------------------
    # A water surface that bends over a straight reach of ground leaves a point
    # on the straight line between its neighbours. It changes no force, but a
    # dload vertex places a slice boundary and seeds a mesh node, so a derived
    # block sampled more finely than the load it describes discretises the model
    # differently from the transcribed block it replaces.
    d, got = res(model([(0, 0), (30, 0)],
                       piezo_line=[(-1, 10), (10, 10), (20, 10), (31, 10)]))
    if len(d['blocks']) != 1 or len(d['blocks'][0]) != 2:
        shape = [len(b) for b in d['blocks']]
        problems.append(f"a flat pool on flat ground is two points; the derivation "
                        f"emitted {shape} -- the water line's own vertices are "
                        f"sampling, not shape")
    if abs(got - 3000.0) > 1e-6:
        problems.append(f"dropping redundant vertices changed the load: {got:.6g} "
                        f"against 3000")
    # a real bend survives: the pool deepens where the ground steps down
    d, _got = res(model([(0, 0), (10, 0), (10, -5), (30, -5)],
                        piezo_line=[(-1, 10), (31, 10)]))
    if len(d['blocks']) != 1 or len(d['blocks'][0]) != 4:
        shape = [len(b) for b in d['blocks']]
        problems.append(f"a step in the ground is a bend in the pressure diagram and "
                        f"must survive; the derivation emitted {shape}")


def run_auto_water_test(test):
    """Manual mode changes nothing; automatic mode is the same analysis."""
    import tempfile
    import warnings as _warnings
    import numpy as np
    from xslope.fileio import load_slope_data, save_slope_data_to_xlsx
    from xslope import water, remedies as rm, preflight as pf

    template = test.get('template', ROUNDTRIP_TEMPLATE)
    problems = []

    def rel(a, b):
        if a is None or b is None:
            return float('inf')
        return abs(b - a) / abs(a) if a else abs(b)

    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')

        # -- the mode, through the sheet ---------------------------------
        # The pre-v22 file is SYNTHESIZED at v21 from the same model rather than
        # borrowed from the corpus, which now ships v22 throughout: the claim
        # being tested is about a file that has no D23 cell at all, and that has
        # to stay testable after the migration. Archived templates never move.
        with tempfile.TemporaryDirectory() as _td:
            _old = provision_legacy_file(AUTO_WATER_CASES[0][0], 21,
                                         os.path.join(_td, 'v21_case0.xlsx'))
            _old_sd = load_slope_data(_old)
            if _old_sd.get('template_version') and int(float(
                    _old_sd['template_version'])) >= 22:
                problems.append("the pre-v22 fixture was not written at v21")
            if _old_sd.get('water_loads') != 'manual':
                problems.append(f"a pre-v22 file loaded as "
                                f"{_old_sd.get('water_loads')!r}; every file that "
                                f"predates the cell means manual")
            saved = _remedy_excel(_old, lambda sd: sd, template)
            if saved.get('water_loads') != 'manual':
                problems.append("saving a manual model into the v22 template let the "
                                "template's own 'auto' leak in — every file that "
                                "round-trips would double its reservoir")
        base = _legacy_manual_water(load_slope_data(AUTO_WATER_CASES[0][0]))
        auto_rt = _remedy_excel(AUTO_WATER_CASES[0][0],
                                lambda sd: dict(sd, water_loads='auto'), template)
        if auto_rt.get('water_loads') != 'auto':
            problems.append("main!D23 = auto did not survive save -> load")
        # an unknown word is refused, never silently defaulted
        try:
            _remedy_excel(AUTO_WATER_CASES[0][0],
                          lambda sd: dict(sd, water_loads='sometimes'), template)
        except Exception as exc:
            if 'D23' not in str(exc):
                problems.append(f"an unknown water-load mode was refused without "
                                f"naming the cell: {str(exc)[:80]}")
        else:
            problems.append("an unknown water-load mode was accepted")

        # -- manual mode is inert ----------------------------------------
        if water.with_water_loads(base) is not base:
            problems.append("a manual model was copied by with_water_loads; manual "
                            "mode must hand back the caller's own model")
        if water.has_derived_loads(water.with_water_loads(base)):
            problems.append("a manual model acquired derived water loads")

        # -- auto == manual, per class -----------------------------------
        # Each case is taken in its pre-v22 expression — the water typed onto the
        # dloads sheet — and the remedy is asked to convert it, which is exactly
        # the conversion the corpus migration performed on the checked-in file.
        for path, what in AUTO_WATER_CASES:
            sd = _legacy_manual_water(load_slope_data(path))
            try:
                proposal = rm.propose(sd, 'switch_to_auto_water')
            except rm.RemedyDeclined as exc:
                problems.append(f"{os.path.basename(path)} ({what}): the "
                                f"switch-to-auto remedy found nothing to do: {exc}")
                continue
            if not proposal.available:
                problems.append(f"{os.path.basename(path)} ({what}): switch-to-auto "
                                f"declined: {proposal.reason[:100]}")
                continue
            flipped, _info = rm.apply_remedy(sd, 'switch_to_auto_water')
            if sd['dloads'] == []:
                problems.append(f"{os.path.basename(path)}: the remedy mutated the "
                                f"caller's model")
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
            try:
                save_slope_data_to_xlsx(flipped, tmp, template=template)
                auto = load_slope_data(tmp)
            finally:
                if os.path.exists(tmp):
                    os.unlink(tmp)
            if auto.get('water_loads') != 'auto':
                problems.append(f"{os.path.basename(path)}: the flipped file came "
                                f"back {auto.get('water_loads')!r}")
            fs_m, load_m, err_m = _auto_water_lem(sd)
            fs_a, load_a, err_a = _auto_water_lem(auto)
            if err_m or err_a:
                problems.append(f"{os.path.basename(path)}: {err_m or err_a}")
            else:
                if rel(fs_m, fs_a) > AUTO_WATER_TOL:
                    problems.append(f"{os.path.basename(path)} ({what}): FS moved "
                                    f"{fs_m:.6f} -> {fs_a:.6f} on the mode change")
                if rel(load_m, load_a) > AUTO_WATER_TOL:
                    problems.append(f"{os.path.basename(path)} ({what}): slice load "
                                    f"resultants moved {load_m:.6g} -> {load_a:.6g}")
            if 'fem' in path:
                mesh = _auto_water_mesh(sd)
                t_m, v_m = _auto_water_fem(sd, mesh)
                t_a, v_a = _auto_water_fem(auto, mesh)
                if not np.array_equal(t_m, t_a):
                    problems.append(f"{os.path.basename(path)}: the automatic model "
                                    f"loaded different NODES than the manual one")
                scale = float(np.abs(v_m).sum()) or 1.0
                if float(np.abs(v_m - v_a).sum()) / scale > AUTO_WATER_TOL:
                    problems.append(f"{os.path.basename(path)}: FEM tractions differ "
                                    f"between the manual and automatic model")
            # -- a derived load seeds the mesh, exactly as a typed one does --
            # Without this the same reservoir on the same geometry would mesh
            # differently under the two modes: the mesher would put no node at the
            # load's own vertices, and the FEM would apply a traction to element
            # edges that were never aligned to it.
            from xslope.mesh import get_material_polygons
            derived_model = water.with_water_loads(auto)
            seeded = set()
            for poly in get_material_polygons(auto):
                g = poly.get('coords', poly.get('polygon')) if isinstance(poly, dict) else poly
                cs = list(g.exterior.coords) if hasattr(g, 'exterior') else list(g)
                seeded |= {(round(float(x), 7), round(float(y), 7)) for x, y in cs}
            edges = []
            for poly in get_material_polygons(sd):
                g = poly.get('coords', poly.get('polygon')) if isinstance(poly, dict) else poly
                cs = list(g.exterior.coords) if hasattr(g, 'exterior') else list(g)
                edges += list(zip(cs, cs[1:]))
            for stage in (1, 2):
                for blk in water.derived_blocks(derived_model, stage):
                    for p in blk:
                        px, py = float(p['X']), float(p['Y'])
                        on_edge = any(
                            abs((bx - ax) * (py - ay) - (by - ay) * (px - ax)) < 1e-6
                            and min(ax, bx) - 1e-9 <= px <= max(ax, bx) + 1e-9
                            and min(ay, by) - 1e-9 <= py <= max(ay, by) + 1e-9
                            for (ax, ay), (bx, by) in edges)
                        if on_edge and (round(px, 7), round(py, 7)) not in seeded:
                            problems.append(
                                f"{os.path.basename(path)}: the derived water load's "
                                f"vertex at ({px:.6g}, {py:.6g}) lies on a polygon edge "
                                f"but was not seeded into the mesh")
                            break

            # the user's own non-water loads pass through auto mode untouched
            keep = [b for i, b in enumerate(sd['dloads'])
                    if i not in {i for i, _j, _m in water.match_water_blocks(
                        sd['dloads'], water.derive_water_loads(sd, 1)['blocks'])['pairs']}]
            if flipped['dloads'] != keep:
                problems.append(f"{os.path.basename(path)}: switching to auto did not "
                                f"preserve the non-water blocks verbatim")

        # -- the double count is reported, not silently carried ----------
        doubled = dict(base, water_loads='auto')       # blocks kept AND derived
        report = pf.preflight(doubled, 'lem', {'surface': 'circular'})
        if not any(f.rule_id == 'water.auto_dload_double_count'
                   for f in report.findings):
            problems.append("an auto-mode model still carrying its transcribed water "
                            "load was not warned about the double count")

        # -- visibility: a derived load is drawn, in its own colour ------
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from xslope.plot import plot_dloads
        from xslope.style import feature_style, resolve_style
        colour = feature_style(resolve_style(None), 'dloads_derived')['color']
        flipped, _ = rm.apply_remedy(base, 'switch_to_auto_water')
        fig, ax = plt.subplots()
        try:
            plot_dloads(ax, flipped)
            drawn = [ln for ln in ax.lines
                     if matplotlib.colors.same_color(ln.get_color(), colour)]
            if not drawn:
                problems.append("a model whose only load is derived drew no water "
                                "load at all — automatic must never mean invisible")
            if any(matplotlib.colors.same_color(
                    ln.get_color(), feature_style(resolve_style(None),
                                                  'dloads')['color'])
                   for ln in ax.lines):
                problems.append("a derived water load was drawn as though the user "
                                "had typed it")
        finally:
            plt.close(fig)

        # -- and it is not editable: derived loads are not user data -----
        try:
            from studio.picking import pick_category
            derived = water.with_water_loads(flipped)
            for blk in water.derived_blocks(derived, 1):
                pt = blk[len(blk) // 2]
                hit = pick_category(derived, pt['X'], pt['Y'], tol=1e-6)
                if hit and hit[0] == 'dloads':
                    problems.append("a derived water load is click-selectable in "
                                    "Studio; there is no sheet row to edit")
                    break
        except ImportError:                       # studio not importable: skip
            pass

    # -- the derivation's geometry, on shapes with closed-form answers ----
    _auto_water_geometry(problems)

    if problems:
        return None, f"{len(problems)} problem(s): " + "; ".join(problems[:6])
    return 0.0, None


# ===========================================================================
# Starting-surface generator (xslope.generators)
#
# The generated set is only useful if a search can be handed it, so the checks
# are the ones a search would impose: every circle daylights on the ground
# INSIDE the model, stays inside the domain polygon, and is accepted by
# generate_slices. The spread covers a single face, a dam (two faces, one
# crest), a layered foundation and a benched dam.
#
# The benched dam also carries the reason the skimming rule exists: on a
# cohesionless face the critical surface is a shallow face-parallel slide whose
# factor of safety is tan(phi)/tan(beta), independent of depth, and a set seeded
# only with toe and base circles cannot reach it. The generated skimming circle
# is required to land on that closed form.
#
# It is then required to SURVIVE being handed to the search, which is a separate
# fact and was separately wrong. A skimming circle's `Depth` is the elevation of
# its nadir — far outside the model, well below the domain floor — while the
# surface it cuts stays inside; the depth optimiser used to clamp that seed up to
# the floor and turn the skin into an ordinary toe circle. On Talbingo the skim
# solves directly to the closed form and the search that was handed it came back
# 8.6% higher. So a search seeded with the generated set may never return a worse
# factor of safety than the best circle in that set already solves to.
# ===========================================================================

#: ``(file, expected face count, must carry a skimming circle)``.
GENERATOR_CASES = [
    (_repo('docs/lem/files/xslope_simple_embankment.xlsx'), 1, False),
    (_repo('docs/lem/files/xslope_simple_mult_layers.xlsx'), 1, False),
    (_repo('docs/lem/files/xslope_eight_layers.xlsx'), 1, False),
    (_repo('docs/inputs/slope/xslope_dam.xlsx'), 2, True),
    (_repo('docs/verification/files/rocscience/vp005.xlsx'), 2, True),
]


def run_generator_circles_test(test):
    """Every generated circle must be one a search could actually be handed."""
    import math
    import warnings as _warnings
    from xslope.fileio import load_slope_data
    from xslope.generators import (generate_starting_circles, slope_geometry,
                                   _analysed_span, _skimming_circle)
    from xslope.slice import generate_slices
    from xslope import solve

    problems = []
    with _warnings.catch_warnings():
        _warnings.simplefilter('ignore')
        for path, n_faces, wants_skim in GENERATOR_CASES:
            if not os.path.exists(path):
                problems.append(f"missing sample file {path}")
                continue
            name = os.path.basename(path)
            sd = load_slope_data(path)
            geom = slope_geometry(sd)
            if len(geom.faces) != n_faces:
                problems.append(f"{name}: found {len(geom.faces)} slope face(s), "
                                f"expected {n_faces}")
            ground = list(geom.ground)
            gx0, gx1 = ground[0][0], ground[-1][0]
            circles = generate_starting_circles(sd)
            if not circles:
                problems.append(f"{name}: no starting circle was generated")
                continue
            for c in circles:
                tag = (f"{name}: circle Xo={c['Xo']:.4g} Yo={c['Yo']:.4g} "
                       f"Depth={c['Depth']:.4g}")
                if abs(c['R'] - (c['Yo'] - c['Depth'])) > 1e-6 * max(1.0, c['R']):
                    problems.append(f"{tag}: R and Depth disagree")
                span = _analysed_span(c['Xo'], c['Yo'], c['R'], ground)
                if span is None:
                    problems.append(f"{tag}: never reaches the ground surface")
                    continue
                if span[0] <= gx0 + 1e-9 or span[1] >= gx1 - 1e-9:
                    problems.append(f"{tag}: daylights at a vertical model edge")
                ok, res = generate_slices(dict(sd, circles=circles), circle=c,
                                          num_slices=20, check_inputs=False)
                if not ok:
                    problems.append(f"{tag}: rejected by generate_slices ({str(res)[:60]})")

            # the centre rule, on the tallest face
            face = geom.primary
            want_x = 0.5 * (face.toe[0] + face.crest[0])
            want_y = face.toe[1] + 2.0 * face.height
            if not any(abs(c['Xo'] - want_x) <= 1e-6 * max(1.0, abs(want_x))
                       and abs(c['Yo'] - want_y) <= 1e-6 * max(1.0, abs(want_y))
                       for c in circles):
                problems.append(f"{name}: no circle sits at the rule's own centre "
                                f"(x = {want_x:.4g}, y = {want_y:.4g})")

            if not wants_skim:
                continue

            # The generated set must survive being handed to a search. Every circle
            # is solved on its own first, so the comparison is against the set's own
            # best member rather than a locked number: a seeded search evaluates its
            # seeds, so it can only come back LOWER — unless something in the search
            # destroys a seed before solving it, which is exactly what the depth
            # clamp did to every skimming circle.
            best_seed, best_seed_c = None, None
            for c in circles:
                ok, res = generate_slices(dict(sd, circles=circles), circle=c,
                                          num_slices=40, check_inputs=False)
                if not ok:
                    continue
                good, out = solve.bishop(res[0])
                if good and out['FS'] > 0 and (best_seed is None or out['FS'] < best_seed):
                    best_seed, best_seed_c = out['FS'], c
            if best_seed is None:
                problems.append(f"{name}: no generated circle solved on its own")
            else:
                from xslope.search import circular_search
                fs_cache, _conv, _sp, _cc = circular_search(
                    dict(sd, circles=circles), 'bishop', num_slices=40, seed='circles')
                if not fs_cache or fs_cache[0]['FS'] >= 9999:
                    problems.append(f"{name}: the seeded search found no valid surface")
                elif fs_cache[0]['FS'] > best_seed * 1.01:
                    problems.append(
                        f"{name}: the search seeded with the generated set returned "
                        f"FS = {fs_cache[0]['FS']:.4f}, WORSE than the seed at "
                        f"Xo={best_seed_c['Xo']:.4g} Depth={best_seed_c['Depth']:.4g} "
                        f"which solves to {best_seed:.4f} on its own — a seed was "
                        f"destroyed before it was solved")

            skims = [c for c in circles if c['R'] > 5.0 * face.height]
            if not skims:
                problems.append(f"{name}: the cohesionless face got no skimming circle")
                continue
            # the closed form the skimming rule exists to reach
            for f in geom.faces:
                sk = _skimming_circle(*f.steepest)
                if sk is None or not any(abs(c['Xo'] - sk['Xo']) < 1e-6 * max(1.0, abs(sk['Xo']))
                                         for c in circles):
                    continue
                ok, res = generate_slices(dict(sd, circles=[sk]), circle=sk,
                                          num_slices=40, check_inputs=False)
                if not ok:
                    problems.append(f"{name}: the skimming circle does not slice")
                    continue
                good, out = solve.bishop(res[0])
                phi = max(float(m.get('phi') or 0.0) for m in sd['materials'])
                closed = (math.tan(math.radians(phi))
                          / math.tan(math.radians(f.steepest_angle)))
                if not good:
                    problems.append(f"{name}: the skimming circle did not solve")
                elif abs(out['FS'] - closed) > 0.05 * closed:
                    problems.append(
                        f"{name}: the skimming circle reads FS = {out['FS']:.4f} against "
                        f"the face-parallel closed form tan(phi)/tan(beta) = "
                        f"{closed:.4f} -- it is not skimming the face")

        # negative controls: nothing to derive from is said, not guessed
        flat = load_slope_data(GENERATOR_CASES[0][0])
        flat = dict(flat, ground_surface=[(0.0, 10.0), (100.0, 10.0)])
        out = generate_starting_circles(flat, report=True)
        if out['circles'] or 'flat' not in out['reason']:
            problems.append("a flat ground surface did not decline with a reason")
        none_ = generate_starting_circles(dict(flat, ground_surface=None), report=True)
        if none_['circles'] or not none_['reason']:
            problems.append("a model with no ground surface did not decline with a reason")

    if problems:
        return None, f"{len(problems)} problem(s): " + "; ".join(problems[:6])
    return 0.0, None


def run_fem_elastic_units_test(test):
    """No FEM material on an English-unit corpus file may carry the metric inherited
    unit-blind elastic default (E = 100,000, nu = 0.3).

    The corpus builders historically filled every FEM material with that single
    default. It is a sane METRIC value (100 MPa), but on an English-unit file (gamma
    in pcf) it is read as 100,000 *psf* ~ 4.8 MPa — ~10x too soft for the rockfill,
    sand and stiff-clay shells it lands on. The SSRM factor of safety is invariant to
    E, so the regression locks never move and no other test notices; what the wrong
    modulus corrupts is displacements, the displacement-vector figure panels, and
    reliability_fem. This is the second time the bug surfaced, so guard against it:
    the builders now assign a soil-type psf modulus via
    elastic_props.assign_elastic_props, and this check makes sure none of the metric
    default slipped back in on a rebuild.

    A magnitude floor cannot separate the leaked default from a legitimately soft
    English modulus (the FEM sample corpus carries a deliberate 60,000 psf soft clay),
    so the check keys on the exact inherited sentinel (E and nu both at the default) —
    the signature of an unmigrated file — in the file's OWN unit system.

    Scope: only files that actually run the FEM (a fem_ssrm / fem_elements /
    fem_reliability test tag). LEM-only English files legitimately leave E at the
    metric default — it is never used — so they are out of scope. Metric FEM files
    keep E = 100,000 kPa (100 MPa), which is correct, so only English files are
    checked.

    Returns (0.0, None) if clean, else (None, message).
    """
    sys.path.insert(0, str(Path(__file__).parent / 'benchmarks' / 'rocscience'))
    from elastic_props import is_imperial, INHERITED_DEFAULT, _finite
    from xslope.fileio import load_slope_data

    tag_re = re.compile(r'<!--\s*test:\s*(.*?)\s*-->')
    fem_types = {'fem_ssrm', 'fem_elements', 'fem_reliability'}
    root = Path(__file__).parent
    fem_files = set()
    for md in sorted(root.glob('docs/**/*.md')):
        for line in md.read_text().splitlines():
            mt = tag_re.search(line)
            if not mt:
                continue
            kv = {}
            for part in mt.group(1).split(','):
                if '=' in part:
                    k, v = part.split('=', 1)
                    kv[k.strip()] = v.strip()
            if kv.get('type') in fem_types and 'file' in kv:
                fem_files.add((md.parent / kv['file']).resolve())
    if not fem_files:
        return None, "no FEM test tags found — cannot verify elastic-property units"

    E0, nu0 = INHERITED_DEFAULT
    offenders = []
    for fp in sorted(fem_files):
        if not fp.exists():
            continue
        try:
            sd = load_slope_data(str(fp))
        except Exception as e:                       # pragma: no cover - defensive
            return None, f"could not load {fp.name}: {e}"
        mats = sd.get('materials', [])
        if not is_imperial(mats):
            continue                                 # metric E=100,000 kPa is 100 MPa — correct
        for m in mats:
            E, nu = _finite(m.get('E')), _finite(m.get('nu'))
            if abs(E - E0) <= 1.0 and abs(nu - nu0) <= 1e-6:
                offenders.append(f"{fp.name}:{m.get('name', '?')} (E={E:g} psf, nu={nu:g})")

    if offenders:
        return None, (
            f"English-unit FEM material carrying the metric inherited default "
            f"(E={E0:g}, nu={nu0:g}) — that is ~4.8 MPa read as psf, ~10x too soft; "
            "rebuild through the classifier-wired builder "
            "(benchmarks/rocscience/build_problems.py or build_rs2.py): "
            + "; ".join(offenders))
    return 0.0, None


def run_dload_direction_test(test):
    """No corpus distributed-load line may come back from the loader running
    right-to-left (decreasing X).

    Point order should never be able to change a load. It did: the FEM built the
    inward normal by rotating the load line's tangent 90 degrees CW, which names
    the inside only for a line written left-to-right, so RS2-67 b/e/f — whose
    downstream pool is traced from the far right boundary back toward the toe —
    had their reservoir pressure applied as UPLIFT on the submerged bench. The
    LEM is exposed the same way for a different reason: slice.py interpolates the
    load intensity with np.interp, which requires an ascending x array and
    silently returns nonsense rather than raising when it is not.

    Both paths are now defended: fem.build_fem_data takes the inward direction
    from the element that owns each loaded edge (and, in the lumped fallback,
    from the elements touching each loaded node) rather than from point order,
    and fileio.load_slope_data reverses any monotone right-to-left line at load
    time. This guard is the cheap standing check on the second one: load every
    corpus file that carries a distributed load and assert no line survives in
    decreasing-X order.

    Out of scope: lines whose X actually turns around (an overhang, or the
    near-vertical toe segment on rs2_67 b/e/f's upstream face, where X backs up
    by 1 cm). Those have no increasing-X form, so the loader deliberately leaves
    them as authored and they are not counted here.

    Returns (0.0, None) if clean, else (None, message).
    """
    from xslope.fileio import load_slope_data

    tag_re = re.compile(r'<!--\s*test:\s*(.*?)\s*-->')
    root = Path(__file__).parent
    files = set()
    for md in sorted(root.glob('docs/**/*.md')):
        for line in md.read_text().splitlines():
            mt = tag_re.search(line)
            if not mt:
                continue
            kv = {}
            for part in mt.group(1).split(','):
                if '=' in part:
                    k, v = part.split('=', 1)
                    kv[k.strip()] = v.strip()
            if 'file' in kv and kv['file'].endswith('.xlsx'):
                files.add((md.parent / kv['file']).resolve())
    if not files:
        return None, "no test tags found — cannot verify distributed-load orientation"

    n_dload_files = 0
    offenders = []
    for fp in sorted(files):
        if not fp.exists():
            continue
        try:
            sd = load_slope_data(str(fp))
        except Exception:
            continue           # load failures are another test's problem
        lines = list(sd.get('dloads') or []) + list(sd.get('dloads2') or [])
        if lines:
            n_dload_files += 1
        for i, ln in enumerate(lines):
            xs = [pt['X'] for pt in ln]
            if all(b <= a for a, b in zip(xs, xs[1:])) and xs[-1] < xs[0]:
                offenders.append(f"{fp.name} line {i + 1} (X {xs[0]:g} -> {xs[-1]:g})")

    if not n_dload_files:
        return None, "no corpus file carries a distributed load — guard is not checking anything"
    if offenders:
        return None, (
            f"distributed-load line still decreasing in X after load "
            f"({len(offenders)} of {n_dload_files} dload-carrying files): "
            + "; ".join(offenders)
            + " — fileio.load_slope_data should have reversed it")
    return 0.0, None


def run_verification_pages_test(test):
    """Standing checks on the six verification pages under docs/verification.

    Three checks run per page (tools/verification_checks): every printed
    percentage and absolute FS difference is re-derived from two numbers the
    page prints in the same sentence or table row; every value a test tag locks
    is printed in the section carrying the tag, and every value the page
    presents as locked has a tag behind it; every caption matches the figure it
    labels.

    Change-gated by tools/verification_checks/certified.json: a page whose
    content hash matches the manifest costs one file read.  A page that has
    changed is re-checked in full, and stays a failure until a developer runs
    `python -m tools.verification_checks.certify --recertify <page>` and commits
    the updated manifest with the page edit.
    """
    import io
    from contextlib import redirect_stdout
    sys.path.insert(0, str(Path(__file__).parent))
    from tools.verification_checks import certify
    buf = io.StringIO()
    with redirect_stdout(buf):
        rc = certify.main(['--quiet'])
    if not rc:
        return 0.0, None
    # something changed and did not pass: re-run verbose so the message names
    # the flags, not just the page
    buf = io.StringIO()
    with redirect_stdout(buf):
        certify.main([])
    lines = [l.strip() for l in buf.getvalue().splitlines() if l.strip()]
    head = [l for l in lines if 'FAILED' in l or 'manifest is stale' in l]
    noise = ('[precision-note]', '[precision-hedged]')
    detail = [l for l in lines
              if (l.startswith(('L', 'DEAD', 'ORPHANED'))
                  or re.match(r'L\d+ ', l))
              and not any(n in l for n in noise)]
    msg = ' | '.join(head + detail[:6])
    return None, msg or ' | '.join(lines[-4:])


def run_corpus_index_test(test):
    """Guard: the generated corpus example index is current.

    tools/make_corpus_index.py derives, from the docs' own test tags and the
    model files those tags name, both docs/verification/corpus_index.json and
    the TOPIC -> EXAMPLES table inside the skill's ``corpus-index`` markers. The
    assistant cites that table when a question matches a topic, so it has to
    keep pace with the corpus: a model added to a page, a support type changed
    in a reinforce sheet, or a heading retitled all move the answer.

    This row regenerates both artifacts in memory and fails if either committed
    copy differs.  The generator is deterministic (everything sorted, no
    timestamps), so a pass means the committed files are exactly what a
    regeneration would produce.
    """
    import io
    import warnings
    from contextlib import redirect_stdout
    sys.path.insert(0, str(Path(__file__).parent))
    from tools import make_corpus_index as mci

    buf = io.StringIO()
    with warnings.catch_warnings(), redirect_stdout(buf):
        warnings.simplefilter('ignore')     # corpus unit-system notes, not our business
        index = mci.build_index(verbose=False)

    stale = []
    index_text = mci.serialize(index)
    if not mci.INDEX_PATH.exists():
        stale.append('docs/verification/corpus_index.json is missing')
    elif mci.INDEX_PATH.read_text(encoding='utf-8') != index_text:
        stale.append('docs/verification/corpus_index.json is stale')

    try:
        block = mci.render_table(index)
        skill_text = mci.SKILL_PATH.read_text(encoding='utf-8')
        if mci.inject(skill_text, block) != skill_text:
            stale.append("the skill's corpus-index block is stale")
    except SystemExit as exc:
        stale.append(str(exc))

    if index['unloadable']:
        names = ', '.join(u['model'] for u in index['unloadable'][:3])
        stale.append(f"{len(index['unloadable'])} model(s) no longer load: {names}")

    if stale:
        return None, '; '.join(stale) + ' — run: python tools/make_corpus_index.py'
    return 0.0, None


def run_mesh_elements_test(test):
    """Lock a published MESH SIZE: the element (and node) count of one model at
    one target size, built and counted, with nothing solved.

    Several verification rows quote XSLOPE's own discretization beside the vendor
    model's — "2 732 elements against the vendor model's 2 204" — to say which of
    the two comparisons is the finer. That is a claim about the mesher, and until
    now nothing checked it: a mesher change moved every one of those counts and
    left the pages asserting stale numbers, because the only rows that touch these
    files are strength-reduction locks that cost minutes each and are tolerant to
    a percent or two of mesh drift.

    This row costs a mesh build — seconds — and is exact. It reuses the same
    ``build_mesh_from_polygons`` call ``build_fem_ssrm_case`` makes, from the same
    ``element_type`` / ``target_size`` / refinement keys, so a count locked here is
    the count the matching SSRM row meshes on and the two cannot drift apart.
    """
    from xslope.fileio import load_slope_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_constraint_line_geometry, extract_point_constraints,
                             extract_size_regions)

    want_el = test.get('expected_elements')
    want_nd = test.get('expected_nodes')
    if want_el is None and want_nd is None:
        return None, 'mesh_elements tag locks nothing: give expected_elements and/or expected_nodes'

    target_size = test.get('target_size')
    if target_size is None:
        return None, 'mesh_elements needs an explicit target_size (a count is meaningless without one)'

    slope_data = load_slope_data(test['file'])
    constraint_lines, _n_reinf, _n_pile = extract_constraint_line_geometry(slope_data)
    polygons = get_material_polygons(slope_data, reinf_lines=constraint_lines)
    mesh = build_mesh_from_polygons(
        polygons, target_size=target_size,
        element_type=test.get('element_type', 'tri6'),
        lines=constraint_lines,
        point_constraints=extract_point_constraints(slope_data),
        size_regions=extract_size_regions(slope_data),
        **_refine_kwargs(test)
    )

    got_el = len(mesh['elements'])
    got_nd = len(mesh['nodes'])
    wrong = []
    if want_el is not None and got_el != want_el:
        wrong.append(f"elements {got_el} != {want_el}")
    if want_nd is not None and got_nd != want_nd:
        wrong.append(f"nodes {got_nd} != {want_nd}")
    if wrong:
        return None, (f"mesh at target_size={target_size}: " + ', '.join(wrong)
                      + ' — the page prints the locked figure, so update both together')
    return 0.0, None


def run_cwd_invariant_test(test):
    """Guard: test discovery must not depend on the process working directory.

    Every fixture, corpus file and docs page the suite reads is at a fixed place
    inside the repo, so `python /abs/path/run_tests.py` from any directory must
    enumerate exactly the same rows as `python run_tests.py` from the repo root.
    The failure this guards against is silent rather than loud: a CWD-relative
    docs path that misses produces no error, just a shorter list, and a suite
    that skipped half its rows still finishes green.

    Enumerates twice through ``--list`` (discovery only — parses tags, solves
    nothing) from the repo root and from a directory that is not the repo, and
    requires byte-identical listings.
    """
    import subprocess
    import tempfile

    def enumerate_from(cwd):
        proc = subprocess.run([sys.executable, str(REPO_ROOT / 'run_tests.py'), '--list'],
                              cwd=cwd, capture_output=True, text=True, timeout=600)
        if proc.returncode != 0:
            return None, f"--list from {cwd} exited {proc.returncode}: {proc.stderr[-300:]}"
        return proc.stdout, None

    from_root, err = enumerate_from(str(REPO_ROOT))
    if err:
        return None, err
    with tempfile.TemporaryDirectory() as elsewhere:
        from_elsewhere, err = enumerate_from(elsewhere)
    if err:
        return None, err

    if from_root != from_elsewhere:
        n_root = from_root.splitlines()[0] if from_root else '(empty)'
        n_else = from_elsewhere.splitlines()[0] if from_elsewhere else '(empty)'
        root_lines = set(from_root.splitlines())
        else_lines = set(from_elsewhere.splitlines())
        missing = sorted(root_lines - else_lines)[:5]
        return None, (f"discovery depends on the working directory: repo root gives "
                      f"{n_root}, another directory gives {n_else}"
                      + (f"; sources differing: {', '.join(missing)}" if missing else ''))
    return 0.0, None


def run_docs_heading_trap_test(test):
    """Guard: no docs .md line may start with '#' followed immediately by a
    non-space, non-'#' character. Python-Markdown treats '#word' as a heading,
    so a hard-wrapped sentence beginning with a vendor model name like '#031'
    renders as a giant H1 mid-prose (caught on rs2.md, 2026-07-28). Legitimate
    headings are '# ', '## ', etc.; inline anchors {#id} are never line-initial.
    """
    import re
    from pathlib import Path
    docs = Path(__file__).parent / 'docs'
    offenders = []
    for md in sorted(docs.rglob('*.md')):
        in_fence = False
        for i, line in enumerate(md.read_text().splitlines(), 1):
            if line.lstrip().startswith('```'):
                in_fence = not in_fence
                continue
            if in_fence:
                continue
            if re.match(r'^#[^#\s]', line):
                offenders.append(f"{md.relative_to(docs.parent)}:{i}: {line[:60]}")
    if offenders:
        return None, "line-initial '#' renders as a heading: " + "; ".join(offenders[:5])
    return 0.0, None


def run_k0_level_ground_test(test):
    """The K0 initial stress must be an exact equilibrium on level ground.

    sigma'_v = -gamma' z with sigma'_h = K0 sigma'_v satisfies equilibrium under self
    weight identically on flat ground, for any K0 — so the solver has nothing to
    redistribute: it must converge on the first iteration with the mesh undisplaced
    to machine precision, reproduce the imposed field, and yield nowhere. That is the
    one configuration where the K0 answer is known in closed form, which makes it the
    standing check on the whole path: the overburden integration, the initial-stress
    load term, the addend at the yield check, the pore-pressure convention, and — via
    a fifth leg — the SSRM's in-situ equilibration, which on level ground must be an
    exact no-op.

    The check itself lives in test/k0_level_ground_check.py (file-less; it builds a
    20 x 10 m block and meshes it).

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'k0_level_ground_check.py'
    if not path.exists():
        return None, f"missing {path}"
    spec = importlib.util.spec_from_file_location('k0_level_ground_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_quad_mesh_test(test):
    """The mesh builder: element quality, delivered size, the size field, and sweeps.

    A quad mesh that comes back at a different element size than the one asked for
    is a silent discretization change, and a structured sweep that lands one node
    off is a mesh that looks right and is not. This guards both, plus the property
    that has to hold for a mesh to be reproducible at all: the same request built
    twice is the same mesh, in free-style and structured alike.

    The check itself lives in test/quad_mesh_check.py: free-style element quality
    and delivered size on seven corpus sections, determinism in both styles,
    structured sweeps on the levee, the ``setTransfiniteCurve`` node-count
    off-by-one, and the size field — no polygon edge discretised coarser than the
    requested size, a pile tip and a reinforcement line meshed at the size the
    refinement asked for, a crack tip detected only where material wraps around a
    slit, and a wedge tip held to the floor its apex angle sets.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'quad_mesh_check.py'
    if not path.exists():
        return None, f"missing {path}"
    spec = importlib.util.spec_from_file_location('quad_mesh_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_fem_1d_details_test(test):
    """The FEM 1D details dialog: gating, badges, the shared capacity envelope,
    the reload path and the PNG+CSV export (test/fem_1d_details_check.py).
    Solves two small FEM models (~6 s); rides with --fem for that reason."""
    import importlib.util
    path = Path(__file__).parent / 'test' / 'fem_1d_details_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips its Studio checks
    spec = importlib.util.spec_from_file_location('fem_1d_details_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_quad_style_dialog_test(test):
    """The Build-mesh dialog's quadrilateral style radio group.

    The style a quad mesh is built with — free-style or structured — is chosen per
    run and stored in no input file, so a broken wire between the radio group and
    ``build_mesh_from_polygons`` leaves no trace anywhere a round-trip guard could
    find it: the dialog would say "structured" and the mesher would build
    free-style, run after run, silently. This guards the group, the quad-only gate
    that hides it for a triangular element type, and the value actually reaching
    the builder.

    The check itself lives in test/quad_style_dialog_check.py; it builds no mesh,
    and skips itself cleanly when PySide6 is absent.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'quad_style_dialog_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips itself
    spec = importlib.util.spec_from_file_location('quad_style_dialog_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_mode_segments_test(test):
    """Studio's analysis-mode switch — the segmented LEM / Seepage / FEM strip.

    The mode decides what the Inputs view draws, what Run runs, and whether Build
    Mesh exists at all; it is picked dozens of times a session and recorded in no
    input file, so a strip wired to nothing would leave no trace anywhere else —
    the highlight would move and the window would not follow. This guards the
    segments themselves, the explicit metrics that keep the strip tight on Windows
    and macOS alike, the Ctrl+N shortcuts, and — against the mode handler as the
    specification — every side effect a click has to reproduce.

    The check itself lives in test/mode_segments_check.py; it opens two sample
    models and runs no analysis, and skips cleanly when PySide6 is absent.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'mode_segments_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips itself
    spec = importlib.util.spec_from_file_location('mode_segments_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_refine_thin_zones_test(test):
    """The Build-mesh dialog's thin-zone refinement toggle, and what it delivers.

    A material zone too thin for the mesh to resolve does not fail — it solves, and
    returns a factor of safety that is too high, because a shear band cannot form
    across one element. The toggle prevents that, so it is on by default, and the
    check guards the whole of it: the control, the derived per-zone size, the wire,
    and — measured on real meshes — the element rows the toggle actually puts across
    a thin band. Both element families reach it the same way, and that is measured
    too: a family-dependent mechanism would mesh one zone two ways from one
    checkbox. So is the refinement cap, together with the preflight warning that
    keeps a capped zone from being a silent under-delivery.

    The check itself lives in test/refine_thin_zones_check.py (~20 s: it builds
    small meshes, because the promise is about resolution and cannot be asserted
    from arguments alone), and skips cleanly when PySide6 is absent.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'refine_thin_zones_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips itself
    spec = importlib.util.spec_from_file_location('refine_thin_zones_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_remedy_panel_test(test):
    """Studio's model-checks panel: its remedies, and the shape it renders in.

    A rule can offer several remedies, and an empty surface sheet is the case that
    made it necessary: circles where the slope's own geometry controls the
    mechanism, a surface tracked along a seam where a weak layer does. A panel that
    renders only the primary leaves the second repair unreachable and says nothing
    about it, which is exactly the kind of silence no other check would catch. This
    guards the finding's own remedy list, the two buttons, the availability gate
    that keeps the second one off a model whose weakest zone is ambiguous, and the
    unchanged propose/confirm/apply and dimming contracts.

    It also guards the panel's PRESENTATION, which the registry cannot see: findings
    that share a rule id render as ONE entry (the line counts and names them, the
    detail states their shared explanation once), the entry still carries every
    remedy of its rule, the Run dialogs lay the controls and the checks out as
    columns, the detail follows the selection, and the notes stay behind their
    counted disclosure.

    The check itself lives in test/remedy_panel_check.py; it opens real Studio
    panels and Run dialogs offscreen, applies one remedy, and skips cleanly when
    PySide6 is absent.

    Returns (0.0, None) on success, else (None, message) -- a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'remedy_panel_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips itself
    spec = importlib.util.spec_from_file_location('remedy_panel_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_polygon_pick_test(test):
    """What a double-click inside a material zone on the Inputs canvas resolves to.

    A polygon-based file has no profile sheet -- its geometry IS the polygon sheet --
    so a zone's interior is the way into the row that defines it, and getting the row
    wrong is silent: a real editor opens on real geometry and the user edits the wrong
    zone. The check asserts the RESOLVED ROW, along with the precedence that keeps a
    line, point or vertex within tolerance ahead of the interior fallback, the
    smallest-containing-zone rule where zones nest, and that a profile-based file
    still answers with its material.

    The check itself lives in test/polygon_pick_check.py; its Qt legs open real
    Studio widgets offscreen and skip cleanly when PySide6 is absent.

    Returns (0.0, None) on success, else (None, message) -- a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'polygon_pick_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips its Studio legs
    spec = importlib.util.spec_from_file_location('polygon_pick_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_transient_seep_test(test):
    """The transient seepage solver's own lock set.

    The march's properties that no corpus row asserts on its own — the stepper, the
    storage terms, the saved-frame schedule — checked directly rather than through
    a verification page's head stations.

    The check itself lives in test/transient_seep_check.py (12 locks, ~60 s: the
    heaviest row in the --tseep group, because it solves several short marches).

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'transient_seep_check.py'
    if not path.exists():
        return None, f"missing {path}"
    spec = importlib.util.spec_from_file_location('transient_seep_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_fs_vs_time_mode_test(test):
    """The FS-versus-time mode's contract (``xslope.sensitivity.fs_vs_time``).

    A curve is read for its shape, so the two ways it can lie are both about rows
    rather than numbers: an instant that produced no result must appear as a
    ``success=False`` row CARRYING ITS REASON, and no requested instant may go
    missing. A silently dropped frame is the dangerous one — drop the instant just
    after a drawdown and the curve's minimum moves to a healthier value with
    nothing in the record to say so.

    The check itself lives in test/fs_vs_time_check.py: row coverage, the reason on
    every failed row, the sentinel/negative screen, the refusal of a field
    interpolated between frames, one re-march for the whole set rather than one per
    instant, the base-model gate before the first solve, and the critical
    time/minimum the curve reports. File-light (it loads the shipped ACADS sample
    and builds synthetic frame ledgers) and solves no seepage.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'fs_vs_time_check.py'
    if not path.exists():
        return None, f"missing {path}"
    spec = importlib.util.spec_from_file_location('fs_vs_time_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_sweep_window_test(test):
    """The search window a sweep searches inside (``xslope.sensitivity``).

    A model may declare a search window on its circles sheet — entry and exit
    ranges, a centre box, a tangent-depth band, a minimum slip depth — and it is
    how an engineer says which mechanism is under study. A parametric sweep
    re-searches at every value, so ignoring the window lets one point settle in a
    different surface FAMILY from its neighbour, and the step in FS between them
    is then read as sensitivity to the parameter rather than as a change in what
    was measured.

    The check itself lives in test/sweep_window_check.py: an undeclared window
    leaves the search unconstrained, a declared one reaches every point including
    the base case, a half-declared range is not a window, an explicit
    ``search_opts`` beats the file, ``use_file_window=False`` turns it off, the
    four studies built on the sweep all inherit it, the non-circular branch takes
    only the limit it understands — and the PARITY that makes it one behaviour:
    a sweep, both reliability engines and Studio's Run LEM path hand the same
    model's search the same constraints. File-light (it loads the shipped ACADS
    sample and replaces the search functions with recorders, because what is
    under test is which constraints reach them).

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'sweep_window_check.py'
    if not path.exists():
        return None, f"missing {path}"
    spec = importlib.util.spec_from_file_location('sweep_window_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_water_hoist_test(test):
    """Where the automatic water load is derived, and how often (``xslope.search``).

    With automatic water loads on, the weight of the standing water is measured
    at solve time from the model's own water definition. The slicer does that
    defensively at the top of every ``generate_slices`` call, which under a search
    means once per TRIAL SURFACE — several thousand identical derivations for one
    factor of safety, because a search varies the circle and touches neither the
    water line nor the ground surface. It is hoisted to the entry of the search
    instead, and this is the performance change's correctness claim.

    The check itself lives in test/water_hoist_check.py: once per search rather
    than once per trial (asserted against a run that really did slice hundreds of
    surfaces), the identical slice table and factor of safety either side of the
    hoist, no reuse ACROSS runs — an edited water definition must be re-derived,
    or Studio's second run answers with the first run's reservoir — and manual
    mode deriving nothing at either place. File-light: two shipped samples and
    short searches.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'water_hoist_check.py'
    if not path.exists():
        return None, f"missing {path}"
    spec = importlib.util.spec_from_file_location('water_hoist_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_stability_time_test(test):
    """Which instant of a transient march a stability run reads, and whether that
    choice survives the file.

    A factor of safety against a time-dependent seepage field belongs to ONE instant,
    and the instant is chosen in several places that have to agree: the run entry
    point's precedence (explicit ``time=`` over the model's ``stability_time`` over
    the engine default), the tseep sheet's round trip, the saved-frame schedule, the
    two Run dialogs' selector, and the run gate that must REFUSE stage times with no
    frames rather than silently reusing the previous run's pore pressures.

    The check itself lives in test/stability_time_check.py (file-less; it builds
    synthetic frame ledgers, solves no seepage, and skips its Studio leg cleanly when
    PySide6 is absent).

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'stability_time_check.py'
    if not path.exists():
        return None, f"missing {path}"
    # The check module builds the QApplication itself only when run as a script, so
    # the harness supplies the offscreen one here — the same idiom every other
    # Studio-touching row uses. Without it the dialog leg aborts the worker process
    # rather than failing a test.
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips its Studio leg
    spec = importlib.util.spec_from_file_location('stability_time_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_steady_seep_save_test(test):
    """Two things a desktop session does between one run and the next: attaching a
    seepage solution it has just computed, and writing the file back out.

    A steady solve that fills only a results tab leaves the model itself carrying no
    pore-pressure field, so the run gate refuses -- correctly for what it can see --
    a run whose seepage solution is on screen. A field belongs to the model the
    moment it is solved, not the moment it is reloaded from a sidecar. The other half
    is the .xlsx writer, which re-zipped the workbook by running ``zip``: there is no
    such executable on Windows, so every Save there raised FileNotFoundError while
    macOS and Linux, where the tool ships with the OS, worked. The check saves with
    every subprocess entry point closed and PATH emptied, and again through a
    frozen-style resource layout, since the bundle's smoke test proved the packaged
    template READABLE, which is not the same as proving a save from it.

    The check itself lives in test/steady_seep_save_check.py. It solves one small
    steady seepage problem for its Studio leg and skips that leg cleanly without
    PySide6.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'steady_seep_save_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication, QMessageBox
        QApplication.instance() or QApplication([])
        # A refusal the Studio leg provokes must not open a modal box in a worker.
        QMessageBox.warning = staticmethod(lambda *a, **k: QMessageBox.Ok)
        QMessageBox.information = staticmethod(lambda *a, **k: QMessageBox.Ok)
        QMessageBox.critical = staticmethod(lambda *a, **k: QMessageBox.Ok)
    except Exception:
        pass                       # no PySide6: the module skips its Studio leg
    spec = importlib.util.spec_from_file_location('steady_seep_save_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_noncircular_generator_test(test):
    """The non-circular starting-surface generator: what it ranks on, when it picks
    versus asks, and whether what it builds is usable.

    A generated surface is a SEED, so the ways it can be wrong are quiet ones. Ranking
    zones on cohesion alone would pick a clean dense sand over the weak seam that
    actually controls the mechanism, and the search would then converge on a perfectly
    valid answer to the wrong question. An end point written with a blank Y, or a
    point with a blank Movement, reaches the slicer as a TypeError or silently freezes
    the search on the surface it was handed.

    The check itself lives in test/noncircular_generator_check.py: the mobilisable-
    strength metric across every strength model, the separation threshold asserted
    from both sides against a measured ratio, the geometry invariants on a slope and
    on its mirror image, outcropping and pinched-out zones, determinism, the reason
    every refusal carries, that a corpus weak-seam surface slices inside its seam, and
    the Studio button and zone picker.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'noncircular_generator_check.py'
    if not path.exists():
        return None, f"missing {path}"
    # The check module builds the QApplication itself only when run as a script, so
    # the harness supplies the offscreen one here, before exec_module — the same idiom
    # every other Studio-touching row uses. Without it the Studio legs abort the
    # worker process rather than failing a test.
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips its Studio legs
    spec = importlib.util.spec_from_file_location('noncircular_generator_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_updater_test(test):
    """Studio's in-app updater: what stands between a version on a web server and
    an installer running on the user's machine.

    Every decision here is invisible when it is right and expensive when it is
    wrong. A version comparison done on strings offers 0.1.9 as an upgrade over
    0.1.10; a ``minimum_version`` nobody asserts on silently stops being a gate; and
    an updater that runs a binary whose checksum did not match is the one bug in
    this feature that actually matters. The check asserts on the COMMAND each
    platform branch would spawn rather than running anything.

    The check itself lives in test/updater_check.py: nine legs, no network at all
    (the manifest and the artifact are files served over ``file://``, the same
    urllib path the real release URL takes), and the two Qt legs skip cleanly when
    PySide6 is absent.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'updater_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips its Qt legs
    spec = importlib.util.spec_from_file_location('updater_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_assistant_models_test(test):
    """Where the assistant's model list comes from: the provider's own catalogue,
    the cached copy, the list this build shipped with, and the curated
    recommendations overlaid on whichever of the three is showing.

    A model list baked into a release goes stale within weeks, so the settings
    dialog builds it at open time — which means the failure modes are silent
    ones: a cache tier that stopped serving leaves an offline user staring at
    year-old model ids, a filter that got greedy hides the model they wanted, and
    a malformed manifest that is NOT ignored takes the dialog down with it. The
    check mutates each of those in turn (kill the live list, kill the cache,
    corrupt the manifest eight ways) and asserts the answer moves exactly one
    step, never to an empty list.

    The check itself lives in test/assistant_models_check.py: eleven legs, no
    network at all (every provider response and the manifest are files served
    over ``file://``, the same urllib path the real endpoints take), and the Qt
    legs build the real runner and the real dialog — the latter with its
    background refresh off, so it cannot reach a provider even on a machine with
    a key in its keychain.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import importlib.util

    path = Path(__file__).parent / 'test' / 'assistant_models_check.py'
    if not path.exists():
        return None, f"missing {path}"
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    try:
        from PySide6.QtWidgets import QApplication
        QApplication.instance() or QApplication([])
    except Exception:
        pass                       # no PySide6: the module skips its Qt leg
    spec = importlib.util.spec_from_file_location('assistant_models_check', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    failures = mod.run()
    if failures:
        return None, "; ".join(failures)
    return 0.0, None


def run_drawdown_tauff_test(test):
    """The Stage-2 undrained strength pipeline, checked against the worked example in
    Duncan, Wright & Brandon, *Soil Strength and Slope Stability*, 2nd ed., Table 9.2.

    An infinite slope, so the whole tau_ff chain -- eqs (4) K1, (5) interpolation,
    (6) Kf, (7)/(8) negative-stress fallback -- is exercised with hand-computable
    inputs and no slice machinery at all.

    gamma = 125 pcf, c' = 0, phi' = 40 deg, d = 2000 psf, psi = 20 deg, beta = 18.4 deg,
    fully submerged under 100 ft of water before a complete drawdown.

    The book carries K1 = 2.0 where the exact value is 2.0304, so its tau_ff (1585 psf
    at z = 5 ft) is a rounded figure. We assert against the exact recomputation and
    record the book's rounded values alongside; substituting K1 = 2.0 and Kf = 4.6
    reproduces 1585 to the digit.

    Also asserts the invariant that motivated the Stage-1 FS >= 1 guard: K1 rises
    monotonically as FS1 falls and equals Kf exactly at FS1 = 1, so K1 <= Kf there and
    eq (5) interpolates rather than extrapolating.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test.
    """
    import numpy as np

    gamma, gamma_w = 125.0, 62.4
    c1, phi1, d_val, psi_val = 0.0, 40.0, 2000.0, 20.0
    # The 3H:1V slope is arctan(1/3) = 18.4349 deg, but the book carries 18.4 deg through
    # eqs (9.19)-(9.42). Use its rounded angle so the printed values are reproducible.
    beta = 18.4
    pr = np.radians(phi1)
    cos_phi, sin_phi = np.cos(pr), np.sin(pr)

    def tau_ff_of(sigma_fc, tau_fc):
        """Mirror of advanced.rapid_drawdown's Stage-2 strength block."""
        t_k1 = d_val + sigma_fc * np.tan(np.radians(psi_val))          # eq (9.26)
        t_kf = c1 + sigma_fc * np.tan(pr)                              # eq (9.28)
        kf_first = sigma_fc - c1 * cos_phi
        if abs(cos_phi) < 1e-12 or abs(kf_first) < 1e-12:
            return t_k1, t_kf, None, None, min(t_k1, t_kf)
        s3_k1 = sigma_fc + tau_fc * (sin_phi - 1) / cos_phi            # eq (9.14)
        s3_kf = kf_first * (1 - sin_phi) / (cos_phi ** 2)              # eq (9.15)
        if s3_k1 <= 0 or s3_kf <= 0:
            return t_k1, t_kf, None, None, min(t_k1, t_kf)
        K1 = (sigma_fc + tau_fc * (sin_phi + 1) / cos_phi) / s3_k1     # eq (9.10)
        Kf = ((sigma_fc + c1 * cos_phi) * (1 + sin_phi)) / (kf_first * (1 - sin_phi))  # eq (9.13)
        t_ff = ((Kf - K1) * t_k1 + (K1 - 1) * t_kf) / (Kf - 1)         # eq (9.11)
        return t_k1, t_kf, K1, Kf, t_ff

    problems = []
    br = np.radians(beta)

    # Book Table 9.2, per depth: sigma'_fc, tau_fc, tau_ff(Kc=1), tau_ff(Kc=Kf), drained
    book = {5: (282.0, 94.0, 2103.0, 237.0, 472.0),
            30: (1691.0, 562.0, 2615.0, 1419.0, 2833.0)}

    for z, (b_sfc, b_tfc, b_k1, b_kf, b_drained) in book.items():
        # Stage 1: submerged infinite slope. With no flow, sigma'_fc = gamma' z cos^2(beta)
        # (eq 9.23) and tau_fc = (gamma - gamma_w) z sin(beta) cos(beta) (eq 9.24).
        sigma_fc = (gamma - gamma_w) * z * np.cos(br) ** 2
        tau_fc = (gamma - gamma_w) * z * np.sin(br) * np.cos(br)
        if abs(sigma_fc - b_sfc) > 1.0:
            problems.append(f"z={z}: sigma'_fc {sigma_fc:.1f} vs book {b_sfc}")
        if abs(tau_fc - b_tfc) > 1.0:
            problems.append(f"z={z}: tau_fc {tau_fc:.1f} vs book {b_tfc}")

        t_k1, t_kf, K1, Kf, t_ff = tau_ff_of(sigma_fc, tau_fc)
        if abs(t_k1 - b_k1) > 1.0:
            problems.append(f"z={z}: tau_ff(Kc=1) {t_k1:.1f} vs book {b_k1}")
        if abs(t_kf - b_kf) > 1.0:
            problems.append(f"z={z}: tau_ff(Kc=Kf) {t_kf:.1f} vs book {b_kf}")
        if K1 is None or not (1.0 <= K1 <= Kf):
            problems.append(f"z={z}: K1={K1} outside [1, Kf={Kf}] — eq (5) would extrapolate")

        # Book's rounded K1=2.0, Kf=4.6 must reproduce its printed tau_ff exactly.
        rounded = ((4.6 - 2.0) * t_k1 + (2.0 - 1) * t_kf) / (4.6 - 1)
        book_tff = {5: 1585.0, 30: 2283.0}[z]
        if abs(rounded - book_tff) > 1.0:
            problems.append(f"z={z}: tau_ff at book's rounded K1/Kf {rounded:.1f} "
                            f"vs book {book_tff}")

        # Stage 3: total drawdown, zero pore pressure -> sigma' = gamma z cos^2(beta).
        drained = c1 + gamma * z * np.cos(br) ** 2 * np.tan(pr)        # eqs (9.37), (9.39)
        if abs(drained - b_drained) > 1.0:
            problems.append(f"z={z}: drained strength {drained:.1f} vs book {b_drained}")

    # K1 == Kf exactly at Stage-1 FS == 1, and K1 > Kf below it. This is why
    # advanced.rapid_drawdown refuses a Stage-1 FS < 1 rather than extrapolating.
    sigma_fc = 282.0
    tau_fail = c1 + sigma_fc * np.tan(pr)          # tau_fc when FS1 == 1
    prev_K1 = None
    for FS1 in (3.0, 2.0, 1.5, 1.0, 0.95, 0.9):
        _, _, K1, Kf, _ = tau_ff_of(sigma_fc, tau_fail / FS1)
        if K1 is None:
            continue
        if prev_K1 is not None and not K1 > prev_K1:
            problems.append(f"K1 not monotonically increasing as FS1 falls (FS1={FS1})")
        prev_K1 = K1
        if abs(FS1 - 1.0) < 1e-12 and abs(K1 - Kf) > 1e-6:
            problems.append(f"K1={K1:.6f} != Kf={Kf:.6f} at FS1 = 1")
        if FS1 < 1.0 and not K1 > Kf:
            problems.append(f"FS1={FS1} < 1 but K1={K1:.4f} <= Kf={Kf:.4f}")
        if FS1 > 1.0 and not K1 < Kf:
            problems.append(f"FS1={FS1} > 1 but K1={K1:.4f} >= Kf={Kf:.4f}")

    if problems:
        return None, "Duncan/Wright/Brandon Table 9.2: " + "; ".join(problems)
    return 0.0, None


def run_drawdown_guard_test(test):
    """`rapid_drawdown` must REFUSE a slope whose Stage-1 (full pool) FS < 1 rather
    than silently extrapolating eq (5) past the Kc=Kf envelope.

    Below FS1 = 1 the mobilized shear tau_fc exceeds the failure envelope, so K1 > Kf
    on every low-K slice and tau_ff falls below the physical floor -- eventually
    negative, where `max(0, tau_ff)` laundered it into a zero-strength slice. A search
    would then rank that trial surface as catastrophic on a fictitious FS ~ 0. The
    negative-stress fallback does not catch it: sigma'_3c stays positive.

    Returns (0.0, None) on success, else (None, message).
    """
    import copy
    from xslope.fileio import load_slope_data
    from xslope.slice import generate_slices
    from xslope.advanced import rapid_drawdown

    path = test['file']
    if not Path(path).exists():
        return None, f"missing fixture {path}"
    base = load_slope_data(path)
    problems = []

    def run(scale):
        data = copy.deepcopy(base)
        for m in data['materials']:
            for key in ('c', 'phi', 'd', 'psi'):
                m[key] = m.get(key, 0) * scale
        ok, payload = generate_slices(data, circle=data['circles'][0],
                                      num_slices=40, debug=False)
        if not ok:
            return None
        return rapid_drawdown(payload[0], 'bishop', debug_level=0)

    # Unweakened: Stage 1 is stable, analysis proceeds.
    got = run(1.0)
    if got is None or not got[0]:
        problems.append(f"unweakened model should succeed, got {got}")
    elif got[1]['stage1_FS'] < 1.0:
        problems.append("fixture no longer has Stage-1 FS >= 1; pick another")

    # Weakened until Stage 1 fails: must be refused, not answered.
    got = run(0.5)
    if got is None:
        problems.append("weakened model failed to generate slices")
    elif got[0]:
        problems.append(f"weakened model (Stage-1 FS < 1) returned FS={got[1]['FS']:.4f} "
                        "instead of being refused")
    elif 'stable before drawdown' not in str(got[1]):
        problems.append(f"refused, but with an unexpected message: {got[1]}")

    if problems:
        return None, "; ".join(problems)
    return 0.0, None


def run_axial_mirror_test(test):
    """Axial (nail) reinforcement must be facing-invariant: a nailed wall and its
    exact left/right mirror must give the same FS in every method, on both a
    circular and a planar surface. Locks the right-facing axial sign conventions
    (vertical force component, facing detection against a vertical wall face,
    spencer/mprice full-vector negation) fixed while building Slide VP47/VP48.

    corps/lowe carry a small inherent directional asymmetry (their interslice
    inclinations are read from the geometry left-to-right), so they get a looser
    tolerance. Pass/fail test: returns (0.0, None) on success.
    """
    from xslope.fileio import load_slope_data
    from xslope.slice import generate_slices
    from xslope import solve

    lp, rp = test['file'], test['file2']
    for fp in (lp, rp):
        if not os.path.exists(fp):
            return None, f"input missing: {fp}"
    left = load_slope_data(lp)
    right = load_slope_data(rp)

    problems = []
    for kind in ('circle', 'plane'):
        if kind == 'circle':
            ok_l, pl = generate_slices(left, circle=left['circles'][0], num_slices=50)
            ok_r, pr = generate_slices(right, circle=right['circles'][0], num_slices=50)
            methods = ('oms', 'bishop', 'janbu', 'spencer', 'mprice', 'corps', 'lowe')
        else:
            ok_l, pl = generate_slices(left, non_circ=left['non_circ'], num_slices=50)
            ok_r, pr = generate_slices(right, non_circ=right['non_circ'], num_slices=50)
            methods = ('janbu', 'spencer', 'mprice', 'corps', 'lowe')
        if not (ok_l and ok_r):
            problems.append(f"{kind}: slice generation failed")
            continue
        df_l, df_r = pl[0], pr[0]
        for m in methods:
            fn = getattr(solve, m)
            ok1, r1 = fn(df_l.copy())
            ok2, r2 = fn(df_r.copy())
            if ok1 != ok2:
                # both-fail is acceptable (method admissibility, mirror-symmetric);
                # one-sided failure is a facing bug
                problems.append(f"{kind}/{m}: one facing failed "
                                f"(L={'ok' if ok1 else 'fail'}, R={'ok' if ok2 else 'fail'})")
                continue
            if not ok1:
                continue
            tol = 0.01 if m in ('corps', 'lowe') else 1e-3
            rel = abs(r1['FS'] - r2['FS']) / max(abs(r1['FS']), 1e-9)
            if rel > tol:
                problems.append(f"{kind}/{m}: L={r1['FS']:.4f} R={r2['FS']:.4f} "
                                f"rel diff {rel:.1e} > {tol:.0e}")
    if problems:
        return None, "; ".join(problems)
    return 0.0, None


def run_mp_spencer_test(test):
    """Morgenstern-Price with f(x) == 1 must reproduce Spencer exactly, on BOTH
    slope facings. Spencer is the f == 1 special case of M-P, so any divergence
    means the M-P march (or its right-facing mirror in `_mp_extract`) is wrong.

    This is the S3b gate. It is asserted here because the right-facing mirror is
    otherwise only exercised implicitly, and a stale docstring once claimed the
    facing was unsupported when it had in fact been implemented.

    Returns (0.0, None) on success, else (None, message) — a pass/fail test, like
    the round-trip checks.
    """
    from xslope.fileio import load_slope_data
    from xslope.slice import generate_slices
    from xslope import solve

    path = test['file']
    if not os.path.exists(path):
        return None, f"input missing: {path}"

    data = load_slope_data(path)
    if not data.get('circles'):
        return None, f"{path} has no circle to analyze"

    ok, payload = generate_slices(data, circle=data['circles'][0],
                                  num_slices=test.get('num_slices', 40), debug=False)
    if not ok:
        return None, f"slice generation failed: {payload}"
    slice_df = payload[0]

    facing = 'right' if slice_df['y_ct'].values[0] > slice_df['y_ct'].values[-1] else 'left'
    expected = test.get('facing')
    if expected and expected != facing:
        return None, f"{path} is {facing}-facing, expected {expected}-facing"

    ok_s, res_s = solve.spencer(slice_df)
    if not ok_s:
        return None, f"spencer failed on {facing}-facing: {res_s}"
    ok_m, res_m = solve.mprice(slice_df, f_type='constant')
    if not ok_m:
        return None, f"mprice(f=constant) failed on {facing}-facing: {res_m}"

    diff = abs(res_s['FS'] - res_m['FS'])
    tol = test.get('tolerance', 1e-8)
    if diff > tol:
        return None, (f"{facing}-facing: mprice(f=1) FS={res_m['FS']:.8f} != "
                      f"spencer FS={res_s['FS']:.8f} (diff {diff:.2e} > {tol:.0e})")
    return 0.0, None


def _default_dxf_mapping(layers):
    """The per-layer mapping the import wizard seeds by default: target from
    ``suggest_dxf_target`` (xslope's own export layer names + geometry kind), and
    the material name from the layer name (PROFILE_ stripped). Mirrors
    studio.dialogs.DxfImportDialog so the suite exercises the same defaults."""
    from xslope.cad import suggest_dxf_target
    out = {}
    for name, geom in layers.items():
        target = suggest_dxf_target(name, geom)
        mat = (name[8:] if name.upper().startswith('PROFILE_') else name)
        out[name] = {'target': target,
                     'material': mat if target in ('material_zone', 'profile') else None}
    return out


def run_dxf_roundtrip_test(test):
    """Round-trip a model through the structured DXF export/import path: load ->
    export_dxf -> read_dxf_layers -> default wizard mapping -> build_from_dxf_mapping
    must recover the geometry (feature counts + circles). Returns (0.0, None) or
    (None, message)."""
    import tempfile
    # Ensure a (headless) QApplication exists — ProjectDocument is a QObject.
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication([])
    from xslope.fileio import load_slope_data
    from xslope.cad import export_dxf, read_dxf_layers
    from studio.document import ProjectDocument

    sd = load_slope_data(test['file'])
    tmp = tempfile.NamedTemporaryFile(suffix='.dxf', delete=False).name
    try:
        export_dxf(sd, tmp)
        layers, _ = read_dxf_layers(tmp)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)

    mapping = _default_dxf_mapping(layers)
    doc = ProjectDocument()
    doc.build_from_dxf_mapping(layers, mapping)
    out = doc.slope_data

    problems = []
    kind = test.get('kind')
    if kind == 'profile':
        n0, n1 = len(sd.get('profile_lines') or []), len(out.get('profile_lines') or [])
        if n1 != n0:
            problems.append(f"profile_lines {n1} vs {n0}")
        if out.get('ground_surface') is None:
            problems.append("ground_surface missing")
    elif kind == 'polygon':
        n0, n1 = len(sd.get('polygons') or []), len(out.get('polygons') or [])
        if n1 != n0:
            problems.append(f"polygons {n1} vs {n0}")
        if not out.get('materials'):
            problems.append("no materials")
        # The in-memory path above never touches the .xlsx, so it cannot see
        # cad.import_dxf writing polygon cells at the wrong ROWS — which is exactly
        # what happened when template v21 inserted the Type and Size rows: the Mat ID
        # landed in the Type cell and the written file no longer loaded at all. Write
        # one and parse the polygon sheet straight back.
        import shutil
        import pandas as _pd
        from xslope.cad import import_dxf as _import_dxf
        from xslope.fileio import (_parse_polygon_sheet as _parse_poly,
                                   _read_template_info as _tinfo)
        _tmpd = tempfile.mkdtemp()
        _dxf2 = os.path.join(_tmpd, 'g.dxf')
        _xl = os.path.join(_tmpd, 'g.xlsx')
        try:
            export_dxf(sd, _dxf2)
            _import_dxf(_dxf2, ROUNDTRIP_TEMPLATE, _xl)
            _ver = _tinfo(_xl)[0]
            _fake_mats = [{'name': f'm{i}'} for i in range(64)]
            _p, _z, _r = _parse_poly(_pd.ExcelFile(_xl), _fake_mats,
                                     template_version=_ver)
            if len(_p) != n0:
                problems.append(f"import_dxf wrote {len(_p)} polygon(s), expected {n0}")
        except Exception as _e:
            problems.append(f"import_dxf -> xlsx: {type(_e).__name__}: {_e}")
        finally:
            shutil.rmtree(_tmpd, ignore_errors=True)
    elif kind == 'reinforce':
        n0 = len(sd.get('reinforcement_lines') or sd.get('reinforce_lines') or [])
        n1 = len(out.get('reinforcement_lines') or [])
        if n1 != n0:                          # rejoin must not split into per-segment lines
            problems.append(f"reinforcement {n1} vs {n0}")
    # Circles are recovered (radius fit from the arc) wherever present.
    if sd.get('circles'):
        n0, n1 = len(sd['circles']), len(out.get('circles') or [])
        if n1 != n0:
            problems.append(f"circles {n1} vs {n0}")

    if problems:
        return None, "DXF round-trip: " + "; ".join(problems[:5])
    return 0.0, None


#: The model the DXF water classifier is exercised on: a dam with a reservoir, whose
#: dloads sheet carries the ponded water and whose piezo sheet carries the line that
#: describes the same pool. Both reach the DXF, on two different layers, which is the
#: whole reason the classifier exists.
DXF_WATER_FILE = _repo('docs/inputs/slope/xslope_dam.xlsx')


def run_dxf_water_test(test):
    """The DXF import's water-load classification, and the double count it prevents.

    A DXF is the only import that round-trips xslope's OWN data, and it reads a
    ``DLOADS`` layer and a ``PIEZO`` layer out of the same file. So the mode cannot be
    a constant: it has to follow what the file carries.

      - **a load block tracing the pool** is ponded water somebody drew, so the model
        imports MANUAL and that block carries the reservoir;
      - **a piezo line and no such block** is a water definition and nothing else, so
        the model imports AUTO and the engine measures the reservoir itself;
      - **a surcharge elsewhere on the ground** is user data: it is kept, and it does
        not make the model manual.

    The classification is geometric because a DXF carries no pressures --
    ``export_dxf`` writes a load block as a bare polyline, so every block returns at
    ``Normal = 0`` and the pressure-based match cannot speak here.

    And the consequence is measured rather than asserted: the imported block is
    priced the way a user would price it, and the same model is solved under both
    modes. Under the mode the classifier chose the slice water force reproduces the
    source file's; under the other one it is doubled.
    """
    import tempfile
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    from PySide6.QtWidgets import QApplication
    QApplication.instance() or QApplication([])
    from xslope.fileio import load_slope_data
    from xslope.cad import export_dxf, read_dxf_layers, water_like_dloads
    from xslope.water import derive_water_loads, _p_at
    from studio.document import ProjectDocument

    problems = []
    # The fixture is the source model in the pre-v22 idiom: a reservoir typed on
    # the dloads sheet beside the piezometric line that defines it. That pairing
    # is what the classifier exists to recognize, and a v22 file no longer states
    # its water twice — so the fixture is built rather than read (the same move
    # as _preflight_apply's manual_water).
    src = _legacy_manual_water(load_slope_data(DXF_WATER_FILE))
    if not src.get('dloads') or not src.get('piezo_line'):
        return None, (f"{DXF_WATER_FILE} no longer carries both a water load and a "
                      f"piezo line, so the classifier is not exercised")
    circle = dict(src['circles'][0])

    def _through_dxf(model, mutate_mapping=None):
        tmp = tempfile.NamedTemporaryFile(suffix='.dxf', delete=False).name
        try:
            export_dxf(model, tmp)
            layers, _ = read_dxf_layers(tmp)
        finally:
            if os.path.exists(tmp):
                os.unlink(tmp)
        mapping = _default_dxf_mapping(layers)
        if mutate_mapping:
            mutate_mapping(mapping)
        doc = ProjectDocument()
        notes = doc.build_from_dxf_mapping(layers, mapping)
        return doc.slope_data, notes

    # (a) the reservoir is drawn on the load layer -> manual.
    got, notes = _through_dxf(src)
    if got.get('water_loads') != 'manual':
        problems.append(
            f"a DXF carrying the reservoir on its DLOADS layer imported as "
            f"{got.get('water_loads')!r}; it must import manual, or pricing that "
            f"block adds a second reservoir to the derived one")
    if len(got.get('dloads') or []) != len(src['dloads']):
        problems.append(f"the water block did not survive the DXF: "
                        f"{len(got.get('dloads') or [])} vs {len(src['dloads'])}")
    if not any('MANUAL' in n for n in notes):
        problems.append("the import did not report that it chose manual, or why")

    # (b) the same file with the load layer ignored: only the piezo line states the
    # water, so the model imports auto and the engine derives the reservoir.
    def _drop_dloads(mapping):
        for lyr, m in mapping.items():
            if lyr.strip().upper() == 'DLOADS':
                m['target'] = 'ignore'
    auto_sd, auto_notes = _through_dxf(src, _drop_dloads)
    if auto_sd.get('water_loads') != 'auto':
        problems.append(f"a DXF with a piezo line and no load block imported as "
                        f"{auto_sd.get('water_loads')!r}, expected 'auto'")
    if not derive_water_loads(auto_sd)['blocks']:
        problems.append("the automatic import derives no water at all — the piezo "
                        "line did not survive, so the reservoir is simply gone")
    if not any('automatic' in n for n in auto_notes):
        problems.append("the automatic import did not report that the engine now "
                        "supplies the reservoir")

    # (c) a surcharge is not water. Put a block on the dry crest, well clear of the
    # pool, and the classifier must leave the model automatic and keep the block.
    ground = list(src['ground_surface'].coords)
    reach = derive_water_loads(src)['reach']
    dry = [(x, y) for x, y in ground if x > reach[1] + 0.3 * (ground[-1][0] - reach[1])]
    if len(dry) < 2:
        problems.append("the fixture has no dry stretch to put a surcharge on")
    else:
        sur = [{'X': x, 'Y': y, 'Normal': 500.0} for x, y in dry[:4]]
        sur_sd, _ = _through_dxf(dict(src, dloads=[sur]))
        if sur_sd.get('water_loads') != 'auto':
            problems.append(
                f"a DXF whose only load block is a crest surcharge imported as "
                f"{sur_sd.get('water_loads')!r}; a surcharge is user data and must "
                f"not be mistaken for the reservoir")
        if len(sur_sd.get('dloads') or []) != 1:
            problems.append("the surcharge was dropped by the DXF import")
        both_sd, _ = _through_dxf(dict(src, dloads=list(src['dloads']) + [sur]))
        if both_sd.get('water_loads') != 'manual':
            problems.append("a DXF carrying BOTH a reservoir block and a surcharge "
                            "imported auto; the reservoir block still rules")
        hits = water_like_dloads(both_sd)['indices']
        if hits != [0]:
            problems.append(f"the classifier picked block(s) {hits} out of a "
                            f"reservoir + surcharge pair, expected [0]")

    # (d) THE DOUBLE COUNT, priced and measured. The DXF brings the block back with
    # no pressures (the file has none), so price it from the pool the way a user
    # would after reading the note. Under the manual mode the classifier chose, the
    # slice water force reproduces the source model's. Under auto -- the mode a
    # permissive classifier would have picked -- the same model carries it twice.
    #
    # The block is the one that came back through the DXF; the rest of the model is
    # the source's, because a DXF carries no material strengths and a slice frame
    # cannot be built on placeholder soils. What is being measured is the water.
    derived = derive_water_loads(src)['blocks']
    priced = [[dict(p, Normal=_p_at(derived[0], p['X'])) for p in blk]
              for blk in (got.get('dloads') or [])]
    if not priced or not derived:
        problems.append("could not price the imported block against the derivation")
    else:
        base = _water_force_columns(src, circle, n=30)
        man = _water_force_columns(dict(src, water_loads='manual', dloads=priced),
                                   circle, n=30)
        dup = _water_force_columns(dict(src, water_loads='auto', dloads=priced),
                                   circle, n=30)
        if any(isinstance(c, str) for c in (base, man, dup)):
            problems.append("the double-count measurement could not build slices")
        else:
            s_base, s_man, s_dup = (sum(c[0]) for c in (base, man, dup))
            if not s_base:
                problems.append("the source model carries no water force on this "
                                "circle, so the measurement proves nothing")
            elif abs(s_man - s_base) > 1e-6 * abs(s_base):
                problems.append(
                    f"the priced DXF import does not reproduce the source model's "
                    f"water force: {s_man:.6g} vs {s_base:.6g}")
            elif abs(s_dup - 2.0 * s_base) > 1e-6 * abs(s_base):
                problems.append(
                    f"forcing the same model to auto gave {s_dup:.6g} against the "
                    f"source's {s_base:.6g}: the double count this classification "
                    f"prevents is not what the check assumes it is")

    if problems:
        return None, "DXF water loads: " + "; ".join(problems[:5])
    return 0.0, None


# Every XML tag path (and attribute) that GeoStudio itself writes, among those
# export_gsz emits. Harvested from Seequent-authored .gsz files; the FILES are
# their copyrighted Materials and are not in this repo, but the tag vocabulary is
# a fact about the format, not a copy of their content.
#
# This guards the failure that a round-trip test CANNOT catch: a reader and writer
# that share the same wrong idea of the schema agree perfectly with each other and
# produce a file GeoStudio rejects. (That is exactly what happened — the first
# exporter put PiezometricSurfaces under <Geometry> instead of the StabilityItem,
# and omitted <Lines> entirely, and the round-trip test passed regardless.)
GSZ_SCHEMA_PATHS = {
    "/GSIData",
    "/GSIData/Analyses",
    "/GSIData/Analyses/Analysis",
    "/GSIData/Analyses/Analysis/GeometryId",
    "/GSIData/Analyses/Analysis/ID",
    "/GSIData/Analyses/Analysis/Kind",
    "/GSIData/Analyses/Analysis/Method",
    "/GSIData/Analyses/Analysis/Name",
    "/GSIData/Analyses/Analysis/SampleSettings",
    "/GSIData/Analyses/Analysis/SampleSettings/NumMonteCarloTrials",
    "/GSIData/Analyses/Analysis/SampleSettings/ProbabilisticModifiers",
    "/GSIData/Analyses/Analysis/SampleSettings/ProbabilisticModifiers/Modifier",
    "/GSIData/Analyses/Analysis/SampleSettings/ProbabilisticModifiers/Modifier/Entry",
    "/GSIData/Analyses/Analysis/SampleSettings/ProbabilisticModifiers/Modifier/Entry@Function",
    "/GSIData/Analyses/Analysis/SampleSettings/ProbabilisticModifiers/Modifier/Key",
    "/GSIData/Analyses/Analysis/SampleSettings/ProbabilisticModifiers@Len",
    "/GSIData/Analyses/Analysis/SampleSettings/Type",
    "/GSIData/Analyses@Len",
    "/GSIData/Contexts",
    "/GSIData/Contexts/Context",
    "/GSIData/Contexts/Context/AnalysisID",
    "/GSIData/Contexts/Context/GeometryUsesMaterials",
    "/GSIData/Contexts/Context/GeometryUsesMaterials/GeometryUsesMaterial",
    "/GSIData/Contexts/Context/GeometryUsesMaterials/GeometryUsesMaterial@Entry",
    "/GSIData/Contexts/Context/GeometryUsesMaterials/GeometryUsesMaterial@ID",
    "/GSIData/Contexts/Context/GeometryUsesMaterials@Len",
    "/GSIData/Contexts/Context/IsDefined",
    "/GSIData/Contexts@Len",
    "/GSIData/Coordinates",
    "/GSIData/Coordinates/EngCoords",
    "/GSIData/Coordinates/EngCoords@HorzScale",
    "/GSIData/Coordinates/EngCoords@LockScales",
    "/GSIData/Coordinates/EngCoords@MaxSnapDist",
    "/GSIData/Coordinates/EngCoords@UnitSystem",
    "/GSIData/Coordinates/EngCoords@VertScale",
    "/GSIData/Coordinates/EngCoords@XPageLeft",
    "/GSIData/Coordinates/EngCoords@XPageOrg",
    "/GSIData/Coordinates/EngCoords@XPageRight",
    "/GSIData/Coordinates/EngCoords@YPageBottom",
    "/GSIData/Coordinates/EngCoords@YPageOrg",
    "/GSIData/Coordinates/EngCoords@YPageTop",
    "/GSIData/Coordinates/PageCoords",
    "/GSIData/Coordinates/PageCoords@PageHeight",
    "/GSIData/Coordinates/PageCoords@PageWidth",
    "/GSIData/Coordinates/PageCoords@PageXOrg",
    "/GSIData/Coordinates/PageCoords@PageYOrg",
    "/GSIData/Coordinates/PageCoords@Units",
    "/GSIData/FileInfo",
    "/GSIData/FileInfo@Comments",
    "/GSIData/FileInfo@Title",
    # SLOPE/W's spatial pore-pressure input: a set of discrete (X, Y, value) points the
    # solver interpolates between, selected on the analysis by the 3DFunction option
    # below. This is how an xslope finite-element seepage field crosses -- as data, with
    # no fabricated SEEP/W parent. GeoStudio's own files carry a USED 3DFunction with
    # <Points> alone (its <Elements>/<Nodes> triangulation cache is not always written),
    # which is why only these paths are emitted.
    "/GSIData/Functions",
    "/GSIData/Functions/Func3Ds",
    "/GSIData/Functions/Func3Ds@Len",
    "/GSIData/Functions/Func3Ds/Fn3D",
    "/GSIData/Functions/Func3Ds/Fn3D/ID",
    "/GSIData/Functions/Func3Ds/Fn3D/Model",
    "/GSIData/Functions/Func3Ds/Fn3D/Name",
    "/GSIData/Functions/Func3Ds/Fn3D/OutputType",
    "/GSIData/Functions/Func3Ds/Fn3D/Points",
    "/GSIData/Functions/Func3Ds/Fn3D/Points@Len",
    "/GSIData/Functions/Func3Ds/Fn3D/Points/Points_",
    "/GSIData/Functions/Func3Ds/Fn3D/Points/Points_@X",
    "/GSIData/Functions/Func3Ds/Fn3D/Points/Points_@Y",
    "/GSIData/Functions/Func3Ds/Fn3D/Points/Points_@Z",
    "/GSIData/Geometries",
    "/GSIData/Geometries/Geometry",
    "/GSIData/Geometries/Geometry/Lines",
    "/GSIData/Geometries/Geometry/Lines/Line",
    "/GSIData/Geometries/Geometry/Lines/Line/ID",
    "/GSIData/Geometries/Geometry/Lines/Line/PointID1",
    "/GSIData/Geometries/Geometry/Lines/Line/PointID2",
    "/GSIData/Geometries/Geometry/Lines@Len",
    "/GSIData/Geometries/Geometry/Name",
    "/GSIData/Geometries/Geometry/Points",
    "/GSIData/Geometries/Geometry/Points/Point",
    "/GSIData/Geometries/Geometry/Points/Point@ID",
    "/GSIData/Geometries/Geometry/Points/Point@Pinned",
    "/GSIData/Geometries/Geometry/Points/Point@X",
    "/GSIData/Geometries/Geometry/Points/Point@Y",
    "/GSIData/Geometries/Geometry/Points@Len",
    "/GSIData/Geometries/Geometry/Regions",
    "/GSIData/Geometries/Geometry/Regions/Region",
    "/GSIData/Geometries/Geometry/Regions/Region/ID",
    "/GSIData/Geometries/Geometry/Regions/Region/Mesh",
    "/GSIData/Geometries/Geometry/Regions/Region/Mesh@Pattern",
    "/GSIData/Geometries/Geometry/Regions/Region/PointIDs",
    "/GSIData/Geometries/Geometry/Regions@Len",
    "/GSIData/Geometries/Geometry/Window",
    "/GSIData/Geometries/Geometry/Window/Base",
    "/GSIData/Geometries/Geometry/Window/Base@X",
    "/GSIData/Geometries/Geometry/Window/Base@Y",
    "/GSIData/Geometries/Geometry/Window/Zoom",
    "/GSIData/Geometries@Len",
    "/GSIData/Materials",
    "/GSIData/Materials/Material",
    "/GSIData/Materials/Material/Color",
    "/GSIData/Materials/Material/ID",
    "/GSIData/Materials/Material/Name",
    "/GSIData/Materials/Material/SlopeModel",
    "/GSIData/Materials/Material/StressStrain",
    "/GSIData/Materials/Material/StressStrain/CohesionPrime",
    "/GSIData/Materials/Material/StressStrain/PhiPrime",
    "/GSIData/Materials/Material/StressStrain/UnitWeight",
    "/GSIData/Materials/Material/StressModel",
    "/GSIData/Materials/Material/StressStrain/JointEffectiveCohesion",
    "/GSIData/Materials/Material/StressStrain/JointEffectiveCohesion@Missing",
    "/GSIData/Materials/Material/StressStrain/OCRatio",
    "/GSIData/Materials/Material/StressStrain/OCRatio@Missing",
    "/GSIData/Materials/Material/StressStrain/TensileStrength",
    "/GSIData/Materials/Material/StressStrain/TensileStrength@Missing",
    "/GSIData/Materials/Material/StressStrain/UseTensileStrength",
    "/GSIData/Materials@Len",
    # The reinforcement product library, and the loads that hang off the analysis. These
    # branches of the exporter went unchecked for a while because the fixture model has
    # none of them; the coupled-export case below carries a surcharge, which is what
    # surfaced the gap. Every path here is one the vendor's own files write.
    "/GSIData/Reinforcements",
    "/GSIData/Reinforcements@Len",
    "/GSIData/Reinforcements/Reinforcement",
    "/GSIData/Reinforcements/Reinforcement/Color",
    "/GSIData/Reinforcements/Reinforcement/ID",
    "/GSIData/Reinforcements/Reinforcement/Name",
    "/GSIData/Reinforcements/Reinforcement/PlateCapacity",
    "/GSIData/Reinforcements/Reinforcement/PulloutResistance",
    "/GSIData/Reinforcements/Reinforcement/Spacing",
    "/GSIData/Reinforcements/Reinforcement/Tensile",
    "/GSIData/Reinforcements/Reinforcement/Type",
    "/GSIData/StabilityItems",
    "/GSIData/StabilityItems/StabilityItem",
    "/GSIData/StabilityItems/StabilityItem/AnalysisID",
    "/GSIData/StabilityItems/StabilityItem/Entry",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints@Len",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints/LineLoadPoint",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints/LineLoadPoint/ID",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints/LineLoadPoint/LineLoad",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints/LineLoadPoint/LineLoad/Direction",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints/LineLoadPoint/LineLoad/Direction@Plunge",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints/LineLoadPoint/LineLoad/Direction@Trend",
    "/GSIData/StabilityItems/StabilityItem/Entry/LineLoadPoints/LineLoadPoint/LineLoad/Value",
    "/GSIData/StabilityItems/StabilityItem/Entry/ReinforcementLines",
    "/GSIData/StabilityItems/StabilityItem/Entry/ReinforcementLines@Len",
    "/GSIData/StabilityItems/StabilityItem/Entry/ReinforcementLines/ReinforcementLine",
    "/GSIData/StabilityItems/StabilityItem/Entry/ReinforcementLines/ReinforcementLine@ID",
    "/GSIData/StabilityItems/StabilityItem/Entry/ReinforcementLines/ReinforcementLine@Point1Id",
    "/GSIData/StabilityItems/StabilityItem/Entry/ReinforcementLines/ReinforcementLine@Point2Id",
    "/GSIData/StabilityItems/StabilityItem/Entry/ReinforcementLines/ReinforcementLine@Reinforcement",
    "/GSIData/StabilityItems/StabilityItem/Entry/Surcharges",
    "/GSIData/StabilityItems/StabilityItem/Entry/Surcharges@Len",
    "/GSIData/StabilityItems/StabilityItem/Entry/Surcharges/Surcharge",
    "/GSIData/StabilityItems/StabilityItem/Entry/Surcharges/Surcharge/DataPoints",
    "/GSIData/StabilityItems/StabilityItem/Entry/Surcharges/Surcharge/DataPoints@Len",
    "/GSIData/StabilityItems/StabilityItem/Entry/Surcharges/Surcharge/DataPoints/DataPoint",
    "/GSIData/StabilityItems/StabilityItem/Entry/Surcharges/Surcharge/ID",
    "/GSIData/StabilityItems/StabilityItem/Entry/Surcharges/Surcharge/Pressure",
    # The analysis's LOCAL point list. The piezometric surface, tension crack, surcharges
    # and reinforcement all index INTO this, by Number — not into the geometry <Points>.
    "/GSIData/StabilityItems/StabilityItem/Entry/DataPoints",
    "/GSIData/StabilityItems/StabilityItem/Entry/DataPoints/DataPoint",
    "/GSIData/StabilityItems/StabilityItem/Entry/DataPoints/DataPoint@Number",
    "/GSIData/StabilityItems/StabilityItem/Entry/DataPoints/DataPoint@X",
    "/GSIData/StabilityItems/StabilityItem/Entry/DataPoints/DataPoint@Y",
    "/GSIData/StabilityItems/StabilityItem/Entry/DataPoints@Len",
    "/GSIData/StabilityItems/StabilityItem/Entry/MaterialUsesPiezs",
    "/GSIData/StabilityItems/StabilityItem/Entry/MaterialUsesPiezs/MaterialUsesPiez",
    "/GSIData/StabilityItems/StabilityItem/Entry/MaterialUsesPiezs/MaterialUsesPiez@ID",
    "/GSIData/StabilityItems/StabilityItem/Entry/MaterialUsesPiezs/MaterialUsesPiez@UsesID",
    "/GSIData/StabilityItems/StabilityItem/Entry/MaterialUsesPiezs@Len",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces/PiezometricSurface",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces/PiezometricSurface/CapSuction",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces/PiezometricSurface/DataPoints",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces/PiezometricSurface/DataPoints/DataPoint",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces/PiezometricSurface/DataPoints@Len",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces/PiezometricSurface/ID",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces/PiezometricSurface/MaxSuction",
    "/GSIData/StabilityItems/StabilityItem/Entry/PiezometricSurfaces@Len",
    "/GSIData/StabilityItems/StabilityItem/Entry/Seismic",
    "/GSIData/StabilityItems/StabilityItem/Entry/Seismic@Horizontal",
    "/GSIData/StabilityItems/StabilityItem/Entry/Seismic@Vertical",
    "/GSIData/StabilityItems@Len",
    "/GSIData/WaterItems",
    "/GSIData/WaterItems/WaterItem",
    "/GSIData/WaterItems/WaterItem/AnalysisID",
    "/GSIData/WaterItems/WaterItem/Entry",
    "/GSIData/WaterItems/WaterItem/Entry/ResultInputInfo",
    "/GSIData/WaterItems/WaterItem/Entry/ResultInputInfo/DataGGID",
    "/GSIData/WaterItems/WaterItem/Entry/ResultInputInfo/Option",
    "/GSIData/WaterItems/WaterItem/Entry/UnitWaterWeight",
    "/GSIData/WaterItems@Len",
    "/GSIData@AppVersion",
    "/GSIData@Version",
}


# The OTHER half of the conformance check, and the one that matters more.
#
# GSZ_SCHEMA_PATHS above says "every tag we write must be one GeoStudio writes". That
# can only catch a tag we invent. It CANNOT catch one we leave out -- and an omission is
# the more dangerous bug, because the file still opens and still draws. We shipped a
# .gsz with no <ComputedPhysics>, so GeoStudio did not know the analysis solved slope
# stability: the geometry rendered, the materials arrived named and correctly coloured,
# and their strengths were simply unreachable -- the Properties button greyed out, with
# nothing said. Only opening it in GeoStudio revealed that.
#
# These are the tag paths GeoStudio writes in EVERY file of a real corpus, minus the ones
# xslope consciously does not write (search definitions, solved results, the saved window
# scroll state, SEEP/W physics -- see tools/gsz_export_diff.py, which regenerates this).
# Like the vocabulary above, a list of tag names is a fact about the format and commits
# cleanly even though the vendor files themselves never can.
GSZ_REQUIRED_PATHS = {
    "/GSIData",
    "/GSIData/Analyses",
    "/GSIData/Analyses/Analysis",
    "/GSIData/Analyses/Analysis/ComputedPhysics",
    "/GSIData/Analyses/Analysis/ComputedPhysics@SlopeStability",
    "/GSIData/Analyses/Analysis/GeometryId",
    "/GSIData/Analyses/Analysis/ID",
    "/GSIData/Analyses/Analysis/IterationControls",
    "/GSIData/Analyses/Analysis/IterationControls/IterationControl",
    "/GSIData/Analyses/Analysis/IterationControls/IterationControl/Entry",
    "/GSIData/Analyses/Analysis/IterationControls/IterationControl/Entry@MaxIterations",
    "/GSIData/Analyses/Analysis/IterationControls/IterationControl/Entry@MaxReviewIterations",
    "/GSIData/Analyses/Analysis/IterationControls/IterationControl/Key",
    "/GSIData/Analyses/Analysis/IterationControls@Len",
    "/GSIData/Analyses/Analysis/Kind",
    "/GSIData/Analyses/Analysis/Method",
    "/GSIData/Analyses/Analysis/Name",
    "/GSIData/Analyses/Analysis/TimeIncrements",
    "/GSIData/Analyses/Analysis/TimeIncrements/Duration",
    "/GSIData/Analyses/Analysis/TimeIncrements/Duration@Missing",
    "/GSIData/Analyses/Analysis/TimeIncrements/TimeSteps",
    "/GSIData/Analyses/Analysis/TimeIncrements/TimeSteps/TimeStep",
    "/GSIData/Analyses/Analysis/TimeIncrements/TimeSteps/TimeStep@Save",
    "/GSIData/Analyses/Analysis/TimeIncrements/TimeSteps@Len",
    "/GSIData/Analyses@Len",
    "/GSIData/Contexts",
    "/GSIData/Contexts/Context",
    "/GSIData/Contexts/Context/AnalysisID",
    "/GSIData/Contexts/Context/GeometryUsesMaterials",
    "/GSIData/Contexts/Context/GeometryUsesMaterials/GeometryUsesMaterial",
    "/GSIData/Contexts/Context/GeometryUsesMaterials/GeometryUsesMaterial@Entry",
    "/GSIData/Contexts/Context/GeometryUsesMaterials/GeometryUsesMaterial@ID",
    "/GSIData/Contexts/Context/GeometryUsesMaterials@Len",
    "/GSIData/Contexts/Context/IsDefined",
    "/GSIData/Contexts@Len",
    "/GSIData/Coordinates",
    "/GSIData/Coordinates/EngCoords",
    "/GSIData/Coordinates/EngCoords@HorzScale",
    "/GSIData/Coordinates/EngCoords@LockScales",
    "/GSIData/Coordinates/EngCoords@MaxSnapDist",
    "/GSIData/Coordinates/EngCoords@UnitSystem",
    "/GSIData/Coordinates/EngCoords@VertScale",
    "/GSIData/Coordinates/EngCoords@XPageLeft",
    "/GSIData/Coordinates/EngCoords@XPageOrg",
    "/GSIData/Coordinates/EngCoords@XPageRight",
    "/GSIData/Coordinates/EngCoords@YPageBottom",
    "/GSIData/Coordinates/EngCoords@YPageOrg",
    "/GSIData/Coordinates/EngCoords@YPageTop",
    "/GSIData/Coordinates/PageCoords",
    "/GSIData/Coordinates/PageCoords@PageHeight",
    "/GSIData/Coordinates/PageCoords@PageWidth",
    "/GSIData/Coordinates/PageCoords@PageXOrg",
    "/GSIData/Coordinates/PageCoords@PageYOrg",
    "/GSIData/Coordinates/PageCoords@Units",
    "/GSIData/FileInfo",
    "/GSIData/FileInfo@AppVersion",
    "/GSIData/FileInfo@Author",
    "/GSIData/FileInfo@Comments",
    "/GSIData/FileInfo@Date",
    "/GSIData/FileInfo@FileVersion",
    "/GSIData/FileInfo@LastAuthor",
    "/GSIData/FileInfo@RevNumber",
    "/GSIData/FileInfo@Time",
    "/GSIData/Geometries",
    "/GSIData/Geometries/Geometry",
    "/GSIData/Geometries/Geometry/Lines",
    "/GSIData/Geometries/Geometry/Lines/Line",
    "/GSIData/Geometries/Geometry/Lines/Line/ID",
    "/GSIData/Geometries/Geometry/Lines/Line/PointID1",
    "/GSIData/Geometries/Geometry/Lines/Line/PointID2",
    "/GSIData/Geometries/Geometry/Lines@Len",
    "/GSIData/Geometries/Geometry/Name",
    "/GSIData/Geometries/Geometry/Points",
    "/GSIData/Geometries/Geometry/Points/Point",
    "/GSIData/Geometries/Geometry/Points/Point@ID",
    "/GSIData/Geometries/Geometry/Points/Point@X",
    "/GSIData/Geometries/Geometry/Points/Point@Y",
    "/GSIData/Geometries/Geometry/Points@Len",
    "/GSIData/Geometries/Geometry/Regions",
    "/GSIData/Geometries/Geometry/Regions/Region",
    "/GSIData/Geometries/Geometry/Regions/Region/ID",
    "/GSIData/Geometries/Geometry/Regions/Region/PointIDs",
    "/GSIData/Geometries/Geometry/Regions@Len",
    "/GSIData/Geometries@Len",
    "/GSIData/Materials",
    "/GSIData/Materials/Material",
    "/GSIData/Materials/Material/Color",
    "/GSIData/Materials/Material/ID",
    "/GSIData/Materials/Material/Name",
    "/GSIData/Materials/Material/SlopeModel",
    "/GSIData/Materials/Material/StressStrain",
    "/GSIData/Materials/Material/StressStrain/CohesionPrime",
    "/GSIData/Materials/Material/StressStrain/DisturbanceFactor",
    "/GSIData/Materials/Material/StressStrain/DisturbanceFactor@Missing",
    "/GSIData/Materials/Material/StressStrain/GeologicalStrengthIndex",
    "/GSIData/Materials/Material/StressStrain/GeologicalStrengthIndex@Missing",
    "/GSIData/Materials/Material/StressStrain/IntactRockParam",
    "/GSIData/Materials/Material/StressStrain/IntactRockParam@Missing",
    "/GSIData/Materials/Material/StressStrain/MaxConfiningStress",
    "/GSIData/Materials/Material/StressStrain/MaxConfiningStress@Missing",
    "/GSIData/Materials/Material/StressStrain/UnitWeight",
    "/GSIData/Materials@Len",
    "/GSIData/StabilityItems",
    "/GSIData/StabilityItems/StabilityItem",
    "/GSIData/StabilityItems/StabilityItem/AnalysisID",
    "/GSIData/StabilityItems/StabilityItem/Entry",
    "/GSIData/StabilityItems/StabilityItem/Entry/Seismic",
    "/GSIData/StabilityItems/StabilityItem/Entry/Seismic@Horizontal",
    "/GSIData/StabilityItems/StabilityItem/Entry/Seismic@Vertical",
    "/GSIData/StabilityItems@Len",
    "/GSIData/WaterItems",
    "/GSIData/WaterItems/WaterItem",
    "/GSIData/WaterItems/WaterItem/AnalysisID",
    "/GSIData/WaterItems/WaterItem/Entry",
    "/GSIData/WaterItems/WaterItem/Entry/UnitWaterWeight",
    "/GSIData/WaterItems/WaterItem/Entry/WaterBulkMod",
    "/GSIData/WaterItems@Len",
    "/GSIData@AppVersion",
    "/GSIData@Version",
}

# Every REQUIRED path is by construction a path GeoStudio writes, so the vocabulary
# must contain it. Keeping the two in sync here means a new required path can never
# be rejected by the allow-list it belongs to.
GSZ_SCHEMA_PATHS |= GSZ_REQUIRED_PATHS



def _write_synthetic_gsz(path):
    """Author a minimal GeoStudio .gsz — a two-analysis, two-material model over one
    region — to exercise the importer.

    The fixture is written here rather than shipped because GeoStudio's own sample and
    verification files are Seequent's copyrighted Materials and must not be committed to
    this repository. Authoring the XML also pins the exact schema the importer relies on
    -- and every fact below was learned by scoring the importer against SLOPE/W's own
    answers over a corpus of real files (tools/gsz_corpus.py). Each was silent when
    wrong, and each cost real accuracy:

      * material assignment lives in <Contexts>, PER ANALYSIS, not on the region;
      * an undrained material keeps its strength in <Cohesion>, not <CohesionPrime>;
      * <PiezometricSurfaces> hangs off StabilityItem/Entry, and indexes that analysis's
        LOCAL <DataPoints> -- not the shared geometry <Points> table (10% of FS);
      * a tension crack is switched off by DROPPING <TensionOption>; GeoStudio keeps the
        crack's geometry, so the points alone mean nothing (2% of FS);
      * a <Surcharge>'s <Pressure> is a UNIT WEIGHT, and the load is the weight of the
        fill between the drawn line and the ground -- not a uniform pressure (4% of FS);
      * reinforcement carries no direction: it is implied by <Type>;
      * a stability analysis whose water <Option> is "Parent" takes a whole pore-pressure
        FIELD from a SEEP/W run, on a binary PLY mesh -- and SLOPE/W raises the reservoir
        standing on the slope from that same field, storing no water surface at all
        (13% of FS, and most of it in the missing reservoir rather than the pressures).
    """
    import zipfile
    xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<GSIData Version="11.11" AppVersion="25.2.1.4">
  <FileInfo Title="synthetic" />
  <Coordinates><EngCoords UnitSystem="Metric" /></Coordinates>
  <Analyses Len="9">
    <Analysis><ID>1</ID><Name>dry</Name><Kind>SLOPE/W</Kind>
      <Method>Morgenstern-Price</Method><GeometryId>1</GeometryId></Analysis>
    <Analysis><ID>2</ID><Name>wet</Name><Kind>SLOPE/W</Kind>
      <Method>Spencer</Method><GeometryId>1</GeometryId></Analysis>
    <Analysis><ID>3</ID><Name>loaded</Name><Kind>SLOPE/W</Kind>
      <Method>Spencer</Method><GeometryId>1</GeometryId></Analysis>
    <Analysis><ID>4</ID><Name>crack off</Name><Kind>SLOPE/W</Kind>
      <Method>Spencer</Method><GeometryId>1</GeometryId></Analysis>
    <Analysis><ID>5</ID><Name>nailed</Name><Kind>SLOPE/W</Kind>
      <Method>Spencer</Method><GeometryId>1</GeometryId></Analysis>
    <Analysis><ID>6</ID><Name>seepage</Name><Kind>SEEP/W</Kind>
      <Method>SteadyState</Method><GeometryId>1</GeometryId></Analysis>
    <Analysis><ID>7</ID><Name>seep stability</Name><Kind>SLOPE/W</Kind>
      <ParentID>6</ParentID><Method>Spencer</Method><GeometryId>1</GeometryId></Analysis>
    <Analysis><ID>8</ID><Name>tall</Name><Kind>SLOPE/W</Kind>
      <Method>Morgenstern-Price</Method><GeometryId>2</GeometryId></Analysis>
    <!-- A PROBABILISTIC analysis. SLOPE/W states uncertainty on the ANALYSIS, not on
         the material: <SampleSettings> holds one <Modifier> per uncertain parameter,
         keyed by material ID, whose <Entry Function> distributes the DEVIATION from the
         applied value (Mean=0, SD=sigma). The same material 1 is deterministic in
         analysis 1 and probabilistic here, which is what makes it per-analysis.
         The modifier list deliberately mixes what CAN be imported with four things
         that cannot, each of which must be reported rather than dropped in silence. -->
    <Analysis><ID>9</ID><Name>probabilistic</Name><Kind>SLOPE/W</Kind>
      <Method>Spencer</Method><GeometryId>1</GeometryId>
      <SampleSettings>
        <Type>Probabilistic</Type>
        <NumMonteCarloTrials>5000</NumMonteCarloTrials>
        <ProbabilisticModifiers Len="8">
          <Modifier><Key>Materials[1].Applied.UnitWeight</Key>
            <Entry Function="Normal(Mean=0,SD=1.5)" /></Modifier>
          <Modifier><Key>Materials[1].Applied.EffectiveCohesion</Key>
            <Entry Function="Normal(Mean=0,SD=5,Min=-10,Max=10)"
                   CorrelationTag="Materials[1].Applied.EffectivePhi"
                   CorrelationCoefficient="-0.5" /></Modifier>
          <Modifier><Key>Materials[1].Applied.EffectivePhi</Key>
            <Entry Function="Normal(Mean=0,SD=3)" /></Modifier>
          <Modifier><Key>Materials[1].StressStrain.Ru</Key>
            <Entry Function="Normal(Mean=0,SD=0.02)" /></Modifier>
          <Modifier><Key>Materials[4].Applied.EffectiveCohesion</Key>
            <Entry Function="Normal(Mean=0,SD=7)" /></Modifier>
          <Modifier><Key>Materials[4].Applied.UnitWeight</Key>
            <Entry Function="Uniform(Min=-1,Max=1)" /></Modifier>
          <Modifier><Key>Materials[2].Applied.EffectiveCohesion</Key>
            <Entry Function="Normal(Mean=0,SD=99)" /></Modifier>
          <Modifier><Key>Loads[1].Magnitude</Key>
            <Entry Function="Normal(Mean=0,SD=4)" /></Modifier>
        </ProbabilisticModifiers>
      </SampleSettings></Analysis>
  </Analyses>
  <Geometries Len="2">
    <Geometry><Name>2D</Name>
      <Points Len="8">
        <Point ID="1" X="0"  Y="0" /><Point ID="2" X="60" Y="0" />
        <Point ID="3" X="60" Y="30" /><Point ID="4" X="40" Y="30" />
        <Point ID="5" X="20" Y="10" /><Point ID="6" X="0"  Y="10" />
        <!-- A bedrock layer beneath the slope. Only the probabilistic analysis
             assigns region 2, so every other analysis is untouched by it. -->
        <Point ID="7" X="0"  Y="-10" /><Point ID="8" X="60" Y="-10" />
      </Points>
      <Regions Len="2">
        <Region><ID>1</ID><PointIDs>1,2,3,4,5,6</PointIDs></Region>
        <Region><ID>2</ID><PointIDs>7,8,2,1</PointIDs></Region>
      </Regions>
    </Geometry>
    <!-- A SECOND geometry in the same file — the "same embankment at another height"
         shape (Borges & Cardoso Case 3). It REUSES point IDs 1-6 and region ID 1, but a
         taller slope: crest at y=45, not y=30. GeoStudio gives the geometry element no
         <ID>; the analysis names it by POSITION via <GeometryId>. Because the IDs collide,
         a single merged point table resolves every analysis to whichever geometry was
         parsed last — here geometry 2 — so analysis 1 would silently import THIS slope. -->
    <Geometry><Name>2D tall</Name>
      <Points Len="6">
        <Point ID="1" X="0"  Y="0" /><Point ID="2" X="60" Y="0" />
        <Point ID="3" X="60" Y="45" /><Point ID="4" X="40" Y="45" />
        <Point ID="5" X="20" Y="15" /><Point ID="6" X="0"  Y="15" />
      </Points>
      <Regions Len="1">
        <Region><ID>1</ID><PointIDs>1,2,3,4,5,6</PointIDs></Region>
      </Regions>
    </Geometry>
  </Geometries>
  <Materials Len="4">
    <Material><ID>1</ID><Name>strong</Name><Color>RGB=(211,201,137)</Color>
      <SlopeModel>MohrCoulomb</SlopeModel>
      <StressStrain><UnitWeight>20</UnitWeight><CohesionPrime>25</CohesionPrime>
        <PhiPrime>30</PhiPrime></StressStrain></Material>
    <Material><ID>2</ID><Name>weak</Name><Color>RGB=(0,128,255)</Color>
      <SlopeModel>MohrCoulomb</SlopeModel>
      <StressStrain><UnitWeight>18</UnitWeight><CohesionPrime>5</CohesionPrime>
        <PhiPrime>20</PhiPrime></StressStrain></Material>
    <Material><ID>3</ID><Name>undrained</Name><SlopeModel>UndrainedPhiZero</SlopeModel>
      <StressStrain><UnitWeight>19</UnitWeight><Cohesion>40</Cohesion></StressStrain></Material>
    <Material><ID>4</ID><Name>bedrock</Name><SlopeModel>Bedrock</SlopeModel>
      <StressStrain><UnitWeight>22</UnitWeight></StressStrain></Material>
  </Materials>
  <Reinforcements Len="1">
    <Reinforcement><ID>1</ID><Name>Soil Nails</Name><Type>Nail</Type>
      <Spacing>2</Spacing><Tensile>100</Tensile><PlateCapacity>60</PlateCapacity>
      <PulloutResistance>10</PulloutResistance></Reinforcement>
  </Reinforcements>
  <Contexts Len="8">
    <Context><AnalysisID>1</AnalysisID><GeometryUsesMaterials Len="1">
        <GeometryUsesMaterial ID="Regions-1" Entry="1" /></GeometryUsesMaterials></Context>
    <Context><AnalysisID>2</AnalysisID><GeometryUsesMaterials Len="1">
        <GeometryUsesMaterial ID="Regions-1" Entry="2" /></GeometryUsesMaterials></Context>
    <Context><AnalysisID>3</AnalysisID><GeometryUsesMaterials Len="1">
        <GeometryUsesMaterial ID="Regions-1" Entry="3" /></GeometryUsesMaterials></Context>
    <Context><AnalysisID>4</AnalysisID><GeometryUsesMaterials Len="1">
        <GeometryUsesMaterial ID="Regions-1" Entry="1" /></GeometryUsesMaterials></Context>
    <Context><AnalysisID>5</AnalysisID><GeometryUsesMaterials Len="1">
        <GeometryUsesMaterial ID="Regions-1" Entry="1" /></GeometryUsesMaterials></Context>
    <Context><AnalysisID>7</AnalysisID><GeometryUsesMaterials Len="1">
        <GeometryUsesMaterial ID="Regions-1" Entry="1" /></GeometryUsesMaterials></Context>
    <Context><AnalysisID>8</AnalysisID><GeometryUsesMaterials Len="1">
        <GeometryUsesMaterial ID="Regions-1" Entry="1" /></GeometryUsesMaterials></Context>
    <Context><AnalysisID>9</AnalysisID><GeometryUsesMaterials Len="2">
        <GeometryUsesMaterial ID="Regions-1" Entry="1" />
        <GeometryUsesMaterial ID="Regions-2" Entry="4" /></GeometryUsesMaterials></Context>
  </Contexts>
  <WaterItems Len="8">
    <WaterItem><AnalysisID>1</AnalysisID>
      <Entry><UnitWaterWeight>9.807</UnitWaterWeight></Entry></WaterItem>
    <WaterItem><AnalysisID>2</AnalysisID>
      <Entry><ResultInputInfo><Option>PiezoSurface</Option></ResultInputInfo>
        <UnitWaterWeight>9.807</UnitWaterWeight></Entry></WaterItem>
    <WaterItem><AnalysisID>3</AnalysisID>
      <Entry><UnitWaterWeight>9.807</UnitWaterWeight></Entry></WaterItem>
    <WaterItem><AnalysisID>4</AnalysisID>
      <Entry><UnitWaterWeight>9.807</UnitWaterWeight></Entry></WaterItem>
    <WaterItem><AnalysisID>5</AnalysisID>
      <Entry><UnitWaterWeight>9.807</UnitWaterWeight></Entry></WaterItem>
    <WaterItem><AnalysisID>7</AnalysisID>
      <Entry><ResultInputInfo><Option>Parent</Option></ResultInputInfo>
        <UnitWaterWeight>9.807</UnitWaterWeight></Entry></WaterItem>
    <WaterItem><AnalysisID>8</AnalysisID>
      <Entry><UnitWaterWeight>9.807</UnitWaterWeight></Entry></WaterItem>
    <WaterItem><AnalysisID>9</AnalysisID>
      <Entry><UnitWaterWeight>9.807</UnitWaterWeight></Entry></WaterItem>
  </WaterItems>
  <StabilityItems Len="8">
    <StabilityItem><AnalysisID>1</AnalysisID>
      <Entry><Seismic Horizontal="" Vertical="" /></Entry></StabilityItem>

    <StabilityItem><AnalysisID>2</AnalysisID>
      <Entry>
        <DataPoints Len="2">
          <DataPoint Number="1" X="0"  Y="8" />
          <DataPoint Number="2" X="60" Y="24" />
        </DataPoints>
        <Seismic Horizontal="0.15" Vertical="" />
        <PiezometricSurfaces Len="1">
          <PiezometricSurface><ID>1</ID>
            <DataPoints Len="2"><DataPoint>1</DataPoint><DataPoint>2</DataPoint></DataPoints>
            <CapSuction>false</CapSuction><MaxSuction>0</MaxSuction>
          </PiezometricSurface>
        </PiezometricSurfaces>
        <MaterialUsesPiezs Len="1"><MaterialUsesPiez ID="2" UsesID="1" /></MaterialUsesPiezs>
      </Entry></StabilityItem>

    <StabilityItem><AnalysisID>3</AnalysisID>
      <Entry>
        <DataPoints Len="4">
          <DataPoint Number="1" X="40" Y="26" />
          <DataPoint Number="2" X="60" Y="26" />
          <DataPoint Number="3" X="42" Y="31" />
          <DataPoint Number="4" X="58" Y="33" />
        </DataPoints>
        <Seismic Horizontal="" Vertical="" />
        <TensionCrack>
          <TensionOption>Surface</TensionOption>
          <PctFilledWithWater>0.5</PctFilledWithWater>
          <DataPoints Len="2"><DataPoint>1</DataPoint><DataPoint>2</DataPoint></DataPoints>
        </TensionCrack>
        <Surcharges Len="1">
          <Surcharge><ID>1</ID>
            <DataPoints Len="2"><DataPoint>3</DataPoint><DataPoint>4</DataPoint></DataPoints>
            <Pressure>20</Pressure>
          </Surcharge>
        </Surcharges>
        <SomeFutureGeoStudioFeature>whatever</SomeFutureGeoStudioFeature>
      </Entry></StabilityItem>

    <StabilityItem><AnalysisID>4</AnalysisID>
      <Entry>
        <DataPoints Len="2">
          <DataPoint Number="1" X="40" Y="26" />
          <DataPoint Number="2" X="60" Y="26" />
        </DataPoints>
        <Seismic Horizontal="" Vertical="" />
        <TensionCrack>
          <PctFilledWithWater>1</PctFilledWithWater>
          <DataPoints Len="2"><DataPoint>1</DataPoint><DataPoint>2</DataPoint></DataPoints>
        </TensionCrack>
      </Entry></StabilityItem>

    <StabilityItem><AnalysisID>5</AnalysisID>
      <Entry>
        <DataPoints Len="2">
          <DataPoint Number="1" X="40" Y="30" />
          <DataPoint Number="2" X="50" Y="20" />
        </DataPoints>
        <Seismic Horizontal="" Vertical="" />
        <ReinforcementLines Len="1">
          <ReinforcementLine ID="1" Reinforcement="1" Point1Id="1" Point2Id="2" />
        </ReinforcementLines>
      </Entry></StabilityItem>

    <StabilityItem><AnalysisID>7</AnalysisID>
      <Entry><Seismic Horizontal="" Vertical="" /></Entry></StabilityItem>

    <StabilityItem><AnalysisID>8</AnalysisID>
      <Entry><Seismic Horizontal="" Vertical="" /></Entry></StabilityItem>

    <StabilityItem><AnalysisID>9</AnalysisID>
      <Entry><Seismic Horizontal="" Vertical="" /></Entry></StabilityItem>
  </StabilityItems>
</GSIData>
"""
    nodes, elements, u = _synthetic_seep_field()
    with zipfile.ZipFile(path, 'w') as zf:
        zf.writestr("synthetic.xml", xml)
        # A SEEP/W-fed stability analysis keeps the TRANSFERRED field in its own result
        # folder, on the mesh SEEP/W solved on. Both are written the way GeoStudio writes
        # them -- a binary PLY, and a node.csv numbered from 1.
        zf.writestr("seep stability/Mesh.ply", _synthetic_ply(nodes, elements))
        zf.writestr("seep stability/000/node.csv",
                    "Node,PoreWaterPressure,XTotalStress\n" +
                    "".join(f"{i + 1},{v}\n" for i, v in enumerate(u)))


# The synthetic seepage model: a flat water table at elevation 15 over the fixture's
# slope, which rises from y=10 at x=20 to y=30 at x=40. So the reservoir stands 5 deep on
# the flat ground left of x=20 and tapers to nothing at x=25, where the face reaches
# elevation 15. Every number the importer must produce is hand-checkable from that.
_SEEP_WATER_LEVEL = 15.0
_SEEP_GAMMA_W = 9.807
_SEEP_COLUMNS = [(0.0, 10.0), (10.0, 10.0), (20.0, 10.0), (25.0, 15.0),
                 (30.0, 20.0), (40.0, 30.0), (50.0, 30.0), (60.0, 30.0)]
_SEEP_ROWS = 5


def _synthetic_seep_field():
    """Nodes, elements and nodal pore pressure for the fixture's SEEP/W analysis.

    The mesh CONFORMS to the slope: its top row of nodes is the ground surface itself,
    so a point sampled just below the ground always lands inside an element. (A mesh that
    merely covered the bounding box would leave the face outside it, and the reservoir
    would silently not be found.)
    """
    nodes, u = [], []
    for x, yg in _SEEP_COLUMNS:
        for r in range(_SEEP_ROWS):
            y = yg * r / (_SEEP_ROWS - 1)
            nodes.append((x, y))
            u.append(_SEEP_GAMMA_W * (_SEEP_WATER_LEVEL - y))    # hydrostatic; <0 above

    def nid(c, r):
        return c * _SEEP_ROWS + r + 1                            # PLY ids are 1-based

    elements = []                                                # (shape_kind, [ids])
    for c in range(len(_SEEP_COLUMNS) - 1):
        for r in range(_SEEP_ROWS - 1):
            quad = [nid(c, r), nid(c + 1, r), nid(c + 1, r + 1), nid(c, r + 1)]
            if c == 0 and r == 0:
                # One quad split into two triangles, so the fixture carries BOTH element
                # types and the importer's element_types has to come out right.
                elements.append((2, [quad[0], quad[1], quad[2]]))
                elements.append((2, [quad[0], quad[2], quad[3]]))
            else:
                elements.append((3, quad))

    # GeoStudio also writes the region's edges and corners into the same block. They are
    # not domain elements; importing them would leave zero-area cells in the mesh.
    elements.append((1, [nid(0, 0), nid(1, 0)]))                 # a line entity
    elements.append((0, [nid(0, 0)]))                            # a point entity
    return nodes, elements, u


def _synthetic_ply(nodes, elements):
    """Write nodes/elements as the binary PLY GeoStudio writes.

    The layout is copied from a real GeoStudio mesh, including the ``version`` block that
    precedes the nodes: a reader that skips it reads every coordinate 4 bytes out of step
    and produces plausible garbage. Writing it here is what pins that.
    """
    import struct
    header = (
        "ply\n"
        "format binary_little_endian 1.0\n"
        "element version 1\n"
        "property ushort major\n"
        "property ushort minor\n"
        f"element node {len(nodes)}\n"
        "property double x\n"
        "property double y\n"
        "property double z\n"
        f"element element {len(elements)}\n"
        "property uchar shape_kind\n"
        "property uchar integration_order\n"
        "property uchar category\n"
        "property uint owner\n"
        "property list uchar uint id\n"
        "element element_line_association 0\n"
        "property uint owner\n"
        "property uint id\n"
        "end_header\n"
    ).encode("ascii")

    body = bytearray(struct.pack("<HH", 1, 0))                   # version major/minor
    for x, y in nodes:
        body += struct.pack("<ddd", x, y, 0.0)
    for shape_kind, ids in elements:
        body += struct.pack("<BBBI", shape_kind, 2, 0, 1)
        body += struct.pack("<B", len(ids))
        body += struct.pack(f"<{len(ids)}I", *ids)
    return bytes(header + body)


def _water_force_columns(slope_data, circle, n=20):
    """The per-slice water/load force columns for one model on one circle.

    The resultant of the distributed load on each slice top, where it acts, how it
    is inclined, and the slice's own weight and boundary pressures -- everything a
    water load reaches the statics through. Returns a tuple of column tuples, or a
    string describing why no frame was built.

    ``check_inputs=False`` because the caller supplies the surface: an imported model
    carries no circles sheet, and the surface rule would refuse a run whose surface
    the caller has already named.
    """
    from xslope.slice import generate_slices
    ok, res = generate_slices(slope_data, circle=dict(circle), num_slices=n,
                              debug=False, check_inputs=False)
    if not ok:
        return f"generate_slices failed: {str(res)[:60]}"
    df = res[0]
    if "dload" not in getattr(df, "columns", []):
        return "degenerate slice frame"
    return tuple(tuple(round(float(v), 12) for v in df[c].tolist())
                 for c in _WATER_FORCE_COLUMNS)


#: The slice-frame columns a distributed water load can move. Compared column by
#: column so a report can say WHICH one moved, rather than that something did.
_WATER_FORCE_COLUMNS = ("dload", "d_x", "d_y", "beta", "qL", "qR", "w")


def run_gsz_import_test(test):
    """Import a synthetic GeoStudio .gsz and check the model that comes out.

    Guards the three schema facts the importer depends on, each of which is easy to
    get wrong and silent when wrong:
      - material assignment is per ANALYSIS (<Contexts>), not per region, so the two
        analyses must yield different materials over the same geometry;
      - a piezometric surface indexes the shared <Points> table by ID;
      - the seismic coefficient rides on the analysis's <StabilityItem>.
    Also checks that an unsolved file reports the missing failure surface as a caveat
    rather than importing a wrong one.

    And the water-load handover (template v22): where the file states a water
    SURFACE the import carries the definition and writes no water dload, because the
    engine derives the same reservoir from the same line — proven here by building
    the slices both ways and requiring the per-slice forces to be identical, not
    close. Where the file states only a solved SEEP/W head field there is no surface
    to carry, so that one keeps its synthesised load and imports manual.
    Returns (0.0, None) or (None, message).
    """
    import tempfile
    from xslope.geostudio import read_gsz, list_analyses, gsz_to_slope_data, gsz_style

    problems = []
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "synthetic.gsz")
        _write_synthetic_gsz(path)
        gsz = read_gsz(path)

        analyses = list_analyses(gsz)
        if len(analyses) != 9:
            return None, f"GeoStudio import: {len(analyses)} analyses, expected 9"

        sd1, cav1 = gsz_to_slope_data(gsz, 1)
        sd2, cav2 = gsz_to_slope_data(gsz, 2)
        sd3, cav3 = gsz_to_slope_data(gsz, 3)
        sd4, cav4 = gsz_to_slope_data(gsz, 4)
        sd5, cav5 = gsz_to_slope_data(gsz, 5)
        sd7, cav7 = gsz_to_slope_data(gsz, 7)
        sd8, cav8 = gsz_to_slope_data(gsz, 8)
        sd9, cav9 = gsz_to_slope_data(gsz, 9)

        # --- probabilistic standard deviations --------------------------------------
        # SLOPE/W puts them on the ANALYSIS. They used to be dropped entirely, so a
        # probabilistic model arrived with every sigma zero and reliability() refused it
        # for having no standard deviations at all — a silent, total loss of the one
        # thing the model was built to state.
        m9 = {m['name']: m for m in sd9['materials']}
        if sorted(m9) != ['bedrock', 'strong']:
            problems.append(f"probabilistic analysis imported materials {sorted(m9)}, "
                            f"expected ['bedrock', 'strong']")
        else:
            got = (m9['strong']['sigma_gamma'], m9['strong']['sigma_c'],
                   m9['strong']['sigma_phi'])
            if got != (1.5, 5.0, 3.0):
                problems.append(f"'strong' imported sigmas (gamma, c, phi) = {got}, "
                                f"expected (1.5, 5.0, 3.0) — SLOPE/W's SD is the "
                                f"deviation's standard deviation, i.e. xslope's sigma")
            # Bedrock is imported as elastic: it cannot fail, so a standard deviation on
            # its strength has nothing to attach to and must not be invented.
            if (m9['bedrock']['option'], m9['bedrock']['sigma_c'],
                    m9['bedrock']['sigma_gamma']) != ('elastic', 0.0, 0.0):
                problems.append(
                    f"bedrock imported as option={m9['bedrock']['option']!r} with "
                    f"sigma_c={m9['bedrock']['sigma_c']}, "
                    f"sigma_gamma={m9['bedrock']['sigma_gamma']}; expected a "
                    f"deterministic elastic material")
        # The SAME material 1 is deterministic in analysis 1. Sampling is per-analysis,
        # so reading it off the material would leak the sigmas into every analysis.
        if any(m['sigma_c'] or m['sigma_phi'] or m['sigma_gamma']
               for m in sd1['materials']):
            problems.append("the deterministic analysis picked up standard deviations — "
                            "sampling is defined per analysis, not per material")
        # A modifier naming a material this analysis does not use is not part of the
        # model: 'weak' (material 2, sigma 99) must not appear anywhere.
        if any(m['sigma_c'] == 99.0 for m in sd9['materials']):
            problems.append("a modifier for an unused material was applied")

        joined9 = " | ".join(cav9)
        for want, why in [
                ("PROBABILISTIC", "the import never says the model is probabilistic"),
                ("truncat", "the Min/Max sampling bounds are not reported"),
                ("correlated", "the c-phi correlation is not reported"),
                ("StressStrain.Ru", "an unmappable modifier (Ru) is not reported"),
                ("Bedrock has no strength", "a sigma on Bedrock strength is not reported"),
                ("Uniform", "a non-normal distribution is not reported"),
                ("Loads[1].Magnitude", "a non-material modifier is not reported")]:
            if want not in joined9:
                problems.append(f"probabilistic caveats: {why} ({want!r} absent)")

        # The reliability guard. A deterministic elastic zone under a probabilistic model
        # is ordinary and legitimate; only an elastic material that itself carries a
        # standard deviation has nowhere to put it.
        from xslope.reliability import _reliability_param_info
        try:
            info9, err9 = _reliability_param_info(sd9['materials'])
        except ValueError as exc:
            info9, err9 = None, f"raised: {exc}"
        if err9 is not None:
            problems.append(f"reliability refused a probabilistic model with a "
                            f"DETERMINISTIC elastic zone: {err9}")
        elif sorted((p['material_name'], p['param']) for p in info9) != \
                [('strong', 'c'), ('strong', 'gamma'), ('strong', 'phi')]:
            problems.append(f"reliability picked up parameters "
                            f"{sorted((p['material_name'], p['param']) for p in info9)}, "
                            f"expected gamma/c/phi on 'strong' alone")
        mutated = [dict(m) for m in sd9['materials']]
        for m in mutated:
            if m['option'] == 'elastic':
                m['sigma_phi'] = 2.0
        try:
            _reliability_param_info(mutated)
            problems.append("reliability accepted a GENUINELY probabilistic elastic "
                            "material — its standard deviation cannot reach the factor "
                            "of safety, so it must be refused, not ignored")
        except ValueError:
            pass

        # --- a stability analysis fed by a parent SEEP/W run -------------------------
        # Its water <Option> is "Parent": pore pressure is a whole finite element FIELD,
        # which SLOPE/W stores on a binary PLY mesh in this analysis's own result folder.
        mesh = sd7.get('mesh')
        n_nodes = len(_SEEP_COLUMNS) * _SEEP_ROWS                    # 8 x 5 = 40
        n_elems = (len(_SEEP_COLUMNS) - 1) * (_SEEP_ROWS - 1) + 1    # 28 quads, one split
        if mesh is None:
            problems.append("the SEEP/W pore-pressure field did not import: no mesh")
        else:
            if len(mesh['nodes']) != n_nodes:
                problems.append(f"SEEP/W mesh has {len(mesh['nodes'])} nodes, expected "
                                f"{n_nodes} — the PLY was misread")
            # The line and point entities in the same PLY block are NOT domain elements.
            if len(mesh['elements']) != n_elems:
                problems.append(f"SEEP/W mesh has {len(mesh['elements'])} elements, "
                                f"expected {n_elems} — GeoStudio writes region edges and "
                                f"corners into the same block, and they are not elements")
            types = sorted(set(int(t) for t in mesh['element_types']))
            if types != [3, 4]:
                problems.append(f"SEEP/W element types {types}, expected [3, 4]")
            # The coordinates themselves: a reader that skips the PLY's <version> block
            # is 4 bytes out of step and returns plausible garbage.
            xs = [float(p[0]) for p in mesh['nodes']]
            ys = [float(p[1]) for p in mesh['nodes']]
            if (min(xs), max(xs), min(ys), max(ys)) != (0.0, 60.0, 0.0, 30.0):
                problems.append(f"SEEP/W mesh spans x {min(xs)}..{max(xs)}, y {min(ys)}.."
                                f"{max(ys)}, expected 0..60 and 0..30 — the binary PLY "
                                f"was decoded at the wrong offset")

        u = sd7.get('seep_u')
        if u is None or len(u) != n_nodes:
            problems.append(f"seep_u is {None if u is None else len(u)}, expected "
                            f"{n_nodes} nodal pore pressures")
        if [m['u'] for m in sd7['materials']] != ['seep']:
            problems.append(f"materials take u={[m['u'] for m in sd7['materials']]}, "
                            f"expected ['seep'] — a Parent water option is a seepage field")

        # THE RESERVOIR. GeoStudio stores no ponded-water object and, with a SEEP/W
        # field, no water surface either — yet SLOPE/W still puts the weight of the water
        # on the submerged face. It can only be deriving it from the head field, and it
        # is: water stands to y + u/gamma_w at the ground. Here that is a flat table at
        # elevation 15 over ground at elevation 10, so 5 deep (49.035) out to x=20, then
        # tapering to zero at x=25 where the slope face reaches 15. Missing this cost 13%
        # of the factor of safety on GeoStudio's own rapid-drawdown example — MORE than
        # the pore pressures did.
        dl7 = sd7['dloads']
        if len(dl7) != 1:
            problems.append(f"the reservoir implied by the SEEP/W field gave {len(dl7)} "
                            f"dload blocks, expected 1")
        else:
            deep = max(p['Normal'] for p in dl7[0])
            xmax = max(p['X'] for p in dl7[0])
            if abs(deep - _SEEP_GAMMA_W * 5.0) > 0.05:
                problems.append(f"ponded water from the SEEP/W field is {deep:.2f} at its "
                                f"deepest, expected {_SEEP_GAMMA_W * 5.0:.2f} (5 deep)")
            if abs(xmax - 25.0) > 0.3:
                problems.append(f"ponded water from the SEEP/W field reaches x={xmax:.2f}, "
                                f"expected 25 — where the face rises to the water level")
            if abs(min(p['Normal'] for p in dl7[0])) > 1e-6:
                problems.append("ponded water must taper to zero pressure at the "
                                "waterline, not stop at full depth")

        # --- the water-load handover (template v22, main!D23) -----------------------
        # THREE analyses, three answers, and the split is the whole point.
        from xslope.water import (with_water_loads, derived_blocks,
                                  ponded_water_dload)

        # (a) A PIEZOMETRIC SURFACE is a water definition, so the import carries it and
        # writes no water dload at all: automatic mode, and the engine measures the
        # same reservoir off the same line at solve time.
        if sd2.get('water_loads') != 'auto':
            problems.append(
                f"an analysis whose water is a piezometric surface imported as "
                f"{sd2.get('water_loads')!r}; it must import as 'auto' — the water "
                f"line is carried and the engine derives the load from it")
        if sd2['dloads']:
            problems.append(
                f"a piezo-surface analysis wrote {len(sd2['dloads'])} distributed "
                f"load(s); under automatic water loads it must write none, or the "
                f"reservoir is stated twice — once on the sheet and once derived")
        auto_blocks = derived_blocks(with_water_loads(sd2), 1)
        if len(auto_blocks) != 1:
            problems.append(f"the engine derived {len(auto_blocks)} water block(s) "
                            f"from the imported piezo line, expected 1 — the water "
                            f"was carried but nothing measures it")
        if not any('derive automatically from the imported water definition' in c
                   for c in cav2):
            problems.append("the import did not say that the water loads now derive "
                            "automatically from the imported water definition")

        # THE EQUIVALENCE GATE. What the engine derives must reproduce what the import
        # used to synthesise, at the level that matters: the force on each slice. A
        # manual twin is built carrying exactly the blocks the old synthesis produced,
        # and the two models are sliced on the same circle. Anything but identical is
        # a regression, and the conformance runs against SLOPE/W's own trial surfaces
        # rest on this being exact rather than close.
        eq_circle = {'Xo': 25.0, 'Yo': 40.0, 'R': 35.0, 'Depth': None, 'Y': None,
                     'Option': 'Depth', 'Movement': 'Left to Right'}
        synth = ponded_water_dload(sd2['ground_surface'], sd2['piezo_line'],
                                   sd2['gamma_water'])
        twin = dict(sd2, water_loads='manual', dloads=synth)
        cols_auto = _water_force_columns(sd2, eq_circle)
        cols_twin = _water_force_columns(twin, eq_circle)
        if isinstance(cols_auto, str) or isinstance(cols_twin, str):
            problems.append(f"the water-load equivalence check could not be made: "
                            f"auto={cols_auto if isinstance(cols_auto, str) else 'ok'}, "
                            f"synthesised="
                            f"{cols_twin if isinstance(cols_twin, str) else 'ok'}")
        else:
            if cols_auto != cols_twin:
                bad = [n for n, a, b in zip(_WATER_FORCE_COLUMNS,
                                            cols_auto, cols_twin) if a != b]
                problems.append(
                    f"the derived water load does not reproduce the synthesised one: "
                    f"the per-slice {', '.join(bad)} differ. The import stopped "
                    f"writing the block on the understanding that the engine builds "
                    f"the same forces from the same line")
            if not any(v for v in cols_auto[0]):
                problems.append("the equivalence circle carries no water force at "
                                "all, so it proves nothing — pick one that crosses "
                                "the loaded ground")
            # ... and the comparison can fail. A twin whose block is 1% heavier must
            # be caught, or the check above is only asserting that two runs of the
            # same code agree.
            heavy = [[dict(p, Normal=p['Normal'] * 1.01) for p in b] for b in synth]
            if _water_force_columns(dict(twin, dloads=heavy), eq_circle) == cols_auto:
                problems.append("a 1% heavier water load compared EQUAL to the "
                                "derived one — the equivalence check cannot fail")

        # (b) A SURCHARGE is user data and keeps arriving as a dload (asserted above);
        # the model still imports auto, and there is no water for the engine to derive.
        if sd3.get('water_loads') != 'auto':
            problems.append(f"a surcharge-only analysis imported as "
                            f"{sd3.get('water_loads')!r}, expected 'auto'")
        if derived_blocks(with_water_loads(sd3), 1):
            problems.append("a model with no water definition derived a water load")

        # (c) A SEEP/W FIELD states no water surface anywhere, and the derivation reads
        # a piezo line or seepage head boundaries — neither of which an imported field
        # is. So this one keeps the reservoir recovered from the field, and says so by
        # importing MANUAL. Flipping it to auto would drop the load entirely: 13% of
        # the factor of safety on GeoStudio's own drawdown example.
        if sd7.get('water_loads') != 'manual':
            problems.append(
                f"a SEEP/W-coupled analysis imported as {sd7.get('water_loads')!r}; "
                f"it must import as 'manual' — its reservoir was recovered from the "
                f"head field, and nothing downstream can re-derive it")
        if derived_blocks(with_water_loads(sd7), 1):
            problems.append("a SEEP/W-coupled model derived a second water load on "
                            "top of the one recovered from its head field")
        if not any('D23 set to manual' in c or 'set to manual' in c for c in cav7):
            problems.append("the SEEP/W-coupled import did not report that it left "
                            "the water-load mode on manual")

        # An UNDRAINED material keeps its strength in <Cohesion>, not <CohesionPrime>.
        # Reading only the drained field gave c = 0, phi = 0 — a soil with no strength
        # and a meaningless factor of safety, with nothing said about it.
        m3 = sd3['materials'][0]
        if (m3['name'], m3['c'], m3['phi']) != ('undrained', 40.0, 0.0):
            problems.append(f"undrained material -> c={m3['c']}, phi={m3['phi']}, "
                            f"expected c=40 (from <Cohesion>), phi=0")

        # Tension crack: depth below ground, and water DEPTH in the crack (not a %).
        # The crest is at y=30 over x=40..60 and the crack line sits at y=26, so the
        # crack is 4 deep and — at PctFilledWithWater=0.5 — holds 2 of water.
        if abs(sd3['tcrack_depth'] - 4.0) > 1e-6:
            problems.append(f"tcrack_depth {sd3['tcrack_depth']}, expected 4")
        if abs(sd3['tcrack_water'] - 2.0) > 1e-6:
            problems.append(f"tcrack_water {sd3['tcrack_water']}, expected 2 "
                            f"(a DEPTH, = 50% of the crack depth, not a fraction)")

        # A crack with NO <TensionOption> is switched OFF — GeoStudio keeps its geometry
        # anyway. Analysis 4 has the very same crack line as analysis 3 and no option,
        # so it must import NO crack. Reading the points as a live crack put a 2 m
        # water-filled crack into models that had none, and cost 2% of FS in c'=0 soil.
        if sd4['tcrack_depth'] != 0.0 or sd4['tcrack_water'] != 0.0:
            problems.append(f"a tension crack with no <TensionOption> is switched OFF, "
                            f"but it imported at depth {sd4['tcrack_depth']}")

        # A <Surcharge> is a WEDGE OF FILL between the drawn line and the ground, and
        # <Pressure> is its unit weight — so the load varies with depth. Here a 20 kN/m3
        # fill runs from 1 deep at x=42 to 3 deep at x=58 over flat ground at y=30, so
        # the pressure must run 20 -> 60, NOT a constant 20.
        dl = sd3['dloads']
        if len(dl) != 1:
            problems.append(f"surcharge -> {len(dl)} dload blocks, expected 1")
        else:
            got = [(round(p['X'], 3), round(p['Normal'], 3)) for p in dl[0]]
            if got != [(42.0, 20.0), (58.0, 60.0)]:
                problems.append(
                    f"surcharge -> {got}, expected [(42.0, 20.0), (58.0, 60.0)]: "
                    f"<Pressure> is a UNIT WEIGHT and the load is the weight of the "
                    f"fill above the ground, not a uniform pressure")

        # The piezometric surface indexes the analysis's LOCAL <DataPoints>, not the
        # shared geometry <Points>. Read against the geometry it silently produced a
        # water table that doubled back on itself, and 10% of FS.
        if sd2['piezo_line'] != [(0.0, 8.0), (60.0, 24.0)]:
            problems.append(f"piezo line {sd2['piezo_line']}, expected "
                            f"[(0.0, 8.0), (60.0, 24.0)] — a piezometric surface indexes "
                            f"the analysis's LOCAL DataPoints, not the geometry points")

        # Reinforcement. GeoStudio quotes per-element values with an out-of-plane
        # spacing; xslope's lines are per unit width. Tensile 100 / spacing 2 = 50;
        # plate 60 / 2 = 30 at the end that is at the face; pullout 10 / 2 = 5 per unit
        # length, so the bond length that develops full capacity is 50 / 5 = 10.
        r5 = (sd5.get('reinforcement_lines') or [None])[0]
        if r5 is None:
            problems.append("reinforcement line was not imported")
        else:
            want = {'t_max': 50.0, 'lp1': 10.0, 'lp2': 10.0, 'tend1': 30.0, 'tend2': 0.0,
                    'spacing': 1.0, 'type': 'nail', 'dir': 'axial', 'appl': 'active'}
            bad = {k: r5.get(k) for k, v in want.items()
                   if not (abs(r5.get(k) - v) < 1e-9 if isinstance(v, float)
                           else r5.get(k) == v)}
            if bad:
                problems.append(f"reinforcement mapped to {bad}, expected "
                                f"{ {k: want[k] for k in bad} }")

        # Material colours come across, so an imported model still looks like the user's.
        if gsz_style(gsz, 1) != {'materials': {'0': {'color': '#d3c989'}}}:
            problems.append(f"material colour not imported: {gsz_style(gsz, 1)}")

        # FAIL LOUD: an element we do not recognise must be reported, never ignored.
        if not any('SomeFutureGeoStudioFeature' in c for c in cav3):
            problems.append("an unrecognised GeoStudio element was silently ignored")

        # Per-analysis material assignment: same geometry, different soil.
        m1 = sd1['materials'][0]
        m2 = sd2['materials'][0]
        if (m1['name'], m1['c'], m1['phi'], m1['gamma']) != ('strong', 25.0, 30.0, 20.0):
            problems.append(f"analysis 1 material {m1['name']}/{m1['c']}/{m1['phi']}")
        if (m2['name'], m2['c'], m2['phi'], m2['gamma']) != ('weak', 5.0, 20.0, 18.0):
            problems.append(f"analysis 2 material {m2['name']}/{m2['c']}/{m2['phi']}")

        # Geometry: the region's 6 points become one zone.
        if len(sd1['polygons']) != 1:
            problems.append(f"{len(sd1['polygons'])} polygons, expected 1")
        elif len(sd1['polygons'][0]['polygon'].exterior.coords) != 7:   # ring closes
            problems.append("zone ring did not round-trip")
        if not sd1['ground_surface'] or sd1['ground_surface'].is_empty:
            problems.append("no ground surface derived")

        # PER-ANALYSIS GEOMETRY. This file carries TWO geometries that reuse the same
        # point IDs (1-6) and region ID (1); each analysis names its own by <GeometryId>.
        # Analyses 1-7 bind geometry 1 (crest at y=30); analysis 8 binds geometry 2 (crest
        # at y=45). If GeometryId is ignored and the point tables are merged, both resolve
        # to whichever geometry was parsed LAST (geometry 2), so analysis 1 silently
        # imports the taller slope — distinct heights, distinct factors of safety, no word
        # said. This was the Borges & Cardoso Case 3 defect (7 m vs 8.75 m embankment).
        def _crest_y(sd):
            return max(xy[1] for pg in sd['polygons']
                       for xy in pg['polygon'].exterior.coords)
        if abs(_crest_y(sd1) - 30.0) > 1e-6:
            problems.append(f"analysis 1 resolved to crest y={_crest_y(sd1)}, expected 30 "
                            f"— its <GeometryId> was ignored, so it took the last geometry "
                            f"parsed instead of its own")
        if abs(_crest_y(sd8) - 45.0) > 1e-6:
            problems.append(f"analysis 8 (GeometryId 2) resolved to crest y={_crest_y(sd8)}, "
                            f"expected 45 — its own geometry, which reuses geometry 1's "
                            f"point IDs, was not selected")

        # Pore pressure: analysis 1 dry, analysis 2 on the piezo surface (points 7-8).
        if sd1['piezo_line'] or sd1['materials'][0]['u'] != 'none':
            problems.append("analysis 1 should be dry")
        if sd2['piezo_line'] != [(0.0, 8.0), (60.0, 24.0)]:
            problems.append(f"piezo line {sd2['piezo_line']}")
        if sd2['materials'][0]['u'] != 'piezo':
            problems.append(f"analysis 2 u={sd2['materials'][0]['u']}, expected piezo")

        # Seismic coefficient rides on the analysis.
        if sd1['k_seismic'] != 0.0 or sd2['k_seismic'] != 0.15:
            problems.append(f"k_seismic {sd1['k_seismic']}/{sd2['k_seismic']}")

        # Unsolved file: no surface invented, and the gap is reported.
        if sd1['circles'] or sd1['non_circ']:
            problems.append("invented a failure surface for an unsolved file")
        if not any('no failure surface' in c for c in cav1):
            problems.append("missing failure surface not reported as a caveat")

        # A SEEP/W field does not fit in a spreadsheet, so import_gsz writes it BESIDE
        # the .xlsx as _mesh.json + _seep.csv, and load_slope_data reads it back. If it
        # did not, the .xlsx would say every material takes its pore pressure from a
        # seepage solution and there would be none -- a dry model wearing a wet one's
        # clothes. Prove the whole product path: .gsz -> .xlsx (+ sidecars) -> reload.
        from xslope.geostudio import import_gsz
        from xslope.fileio import (load_slope_data, default_template_path,
                                   save_slope_data_to_xlsx)
        xlsx = os.path.join(td, "seep_roundtrip.xlsx")
        rcav = import_gsz(path, default_template_path(), xlsx, analysis_id=7)
        side = {os.path.basename(p) for p in os.listdir(td)}
        missing_side = {"seep_roundtrip_mesh.json", "seep_roundtrip_seep.csv"} - side
        if missing_side:
            problems.append(f"import_gsz did not write the SEEP/W sidecar(s) "
                            f"{sorted(missing_side)} — the .xlsx alone cannot carry a "
                            f"pore-pressure field, so the model would reload dry")
        try:
            reload = load_slope_data(xlsx)
        except Exception as e:
            # load_slope_data waives the failure-surface requirement only when a mesh is
            # present; if the sidecar was not written, it raises here instead. Report that
            # as the sidecar loss it is, rather than letting it read as an unrelated crash.
            reload = None
            problems.append(f"a SEEP/W field imported to .xlsx would not reload ({e}) — "
                            f"most likely its mesh/seep sidecar was not written beside it")
        if reload is not None:
            if reload.get("mesh") is None or reload.get("seep_u") is None:
                problems.append("a SEEP/W field imported to .xlsx did not reload — the "
                                "mesh or the pore pressure was lost between save and open")
            elif len(reload["seep_u"]) != n_nodes:
                problems.append(f"reloaded seep_u has {len(reload['seep_u'])} values, "
                                f"expected {n_nodes}")
            elif [m['u'] for m in reload['materials']] != ['seep']:
                problems.append(f"reloaded material u={[m['u'] for m in reload['materials']]}"
                                f", expected ['seep'] — the field reloaded but nothing reads it")
            else:
                # export_seep_u writes head+u only. import_seep_solution must load u as
                # real and the flow-net columns as NaN (not zeros), so a flow-net plot
                # fails visibly rather than drawing a field that was never solved.
                from xslope.seep import build_seep_data, import_seep_solution
                import numpy as _np
                sday = build_seep_data(reload["mesh"], reload, seep_bc=1)
                sol = import_seep_solution(sday, os.path.join(td, "seep_roundtrip_seep.csv"))
                if _np.any(_np.isnan(sol["u"])):
                    problems.append("imported seepage field lost its pore pressure to NaN")
                if not _np.all(_np.isnan(sol["phi"])):
                    problems.append("a field with no flow net came back with fabricated "
                                    "phi instead of NaN — a flow-net plot would draw a lie")

        # The mode has to reach the FILE, not just the dict. main!D23 is where an
        # imported model states who supplies the weight of its standing water, and a
        # mode that does not survive the write is a reservoir lost (auto -> a file
        # that reloads manual with an empty dloads sheet) or doubled (manual -> auto
        # over a synthesised block).
        if reload is not None and reload.get('water_loads') != 'manual':
            problems.append(
                f"the SEEP/W-coupled model reloaded as "
                f"{reload.get('water_loads')!r}; it was written as manual, and any "
                f"other answer either drops or doubles the reservoir it carries")
        # The piezo-surface analysis of this fixture is unsolved, so it imports with no
        # failure surface and load_slope_data would refuse the file. Give it one, as a
        # user must, and write the same model the importer produced.
        xlsx2 = os.path.join(td, "piezo_roundtrip.xlsx")
        save_slope_data_to_xlsx(
            dict(sd2, circular=True,
                 circles=[{'Xo': 25.0, 'Yo': 40.0, 'R': 35.0, 'Depth': 5.0}]),
            xlsx2, template=default_template_path())
        try:
            reload2 = load_slope_data(xlsx2)
        except Exception as e:
            reload2 = None
            problems.append(f"a piezo-surface import would not reload ({e})")
        if reload2 is not None:
            if reload2.get('water_loads') != 'auto':
                problems.append(f"a piezo-surface import reloaded as "
                                f"{reload2.get('water_loads')!r}, expected 'auto'")
            if reload2.get('dloads'):
                problems.append("a piezo-surface import reloaded carrying a "
                                "distributed load; its dloads sheet must be empty")
            if len(derived_blocks(with_water_loads(reload2), 1)) != 1:
                problems.append("the reloaded piezo-surface model derives no water "
                                "load — the reservoir did not survive the file")

        # A pre-v22 template has no D23, so an automatic-water model cannot be
        # written into one: the file would come back manual with an empty dloads
        # sheet and no water at all. Refused, and the refusal names the cell.
        v21 = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           'docs', 'inputs', 'input_template_v21.xlsx')
        if os.path.exists(v21):
            old_out = os.path.join(td, "old_template.xlsx")
            try:
                save_slope_data_to_xlsx(sd2, old_out, template=v21)
            except ValueError as e:
                if 'D23' not in str(e):
                    problems.append(f"saving an automatic-water model into a v21 "
                                    f"template was refused without naming the cell: "
                                    f"{str(e)[:90]}")
                if os.path.exists(old_out):
                    problems.append("the refused save left a half-written file behind")
            else:
                problems.append(
                    "an automatic-water model saved into a v21 template without "
                    "complaint — that file reloads as manual with no water load, "
                    "and the reservoir is silently gone")

        # Export round-trip: slope_data -> .gsz -> slope_data must preserve the
        # geometry, the materials, and the piezo line.
        from xslope.geostudio import export_gsz
        out = os.path.join(td, "exported.gsz")
        export_gsz(sd2, out, analysis_name="rt")

        # SCHEMA CONFORMANCE — the check the round-trip cannot make. Every tag path we
        # write must be one GeoStudio itself writes; otherwise the file round-trips
        # through our own reader perfectly and GeoStudio still refuses it.
        import zipfile as _zip, xml.etree.ElementTree as _ET
        zf = _zip.ZipFile(out)
        root = _ET.fromstring(zf.read(zf.namelist()[0]))
        emitted = set()

        def _walk(e, p=""):
            q = f"{p}/{e.tag}"
            emitted.add(q)
            for a in e.attrib:
                emitted.add(f"{q}@{a}")
            for c in e:
                _walk(c, q)
        _walk(root)
        unknown = sorted(emitted - GSZ_SCHEMA_PATHS)
        if unknown:
            problems.append("export writes tag(s) GeoStudio does not use: "
                            + ", ".join(unknown[:3]))

        # THE OTHER DIRECTION, and the one that actually bites. The check above can only
        # catch a tag we write that GeoStudio does not; it is structurally incapable of
        # catching one we FAIL to write. Everything GeoStudio needs and we omitted got
        # through it — and the file then opens, draws, and looks fine while quietly
        # missing its physics. GeoStudio was the only thing that could tell us.
        #
        # So: every path GeoStudio writes in EVERY vendor file must be one we write too,
        # unless it is on the list of things we consciously do not write (a search
        # definition, solved results, the window scroll state, SEEP/W). Anything else
        # missing is a defect. This is what would have caught <ComputedPhysics> — without
        # it GeoStudio does not know the analysis solves slope stability, so the materials
        # arrive named and coloured with their strengths unreachable and nothing said.
        missing = sorted(GSZ_REQUIRED_PATHS - emitted)
        if missing:
            problems.append("export OMITS path(s) GeoStudio always writes: "
                            + ", ".join(missing[:3]))
        # And the structural facts a wrong-schema writer gets wrong:
        if not root.findall("./Geometries/Geometry/Lines/Line"):
            problems.append("export wrote no <Lines> — GeoStudio draws regions from them")
        if not root.findall("./StabilityItems/StabilityItem/Entry/"
                            "PiezometricSurfaces/PiezometricSurface"):
            problems.append("export put the piezometric surface outside the StabilityItem")
        if root.find("./Coordinates/EngCoords") is None:
            problems.append("export declared no unit system / view extent")
        # GeoStudio wants a piezometric surface's point IDs to ascend along the
        # polyline. Reusing a coincident region vertex hands it a low ID out of
        # sequence and it reports the surface as corrupt — so the surface gets its
        # own points, in order.
        pz = [int(d.text) for d in root.findall(
            "./StabilityItems/StabilityItem/Entry/PiezometricSurfaces/"
            "PiezometricSurface/DataPoints/DataPoint")]
        if pz != sorted(pz):
            problems.append(f"piezo surface point IDs not ascending: {pz}")
        # Materials must be visually distinguishable, not all on GeoStudio's default.
        cols = [m.findtext("Color") for m in root.findall("./Materials/Material")]
        if len(cols) > 1 and len(set(cols)) != len(cols):
            problems.append("materials exported without distinct colours")

        back, _ = gsz_to_slope_data(read_gsz(out), 1)
        b = back['materials'][0]
        if (b['name'], b['c'], b['phi'], b['gamma']) != ('weak', 5.0, 20.0, 18.0):
            problems.append(f"export round-trip material {b['name']}/{b['c']}/{b['phi']}")
        a0 = round(sd2['polygons'][0]['polygon'].area, 6)
        a1 = round(back['polygons'][0]['polygon'].area, 6)
        if a0 != a1:
            problems.append(f"export round-trip zone area {a1} vs {a0}")
        if back['piezo_line'] != sd2['piezo_line']:
            problems.append(f"export round-trip piezo {back['piezo_line']}")
        if back['k_seismic'] != sd2['k_seismic']:
            problems.append(f"export round-trip k_seismic {back['k_seismic']}")

        # v16/v17 material features across the export round-trip. The one SLOPE/W can
        # carry is 'elastic', as its impenetrable Bedrock model: an elastic material must
        # survive export -> import as elastic. t_cut, phi_b and s_cap have no SLOPE/W
        # material encoding — they must be REPORTED and dropped, never silently written
        # (an invented tag would fail the schema conformance check above) and never
        # corrupt the file. This is the guard the mission asks for: elastic + phi_b
        # materials survive the trip with the mapped semantics.
        from shapely.geometry import Polygon as _Poly
        from xslope.geostudio import _blank_material as _gm
        from xslope.fileio import build_ground_surface_from_polygons as _bg

        def _mk(name, **kw):
            m = _gm(name); m.update(kw); return m
        feat_polys = [
            {"polygon": _Poly([(0, 0), (40, 0), (40, 6), (0, 6)]), "mat_id": 0},
            {"polygon": _Poly([(0, 6), (40, 6), (40, 16), (20, 16), (0, 10)]), "mat_id": 1},
        ]
        _gs, _dom = _bg(feat_polys)
        feat_sd = dict(sd2)
        feat_sd.update(
            polygons=feat_polys, domain_polygon=_dom, ground_surface=_gs, piezo_line=[],
            circles=[], non_circ=[], circular=False, dloads=[],
            materials=[_mk("bedrock", option="elastic", gamma=22.0, sigma_c=4.0),
                       _mk("silt", option="mc", gamma=18.0, c=6.0, phi=27.0,
                           phi_b=15.0, t_cut=8.0, s_cap=50.0,
                           sigma_gamma=0.8, sigma_c=1.2, sigma_phi=2.5)])
        feat_out = os.path.join(td, "features.gsz")
        fcav = export_gsz(feat_sd, feat_out, analysis_name="features")
        # SlopeModel written: the elastic material is Bedrock, the soil is MohrCoulomb.
        froot = _ET.fromstring(_zip.ZipFile(feat_out).read(
            _zip.ZipFile(feat_out).namelist()[0]))
        models = [e.text for e in froot.findall("./Materials/Material/SlopeModel")]
        if models != ["Bedrock", "MohrCoulomb"]:
            problems.append(f"elastic did not export as Bedrock: SlopeModel={models}")
        # No invented tags (t_cut/phi_b are dropped, not written).
        femit = set()
        def _walk2(e, p=""):
            q = f"{p}/{e.tag}"; femit.add(q)
            for a in e.attrib:
                femit.add(f"{q}@{a}")
            for c in e:
                _walk2(c, q)
        _walk2(froot)
        if sorted(femit - GSZ_SCHEMA_PATHS):
            problems.append("feature export wrote tag(s) GeoStudio does not use: "
                            + ", ".join(sorted(femit - GSZ_SCHEMA_PATHS)[:3]))
        # The dropped features must be reported, not silent.
        if not any("elastic" in c and "Bedrock" in c for c in fcav):
            problems.append("elastic -> Bedrock export was not reported as a caveat")
        if not any("t_cut" in c for c in fcav):
            problems.append("dropped t_cut was not reported on export")
        if not any("phi_b" in c or "suction" in c.lower() for c in fcav):
            problems.append("dropped phi_b/s_cap was not reported on export")
        # Re-import: the elastic material comes back elastic (and reported so).
        fback, fbcav = gsz_to_slope_data(read_gsz(feat_out), 1)
        fmats = {m["name"]: m for m in fback["materials"]}
        if fmats.get("bedrock", {}).get("option") != "elastic":
            problems.append("Bedrock did not re-import as elastic across the round-trip")
        if round(fmats.get("bedrock", {}).get("gamma", 0), 3) != 22.0:
            problems.append("the elastic material's unit weight was lost")
        if fmats.get("silt", {}).get("option") != "mc":
            problems.append("the ordinary soil did not survive the feature round-trip")
        if not any("Bedrock" in c and "elastic" in c for c in fbcav):
            problems.append("Bedrock -> elastic re-import was not reported")
        # Standard deviations cross both ways, as SLOPE/W's own probabilistic modifiers.
        # The one on the elastic material does NOT: it is written as Bedrock, which has
        # no strength to vary, so it is reported rather than written into a dead slot.
        fsig = (fmats.get("silt", {}).get("sigma_gamma"),
                fmats.get("silt", {}).get("sigma_c"),
                fmats.get("silt", {}).get("sigma_phi"))
        if fsig != (0.8, 1.2, 2.5):
            problems.append(f"standard deviations did not survive the export round-trip: "
                            f"got {fsig}, expected (0.8, 1.2, 2.5)")
        if fmats.get("bedrock", {}).get("sigma_c"):
            problems.append("a sigma on an elastic material was written as Bedrock "
                            "strength uncertainty, which cannot exist")
        if not any("standard deviation" in c and "not written" in c for c in fcav):
            problems.append("the dropped sigma on the elastic material was not reported")

        # --- a coupled model, through the real file path ----------------------------
        # A finite-element seepage field used to be dropped on export: the .gsz opened
        # DRY, and only a caveat said so. It now crosses as SLOPE/W's own spatial
        # pore-pressure input -- <Fn3D> points of pressure head, selected by the
        # analysis's 3DFunction option. Exercised on a real seep-fed corpus model read
        # off disk with its sidecars, not a hand-built dict, because the field only
        # exists once load_slope_data has found _mesh.json and _seep.csv beside the
        # .xlsx: a fixture would prove the writer works on data the product never
        # produces. The model is imperial, so it also pins the u -> head conversion to
        # the file's OWN unit weight of water rather than a metric constant.
        seep_xlsx = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 "docs", "lem", "files", "xslope_gsat_seep.xlsx")
        if not os.path.exists(seep_xlsx):
            problems.append(f"missing coupled-export fixture {seep_xlsx}")
        else:
            import numpy as _np
            sd_seep = load_slope_data(seep_xlsx)
            gw = sd_seep["gamma_water"]
            nodes, ufield = sd_seep["mesh"]["nodes"], sd_seep["seep_u"]
            seep_out = os.path.join(td, "coupled.gsz")
            scav = export_gsz(sd_seep, seep_out, analysis_name="coupled")
            sroot = _ET.fromstring(_zip.ZipFile(seep_out).read(
                _zip.ZipFile(seep_out).namelist()[0]))

            fpts = sroot.findall("./Functions/Func3Ds/Fn3D/Points/Points_")
            if len(fpts) != len(nodes):
                problems.append(f"coupled export wrote {len(fpts)} pore-pressure points "
                                f"for {len(nodes)} seepage nodes — the field was "
                                f"truncated, and a thinned field is a different model")
            else:
                # Values, not just count: the function is a PRESSURE HEAD function in
                # every vendor file that defines one, so Z is u/gamma_w. Writing u
                # itself would be silently wrong by a factor of 62.4 here.
                z = _np.array([float(p.get("Z")) for p in fpts])
                x = _np.array([float(p.get("X")) for p in fpts])
                if not _np.allclose(z, _np.asarray(ufield, dtype=float) / gw, atol=1e-6,
                                    rtol=1e-9):
                    problems.append("coupled export wrote pore PRESSURE where GeoStudio's "
                                    "spatial water function expects pressure HEAD")
                if not _np.allclose(x, nodes[:, 0], atol=1e-6):
                    problems.append("coupled export's pore-pressure points are not on the "
                                    "seepage mesh nodes")
            if sroot.findtext("./Functions/Func3Ds/Fn3D/OutputType") != "PressureHead":
                problems.append("the spatial water function was not declared PressureHead")
            # The analysis must actually SELECT it. Without the option and the GGID the
            # points sit in the file unused and the model still solves — dry.
            if sroot.findtext("./WaterItems/WaterItem/Entry/ResultInputInfo/Option") \
                    != "3DFunction":
                problems.append("coupled export wrote the pore-pressure points but did "
                                "not point the analysis at them — the model opens dry")
            if sroot.findtext("./WaterItems/WaterItem/Entry/ResultInputInfo/DataGGID") \
                    != "3DFns-1":
                problems.append("coupled export's water source has no/!= 3DFns-1 GGID")
            # Schema conformance again, on the branch the other exports never take.
            semit = set()

            def _walk3(e, p=""):
                q = f"{p}/{e.tag}"; semit.add(q)
                for a in e.attrib:
                    semit.add(f"{q}@{a}")
                for c in e:
                    _walk3(c, q)
            _walk3(sroot)
            if sorted(semit - GSZ_SCHEMA_PATHS):
                problems.append("coupled export writes tag(s) GeoStudio does not use: "
                                + ", ".join(sorted(semit - GSZ_SCHEMA_PATHS)[:3]))
            if sorted(GSZ_REQUIRED_PATHS - semit):
                problems.append("coupled export OMITS path(s) GeoStudio always writes: "
                                + ", ".join(sorted(GSZ_REQUIRED_PATHS - semit)[:3]))
            # The caveat has to SAY what was done, with the count — the point of the
            # exercise is that the user knows the pressures crossed and the model did not.
            if not any(str(len(nodes)) in c and "SEEP/W" in c for c in scav):
                problems.append("the coupled export did not report how many pore-pressure "
                                "points it wrote, or that no SEEP/W analysis crossed")
            # And it must not still claim the old behaviour.
            if any("NO pore pressure" in c for c in scav):
                problems.append("the coupled export still reports the model as dry")
            # It must round-trip as a file: geometry and materials survive alongside the
            # new block. (Our reader does not read a spatial function back — it reports
            # it, loudly — so this proves the file, not the field.)
            sback, sbcav = gsz_to_slope_data(read_gsz(seep_out), 1)
            if len(sback["polygons"]) != len(sd_seep["polygons"]):
                problems.append("the coupled export lost zones")
            if not any("3DFunction" in c for c in sbcav):
                problems.append("re-importing the coupled export did not report that its "
                                "spatial pore-pressure function cannot be read back")

    if problems:
        return None, "GeoStudio import: " + "; ".join(problems[:5])
    return 0.0, None


# A minimal Slide2 model in the sectioned ASCII grammar Slide2 writes. Authored here
# rather than shipped because Rocscience's own tutorial files are their copyrighted
# material and are not in this repository (the same reason the .gsz fixture above is
# synthesised). Two materials over a two-layer slope, a water table the upper layer
# draws from, and one specified circle — enough to exercise the cell-union geometry,
# the shared-strength/per-scenario-flag material merge, the 'hu' water mapping, and a
# round-trip through the .xlsx writer and load_slope_data.
_SYNTH_SLI = """
model description:
  version: 9.027
  methods: 6
  units: metric
  seismic: 0
  seismicv: 0
  water: hu
  direction: right to left
  gammaw: 9.81
  nummaterials: 2
  transient: no
  design_selection: 0

material types:
  soil1 = type: 0 water: 1 wtable: 1 c: 5 phi: 30 uw: 19 hutype: 0 withru: 0 phib: 15
  soil2 = type: 5 water: 1 wtable: 1 c: 25 phi: 32 uw: 20 hutype: 0 withru: 0

vertices:
  1 x: 0  y: 0
  2 x: 40  y: 0
  3 x: 40  y: 6
  4 x: 0  y: 6
  5 x: 40  y: 18
  6 x: 20  y: 18
  7 x: 0  y: 9
  8 x: 40  y: 9

cells:
  1  vertices: [1,2,3] material: soil2
  2  vertices: [1,3,4] material: soil2
  3  vertices: [4,3,5] material: soil1
  4  vertices: [4,5,6] material: soil1

water table:
  from_boreholes: 0  vertices: [7,8]

exterior:
  1  vertices: [1,2,5,6,4]

centers:
  1 x: 12 y: 30 r: 26 unique_id: {AAAAAAAA-0000-0000-0000-000000000001}

grids:

surfaces:

anchors:
"""


def _write_synthetic_slim(path):
    """Author a minimal Slide2 .slim — a ZIP of one .sli plus empty sidecars."""
    import zipfile
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("synthetic.sli", _SYNTH_SLI)
        zf.writestr("synthetic.slv", "")     # view options — ignored by the reader
        zf.writestr("synthetic.sltm", "")    # theme — ignored


def run_slide2_import_test(test):
    """Import a synthetic Slide2 .slim and check the model that comes out.

    Guards the schema facts the importer depends on, each silent when wrong:
      - triangular cells are unioned per material into one zone each, so two
        materials yield two polygons over the shared vertex table;
      - c/phi/uw come across, and only Mohr-Coulomb (type 0) maps cleanly;
      - a 'hu' water table becomes a piezo line, and only materials with wtable = 1
        draw pore pressure from it;
      - the water table is carried as a water DEFINITION rather than converted to a
        load: main!D23 = auto, an empty dloads sheet, and the engine deriving the
        same block the import used to synthesise;
      - a specified circle is imported (a search would not be);
      - the whole model round-trips through the .xlsx writer and load_slope_data.
    """
    import tempfile
    from xslope.slide2 import (read_slim, read_slide2, list_scenarios,
                               slide2_to_slope_data, import_slmd)
    from xslope.fileio import (save_slope_data_to_xlsx, load_slope_data,
                               default_template_path)

    problems = []
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "synthetic.slim")
        _write_synthetic_slim(path)

        d = read_slim(path)
        scens = list_scenarios(d)
        if len(scens) != 1:
            return None, f"Slide2 import: {len(scens)} scenarios, expected 1"

        sd, caveats = slide2_to_slope_data(d)

        # Two materials, two zones. soil1 is Mohr-Coulomb (with a phi_b); soil2 is Slide2's
        # Infinite Strength type, which must import as an ELASTIC (impenetrable) material.
        mats = sd["materials"]
        if len(mats) != 2:
            problems.append(f"{len(mats)} materials, expected 2")
        else:
            if (round(mats[0]["c"], 3), round(mats[0]["phi"], 3),
                    round(mats[0]["gamma"], 3)) != (5.0, 30.0, 19.0):
                problems.append(f"soil1 came across as c={mats[0]['c']} phi="
                                f"{mats[0]['phi']} gamma={mats[0]['gamma']}, "
                                f"expected 5/30/19")
            if mats[0].get("option") != "mc":
                problems.append(f"soil1 option={mats[0].get('option')!r}, expected 'mc'")
            # v17: a Slide2 unsaturated friction angle maps to phi_b (encoding unverified).
            if mats[0].get("phi_b") != 15.0:
                problems.append(f"soil1 phi_b={mats[0].get('phi_b')!r}, expected 15.0")
            # soil2 type 5 -> elastic, unit weight kept, no invented strength.
            if mats[1].get("option") != "elastic":
                problems.append(f"soil2 option={mats[1].get('option')!r}, expected "
                                f"'elastic' (Slide2 Infinite Strength did not map)")
            if round(mats[1].get("gamma", 0), 3) != 20.0:
                problems.append(f"soil2 gamma={mats[1].get('gamma')}, expected 20 "
                                f"(the elastic material lost its unit weight)")
        # The mappings must be reported, and the elastic material must NOT be flagged as
        # zero-strength.
        if not any("Infinite Strength" in c and "soil2" in c for c in caveats):
            problems.append("Slide2 Infinite Strength -> elastic was not reported")
        if not any("unsaturated" in c.lower() and "soil1" in c for c in caveats):
            problems.append("the Slide2 phi_b import was not reported")
        if any("NO STRENGTH" in c and "soil2" in c for c in caveats):
            problems.append("the elastic material was wrongly flagged as zero-strength")
        if len(sd["polygons"]) != 2:
            problems.append(f"{len(sd['polygons'])} polygons, expected 2 (the "
                            f"triangular cells did not union per material)")

        # 'hu' water table -> piezo line; both materials have wtable = 1.
        if len(sd["piezo_line"]) != 2:
            problems.append(f"piezo_line has {len(sd['piezo_line'])} points, "
                            f"expected 2 (the water table did not import)")
        if [m["u"] for m in mats] != ["piezo", "piezo"]:
            problems.append(f"materials take u={[m['u'] for m in mats]}, expected "
                            f"['piezo', 'piezo'] — a 'hu' water table is a piezo line")

        # Ponded water: the water table (y=9) stands above the left ground (y=6), so
        # the reservoir weight must reach the statics — but NOT as a row on the dloads
        # sheet. Slide2 stores no ponded-water object and carries the weight from the
        # water table itself; under template v22 xslope does the same, so the import
        # carries the table, sets main!D23 to auto, and the engine derives the load.
        # Peak intensity is the closed-form hydrostatic gamma_w * max_depth,
        # 9.81 * (9 - 6) = 29.43, tapering to zero where the water meets the ground.
        from xslope.water import (with_water_loads, derived_blocks,
                                  ponded_water_dload)
        if sd.get("water_loads") != "auto":
            problems.append(f"the model imported as {sd.get('water_loads')!r}; a "
                            f"Slide2 water table is a water definition, so the "
                            f"import must set 'auto' and let the engine measure it")
        if sd["dloads"]:
            problems.append(
                f"the import wrote {len(sd['dloads'])} distributed load(s); under "
                f"automatic water loads the dloads sheet carries non-water loads "
                f"only, and Slide2's own external loads are not imported at all")
        auto_blocks = derived_blocks(with_water_loads(sd), 1)
        if len(auto_blocks) != 1:
            problems.append(f"the engine derived {len(auto_blocks)} water block(s) "
                            f"from the imported water table, expected 1 — the "
                            f"reservoir weight is lost from the statics")
        else:
            peak = max(pt["Normal"] for pt in auto_blocks[0])
            if round(peak, 2) != round(9.81 * 3.0, 2):
                problems.append(f"derived ponded-water peak Normal {peak:.2f}, "
                                f"expected {9.81 * 3.0:.2f} (gamma_w * max_depth)")
        # EQUIVALENCE: what the engine derives is what the import used to synthesise,
        # vertex for vertex. The import stopped writing the block on that basis.
        synth = ponded_water_dload(sd["ground_surface"], sd["piezo_line"],
                                   sd["gamma_water"])
        if [[(round(p["X"], 12), round(p["Y"], 12), round(p["Normal"], 12))
             for p in b] for b in synth] != \
                [[(round(p["X"], 12), round(p["Y"], 12), round(p["Normal"], 12))
                  for p in b] for b in auto_blocks]:
            problems.append("the derived water load does not reproduce the block the "
                            "import used to synthesise from the same water table")
        if not any("ponded water" in c and "derive automatically" in c
                   for c in caveats):
            problems.append("the import did not report that the ponded water now "
                            "derives automatically from the imported water table")

        # A specified circle is imported; a search would not be.
        if len(sd["circles"]) != 1:
            problems.append(f"{len(sd['circles'])} circles, expected 1")
        elif round(sd["circles"][0]["R"], 3) != 26.0:
            problems.append(f"circle R={sd['circles'][0]['R']}, expected 26")

        # A search-only model must report its missing surface, never invent one.
        d2 = read_slim(path)
        d2["scenarios"][0]["sections"]["centers"] = []
        sd_ns, cav_ns = slide2_to_slope_data(d2)
        if sd_ns["circles"] or sd_ns["non_circ"]:
            problems.append("a search-only scenario imported a surface it should not")
        if not any("no failure surface" in c for c in cav_ns):
            problems.append("a search-only scenario did not report its missing surface")

        # read_slide2 dispatches on extension and agrees with read_slim.
        if len(list_scenarios(read_slide2(path))) != 1:
            problems.append("read_slide2 did not dispatch a .slim correctly")

        # Round-trip: the model must survive the .xlsx writer AND reload — solving the
        # raw dict is not enough (load_slope_data is the real consumer).
        xlsx = os.path.join(td, "synthetic.xlsx")
        rcav = import_slmd(path, default_template_path(), xlsx)
        if not os.path.exists(xlsx):
            problems.append("import_slmd did not write the .xlsx")
        else:
            reloaded = load_slope_data(xlsx)
            if len(reloaded["materials"]) != 2:
                problems.append(f"reloaded model has {len(reloaded['materials'])} "
                                f"materials, expected 2")
            if not reloaded.get("circular") or len(reloaded["circles"]) != 1:
                problems.append("the circle did not survive the .xlsx round-trip")
            if len(reloaded.get("piezo_line") or []) != 2:
                problems.append("the piezo line did not survive the .xlsx round-trip")
            # The reservoir crosses the file as the water table plus the mode cell,
            # not as a load row. Both halves have to arrive: a reloaded model that
            # says 'manual' with an empty dloads sheet has lost the water entirely.
            if reloaded.get("water_loads") != "auto":
                problems.append(f"the reloaded model says water_loads="
                                f"{reloaded.get('water_loads')!r}, expected 'auto' — "
                                f"main!D23 did not survive the .xlsx round-trip, and "
                                f"the reservoir is gone with it")
            if reloaded.get("dloads"):
                problems.append("the reloaded model carries a distributed load; an "
                                "automatic-water import writes none")
            if len(derived_blocks(with_water_loads(reloaded), 1)) != 1:
                problems.append("the reloaded model derives no water load — the "
                                "reservoir did not survive the .xlsx round-trip")
        if not isinstance(rcav, list):
            problems.append("import_slmd did not return a caveat list")

    if problems:
        return None, "Slide2 import: " + "; ".join(problems[:5])
    return 0.0, None


# --------------------------------------------------------------------------------------
# Vendor-file water conformance: the same question, asked of real models
# --------------------------------------------------------------------------------------
#
# A synthetic fixture proves the rule; a vendor corpus proves the rule survives every
# shape a real modeller draws. Both sweeps below read a LOCAL folder of the vendor's
# own files -- Seequent's .gsz and Rocscience's Slide2 tutorials are their copyrighted
# material and are not in this repository -- and skip cleanly when it is absent.

#: Local folder of GeoStudio .gsz models (git-ignored, not redistributable).
GSZ_VENDOR_CORPUS = os.path.expanduser(
    '~/python_projects/vendor_files/gsz_corpus')
#: Local folder of Slide2 tutorial models (git-ignored, not redistributable).
SLIDE2_VENDOR_CORPUS = os.path.expanduser(
    '~/python_projects/vendor_files/rocscience_downloads/work')


def _block_shape(blocks):
    """A dload block list reduced to comparable numbers, vertex for vertex."""
    return [[(round(float(p["X"]), 12), round(float(p["Y"]), 12),
              round(float(p["Normal"]), 12)) for p in b] for b in blocks]


def run_gsz_water_corpus_test(test):
    """Every GeoStudio model in the local corpus, on the water-load handover.

    Two questions per analysis, and they are the two that decide whether the change
    of expression changed the model:

      - **the mode follows the vendor's own water statement.** A file that states a
        water SURFACE imports auto with an empty water sheet; a SEEP/W-coupled file,
        which states only a solved head field, keeps its recovered load and imports
        manual. Nothing imports auto while also carrying a synthesised block.
      - **the derivation reproduces the synthesis.** On every analysis that carries a
        piezometric surface above the ground, the blocks the engine derives are
        compared vertex for vertex against the blocks the import used to write --
        and where SLOPE/W saved trial circles, the models are sliced BOTH ways and
        the per-slice forces compared. A pool the derivation discards as coordinate
        round-off (POOL_MIN_DEPTH_FRAC) is allowed to be missing, and reported.

    Skips cleanly (0.0, None) when the local corpus is absent.
    """
    import glob
    from xslope.geostudio import read_gsz, gsz_to_slope_data, read_gsz_results
    from xslope.water import (ponded_water_dload, with_water_loads, derived_blocks,
                              block_resultant, POOL_MIN_DEPTH_FRAC)

    files = sorted(glob.glob(os.path.join(GSZ_VENDOR_CORPUS, '*.gsz')))
    if not files:
        return 0.0, None                  # no local vendor corpus: nothing to read

    problems, n_water, n_surf, n_dropped = [], 0, 0, 0
    for path in files:
        name = os.path.basename(path)
        try:
            gsz = read_gsz(path)
        except Exception as e:
            problems.append(f"{name}: unreadable ({str(e)[:50]})")
            continue
        for a in gsz["analyses"]:
            try:
                sd, _cav = gsz_to_slope_data(gsz, a["id"], critical_surface=False)
            except Exception:
                continue                  # import failures are the other tests' business
            label = f"{name} / {a['name'][:28]}"
            mode = sd.get("water_loads")
            if mode not in ("auto", "manual"):
                problems.append(f"{label}: imported water_loads={mode!r}")
                continue
            synth = ponded_water_dload(sd["ground_surface"], sd.get("piezo_line") or [],
                                       sd["gamma_water"])
            derived = derived_blocks(with_water_loads(sd), 1)
            if mode == "auto" and synth and sd["dloads"]:
                # The dloads sheet may legitimately carry a surcharge; what it must
                # never carry in auto mode is the water the engine also derives.
                from xslope.water import match_water_blocks
                dup = match_water_blocks(sd["dloads"], derived)
                if dup["pairs"]:
                    problems.append(f"{label}: imported auto AND wrote the water as a "
                                    f"dload — the reservoir is counted twice")
            if not synth:
                continue
            n_water += 1
            if mode != "auto":
                problems.append(f"{label}: states its water as a piezometric surface "
                                f"but imported {mode!r}")
                continue
            if _block_shape(synth) != _block_shape(derived):
                # The one legitimate difference: a block the derivation discards as
                # coordinate round-off rather than a pool. Anything else is a defect.
                extra = [b for b in synth if _block_shape([b])[0]
                         not in _block_shape(derived)]
                total = sum(block_resultant(b) for b in synth) or 1.0
                small = sum(block_resultant(b) for b in extra)
                if len(derived) + len(extra) != len(synth) or \
                        abs(small / total) > 1e-4:
                    problems.append(
                        f"{label}: the derived water load differs from the "
                        f"synthesised one by {100 * abs(small / total):.3g}% of the "
                        f"reservoir, which is more than a discarded hairline")
                else:
                    n_dropped += 1
            # Per-slice forces, on SLOPE/W's own trial circles where it saved any.
            rows = [r for r in read_gsz_results(gsz, a["name"], None) if r["circular"]]
            # Spread the sample over SLOPE/W's whole trial set. A search writes its
            # surfaces in grid order, so the first N of them are neighbours -- and on
            # several of these models none of that first corner rebuilds at all.
            if len(rows) > 40:
                rows = [rows[int(i * len(rows) / 40.0)] for i in range(40)]
            twin = dict(sd, water_loads="manual", dloads=synth)
            for row in rows:
                circle = {"Xo": row["xo"], "Yo": row["yo"], "R": row["r"],
                          "Depth": None, "Y": None, "Option": "Depth",
                          "Movement": "Left to Right"}
                ca = _water_force_columns(sd, circle, n=30)
                cm = _water_force_columns(twin, circle, n=30)
                if isinstance(ca, str) or isinstance(cm, str):
                    continue              # a circle that does not build on this model
                n_surf += 1
                if ca != cm:
                    bad = [c for c, x, y in zip(_WATER_FORCE_COLUMNS, ca, cm)
                           if x != y]
                    problems.append(
                        f"{label}: on SLOPE/W's circle at ({row['xo']:.1f}, "
                        f"{row['yo']:.1f}) r={row['r']:.1f} the per-slice "
                        f"{', '.join(bad)} differ between the derived water load and "
                        f"the synthesised one")
                    break

    # Coverage is asserted, not assumed: a sweep that silently stops finding water,
    # or stops rebuilding any of SLOPE/W's circles, passes while testing nothing.
    if n_water == 0:
        return None, ("GeoStudio water corpus: no analysis in the local corpus "
                      "carries water above the ground, so nothing was tested")
    if n_surf == 0:
        return None, ("GeoStudio water corpus: not one of SLOPE/W's own trial "
                      "circles rebuilt on a water-carrying model, so the per-slice "
                      "half of the comparison never ran")
    if problems:
        return None, (f"GeoStudio water corpus ({n_water} water analyses, {n_surf} "
                      f"surfaces): " + "; ".join(problems[:4]))
    return 0.0, None


def run_slide2_water_corpus_test(test):
    """Every Slide2 tutorial model in the local corpus, on the water-load handover.

    Slide2 has one water statement and one treatment: the water table is a water
    definition, so every scenario imports auto with an empty dloads sheet, and what
    the engine derives is what the import used to synthesise, vertex for vertex.
    Skips cleanly (0.0, None) when the local corpus is absent.
    """
    import glob
    from xslope.slide2 import read_slide2, list_scenarios, slide2_to_slope_data
    from xslope.water import ponded_water_dload, with_water_loads, derived_blocks

    files = sorted(glob.glob(os.path.join(SLIDE2_VENDOR_CORPUS, '**', '*.slmd'),
                             recursive=True))
    files += sorted(glob.glob(os.path.join(SLIDE2_VENDOR_CORPUS, '**', '*.slim'),
                              recursive=True))
    if not files:
        return 0.0, None

    problems, n_model, n_water = [], 0, 0
    for path in files:
        name = os.path.basename(path)
        try:
            d = read_slide2(path)
            scenarios = list_scenarios(d)
        except Exception:
            continue
        for sc in scenarios:
            sname = sc["name"] if isinstance(sc, dict) else sc
            try:
                sd, _cav = slide2_to_slope_data(d, sname)
            except Exception:
                continue                  # unimportable models are another test's job
            n_model += 1
            label = f"{name[:34]} / {str(sname)[:16]}"
            if sd.get("water_loads") != "auto":
                problems.append(f"{label}: imported "
                                f"{sd.get('water_loads')!r}, expected 'auto'")
            if sd["dloads"]:
                problems.append(f"{label}: imported {len(sd['dloads'])} distributed "
                                f"load(s); a Slide2 import writes none")
            synth = ponded_water_dload(sd["ground_surface"],
                                       sd.get("piezo_line") or [], sd["gamma_water"])
            if not synth:
                continue
            n_water += 1
            if _block_shape(synth) != _block_shape(derived_blocks(
                    with_water_loads(sd), 1)):
                problems.append(f"{label}: the derived water load does not reproduce "
                                f"the block the import used to synthesise")

    if n_model == 0:
        return None, "Slide2 water corpus: no model in the local corpus imported"
    if n_water == 0:
        return None, ("Slide2 water corpus: no scenario carries water above the "
                      "ground, so the derivation was never exercised")
    if problems:
        return None, (f"Slide2 water corpus ({n_model} scenarios, {n_water} with "
                      f"water): " + "; ".join(problems[:4]))
    return 0.0, None


# --------------------------------------------------------------------------------------
# RS2 (.fez) import test
# --------------------------------------------------------------------------------------

# A hand-authored RS2 model, synthesised so the test carries NO Rocscience file. The
# real .fea (megabytes of finite-element mesh) is reduced to just the sections the
# importer reads: the model description with its SSR settings and the version-on-the-
# next-line dialect, two Mohr-Coulomb materials, a two-zone geometry (an external
# boundary cut by one material boundary), the coarse material-mesh seed triangles that
# label the zones, and a piezometric line the two materials draw from. Enough to
# exercise the boundary polygonisation, the seed-triangle material assignment, the
# next-line value dialect, the water mapping, and a round-trip through the .xlsx writer
# and load_slope_data.
_SYNTH_FEA = """model description:
version:
11.028
title:
analysis:  solid
strength_reduction_analysis: ON
auto_SRF: ON
change_in_SRF: 0.2
initial_SRF: 1
final_SRF: 2
maxiter_SRF: 500
tolerance_SRF: 0.001
delta_FS: 0.01
tensilestrength_SRF: 1
RFCunits: Metric kPa
stages: 1

material types:
material 1: soilA
 solid properties:
  rhoS: 2 rhoF: 1 porosity: 0.5
 Elastic Properties: LinearElastic
  nu: 0.3 E: 50000
 Plasticity Specifications: MohrCoulomb
  C: 5 phi: 30 dil: 0 T: 5 Cr: 5 phir: 30 Tr: 5 Apply_SSR: 1 Phi_b: 12 Air_Entry: 5 UseUnsaturated: 1
material 2: soilB
 solid properties:
  rhoS: 2 rhoF: 1 porosity: 0.5
 Elastic Properties: LinearElastic
  nu: 0.3 E: 50000
 Plasticity Specifications: Non

material properties:
soilA
1 19 0 50000 8333 0.3 0 20000 20000 20000 0.2 0.2 0.2 1 0 5 0 30 30 5 5 100000
soilB
1 20 0 50000 8333 0.3 0 20000 20000 20000 0.2 0.2 0.2 1 0 25 0 32 32 25 25 100000

element types:
cst_element = type: solid order: linear shape: tri
lst_element = type: solid order: quadratic shape: tri

nodes:  total: 8
1 x: 20 y: 0
2 x: 30 y: 0
3 x: 40 y: 0
4 x: 25 y: 5
5 x: 0 y: 4.5
6 x: 20 y: 4.5
7 x: 40 y: 4.5
8 x: 20 y: 9

elements:  total: 4
1 type: cst_element nodes: [1,2,4] material: soilB materialID: 2
2 type: cst_element nodes: [2,3,4] material: soilB materialID: 2
3 type: cst_element nodes: [5,6,8] material: soilA materialID: 1
4 type: cst_element nodes: [6,7,8] material: soilA materialID: 1

tractions:
e: 1 side: 1 qx1: 0 qy1: -30 qx2: 0 qy2: -30 flag_water: 0
e: 2 side: 1 qx1: 0 qy1: -30 qx2: 0 qy2: -30 flag_water: 0
e: 3 side: 1 qx1: 0 qy1: 12 qx2: 0 qy2: 12 flag_water: 0
e: 4 side: 1 qx1: 0 qy1: 12 qx2: 0 qy2: 12 flag_water: 0

material piezos:
 1
 1

groundwater setup:
  gw_type: "Static Analysis"

piezos:
1
1
2
0 6
40 6

materials mesh start:
  num elements: 2
  num valid stages: 1
  element 0 start:
    P1: 20, 10
    P2: 21, 10
    P3: 20, 11
    materials start:
      stage 1: 1
    materials end:
  element 0 end:
  element 1 start:
    P1: 20, 2
    P2: 21, 2
    P3: 20, 3
    materials start:
      stage 1: 2
    materials end:
  element 1 end:
materials mesh end:

v6 geometry start:
  num boundaries: 2
  boundary 1 start:
    type: "external"
    guid: "{00000000-0000-0000-0000-000000000001}"
    vertices start:
      dpoint array start:
        num points: 5
        0: 0, 0
        1: 40, 0
        2: 40, 18
        3: 20, 18
        4: 0, 9
      dpoint array end:
    vertices end:
  boundary 1 end:
  boundary 2 start:
    type: "material"
    guid: "{00000000-0000-0000-0000-000000000002}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 0, 4.5
        1: 40, 4.5
      dpoint array end:
    vertices end:
  boundary 2 end:
v6 geometry end:

new distributed loads start:
  num distributed loads: 8
  distributed load 1 start:
    unique_id: "{00000000-0000-0000-0000-0000000000D1}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 0, 0
        1: 20, 0
      dpoint array end:
    vertices end:
    strLoadName: "Ponded Water Load 1"
    Dist Load Settings start:
      type: "normal"
      triangular: no
      angle_to_bound: 0
      angle: 0
      flip angle: no
      magnitude1: 1
      magnitude2: 1
      is staged: no
      is_groundwater: yes
      use_calculated_pwp: no
      usesPiezos: yes
      usesGrids: no
      piezoID: 1
      gridID: 0
      totalhead1: 0
      num_stages: 1
    Dist Load Settings end:
  distributed load 1 end:
  distributed load 2 start:
    unique_id: "{00000000-0000-0000-0000-0000000000D2}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 25, 18
        1: 35, 18
      dpoint array end:
    vertices end:
    strLoadName: "Surcharge"
    Dist Load Settings start:
      type: "normal"
      triangular: no
      angle_to_bound: 0
      angle: 0
      flip angle: no
      magnitude1: 40
      magnitude2: 0
      is staged: no
      is_groundwater: no
      use_calculated_pwp: no
      usesPiezos: no
      usesGrids: no
      piezoID: 0
      gridID: 0
      totalhead1: 0
      num_stages: 1
    Dist Load Settings end:
  distributed load 2 end:
  distributed load 3 start:
    unique_id: "{00000000-0000-0000-0000-0000000000D3}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 20, 18
        1: 25, 18
      dpoint array end:
    vertices end:
    strLoadName: "Angled Load"
    Dist Load Settings start:
      type: "normal"
      triangular: no
      angle_to_bound: 0
      angle: 30
      flip angle: no
      magnitude1: 10
      magnitude2: 10
      is staged: no
      is_groundwater: no
      use_calculated_pwp: no
      usesPiezos: no
      usesGrids: no
      piezoID: 0
      gridID: 0
      totalhead1: 0
      num_stages: 1
    Dist Load Settings end:
  distributed load 3 end:
  distributed load 4 start:
    unique_id: "{00000000-0000-0000-0000-0000000000D4}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 40, 18
        1: 35, 18
      dpoint array end:
    vertices end:
    strLoadName: "Cap Weight"
    Dist Load Settings start:
      type: "vertical"
      triangular: yes
      angle_to_bound: 0
      angle: 0
      flip angle: no
      magnitude1: 0
      magnitude2: 60
      is staged: no
      is_groundwater: no
      use_calculated_pwp: no
      usesPiezos: no
      usesGrids: no
      piezoID: 0
      gridID: 0
      totalhead1: 0
      num_stages: 1
    Dist Load Settings end:
  distributed load 4 end:
  distributed load 5 start:
    unique_id: "{00000000-0000-0000-0000-0000000000D5}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 0, 9
        1: 0, 0
      dpoint array end:
    vertices end:
    strLoadName: "Flipped Load"
    Dist Load Settings start:
      type: "normal"
      triangular: no
      angle_to_bound: 0
      angle: 0
      flip angle: yes
      magnitude1: 15
      magnitude2: 15
      is staged: no
      is_groundwater: no
      use_calculated_pwp: no
      usesPiezos: no
      usesGrids: no
      piezoID: 0
      gridID: 0
      totalhead1: 0
      num_stages: 1
    Dist Load Settings end:
  distributed load 5 end:
  distributed load 6 start:
    unique_id: "{00000000-0000-0000-0000-0000000000D6}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 40, 0
        1: 20, 0
      dpoint array end:
    vertices end:
    strLoadName: "Global Down"
    Dist Load Settings start:
      type: "global angle"
      triangular: no
      angle_to_bound: 0
      angle: 270
      flip angle: yes
      magnitude1: 30
      magnitude2: 30
      is staged: no
      is_groundwater: no
      use_calculated_pwp: no
      usesPiezos: no
      usesGrids: no
      piezoID: 0
      gridID: 0
      totalhead1: 0
      num_stages: 1
    Dist Load Settings end:
  distributed load 6 end:
  distributed load 7 start:
    unique_id: "{00000000-0000-0000-0000-0000000000D7}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 0, 4.5
        1: 40, 4.5
      dpoint array end:
    vertices end:
    strLoadName: "Global Uplift"
    Dist Load Settings start:
      type: "global angle"
      triangular: no
      angle_to_bound: 0
      angle: -90
      flip angle: no
      magnitude1: 12
      magnitude2: 12
      is staged: no
      is_groundwater: no
      use_calculated_pwp: no
      usesPiezos: no
      usesGrids: no
      piezoID: 0
      gridID: 0
      totalhead1: 0
      num_stages: 1
    Dist Load Settings end:
  distributed load 7 end:
  distributed load 8 start:
    unique_id: "{00000000-0000-0000-0000-0000000000D8}"
    vertices start:
      dpoint array start:
        num points: 2
        0: 0, 0
        1: 0, 9
      dpoint array end:
    vertices end:
    strLoadName: "Global Slanted"
    Dist Load Settings start:
      type: "global angle"
      triangular: no
      angle_to_bound: 0
      angle: 45
      flip angle: no
      magnitude1: 18
      magnitude2: 18
      is staged: no
      is_groundwater: no
      use_calculated_pwp: no
      usesPiezos: no
      usesGrids: no
      piezoID: 0
      gridID: 0
      totalhead1: 0
      num_stages: 1
    Dist Load Settings end:
  distributed load 8 end:
new distributed loads end:
"""


def _write_synthetic_fez(path):
    """Author a minimal RS2 .fez — a ZIP of one .fea plus empty sidecars."""
    import zipfile
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("synthetic.fea", _SYNTH_FEA)
        zf.writestr("synthetic.p2m", "")       # binary mesh sidecar — ignored
        zf.writestr("synthetic.config", "")    # view config — ignored


def run_rs2_import_test(test):
    """Import a synthetic RS2 .fez and check the model that comes out.

    Guards the schema facts the importer depends on, each silent when wrong:
      - the 'version:' then '11.028' next-line dialect parses;
      - two Mohr-Coulomb materials come across with c/phi (from 'material types:')
        and unit weight (from the 'material properties:' numeric row);
      - an external boundary cut by one material boundary polygonises into two zones,
        each labelled by its material-mesh seed triangle;
      - a piezometric line becomes the piezo line and the materials draw from it;
      - RS2's explicit ponded-water / distributed loads are PRICED into dloads
        (water pressure gamma_w*depth via the referenced piezo; plain loads direct;
        angled/reversed loads reported-not-imported), never count-and-dropped;
      - a ``type: "vertical"`` load imports as a Direction='vertical' dload — a dead
        weight straight down, not a thrust perpendicular to the crest — and that
        Direction survives the .xlsx write and reload;
      - a ``global angle`` load at a quarter turn imports as Direction='vertical' ONLY
        when the file's own solved ``tractions:`` deck reproduces it: the fixture's
        downward one crosses, its uplift twin (deck qy > 0) is reported-and-skipped,
        and one at 45 degrees never reaches the deck at all;
      - the magnitude rules hold: ``magnitude1`` sits at the FIRST stored vertex, and
        a uniform (``triangular: no``) load ignores ``magnitude2`` entirely;
      - each material's linear-elastic ``nu``/``E`` pair comes across from the
        ``Elastic Properties`` line, so no material lands in the file with E = 0;
      - the SSR settings are surfaced as metadata, never turned into an LEM search;
      - NO failure surface is imported (an SSR analysis has none), and that is said;
      - the model round-trips through the .xlsx writer and load_slope_data.
    """
    import tempfile
    from xslope.rs2 import read_fez, fez_to_slope_data, import_fez
    from xslope.fileio import (save_slope_data_to_xlsx, load_slope_data,
                               default_template_path)

    problems = []
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "synthetic.fez")
        _write_synthetic_fez(path)

        d = read_fez(path)
        if d["version"] != "11.028":
            problems.append(f"version parsed as {d['version']!r}, expected '11.028' "
                            f"(the next-line value dialect did not parse)")
        if d["srf"].get("strength_reduction_analysis") != "ON":
            problems.append("SSR settings did not parse from the model description")

        sd, caveats = fez_to_slope_data(d)

        # Two materials, two zones. soilA is Mohr-Coulomb; soilB is RS2's 'Non', which
        # must import as an ELASTIC (impenetrable) material, not a zero-strength soil.
        mats = sd["materials"]
        if len(mats) != 2:
            problems.append(f"{len(mats)} materials, expected 2")
        else:
            if (round(mats[0]["c"], 3), round(mats[0]["phi"], 3),
                    round(mats[0]["gamma"], 3)) != (5.0, 30.0, 19.0):
                problems.append(f"soilA came across as c={mats[0]['c']} phi="
                                f"{mats[0]['phi']} gamma={mats[0]['gamma']}, "
                                f"expected 5/30/19")
            if mats[0].get("option") != "mc":
                problems.append(f"soilA option={mats[0].get('option')!r}, expected 'mc'")
            # v16: RS2's per-material T (5) is a Rankine tension cutoff = xslope's t_cut.
            if mats[0].get("t_cut") != 5.0:
                problems.append(f"soilA t_cut={mats[0].get('t_cut')!r}, expected 5.0 "
                                f"(RS2's per-material T did not map to t_cut)")
            # v17: RS2's Phi_b (12) with UseUnsaturated on must import as phi_b.
            if mats[0].get("phi_b") != 12.0:
                problems.append(f"soilA phi_b={mats[0].get('phi_b')!r}, expected 12.0 "
                                f"(RS2 unsaturated Phi_b did not map)")
            # soilB is 'Non' -> elastic, with its unit weight kept and no invented c/phi.
            if mats[1].get("option") != "elastic":
                problems.append(f"soilB option={mats[1].get('option')!r}, expected "
                                f"'elastic' (RS2's 'Non' did not map to elastic)")
            if round(mats[1].get("gamma", 0), 3) != 20.0:
                problems.append(f"soilB gamma={mats[1].get('gamma')}, expected 20 "
                                f"(the elastic material lost its unit weight)")
            if mats[1].get("t_cut") is not None:
                problems.append("the elastic soilB was given a t_cut, which is meaningless")
            # Both materials state ``nu: 0.3 E: 50000`` on their Elastic Properties
            # line, and RS2 solved this model with them: they are inputs, and a
            # material arriving with E = 0 is a singular stiffness matrix the FEM
            # would hit later. The elastic ('Non') material needs its pair MOST —
            # elastic constants are the only thing it has.
            for m in mats:
                if (round(m.get("E") or 0, 3), round(m.get("nu") or 0, 3)) != (50000.0, 0.3):
                    problems.append(
                        f"material {m['name']!r} came across with E={m.get('E')!r} "
                        f"nu={m.get('nu')!r}, expected 50000/0.3 from the RS2 "
                        f"'Elastic Properties' line")
        # Every material states its elastic pair, so nothing may be attributed to the
        # soil-type fallback — a caveat here would mean the parse silently missed it.
        if any("soil-type table" in c for c in caveats):
            problems.append("the E/nu soil-type fallback fired on a file that states "
                            "nu/E for every material")
        # The elastic material must NOT trip the zero-strength warning, and the mapping
        # notes must all be reported (caveat discipline, never silent).
        if any("NO STRENGTH" in c and "soilB" in c for c in caveats):
            problems.append("the elastic material was wrongly flagged as zero-strength")
        if not any("elastic" in c and "soilB" in c for c in caveats):
            problems.append("the 'Non' -> elastic mapping was not reported as a caveat")
        if not any("unsaturated" in c.lower() and "soilA" in c for c in caveats):
            problems.append("the RS2 unsaturated (phi_b) import was not reported")
        if not any("brittle" in c.lower() or "residual" in c.lower() for c in caveats):
            problems.append("RS2's residual/brittle tensile Tr was not reported")
        if not any("tensilestrength_SRF" in c or "reduced tensile" in c for c in caveats):
            problems.append("the RS2 tensilestrength_SRF run flag was not reported")
        if len(sd["polygons"]) != 2:
            problems.append(f"{len(sd['polygons'])} polygons, expected 2 (the external "
                            f"boundary did not polygonise into two zones)")

        # A piezometric line -> piezo line; both materials draw from it.
        if len(sd["piezo_line"]) != 2:
            problems.append(f"piezo_line has {len(sd['piezo_line'])} points, "
                            f"expected 2 (the water table did not import)")
        if [m["u"] for m in mats] != ["piezo", "piezo"]:
            problems.append(f"materials take u={[m['u'] for m in mats]}, expected "
                            f"['piezo', 'piezo']")

        # Distributed loads: RS2 stores ponded water as explicit load objects, and the
        # importer must PRICE them, not count-and-drop them. The fixture carries eight:
        #   1. a piezo-driven ponded-water load on the y=0 boundary — priced as the
        #      water pressure gamma_w * (piezo_head 6 - y 0) = 9.81 * 6 = 58.86;
        #   2. a plain numeric normal load (a surcharge), magnitude1 = 40 with a STALE
        #      magnitude2 = 0 — imported as a uniform 40, because RS2 reads magnitude2
        #      only on a triangular load (reading it always would halve this one);
        #   3. an angled (30 deg) load — reported, NOT imported (no xslope equivalent);
        #   4. a vertical triangular load on the crest, stored right-to-left with
        #      magnitude1 = 0 at (40,18) and magnitude2 = 60 at (35,18) — imported as a
        #      Direction='vertical' dload, and left-to-right after the writer's
        #      orientation pass, so 60 lands at x = 35 and 0 at x = 40;
        #   5. a reversed ('flip angle') load — reported, NOT imported, since RS2 aims
        #      it away from the boundary and importing it as a surcharge would flip
        #      the sign silently.
        #   6. a 'global angle' load at 270 with flip on, over y = 0, x = 20..40. The
        #      angle/flip pair is unreadable on its own (the archive writes the same
        #      downward load both ways), so the file's solved tractions: deck decides:
        #      it carries qx = 0, qy = -30 on those two element sides, which is exactly
        #      the mapping, so the load imports as a uniform Direction='vertical' 30;
        #   7. its uplift twin over y = 4.5 — same shape, but the deck there reads
        #      qy = +12, AGAINST the downward mapping, so it is reported and skipped
        #      rather than imported with a silently flipped sign;
        #   8. a 'global angle' load at 45 deg — never a candidate at all (it has a
        #      horizontal component xslope's distributed load cannot carry), so it is
        #      skipped without the deck being consulted.
        dloads = sd["dloads"]
        dirs = sd.get("dload_dirs") or []
        if len(dloads) != 4:
            problems.append(f"{len(dloads)} distributed load(s) imported, expected 4 "
                            f"(ponded water, a plain load, a vertical load and the "
                            f"deck-confirmed 'global angle' one; the angled, reversed, "
                            f"uplift and 45-degree loads are skipped)")
        else:
            peak = max(pt["Normal"] for pt in dloads[0])
            if round(peak, 2) != round(9.81 * 6.0, 2):
                problems.append(f"ponded-water peak Normal {peak:.2f}, expected "
                                f"{9.81 * 6.0:.2f} (gamma_w * depth)")
            if not any(all(round(pt["Normal"], 3) == 40.0 for pt in b) for b in dloads):
                problems.append("the plain numeric distributed load did not import as a "
                                "uniform 40 (magnitude2 is read only when triangular)")
            if dirs != ["normal", "normal", "vertical", "vertical"]:
                problems.append(f"dload_dirs={dirs}, expected ['normal', 'normal', "
                                f"'vertical', 'vertical'] (an RS2 'vertical' or "
                                f"quarter-turn 'global angle' load did not become a "
                                f"Direction='vertical' distributed load)")
            vert = dloads[2]
            got = [(round(p["X"], 3), round(p["Normal"], 3)) for p in vert]
            if got != [(40.0, 0.0), (35.0, 60.0)]:
                problems.append(f"the vertical triangular load imported as {got}, "
                                f"expected [(40.0, 0.0), (35.0, 60.0)] — magnitude1 "
                                f"belongs at the FIRST stored vertex")
            ga = [(round(p["X"], 3), round(p["Y"], 3), round(p["Normal"], 3))
                  for p in dloads[3]]
            if ga != [(40.0, 0.0, 30.0), (20.0, 0.0, 30.0)]:
                problems.append(f"the deck-confirmed 'global angle' load imported as "
                                f"{ga}, expected a uniform 30 over (40,0)-(20,0)")
        if not any("ponded-water load" in c for c in caveats):
            problems.append("the ponded-water load import was not reported as a caveat")
        if not any("Direction='vertical'" in c for c in caveats):
            problems.append("the vertical load import was not reported as a caveat")
        if not any("solved traction deck" in c and "global angle" in c for c in caveats):
            problems.append("the 'global angle' import did not report that it was "
                            "checked against the file's own solved traction deck")
        if not any("distributed load(s) were NOT imported" in c and "flip angle" in c
                   for c in caveats):
            problems.append("the skipped angled/reversed loads were not reported")
        _skipped = next((c for c in caveats
                         if "distributed load(s) were NOT imported" in c), "")
        if "Global Uplift" not in _skipped:
            problems.append("the uplift 'global angle' load (whose deck contradicts the "
                            "downward mapping) was not reported by name as skipped")
        if "Global Slanted" not in _skipped:
            problems.append("the 45-degree 'global angle' load was not reported by name "
                            "as skipped")
        if "Global Down" in _skipped:
            problems.append("the deck-confirmed 'global angle' load was reported as "
                            "skipped as well as imported")

        # An RS2 model imports NO failure surface, and must say so.
        if sd["circles"] or sd["non_circ"]:
            problems.append("an RS2 model imported a failure surface it should not have")
        if not any("no failure surface" in c for c in caveats):
            problems.append("the missing failure surface was not reported")
        if not any("Shear-Strength-Reduction" in c for c in caveats):
            problems.append("the SSR analysis metadata was not reported as a caveat")

        # --- water loads stay MANUAL, and that is the point of this importer -------
        # RS2 is the exception to the automatic-water rule, for a reason in ITS data
        # model: it stores ponded water as an explicit load object (imported above),
        # and its piezometric surface is a whole-domain surface rather than a water
        # table drawn on the ground. Deriving ground-against-piezo here would invent a
        # plateau of water the model never had -- wrong, not merely redundant. So the
        # import pins D23 to manual, keeps the vendor's own loads, and says why.
        from xslope.water import with_water_loads, derived_blocks, derive_water_loads
        if sd.get("water_loads") != "manual":
            problems.append(
                f"an RS2 model imported as {sd.get('water_loads')!r}; it must import "
                f"as 'manual' — its ponded-water loads are explicit vendor objects, "
                f"and its whole-domain piezo would derive a load the model never had")
        if derived_blocks(with_water_loads(sd), 1):
            problems.append("an RS2 import derived a water load on top of the "
                            "vendor's own explicit ponded-water load")
        if not any("D23" in c and "MANUAL" in c.upper() for c in caveats):
            problems.append("the import did not report that it left the water-load "
                            "mode on manual, or did not name the cell")
        # The REASON for pinning the mode is a claim about RS2's whole-domain piezo,
        # and a hand-built fixture cannot settle it. It is measured instead on the
        # vendor's own archive, in run_rs2_water_mode_test.

        # Round-trip: the geometry must survive the .xlsx writer AND reload. An RS2
        # import carries no surface, so give it one (as a user must) before saving —
        # load_slope_data rejects a surface-less file, which is the real consumer.
        sd["circular"] = True
        sd["circles"] = [{"Xo": 20.0, "Yo": 30.0, "R": 26.0, "Depth": 4.0}]
        xlsx = os.path.join(td, "synthetic.xlsx")
        save_slope_data_to_xlsx(sd, xlsx, template=default_template_path())
        reloaded = load_slope_data(xlsx)
        if len(reloaded["materials"]) != 2:
            problems.append(f"reloaded model has {len(reloaded['materials'])} "
                            f"materials, expected 2")
        if len(reloaded["polygons"]) != 2:
            problems.append("the two zones did not survive the .xlsx round-trip")
        if len(reloaded.get("piezo_line") or []) != 2:
            problems.append("the piezo line did not survive the .xlsx round-trip")
        if len(reloaded.get("dloads") or []) != 4:
            problems.append("the distributed loads did not survive the .xlsx round-trip")
        # The Direction cell is the point of the vertical load: it must come back off
        # the sheet, and the writer's left-to-right orientation pass must carry each
        # (X, Y, Normal) triple intact rather than re-shaping the ramp.
        elif (reloaded.get("dload_dirs") or [])[2:] != ["vertical", "vertical"]:
            problems.append(
                f"reloaded dload_dirs={reloaded.get('dload_dirs')}, expected both "
                f"vertical loads' Direction to survive the .xlsx round-trip")
        else:
            back = [(round(p["X"], 3), round(p["Normal"], 3))
                    for p in reloaded["dloads"][2]]
            if back != [(35.0, 60.0), (40.0, 0.0)]:
                problems.append(f"the vertical load reloaded as {back}, expected "
                                f"[(35.0, 60.0), (40.0, 0.0)]")
            back_ga = [(round(p["X"], 3), round(p["Normal"], 3))
                       for p in reloaded["dloads"][3]]
            if back_ga != [(20.0, 30.0), (40.0, 30.0)]:
                problems.append(f"the 'global angle' load reloaded as {back_ga}, "
                                f"expected [(20.0, 30.0), (40.0, 30.0)]")
        # E/nu must survive the round-trip too — the FEM reads them off the sheet.
        if [(round(m.get("E") or 0, 3), round(m.get("nu") or 0, 3))
                for m in reloaded["materials"]] != [(50000.0, 0.3), (50000.0, 0.3)]:
            problems.append("the imported E/nu did not survive the .xlsx round-trip")

        # import_fez writes an .xlsx and returns the caveat list (the surface-less file
        # it writes is intentionally incomplete — an SSR model has no LEM surface).
        out2 = os.path.join(td, "direct.xlsx")
        rcav = import_fez(path, default_template_path(), out2)
        if not os.path.exists(out2):
            problems.append("import_fez did not write the .xlsx")
        if not isinstance(rcav, list):
            problems.append("import_fez did not return a caveat list")

    if problems:
        return None, "RS2 import: " + "; ".join(problems[:5])
    return 0.0, None


# Rocscience's public RS2 verification models, used READ-ONLY by the water-mode and
# load-deck tests below. They are Rocscience's copyrighted material and are NOT in this
# repository — the tests read them from a local copy of the downloads and skip cleanly
# when that copy is absent (any clone but the author's).
RS2_VENDOR_ARCHIVE = os.path.expanduser(
    '~/python_projects/vendor_files/rocscience_downloads')
RS2_VENDOR_ZIP = 'RS2_Slope-Stability-Verification-RS2-and-Slide2-Import.zip'
# member basename -> what the importer must produce. Each pins one decoded
# iStaticWaterMode against a file whose water source is known from the printed
# verification manual and from RS2's own solved nodal pore-pressure field:
#   #017_01/02/03 are the controlled triple — the SAME slope authored dry, with
#   ru = 0.25, and with a piezometric line, differing in nothing else;
#   #007 takes its pore pressure from a 185-point grid xslope cannot read;
#   #016 is a MIXED model (one dry material + three on the piezo line), which a
#   chained if/elif reports only half of.
RS2_WATER_CASES = {
    'slope stability #017_01.fez': {'u': ['none'], 'ru': [0.0], 'piezo_pts': 0},
    'slope stability #017_02.fez': {'u': ['ru'], 'ru': [0.25], 'piezo_pts': 0},
    'slope stability #017_03.fez': {'u': ['piezo'], 'ru': [0.0], 'piezo_pts': 3},
    'slope stability #007.fez': {'u': ['none'], 'ru': [0.0], 'piezo_pts': 0,
                                 'caveat': 'grid'},
    'slope stability #016.fez': {'u': ['none', 'piezo', 'piezo', 'piezo'],
                                 'ru': [0.0, 0.0, 0.0, 0.0], 'piezo_pts': 7},
}


def run_rs2_water_mode_test(test):
    """RS2 per-material water source (iStaticWaterMode), against the vendor's own files.

    RS2 picks the pore-pressure source PER MATERIAL, and the importer read only one of
    them (the piezo line): an ru ratio or a pore-pressure grid silently imported dry,
    which removes u everywhere and drives the factor of safety non-conservatively high.
    This pins the decode on files whose water source is independently known:

      - mode 2 with ru > 0  -> u = 'ru' carrying the ratio from the material-property
        row (#017_02: ru = 0.25), and mode 2 with ru = 0 stays dry (#017_01);
      - mode 3 -> u = 'piezo' with the referenced line's points (#017_03);
      - mode 4 -> still dry (xslope has no grid water source) but LOUDLY caveated as a
        grid, never silent (#007);
      - a MIXED model assigns each material its own source, and the caveats are
        de-chained so every group is reported (#016).

    Skips cleanly (0.0, None) when the vendor archive is not present locally — the
    files are Rocscience's copyrighted material and are not in this repository.
    """
    import tempfile
    import zipfile
    from xslope.rs2 import read_fez, fez_to_slope_data

    zpath = os.path.join(RS2_VENDOR_ARCHIVE, RS2_VENDOR_ZIP)
    if not os.path.exists(zpath):
        return 0.0, None                       # no local vendor archive: nothing to read

    problems = []
    with zipfile.ZipFile(zpath) as oz:
        # The dual-authored archive holds each model twice (native + Slide2-import);
        # take the native "Slope Stability Verification/" copy.
        members = {os.path.basename(n): n for n in oz.namelist()
                   if n.lower().endswith('.fez') and '(Slide2 Import)' not in n}
        with tempfile.TemporaryDirectory() as td:
            for base, want in RS2_WATER_CASES.items():
                if base not in members:
                    problems.append(f"{base} is missing from the vendor archive")
                    continue
                fez = os.path.join(td, 'case.fez')
                with open(fez, 'wb') as fh:
                    fh.write(oz.read(members[base]))
                sd, caveats = fez_to_slope_data(read_fez(fez))
                got_u = [m['u'] for m in sd['materials']]
                got_ru = [round(float(m['ru']), 6) for m in sd['materials']]
                if got_u != want['u']:
                    problems.append(f"{base}: u = {got_u}, expected {want['u']}")
                if got_ru != want['ru']:
                    problems.append(f"{base}: ru = {got_ru}, expected {want['ru']}")
                if len(sd['piezo_line']) != want['piezo_pts']:
                    problems.append(f"{base}: piezo line has {len(sd['piezo_line'])} "
                                    f"points, expected {want['piezo_pts']}")
                key = want.get('caveat')
                if key and not any(key in c and 'WILL BE WRONG' in c for c in caveats):
                    problems.append(f"{base}: no loud '{key}' caveat — the unimported "
                                    f"water source was dropped silently")
                # An imported ru must say so; a dry model must not invent the note.
                said_ru = any('ru pore-pressure ratio' in c for c in caveats)
                if ('ru' in want['u']) != said_ru:
                    problems.append(f"{base}: ru caveat reported={said_ru}, expected "
                                    f"{'ru' in want['u']}")

            # --- THE RS2 EXCEPTION, MEASURED ON THE WHOLE ARCHIVE -----------------
            # Every other importer hands its water to the engine's derivation; RS2
            # does not, and the reason is a claim about RS2's data model rather than
            # about ours: its piezometric surface is a WHOLE-DOMAIN surface, so
            # measuring the ground against it invents a reservoir the model never
            # had. That claim is checkable, so it is checked -- on the vendor's own
            # files, by asking how many of them WOULD acquire a water load under
            # automatic mode while RS2 itself defines no ponded-water load at all.
            # If the answer ever became zero, pinning the mode would be superstition.
            from xslope.water import derive_water_loads, block_resultant
            n_read, spurious, worst = 0, [], 0.0
            for base, member in sorted(members.items()):
                fez = os.path.join(td, 'sweep.fez')
                with open(fez, 'wb') as fh:
                    fh.write(oz.read(member))
                try:
                    sd, _cav = fez_to_slope_data(read_fez(fez))
                except Exception:
                    continue
                n_read += 1
                if sd.get('water_loads') != 'manual':
                    problems.append(f"{base}: imported "
                                    f"{sd.get('water_loads')!r}, expected 'manual'")
                would = derive_water_loads(dict(sd, water_loads='auto'))
                if would['blocks'] and not sd['dloads']:
                    spurious.append(base)
                    worst = max(worst, sum(block_resultant(b)
                                           for b in would['blocks']))
            if n_read < 50:
                problems.append(f"only {n_read} vendor models imported, so the "
                                f"archive-wide check proves little")
            elif not spurious:
                problems.append(
                    f"none of {n_read} vendor RS2 models would acquire a water load "
                    f"under automatic mode, so the recorded reason for pinning them "
                    f"to manual — a whole-domain piezo inventing a reservoir — is no "
                    f"longer demonstrated by the archive")

    if problems:
        return None, "RS2 water modes: " + "; ".join(problems[:5])
    return 0.0, None


# Every ``global angle`` load in the public RS2 verification archive: 11 loads across 8
# models, all of them a quarter turn. They are the reason the traction deck is read at
# all — the stored angle/flip pair contradicts itself across the set, and both readings
# are represented here:
#
#   #006 / #020 / #021 / #054   angle 270 with ``flip angle: yes``
#   #009 / #025 / #026 / #060   angle -90, #009 and #025 with no flip, #026 and #060
#                               with it
#
# and all eleven solved decks push straight DOWN. ``slide2`` picks which copy of the
# dual-authored archive the model lives in. Each entry is the full imported dload list:
# every (X, Y, Normal) triple, in stored order, so a magnitude, a ramp direction or a
# sign that moves is caught. #060-slope7 is the coverage case — two collinear loads
# sharing the vertex (5, 25), where a deck side must be attributed to exactly one.
RS2_LOAD_DECK_CASES = {
    'slope stability #006.fez': {
        'slide2': False,
        'blocks': [[(43.0, 27.75, 20.0), (23.0, 27.75, 20.0)],
                   [(80.0, 40.0, 40.0), (70.0, 40.0, 20.0)]],
        'e_nu': [(50000.0, 0.4), (50000.0, 0.4)]},
    'slope stability #009.fez': {
        'slide2': True,
        'blocks': [[(23.0, 27.75, 20.0), (43.0, 27.75, 20.0)],
                   [(80.0, 40.0, 40.0), (70.0, 40.0, 20.0)]]},
    'slope stability #020.fez': {
        'slide2': False,
        'blocks': [[(15.7735, 10.0, 149.3127), (5.7735, 10.0, 149.3127)]]},
    'slope stability #021.fez': {
        'slide2': False,
        'blocks': [[(5.0, 10.0, 102.83), (-5.0, 10.0, 102.83)]]},
    'slope stability #025.fez': {
        'slide2': True,
        'blocks': [[(5.7735, 10.0, 149.3127), (15.7735, 10.0, 149.3127)]]},
    'slope stability #026.fez': {
        'slide2': True,
        'blocks': [[(-5.0, 10.0, 102.83), (5.0, 10.0, 102.83)]]},
    'slope stability #054.fez': {
        'slide2': False,
        'blocks': [[(24.0, 15.0, 20.0), (8.7, 15.0, 20.0)]]},
    'slope stability #060-slope7.fez': {
        'slide2': True,
        'blocks': [[(0.0, 25.0, 500.0), (5.0, 25.0, 500.0)],
                   [(5.0, 25.0, 250.0), (50.0, 25.0, 250.0)]]},
}

#: The model the deck-mutation checks run on — the smallest deck in the set.
RS2_DECK_MUTATION_CASE = 'slope stability #021.fez'


def run_rs2_load_deck_test(test):
    """RS2 'global angle' loads and material E/nu, against the vendor's own files.

    Both are things RS2 states and the importer used to drop. A ``global angle`` load
    at a quarter turn is a plain vertical surcharge, but which WAY it points cannot be
    read from what RS2 stores: #006 writes 270 degrees with ``flip angle: yes`` and its
    Slide2-import twin #009 writes -90 with no flip, and both apply the same load
    downward. So the importer scores its mapping against the file's OWN solved
    ``tractions:`` deck — the edge traction RS2 assembled at every mesh node of the
    loaded boundary — and imports only what the deck reproduces.

    This runs that through the real file path on all 11 such loads (8 models, both
    dialects, quadratic meshes, and the collinear pair that shares a vertex), then
    MUTATES one file's deck four ways to prove the check can fail:

      - the deck reversed (uplift) -> refused, not imported with a flipped sign;
      - the deck halved -> refused (a magnitude the mapping cannot reproduce);
      - the deck given a horizontal component -> refused (not vertical after all);
      - the deck deleted, as in an unsolved model -> refused, because no deck is not
        permission to guess.

    E/nu ride along on the same files: RS2 states ``nu``/``E`` per material on its
    ``Elastic Properties`` line and solves the published SSR with them, so they are
    inputs. A material arriving with E = 0 is the silent-default defect class the
    preflight audit traced 16 deleted elements to.

    Skips cleanly (0.0, None) when the vendor archive is not present locally — the
    files are Rocscience's copyrighted material and are not in this repository.
    """
    import io
    import re as _re
    import tempfile
    import zipfile
    from xslope.rs2 import read_fez, fez_to_slope_data

    zpath = os.path.join(RS2_VENDOR_ARCHIVE, RS2_VENDOR_ZIP)
    if not os.path.exists(zpath):
        return 0.0, None                       # no local vendor archive: nothing to read

    def _blocks(sd):
        return [[(round(p['X'], 4), round(p['Y'], 4), round(p['Normal'], 4))
                 for p in b] for b in sd['dloads']]

    problems = []
    with zipfile.ZipFile(zpath) as oz:
        members = {}
        for n in oz.namelist():
            if n.lower().endswith('.fez'):
                members[(os.path.basename(n), '(Slide2 Import)' in n)] = n
        with tempfile.TemporaryDirectory() as td:
            fez = os.path.join(td, 'case.fez')
            for base, want in RS2_LOAD_DECK_CASES.items():
                key = (base, want['slide2'])
                if key not in members:
                    problems.append(f"{base} is missing from the vendor archive")
                    continue
                raw = oz.read(members[key])
                with open(fez, 'wb') as fh:
                    fh.write(raw)
                sd, caveats = fez_to_slope_data(read_fez(fez))
                got = _blocks(sd)
                if got != want['blocks']:
                    problems.append(f"{base}: imported dloads {got}, expected "
                                    f"{want['blocks']}")
                dirs = sd.get('dload_dirs') or []
                if dirs != ['vertical'] * len(want['blocks']):
                    problems.append(f"{base}: dload_dirs = {dirs}, expected every "
                                    f"'global angle' load to import as 'vertical'")
                if not any('solved traction deck' in c for c in caveats):
                    problems.append(f"{base}: the deck check that authorised the import "
                                    f"was not reported as a caveat")
                if any('distributed load(s) were NOT imported' in c for c in caveats):
                    problems.append(f"{base}: a load was skipped that the deck confirms")
                # E/nu: stated per material, so nothing may be zero and nothing may be
                # attributed to the soil-type fallback.
                zero = [m['name'] for m in sd['materials'] if not m.get('E')]
                if zero:
                    problems.append(f"{base}: material(s) {zero} imported with E = 0")
                if any('soil-type table' in c for c in caveats):
                    problems.append(f"{base}: the E/nu soil-type fallback fired on a "
                                    f"file that states nu/E for every material")
                if 'e_nu' in want:
                    got_e = [(round(m.get('E') or 0, 4), round(m.get('nu') or 0, 4))
                             for m in sd['materials']]
                    if got_e != want['e_nu']:
                        problems.append(f"{base}: E/nu = {got_e}, expected "
                                        f"{want['e_nu']}")

            # --- mutation: the deck must be able to REFUSE ------------------------
            base = RS2_DECK_MUTATION_CASE
            key = (base, RS2_LOAD_DECK_CASES[base]['slide2'])
            if key in members:
                inner = zipfile.ZipFile(io.BytesIO(oz.read(members[key])))
                fea = next(n for n in inner.namelist() if n.lower().endswith('.fea'))
                lines = inner.read(fea).decode('latin-1').splitlines()
                start = lines.index('tractions:')
                stop = start + 1
                while lines[stop].startswith('e:'):
                    stop += 1
                deck = lines[start + 1:stop]

                def _rewrite(new_deck):
                    """The same model with its traction deck replaced."""
                    body = '\n'.join(lines[:start + 1] + new_deck + lines[stop:])
                    buf = io.BytesIO()
                    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for n in inner.namelist():
                            zf.writestr(n, body.encode('latin-1') if n == fea
                                        else inner.read(n))
                    with open(fez, 'wb') as fh:
                        fh.write(buf.getvalue())
                    return fez_to_slope_data(read_fez(fez))

                mutations = {
                    'reversed (uplift)':
                        [_re.sub(r'qy(\d): -', r'qy\1: ', ln) for ln in deck],
                    'halved':
                        [ln.replace('-102.83', '-51.415') for ln in deck],
                    'given a horizontal component':
                        [_re.sub(r'qx(\d): 0', r'qx\1: 5', ln) for ln in deck],
                    'deleted (an unsolved model)': [],
                }
                for label, new_deck in mutations.items():
                    if new_deck == deck:
                        problems.append(f"the '{label}' mutation did not change the "
                                        f"deck — the check would pass vacuously")
                        continue
                    msd, mcav = _rewrite(new_deck)
                    if msd['dloads']:
                        problems.append(
                            f"{base} with its deck {label}: the load still imported as "
                            f"{_blocks(msd)} — the deck check does not bite")
                    if not any('distributed load(s) were NOT imported' in c for c in mcav):
                        problems.append(f"{base} with its deck {label}: the refusal was "
                                        f"not reported as a caveat")
                # And unmutated, the same rewrite path still imports — so the four
                # refusals above are the mutations talking, not the rewrite.
                bsd, _ = _rewrite(deck)
                if _blocks(bsd) != RS2_LOAD_DECK_CASES[base]['blocks']:
                    problems.append(f"{base} rebuilt with its own deck no longer "
                                    f"imports — the mutation harness is not neutral")

    if problems:
        return None, "RS2 load deck: " + "; ".join(problems[:5])
    return 0.0, None


def run_submerged_oracle_test(test):
    """Dry-buoyant still-water oracle (benchmarks/submerged_oracle_guard.py).

    A fully submerged slope under still water must read identically to the same slope
    run dry with gamma' = gamma_sat - gamma_w. This is the exact effective-stress
    equivalence that validates the perpendicular-distributed-load water convention the
    Slide2 and RS2 importers now convert ponded water into. Returns (0.0, None) on
    pass, else (None, message). File-less: builds on xslope_acads_simple.xlsx."""
    import importlib
    bench = str(Path(__file__).parent / 'benchmarks')
    if bench not in sys.path:
        sys.path.insert(0, bench)
    mod = importlib.import_module('submerged_oracle_guard')
    if not os.path.exists(mod._BASE_XLSX):
        return 0.0, None                      # engine-only clone without the docs file
    failures = mod.check()
    if failures:
        return None, "submerged oracle: " + "; ".join(failures[:4])
    return 0.0, None


def run_suction_guard_test(test):
    """Matric-suction apparent-cohesion guard (benchmarks/suction_guard.py).

    A simple homogeneous slope with a piezo line well below the surface, so the
    upper failure surface carries matric suction. Freezes the opt-in Fredlund
    extended-Mohr-Coulomb option: OFF (None / phi_b=0) is bit-identical to the
    clamped baseline, phi_b>0 strictly raises FS, suction_cap gives baseline <
    capped < uncapped, and c_suction = max(0, gamma_w*(y_base - y_piezo))*tan(phi_b)
    exactly. Returns (0.0, None) on pass, else (None, message). File-less: builds
    on xslope_acads_simple.xlsx."""
    import importlib
    bench = str(Path(__file__).parent / 'benchmarks')
    if bench not in sys.path:
        sys.path.insert(0, bench)
    mod = importlib.import_module('suction_guard')
    if not os.path.exists(mod._BASE_XLSX):
        return 0.0, None                      # engine-only clone without the docs file
    # LEM piezo strength + seep u-source delivery + v17 file auto-wiring, plus the
    # FEM/SSRM branch (off-by-default bit-identical, phi_b engages above the WT, the
    # apparent cohesion reduced by the trial F — the RS2/SIGMA-W SRF treatment).
    failures = mod.check()
    failures += mod.check_seep()
    failures += mod.check_autowire()
    failures += mod.check_fem()
    if failures:
        return None, "matric suction: " + "; ".join(failures[:4])
    return 0.0, None


def run_piezo_u_guard_test(test):
    """Silent-zero pore-pressure guards (benchmarks/piezo_extent_guard.py).

    Three paths that used to deliver u = 0 without saying so, each of which
    over-predicts the factor of safety: an unrecognized ``u`` option on the mat
    sheet (a typo fell through every branch), ``u = piezo`` in a file with no
    piezometric line at all, and a point that samples a piezometric line from
    outside the line's own x-extent (the interpolation returns NaN there, which
    was read as zero). Checks the load-time u-option rejection by material name,
    the LEM slicer's per-slice no-line and extent errors, the FEM's nodal and
    Gauss-point versions of both, and — the other half of the contract — that a
    line stopping short of the domain but still covering what samples it solves
    unchanged. Returns (0.0, None) on pass, else (None, message). File-less
    (builds on xslope_acads_simple); the FEM half skips without gmsh."""
    import importlib
    bench = str(Path(__file__).parent / 'benchmarks')
    if bench not in sys.path:
        sys.path.insert(0, bench)
    mod = importlib.import_module('piezo_extent_guard')
    if not os.path.exists(mod._BASE_XLSX):
        return 0.0, None                      # engine-only clone without the docs file
    failures = mod.check()
    if failures:
        return None, "silent-zero pore pressure: " + "; ".join(failures[:4])
    return 0.0, None


def run_kernel_xcheck_test(test):
    """Fast-kernel divergence fence (benchmarks/kernel_xcheck.py).

    The pure-NumPy Step-6 path is the oracle; when the optional compiled kernel is
    built (setup_kernel.py), it must reproduce it bit-for-bit. Solves three small
    coarse cases (plain Mohr-Coulomb, Rankine tension cutoff, matric suction) both
    ways and fails on any factor-of-safety disagreement or displacement-field
    max-abs difference above 1e-8. Returns (0.0, None) on agreement, else
    (None, message). Cheap by construction (coarse meshes, <2 min).

    This test is only appended to the suite when the compiled kernel imports (see
    main()), so reaching here means the kernel is present. Still guards the input
    files so an engine-only clone without the docs corpus skips cleanly."""
    import importlib
    bench = str(Path(__file__).parent / 'benchmarks')
    if bench not in sys.path:
        sys.path.insert(0, bench)
    mod = importlib.import_module('kernel_xcheck')
    if not mod.kernel_available():
        return 0.0, None                      # kernel not built: nothing to cross-check
    if not mod.files_present():
        return 0.0, None                      # engine-only clone without the docs corpus
    failures = mod.check()
    if failures:
        return None, "fast kernel diverged: " + "; ".join(failures[:4])
    return 0.0, None


def run_no_void_test(test):
    """Corpus material-tiling void guard (benchmarks/void_guard.py).

    Every corpus input file's material polygons must tile its domain with no
    interior void — no enclosed column-gap that is both wide and tall. This is the
    defect that excised Slide2 VP42's downstream toe wedge and starved its
    reservoir-side surfaces of resisting material. Returns (0.0, None) on a clean
    corpus, else (None, message). File-less: sweeps the corpus itself."""
    import importlib
    bench = str(Path(__file__).parent / 'benchmarks')
    if bench not in sys.path:
        sys.path.insert(0, bench)
    mod = importlib.import_module('void_guard')
    if not mod.corpus_files():
        return 0.0, None                      # engine-only clone without the docs corpus
    failures = mod.check()
    if failures:
        return None, "corpus void: " + "; ".join(failures[:4])
    return 0.0, None


def run_vg_kr_test(test):
    """Unit check for the van Genuchten relative-permeability function and the
    kr-model dispatch (xslope.seep). Verifies kr_vg_vec against an independent
    scalar evaluation of the Mualem-vG formula across pressure heads and (alpha, n),
    the saturation / floor / monotonicity behavior, and that the dispatcher reduces
    EXACTLY to the linear-front kr when no vG model is active. Returns (0.0, None)
    on success, else (None, message). No mesh / gmsh needed."""
    import numpy as np
    from xslope.seep import (kr_vg_vec, kr_relative_vec, kr_relative,
                             kr_frontal_vec, KR_LF, KR_VG)
    KR_MIN = 1e-4

    def vg_scalar(p, a, n):
        # Independent scalar reference for Mualem-van Genuchten kr.
        n = max(n, 1.0 + 1e-6)
        m = 1.0 - 1.0 / n
        if p >= 0.0:
            return 1.0
        se = (1.0 + (a * abs(p)) ** n) ** (-m)
        kr = (se ** 0.5) * (1.0 - (1.0 - se ** (1.0 / m)) ** m) ** 2
        return min(max(kr, KR_MIN), 1.0)

    problems = []
    heads = np.array([-200., -50., -10., -3., -1., -0.3, -0.05, 0.0, 2.0])
    for a, n in [(0.075, 1.89), (2.286, 1.89), (1.097, 1.56), (0.5, 1.1), (3.0, 4.0)]:
        got = kr_vg_vec(heads, a, n)
        ref = np.array([vg_scalar(float(p), a, n) for p in heads])
        if not np.allclose(got, ref, atol=1e-12, rtol=1e-9):
            problems.append(f"kr_vg_vec != Mualem-vG for (a={a}, n={n}): max|d|={np.max(np.abs(got-ref)):.2e}")
        if got[7] != 1.0 or got[8] != 1.0:
            problems.append(f"kr not 1.0 at saturation for (a={a}, n={n})")
        unsat = got[:7]
        if np.any(unsat < KR_MIN - 1e-15) or np.any(unsat > 1.0 + 1e-12):
            problems.append(f"kr out of [{KR_MIN},1] for (a={a}, n={n})")
        if np.any(np.diff(unsat) < -1e-12):     # drier (more negative p) -> lower kr
            problems.append(f"kr not monotonic in p for (a={a}, n={n})")

    # Dispatch must reduce EXACTLY to the linear front when no vG model is active.
    P = np.array([[-2.0, -1.0, 0.5], [-5.0, -0.2, 1.0]])
    kr0 = np.array([0.001, 0.01]); h0 = np.array([-1.0, -2.0])
    base = kr_frontal_vec(P, kr0[:, None], h0[:, None])
    if not np.array_equal(kr_relative_vec(P, kr0[:, None], h0[:, None]), base):
        problems.append("kr_relative_vec(model=None) != kr_frontal_vec (lf not bit-identical)")
    alllf = kr_relative_vec(P, kr0[:, None], h0[:, None],
                            vg_a=np.zeros((2, 1)), vg_n=np.zeros((2, 1)),
                            model=np.array([[KR_LF], [KR_LF]]))
    if not np.array_equal(alllf, base):
        problems.append("kr_relative_vec(all-lf) != kr_frontal_vec")
    mixed = kr_relative_vec(P, kr0[:, None], h0[:, None],
                            vg_a=np.array([[0.075], [0.0]]), vg_n=np.array([[1.89], [0.0]]),
                            model=np.array([[KR_VG], [KR_LF]]))
    if not np.array_equal(mixed[1], base[1]):
        problems.append("mixed dispatch perturbed the lf row")
    if not np.allclose(mixed[0], kr_vg_vec(P[0], 0.075, 1.89)):
        problems.append("mixed dispatch wrong on the vG row")

    # Scalar kr_relative agrees with the vector path and dispatches on model.
    if abs(kr_relative(-1.5, 0.01, -1.0, model=KR_LF) - float(kr_frontal_vec(np.array(-1.5), 0.01, -1.0))) > 1e-12:
        problems.append("scalar kr_relative(lf) != kr_frontal")
    if abs(kr_relative(-1.5, 0.0, 0.0, vg_a=0.075, vg_n=1.89, model=KR_VG) - vg_scalar(-1.5, 0.075, 1.89)) > 1e-12:
        problems.append("scalar kr_relative(vg) != Mualem-vG")

    if problems:
        return None, "; ".join(problems[:5])
    return 0.0, None


def run_pinchout_lobes_test(test):
    """Guard the self-touching-ring repair in mesh._clean_pinchouts: a material zone
    that buffer(0) resolves into SEVERAL real lobes must keep every one of them.

    The geometry is the corpus case (Slide2 VP65/VP66): a sand layer whose top line
    dips to the base of the layer through a cutoff trench, so the ring touches itself
    at the trench and splits into two equal lobes. Keeping only the largest deleted
    the other half of the layer, and the mesh carried a full-depth void under the
    downstream half of the dam with no error and no warning.

    Checks (a) both lobes come back, under the same mat_id, with the ring's whole
    area, (b) they mesh as two regions covering that area, (c) an ordinary valid
    polygon is passed through untouched, and (d) a ring whose repair really does lose
    area raises instead of returning the survivor. Returns (0.0, None) on success,
    else (None, message). Builds its own geometry — no input file needed."""
    import numpy as np
    from shapely.geometry import Polygon
    from xslope.mesh import _clean_pinchouts, build_mesh_from_polygons

    # Layer between y = -5 and y = 0, pinched to zero thickness by a trench at x = 0.
    ring = [(-20.0, 0.0), (-4.0, 0.0), (-2.0, -5.0), (2.0, -5.0), (4.0, 0.0),
            (20.0, 0.0), (20.0, -5.0), (-20.0, -5.0)]
    ring_area = Polygon(ring).area
    square = [(0.0, 0.0), (10.0, 0.0), (10.0, 10.0), (0.0, 10.0)]

    problems = []
    cleaned = _clean_pinchouts([{'coords': ring, 'mat_id': 7},
                                {'coords': square, 'mat_id': 3}])
    lobes = [p for p in cleaned if p['mat_id'] == 7]
    if len(lobes) != 2:
        problems.append(f"trench-split layer came back as {len(lobes)} polygon(s), expected 2")
    kept = sum(Polygon(p['coords']).area for p in lobes)
    if abs(kept - ring_area) > 1e-9 * ring_area:
        problems.append(f"repair lost area: kept {kept:.4f} of {ring_area:.4f}")
    if any(not Polygon(p['coords']).is_valid for p in lobes):
        problems.append("a repaired lobe is not a valid polygon")
    passthrough = [p for p in cleaned if p['mat_id'] == 3]
    if len(passthrough) != 1 or list(passthrough[0]['coords']) != square:
        problems.append("a valid polygon was not passed through untouched")

    if len(lobes) == 2:
        mesh = build_mesh_from_polygons(lobes, 2.0, 'tri3')
        nodes = np.asarray(mesh['nodes'])
        el = np.asarray(mesh['elements']); et = np.asarray(mesh['element_types'])
        meshed = 0.0
        left = right = 0
        for ei in range(len(el)):
            tri = Polygon([nodes[nd] for nd in el[ei][:int(et[ei])]])
            meshed += tri.area
            if tri.centroid.x < 0:
                left += 1
            else:
                right += 1
        if abs(meshed - ring_area) > 1e-6 * ring_area:
            problems.append(f"meshed area {meshed:.4f} != layer area {ring_area:.4f}")
        if not left or not right:
            problems.append(f"only one side of the trench meshed (left={left}, right={right})")

    # A ring traced twice encloses half the area its coordinates claim: the repair
    # genuinely loses area, which is the silent-void signature and must raise.
    try:
        _clean_pinchouts([{'coords': square + square, 'mat_id': 1}])
        problems.append("area-losing repair did not raise")
    except ValueError:
        pass

    if problems:
        return None, "; ".join(problems[:5])
    return 0.0, None


def run_side_roller_test(test):
    """Guard the x-roller assignment in fem.build_fem_data on an off-vertical side.

    A far-field truncation boundary is often digitized a little off plumb (RS2-28's
    left face drifts 0.13 m over 30.7 m of height). Selecting side nodes by
    x == x_extreme rollered only the single extreme node and left the rest of that
    face traction-free, which cannot equilibrate at any strength-reduction factor
    once the seepage field prescribes uplift on it — and the face is part of the
    ground surface, so the bottom rule does not pick it up either.

    Builds that geometry (a 100 x 30 block whose left face runs from (0, 0) to
    (0.13, 30), with the face inside the ground-surface polyline) and checks that
    EVERY node on it is restrained, that exactly the face nodes are — no node one
    element in from it — and that the vertical right face is unchanged. Returns
    (0.0, None) on success, else (None, message)."""
    import copy
    import numpy as np
    from shapely.geometry import LineString, Point, Polygon
    from xslope.fem import build_fem_data
    from xslope.fileio import load_slope_data
    from xslope.mesh import build_mesh_from_polygons

    base = Path(__file__).parent / 'docs' / 'fem' / 'files' / 'xslope_griffiths1.xlsx'
    if not base.exists():
        return 0.0, None                      # engine-only clone without the docs files

    W, H, DRIFT = 100.0, 30.0, 0.13
    ring = [(0.0, 0.0), (DRIFT, H), (W, H), (W, 0.0)]
    face = LineString([(0.0, 0.0), (DRIFT, H)])

    d = copy.deepcopy(load_slope_data(str(base)))
    poly = Polygon(ring)
    d['profile_lines'] = []
    d['polygons'] = [{'polygon': poly, 'mat_id': 0}]
    d['domain_polygon'] = poly
    # The truncation face is part of the ground-surface polyline, as in rs2_28a —
    # that is what excludes it from the fixed-base polyline.
    d['ground_surface'] = LineString([(0.0, 0.0), (DRIFT, H), (W, H)])
    d['max_depth'] = 0.0
    d['circles'] = []; d['non_circ'] = []
    d['dloads'] = []; d['dloads2'] = []
    d['piezo_line'] = []; d['piezo_phreatic'] = False
    d['reinforce_lines'] = []; d['pile_lines'] = []

    mesh = build_mesh_from_polygons([{'coords': ring, 'mat_id': 0}], 5.0, 'tri3')
    bc = build_fem_data(d, mesh)['bc_type']
    nodes = np.asarray(mesh['nodes'])

    problems = []
    on_face = np.array([face.distance(Point(x, y)) < 1e-6 * W for x, y in nodes])
    free_on_face = int(((bc == 0) & on_face).sum())
    if free_on_face:
        problems.append(f"{free_on_face} of {int(on_face.sum())} nodes on the "
                        f"off-vertical left face left unrestrained")
    stray = int(((bc == 2) & ~on_face & (nodes[:, 0] < W - 1e-6)).sum())
    if stray:
        problems.append(f"{stray} node(s) off the side faces given x-rollers")
    right = np.abs(nodes[:, 0] - W) < 1e-6
    if int(((bc == 0) & right).sum()):
        problems.append("a node on the vertical right face lost its roller")

    if problems:
        return None, "; ".join(problems[:5])
    return 0.0, None


def run_mesh_conform_test(test):
    """Guard the conforming-edge preprocessing in build_mesh_from_polygons: a vertex
    in the interior of a neighbour's shared edge (a T-junction) must be inserted so
    the interface meshes without a slit. Builds the earth-dam shell/core case with a
    T-junction at (55,18) on the core-top interface and checks (a) the preprocessor
    inserts it into the core edge and (b) every interface node is shared by elements
    of BOTH materials. Returns (0.0, None) on success, else (None, message)."""
    import numpy as np
    from collections import defaultdict
    from xslope.mesh import build_mesh_from_polygons, make_polygons_conforming

    core = [(46.5, 0.0), (63.5, 0.0), (59.0, 18.0), (51.0, 18.0)]
    shell = [(0.0, 0.0), (51.0, 22.0), (59.0, 22.0), (110.0, 0.0), (63.5, 0.0),
             (59.0, 18.0), (55.0, 18.0), (51.0, 18.0), (46.5, 0.0)]   # T-junction (55,18)

    problems = []
    pc = [list(shell), list(core)]
    make_polygons_conforming(pc)
    if (55.0, 18.0) not in [tuple(p) for p in pc[1]]:
        problems.append("preprocessor did not insert the T-junction vertex (55,18) into the core edge")

    mesh = build_mesh_from_polygons([{'coords': shell, 'mat_id': 0},
                                     {'coords': core, 'mat_id': 1}], 3.0, 'tri3')
    nodes = np.asarray(mesh['nodes']); el = np.asarray(mesh['elements'])
    et = np.asarray(mesh['element_types']); em = np.asarray(mesh['element_materials'])
    node_mats = defaultdict(set)
    for ei in range(len(el)):
        for nd in el[ei][:int(et[ei])]:
            node_mats[int(nd)].add(int(em[ei]))
    nonconf = sum(1 for i, (x, y) in enumerate(nodes)
                  if abs(y - 18.0) < 1e-6 and 51.0 + 1e-6 < x < 59.0 - 1e-6
                  and len(node_mats.get(i, set())) < 2)
    if nonconf:
        problems.append(f"{nonconf} non-conforming interface node(s) on the core-top edge")

    if problems:
        return None, "; ".join(problems[:5])
    return 0.0, None


def run_tseep_exit_cycle_test(test):
    """Guard the transient exit-face anti-cycling rule (task #71).

    A quadratic (tri6/quad8/quad9) exit edge is tracked ALL-OR-NOTHING, so the
    state the face is genuinely in while the drainage front's exit point descends
    THROUGH that edge — partly wet — has no representation. Both available states
    are then inadmissible and the active set alternates between them forever: the
    Picard set-stability test can never close, dt collapses to its floor and the
    march force-accepts steps microseconds wide. On the RS2 VP102 drawdown march
    that hard-stalled at t = 320 h with 823 force-accepted steps, which is why the
    builder still runs it on a halved head-change limiter.

    Two checks, because the rule has to do exactly one thing and nothing else:

    * the RESOLUTION is the partly-wet state. Fed a cycling edge whose lower corner
      and midside pass their own SEEP2D h/q test and whose upper corner does not,
      _resolve_exit_cycle must pin wet-lower/dry-upper — the transition landing AT
      the midside node. Pinning the whole edge either way (the two states that were
      already available) is what the fix exists to stop.

    * it stays BEHIND A FAILED STEP. The fallback runs only on the retry of a step
      that has already failed with the set cycling, which is what keeps every march
      that converges the ordinary way on exactly the path it followed before the
      rule existed — and every locked transient trajectory with it. That gating is
      visible in the counters: a resolution is preceded by a detection on the failed
      sweep AND another on the retry, so ``2 * damped <= cycles`` whenever the rule
      engages at all. Resolving on first detection instead makes the two counts
      equal. A tri6 model that never cycles must report neither.

    Returns (0.0, None) on success, else (None, message)."""
    import io
    import contextlib
    import numpy as np
    from xslope.seep import (_resolve_exit_cycle, build_seep_data, build_tseep_data,
                             run_transient_seepage)

    problems = []

    # --- the resolution is the partly-wet state -------------------------------
    # Two quadratic exit edges sharing corner 2, stacked up the face: (0,1,2) is the
    # edge the exit point is crossing and (2,3,4) is the dry edge above it. The whole
    # lower edge is active now and inactive on the next sweep — the period-2 cycle.
    edges = [(0, 1, 2), (2, 3, 4)]
    n = 6
    linear = np.zeros(n, dtype=bool)
    act = np.array([1, 1, 1, 0, 0, 0], dtype=bool)
    new_act = np.zeros(n, dtype=bool)
    corner_candidate = np.array([1, 0, 0, 0, 0, 0], dtype=bool)
    mid_cands = [True, False]          # the midside is still wet, the one above is not
    pinned = np.zeros(n, dtype=bool)
    pinned_state = np.zeros(n, dtype=bool)
    _resolve_exit_cycle(act, new_act, corner_candidate, mid_cands, edges, linear,
                        pinned, pinned_state)
    got = tuple(bool(pinned_state[i]) if pinned[i] else None for i in range(5))
    want = (True, True, False, False, False)
    if got != want:
        problems.append(
            f"a cycling quadratic exit edge resolved to {got}, expected {want} — the "
            f"wet lower half with the seepage transition at the midside node, which is "
            f"the state the all-or-nothing rule cannot express")
    if pinned[5]:
        problems.append("a node on no exit edge was pinned")

    # An edge that is NOT oscillating must keep the all-or-nothing rule.
    pinned2 = np.zeros(n, dtype=bool)
    _resolve_exit_cycle(act, act.copy(), corner_candidate, mid_cands, edges, linear,
                        pinned2, np.zeros(n, dtype=bool))
    if pinned2.any():
        problems.append("a settled exit-face set was pinned — the rule must only "
                        "touch nodes that are actually oscillating")

    # --- the rule stays behind a failed step ----------------------------------
    from xslope.fileio import load_slope_data
    from xslope.mesh import (get_material_polygons, build_mesh_from_polygons,
                             extract_size_regions)
    here = os.path.dirname(os.path.abspath(__file__))

    def _march(rel, target_size):
        sd = load_slope_data(os.path.join(here, *rel))
        with contextlib.redirect_stdout(io.StringIO()):
            mesh = build_mesh_from_polygons(get_material_polygons(sd), target_size,
                                            'tri6',
                                            size_regions=extract_size_regions(sd))
            return run_transient_seepage(build_seep_data(mesh, sd),
                                         build_tseep_data(sd), verbose=False)

    # A march whose exit set never revisits a state must see neither counter move.
    quiet = _march(('docs', 'verification', 'files', 'geostudio', 'gs2_pond.xlsx'), 3.0)
    if not quiet.get('converged'):
        problems.append("the quiet tri6 march did not converge")
    if quiet.get('exit_face_cycles') or quiet.get('exit_face_damped'):
        problems.append(
            f"a march with no exit-face limit cycle reported "
            f"cycles={quiet.get('exit_face_cycles')}, "
            f"damped={quiet.get('exit_face_damped')} — the detector is firing on an "
            f"exit set that is merely settling")

    # A march that DOES cycle must resolve only on the retry of a failed step.
    cyc = _march(('docs', 'verification', 'files', 'rocscience_gw', 'gw018.xlsx'), 3.0)
    if not cyc.get('converged'):
        problems.append("the cycling tri6 march did not converge")
    n_cyc, n_damp = cyc.get('exit_face_cycles', 0), cyc.get('exit_face_damped', 0)
    if not n_damp:
        problems.append(f"the cycling fixture no longer exercises the rule "
                        f"(cycles={n_cyc}, damped={n_damp}) — it guards nothing")
    elif 2 * n_damp > n_cyc:
        problems.append(
            f"the anti-cycling rule resolved {n_damp} edge(s) against {n_cyc} detected "
            f"cycle(s): it is firing on FIRST detection instead of on the retry of a "
            f"step that has already failed, so every march that cycles and recovers by "
            f"itself — and every locked trajectory with it — moves")

    if problems:
        return None, "; ".join(problems[:4])
    return 0.0, None


def run_seep_exit_collapse_test(test):
    """Guard the tri6 exit-face active-set logic on a thin flat domain — the
    geometry class that exposed two coupled seepage defects (found during the
    GW8 build).

    Both checks run on a 1.0 x 0.5 m domain (target 0.25, so the full-height
    left-wall exit face is only two quadratic edges) meshed as tri3 and tri6:

    #51  Rain flux on the top (q = 4.4e-6 over L = 1, so Q = q*L is exact by
         construction) + the left-wall exit face + NO specified head. The tri3
         oracle drains through the bottom of the face and converges. Pre-fix,
         tri6 collapsed the WHOLE exit face to inactive — the edge-based
         active-set rule let the borderline transition corner veto the wet edge
         below it — leaving no Dirichlet node, so spsolve diverged to ~1e15 and
         converged=False. The fix must make tri6 CONVERGE to a finite field with
         at least one active exit node and the same construction-exact flowrate.

    #53  A deliberately SINGULAR flux-only + exit-face model: an extraction
         (outflow) flux with the same exit face and no specified head. The exit
         face legitimately empties (nothing to rescue) so the effective Dirichlet
         set is empty. Pre-fix that silently diverged to ~1e15; the runtime guard
         must now raise a clear SeepInputError naming the cause.

    Returns (0.0, None) on success, else (None, message). Builds its own mesh —
    no input file needed; gmsh is already a seep-suite dependency."""
    import io
    import contextlib
    import numpy as np
    from xslope.mesh import build_mesh_from_polygons
    from xslope.seep import solve_unsaturated, assemble_flux_nodal, SeepInputError

    poly = [(0.0, 0.0), (1.0, 0.0), (1.0, 0.5), (0.0, 0.5)]

    def _solve(element_type, flux):
        with contextlib.redirect_stdout(io.StringIO()):
            mesh = build_mesh_from_polygons([{'coords': poly, 'mat_id': 0}], 0.25, element_type)
            nodes = np.asarray(mesh['nodes'], dtype=float)
            x, y = nodes[:, 0], nodes[:, 1]
            bc_type = np.zeros(len(nodes), dtype=int)
            bc_values = np.zeros(len(nodes), dtype=float)
            left = np.abs(x) <= 1e-6            # full-height left-wall exit face
            bc_type[left] = 2
            bc_values[left] = y[left]
            fluxes = [{'coords': [(0.0, 0.5), (1.0, 0.5)], 'flux': flux}]   # top
            flux_nodal = assemble_flux_nodal(nodes, mesh['elements'],
                                             mesh['element_types'], fluxes, 1e-6)
            res = solve_unsaturated(
                nodes=nodes, elements=mesh['elements'], bc_type=bc_type,
                bc_values=bc_values, kr0=0.001, h0=-1.0, k1_vals=1.0, k2_vals=1.0,
                angles=0.0, element_types=mesh['element_types'], tol=1e-6,
                max_iter=400, closure_tol=1e-3, flux_nodal=flux_nodal)
        head, _A, _q, total_flow, exit_active, converged, _closure = res
        n_active = int(np.sum(exit_active & (bc_type == 2)))
        return head, total_flow, n_active, converged

    problems = []

    # ---- #51: thin-domain exit face must CONVERGE (tri3 is the oracle) --------
    q_rain = 4.4e-6
    try:
        h3, flow3, act3, conv3 = _solve('tri3', q_rain)
    except Exception as e:
        return None, f"#51 tri3 oracle unexpectedly failed: {type(e).__name__}: {e}"
    if not conv3:
        problems.append("#51 tri3 oracle did not converge")
    try:
        h6, flow6, act6, conv6 = _solve('tri6', q_rain)
    except Exception as e:
        problems.append(f"#51 tri6 raised instead of converging: {type(e).__name__}: {e}")
        conv6 = False; h6 = np.array([np.nan]); flow6 = 0.0; act6 = 0
    if not conv6:
        problems.append("#51 tri6 did not converge (exit face collapsed -> singular)")
    if not np.all(np.isfinite(h6)):
        problems.append("#51 tri6 head is not finite (diverged)")
    elif np.max(np.abs(h6)) > 1.0:
        problems.append(f"#51 tri6 head unphysical (max|h|={np.max(np.abs(h6)):.3e}, expected <1)")
    if act6 < 1:
        problems.append("#51 tri6 finished with no active exit-face node")
    # Flowrate is Q = q*L = 4.4e-6, exact by construction for every element type.
    if abs(flow6 - q_rain) > 1e-9 or abs(flow3 - q_rain) > 1e-9:
        problems.append(f"#51 flowrate off q*L: tri3={flow3:.3e}, tri6={flow6:.3e}, expected {q_rain:.3e}")

    # ---- #53: deliberately singular flux-only + exit face must RAISE ----------
    raised = False
    try:
        _solve('tri6', -q_rain)     # extraction: face empties, no head anchor
    except SeepInputError as e:
        raised = True
        msg = str(e).lower()
        if 'singular' not in msg or 'exit face' not in msg:
            problems.append(f"#53 raised SeepInputError but message unclear: {str(e)[:60]}")
    except Exception as e:
        problems.append(f"#53 raised {type(e).__name__}, expected SeepInputError")
    if not raised:
        problems.append("#53 singular flux-only+exit-face model did NOT raise (diverged silently)")

    if problems:
        return None, "; ".join(problems[:6])
    return 0.0, None


def run_test(test):
    """Run a single test. Returns ``(computed_value, error_msg, annotation)``.

    ``annotation`` is ``None`` for every test type except ``fem_ssrm``, whose
    two-tier fast-first-with-fallback runner returns a ``(bucket, text)`` routing
    note (see ``_run_fem_ssrm``). Keeping the annotation IN the return value —
    rather than mutating the test dict — is deliberate: parallel workers run under
    the ``spawn`` start method, so only the returned tuple crosses the process
    boundary back to the summary."""
    if test.get('type', '') == 'fem_ssrm':
        return _run_fem_ssrm(test)
    computed, error_msg = _dispatch_test(test)
    return computed, error_msg, None


def _dispatch_test(test):
    """Route a single test to its type-specific runner, returning
    ``(computed_value, error_msg)``. ``fem_ssrm`` is intercepted upstream in
    ``run_test`` by the two-tier kernel router, so the branch below is only the
    single-solve fallback for any direct caller."""
    test_type = test.get('type', '')
    if test_type == 'seep_exit_collapse':
        return run_seep_exit_collapse_test(test)
    if test_type == 'tseep_exit_cycle':
        return run_tseep_exit_cycle_test(test)
    if test_type == 'mesh_conform':
        return run_mesh_conform_test(test)
    if test_type == 'pinchout_lobes':
        return run_pinchout_lobes_test(test)
    if test_type == 'side_roller':
        return run_side_roller_test(test)
    if test_type == 'vg_kr':
        return run_vg_kr_test(test)
    if test_type == 'roundtrip':
        return run_roundtrip_test(test)
    if test_type == 'v19_roundtrip':
        return run_v19_roundtrip_test(test)
    if test_type == 'ssr_zone_roundtrip':
        return run_ssr_zone_roundtrip_test(test)
    if test_type == 'surface_family_roundtrip':
        return run_surface_family_roundtrip_test(test)
    if test_type == 'v21_roundtrip':
        return run_v21_roundtrip_test(test)
    if test_type == 'editor_roundtrip':
        return run_editor_roundtrip_test(test)
    if test_type == 'dxf':
        return run_dxf_roundtrip_test(test)
    if test_type == 'dxf_water':
        return run_dxf_water_test(test)
    if test_type == 'gsz':
        return run_gsz_import_test(test)
    if test_type == 'gsz_water':
        return run_gsz_water_corpus_test(test)
    if test_type == 'slide2':
        return run_slide2_import_test(test)
    if test_type == 'slide2_water':
        return run_slide2_water_corpus_test(test)
    if test_type == 'rs2':
        return run_rs2_import_test(test)
    if test_type == 'rs2_water':
        return run_rs2_water_mode_test(test)
    if test_type == 'rs2_loads':
        return run_rs2_load_deck_test(test)
    if test_type == 'submerged_oracle':
        return run_submerged_oracle_test(test)
    if test_type == 'no_void':
        return run_no_void_test(test)
    if test_type == 'suction_guard':
        return run_suction_guard_test(test)
    if test_type == 'piezo_u_guard':
        return run_piezo_u_guard_test(test)
    if test_type == 'kernel_xcheck':
        return run_kernel_xcheck_test(test)
    if test_type == 'preflight_rules':
        return run_preflight_rules_test(test)
    if test_type == 'preflight_corpus':
        return run_preflight_corpus_test(test)
    if test_type == 'preflight_contract':
        return run_preflight_contract_test(test)
    if test_type == 'sweep_gate':
        return run_sweep_gate_test(test)
    if test_type == 'preflight_remedies':
        return run_preflight_remedies_test(test)
    if test_type == 'generator_circles':
        return run_generator_circles_test(test)
    if test_type == 'auto_water':
        return run_auto_water_test(test)
    if test_type == 'template_sync':
        return run_template_sync_test(test)
    if test_type == 'deps_declared':
        return run_deps_declared_test(test)
    if test_type == 'v16_backcompat':
        return run_v16_backcompat_test(test)
    if test_type == 'fem_elastic_units':
        return run_fem_elastic_units_test(test)
    if test_type == 'dload_direction':
        return run_dload_direction_test(test)
    if test_type == 'k0_level_ground':
        return run_k0_level_ground_test(test)
    if test_type == 'stability_time':
        return run_stability_time_test(test)
    if test_type == 'steady_seep_save':
        return run_steady_seep_save_test(test)
    if test_type == 'noncircular_generator':
        return run_noncircular_generator_test(test)
    if test_type == 'updater':
        return run_updater_test(test)
    if test_type == 'assistant_models':
        return run_assistant_models_test(test)
    if test_type == 'quad_mesh':
        return run_quad_mesh_test(test)
    if test_type == 'quad_style_dialog':
        return run_quad_style_dialog_test(test)
    if test_type == 'fem_1d_details':
        return run_fem_1d_details_test(test)
    if test_type == 'mode_segments':
        return run_mode_segments_test(test)
    if test_type == 'refine_thin_zones':
        return run_refine_thin_zones_test(test)
    if test_type == 'remedy_panel':
        return run_remedy_panel_test(test)
    if test_type == 'polygon_pick':
        return run_polygon_pick_test(test)
    if test_type == 'transient_seep':
        return run_transient_seep_test(test)
    if test_type == 'fs_vs_time_mode':
        return run_fs_vs_time_mode_test(test)
    if test_type == 'sweep_window':
        return run_sweep_window_test(test)
    if test_type == 'water_hoist':
        return run_water_hoist_test(test)
    if test_type == 'mesh_elements':
        return run_mesh_elements_test(test)
    if test_type == 'cwd_invariant':
        return run_cwd_invariant_test(test)
    if test_type == 'docs_heading_trap':
        return run_docs_heading_trap_test(test)
    if test_type == 'verification_pages':
        return run_verification_pages_test(test)
    if test_type == 'corpus_index':
        return run_corpus_index_test(test)
    if test_type == 'gsat_pair':
        return run_gsat_pair_test(test)
    if test_type == 'axial_mirror':
        return run_axial_mirror_test(test)
    if test_type == 'mp_spencer':
        return run_mp_spencer_test(test)
    if test_type == 'drawdown_tauff':
        return run_drawdown_tauff_test(test)
    if test_type == 'drawdown_guard':
        return run_drawdown_guard_test(test)
    if test_type == 'fem_ssrm':
        # Unreachable in practice (run_test intercepts fem_ssrm upstream), but the
        # kernel is pinned explicitly here too: no path in this suite may inherit
        # solve_fem's 'auto' default, because the lock is a property of the
        # reference path.
        return run_fem_test(test, fast_kernel=False)
    if test_type == 'seep_elements':
        return run_seep_elements_test(test)
    if test_type == 'fem_elements':
        return run_fem_elements_test(test)
    if test_type == 'fem_reliability':
        return run_fem_reliability_test(test)
    elif test_type == 'seep':
        return run_seep_test(test)
    elif test_type == 'seep_head':
        return run_seep_head_test(test)
    elif test_type == 'tseep_head':
        return run_tseep_head_test(test)
    elif test_type == 'fs_vs_time':
        return run_fs_vs_time_test(test)
    elif test_type == 'reliability':
        return run_reliability_test(test)
    elif test_type == 'reliability_mc':
        return run_reliability_mc_test(test)
    elif test_type == 'design_search':
        return run_design_test(test)
    elif test_type == 'design_callable':
        return run_design_callable_test(test)
    elif test_type == 'critical_kc':
        return run_critical_kc_test(test)
    elif test_type == 'sensitivity':
        return run_sensitivity_test(test)
    else:
        return run_lem_test(test)


def _expected_and_tol(test, default_tolerance):
    """Expected value and tolerance for a test dict (shared by both runners)."""
    test_type = test.get('type', '?')
    if test_type == 'seep':
        expected = test.get('expected_flowrate')
        tol = test.get('tolerance', 0.05) * abs(expected) if expected else 0
    elif test_type in ('reliability', 'fem_reliability', 'reliability_mc'):
        expected = test.get('expected_beta')
        tol = test.get('tolerance', 0.02)
    elif test_type == 'critical_kc':
        expected = test.get('expected_kc')
        tol = test.get('kc_tol', 0.01)
    elif test_type == 'sensitivity':
        # the runner checks base/low/high internally; the framework-level
        # comparison re-checks the base row
        expected = float(test['expected_base']) if 'expected_base' in test else None
        tol = float(test.get('tolerance', 0.01))
    elif test_type in ('preflight_rules', 'preflight_corpus', 'preflight_contract',
                       'preflight_remedies', 'generator_circles', 'auto_water',
                       'sweep_gate', 'steady_seep_save',
                       'roundtrip', 'v19_roundtrip', 'ssr_zone_roundtrip', 'v21_roundtrip', 'surface_family_roundtrip', 'editor_roundtrip', 'template_sync', 'deps_declared', 'v16_backcompat', 'fem_elastic_units', 'dload_direction', 'k0_level_ground', 'stability_time', 'docs_heading_trap', 'cwd_invariant', 'mesh_elements', 'verification_pages', 'corpus_index', 'dxf', 'dxf_water', 'gsz', 'gsz_water', 'slide2', 'slide2_water', 'rs2', 'rs2_water', 'rs2_loads', 'vg_kr',
                       'mesh_conform', 'pinchout_lobes', 'quad_mesh', 'side_roller',
                       'quad_style_dialog', 'mode_segments',
                       'refine_thin_zones', 'remedy_panel',
                       'polygon_pick', 'transient_seep',
                       'fs_vs_time_mode', 'sweep_window', 'water_hoist',
                       'noncircular_generator', 'updater', 'fem_1d_details',
                       'assistant_models',
                       'fs_vs_time',
                       'seep_elements', 'seep_exit_collapse', 'tseep_exit_cycle',
                       'fem_elements',
                       'mp_spencer', 'axial_mirror', 'drawdown_tauff', 'drawdown_guard',
                       'submerged_oracle', 'no_void', 'suction_guard', 'piezo_u_guard',
                       'gsat_pair', 'seep_head',
                       'tseep_head', 'design_callable', 'kernel_xcheck'):
        expected = 0.0          # these return 0.0 on success (pass/fail tests)
        tol = 0.0
    else:
        expected = test.get('expected_fs')
        tol = test.get('tolerance', default_tolerance)
    return expected, tol


# Rough per-type cost ranks so the parallel scheduler starts the slow tests
# first (wall time is otherwise dominated by an FEM case landing last).
_COST_RANK = {'fem_reliability': 6, 'reliability_mc': 6, 'fem_ssrm': 5, 'fem_elements': 5,
              'preflight_corpus': 5, 'preflight_rules': 4,
              'reliability': 4, 'critical_kc': 4, 'tseep_head': 4, 'fs_vs_time': 5,
              'transient_seep': 4, 'seep_elements': 3, 'seep': 3,
              'noncircular_search': 2, 'circular_search': 2}


def _parallel_worker(item):
    """Run one test in a worker process, stdout/stderr suppressed."""
    i, test = item
    import io as _io
    import contextlib as _ctx
    import warnings as _w
    _w.filterwarnings('ignore')
    t0 = time.time()
    buf = _io.StringIO()
    annotation = None
    try:
        with _ctx.redirect_stdout(buf), _ctx.redirect_stderr(buf):
            computed, error_msg, annotation = run_test(test)
    except Exception as e:
        computed, error_msg = None, str(e)
    return i, computed, error_msg, annotation, time.time() - t0


def main():
    parser = argparse.ArgumentParser(description='xslope regression test suite')
    parser.add_argument('--lem', action='store_true', help='Run only LEM tests')
    parser.add_argument('--fem', action='store_true', help='Run only FEM tests')
    parser.add_argument('--seep', action='store_true', help='Run only (steady) seepage tests')
    parser.add_argument('--tseep', action='store_true',
                        help='Run only transient-seepage tests (type=tseep_head)')
    parser.add_argument('--roundtrip', action='store_true',
                        help='Run only the Excel save/load round-trip tests')
    parser.add_argument('--dxf', action='store_true',
                        help='Run only the structured DXF export/import round-trip '
                             'tests (needs ezdxf + PySide6)')
    parser.add_argument('--gsz', action='store_true',
                        help='Run only the GeoStudio (.gsz) import test')
    parser.add_argument('--slide2', action='store_true',
                        help='Run only the Slide2 (.slim/.slmd) import test')
    parser.add_argument('--rs2', action='store_true',
                        help='Run only the RS2 (.fez) import test')
    parser.add_argument('--preflight', action='store_true',
                        help='Run only the preflight (input dependency check) tests')
    parser.add_argument('--tolerance', type=float, default=0.01,
                        help='Default tolerance for FS comparison (default: 0.01)')
    parser.add_argument('--verbose', action='store_true',
                        help='Print details for passing tests')
    parser.add_argument('--benchmark', default=None,
                        help='Run only tests whose benchmark id matches the '
                             'given comma-separated list; each entry matches '
                             'exactly or as a family prefix (VP28 runs VP28a, '
                             'VP28a-beta, ... but not VP280). For tag-only or '
                             'builder-only changes, run just the new tags '
                             'instead of the whole suite.')
    parser.add_argument('--skip-benchmarks', action='store_true',
                        help='Skip verification benchmark tests (annotations '
                             'carrying a benchmark=<ID> tag), e.g. the slow '
                             'SSRM dam cases')
    parser.add_argument('--jobs', default='auto',
                        help="Parallel worker processes: an integer or 'auto' "
                             "(cpu_count - 2, the default). Use --jobs 1 for "
                             "sequential runs (in-order rows, full solver "
                             "output — the debugging mode).")
    parser.add_argument('--quick', action='store_true',
                        help='For LEM problems that list several methods, check '
                             'only one method per problem (prefers Spencer) so '
                             'routine runs stay fast')
    parser.add_argument('--reference-only', action='store_true',
                        help='Verify every FEM SSRM row on the pure reference '
                             'kernel only, disabling the fast-first-with-fallback '
                             'path. Use for strict runs: pre-release, or right '
                             'after a constitutive-physics change, when the oracle '
                             'should speak directly (see run_tests._run_fem_ssrm).')
    parser.add_argument('--mesh', action='store_true',
                        help='Run only the mesh-size locks (type=mesh_elements): '
                             'the element and node counts the verification pages '
                             'print. A mesh build each, no solve — seconds.')
    parser.add_argument('--list', action='store_true',
                        help='Print the discovered test count and a per-source '
                             'breakdown, then exit without running anything. '
                             'Discovery reads files and parses tags only, so this '
                             'costs a second; it is what makes the suite\'s '
                             'independence from the working directory checkable.')
    args = parser.parse_args()

    # If no specific flags, run all
    run_all = not (args.lem or args.fem or args.seep or args.tseep or args.roundtrip
                   or args.dxf or args.gsz or args.slide2 or args.rs2
                   or args.preflight or args.mesh)
    run_lem = args.lem or run_all
    run_fem = args.fem or run_all
    run_seep = args.seep or run_all
    run_tseep = args.tseep or run_all
    run_roundtrip = args.roundtrip or run_all
    run_dxf = args.dxf or run_all
    run_gsz = args.gsz or run_all
    run_slide2 = args.slide2 or run_all
    run_rs2 = args.rs2 or run_all
    run_preflight = args.preflight or run_all
    run_mesh = args.mesh or run_all

    # Discover tests from markdown files
    tests = []
    if run_lem:
        lem_samples = Path(_repo('docs/lem/samples.md'))
        if lem_samples.exists():
            tests.extend(parse_test_tags(lem_samples))
        # The design example carries a design_search regression (guards the
        # profile-edit / polygon-rebuild bug found in main_design.py).
        lem_design = Path(_repo('docs/lem/design.md'))
        if lem_design.exists():
            tests.extend(parse_test_tags(lem_design))
        # Parametric studies (sensitivity/design/back-analysis) live under
        # docs/parametric/ as three pages; scan all of them for tags.
        for parametric_md in sorted(Path(_repo('docs/parametric')).glob('*.md')):
            tests.extend(parse_test_tags(parametric_md))
        # File-less unit guard for the design()/back_analysis() modify= callable
        # path: param-vs-equivalent-callable equivalence plus a hand-computable
        # (OMS fixed-surface) crossing. A callable cannot be written in a doc tag,
        # so this check is registered directly rather than parsed from markdown.
        tests.append({'type': 'design_callable', 'file': 'design modify= (unit)',
                      'method': '-', 'source': 'design_callable'})

    if run_fem:
        fem_samples = Path(_repo('docs/fem/samples.md'))
        if fem_samples.exists():
            tests.extend(parse_test_tags(fem_samples))
        # Side-roller assignment on an off-vertical truncation face.
        tests.append({'type': 'side_roller', 'file': 'off-vertical side face (fem)',
                      'method': '-', 'source': 'side_roller'})
        # The 1D details dialog: gating, envelope sharing, reload, export.
        tests.append({'type': 'fem_1d_details',
                      'file': 'FEM 1D solution details (Studio dialog)',
                      'method': '-', 'source': 'fem_1d_details'})
        # Fast-kernel divergence fence: only meaningful when the optional compiled
        # Mohr-Coulomb kernel is built (setup_kernel.py). It cross-checks the
        # fast path against the NumPy oracle (bit-identical FS + field max-diff
        # < 1e-8) on three small coarse cases. Skipped with a notice when the
        # kernel isn't built — the NumPy path needs no cross-check against itself.
        try:
            from xslope import _fem_kernel  # noqa: F401
            tests.append({'type': 'kernel_xcheck',
                          'file': 'fast-kernel divergence fence',
                          'method': '-', 'source': 'kernel_xcheck'})
            if not run_all:
                print("Including fast-kernel cross-check gate")
        except ImportError:
            print("Skipping fast-kernel cross-check (compiled xslope._fem_kernel "
                  "not built; opt-in, see setup_kernel.py)")
    # Verification corpus pages (docs/verification/*.md): tags are routed by
    # type so LEM, SSRM, and seepage assertions can live on the same page.
    for verification_md in sorted(Path(_repo('docs/verification')).glob('*.md')):
        for t in parse_test_tags(verification_md):
            ttype = t.get('type', '')
            if ttype in ('fem_ssrm', 'fem_elements', 'fem_reliability'):
                if run_fem:
                    tests.append(t)
            elif ttype == 'mesh_elements':
                # A mesh build and a count, no solve: seconds, so it rides the
                # DEFAULT set rather than the minutes-per-row --fem group the rest
                # of its file's tags belong to.
                if run_mesh:
                    tests.append(t)
            elif ttype in ('tseep_head', 'fs_vs_time'):
                # fs_vs_time is a stability lock riding a transient march: the
                # march dominates its cost and it must re-solve the same field the
                # tseep_head rows on that page lock, so it belongs with them.
                if run_tseep:
                    tests.append(t)
            elif ttype in ('seep', 'seep_elements', 'seep_head'):
                if run_seep:
                    tests.append(t)
            elif run_lem:
                tests.append(t)

    if run_seep:
        # Fast, file-less unit check for the van Genuchten kr function + dispatch.
        tests.append({'type': 'vg_kr', 'file': 'kr_vg_vec (unit)', 'method': '-',
                      'source': 'vg_kr'})
        # Mesh conforming-edge (T-junction) regression.
        tests.append({'type': 'mesh_conform', 'file': 'conforming edges (mesh)',
                      'method': '-', 'source': 'mesh_conform'})
        # Multi-lobe material zone (trench-split layer) must survive the
        # self-touching-ring repair instead of meshing as a void.
        tests.append({'type': 'pinchout_lobes', 'file': 'trench-split zone (mesh)',
                      'method': '-', 'source': 'pinchout_lobes'})
        # The mesh builder: free-style element quality and DELIVERED size on seven
        # corpus sections, determinism in both styles, structured sweeps on the
        # levee, the setTransfiniteCurve node-count off-by-one, and what the size
        # field delivers — boundary spacing, a refined feature tip, a reinforcement
        # line, the crack-tip detector, and a wedge tip against the limit its own
        # apex angle sets. A mesh that comes back at a size nobody asked for is a
        # silent discretization change, and a sweep one node off looks right and
        # is not.
        tests.append({'type': 'quad_mesh',
                      'file': 'mesh quality, size field + structured sweeps',
                      'method': '-', 'source': 'quad_mesh'})
        # Thin-domain tri6 exit-face regression: must converge (#51) and the
        # singular flux-only+exit-face model must raise a clear error (#53).
        tests.append({'type': 'seep_exit_collapse',
                      'file': 'tri6 thin-domain exit face (#51 #53)',
                      'method': '-', 'source': 'seep_exit_collapse'})

    if run_tseep:
        # Transient exit-face anti-cycling: the partly-wet resolution of a cycling
        # quadratic exit edge, and that the rule stays inert on a healthy march.
        tests.append({'type': 'tseep_exit_cycle',
                      'file': 'tri6 exit-edge limit cycle (transient)',
                      'method': '-', 'source': 'tseep_exit_cycle'})
        # The transient solver's own lock set (test/transient_seep_check.py): 12
        # locks on the stepper and the storage terms. It solves several short
        # marches, which makes it the heaviest single row in this group (~60 s).
        tests.append({'type': 'transient_seep',
                      'file': 'transient seepage solver (12 locks)',
                      'method': '-', 'source': 'transient_seep'})

    # docs/seep/samples.md carries the steady seep sample locks AND (Problems 8-9)
    # the transient tseep_head locks — route each by type so --tseep picks up the
    # transient sample locks and --seep the steady ones (mirrors the verification
    # pages and seep_slope.md).
    seep_samples = Path(_repo('docs/seep/samples.md'))
    if seep_samples.exists():
        for t in parse_test_tags(seep_samples):
            if t.get('type', '') == 'tseep_head':
                if run_tseep:
                    tests.append(t)
            elif run_seep:
                tests.append(t)

    # docs/seep/seep_slope.md mixes seepage, LEM, and FEM tests (the combined
    # Johnson Reservoir example) — always scan it and route each test by type.
    seep_slope = Path(_repo('docs/seep/seep_slope.md'))
    if seep_slope.exists():
        for t in parse_test_tags(seep_slope):
            ttype = t.get('type', '')
            if ttype == 'fem_ssrm':
                if run_fem:
                    tests.append(t)
            elif ttype == 'seep':
                if run_seep:
                    tests.append(t)
            elif run_lem:
                tests.append(t)

    # Private test problems (CE 544 homework/exam keys, plus synthetic variants)
    # live in a separate repo kept out of the public xslope tree. Look for it as a
    # sibling directory or at $XSLOPE_PRIVATE_TESTS; scan its markdown files for
    # test tags and route by type. Silently skipped when the repo is absent
    # (public CI, other clones), so the public suite is unaffected.
    private_dir = _private_dir()
    # No env var and no sibling repo means there are no private tests — say that with
    # None, not with an empty path. Path('') is Path('.'), which is a directory, so an
    # empty fallthrough turned the private scan loose on the public tree itself and
    # every docs/ fixture was collected a second time (26 duplicate rows).
    private_path = Path(private_dir) if private_dir else None
    if private_path is not None and private_path.is_dir():
        n_priv = 0
        # Fixtures live under tests/. Scope the walk there rather than over the whole
        # repo: sibling trees (plans/, ref_docs/) hold prose that quotes the
        # `<!-- test: ... -->` syntax as documentation, and parsing those yields
        # garbage entries with no 'file' key.
        scan_root = private_path / 'tests'
        if not scan_root.is_dir():
            scan_root = private_path
        for md in sorted(scan_root.rglob('*.md')):
            if md.name.lower() == 'readme.md':
                continue
            if 'ref_docs' in md.relative_to(private_path).parts:
                continue
            for t in parse_test_tags(md):
                if not t.get('file'):
                    continue    # a prose example of the tag syntax, not a fixture
                ttype = t.get('type', '')
                if ttype == 'fem_ssrm':
                    if run_fem:
                        tests.append(t); n_priv += 1
                elif ttype == 'seep':
                    if run_seep:
                        tests.append(t); n_priv += 1
                elif run_lem:
                    tests.append(t); n_priv += 1
        if n_priv:
            print(f"Including {n_priv} private tests from {private_path.name}/")

    # Spencer is the f(x) == 1 special case of Morgenstern-Price, so the two must
    # agree to machine precision. Asserted on BOTH facings: the right-facing mirror
    # in _mp_extract is otherwise only exercised implicitly.
    if run_lem:
        for _fp, _facing in ((MP_SPENCER_LEFT, 'left'), (MP_SPENCER_RIGHT, 'right')):
            if Path(_fp).exists():
                tests.append({'type': 'mp_spencer', 'file': _fp, 'facing': _facing,
                              'method': 'mprice=spencer', 'source': f'{_facing}-facing'})

        # Axial reinforcement facing-invariance: nailed wall vs its exact mirror
        if Path(AXIAL_MIRROR_LEFT).exists() and Path(AXIAL_MIRROR_RIGHT).exists():
            tests.append({'type': 'axial_mirror', 'file': AXIAL_MIRROR_LEFT,
                          'file2': AXIAL_MIRROR_RIGHT, 'method': 'all-methods L==R',
                          'source': 'axial_mirror'})

        # Rapid-drawdown Stage-2 strength pipeline, against the worked example in
        # Duncan/Wright/Brandon Table 9.2. File-less: an infinite slope, hand-computable.
        tests.append({'type': 'drawdown_tauff', 'file': 'DWB Table 9.2 (unit)',
                      'method': 'tau_ff', 'source': 'drawdown_tauff'})
        # And the guard that stops eq (5) extrapolating past Kc=Kf when Stage-1 FS < 1.
        if Path(DRAWDOWN_GUARD_FILE).exists():
            tests.append({'type': 'drawdown_guard', 'file': DRAWDOWN_GUARD_FILE,
                          'method': 'stage1 FS>=1', 'source': 'drawdown_guard'})
        # Dry-buoyant still-water oracle: the exact effective-stress equivalence that
        # validates the perpendicular-dload water convention the vendor importers convert
        # ponded water into. File-less (builds on xslope_acads_simple).
        tests.append({'type': 'submerged_oracle',
                      'file': 'dry-buoyant still-water oracle',
                      'method': 'bishop/spencer', 'source': 'submerged_oracle'})
        # No-void tiling guard: every corpus file's material zones must tile its
        # domain with no interior material void (the VP42 toe-wedge class). Sweeps
        # the corpus itself; file-less.
        tests.append({'type': 'no_void',
                      'file': 'corpus material-tiling void audit',
                      'method': '-', 'source': 'no_void'})
        # Matric-suction apparent-cohesion guard: the opt-in Fredlund extended
        # Mohr-Coulomb option is off-by-default bit-identical, raises FS for
        # phi_b>0, respects suction_cap, and prices c_suction exactly. File-less
        # (builds on xslope_acads_simple).
        tests.append({'type': 'suction_guard',
                      'file': 'matric-suction apparent-cohesion guard',
                      'method': '-', 'source': 'suction_guard'})
        # Silent-zero pore pressure: an unrecognized mat-sheet `u` option, u='piezo'
        # with no piezometric line at all, and a slice base / mesh node / Gauss
        # point that samples one from outside its own x-extent. All must raise, not
        # quietly read u = 0; a line that stops short but still covers what samples
        # it must solve unchanged. File-less (builds on xslope_acads_simple).
        tests.append({'type': 'piezo_u_guard',
                      'file': 'silent-zero pore-pressure guards',
                      'method': '-', 'source': 'piezo_u_guard'})

    # Preflight (xslope.preflight) — the rule registry's own regression family.
    # The contract and mutation checks are file-less; the corpus check is one row
    # per markdown source, so the parallel scheduler can spread the workbook loads.
    if run_preflight:
        tests.append({'type': 'preflight_contract', 'file': '(rule registry)',
                      'method': '-', 'source': 'preflight'})
        tests.append({'type': 'preflight_rules', 'file': '(one mutation per rule)',
                      'method': '-', 'source': 'preflight'})
        # The two-stage sweep contract: a full base-model preflight once, a
        # per-step re-check of only the rules the substituted value touches, a
        # skipped step that states its reason, and the post-step sentinel screen.
        tests.append({'type': 'sweep_gate', 'file': '(the sweep gate)',
                      'method': '-', 'source': 'preflight'})
        # The remedies and the starting-surface generator ride with preflight: a
        # remedy is a rule's own offer, and the generator is one of them.
        tests.append({'type': 'preflight_remedies', 'file': '(the remedy contract)',
                      'method': '-', 'source': 'preflight'})
        tests.append({'type': 'generator_circles', 'file': '(starting circles)',
                      'method': '-', 'source': 'preflight'})
        # Automatic water loads ride with preflight too: the mode is a preflight
        # concern (which rules apply), and the remedy that flips it is one of these.
        tests.append({'type': 'auto_water', 'file': '(automatic water loads)',
                      'method': '-', 'source': 'preflight'})
        # A steady seepage solution the session has just computed is a
        # pore-pressure field the run can have, and the gate must see it without a
        # sidecar being reloaded first. Rides with preflight because that is the
        # half that was reported; the save-path half rides along because it is the
        # same session and the same check module.
        tests.append({'type': 'steady_seep_save', 'file': '(steady field + save path)',
                      'method': '-', 'source': 'preflight'})
        _preflight_sources = ([
            _repo('docs/lem/samples.md'), _repo('docs/lem/design.md'),
            _repo('docs/parametric/sensitivity.md'), _repo('docs/parametric/reliability.md'),
            _repo('docs/fem/samples.md'), _repo('docs/seep/samples.md'),
            _repo('docs/seep/seep_slope.md'),
        ] + sorted(glob.glob(_repo('docs/verification/*.md'))))
        # The private fixtures are corpus too: they carry the same tags and the same
        # standing locks, so a rule that refuses one of them is equally miscalibrated.
        _pf_priv = _private_dir()
        if _pf_priv and (Path(_pf_priv) / 'tests').is_dir():
            _preflight_sources += [str(m) for m in
                                   sorted((Path(_pf_priv) / 'tests').rglob('*.md'))
                                   if m.name.lower() != 'readme.md']
        for _src in _preflight_sources:
            if not os.path.exists(_src):
                continue
            # Skip a markdown file that carries no runnable tags (the verification
            # index and reference pages) — there is nothing for it to check.
            if not any(t.get('type', 'single_circle') in PREFLIGHT_TAG_ANALYSIS
                       and str(t.get('file', '')).endswith('.xlsx')
                       for t in parse_test_tags(_src)):
                continue
            tests.append({'type': 'preflight_corpus', 'file': _src,
                          'method': '-', 'source': 'preflight'})

    # Excel round-trip tests (save_slope_data_to_xlsx). Built from a curated file
    # list rather than markdown tags, since they check load/save fidelity, not FS.
    if run_roundtrip:
        # Guard that the packaged copies (shipped in the wheel) haven't drifted
        # from their editable docs masters.
        tests.append({'type': 'template_sync', 'file': BUNDLED_TEMPLATE,
                      'method': '-', 'source': 'template'})
        # Guard that every module-scope third-party import is a declared runtime dep,
        # so `pip install xslope` yields a package that actually imports.
        tests.append({'type': 'deps_declared', 'file': 'pyproject.toml',
                      'method': '-', 'source': 'packaging'})
        # Guard that pre-v16 files still load identically after the v16 mat-sheet
        # reshuffle (E/nu moved, t_cut inserted) — the column move is header-name-
        # keyed and so must be transparent to old files (plus a new t_cut=None).
        tests.append({'type': 'v16_backcompat', 'file': '(pre-v16 corpus files)',
                      'method': '-', 'source': 'v16_backcompat'})
        # Guard that no English-unit FEM corpus file carries the metric inherited
        # elastic default (E=100,000 psf ~ 4.8 MPa, ~10x too soft) — the unit-blind
        # bug that corrupts displacements and the RS2 displacement-vector panels.
        tests.append({'type': 'fem_elastic_units', 'file': 'docs/verification/files (FEM corpus)',
                      'method': '-', 'source': 'fem_units'})
        # Guard that no corpus distributed-load line comes back from the loader
        # running right-to-left. Point order silently reversed the reservoir
        # pressure on RS2-67 b/e/f (applied as uplift); the loader now orients
        # every monotone line to increasing X, and this is the standing check.
        tests.append({'type': 'dload_direction', 'file': 'corpus distributed-load orientation',
                      'method': '-', 'source': 'dload_direction'})
        # Guard that the K0 initial stress stays an exact equilibrium on level
        # ground — the one configuration where the answer is known in closed form —
        # both as built and after the SSRM's in-situ equilibration, which must be a
        # no-op there. File-less (builds a 20 x 10 m block).
        tests.append({'type': 'k0_level_ground', 'file': 'K0 level-ground equilibrium',
                      'method': '-', 'source': 'k0_level_ground'})
        # Guard the transient stability TIME: the precedence a run applies, the
        # tseep sheet's round trip of stability_time, the saved-frame schedule, the
        # two Run dialogs' selector, and the refusal of stage times with no frames.
        # It rides with the round-trip group rather than --tseep because it solves
        # no seepage at all — it exercises the reader, the writer and the dialogs.
        tests.append({'type': 'stability_time', 'file': 'transient stability time',
                      'method': '-', 'source': 'stability_time'})
        # Guard the FS-versus-time mode's row contract: every requested instant
        # appears, an instant that produced no result carries its reason, the
        # sentinel screen, one re-march for the whole set, and the base-model gate
        # before the first solve. Rides here rather than with --tseep because it
        # solves no seepage — it feeds the mode synthetic frame ledgers.
        tests.append({'type': 'fs_vs_time_mode', 'file': 'FS-vs-time mode contract',
                      'method': '-', 'source': 'fs_vs_time_mode'})
        # Guard the search window a sweep searches inside: read from the model by
        # default so every point of a curve measures one mechanism, and read the
        # SAME way by the sweeps, both reliability engines and Studio's Run LEM
        # path. Rides here rather than with --lem because it runs no search at all
        # — the search functions are recorders, and what is under test is which
        # constraints reach them.
        tests.append({'type': 'sweep_window', 'file': "a sweep's search window",
                      'method': '-', 'source': 'sweep_window'})
        # Guard the automatic water load's derivation point: once at the entry to a
        # search rather than once per trial surface, the identical slice table and
        # factor of safety either side of the hoist, and no reuse ACROSS runs — an
        # edited water definition must be re-derived, or a second run answers with
        # the first run's reservoir.
        tests.append({'type': 'water_hoist',
                      'file': 'water load derived once per search',
                      'method': '-', 'source': 'water_hoist'})
        # Guard the Build-mesh dialog's quadrilateral style radio group: the choice
        # is per-run and stored in no file, so a broken wire between the group and
        # build_mesh_from_polygons leaves no trace in any input file. Builds no mesh.
        tests.append({'type': 'quad_style_dialog',
                      'file': 'quad mesh style (Studio dialog)',
                      'method': '-', 'source': 'quad_style_dialog'})
        # Guard the toolbar's analysis-mode strip: the segments, the exclusivity,
        # the Ctrl+N shortcuts, and every side effect a click has to reproduce —
        # measured against the mode handler itself, so a strip wired to nothing
        # fails here rather than in a user's session. Runs no analysis.
        tests.append({'type': 'mode_segments',
                      'file': 'analysis-mode switch (Studio toolbar)',
                      'method': '-', 'source': 'mode_segments'})
        # Guard the Build-mesh dialog's thin-zone refinement toggle. Same reason as
        # the style group — a per-run choice no input file records — with one more:
        # what it promises is a RESOLUTION, and a mechanism that stopped delivering
        # it would leave every argument looking right. So this one builds four small
        # meshes and measures the element rows across a real thin band.
        tests.append({'type': 'refine_thin_zones',
                      'file': 'thin-zone refinement (Studio dialog)',
                      'method': '-', 'source': 'refine_thin_zones'})
        # Guard what a double-click inside a material zone resolves to: the polygon
        # row it opens the editor at, the precedence that keeps lines, points and
        # vertices ahead of the interior, and the smallest containing zone where
        # they nest. A wrong row is silent — a real editor on the wrong geometry.
        tests.append({'type': 'polygon_pick',
                      'file': 'polygon interior picking (Studio)',
                      'method': '-', 'source': 'polygon_pick'})
        # Guard the model-checks panel where a rule offers more than one repair: a
        # panel that renders only the primary leaves the second one unreachable and
        # says nothing about it. Rides with the Studio Qt rows rather than with
        # --preflight because it opens real panels; the remedies' own behaviour is
        # checked by the preflight_remedies row.
        tests.append({'type': 'remedy_panel',
                      'file': 'multi-remedy checks panel (Studio)',
                      'method': '-', 'source': 'remedy_panel'})
        # Guard the non-circular starting-surface generator: the mobilisable-strength
        # metric it ranks zones on, the separation threshold that decides whether it
        # picks or raises the zone picker, the geometry invariants that make the
        # surface usable (explicit Y and Movement on every point), and the Studio
        # button. Rides here rather than with --lem because it searches nothing — it
        # builds seeds and slices one of them.
        tests.append({'type': 'noncircular_generator',
                      'file': 'non-circular starting surface',
                      'method': '-', 'source': 'noncircular_generator'})
        # Guard Studio's in-app updater: the version comparison, the platform
        # artifact key, the minimum_version gate, the checksum refusal, and the
        # command each platform branch would spawn. Touches no network — the
        # manifest and the artifact are served over file:// — and runs no installer.
        tests.append({'type': 'updater', 'file': 'Studio updater',
                      'method': '-', 'source': 'updater'})
        # Guard where the assistant's model list comes from: the provider's own
        # list-models endpoint, the cached copy, the shipped fallback, and the
        # curated recommendations overlaid on whichever is showing. Same reason
        # the updater rides here and touches no network — every provider response
        # and the manifest are served over file://.
        tests.append({'type': 'assistant_models', 'file': 'assistant model list',
                      'method': '-', 'source': 'assistant_models'})
        # Guard against the Markdown heading trap: this theme's parser accepts
        # '#word' with no space as a heading, so a wrapped docs line starting
        # with a vendor model name ('#031 .fez ...') becomes an H1 mid-sentence.
        tests.append({'type': 'docs_heading_trap', 'file': 'docs line-initial # heading trap',
                      'method': '-', 'source': 'docs_heading_trap'})
        # The suite must discover the same tests from any working directory. A
        # CWD-relative fixture path does not error when it misses — the .exists()
        # guards around the docs scan just collect fewer rows — so a shrunken run
        # still reports all green. This row enumerates from two directories and
        # requires the two listings to match exactly.
        tests.append({'type': 'cwd_invariant', 'file': 'test discovery, any working directory',
                      'method': '-', 'source': 'cwd_invariant'})
        # Standing checks on the six verification pages: printed deltas
        # re-derived from the page's own numbers, tag/text agreement both ways,
        # and caption-vs-figure. Change-gated on a committed content hash, so an
        # unchanged page costs one file read. See tools/verification_checks.
        tests.append({'type': 'verification_pages', 'file': 'docs/verification (6 pages)',
                      'method': '-', 'source': 'verification_pages'})
        tests.append({'type': 'template_sync', 'file': BUNDLED_SKILL,
                      'master': SKILL_MASTER, 'copy': BUNDLED_SKILL,
                      'method': '-', 'source': 'skill'})
        # Guard that the generated corpus example index — the machine index and
        # the skill's TOPIC -> EXAMPLES table — matches what the docs' test tags
        # and the models they name would produce today. Regenerates in memory;
        # the fix is always `python tools/make_corpus_index.py`.
        tests.append({'type': 'corpus_index', 'file': 'docs/verification/corpus_index.json',
                      'method': '-', 'source': 'corpus_index'})
        # Editor no-drop guard (studio.editors). Touches the studio/Qt layer, so
        # it's skipped cleanly when PySide6 is absent (engine-only installs), like
        # the DXF round-trip tests.
        try:
            import PySide6          # noqa: F401
            tests.append({'type': 'editor_roundtrip', 'file': '(studio editors)',
                          'method': '-', 'source': 'editor_roundtrip'})
            if not run_all:
                print("Including 1 editor round-trip guard")
        except ImportError:
            print("Skipping editor round-trip guard (PySide6 not installed)")
        if Path(ROUNDTRIP_TEMPLATE).exists():
            n_rt = 0
            for fp, legacy_version in ROUNDTRIP_FILES:
                if Path(fp).exists() and Path(
                        LEGACY_TEMPLATE_FMT.format(legacy_version)).exists():
                    tests.append({'type': 'roundtrip', 'file': fp,
                                  'legacy_version': legacy_version,
                                  'template': ROUNDTRIP_TEMPLATE,
                                  'method': f'v{legacy_version}',
                                  'source': 'roundtrip'})
                    n_rt += 1
            if Path(V19_ROUNDTRIP_BASE).exists():
                tests.append({'type': 'v19_roundtrip', 'file': V19_ROUNDTRIP_BASE,
                              'template': ROUNDTRIP_TEMPLATE, 'method': '-',
                              'source': 'roundtrip'})
                n_rt += 1
            if Path(SSR_ZONE_ROUNDTRIP_BASE).exists():
                tests.append({'type': 'ssr_zone_roundtrip',
                              'file': SSR_ZONE_ROUNDTRIP_BASE,
                              'template': ROUNDTRIP_TEMPLATE, 'method': '-',
                              'source': 'roundtrip'})
                n_rt += 1
            if Path(SURFACE_FAMILY_BASE).exists():
                tests.append({'type': 'surface_family_roundtrip',
                              'file': SURFACE_FAMILY_BASE,
                              'template': ROUNDTRIP_TEMPLATE, 'method': '-',
                              'source': 'roundtrip'})
                n_rt += 1
            if Path(V21_ROUNDTRIP_BASE).exists():
                tests.append({'type': 'v21_roundtrip', 'file': V21_ROUNDTRIP_BASE,
                              'profile_file': V21_ROUNDTRIP_PROFILE_BASE,
                              'tseep_file': V21_ROUNDTRIP_TSEEP_BASE,
                              'template': ROUNDTRIP_TEMPLATE, 'method': '-',
                              'source': 'roundtrip'})
                n_rt += 1
            if n_rt and not run_all:
                print(f"Including {n_rt} Excel round-trip tests")
        else:
            print(f"Skipping round-trip tests (template not found: {ROUNDTRIP_TEMPLATE})")

    # Structured DXF export/import round-trip tests. These touch the studio layer
    # (build_from_dxf_mapping) and ezdxf, so they're skipped cleanly when either
    # dependency is absent — the engine-only suite stays runnable without the GUI.
    if run_dxf:
        try:
            import ezdxf            # noqa: F401
            import PySide6          # noqa: F401
            dxf_ok = True
        except ImportError:
            dxf_ok = False
            print("Skipping DXF round-trip tests (ezdxf/PySide6 not installed)")
        if dxf_ok:
            n_dxf = 0
            for fp, kind in DXF_FILES:
                if Path(fp).exists():
                    tests.append({'type': 'dxf', 'file': fp, 'kind': kind,
                                  'method': '-', 'source': 'dxf'})
                    n_dxf += 1
            # The water-load classifier rides with the DXF group: it is the same
            # export -> read -> map -> build path, asked which mode the file implies.
            if Path(DXF_WATER_FILE).exists():
                tests.append({'type': 'dxf_water', 'file': '(DXF water loads)',
                              'method': '-', 'source': 'dxf'})
                n_dxf += 1
            if n_dxf and not run_all:
                print(f"Including {n_dxf} DXF round-trip tests")

    # The GeoStudio import test authors its own .gsz fixture (GeoStudio's own sample
    # files are Seequent's copyrighted Materials and are not in this repository), so it
    # needs no input file and no GUI — engine-only, always runnable.
    if run_gsz:
        tests.append({'type': 'gsz', 'file': '(synthetic .gsz)',
                      'method': '-', 'source': 'gsz'})
        # The water-load handover is pinned on Seequent's own models as well as on the
        # fixture: it reads a local, git-ignored copy of the .gsz corpus and skips
        # when absent (the files are Seequent's copyrighted material).
        tests.append({'type': 'gsz_water', 'file': '(GeoStudio corpus)',
                      'method': 'water loads', 'source': 'gsz'})
        if not run_all:
            print("Including 2 GeoStudio import tests")

    # The Slide2 import test authors its own .slim fixture (Rocscience's own tutorial
    # files are their copyrighted material and are not in this repository), so it needs
    # no input file and no GUI — engine-only, always runnable.
    if run_slide2:
        tests.append({'type': 'slide2', 'file': '(synthetic .slim)',
                      'method': '-', 'source': 'slide2'})
        # Same discipline as the .gsz corpus: the water handover is checked on
        # Rocscience's own tutorial models from a local copy, skipped when absent.
        tests.append({'type': 'slide2_water', 'file': '(Slide2 tutorial models)',
                      'method': 'water loads', 'source': 'slide2'})
        if not run_all:
            print("Including 2 Slide2 import tests")

    # The RS2 import test authors its own .fez fixture (Rocscience's own verification
    # files are their copyrighted material and are not in this repository), so it needs
    # no input file and no GUI — engine-only, always runnable.
    if run_rs2:
        tests.append({'type': 'rs2', 'file': '(synthetic .fez)',
                      'method': '-', 'source': 'rs2'})
        # The per-material water-source (iStaticWaterMode) decode is pinned against
        # Rocscience's own verification models, which cannot live in this repository —
        # it reads them from a local copy of the downloads and skips when absent.
        tests.append({'type': 'rs2_water', 'file': '(RS2 verification models)',
                      'method': 'water modes', 'source': 'rs2'})
        # Same discipline for the 'global angle' load mapping and the material E/nu
        # transcription: both are scored against the vendor's own solved files.
        tests.append({'type': 'rs2_loads', 'file': '(RS2 verification models)',
                      'method': 'load deck', 'source': 'rs2'})
        if not run_all:
            print("Including 3 RS2 import tests")

    if args.skip_benchmarks:
        n_before = len(tests)
        tests = [t for t in tests if 'benchmark' not in t]
        n_skipped = n_before - len(tests)
        if n_skipped:
            print(f"Skipping {n_skipped} benchmark tests (--skip-benchmarks)")

    if args.benchmark:
        wanted = [w.strip().lower() for w in args.benchmark.split(',') if w.strip()]

        def _matches(bench_id):
            b = str(bench_id).lower()
            for w in wanted:
                # exact id, or a family prefix followed by a non-digit
                # (VP28 matches VP28a and VP28-beta but not VP280; VP3 does
                # not match VP33)
                if b == w or (b.startswith(w) and not b[len(w):][:1].isdigit()):
                    return True
            return False

        tests = [t for t in tests if 'benchmark' in t and _matches(t['benchmark'])]
        print(f"Running {len(tests)} tests matching --benchmark {args.benchmark}")
        if not tests:
            print("No benchmark tags matched.")
            return 1

    if args.quick:
        # Collapse each problem's per-method expansion to a single check
        # (prefer Spencer). Singleton tests (FEM, seep, single-method LEM) pass through.
        from collections import OrderedDict
        groups = OrderedDict()
        for t in tests:
            groups.setdefault((t.get('source'), t.get('file'), t.get('type')), []).append(t)
        tests = [next((t for t in grp if t.get('method') == 'spencer'), grp[0])
                 for grp in groups.values()]
        print(f"--quick: checking one method per problem ({len(tests)} tests)")

    if not tests:
        print("No test tags found in documentation files.")
        sys.exit(1)

    if args.list:
        # Enumeration only. Sorted and fully qualified so two runs from two
        # different working directories produce byte-identical output, which is
        # the property the cwd_invariant row checks.
        by_source = {}
        for t in tests:
            by_source[str(t.get('source', '-'))] = by_source.get(str(t.get('source', '-')), 0) + 1
        print(f"TOTAL {len(tests)}")
        for name in sorted(by_source):
            print(f"{by_source[name]:5d}  {Path(name).name if os.sep in name else name}")
        return

    # Thread the run-level context the two-tier fem_ssrm kernel router needs onto
    # each fem_ssrm row. It must ride ON the test dict (not a module global):
    # parallel workers run under the 'spawn' start method, so only the pickled test
    # dict crosses into the worker. `_default_tol` matches what the summary compares
    # with, so the fast-hit test uses the identical expected/tolerance the framework
    # applies to the returned FS.
    _n_ssrm = 0
    for t in tests:
        if t.get('type') == 'fem_ssrm':
            t['_default_tol'] = args.tolerance
            t['_reference_only'] = args.reference_only
            _n_ssrm += 1
    if _n_ssrm:
        print(_ssrm_mode_notice(_n_ssrm, args.reference_only))

    print(f"Found {len(tests)} tests\n")
    print(f"{'#':<4} {'File':<45} {'Type':<20} {'Method':<10} {'Expected':>10}  {'Computed':>10}  {'Status'}")
    print("-" * 120)

    passed = 0
    failed = 0
    errors = 0
    total_time = 0
    wall_t0 = time.time()
    # Two-tier fem_ssrm kernel-routing tallies (see _run_fem_ssrm). Reported at the
    # end so kernel-drift trends are visible run over run.
    route_fast = 0            # PASS via the fast kernel (Tier 1 hit)
    route_fallback = 0        # PASS via the reference kernel after a fast miss (Tier 2)
    route_direct = 0          # PASS via reference with no fast attempt (--reference-only / no kernel)
    route_fail = 0            # fem_ssrm row that did not pass (reference verdict FINAL)

    def _fmt(v):
        """Expected/computed in a 10-wide column, readable at ANY magnitude.

        Fixed-point %.3f renders a seepage flowrate of 2.5e-08 as '0.000' — every
        such row printed '0.000 0.000' whatever it did, and a real miss showed as
        'diff=+0.0000'. The comparison was always relative for those rows
        (_expected_and_tol scales by |expected|), so they did fail correctly; only
        the report was unreadable. Switch to scientific outside the range where
        fixed-point carries three real digits."""
        if v is None:
            return "    --    "
        av = abs(v)
        return f"{v:10.3e}" if (av and (av < 1e-3 or av >= 1e6)) else f"{v:10.3f}"

    def report(i, test, computed, error_msg, annotation, elapsed):
        nonlocal passed, failed, errors
        nonlocal route_fast, route_fallback, route_direct, route_fail
        file_name = Path(test['file']).name
        test_type = test.get('type', '?')
        method = test.get('method', '-')
        expected, tol = _expected_and_tol(test, args.tolerance)
        if error_msg:
            status = f"ERROR: {error_msg}"
            errors += 1
            comp_str = "    --    "
        elif expected is None:
            # A tag with no expected value for its type (e.g. a typo'd key or a
            # type missing from the pass/fail list above) is a broken tag, not
            # a pass — surface it instead of crashing the summary.
            status = "ERROR: tag has no expected value for this test type"
            errors += 1
            comp_str = _fmt(computed)
        elif abs(computed - expected) <= tol:
            status = f"PASS ({elapsed:.1f}s)"
            passed += 1
            comp_str = _fmt(computed)
        else:
            diff = computed - expected if computed is not None else float('nan')
            # Quote the miss the way its own tolerance reads it: relative when the
            # tolerance was scaled by |expected| (flowrates), absolute otherwise.
            if expected and abs(expected) < 1e-3:
                status = (f"FAIL (diff={diff:+.3e} = {100.0 * diff / expected:+.2f}%, "
                          f"{elapsed:.1f}s)")
            else:
                status = f"FAIL (diff={diff:+.4f}, {elapsed:.1f}s)"
            failed += 1
            comp_str = _fmt(computed)
        # fem_ssrm two-tier routing: annotate the row and tally which kernel decided
        # it, so a run-over-run rise in the fallback count flags creeping kernel drift.
        if annotation is not None:
            bucket, ann_text = annotation
            status = f"{status} [{ann_text}]"
            if status.startswith('PASS'):
                if bucket == 'fast':
                    route_fast += 1
                elif bucket == 'fallback':
                    route_fallback += 1
                else:
                    route_direct += 1
            elif status.startswith('FAIL'):
                route_fail += 1
        exp_str = _fmt(expected)
        print(f"{i:<4} {file_name:<45} {test_type:<20} {method:<10} {exp_str}  {comp_str}  {status}",
              flush=True)

    jobs = os.cpu_count() - 2 if str(args.jobs) == 'auto' else int(args.jobs)
    jobs = max(1, min(jobs, len(tests) or 1))

    if jobs > 1:
        # Keep numpy's BLAS single-threaded inside each worker so N processes
        # don't oversubscribe the CPU (inherited by spawned children).
        for _v in ('OMP_NUM_THREADS', 'OPENBLAS_NUM_THREADS',
                   'MKL_NUM_THREADS', 'VECLIB_MAXIMUM_THREADS'):
            os.environ.setdefault(_v, '1')
        from concurrent.futures import ProcessPoolExecutor, as_completed
        indexed = list(enumerate(tests, 1))
        # slowest types first: wall time is set by the last straggler
        indexed.sort(key=lambda it: -(_COST_RANK.get(it[1].get('type'), 1)
                                      + (2 if it[1].get('rapid') else 0)))
        print(f"Running {len(tests)} tests on {jobs} workers "
              f"(rows print as they finish)")
        with ProcessPoolExecutor(max_workers=jobs) as ex:
            futures = [ex.submit(_parallel_worker, item) for item in indexed]
            for fut in as_completed(futures):
                i, computed, error_msg, annotation, elapsed = fut.result()
                total_time += elapsed
                report(i, tests[i - 1], computed, error_msg, annotation, elapsed)
    else:
        for i, test in enumerate(tests, 1):
            t0 = time.time()
            annotation = None
            try:
                computed, error_msg, annotation = run_test(test)
            except Exception as e:
                computed = None
                error_msg = str(e)
            elapsed = time.time() - t0
            total_time += elapsed
            report(i, test, computed, error_msg, annotation, elapsed)

    print("-" * 120)
    print(f"\nResults: {passed} passed, {failed} failed, {errors} errors out of {len(tests)} tests")
    # FEM SSRM kernel routing: which kernel decided each fem_ssrm row. A run-over-run
    # rise in the fallback count is the visible signal of fast-kernel drift (see
    # _run_fem_ssrm); its companion guard is the kernel_xcheck gate.
    if route_fast or route_fallback or route_direct or route_fail:
        print(f"FEM SSRM kernel routing: {route_fast} via fast kernel, "
              f"{route_fallback} via reference fallback, {route_direct} reference-only pass, "
              f"{route_fail} failed")
        if route_fallback:
            print("  (fast-kernel misses fell back to the reference kernel and it "
                  "decided them — watch this count for kernel drift)")
    if jobs > 1:
        print(f"Total time: {time.time() - wall_t0:.1f}s wall on {jobs} workers "
              f"({total_time:.1f}s of test time)")
    else:
        print(f"Total time: {total_time:.1f}s")

    if failed > 0 or errors > 0:
        sys.exit(1)


if __name__ == '__main__':
    main()
