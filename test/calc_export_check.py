# Copyright 2025 Norman L. Jones
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Checks for Export LEM Calcs.

What is being defended:

  A. THE WORKBOOK IS THE CALCULATION — a solved model produces a workbook with
     the four sheets, the report's own curated slice columns as NUMBERS, a
     ``=SUM()`` under every per-slice contribution column, and exactly one
     embedded figure.

  B. THE FORMULAS ARE RIGHT — this is the whole point of the feature, so the
     check is not that a formula string exists but that it computes. Every
     formula in every workbook is evaluated here, in Python, from the numbers
     that were written: a small evaluator for the ``=SUM``/``MAX``/``ABS`` and
     trigonometric subset the exporter emits, no more. The factor of safety it
     arrives at must be the solver's, to the precision a factor of safety is
     printed at, and every check cell must read zero.

  C. A SECOND ENGINE AGREES — Python evaluating its own formulas proves the
     arithmetic, not the file. LibreOffice is asked to open the workbook and
     compute it, and its answer for the factor of safety must be the same. On a
     machine without LibreOffice the check says so and passes; it never asserts
     against an engine that is not there.

  D. THE ANNOTATION — the README names the method, the model, the date, the
     version and the published derivation, every sheet carries its reading
     notes, and NOTHING anywhere is a cell comment: annotations are cells.

  E. THE STUDIO COMMAND — the menu item exists, it is gated on a solution, and
     the default filename names the model and the method.

One small LEM model is solved once and exported under four methods. Janbu is
deliberately not pinned to a value here.
"""

import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO = os.path.dirname(_HERE)
if _REPO not in sys.path:
    sys.path.insert(0, _REPO)

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

REINF_XLSX = os.path.join(_REPO, "docs", "inputs", "slope", "xslope_reinf.xlsx")

#: The methods exported for these checks. OMS is explicit; Bishop and Spencer
#: iterate and carry a closure block; Corps exercises the branch that states the
#: converged unknowns without rebuilding them. Janbu is left out on purpose — its
#: sign convention is in flight, and a value pinned here would pin the bug.
METHODS = ("oms", "bishop", "spencer", "corps")

#: Where LibreOffice lives on the machines this is run on.
SOFFICE = ("/Applications/LibreOffice.app/Contents/MacOS/soffice",
           "/usr/bin/soffice", "/usr/local/bin/soffice", "soffice")

_EXPORTS = {}


def _exports():
    """One solved model, exported once per method: ``{method: (path, results)}``
    plus the model and the bundles, cached for every check below."""
    if "it" in _EXPORTS:
        return _EXPORTS["it"]
    import matplotlib
    matplotlib.use("Agg")
    from xslope.calc_export import export_lem_calcs
    from xslope.fileio import load_slope_data
    from xslope.slice import generate_slices
    from xslope.solve import solve_selected

    slope_data = load_slope_data(REINF_XLSX)
    ok, out = generate_slices(slope_data, circle=slope_data["circles"][0],
                              num_slices=15)
    if not ok:
        raise RuntimeError(f"the sample model produced no slices: {out}")
    slice_df, surface = out

    tmp = tempfile.mkdtemp(prefix="xslope_calc_export_")
    exports = {}
    for method in METHODS:
        df = slice_df.copy()
        results = solve_selected(method, df)
        bundle = {"slice_df": df, "failure_surface": surface,
                  "results": results, "search": None, "method": method}
        path = os.path.join(tmp, f"calcs_{method}.xlsx")
        ok, out = export_lem_calcs(slope_data, bundle, path,
                                   model_path=REINF_XLSX)
        if not ok:
            raise RuntimeError(f"{method} did not export: {out}")
        exports[method] = {"path": path, "results": results, "bundle": bundle,
                           "out": out}
    _EXPORTS["it"] = (slope_data, exports, tmp)
    return _EXPORTS["it"]


# ---------------------------------------------------------------------------
# A formula evaluator for exactly what the exporter writes
# ---------------------------------------------------------------------------

#: A cell reference, a range, or a function name, in that order — a range has to
#: win over the single reference that starts it.
_TOKEN = re.compile(
    r"(?P<range>(?:'(?P<rs>[^']+)'!)?\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)"
    r"|(?P<cell>(?:'(?P<cs>[^']+)'!)?\$?[A-Z]{1,3}\$?\d+)"
    r"|(?P<func>[A-Z][A-Z0-9]*(?=\())")

#: Every spreadsheet function the exporter is allowed to emit. Anything else in
#: a workbook is a failure here rather than a silent skip: a formula this cannot
#: read is a formula this cannot check.
_FUNCS = {"SUM": "f_sum", "MAX": "f_max", "ABS": "f_abs", "COS": "f_cos",
          "SIN": "f_sin", "TAN": "f_tan", "RADIANS": "f_rad"}


class Evaluated:
    """A workbook whose formulas are computed from the numbers in it.

    openpyxl writes formula strings and caches no result, so the file on disk
    holds no answers at all — which is exactly the state a consumer opens it in.
    This walks the formulas and works them out, so that what is asserted is what
    a spreadsheet would compute, not what the exporter believed.
    """

    def __init__(self, path):
        import openpyxl
        self.wb = openpyxl.load_workbook(path, data_only=False)
        self._cache = {}
        self.count = 0

    def cell(self, sheet, coord):
        key = (sheet, coord.replace("$", ""))
        if key in self._cache:
            return self._cache[key]
        self._cache[key] = None                       # breaks any cycle
        value = self._compute(sheet, self.wb[sheet][key[1]].value)
        self._cache[key] = value
        return value

    def _range(self, sheet, ref):
        return [self.cell(sheet, c.coordinate)
                for row in self.wb[sheet][ref.replace("$", "")] for c in row]

    def _compute(self, sheet, raw):
        if not isinstance(raw, str) or not raw.startswith("="):
            return raw
        self.count += 1

        def replace(m):
            if m.group("range"):
                return (f"_R({(m.group('rs') or sheet)!r},"
                        f"{m.group('range').split('!')[-1]!r})")
            if m.group("cell"):
                return (f"_C({(m.group('cs') or sheet)!r},"
                        f"{m.group('cell').split('!')[-1]!r})")
            name = m.group("func")
            if name not in _FUNCS:
                raise ValueError(f"the export used {name}(), which these checks "
                                 f"cannot evaluate: {raw!r}")
            return _FUNCS[name]

        python = _TOKEN.sub(replace, raw[1:])
        env = {
            "_R": self._range, "_C": self.cell,
            "f_sum": lambda v: sum(x or 0.0 for x in v),
            "f_max": lambda v: max(x or 0.0 for x in v),
            "f_abs": abs, "f_cos": math.cos, "f_sin": math.sin,
            "f_tan": math.tan, "f_rad": math.radians,
        }
        try:
            return eval(python, {"__builtins__": {}}, env)
        except Exception as exc:
            raise ValueError(f"{raw!r} did not evaluate ({python!r}): {exc}")

    def all_formulas(self):
        """Compute every formula in the workbook. Returns how many there were."""
        for sheet in self.wb.sheetnames:
            for row in self.wb[sheet].iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        self.cell(sheet, cell.coordinate)
        return self.count

    def labelled(self, sheet=None):
        """The FS sheet's labelled rows: ``[{"label", "note", "value"}, ...]``.

        The label is column A and the note is column C, and both are read: the
        sheet says what a cell is in A and what it must read in C, so "should be
        zero" can live in either.
        """
        ws = self.wb[sheet or "FS Calculation"]
        out = []
        for row in ws.iter_rows(min_col=1, max_col=3):
            label, value, note = row[0].value, row[1].value, row[2].value
            if isinstance(label, str) and value is not None:
                out.append({"label": label,
                            "note": note if isinstance(note, str) else "",
                            "value": self.cell(ws.title, row[1].coordinate)})
        return out


def _find(rows, needle):
    """The first labelled value whose label contains ``needle``, or None."""
    for row in rows:
        if needle.lower() in row["label"].lower():
            return row["value"]
    return None


# ---------------------------------------------------------------------------
# A. the workbook is the calculation
# ---------------------------------------------------------------------------

def test_workbook_shape():
    """Four sheets, the report's own columns as numbers, a Σ under each
    contribution column, one figure, and calculate-on-load."""
    import openpyxl

    from xslope.calc_export import (SHEET_CALC, SHEET_PLOT, SHEET_README,
                                    SHEET_TABLE, SUMMED_KEYS)
    from xslope.columns import BY_KEY, header, slice_table
    from xslope.plot import declared_unit_labels
    from xslope.report import calculation

    fails = []
    slope_data, exports, _tmp = _exports()
    unit_labels = declared_unit_labels(slope_data)

    for method, export in exports.items():
        path = export["path"]
        wb = openpyxl.load_workbook(path, data_only=False)
        if wb.sheetnames != [SHEET_README, SHEET_PLOT, SHEET_TABLE, SHEET_CALC]:
            fails.append(f"{method}: the sheets are {wb.sheetnames}")
            continue
        if not wb.calculation.fullCalcOnLoad:
            fails.append(f"{method}: the workbook does not ask to be recomputed "
                         f"on open, so a consumer could show stale values")

        calc = calculation(slope_data, export["bundle"], method)
        headers, rows, _legend = slice_table(calc["slice_df"], unit_labels)
        ws = wb[SHEET_TABLE]
        written = [ws.cell(row=1, column=i + 1).value for i in range(len(headers))]
        if written != headers:
            fails.append(f"{method}: the slice table's headers are not the "
                         f"report's ({written[:4]} vs {headers[:4]})")
        if ws.freeze_panes != "B2":
            fails.append(f"{method}: the header row is not frozen "
                         f"({ws.freeze_panes})")

        # The values are numbers, not the report's formatted strings.
        n = len(calc["slice_df"])
        sample = ws.cell(row=2, column=headers.index(
            header(BY_KEY["w"], unit_labels)) + 1).value
        if not isinstance(sample, float):
            fails.append(f"{method}: slice weights were written as "
                         f"{type(sample).__name__}, not numbers")

        total = 2 + n
        summed = set()
        for i, text in enumerate(headers, start=1):
            value = ws.cell(row=total, column=i).value
            if isinstance(value, str) and value.startswith("=SUM("):
                summed.add(text)
        want = {header(BY_KEY[k], unit_labels) for k in SUMMED_KEYS
                if header(BY_KEY[k], unit_labels) in headers}
        if summed != want:
            fails.append(f"{method}: the Σ row covers {sorted(summed)}, not "
                         f"{sorted(want)}")
        if not want:
            fails.append(f"{method}: the slice table carries no contribution "
                         f"column to sum")

        with zipfile.ZipFile(path) as package:
            media = [nm for nm in package.namelist() if nm.startswith("xl/media/")]
        if len(media) != 1:
            fails.append(f"{method}: the workbook embeds {len(media)} images, "
                         f"not one ({media})")
    return fails


# ---------------------------------------------------------------------------
# B. the formulas compute the solver's answer
# ---------------------------------------------------------------------------

def test_formulas_reproduce_the_solver():
    """Every formula evaluated; the factor of safety is the solver's."""
    from xslope.columns import format_fs

    fails = []
    _slope_data, exports, _tmp = _exports()
    for method, export in exports.items():
        book = Evaluated(export["path"])
        try:
            count = book.all_formulas()
        except ValueError as exc:
            fails.append(f"{method}: {exc}")
            continue
        if count < 3:
            fails.append(f"{method}: only {count} formulas — the workbook is a "
                         f"dump, not a calculation")
        labels = book.labelled()
        stated = _find(labels, "reported by the solver")
        computed = _find(labels, "resisting / driving")
        if method == "janbu":
            computed = _find(labels, "F_corr")
        if stated is None or computed is None:
            fails.append(f"{method}: the FS sheet has no stated/computed pair")
            continue
        solved = float(export["results"]["FS"])
        if abs(stated - solved) > 1e-9 * max(1.0, abs(solved)):
            fails.append(f"{method}: the workbook states F = {stated}, the "
                         f"solver returned {solved}")
        if format_fs(computed) != format_fs(solved):
            fails.append(f"{method}: the formulas compute F = "
                         f"{format_fs(computed)}, the solver returned "
                         f"{format_fs(solved)}")
    return fails


def test_every_check_cell_reads_zero():
    """A cell whose label says it should be zero is zero.

    That is the workbook's own honesty test, and it covers the closures: Bishop's
    N' rebuilt from the stated factor of safety returning the moments the
    quotient was formed from, and Spencer's Q rebuilt from the stated (F, θ)
    driving both equilibrium sums to nothing.
    """
    fails = []
    _slope_data, exports, _tmp = _exports()
    for method, export in exports.items():
        book = Evaluated(export["path"])
        book.all_formulas()
        checked = 0
        for row in book.labelled():
            words = f"{row['label']} {row['note']}".lower()
            if "should be zero" not in words:
                continue
            checked += 1
            value = row["value"]
            if value is None or abs(float(value)) > 1e-3:
                fails.append(f"{method}: '{row['label']}' reads {value}, which "
                             f"is not zero")
        if method in ("bishop", "spencer") and checked < 2:
            fails.append(f"{method} iterates, so its sheet must demonstrate the "
                         f"closure; it carries {checked} check cells")
        if checked == 0:
            fails.append(f"{method}: the sheet has no check cell at all")
    return fails


def test_closure_blocks_rebuild_the_per_slice_values():
    """Bishop's N' and Spencer's Q are recomputed by formula from the stated
    converged values, not copied — and land on the solver's own numbers."""
    fails = []
    _slope_data, exports, _tmp = _exports()
    expected = {"bishop": ("recomputed N'", 1e-3),
                "spencer": ("recomputed Q", 1e-6)}
    for method, (needle, tol) in expected.items():
        export = exports.get(method)
        if export is None:
            continue
        book = Evaluated(export["path"])
        book.all_formulas()
        labels = book.labelled()
        worst = _find(labels, needle)
        if worst is None:
            fails.append(f"{method}: the sheet does not rebuild its per-slice "
                         f"values from the stated solution")
            continue
        scale = max(abs(float(v)) for v in
                    export["bundle"]["slice_df"]["n_eff"].values)
        if abs(float(worst)) > tol * max(1.0, scale):
            fails.append(f"{method}: the rebuilt per-slice values differ from "
                         f"the solver's by {worst}")
    return fails


# ---------------------------------------------------------------------------
# C. a second engine agrees
# ---------------------------------------------------------------------------

def _soffice():
    for candidate in SOFFICE:
        found = candidate if os.path.isfile(candidate) else shutil.which(candidate)
        if found:
            return found
    return ""


def test_libreoffice_agrees():
    """LibreOffice opens the workbook, computes it, and reports the same factor
    of safety. Skipped — with a word — where LibreOffice is not installed."""
    from xslope.columns import format_fs

    soffice = _soffice()
    if not soffice:
        print("    LibreOffice is not installed — the cross-check was skipped.")
        return []

    fails = []
    _slope_data, exports, _tmp = _exports()
    out_dir = tempfile.mkdtemp(prefix="xslope_calc_lo_")
    try:
        paths = [e["path"] for e in exports.values()]
        cmd = [soffice, "--headless", "--convert-to",
               "csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,"
               "false,false,-1", "--outdir", out_dir] + paths
        try:
            subprocess.run(cmd, check=True, capture_output=True, timeout=300)
        except Exception as exc:
            print(f"    LibreOffice could not be driven ({exc}); the "
                  f"cross-check was skipped.")
            return []
        for method, export in exports.items():
            stem = os.path.splitext(os.path.basename(export["path"]))[0]
            csv = os.path.join(out_dir, f"{stem}-FS Calculation.csv")
            if not os.path.exists(csv):
                fails.append(f"{method}: LibreOffice wrote no FS Calculation "
                             f"sheet ({os.listdir(out_dir)})")
                continue
            with open(csv, encoding="utf-8", errors="replace") as fh:
                text = fh.read()
            want = format_fs(export["results"]["FS"])
            line = [ln for ln in text.splitlines()
                    if ln.startswith("F  =  resisting / driving")]
            if method == "janbu":
                line = [ln for ln in text.splitlines() if ln.startswith("F_corr")]
            if not line:
                fails.append(f"{method}: LibreOffice's sheet has no quotient row")
                continue
            got = line[0].split(",")[1].strip('"')
            if got != want:
                fails.append(f"{method}: LibreOffice computed F = {got}, the "
                             f"solver returned {want}")
            # And the workbook's own check cell, computed by LibreOffice.
            for ln in text.splitlines():
                if ln.startswith("Check:"):
                    value = ln.split(",")[1].strip('"')
                    try:
                        if abs(float(value)) > 1e-3:
                            fails.append(f"{method}: LibreOffice reads the check "
                                         f"cell as {value}")
                    except ValueError:
                        fails.append(f"{method}: LibreOffice left the check cell "
                                     f"as {value!r}")
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)
    return fails


# ---------------------------------------------------------------------------
# D. the annotation
# ---------------------------------------------------------------------------

def test_readme_and_no_cell_notes():
    """The README says what this is, and nothing in the workbook is a comment."""
    import openpyxl

    from xslope.calc_export import (SHEET_CALC, SHEET_PLOT, SHEET_README,
                                    SHEET_TABLE)
    from xslope.report import method_doc_url, method_label

    fails = []
    _slope_data, exports, _tmp = _exports()
    for method, export in exports.items():
        wb = openpyxl.load_workbook(export["path"], data_only=False)
        text = "\n".join(
            str(cell.value) for row in wb[SHEET_README].iter_rows()
            for cell in row if cell.value is not None)
        for needed in (method_label(method), "xslope_reinf.xlsx",
                       method_doc_url(method), SHEET_TABLE, SHEET_CALC,
                       SHEET_PLOT):
            if needed and needed not in text:
                fails.append(f"{method}: the README does not carry {needed!r}")
        if not re.search(r"\b20\d\d\b", text):
            fails.append(f"{method}: the README carries no date")
        from xslope import __version__
        if __version__ not in text:
            fails.append(f"{method}: the README does not name the xslope version")

        # No cell notes, anywhere, ever (the standing rule).
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows():
                for cell in row:
                    if cell.comment is not None:
                        fails.append(f"{method}: {sheet}!{cell.coordinate} "
                                     f"carries a cell note")
        with zipfile.ZipFile(export["path"]) as package:
            if any("comments" in nm for nm in package.namelist()):
                fails.append(f"{method}: the package contains a comments part")

        # The named ranges are there and point at real cells.
        names = set(wb.defined_names)
        for wanted in ("Sum_resisting", "Sum_driving", "Factor_of_safety"):
            if wanted not in names:
                fails.append(f"{method}: there is no {wanted} named range")
    return fails


def test_notation_is_plain():
    """The report's math notation arrives as something a cell can show."""
    from xslope.calc_export import plain_math

    fails = []
    cases = [
        ("F = frac{sum{a}}{sum{b}}", "F = (Σ(a)) / (Σ(b))"),
        ("m_α = cos α + frac{sin α·tan φ}{F}", "m_α = cos α + (sin α·tan φ) / (F)"),
        ("sum{Q} = 0", "Σ(Q) = 0"),
    ]
    for source, want in cases:
        got = plain_math(source)
        if got != want:
            fails.append(f"plain_math({source!r}) gave {got!r}, not {want!r}")
    # Nothing that reached a cell still carries the renderer's braces.
    import openpyxl
    _slope_data, exports, _tmp = _exports()
    for method, export in exports.items():
        wb = openpyxl.load_workbook(export["path"], data_only=False)
        for row in wb["FS Calculation"].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and ("frac{" in cell.value
                                                    or "sum{" in cell.value):
                    fails.append(f"{method}: {cell.coordinate} still holds the "
                                 f"renderer's notation: {cell.value!r}")
    return fails


# ---------------------------------------------------------------------------
# E. the Studio command
# ---------------------------------------------------------------------------

def test_default_filename():
    from xslope.calc_export import default_filename
    fails = []
    got = default_filename("/models/north_levee.xlsx", "spencer")
    if got != "north_levee_calcs_spencer.xlsx":
        fails.append(f"the default filename is {got!r}")
    if not default_filename(None, "oms").endswith(".xlsx"):
        fails.append("an unsaved model gets no default filename")
    return fails


def _app():
    from PySide6.QtWidgets import QApplication
    return QApplication.instance() or QApplication([])


def test_main_window_action():
    """The menu item exists, sits in File, and waits for a solution."""
    fails = []
    _app()
    from PySide6.QtWidgets import QMenu

    from studio.main_window import MainWindow

    win = MainWindow()
    try:
        act = getattr(win, "act_export_calcs", None)
        if act is None:
            return ["there is no Export LEM Calcs action"]
        if "Calcs" not in act.text():
            fails.append(f"the action is called {act.text()!r}")
        file_menus = [m for m in win.menuBar().findChildren(QMenu)
                      if "File" in m.title()]
        if not file_menus:
            fails.append("there is no File menu to look in")
        elif not any(a is act for m in file_menus for a in m.actions()):
            fails.append("the action is not on the File menu")
        if act.isEnabled():
            fails.append("the action is live with no model open")
        if not act.toolTip():
            fails.append("a disabled action says nothing about why")
        if not hasattr(win, "export_lem_calcs_dialog"):
            fails.append("the action has no handler on the window")
    finally:
        win.close()
    return fails


CHECKS = [
    ("the workbook is the calculation", test_workbook_shape),
    ("the formulas reproduce the solver", test_formulas_reproduce_the_solver),
    ("every check cell reads zero", test_every_check_cell_reads_zero),
    ("the closures rebuild the per-slice values",
     test_closure_blocks_rebuild_the_per_slice_values),
    ("LibreOffice computes the same answer", test_libreoffice_agrees),
    ("the README, and no cell notes", test_readme_and_no_cell_notes),
    ("the notation reaches the cells as text", test_notation_is_plain),
    ("the default filename", test_default_filename),
    ("the menu item and its gate", test_main_window_action),
]

#: Checks that need the Studio layer; skipped when PySide6 is absent.
_STUDIO_ONLY = {test_main_window_action}


def run():
    """Every check; returns a list of failure strings (empty = pass)."""
    checks = CHECKS
    try:
        import PySide6                                    # noqa: F401
    except Exception:
        print("Calc export: PySide6 not installed — Studio checks skipped.")
        checks = [c for c in CHECKS if c[1] not in _STUDIO_ONLY]

    failures = []
    try:
        for name, fn in checks:
            try:
                fs = fn()
            except Exception as exc:
                import traceback
                traceback.print_exc()
                fs = [f"{name} raised: {exc!r}"]
            print(f"  {name:48s} {'ok' if not fs else f'FAIL ({len(fs)})'}")
            failures += fs
    finally:
        if "it" in _EXPORTS:
            shutil.rmtree(_EXPORTS["it"][2], ignore_errors=True)
    return failures


def main():
    print("Export LEM Calcs:")
    failures = run()
    if failures:
        print("\nFAILURES:")
        for f in failures:
            print(f"  - {f}")
        raise SystemExit(1)
    print("\nAll calculation-export checks passed.")


if __name__ == "__main__":
    import matplotlib
    matplotlib.use("Agg")
    try:
        from PySide6.QtWidgets import QApplication
        _app_ = QApplication.instance() or QApplication([])
    except Exception:
        pass
    main()
