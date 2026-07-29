"""Build the per-sheet SAMPLE workbooks that the input-template docs images are
rendered from (see ``tools/render_sheet.py`` and ``tools/docs_sheets_manifest.py``).

Most template sheets (profile, polygon, circles, reinforce, piles, lloads,
non-circ, dloads) have an identical layout in template v12 and the current v15,
so the manifest points the renderer straight at existing committed sample files
in ``docs/lem/files`` / ``docs/seep/files``. Three sheets changed shape (or need
data no committed file carries) and so are (re)built here as current-version
files under ``docs/usage/sample_sheets/``:

  * ``sheets_mat.xlsx``   — a materials table exercising every strength option
    (mc / cp / pow / hb), plus gamma_sat, the ru pore option, both unsaturated
    laws (lf and van Genuchten), a couple of standard deviations, stiffness, the
    v17 matric-suction pair phi_b/s_cap on the one material where they're live
    (mc + u=piezo), and the v18 transient-storage pair Ss/Sy (Ss on every
    water-bearing material; Sy left blank on the confined bedrock).
    Written cell-by-cell into a fresh template via :func:`mat_header_cols` +
    :func:`write_cells_to_xlsx`, so N/A cells stay *blank* (matching how the
    conditional formatting greys them) rather than the ``0`` that a full
    ``save_slope_data_to_xlsx`` round-trip would leave.
  * ``sheets_rapid.xlsx`` — a rapid-drawdown model (two piezometric lines, two
    distributed-load stages) round-tripped to v15 so the piezo sheet carries the
    v13+ Type row; line 2 is flipped to ``phreatic`` to show both line types.
  * ``sheets_seepbc.xlsx`` — a two-stage seepage model round-tripped to v15, with
    a specified-flux (infiltration) boundary added to stage 1 so the seep-bc image
    shows all three BC kinds (head, flux, exit face).
  * ``sheets_tseep.xlsx`` — a seepage model round-tripped through the v18 template
    with a transient-seepage (``tseep``) block attached: a time axis, two named
    series (a reservoir-drawdown head series and a rainfall-flux series with a
    gap), and the run controls (duration, save_interval, save_times, stage_1/
    stage_2). Illustrates the tseep sheet layout only — parse + round-trip; the
    transient solver is a later phase.

Run:  python tools/build_docs_sheet_samples.py     # regenerate the three files
"""

from __future__ import annotations

import os

from xslope.fileio import (
    load_slope_data, save_slope_data_to_xlsx, default_template_path,
    mat_header_cols, write_cells_to_xlsx, cell_ref, _read_template_info,
)


def _template_version(path):
    """Template version of a workbook (main!D5), or 0 when unreadable."""
    return _read_template_info(path)[0]

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(REPO_ROOT, "docs", "usage", "sample_sheets")


# --------------------------------------------------------------------------- #
# 1. mat — write only the applicable cells (blank N/A cells, like a hand-filled
#    template), so the conditional-format greying reads clean in the docs image.
# --------------------------------------------------------------------------- #
# Each material is a {header-key: value} dict; header keys match mat_header_cols()
# (underscore-stripped: 'powa', 'hbsci', 's(g)', ...). Only keys present are
# written; everything else is left blank.
#
# t_cut (v16) shows all three tri-state readings across the showcase: a positive
# Rankine cap on Embankment, the RS2 VP2-style T=0 "no tension" idiom on the
# phi=0 Soft Clay (cp) envelope (see the dependency-matrix note in the docs),
# and blank (today's unbounded-tension default) on the rest.
#
# phi_b/s_cap (v17) show on Embankment: mc + u=piezo is exactly the combination
# where the columns are active (mc/pow/hb strength with a signed u source) AND
# the piezo-cap caution applies (a piezometric line's hydrostatic suction above
# the line is unbounded, so s_cap is load-bearing there, unlike u=seep's
# self-bounded field). Every other showcase material leaves phi_b/s_cap blank —
# cp/elastic are dependency-inert (greyed) and Rockfill/Weathered Rock/Drain Sand
# don't carry a piezo/seep u option, so the columns would grey there too.
# Ss/Sy (v18) are the transient-seepage storage columns (AO/AP). Ss (specific
# storage, 1/length) is shown on every water-bearing material; Sy (specific yield,
# dimensionless) on the drainable ones. Bedrock leaves Sy blank on purpose — it is
# the confined / always-saturated zone, the case the plan lets skip Sy — so the
# image shows both the filled and the legitimately-blank reading side by side.
MAT_SHOWCASE = [
    # Embankment fill — Mohr-Coulomb, moist+saturated unit weights, a Kc=1
    # rapid-drawdown envelope (d, psi), a Rankine tension cutoff, a matric-
    # suction friction angle capped at 1000 (stress units), and two reliability
    # std deviations.
    {"name": "Embankment", "g": 125, "gsat": 128, "option": "mc",
     "c": 50, "f": 32, "d": 300, "psi": 20, "phib": 15, "scap": 1000,
     "tcut": 20, "u": "piezo",
     "s(c)": 15, "s(f)": 3,
     "k1": 5e-5, "k2": 5e-5, "alpha": 0, "unsat": "lf", "kr0": 0.001, "h0": -1,
     "Ss": 3e-5, "Sy": 0.12,
     "E": 30000, "nu": 0.33},
    # Soft foundation clay — c/p undrained strength increasing below a reference
    # elevation. t_cut=0 is the RS2 VP2 crack-layer idiom: a phi=0 envelope
    # implies unlimited tension unless explicitly capped at zero.
    {"name": "Soft Clay", "g": 118, "gsat": 122, "option": "cp",
     "c": 600, "c/p": 12, "r-elev": 95, "tcut": 0, "u": "piezo",
     "s(c)": 60, "s(c/p)": 2,
     "k1": 1e-7, "k2": 1e-7, "alpha": 0, "unsat": "lf", "kr0": 0.001, "h0": -1,
     "Ss": 1e-4, "Sy": 0.05,
     "E": 8000, "nu": 0.40},
    # Rockfill shell — nonlinear power-curve envelope.
    {"name": "Rockfill", "g": 140, "option": "pow",
     "powa": 5.5, "powb": 0.82, "powc": 0, "powd": 10, "u": "none",
     "k1": 1e-2, "k2": 1e-2, "alpha": 0, "unsat": "lf", "kr0": 0.001, "h0": -1,
     "Ss": 1e-6, "Sy": 0.25,
     "E": 50000, "nu": 0.30},
    # Weathered bedrock — generalized Hoek-Brown.
    {"name": "Weathered Rock", "g": 155, "option": "hb",
     "hbsci": 25000, "hbgsi": 45, "hbmi": 8, "hbd": 0, "u": "none",
     "k1": 1e-6, "k2": 1e-6, "alpha": 0, "unsat": "lf", "kr0": 0.001, "h0": -1,
     "Ss": 5e-6, "Sy": 0.03,
     "E": 500000, "nu": 0.25},
    # Drainage sand — Mohr-Coulomb with the ru pore-pressure option and a van
    # Genuchten unsaturated curve.
    {"name": "Drain Sand", "g": 120, "option": "mc",
     "c": 0, "f": 34, "u": "ru", "ru": 0.15,
     "k1": 5e-2, "k2": 5e-2, "alpha": 0, "unsat": "vg", "a": 0.05, "n": 1.8,
     "Ss": 5e-6, "Sy": 0.28,
     "E": 40000, "nu": 0.30},
    # Sound bedrock — option='elastic' (cannot fail): only g/gsat, E, nu, and the
    # seepage columns are read; strength columns are left blank (conditional
    # formatting greys them out automatically). Confined/always-saturated, so Sy
    # is left blank (only Ss is required for a fully-saturated transient zone).
    {"name": "Bedrock", "g": 165, "gsat": 168, "option": "elastic", "u": "none",
     "k1": 1e-8, "k2": 1e-8, "alpha": 0, "unsat": "lf", "kr0": 0.001, "h0": -1,
     "Ss": 1e-7,
     "E": 2_000_000, "nu": 0.22},
]


def build_mat(out_path):
    # Source the master template the docs page links for download
    # (docs/inputs/input_template.xlsx), NOT the bundled package copy from
    # default_template_path(). The two can lag each other between a master edit and
    # the next resource sync; sourcing the documented master keeps the rendered mat
    # images matching the template a reader actually downloads. Cells are written by
    # header NAME below (mat_header_cols), so this is layout-order independent.
    template = os.path.join(REPO_ROOT, "docs", "inputs", "input_template.xlsx")
    import shutil
    shutil.copy(template, out_path)
    header_row, cols = mat_header_cols(out_path)

    def col_for(key):
        return cols.get(str(key).replace("_", ""))

    updates = {}
    for i, mat in enumerate(MAT_SHOWCASE):
        row = header_row + 1 + i
        updates[cell_ref(row, cols["mat"])] = i + 1
        for key, val in mat.items():
            c = col_for(key)
            if c is None:
                raise KeyError("mat header %r not found in template" % key)
            updates[cell_ref(row, c)] = val
    write_cells_to_xlsx(out_path, {"mat": updates})
    return out_path


# --------------------------------------------------------------------------- #
# 2 & 3. round-trip existing rich files to v15, augmenting where a feature that
#        no committed file carries is needed for the docs.
# --------------------------------------------------------------------------- #
def _fill_material_names(path, sheet, materials):
    """Write the profile/polygon material-NAME echo as literal text.

    Those cells hold an XLOOKUP against the mat sheet; a file written by xslope
    (never opened in Excel) has no cached formula result, so the name row renders
    blank. Overwriting with the resolved name is what Excel would display and keeps
    the docs image complete. Mat-ID values sit every third column (B, E, H, ...);
    the name is a merged cell anchored one column to the LEFT of its Mat-ID value
    (A6:B6 looks up B5, D6:E6 looks up E5, ...), written at col-1.

    The rows differ by sheet on v21: the profile sheet keeps Mat ID on row 5 with the
    echo on row 6, but the polygon sheet inserted a Type row above them, so its Mat ID
    is row 6 and its echo row 7. On the polygon sheet the echo is deliberately BLANK
    for a non-material Type — the dropdown already says what the block is, so echoing
    it again would be redundant (Norm's v21 review) — which the formula does by
    testing the Type cell, and which is reproduced here by simply not writing one.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    ver = _template_version(path)
    if sheet == "polygon" and ver >= 21:
        type_row, id_row, name_row = 5, 6, 7
    else:
        type_row, id_row, name_row = None, 5, 6
    id_to_name = {i + 1: str(m.get("name", "")) for i, m in enumerate(materials)}
    from xslope.fileio import SSR_ZONE_LABELS, SSR_ZONE_SENTINELS
    zone_names = {mid: SSR_ZONE_LABELS[kind]
                  for mid, kind in SSR_ZONE_SENTINELS.items()}
    updates = {}
    for c in range(2, ws.max_column + 1, 3):
        if type_row is not None:
            word = str(ws.cell(row=type_row, column=c).value or "").strip().lower()
            if word not in ("", "material"):
                continue                      # overlay block: the echo stays blank
        mid = ws.cell(row=id_row, column=c).value
        if isinstance(mid, (int, float)):
            # Pre-v21: a NEGATIVE Mat ID is an SSR zone sentinel, and the echo
            # formula shows its display code rather than a material name.
            name = zone_names.get(int(mid)) or id_to_name.get(int(mid))
            if name:
                updates[cell_ref(name_row, c - 1)] = name
    wb.close()
    if updates:
        write_cells_to_xlsx(path, {sheet: updates})


def build_rapid(out_path):
    sd = load_slope_data(os.path.join(REPO_ROOT, "docs/lem/files/xslope_gsat_rapid.xlsx"))
    # Show both piezometric-line types: keep line 1 as a piezometric line, mark
    # line 2 (the drawdown line) as a phreatic surface.
    sd["piezo_phreatic2"] = True
    # v21 profile Size on one line only — optional per line, so the image shows both
    # a set and an unset cell rather than a column of identical values.
    if sd.get("profile_lines"):
        sd["profile_lines"][0]["size"] = 2.0
    # v21 dload Direction on the FIRST block of each sheet, left blank on the rest,
    # so the image shows the dropdown filled beside the default it replaces. This
    # file is a sample for the sheet layout only — nothing solves it — so the
    # direction is a display value, not a re-statement of the drawdown physics.
    for key in ("dload_dirs", "dload2_dirs"):
        if sd.get(key):
            sd[key][0] = "vertical"
    save_slope_data_to_xlsx(sd, out_path)
    _fill_material_names(out_path, "profile", sd.get("materials", []))
    return out_path


def build_polygon(out_path):
    sd = load_slope_data(os.path.join(REPO_ROOT, "docs/seep/files/xslope_levee_poly.xlsx"))
    # One block of EVERY polygon Type, so the docs image shows the whole vocabulary
    # side by side: ordinary material zones, then the three SSR overlays ("ssr
    # reduce" / "ssr hold" / "ssr elastic") and a mesh "refine" region. Placed over
    # the levee's own section — a search area over the embankment, a hold over the
    # left foundation, an elastic hold over the right, and a refine box at the toe.
    #
    # Sizes are set on two of them (a material zone and one overlay) but not the
    # rest, because that IS the option: Size is optional on any polygon and layers
    # on top of the Type rather than replacing it. The refine block carries one
    # necessarily — it has no other effect.
    sd["polygons"][0]["size"] = 1.5
    sd["ssr_zones"] = [
        {"kind": "reduce", "polygon": [(30.0, 0.0), (110.0, 0.0),
                                       (110.0, 30.0), (30.0, 30.0)], "size": 0.75},
        {"kind": "hold", "polygon": [(0.0, 0.0), (30.0, 0.0),
                                     (30.0, 12.0), (0.0, 12.0)]},
        {"kind": "hold_elastic", "polygon": [(110.0, 0.0), (140.0, 0.0),
                                             (140.0, 12.0), (110.0, 12.0)]},
    ]
    sd["refine_zones"] = [
        {"polygon": [(96.0, 8.0), (120.0, 8.0), (120.0, 20.0), (96.0, 20.0)],
         "size": 0.5},
    ]
    save_slope_data_to_xlsx(sd, out_path)
    _fill_material_names(out_path, "polygon", sd.get("materials", []))
    return out_path


def build_seepbc(out_path):
    sd = load_slope_data(os.path.join(REPO_ROOT, "docs/lem/files/xslope_johnson_rapid_KEY.xlsx"))
    # Add a specified-flux (infiltration) boundary on the crest of stage 1, so the
    # seep-bc image shows all three BC kinds alongside the two heads + exit face.
    bc = sd.setdefault("seepage_bc", {})
    bc.setdefault("specified_fluxes", []).append(
        {"flux": 1e-6, "coords": [(320.0, 160.0), (360.0, 180.0), (380.0, 180.0)]}
    )
    save_slope_data_to_xlsx(sd, out_path)
    return out_path


def build_tseep(out_path):
    # Round-trip a seepage model through the v18 template (default template is v18),
    # attaching a transient-seepage block so the tseep sheet renders with example
    # data. The declared time unit + per-material Ss keep the file coherent, but the
    # image only documents the tseep sheet layout — the transient solver is a later
    # phase, so nothing here is run.
    sd = load_slope_data(os.path.join(REPO_ROOT, "docs/lem/files/xslope_johnson_rapid_KEY.xlsx"))
    sd["unit_system"] = "imperial"
    sd["time_unit"] = "day"
    # Give every material storage props so the file is a valid, loadable transient
    # model (Ss required for all, Sy required on the unconfined ones). The loaded
    # materials already carry Ss/Sy = None, so assign unconditionally, not setdefault.
    for m in sd.get("materials", []):
        m["Ss"] = 1e-5
        m["Sy"] = 0.15
    # A reservoir head drawn down over 40 days, and a rainfall flux applied only over
    # the middle of the window (blank anchors = "no breakpoint here" — the series
    # interpolates straight through, so gaps are legitimate and shown).
    sd["tseep"] = {
        "times": [0.0, 5.0, 10.0, 20.0, 40.0],
        "series": {
            "reservoir": [100.0, 100.0, 70.0, 55.0, 55.0],
            "rain": [None, None, 1.0e-6, 1.0e-6, None],
        },
        "duration": 40.0,
        "save_interval": 5.0,
        "save_times": [12.0, 30.0],
        "stage_1": 0.0,
        "stage_2": 20.0,
    }
    # The tseep sheet + Ss/Sy columns exist only in the v18 master template, so write
    # against it explicitly (default_template_path() may still ship the prior version).
    template = os.path.join(REPO_ROOT, "docs", "inputs", "input_template.xlsx")
    save_slope_data_to_xlsx(sd, out_path, template=template)
    return out_path


BUILDERS = {
    "sheets_mat.xlsx": build_mat,
    "sheets_rapid.xlsx": build_rapid,
    "sheets_polygon.xlsx": build_polygon,
    "sheets_seepbc.xlsx": build_seepbc,
    "sheets_tseep.xlsx": build_tseep,
}


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    for name, fn in BUILDERS.items():
        out = os.path.join(OUT_DIR, name)
        fn(out)
        print("built", os.path.relpath(out, REPO_ROOT))


if __name__ == "__main__":
    main()
