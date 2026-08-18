"""Store cached results for the profile/polygon sheets' material-name formulas.

The template's ``profile`` row 6 (A6, D6, ... — one per block) holds
``XLOOKUP(<Mat ID>, mat!A:A, mat!B:B)``, and ``polygon`` row 7 holds the same
lookup guarded by the block's Type. A workbook written by xslope's writer
carries the formulas but no cached results, so Excel shows the names once it
opens the file while every other reader — ``render_sheet``, a docs screenshot,
a pandas dump — sees blank cells. This script adds the cached ``<v>`` beside
each ``<f>`` by the same sheet-XML surgery ``fill_reinforce_formula_cache.py``
uses (openpyxl cannot write a value behind a formula), leaving the formulas
byte-identical. ``fullCalcOnLoad`` still makes Excel recompute on open, so a
wrong cache could never survive a real session.

The cached names are not guessed: each block's Mat ID (and, on polygon, its
Type guard) is read from the sheet and resolved through ``xslope.fileio``'s
own loaded materials, and the loader's output is asserted identical before
and after.

Run:  python3 tools/fill_profile_name_cache.py <workbook.xlsx> [...]
"""

import contextlib
import io
import os
import re
import shutil
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

#: One name cell per block: (sheet, Mat ID cell, name-formula cell, Type cell).
#: profile has no Type guard; polygon's name shows only when Type is blank or
#: 'material' (the template's own IF).
BLOCKS = ([("profile", f"{c}5", f"{c0}6", None)
           for c0, c in (("A", "B"), ("D", "E"), ("G", "H"),
                         ("J", "K"), ("M", "N"), ("P", "Q"))]
          + [("polygon", f"{c}6", f"{c0}7", f"{c}5")
             for c0, c in (("A", "B"), ("D", "E"), ("G", "H"),
                           ("J", "K"), ("M", "N"), ("P", "Q"))])


def load_state(path):
    from xslope.fileio import load_slope_data
    with contextlib.redirect_stdout(io.StringIO()):
        sd = load_slope_data(path)
    mats = [m.get("name") for m in sd.get("materials") or []]
    lines = [len(pl) for pl in sd.get("profile_lines") or []]
    return mats, lines


def fill(path):
    import openpyxl
    before = load_state(path)
    mats, _ = before

    wb = openpyxl.load_workbook(path)          # read-only pass for the Mat IDs
    wanted = {}                                 # sheet -> {cell: name}
    for sheetname, id_cell, name_cell, type_cell in BLOCKS:
        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]
        if type_cell is not None:               # polygon's Type guard
            t = ws[type_cell].value
            if t is not None and str(t).strip().lower() not in ("", "material"):
                continue
        mid = ws[id_cell].value
        try:
            name = mats[int(mid) - 1]
        except (TypeError, ValueError, IndexError):
            continue
        if name:
            wanted.setdefault(sheetname, {})[name_cell] = str(name)
    wb.close()

    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        names = zin.namelist()
        wbxml = zin.read("xl/workbook.xml").decode("utf-8")
        rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        edited = {}                              # zip member -> new xml
        filled = []
        for sheetname, cells in wanted.items():
            m = re.search(r'<sheet[^>]*name="%s"[^>]*?r:id="rId(\d+)"' % sheetname,
                          wbxml)
            relel = re.search(r'<Relationship\b[^>]*\bId="rId%s"[^>]*/>' % m.group(1),
                              rels)
            target = re.search(r'Target="([^"]+)"', relel.group(0)).group(1).lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            sheet = edited.get(target, zin.read(target).decode("utf-8"))
            sheet, refs = _fill_sheet(sheet, cells)
            edited[target] = sheet
            filled += ["%s!%s" % (sheetname, r) for r in refs]

        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in names:
                data = (edited[n].encode("utf-8") if n in edited
                        else zin.read(n))
                zout.writestr(n, data)
    shutil.move(tmp, path)
    after = load_state(path)
    assert before == after, f"loader output changed on {path}"
    assert filled, f"{path}: nothing was filled -- check the cell matching"
    print(f"{path}: cached {', '.join(filled)}")


def _fill_sheet(sheet, cells):
    refs = []
    for ref, val in cells.items():
            pat = re.compile(r'(<c r="%s"[^>]*?)(/>|>(.*?)</c>)' % ref, re.S)
            mm = pat.search(sheet)
            if not mm:
                continue
            cell = mm.group(0)
            if "<f" not in cell:
                continue                        # not a formula cell; leave alone
            if re.search(r"<v>[^<]+</v>", cell):
                continue                        # already carries a cache
            head = mm.group(1)
            if 't="' not in head:
                head = head.rstrip() + ' t="str"'
            inner = re.search(r">(.*)</c>", cell, re.S)
            body = inner.group(1) if inner else ""
            body = re.sub(r"<v\s*/>|<v></v>", "", body)
            sheet = (sheet[:mm.start()] + head + ">" + body
                     + f"<v>{val}</v></c>" + sheet[mm.end():])
            refs.append(ref)
    return sheet, refs


if __name__ == "__main__":
    for p in sys.argv[1:]:
        fill(p)
