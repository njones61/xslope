"""Render an XSLOPE input-template worksheet to a PNG that reads like an Excel
screen capture — for the docs (``docs/usage/input_template.md``).

Why this exists
---------------
The material in ``docs/usage/input_template.md`` used to be illustrated with
hand-taken Excel screenshots. They drifted out of date every time the template
changed (new ``mat`` strength columns, the ``lloads`` sheet, reworked ``seep bc``
blocks, ...). This tool regenerates every ``docs/usage/images/sheet_*.png`` from a
manifest so a single command refreshes them all:

    python tools/render_sheet.py                 # rebuild everything in the manifest
    python tools/render_sheet.py --only mat1     # rebuild one manifest entry
    python tools/render_sheet.py --list          # list manifest entries

The manifest lives beside this file in ``tools/docs_sheets_manifest.py``; the
source sample workbooks it points at are (re)built by
``tools/build_docs_sheet_samples.py``.

Approach — pure-Python (openpyxl -> PIL), not LibreOffice
---------------------------------------------------------
Two paths were evaluated:

  (a) pure-Python: read the real workbook with openpyxl (values, theme/indexed
      colours, fonts, borders, merged cells, column widths, number formats,
      conditional-format greying) and draw the grid with PIL;
  (b) LibreOffice headless (``soffice --convert-to png``).

Path (b) was rejected for the cell sheets: ``soffice`` renders only the active
sheet, with no column-letter header row, no row-number gutter, no gridlines on
empty cells, no sheet-tab strip, and no way to crop to a column range — i.e. none
of the "chrome" that makes these read as spreadsheet captures, and no per-sheet
selection from the CLI. Path (a) gives full, deterministic control of all of it,
with no runtime dependency beyond openpyxl + Pillow (already vendored via
matplotlib for the fonts).

The one exception is the ``plot`` sheet, which is a live Excel *chart* (a
ScatterChart drawing object, not cell data) that openpyxl cannot rasterise. For
that single sheet the manifest sets ``renderer="libreoffice"`` and we fall back to
``soffice`` — patching the workbook's ``activeTab`` at the XML level (so the chart
survives, unlike an openpyxl re-save) and cropping the result. If ``soffice`` is
absent the plot image is skipped with a warning; every other sheet renders without
it.

Fidelity notes
--------------
* Theme colours are resolved from the workbook theme (``xl/theme/theme1.xml``) with
  the OOXML tint algorithm, so the black group bars, the accent-green / accent-
  purple header fills and the conditional-format grey all come out right.
* ``Symbol``-font runs (how the template stores the Greek header glyphs — a cell
  whose text is ``g`` in the Symbol font renders as gamma) are translated to the
  matching Unicode Greek letters and drawn with DejaVu Sans, which is bundled with
  matplotlib and therefore deterministic across machines.
* Columns auto-size to their contained header text so nothing is cropped, while
  free "legend" text (e.g. the strength-option descriptions) overflows into the
  empty cells to its right exactly as Excel draws it.
"""

from __future__ import annotations

import argparse
import colorsys
import importlib.util
import os
import re
import subprocess
import tempfile
import zipfile

import matplotlib
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.colors import COLOR_INDEX
from PIL import Image, ImageDraw, ImageFont

# --------------------------------------------------------------------------- #
# Paths / constants
# --------------------------------------------------------------------------- #
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMAGES_DIR = os.path.join(REPO_ROOT, "docs", "usage", "images")
SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"

_FONT_DIR = os.path.join(matplotlib.get_data_path(), "fonts", "ttf")
_FONT_REGULAR = os.path.join(_FONT_DIR, "DejaVuSans.ttf")
_FONT_BOLD = os.path.join(_FONT_DIR, "DejaVuSans-Bold.ttf")

SS = 2  # supersample factor: render at 2x then downscale for crisp anti-aliasing
PT_TO_PX = 96.0 / 72.0  # points -> screen pixels

# Excel column-/row-header (gutter) chrome
HEADER_FILL = "#E9ECEF"
HEADER_LINE = "#B7BEC6"
HEADER_TEXT = "#3C4043"
GRIDLINE = "#D4D7DC"
BORDER_DEFAULT = "#000000"
TAB_ACTIVE_FILL = "#FFFFFF"
TAB_INACTIVE_FILL = "#DDE1E6"
TAB_LINE = "#B7BEC6"
TAB_ACCENT = "#1D7A46"  # Excel-ish green underline on the active tab

# Framing convention (measured from Norm's original captures): frame to the
# formatted-content bounding box, then leave one spare empty row/column of margin.
# The sheet-tab strip is drawn at one fixed scale on every sheet (never shrunk to
# fit a narrow grid — that is what made the piezo strip render tiny).
MARGIN_COLS = 1          # spare empty columns kept to the right of content
MARGIN_ROWS = 1          # spare empty rows kept above/below content (auto-rows only)
OVERFLOW_BUFFER = 40     # empty columns probed to the right so overflow text completes
TAB_FONT_PT = 10         # fixed sheet-tab font size (points)

# Adobe "Symbol" font: ASCII letter -> Unicode Greek/misc glyph.
SYMBOL_MAP = {
    "a": "α", "b": "β", "c": "χ", "d": "δ", "e": "ε",
    "f": "φ", "g": "γ", "h": "η", "i": "ι", "j": "ϕ",
    "k": "κ", "l": "λ", "m": "μ", "n": "ν", "o": "ο",
    "p": "π", "q": "θ", "r": "ρ", "s": "σ", "t": "τ",
    "u": "υ", "v": "ϖ", "w": "ω", "x": "ξ", "y": "ψ",
    "z": "ζ",
    "A": "Α", "B": "Β", "C": "Χ", "D": "Δ", "E": "Ε",
    "F": "Φ", "G": "Γ", "H": "Η", "I": "Ι", "J": "ϑ",
    "K": "Κ", "L": "Λ", "M": "Μ", "N": "Ν", "O": "Ο",
    "P": "Π", "Q": "Θ", "R": "Ρ", "S": "Σ", "T": "Τ",
    "U": "Υ", "V": "ς", "W": "Ω", "X": "Ξ", "Y": "Ψ",
    "Z": "Ζ",
}

# Standard Office 2013+ theme, used only if a workbook has no embedded theme.
_THEME_FALLBACK = [
    "FFFFFF", "000000", "E8E8E8", "0E2841", "156082", "E97132",
    "196B24", "0F9ED5", "A02B93", "4EA72E", "467886", "96607D",
]

_FONT_CACHE: dict = {}


def _font(size_px: int, bold: bool) -> ImageFont.FreeTypeFont:
    key = (size_px, bold)
    if key not in _FONT_CACHE:
        path = _FONT_BOLD if bold else _FONT_REGULAR
        _FONT_CACHE[key] = ImageFont.truetype(path, size_px)
    return _FONT_CACHE[key]


# --------------------------------------------------------------------------- #
# Colour resolution (rgb / theme+tint / indexed)
# --------------------------------------------------------------------------- #
def load_theme_colors(wb) -> list:
    """Return the 12 workbook theme colours as 'RRGGBB' hex, indexed the way cell
    ``theme=`` attributes reference them (0=lt1, 1=dk1, 2=lt2, 3=dk2, 4..9=accent1..6,
    10=hlink, 11=folHlink)."""
    xml = getattr(wb, "loaded_theme", None)
    if not xml:
        return list(_THEME_FALLBACK)
    if isinstance(xml, bytes):
        xml = xml.decode("utf-8", "replace")
    m = re.search(r"<a:clrScheme.*?</a:clrScheme>", xml, re.S)
    if not m:
        return list(_THEME_FALLBACK)
    block = m.group(0)
    scheme = {}
    for name, body in re.findall(
        r"<a:(dk1|lt1|dk2|lt2|accent[1-6]|hlink|folHlink)>(.*?)</a:\1>", block, re.S
    ):
        srgb = re.search(r'srgbClr val="([0-9A-Fa-f]{6})"', body)
        if srgb:
            scheme[name] = srgb.group(1).upper()
            continue
        sysc = re.search(r'sysClr[^>]*lastClr="([0-9A-Fa-f]{6})"', body)
        if sysc:
            scheme[name] = sysc.group(1).upper()
    order = ["lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3",
             "accent4", "accent5", "accent6", "hlink", "folHlink"]
    return [scheme.get(n, _THEME_FALLBACK[i]) for i, n in enumerate(order)]


def _apply_tint(hex6: str, tint: float) -> str:
    r, g, b = int(hex6[0:2], 16) / 255, int(hex6[2:4], 16) / 255, int(hex6[4:6], 16) / 255
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        l = l * (1 + tint)
    else:
        l = l * (1 - tint) + tint
    l = max(0.0, min(1.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return "#%02X%02X%02X" % (round(r * 255), round(g * 255), round(b * 255))


def resolve_color(color, theme_colors):
    """openpyxl Color -> '#RRGGBB' or None (None = no fill / automatic)."""
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        rgb = color.rgb
        if not isinstance(rgb, str):
            return None
        if len(rgb) == 8:
            if rgb[:2] == "00":  # fully transparent -> automatic/none
                return None
            return "#" + rgb[2:]
        if len(rgb) == 6:
            return "#" + rgb
        return None
    if ctype == "theme":
        idx = color.theme
        if not isinstance(idx, int) or idx >= len(theme_colors):
            return None
        base = theme_colors[idx]
        return _apply_tint(base, getattr(color, "tint", 0.0) or 0.0)
    if ctype == "indexed":
        idx = color.indexed
        if not isinstance(idx, int):
            return None
        if idx in (64, 65):  # system foreground / background -> automatic
            return None
        try:
            rgb = COLOR_INDEX[idx]
        except (IndexError, TypeError):
            return None
        if isinstance(rgb, str) and len(rgb) == 8:
            return "#" + rgb[2:]
        return None
    return None


def _fill_color(cell, theme_colors):
    fill = cell.fill
    if not fill or fill.patternType not in ("solid",):
        return None
    return resolve_color(fill.fgColor, theme_colors)


# --------------------------------------------------------------------------- #
# Number formatting (a pragmatic subset of Excel number formats)
# --------------------------------------------------------------------------- #
def _format_number(value, fmt):
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    fmt = (fmt or "General").strip()
    if fmt.lower() == "general" or fmt == "@":
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
        return ("%g" % value)
    # take the positive section only
    section = fmt.split(";")[0]
    is_percent = "%" in section
    v = value * 100 if is_percent else value
    core = re.sub(r'[^0#.,]', '', section)
    if "." in core:
        decimals = len(core.split(".")[1].replace(",", ""))
    else:
        decimals = 0
    thousands = "," in core.split(".")[0]
    try:
        if thousands:
            s = ("{:,.%df}" % decimals).format(v)
        else:
            s = ("{:.%df}" % decimals).format(v)
    except (ValueError, TypeError):
        s = str(v)
    if is_percent:
        s += "%"
    return s


def display_string(cell, value):
    """A cell's value as the text Excel would show (number format applied)."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        return _format_number(value, cell.number_format)
    return str(value)


# --------------------------------------------------------------------------- #
# Rich text -> runs (text, font_name, bold, color)
# --------------------------------------------------------------------------- #
def _cell_runs(cell, value, theme_colors):
    base_name = cell.font.name if cell.font else None
    base_bold = bool(cell.font.bold) if cell.font else False
    base_color = resolve_color(cell.font.color, theme_colors) if cell.font else None

    runs = []
    # CellRichText: a list of str and TextBlock
    if not isinstance(value, (str, int, float, bool)) and value is not None and hasattr(value, "__iter__"):
        for item in value:
            if isinstance(item, str):
                runs.append((item, base_name, base_bold, base_color))
            else:
                f = getattr(item, "font", None)
                name = getattr(f, "rFont", None) or base_name
                bold = bool(getattr(f, "b", None)) or base_bold
                col = getattr(f, "color", None)
                color = resolve_color(col, theme_colors) if col is not None else base_color
                runs.append((getattr(item, "text", ""), name, bold, color))
    else:
        runs.append((display_string(cell, value), base_name, base_bold, base_color))

    # apply Symbol-font translation
    out = []
    for text, name, bold, color in runs:
        if name == "Symbol" and text:
            text = "".join(SYMBOL_MAP.get(ch, ch) for ch in text)
        out.append((text, bold, color))
    return out


def _runs_width(runs, size_px):
    w = 0
    for text, bold, _ in runs:
        if not text:
            continue
        f = _font(size_px, bold)
        w += f.getbbox(text)[2]
    return w


# --------------------------------------------------------------------------- #
# Conditional formatting: evaluate the template's greying rules
# --------------------------------------------------------------------------- #
def _split_top_commas(s):
    out, depth, cur = [], 0, ""
    for ch in s:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "," and depth == 0:
            out.append(cur)
            cur = ""
        else:
            cur += ch
    if cur.strip():
        out.append(cur)
    return out


def _eval_cf_expr(expr, getcol):
    """Evaluate the small boolean grammar used by the template's CF rules:
    AND(...), OR(...), and `$COL<row> = / <> "literal"`.

    getcol(colletter, row=None) -> value. A row-RELATIVE reference (``$E11`` in a
    rule that spans many rows, the mat/circles idiom) is read from the row being
    drawn, so ``row`` is None; a row-ANCHORED reference (``$B$5`` — the polygon
    sheet's per-block option cell, which sits in a fixed row above the cells it
    controls) is read from that literal row instead."""
    expr = expr.strip()
    if expr.startswith("="):
        expr = expr[1:].strip()
    m = re.match(r"^(AND|OR)\((.*)\)$", expr, re.S)
    if m:
        op, args = m.group(1), _split_top_commas(m.group(2))
        vals = [_eval_cf_expr(a, getcol) for a in args]
        return all(vals) if op == "AND" else any(vals)
    m = re.match(r'^\$?([A-Za-z]+)(\$?)(\d+)\s*(<>|=)\s*"(.*)"$', expr)
    if m:
        col, anchored, ref_row = m.group(1).upper(), m.group(2), int(m.group(3))
        op, lit = m.group(4), m.group(5)
        cur = getcol(col, ref_row if anchored else None)
        cur_s = "" if cur is None else str(cur).strip()
        if op == "=":
            return cur_s.lower() == lit.lower()
        return cur_s.lower() != lit.lower()
    raise ValueError("unparseable CF expression: %r" % expr)


def build_cf_rules(ws, theme_colors):
    """List of (MultiCellRange, formula, fill_hex) for expression rules we can draw."""
    rules = []
    cf = ws.conditional_formatting
    for rng in cf:
        for rule in cf[rng]:
            if rule.type != "expression" or not rule.formula:
                continue
            dxf = rule.dxf
            if not dxf or not dxf.fill:
                continue
            fill = dxf.fill
            color = None
            for cand in (fill.bgColor, fill.fgColor):
                color = resolve_color(cand, theme_colors)
                if color:
                    break
            if not color:
                continue
            rules.append((rng, rule.formula[0], color))
    return rules


def cf_fill_for(r, c, rules, value_at):
    """Grey (or other CF) fill for cell (r,c), or None. value_at(row,col)->value."""
    ref = "%s%d" % (get_column_letter(c), r)
    for rng, formula, color in rules:
        if ref not in rng:
            continue
        try:
            ok = _eval_cf_expr(
                formula,
                lambda col, row=None: value_at(row if row else r,
                                               column_index_from_string(col)))
        except ValueError:
            continue
        if ok:
            return color
    return None


# --------------------------------------------------------------------------- #
# Geometry helpers
# --------------------------------------------------------------------------- #
def col_width_px(ws, c, default_char_width):
    cd = ws.column_dimensions.get(get_column_letter(c))
    w = cd.width if (cd and cd.width) else default_char_width
    return int(round(w * 7)) + 5


def row_height_px(ws, r, default_row_height):
    rd = ws.row_dimensions.get(r)
    h = rd.height if (rd and rd.height) else default_row_height
    return int(round(h * PT_TO_PX))


# --------------------------------------------------------------------------- #
# Core render
# --------------------------------------------------------------------------- #
def _parse_cols(spec):
    """'A:U' or 'A' -> list of 1-based column indices."""
    out = []
    for part in str(spec).split(","):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            a, b = part.split(":")
            out.extend(range(column_index_from_string(a), column_index_from_string(b) + 1))
        else:
            out.append(column_index_from_string(part))
    return out


def _is_wrapped(cell):
    a = cell.alignment
    return bool(a and a.wrapText)


def render_sheet(xlsx_path, sheet, out_path, rows=None, cols=None,
                 identity_cols=None, base_font_pt=11):
    """Render one worksheet to a PNG that reads like an Excel screen capture.

    ``rows`` / ``cols`` select a region of interest (needed for the wide ``mat``
    sheet's three views and the ``seep bc`` split). Within that region the frame is
    computed from the *formatted-content* bounding box (any cell with a value, fill,
    border or merge) plus one spare row/column of margin, and the right edge is
    auto-extended so Excel-style overflow text (the option/help legends) renders
    complete. ``rows`` is honoured verbatim when given (so the three ``mat`` views
    stay row-aligned); otherwise rows auto-frame to content too.
    """
    wb_f = openpyxl.load_workbook(xlsx_path, data_only=False, rich_text=True)
    wb_v = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet not in wb_f.sheetnames:
        raise ValueError("sheet %r not in %s" % (sheet, xlsx_path))
    ws = wb_f[sheet]
    ws_v = wb_v[sheet]
    theme = load_theme_colors(wb_f)

    default_cw = ws.sheet_format.defaultColWidth or 8.43
    default_rh = ws.sheet_format.defaultRowHeight or 15.0

    # ----- resolved value + display caches --------------------------------- #
    def value_at(r, c):
        raw = ws.cell(row=r, column=c).value
        if isinstance(raw, str) and raw.startswith("="):
            cached = ws_v.cell(row=r, column=c).value
            return cached
        return raw

    def display_at(r, c):
        cell = ws.cell(row=r, column=c)
        return display_string(cell, value_at(r, c))

    fs_px = int(round(base_font_pt * PT_TO_PX * SS))
    pad = 6 * SS
    pad_in = 3 * SS

    # ----- merged cells ---------------------------------------------------- #
    merged_anchor = {}   # (r,c) -> (min_r,min_c,max_r,max_c) for the top-left cell
    merged_covered = set()  # (r,c) covered by a merge but NOT the top-left
    for m in ws.merged_cells.ranges:
        for rr in range(m.min_row, m.max_row + 1):
            for cc in range(m.min_col, m.max_col + 1):
                if (rr, cc) == (m.min_row, m.min_col):
                    merged_anchor[(rr, cc)] = (m.min_row, m.min_col, m.max_row, m.max_col)
                else:
                    merged_covered.add((rr, cc))

    cf_rules = build_cf_rules(ws, theme)

    def align_of(cell, val, merged):
        h = cell.alignment.horizontal if cell.alignment else None
        if h in ("left", "center", "right"):
            return h
        if merged:
            return "center"
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return "right"
        return "left"

    def cell_visible(r, c):
        """True if the cell belongs to the formatted-content bbox: it carries a
        value, a solid fill, any styled border, a conditional-format fill, or is
        part of a merge. Empty-but-bordered table cells count, so the frame keeps
        the whole formatted table (as Norm's captures do), not just filled cells."""
        if (r, c) in merged_anchor or (r, c) in merged_covered:
            return True
        cell = ws.cell(row=r, column=c)
        if value_at(r, c) not in (None, ""):
            return True
        if _fill_color(cell, theme) is not None:
            return True
        b = cell.border
        if any(getattr(b, s) and getattr(b, s).style
               for s in ("left", "right", "top", "bottom")):
            return True
        if cf_fill_for(r, c, cf_rules, value_at) is not None:
            return True
        return False

    # ----- region of interest --------------------------------------------- #
    if cols:
        roi_cols = _parse_cols(cols)
    else:
        roi_cols = list(range(1, ws.max_column + 1))
    ident = _parse_cols(identity_cols) if identity_cols else []
    if ident:
        roi_cols = ident + [c for c in roi_cols if c not in ident]
    roi_rows = list(range(rows[0], rows[1] + 1)) if rows else list(range(1, ws.max_row + 1))

    # ----- formatted-content bbox within the ROI -------------------------- #
    content_cols = [c for c in roi_cols if any(cell_visible(r, c) for r in roi_rows)]
    content_rows = [r for r in roi_rows if any(cell_visible(r, c) for c in roi_cols)]
    if not content_cols:
        content_cols = list(roi_cols)
    if not content_rows:
        content_rows = list(roi_rows)

    # ----- final row list -------------------------------------------------- #
    if rows:
        row_list = list(roi_rows)                    # explicit rows: honour verbatim
    else:
        top = max(1, min(content_rows) - MARGIN_ROWS)
        row_list = list(range(top, max(content_rows) + MARGIN_ROWS + 1))

    # ----- candidate columns: ROI trimmed to content, + a right buffer ----- #
    cmax = max(content_cols)
    left_edge = min(roi_cols)                        # keep ROI's left (1-col margin)
    base_cols = [c for c in roi_cols if left_edge <= c <= cmax]
    gmax = max(base_cols)
    cand_cols = base_cols + list(range(gmax + 1, gmax + 1 + OVERFLOW_BUFFER))

    # ----- base widths + "contained text" autofit -------------------------- #
    col_px = {c: col_width_px(ws, c, default_cw) * SS for c in cand_cols}
    for c in base_cols:
        need = col_px[c]
        for r in row_list:
            if (r, c) in merged_covered:
                continue
            anc = merged_anchor.get((r, c))
            if anc and (anc[3] - anc[1]) > 0:
                continue  # multi-column merge: doesn't drive single-col width
            val = value_at(r, c)
            if val in (None, ""):
                continue
            cell = ws.cell(row=r, column=c)
            is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
            has_fill = _fill_color(cell, theme) is not None
            right_full = value_at(r, c + 1) not in (None, "")
            # Numbers never overflow in Excel (they show ### if too narrow), so a
            # numeric cell always drives its column width; text only does so when it
            # is "contained" (a fill, or a non-empty right neighbour clips it).
            if not (is_num or has_fill or right_full):
                continue
            runs = _cell_runs(cell, val, theme)
            need = max(need, _runs_width(runs, fs_px) + 2 * pad)
        col_px[c] = int(need)

    # A merged header whose (centered) text is wider than its span is clipped at the
    # merge boundary; grow the span's columns so it fits — the multi-column analogue
    # of the single-column autofit above (e.g. "Piezometric Line #1" over A:B).
    for (ar, ac), (mnr, mnc, mxr, mxc) in merged_anchor.items():
        if mxc <= mnc or ar not in row_list:
            continue
        cols_in = [c for c in range(mnc, mxc + 1) if c in col_px and c in base_cols]
        if len(cols_in) < 2:
            continue
        val = value_at(ar, ac)
        if val in (None, ""):
            continue
        need = _runs_width(_cell_runs(ws.cell(row=ar, column=ac), val, theme), fs_px) + 2 * pad
        have = sum(col_px[c] for c in cols_in)
        if need > have:
            per = (need - have) // len(cols_in) + 1
            for c in cols_in:
                col_px[c] += per

    # ----- Excel-style horizontal overflow: complete every legend string ---- #
    # Gather left-aligned, non-wrapped text cells whose right neighbour is empty
    # (so Excel would let the text flow rightward). Record the first non-empty cell
    # downstream (the blocker) if any.
    overflow = []
    for r in row_list:
        for c in base_cols:
            if (r, c) in merged_covered or (r, c) in merged_anchor:
                continue
            val = value_at(r, c)
            if val in (None, "") or (isinstance(val, (int, float)) and not isinstance(val, bool)):
                continue
            cell = ws.cell(row=r, column=c)
            if _is_wrapped(cell) or align_of(cell, val, False) != "left":
                continue
            if value_at(r, c + 1) not in (None, ""):
                continue
            tw = _runs_width(_cell_runs(cell, val, theme), fs_px)
            blocker = next((nc for nc in cand_cols
                            if nc > c and value_at(r, nc) not in (None, "")), None)
            overflow.append((r, c, tw, blocker))

    def _cumx():
        xs, x = {}, 0
        for c in cand_cols:
            xs[c] = x
            x += col_px[c]
        return xs

    # (a) BLOCKED overflow (a non-empty cell lies downstream, e.g. mat's pow/hb help
    #     capped by the pore-pressure column): widen the empty spanned columns so the
    #     string fits before the blocker — reproducing Excel, where the real
    #     (condensed) font fits it in that span.
    for _ in range(6):
        xs = _cumx()
        changed = False
        for (r, c, tw, blocker) in overflow:
            if blocker is None:
                continue
            deficit = (xs[c] + pad_in + tw + pad_in) - xs[blocker]
            span = [k for k in cand_cols if c < k < blocker]
            if deficit > 1 and span:
                per = int(deficit // len(span)) + 1
                for k in span:
                    col_px[k] += per
                changed = True
        if not changed:
            break

    # (b) UNBLOCKED overflow (nothing downstream — e.g. the piezo/circles/seep-bc
    #     legends): extend the frame rightward with empty columns until the string
    #     ends, then keep the standard margin.
    xs = _cumx()
    needed_right = xs[gmax] + col_px[gmax]
    for (r, c, tw, blocker) in overflow:
        if blocker is None:
            needed_right = max(needed_right, xs[c] + pad_in + tw + pad_in)
    extent = max((c for c in cand_cols if xs[c] < needed_right - 1), default=gmax)

    col_list = [c for c in cand_cols if c <= extent]
    # Columns pulled in ONLY to let an unblocked string spill (above) sit outside the
    # manifest's region of interest, so they render as chrome — gridlines and a column
    # letter — exactly like the tab-strip fill below. Without this, a long header note
    # drags the next unused table block into the frame and it appears half-drawn at the
    # right edge (the polygon sheet's row-2 Type/Size note reaches column AA and would
    # otherwise expose a clipped, empty "Polygon #9").
    overflow_filler = {c for c in col_list if c > gmax}
    for i in range(MARGIN_COLS):                      # spare empty margin column(s)
        nc = extent + 1 + i
        if nc in col_px and not any(cell_visible(r, nc) for r in row_list):
            col_list.append(nc)
        else:
            break

    row_px = {r: row_height_px(ws, r, default_rh) * SS for r in row_list}

    # ----- canvas layout --------------------------------------------------- #
    gutter_w = _font(fs_px, False).getbbox(str(max(row_list)))[2] + 10 * SS
    header_h = int(round(default_rh * PT_TO_PX * SS))
    grid_w = gutter_w + sum(col_px[c] for c in col_list)
    grid_h = header_h + sum(row_px[r] for r in row_list)

    # ----- fill a narrow grid with empty gridlined columns ----------------- #
    # The sheet-tab strip is drawn at one fixed scale on every sheet. When the grid
    # is narrower than that strip, a real Excel window shows no white void to the
    # right of the last content column: it just continues empty, gridlined columns
    # (each with its column letter) out to the window edge. Reproduce that — append
    # default-width columns (continuing the column letters) until the grid is at
    # least as wide as the tab strip, so the grid, not a blank void, meets it. These
    # filler columns render as pure chrome — gridlines + column letter only, never
    # any value/fill/border/merge, even where the underlying sheet holds data there
    # (e.g. a mat view's next-view columns): `filler_cols` marks them so the content
    # passes below skip them. Wide grids append nothing and stay byte-identical.
    tab_strip_w = _tab_strip_width(wb_f.sheetnames, sheet)
    filler_cols = set(overflow_filler)
    if grid_w < tab_strip_w:
        nc = max(col_list) + 1
        while grid_w < tab_strip_w:
            col_px[nc] = col_width_px(ws, nc, default_cw) * SS
            col_list.append(nc)
            filler_cols.add(nc)
            grid_w += col_px[nc]
            nc += 1

    tab_h = int(round(20 * PT_TO_PX * SS))
    tab_gap = 6 * SS
    total_h = grid_h + tab_gap + tab_h
    # After the fill above, grid_w >= tab_strip_w for narrow sheets too, so the grid
    # spans the whole canvas and the tab strip meets gridlined columns, not a void.
    canvas_w = max(grid_w, tab_strip_w)

    img = Image.new("RGB", (canvas_w, total_h), "#FFFFFF")
    d = ImageDraw.Draw(img)

    # x/y pixel positions of each column/row boundary
    xs = {}
    x = gutter_w
    for c in col_list:
        xs[c] = x
        x += col_px[c]
    ys = {}
    y = header_h
    for r in row_list:
        ys[r] = y
        y += row_px[r]

    def cell_box(r, c):
        return xs[c], ys[r], xs[c] + col_px[c], ys[r] + row_px[r]

    # ----- 1. light gridlines over the whole cell area --------------------- #
    for c in col_list:
        d.line([(xs[c], header_h), (xs[c], grid_h)], fill=GRIDLINE, width=SS)
    d.line([(grid_w - 1, header_h), (grid_w - 1, grid_h)], fill=GRIDLINE, width=SS)
    for r in row_list:
        d.line([(gutter_w, ys[r]), (grid_w, ys[r])], fill=GRIDLINE, width=SS)
    d.line([(gutter_w, grid_h - 1), (grid_w, grid_h - 1)], fill=GRIDLINE, width=SS)

    # ----- 2. fills (static, then conditional) ----------------------------- #
    for r in row_list:
        for c in col_list:
            if c in filler_cols or (r, c) in merged_covered:
                continue
            cell = ws.cell(row=r, column=c)
            # Excel draws a conditional format ON TOP of the cell's own fill, so
            # the CF rule is asked first and the static fill is the fallback —
            # otherwise a greying rule over an already-filled cell (the polygon
            # sheet's Mat ID / name cells) would never show.
            color = cf_fill_for(r, c, cf_rules, value_at)
            if color is None:
                color = _fill_color(cell, theme)
            if color is None:
                continue
            anc = merged_anchor.get((r, c))
            if anc:
                x0, y0 = xs[anc[1]], ys[anc[0]]
                x1 = xs[anc[3]] + col_px.get(anc[3], 0) if anc[3] in col_px else xs[c] + col_px[c]
                y1 = ys[anc[2]] + row_px.get(anc[2], 0) if anc[2] in row_px else ys[r] + row_px[r]
                d.rectangle([x0, y0, x1, y1], fill=color)
            else:
                x0, y0, x1, y1 = cell_box(r, c)
                d.rectangle([x0, y0, x1, y1], fill=color)

    # ----- 3. explicit cell borders --------------------------------------- #
    def draw_borders(r, c, box):
        cell = ws.cell(row=r, column=c)
        b = cell.border
        x0, y0, x1, y1 = box
        for side, pts in (("left", [(x0, y0), (x0, y1)]),
                          ("right", [(x1, y0), (x1, y1)]),
                          ("top", [(x0, y0), (x1, y0)]),
                          ("bottom", [(x0, y1), (x1, y1)])):
            s = getattr(b, side)
            if s and s.style:
                col = resolve_color(s.color, theme) or BORDER_DEFAULT
                w = SS * (2 if s.style in ("medium", "thick") else 1)
                d.line(pts, fill=col, width=w)

    for r in row_list:
        for c in col_list:
            if c in filler_cols or (r, c) in merged_covered:
                continue
            anc = merged_anchor.get((r, c))
            if anc:
                box = (xs[anc[1]], ys[anc[0]],
                       xs[anc[3]] + col_px.get(anc[3], 0), ys[anc[2]] + row_px.get(anc[2], 0))
            else:
                box = cell_box(r, c)
            draw_borders(r, c, box)

    # ----- 4. text (with Excel-style horizontal overflow) ------------------ #
    def text_for(r, c):
        cell = ws.cell(row=r, column=c)
        val = value_at(r, c)
        if val in (None, ""):
            return None, cell
        return _cell_runs(cell, val, theme), cell

    for r in row_list:
        for c in col_list:
            if c in filler_cols or (r, c) in merged_covered:
                continue
            runs, cell = text_for(r, c)
            if not runs:
                continue
            val = value_at(r, c)
            anc = merged_anchor.get((r, c))
            merged = anc is not None
            align = align_of(cell, val, merged)
            if merged:
                x0, y0 = xs[anc[1]], ys[anc[0]]
                x1 = xs[anc[3]] + col_px.get(anc[3], 0)
                y1 = ys[anc[2]] + row_px.get(anc[2], 0)
                clip_x0, clip_x1 = x0, x1
            else:
                x0, y0, x1, y1 = cell_box(r, c)
                # Excel lets non-wrapped text overflow across EMPTY neighbours until a
                # non-empty cell: left/general/centre rightward, right/centre leftward.
                # Numbers never overflow (they print ### instead).
                clip_x0, clip_x1 = x0, x1
                is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
                if not is_num and not _is_wrapped(cell):
                    idx = col_list.index(c)
                    if align in ("left", "center"):
                        for nc in col_list[idx + 1:]:
                            if value_at(r, nc) not in (None, ""):
                                break
                            clip_x1 += col_px[nc]
                    if align in ("right", "center"):
                        for nc in reversed(col_list[:idx]):
                            if value_at(r, nc) not in (None, ""):
                                break
                            clip_x0 -= col_px[nc]

            tw = _runs_width(runs, fs_px)
            th = max(_font(fs_px, b).getbbox(t or "x")[3] for t, b, _ in runs)
            if align == "left":
                tx = x0 + pad_in
            elif align == "right":
                tx = x1 - pad_in - tw
            else:
                tx = x0 + (x1 - x0 - tw) / 2
            ty = y0 + (y1 - y0 - th) / 2

            # Draw onto a per-cell RGBA layer and composite only the clip window so
            # text never paints over a neighbour beyond where Excel would allow it.
            clip_l = int(min(clip_x0, x0))
            clip_r = int(max(clip_x1, x1))
            cw = max(1, clip_r - clip_l)
            layer = Image.new("RGBA", (cw, int(y1 - y0)), (0, 0, 0, 0))
            ld = ImageDraw.Draw(layer)
            cx = tx - clip_l
            for text, bold, color in runs:
                if not text:
                    continue
                f = _font(fs_px, bold)
                ld.text((cx, ty - y0), text, font=f, fill=(color or BORDER_DEFAULT))
                cx += f.getbbox(text)[2]
            img.paste(layer, (clip_l, int(y0)), layer)

    # ----- 5. column-letter header row + row-number gutter ----------------- #
    d.rectangle([0, 0, grid_w, header_h], fill=HEADER_FILL)
    d.rectangle([0, 0, gutter_w, grid_h], fill=HEADER_FILL)
    hfont = _font(int(fs_px * 0.92), False)
    for c in col_list:
        x0 = xs[c]
        d.line([(x0, 0), (x0, header_h)], fill=HEADER_LINE, width=SS)
        letter = get_column_letter(c)
        lw = hfont.getbbox(letter)[2]
        lh = hfont.getbbox(letter)[3]
        d.text((x0 + (col_px[c] - lw) / 2, (header_h - lh) / 2 - SS), letter,
               font=hfont, fill=HEADER_TEXT)
    for r in row_list:
        y0 = ys[r]
        d.line([(0, y0), (gutter_w, y0)], fill=HEADER_LINE, width=SS)
        num = str(r)
        nw = hfont.getbbox(num)[2]
        nh = hfont.getbbox(num)[3]
        d.text((gutter_w - nw - 4 * SS, y0 + (row_px[r] - nh) / 2 - SS), num,
               font=hfont, fill=HEADER_TEXT)
    d.line([(gutter_w, 0), (gutter_w, grid_h)], fill=HEADER_LINE, width=SS)
    d.line([(0, header_h), (grid_w, header_h)], fill=HEADER_LINE, width=SS)
    d.rectangle([0, 0, gutter_w, header_h], fill=HEADER_FILL)
    d.line([(0, header_h), (gutter_w, header_h)], fill=HEADER_LINE, width=SS)
    d.line([(gutter_w, 0), (gutter_w, header_h)], fill=HEADER_LINE, width=SS)

    # ----- 6. sheet-tab strip --------------------------------------------- #
    _draw_tab_strip(d, wb_f.sheetnames, sheet, grid_h + tab_gap, tab_h)

    # ----- downscale + save ------------------------------------------------ #
    final = img.resize((canvas_w // SS, total_h // SS), Image.LANCZOS)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    final.save(out_path)
    return out_path


def _tab_strip_width(sheetnames, active):
    """Total pixel width of the sheet-tab strip at the fixed tab scale."""
    fs = int(round(TAB_FONT_PT * PT_TO_PX * SS))
    pad = 8 * SS
    total = 2 * SS
    for name in sheetnames:
        total += _font(fs, name == active).getbbox(name)[2] + 2 * pad
    return total


def _draw_tab_strip(d, sheetnames, active, y0, height):
    fs = int(round(TAB_FONT_PT * PT_TO_PX * SS))
    pad = 8 * SS
    f = _font(fs, False)
    fb = _font(fs, True)
    x = 2 * SS
    for name in sheetnames:
        is_active = (name == active)
        ff = fb if is_active else f
        tw = ff.getbbox(name)[2]
        w = tw + 2 * pad
        fill = TAB_ACTIVE_FILL if is_active else TAB_INACTIVE_FILL
        d.rectangle([x, y0, x + w, y0 + height], fill=fill, outline=TAB_LINE, width=SS)
        if is_active:
            d.rectangle([x, y0, x + w, y0 + 3 * SS], fill=TAB_ACCENT)
        th = ff.getbbox(name)[3]
        d.text((x + pad, y0 + (height - th) / 2 - SS), name, font=ff,
               fill="#111111" if is_active else "#4A4A4A")
        x += w


# --------------------------------------------------------------------------- #
# LibreOffice fallback (for the live-chart 'plot' sheet)
# --------------------------------------------------------------------------- #
def render_via_libreoffice(xlsx_path, sheet, out_path):
    if not os.path.exists(SOFFICE):
        print("  ! LibreOffice not found at %s; skipping %s" % (SOFFICE, out_path))
        return None
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        raise ValueError("sheet %r not in %s" % (sheet, xlsx_path))
    idx = wb.sheetnames.index(sheet)
    with tempfile.TemporaryDirectory() as td:
        patched = os.path.join(td, "patched.xlsx")
        _patch_active_tab(xlsx_path, patched, idx)
        subprocess.run([SOFFICE, "--headless", "--convert-to", "png",
                        "--outdir", td, patched],
                       check=True, capture_output=True)
        png = os.path.join(td, "patched.png")
        if not os.path.exists(png):
            print("  ! soffice produced no PNG for %s" % out_path)
            return None
        im = Image.open(png).convert("RGB")
        im = _autocrop(im)
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        im.save(out_path)
    return out_path


def _patch_active_tab(src, dst, idx):
    """Copy the xlsx and make sheet `idx` the FIRST sheet, at the zip/XML level —
    without an openpyxl re-save, which would drop the chart drawing. LibreOffice's
    image export always rasterises the first sheet regardless of activeTab /
    tabSelected, so reordering the ``<sheet>`` list in workbook.xml (which only
    changes tab order, not the chart's worksheet relationships) is what actually
    selects it."""
    with zipfile.ZipFile(src) as zin:
        names = zin.namelist()
        data = {n: zin.read(n) for n in names}

    wbxml = data["xl/workbook.xml"].decode("utf-8")
    m = re.search(r"(<sheets>)(.*?)(</sheets>)", wbxml, re.S)
    if m:
        sheet_elems = re.findall(r"<sheet\b[^>]*/>", m.group(2))
        if 0 <= idx < len(sheet_elems):
            reordered = [sheet_elems[idx]] + sheet_elems[:idx] + sheet_elems[idx + 1:]
            wbxml = wbxml[:m.start(2)] + "".join(reordered) + wbxml[m.end(2):]
    # point activeTab at the (now first) target sheet
    if "<workbookView" in wbxml:
        wbxml = re.sub(r'\s*activeTab="\d+"', "", wbxml)
        wbxml = wbxml.replace("<workbookView", '<workbookView activeTab="0"', 1)
    data["xl/workbook.xml"] = wbxml.encode("utf-8")

    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, data[n])


def _autocrop(im, bg=(255, 255, 255), margin=8):
    from PIL import ImageChops
    diff = ImageChops.difference(im, Image.new("RGB", im.size, bg))
    bbox = diff.getbbox()
    if not bbox:
        return im
    l, t, r, b = bbox
    l, t = max(0, l - margin), max(0, t - margin)
    r, b = min(im.width, r + margin), min(im.height, b + margin)
    return im.crop((l, t, r, b))


# --------------------------------------------------------------------------- #
# Manifest driver
# --------------------------------------------------------------------------- #
def load_manifest():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "docs_sheets_manifest.py")
    spec = importlib.util.spec_from_file_location("docs_sheets_manifest", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.SHEETS


def build_one(entry):
    src = entry["src"]
    if not os.path.isabs(src):
        src = os.path.join(REPO_ROOT, src)
    out = os.path.join(IMAGES_DIR, entry["out"])
    renderer = entry.get("renderer", "grid")
    if renderer == "manual":
        # A hand-taken capture that must not be regenerated (e.g. the live Excel
        # chart on the 'plot' sheet, which LibreOffice rasterises badly).
        print("-> %-22s  (keep-manual: %s)" % (entry["out"], entry.get("note", "manual capture")))
        return out
    print("-> %-22s  %s!%s" % (entry["out"], os.path.relpath(src, REPO_ROOT), entry["sheet"]))
    if renderer == "libreoffice":
        return render_via_libreoffice(src, entry["sheet"], out)
    return render_sheet(src, entry["sheet"], out,
                        rows=entry.get("rows"), cols=entry.get("cols"),
                        identity_cols=entry.get("identity_cols"))


def main(argv=None):
    ap = argparse.ArgumentParser(description="Render XSLOPE template sheets to docs PNGs.")
    ap.add_argument("--only", help="build a single manifest entry by its 'out' key "
                                    "(with or without the .png / sheet_ affixes)")
    ap.add_argument("--list", action="store_true", help="list manifest entries and exit")
    ap.add_argument("--xlsx", help="ad-hoc: render this workbook")
    ap.add_argument("--sheet", help="ad-hoc: sheet name (with --xlsx)")
    ap.add_argument("--out", help="ad-hoc: output PNG path (with --xlsx)")
    ap.add_argument("--rows", help="ad-hoc: 'r0:r1'")
    ap.add_argument("--cols", help="ad-hoc: 'A:U'")
    ap.add_argument("--identity-cols", help="ad-hoc: 'A:B'")
    args = ap.parse_args(argv)

    if args.xlsx:
        rows = tuple(int(x) for x in args.rows.split(":")) if args.rows else None
        render_sheet(args.xlsx, args.sheet, args.out, rows=rows, cols=args.cols,
                     identity_cols=args.identity_cols)
        print("wrote", args.out)
        return

    manifest = load_manifest()
    if args.list:
        for e in manifest:
            print("%-22s <- %s!%s" % (e["out"], e["src"], e["sheet"]))
        return

    if args.only:
        key = args.only.replace("sheet_", "").replace(".png", "")
        manifest = [e for e in manifest
                    if e["out"].replace("sheet_", "").replace(".png", "") == key]
        if not manifest:
            print("no manifest entry matching %r" % args.only)
            return

    for e in manifest:
        build_one(e)
    print("done (%d image%s)" % (len(manifest), "" if len(manifest) == 1 else "s"))


if __name__ == "__main__":
    main()
