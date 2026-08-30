"""Set the support-type preset on a workbook's ``reinforce`` rows.

The ``Type`` column (``reinforce!G``) is an input; ``Dir`` (H) and ``Appl`` (I)
are formula columns that look the type up in the sheet's own preset block, so a
row that names its support type does not spell out the two settings the preset
implies. A model whose lines are geogrids should say *Geosynthetic* rather than
leave the column empty and rely on the loader's blank default — the two produce
the same forces, but only one of them says what the reinforcement is.

Writes column G and nothing else: H and I keep their formulas untouched, and the
workbook is marked for a full recalculation on load, which is what a file needs
when a formula's precedent changes outside Excel.

Run:  python3 tools/set_reinforce_type.py <file.xlsx> <Type>
      python3 tools/set_reinforce_type.py docs/lem/files/xslope_reinforce.xlsx Geosynthetic
"""

from __future__ import annotations

import os
import sys

import openpyxl

#: The sheet's own preset block — Type, Dir, Appl per row. It sits at AB8:AD11
#: from template v24, at Z8:AB11 before that (Adhesion/Delta pushed the sheet two
#: columns right), so both are tried and whichever holds the block wins.
PRESET_RANGES = (((8, 11), (28, 30)), ((8, 11), (26, 28)))
TYPE_COL, DIR_COL, APPL_COL = 7, 8, 9
LABEL_COL = 2
HEADERS = {TYPE_COL: "Type", DIR_COL: "Dir", APPL_COL: "Appl"}


def presets(ws):
    """{type: (dir, appl)} read from the sheet's lookup block."""
    for rng in PRESET_RANGES:
        out = _presets_at(ws, rng)
        if out:
            return out
    return {}


def _presets_at(ws, rng):
    (r0, r1), (c0, c1) = rng
    out = {}
    for r in range(r0, r1 + 1):
        row = [ws.cell(row=r, column=c).value for c in range(c0, c1 + 1)]
        if row[0]:
            out[str(row[0])] = (row[1], row[2])
    return out


def set_type(path, type_name, sheet="reinforce", header_row=2):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet]

    for col, name in HEADERS.items():
        got = ws.cell(row=header_row, column=col).value
        if got != name:
            raise SystemExit(f"{path}: {sheet} column {col} heads {got!r}, not "
                             f"{name!r} — the sheet layout is not the one this "
                             f"script writes")
    table = presets(ws)
    if type_name not in table:
        raise SystemExit(f"{type_name!r} is not one of the sheet's presets "
                         f"({', '.join(sorted(table))})")

    rows, formulas = [], {}
    r = header_row + 1
    while ws.cell(row=r, column=LABEL_COL).value not in (None, ""):
        for col in (DIR_COL, APPL_COL):
            val = ws.cell(row=r, column=col).value
            if not (isinstance(val, str) and val.startswith("=")):
                raise SystemExit(f"{path}: {sheet}!{ws.cell(row=r, column=col).coordinate} "
                                 f"is not a formula ({val!r}) — this script only "
                                 f"writes files whose Dir/Appl read the preset")
            formulas[(r, col)] = val
        rows.append(r)
        r += 1
    if not rows:
        raise SystemExit(f"{path}: {sheet} has no reinforcement rows to set")

    for r in rows:
        ws.cell(row=r, column=TYPE_COL).value = type_name
    # Writing a formula's precedent outside Excel leaves the cached results
    # stale; a full recalculation on load is what clears them.
    wb.calculation.fullCalcOnLoad = True
    wb.save(path)

    check = openpyxl.load_workbook(path, data_only=False)[sheet]
    for r in rows:
        if check.cell(row=r, column=TYPE_COL).value != type_name:
            raise SystemExit(f"{path}: row {r} did not take the type")
        for col in (DIR_COL, APPL_COL):
            if check.cell(row=r, column=col).value != formulas[(r, col)]:
                raise SystemExit(f"{path}: row {r} column {col} lost its formula")

    from xslope.fileio import load_slope_data
    want = (type_name.lower(),) + tuple(str(v).lower() for v in table[type_name])
    for line in load_slope_data(path).get("reinforcement_lines") or []:
        got = (str(line["type"]), str(line["dir"]), str(line["appl"]))
        if got != want:
            raise SystemExit(f"{path}: a line loads as {got}, not {want}")

    print("%s: %d reinforcement rows set to %s (Dir/Appl resolve to %s/%s)"
          % (os.path.relpath(path), len(rows), type_name, *table[type_name]))
    return rows


def main(argv=None):
    argv = list(sys.argv[1:] if argv is None else argv)
    if len(argv) != 2:
        print(__doc__)
        return 1
    repo = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if repo not in sys.path:
        sys.path.insert(0, repo)
    set_type(argv[0], argv[1])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
