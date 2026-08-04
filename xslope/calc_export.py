# Copyright 2025 Norman L. Jones
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Export LEM Calcs — the factor of safety as a live spreadsheet.

A report states a factor of safety and shows the sums it came from. This writes
the same calculation as a workbook a reader can take apart: every slice on its
own row, every sum a real ``=SUM()`` over those rows, and the factor of safety a
real quotient of two cells. Change a number and the workbook answers. That is
the point — it is a teaching tool, and a reviewer's audit trail, not a dump.

The content is the report's own. :func:`xslope.report.calculation` decides which
per-slice terms the equation carries for this model and this method, and this
module writes them; the curated column set is
:func:`xslope.columns.slice_table`'s, so the spreadsheet's slice table and the
report's are the same table.

Four sheets:

``README``
    What the workbook is, which method, which model, when, which version of
    xslope, a link to the method's published derivation, and how to read each of
    the other sheets. Annotations are cells — this workbook carries no cell
    notes, ever.
``Slice Plot``
    The critical surface with its slices numbered, so the Slice column has a
    visual key.
``Slice Table``
    One row per slice, values as numbers, with a ``=SUM()`` under each of the
    per-slice contribution columns.
``FS Calculation``
    The governing equation in the notation the documentation uses, then the
    formula chain from those sums to F, then — for a method that iterates — the
    closure: the converged unknowns stated as values, the per-slice quantities
    that depend on them recomputed by formula, and a check cell that must read
    zero.

Headless::

    ok, out = export_lem_calcs(slope_data, bundle, "calcs.xlsx")

``bundle`` is Studio's LEM solution — ``{"slice_df", "failure_surface",
"results", "search", "method"}``.
"""

from __future__ import annotations

import os
import re
from datetime import datetime
from math import ceil

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

#: The sheets, in the order they are written.
SHEET_README = "README"
SHEET_PLOT = "Slice Plot"
SHEET_TABLE = "Slice Table"
SHEET_CALC = "FS Calculation"

#: The contribution columns that get a ``=SUM()`` under them. Every one of these
#: is a term of the factor of safety equation, so its total is a number the FS
#: sheet uses. ``y_Q`` is a per-slice elevation and is deliberately not summed —
#: a total under it would be arithmetic that means nothing.
SUMMED_KEYS = ("m_res", "m_drv", "f_res", "f_drv", "q_s")

#: Column widths on the annotated sheets, in characters. Column A carries the
#: labels, so it is wide; the rest carry values and notes.
README_WIDTHS = (30, 14, 14, 14, 14, 14, 14, 14)
CALC_WIDTHS = (46, 16, 60)

#: How wide the embedded figure is drawn, in pixels at the spreadsheet's own 96
#: dpi. About eleven default columns — readable on screen, and it does not bury
#: the sheet below it.
FIGURE_WIDTH_PX = 720

#: The figure is rendered at this size and resolution before it is scaled to
#: :data:`FIGURE_WIDTH_PX`, so the text in it stays sharp when a reader zooms.
FIGURE_SIZE = (10.0, 5.5)
FIGURE_DPI = 150

#: Number formats. A factor of safety is written the way the report writes one;
#: a residual is written in exponent notation, because a residual being small is
#: the whole statement.
FS_FORMAT = "0.000"
SUM_FORMAT = "#,##0.000"
RESIDUAL_FORMAT = "0.000E+00"

_BODY = Font(name="Calibri", size=11)
_HEAD = Font(name="Calibri", size=11, bold=True)
_TITLE = Font(name="Calibri", size=14, bold=True)
_SECTION = Font(name="Calibri", size=12, bold=True)
_NOTE = Font(name="Calibri", size=10, italic=True, color="FF555555")
_MATH = Font(name="Consolas", size=11)
_LINK = Font(name="Calibri", size=11, color="FF0563C1", underline="single")
_SUM_FILL = PatternFill("solid", fgColor="FFEFF3F8")
_HEAD_FILL = PatternFill("solid", fgColor="FFDDE5EF")
_TOP_RULE = Border(top=Side(style="thin", color="FF888888"))
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")


# ---------------------------------------------------------------------------
# Notation
# ---------------------------------------------------------------------------

def _brace(text, i):
    """The braced group starting at ``text[i] == "{"``: ``(inside, next index)``."""
    depth = 0
    for j in range(i, len(text)):
        if text[j] == "{":
            depth += 1
        elif text[j] == "}":
            depth -= 1
            if depth == 0:
                return text[i + 1:j], j + 1
    return text[i + 1:], len(text)


def plain_math(text):
    """The report's math notation as one line of plain text.

    The report writes ``frac{a}{b}`` and ``sum{x}`` so that a Word renderer can
    build native equations from them. A spreadsheet cell has no equation
    renderer, so the same string is written the way it is spoken: a quotient in
    parentheses and a capital sigma.
    """
    out, i, text = "", 0, str(text or "")
    while i < len(text):
        if text.startswith("frac{", i):
            top, j = _brace(text, i + 4)
            bottom, k = _brace(text, j)
            out += f"({plain_math(top)}) / ({plain_math(bottom)})"
            i = k
        elif text.startswith("sum{", i):
            inside, j = _brace(text, i + 3)
            out += f"Σ({plain_math(inside)})"
            i = j
        else:
            out += text[i]
            i += 1
    return out


def _number_format(fmt):
    """``"{:.2f}"`` -> ``"0.00"`` — a column's print format, as Excel writes it."""
    m = re.match(r"\{:\.(\d+)f\}", str(fmt or ""))
    if not m:
        return "General"
    decimals = int(m.group(1))
    return "0" if decimals == 0 else "0." + "0" * decimals


def _float(value):
    """A float, or None for a blank or a NaN — what a numeric cell may hold."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    return None if v != v else v


# ---------------------------------------------------------------------------
# Sheet-writing helpers
# ---------------------------------------------------------------------------

def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _text(ws, row, text, font=None, col=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font or _BODY
    cell.alignment = _TOP
    return row + 1


def _para(ws, row, text, span=8, chars=None):
    """One paragraph, wrapped across ``span`` merged columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _BODY
    cell.alignment = _WRAP
    if chars is None:
        chars = sum(ws.column_dimensions[get_column_letter(i)].width or 8.43
                    for i in range(1, span + 1))
    ws.row_dimensions[row].height = 15 * max(1, ceil(len(text) / max(20, chars - 6)))
    return row + 1


def _row(ws, row, label, value=None, note="", fmt=None, font=None):
    """A labelled row: the label in A, the value or formula in B, a note in C."""
    ws.cell(row=row, column=1, value=label).font = font or _BODY
    ws.cell(row=row, column=1).alignment = _TOP
    if value is not None:
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = font or _BODY
        if fmt:
            cell.number_format = fmt
    if note:
        cell = ws.cell(row=row, column=3, value=note)
        cell.font = _NOTE
        cell.alignment = _WRAP
    return row + 1


def _ref(sheet, column_letter, row):
    """``'Slice Table'!$C$7`` — an absolute reference to another sheet."""
    return f"'{sheet}'!${column_letter}${row}"


#: Where a closure block's per-slice table starts on the FS Calculation sheet.
#: Columns A to C carry that sheet's labels, values and notes and are shaped for
#: sentences; the closure table needs a dozen narrow numeric columns, so it is
#: laid out beside them rather than under them.
CLOSURE_COL0 = 5


def _closure_table(ws, head_row, headers):
    """Write a closure table's two header rows. Returns ``(letters, body_row)``,
    where ``letters`` is one column letter per header, in order."""
    letters = []
    for i, (text, note) in enumerate(headers):
        column = CLOSURE_COL0 + i
        letters.append(get_column_letter(column))
        cell = ws.cell(row=head_row, column=column, value=text)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
        note_cell = ws.cell(row=head_row + 1, column=column, value=note)
        note_cell.font = _NOTE
        ws.column_dimensions[letters[-1]].width = 20
    ws.row_dimensions[head_row].height = 46
    return letters, head_row + 2


def _closure_total(ws, row, letters, entries):
    """The ruled total row under a closure table. ``entries`` is
    ``{letter: (formula, number format)}``."""
    for letter in letters:
        cell = ws[f"{letter}{row}"]
        cell.fill = _SUM_FILL
        cell.border = _TOP_RULE
    ws[f"{letters[0]}{row}"] = "Σ"
    ws[f"{letters[0]}{row}"].font = _HEAD
    for letter, (formula, fmt) in entries.items():
        cell = ws[f"{letter}{row}"]
        cell.value = formula
        cell.font = _HEAD
        cell.number_format = fmt


# ---------------------------------------------------------------------------
# The slice-key figure
# ---------------------------------------------------------------------------

#: A dedicated slice-key renderer, if :mod:`xslope.plot` grows one. It is called
#: with the same four positional arguments :func:`~xslope.plot.plot_solution`
#: takes, and anything it raises sends the seam back to the general one.
SLICE_KEY_FUNCS = ("plot_slice_key",)

#: The keyword that frames the solution plot tightly around the sliced mass,
#: under whichever name it arrives; the first one the signature has is used.
TIGHT_FRAME_KWARGS = ("tight_frame", "frame_slices", "tight", "zoom_to_slices")


def slice_plot_renderer():
    """The best slice-key figure this build of :mod:`xslope.plot` can draw.

    Returns ``(callable, kwargs, name)``. The figure wanted here is the solution
    plot framed tightly around the sliced mass with the slice numbers on, so the
    workbook's Slice column has a visual key. Where the plotting module offers
    that as its own function it is used; otherwise the solution plot is called
    with whichever of those two options it accepts. A build that has neither
    still gets a figure — the plain solution plot — and the caller is told which
    it got, so nothing fails silently.
    """
    import inspect

    from . import plot as _plot

    for name in SLICE_KEY_FUNCS:
        fn = getattr(_plot, name, None)
        if callable(fn):
            return fn, {}, name

    fn = _plot.plot_solution
    params = inspect.signature(fn).parameters
    kwargs = {}
    if "slice_numbers" in params:
        kwargs["slice_numbers"] = True
    for key in TIGHT_FRAME_KWARGS:
        if key in params:
            kwargs[key] = True
            break
    return fn, kwargs, "plot_solution"


def render_slice_plot(slope_data, bundle, path, style=None):
    """Draw the slice-key figure to ``path``. Returns the renderer's name, or
    ``""`` when no figure could be drawn."""
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.figure import Figure

    slice_df = bundle.get("slice_df")
    if slice_df is None:
        return ""
    args = (slope_data, slice_df, bundle.get("failure_surface"),
            bundle.get("results") or {})
    fn, kwargs, name = slice_plot_renderer()
    for attempt in ((fn, kwargs, name), (None, None, None)):
        call, extra, label = attempt
        if call is None:
            if name == "plot_solution":
                return ""
            from .plot import plot_solution as call
            extra, label = {"slice_numbers": True}, "plot_solution"
        fig = Figure(figsize=FIGURE_SIZE, dpi=FIGURE_DPI)
        FigureCanvasAgg(fig)
        try:
            call(*args, fig=fig, show_title=False, style=style, **extra)
            fig.savefig(path, dpi=FIGURE_DPI, bbox_inches="tight")
            return label
        except Exception:
            import traceback
            traceback.print_exc()
    return ""


# ---------------------------------------------------------------------------
# Sheet 1 — README
# ---------------------------------------------------------------------------

def _readme_sheet(ws, calc, slope_data, meta, layout):
    from .report import method_doc_url, method_label

    method = calc["method"]
    label = method_label(method)
    _widths(ws, README_WIDTHS)
    row = _text(ws, 1, "XSLOPE — Limit Equilibrium Calculation Export", _TITLE)
    row += 1

    row = _para(ws, row,
                "This workbook is the factor of safety on the critical surface, "
                "written out so it can be checked cell by cell. Every sum is a "
                "live =SUM() over the slice rows and the factor of safety is a "
                "live quotient of two cells, so a number that is changed is a "
                "number the workbook answers for. It is generated from the same "
                "solution the analysis report documents.")
    row += 1

    row = _text(ws, row, "This export", _SECTION)
    url = method_doc_url(method)
    items = [
        ("Method", label),
        ("Factor of safety", _float(calc["FS"])),
        ("Slices", len(calc["slice_df"])),
        ("Surface", "circular" if slope_data.get("circular") else "non-circular"),
        ("Model file", meta.get("model") or "(not saved to a file)"),
        ("Exported", meta.get("date")),
        ("xslope version", meta.get("version")),
    ]
    if calc.get("stage"):
        items.append(("Governing stage", calc["stage"]))
    for name, value in items:
        fmt = FS_FORMAT if name == "Factor of safety" else None
        ws.cell(row=row, column=1, value=name).font = _HEAD
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _BODY
        if fmt:
            cell.number_format = fmt
        row += 1
    if url:
        ws.cell(row=row, column=1, value="Method derivation").font = _HEAD
        cell = ws.cell(row=row, column=2, value=url)
        cell.hyperlink = url
        cell.font = _LINK
        row += 1
    row += 1

    row = _text(ws, row, "How to read the sheets", _SECTION)
    notes = [
        (SHEET_PLOT,
         "The critical surface and its slices. The numbers on the slices are the "
         "Slice column of the Slice Table sheet, so a row in the table can be "
         "found on the section."),
        (SHEET_TABLE,
         "One row per slice, left to right along the failure surface. The "
         "geometry, force and strength columns are the ones the analysis report "
         "prints, in the same order and under the same symbols; each header "
         "carries the model's own units. The last columns are this method's "
         "per-slice contributions to the two sides of the factor of safety "
         "equation, and under each of those is a =SUM() over the slice rows. "
         "Those sums are what the FS Calculation sheet divides."),
        (SHEET_CALC,
         "The equation the solver evaluates, written in the notation the "
         "published derivation uses, then the arithmetic: the two sums pulled "
         "from the Slice Table, their quotient, and the factor of safety. Every "
         "value in the chain is a formula except the ones marked as stated, "
         "which are the converged values the solver reported."),
    ]
    for name, note in notes:
        ws.cell(row=row, column=1, value=name).font = _HEAD
        row += 1
        row = _para(ws, row, note)
    row += 1

    row = _text(ws, row, "What is stated and what is computed", _SECTION)
    row = _para(ws, row,
                "A method that iterates cannot be written as one pass of "
                "arithmetic: the base normal force on each slice depends on the "
                "factor of safety, which depends on the base normal forces. The "
                "converged values are therefore stated — they are the solver's "
                "answer — and the sheet then demonstrates the closure: it "
                "recomputes, by formula, the per-slice quantities that depend on "
                "those values, and shows that they return what was stated. Each "
                "check cell is labelled in words and reads zero to within the "
                "solver's own convergence tolerance.")
    row = _para(ws, row,
                "Nothing in this workbook is written as a cell comment. Every "
                "explanation is a cell you can read, print and search.")
    row += 1

    row = _text(ws, row, "Named ranges", _SECTION)
    row = _para(ws, row,
                "The sums and the factor of safety carry workbook names, so a "
                "formula of your own can refer to them by name instead of by "
                "address: " + ", ".join(layout["names"]) + ".")
    return row


# ---------------------------------------------------------------------------
# Sheet 2 — the slice-key figure
# ---------------------------------------------------------------------------

def _plot_sheet(ws, calc, png, renderer):
    from openpyxl.drawing.image import Image

    from .report import method_label

    _text(ws, 1, f"Critical surface and slices — {method_label(calc['method'])}",
          _TITLE)
    _text(ws, 2,
          "The slices are numbered as the Slice Table numbers them, left to "
          f"right along the failure surface.  (drawn by {renderer})", _NOTE)
    img = Image(png)
    scale = FIGURE_WIDTH_PX / float(img.width)
    img.width = FIGURE_WIDTH_PX
    img.height = int(round(img.height * scale))
    ws.add_image(img, "A4")


# ---------------------------------------------------------------------------
# Sheet 3 — the slice table
# ---------------------------------------------------------------------------

def _slice_sheet(ws, calc, unit_labels):
    """The curated slice table as numbers, with a ``=SUM()`` under each
    contribution column. Returns the layout the FS sheet needs."""
    from .columns import header, report_columns, slice_table

    df = calc["slice_df"]
    headers, _rows, _legend = slice_table(df, unit_labels)
    by_header = {header(c, unit_labels): c for c in report_columns()}
    cols = [by_header[h] for h in headers]

    for i, (text, column) in enumerate(zip(headers, cols), start=1):
        cell = ws.cell(row=1, column=i, value=text)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="bottom",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = max(9, len(text) + 3)

    first = 2
    for r, (_i, record) in enumerate(df.iterrows(), start=first):
        for i, column in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=i, value=_float(record[column.key]))
            cell.font = _BODY
            cell.number_format = _number_format(column.fmt)
    last = first + len(df) - 1

    letters = {c.key: get_column_letter(i) for i, c in enumerate(cols, start=1)}
    total_row = last + 1
    label = ws.cell(row=total_row, column=1, value=f"Σ  ({len(df)} slices)")
    label.font = _HEAD
    sums = {}
    for i, column in enumerate(cols, start=1):
        cell = ws.cell(row=total_row, column=i)
        cell.fill = _SUM_FILL
        cell.border = _TOP_RULE
        if column.key in SUMMED_KEYS:
            letter = get_column_letter(i)
            cell.value = f"=SUM({letter}{first}:{letter}{last})"
            cell.font = _HEAD
            # Spencer's Q sums to zero by construction — that is the force
            # equilibrium of the whole mass — so its total is written the way a
            # residual is written, not the way a force is.
            cell.number_format = (RESIDUAL_FORMAT if column.key == "q_s"
                                  else SUM_FORMAT)
            sums[column.key] = _ref(SHEET_TABLE, letter, total_row)

    ws.freeze_panes = "B2"

    row = total_row + 2
    row = _text(ws, row,
                "The Σ row is a live =SUM() over the slice rows above it, for "
                "each per-slice contribution column. The FS Calculation sheet "
                "divides those totals.", _NOTE)
    _text(ws, row,
          "There is no total under y_Q: it is an elevation, and a sum of "
          "elevations is not a quantity.", _NOTE)

    return {"first": first, "last": last, "total_row": total_row,
            "letters": letters, "sums": sums, "columns": cols}


# ---------------------------------------------------------------------------
# Sheet 4 — the factor of safety
# ---------------------------------------------------------------------------

def _preamble_blocks(calc):
    """The report's own preamble for this method, as ``(kind, text)`` pairs.

    Read from :mod:`xslope.report` so the two documents cannot drift apart. A
    build whose report does not offer it simply gets no preamble.
    """
    try:
        from .report import _method_preamble
        blocks = _method_preamble(calc, calc["method"])
    except Exception:
        return []
    out = []
    for block in blocks:
        kind = getattr(block, "kind", "")
        text = getattr(block, "notation", "") if kind == "math" \
            else getattr(block, "text", "")
        if kind in ("prose", "math") and text:
            out.append((kind, text))
    return out


def _calc_sheet(ws, calc, layout, closure):
    """The equation, the sums, the quotient, and the closure."""
    from .report import method_label

    method = calc["method"]
    _widths(ws, CALC_WIDTHS)
    row = _text(ws, 1, f"Factor of Safety — {method_label(method)}", _TITLE)
    row += 1

    row = _text(ws, row, "The equation the solver evaluates", _SECTION)
    for kind, text in _preamble_blocks(calc):
        if kind == "math":
            row = _text(ws, row, plain_math(text), _MATH)
        else:
            row = _para(ws, row, text, span=3, chars=sum(CALC_WIDTHS))
    row = _text(ws, row, plain_math(calc["equation"]), _MATH)
    row = _para(ws, row,
                "The equation carries only the terms this model exercises. A "
                "force the model does not have is dropped rather than printed as "
                "a column of zeros; the per-slice terms that remain are the "
                "contribution columns of the Slice Table.",
                span=3, chars=sum(CALC_WIDTHS))
    row += 1

    row = _text(ws, row, "The arithmetic", _SECTION)
    for i, text in enumerate(("Quantity", "Value", "How it is computed")):
        cell = ws.cell(row=row, column=i + 1, value=text)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL
    row += 1

    res_key, drv_key = calc["res_key"], calc["drv_key"]
    from .columns import BY_KEY
    res_label = BY_KEY[res_key].label
    drv_label = BY_KEY[drv_key].label
    moment = method in ("oms", "bishop")
    what = "moment" if moment else "force"

    res_row = row
    row = _row(ws, row, f"Sum of the resisting {what}s,  Σ {res_label}",
               f"={layout['sums'][res_key]}",
               f"the Σ cell under column {res_label} of the Slice Table",
               fmt=SUM_FORMAT)
    drv_row = row
    row = _row(ws, row, f"Sum of the net driving {what}s,  Σ {drv_label}",
               f"={layout['sums'][drv_key]}",
               f"the Σ cell under column {drv_label} of the Slice Table",
               fmt=SUM_FORMAT)
    quotient_row = row
    row = _row(ws, row, "F  =  resisting / driving",
               f"=B{res_row}/B{drv_row}",
               f"the quotient of the two cells above:  B{res_row} / B{drv_row}",
               fmt=FS_FORMAT, font=_HEAD)

    final_row = quotient_row
    if calc.get("fo"):
        fo_row = row
        row = _row(ws, row, "Janbu correction factor  f_o", _float(calc["fo"]),
                   "stated — read from the method's chart fit for this "
                   "surface's depth-to-length ratio and soil type",
                   fmt="0.0000")
        final_row = row
        row = _row(ws, row, "F_corr  =  f_o · F",
                   f"=B{fo_row}*B{quotient_row}",
                   f"the corrected factor of safety:  B{fo_row} × B{quotient_row}",
                   fmt=FS_FORMAT, font=_HEAD)

    solver_row = row
    row = _row(ws, row, "Factor of safety reported by the solver",
               _float(calc["FS"]),
               "stated — the converged value this workbook was written from",
               fmt=FS_FORMAT)
    check_row = row
    row = _row(ws, row,
               "Check: this sheet minus the solver (should be zero)",
               f"=B{final_row}-B{solver_row}",
               "the workbook's own arithmetic against the solver's answer",
               fmt=RESIDUAL_FORMAT)
    row += 1

    cells = {"resisting": f"$B${res_row}", "driving": f"$B${drv_row}",
             "quotient": f"$B${quotient_row}", "fs": f"$B${final_row}",
             "solver": f"$B${solver_row}", "check": f"$B${check_row}"}

    if closure is not None:
        row = closure["write"](ws, row, calc, layout, cells)
    elif method in ("corps", "lowe", "mprice", "janbu"):
        row = _closure_stated(ws, row, calc)
    else:
        row = _para(ws, row,
                    "This method is explicit: nothing above depends on the "
                    "factor of safety, so the two sums are formed once and "
                    "divided once. There is no iteration to close.",
                    span=3, chars=sum(CALC_WIDTHS))
    return cells


def _closure_stated(ws, row, calc):
    """For a method whose converged unknowns are stated but whose per-slice
    closure is not rebuilt here: state them, and say what they are."""
    row = _text(ws, row, "The converged unknowns", _SECTION)
    stated = []
    if calc.get("theta") is not None:
        stated.append(("Interslice force inclination  θ (degrees)",
                       _float(calc["theta"]),
                       "stated — the method's own convention fixes it from the "
                       "geometry"))
    if calc.get("lambda") is not None:
        stated.append(("Interslice force scale  λ", _float(calc["lambda"]),
                       "stated — solved jointly with F"))
    residuals = calc.get("residuals")
    if residuals is not None:
        stated.append(("Force closure at the far end of the march  Z_n",
                       _float(residuals[0]),
                       "stated — what the converged iteration left behind"))
        stated.append(("Moment of the sliding mass about the origin  Σ M_o",
                       _float(residuals[1]),
                       "stated — what the converged iteration left behind"))
    if not stated:
        stated.append(("(no additional unknowns)", None, ""))
    for label, value, note in stated:
        fmt = RESIDUAL_FORMAT if "closure" in label or "Moment" in label else "0.0000"
        row = _row(ws, row, label, value, note, fmt=fmt if value is not None else None)
    row += 1
    row = _para(ws, row,
                "The base normal force on each slice was evaluated at the "
                "converged factor of safety, which is what makes the quotient "
                "above close on itself: the sums formed from those normals "
                "return the F they were evaluated at.",
                span=3, chars=sum(CALC_WIDTHS))
    return row


# --- Bishop -----------------------------------------------------------------

def bishop_closure(calc):
    """The per-slice values Bishop's closure block needs, or None.

    Bishop solves the base normal force and the quotient together: N' depends on
    F through m_α and through the cohesion term, and the moments are formed from
    N'. The block recomputes N' from the STATED F by formula, rebuilds both
    moments from it, and divides — so the check cell at the bottom is the
    iteration closing on itself, done in the spreadsheet.

    ``V`` is the part of vertical equilibrium that does not depend on F: the
    slice weight and the surface loads. The pore-pressure and cohesion terms are
    left to the formula, which takes them from the Slice Table.

    Returns None when the model carries passive support — whose vertical
    components mobilise with the soil and so divide by F as well — or apparent
    cohesion from matric suction, which the solver adds to c and the Slice
    Table's c column does not carry. Neither closure is written here.
    """
    import numpy as np

    from .report import PASSIVE_COLUMNS
    from .solve import _moment_arms

    df = calc["slice_df"]
    if calc["method"] != "bishop" or "xo" not in df.columns:
        return None
    for name in PASSIVE_COLUMNS + ("c_suction",):
        if name in df.columns and np.any(np.abs(df[name].values.astype(float)) > 1e-12):
            return None

    def col(name):
        return (df[name].values.astype(float) if name in df.columns
                else np.zeros(len(df)))

    alpha = np.radians(df["alpha"].values.astype(float))
    # The solver reads the facing two ways: an explicit override for the moment
    # arms, and the raw geometric test for the axial reinforcement's vertical
    # component. Both are followed here, so V is the solver's own.
    right_facing = bool(df.attrs.get(
        "right_facing", df["y_lb"].iat[0] > df["y_rb"].iat[-1]))
    geometric = bool(df["y_lb"].iat[0] > df["y_rb"].iat[-1])
    Xo, Yo = float(df["xo"].iloc[0]), float(df["yo"].iloc[0])
    _xr, a_S, a_N = _moment_arms(df, Xo, Yo, alpha, right_facing)

    beta = np.radians(col("beta"))
    ll_b = np.radians(col("ll_beta"))
    pa_cy = col("pa_cy")
    pa_cy_v = -pa_cy if geometric else pa_cy
    V = (col("w") + col("dload") * np.cos(beta) + col("lload") * np.cos(ll_b)
         - col("p") * np.sin(alpha) - pa_cy_v
         - col("h_pile") * np.sin(col("theta_p")))

    N = col("n_eff")
    driving_fixed = (calc["slice_df"][calc["drv_key"]].values.astype(float)
                     + (N + col("u") * col("dl")) * a_N)

    return {"V": V, "a_S": a_S, "a_N": a_N, "driving_fixed": driving_fixed,
            "write": _write_bishop_closure}


def _write_bishop_closure(ws, row, calc, layout, cells):
    state = calc["_closure"]
    n = len(calc["slice_df"])
    first, letters = layout["first"], layout["letters"]

    def st(key, r):
        return _ref(SHEET_TABLE, letters[key], r)

    row = _text(ws, row, "Closure — the iteration, done in the spreadsheet",
                _SECTION)
    row = _para(ws, row,
                "The base normal force N' comes from vertical equilibrium of "
                "the slice and depends on the factor of safety itself, so the "
                "two are solved together. Below, N' is recomputed by formula "
                "from the STATED factor of safety, the two moments are rebuilt "
                "from that N', and their quotient is compared with the stated "
                "value. The three check cells at the bottom are the closure.",
                span=3, chars=sum(CALC_WIDTHS))
    row += 1

    fs_cell = cells["fs"]
    head_row = row
    (SL, V, AS, AN, DF, MA, NP, NS, DN, MR, MD), body = _closure_table(
        ws, head_row, [
            ("Slice", "from the Slice Table"),
            ("V  (the F-independent vertical forces)", "stated"),
            ("a_S  (moment arm of the base shear)", "stated"),
            ("a_N  (moment arm of the base normal)", "stated"),
            ("M_D, the part that does not depend on N'", "stated"),
            ("m_α  = cos α + sin α·tan φ / F", "formula"),
            ("N'  = (V − u·Δl·cos α − c·Δl·sin α/F) / m_α", "formula"),
            ("N' from the solver", "Slice Table"),
            ("|ΔN'|", "formula"),
            ("M_R  = (c·Δl + N'·tan φ)·a_S", "formula"),
            ("M_D  = fixed part − (N' + u·Δl)·a_N", "formula"),
        ])

    for i in range(n):
        r = body + i
        sr = first + i
        a = st("alpha", sr)
        u_dl = f"{st('u', sr)}*{st('dl', sr)}"
        c_dl = f"{st('c', sr)}*{st('dl', sr)}"
        tan_phi = f"TAN(RADIANS({st('phi', sr)}))"
        ws[f"{SL}{r}"] = i + 1
        ws[f"{SL}{r}"].number_format = "0"
        for letter, value, fmt in ((V, state["V"][i], "0.0"),
                                   (AS, state["a_S"][i], "0.000"),
                                   (AN, state["a_N"][i], "0.000"),
                                   (DF, state["driving_fixed"][i], "0.0")):
            ws[f"{letter}{r}"] = _float(value)
            ws[f"{letter}{r}"].number_format = fmt
        for letter, formula, fmt in (
                (MA, f"=COS(RADIANS({a}))+SIN(RADIANS({a}))*{tan_phi}/{fs_cell}",
                 "0.00000"),
                (NP, f"=({V}{r}-{u_dl}*COS(RADIANS({a}))"
                     f"-{c_dl}*SIN(RADIANS({a}))/{fs_cell})/{MA}{r}", "0.0"),
                (NS, f"={st('n_eff', sr)}", "0.0"),
                (DN, f"=ABS({NP}{r}-{NS}{r})", RESIDUAL_FORMAT),
                (MR, f"=({c_dl}+{NP}{r}*{tan_phi})*{AS}{r}", "0.0"),
                (MD, f"={DF}{r}-({NP}{r}+{u_dl})*{AN}{r}", "0.0")):
            ws[f"{letter}{r}"] = formula
            ws[f"{letter}{r}"].number_format = fmt
    end = body + n - 1

    total = end + 1
    _closure_total(ws, total, [SL, V, AS, AN, DF, MA, NP, NS, DN, MR, MD], {
        DN: (f"=MAX({DN}{body}:{DN}{end})", RESIDUAL_FORMAT),
        MR: (f"=SUM({MR}{body}:{MR}{end})", SUM_FORMAT),
        MD: (f"=SUM({MD}{body}:{MD}{end})", SUM_FORMAT)})
    row = total + 2

    row = _row(ws, row,
               "Largest difference between the recomputed N' and the solver's",
               f"={DN}{total}",
               "should be zero — the base normals really were evaluated at the "
               "stated factor of safety", fmt=RESIDUAL_FORMAT)
    q_row = row
    row = _row(ws, row, "F recomputed from the rebuilt moments",
               f"={MR}{total}/{MD}{total}",
               f"the rebuilt sums divided:  {MR}{total} / {MD}{total}",
               fmt=FS_FORMAT, font=_HEAD)
    row = _row(ws, row,
               "Check: recomputed F minus the stated F (should be zero)",
               f"=B{q_row}-{cells['solver']}",
               "the iteration closing on itself", fmt=RESIDUAL_FORMAT)
    return row + 1


# --- Spencer ----------------------------------------------------------------

def spencer_closure(calc):
    """The per-slice values Spencer's closure block needs, or None.

    Spencer's two equilibrium equations are sums of the per-slice resultant Q,
    which is a function of the converged pair (F, θ). The block recomputes Q from
    the stated pair by formula and sums it twice — once as a force, once as a
    moment about the origin — so both closures are arithmetic a reader can
    follow.

    ``F_h`` and ``F_v`` are the sums of the forces on the slice other than the
    base normal, the base shear and the interslice forces; they come from the
    solver, whose debug level writes them into the table it is given. On a
    right-facing slope the solver mirrors α, c and tan φ, so those are carried
    here with the sign the solver gives them rather than the table's.
    """
    import contextlib
    import io

    import numpy as np

    from .solve import spencer

    df = calc["slice_df"]
    if calc["method"] != "spencer":
        return None
    work = df.copy()
    try:
        with contextlib.redirect_stdout(io.StringIO()):
            ok, res = spencer(work, debug_level=2)
    except Exception:
        return None
    if not ok or "Fh" not in work.columns:
        return None

    right_facing = bool(df.attrs.get(
        "right_facing", df["y_lb"].iat[0] > df["y_rb"].iat[-1]))
    sign = -1.0 if right_facing else 1.0
    c = df["c"].values.astype(float)
    if "c_suction" in df.columns:
        c = c + df["c_suction"].values.astype(float)
    return {
        "Fh": work["Fh"].values.astype(float),
        "Fv": work["Fv"].values.astype(float),
        "alpha": sign * df["alpha"].values.astype(float),
        "c": sign * c,
        "tan_phi": sign * np.tan(np.radians(df["phi"].values.astype(float))),
        "theta": float(res["theta"]), "FS": float(res["FS"]),
        "right_facing": right_facing,
        "write": _write_spencer_closure,
    }


def _write_spencer_closure(ws, row, calc, layout, cells):
    state = calc["_closure"]
    n = len(calc["slice_df"])
    first, letters = layout["first"], layout["letters"]

    def st(key, r):
        return _ref(SHEET_TABLE, letters[key], r)

    row = _text(ws, row, "Closure — both equilibrium sums, from the stated "
                         "(F, θ)", _SECTION)
    row = _para(ws, row,
                "Spencer's method lumps the interslice forces on each slice "
                "into a single resultant Q at the constant inclination θ, so "
                "that force and moment equilibrium of the whole sliding mass "
                "are two equations in the two unknowns F and θ. Below, Q is "
                "recomputed by formula from the stated pair and summed twice. "
                "Both sums must read zero; neither is exactly zero, because "
                "that is what a converged iteration leaves behind.",
                span=3, chars=sum(CALC_WIDTHS))
    row += 1

    fs_row, theta_row = row, row + 1
    row = _row(ws, row, "Factor of safety at the solution  F", _float(state["FS"]),
               "stated", fmt=FS_FORMAT)
    row = _row(ws, row, "Interslice inclination at the solution  θ (degrees)",
               _float(state["theta"]), "stated", fmt="0.0000")
    if state["right_facing"]:
        row = _row(ws, row, "Slope faces", "right",
                   "α, c and tan φ below carry the sign the solver gives them "
                   "for this facing, which is the mirror of the Slice Table's")
    row += 1

    fs_cell, th_cell = f"$B${fs_row}", f"$B${theta_row}"
    head_row = row
    (SL, AL, CC, TP, FH, FV, MA, QQ, QS, DQ, MO, AQ), body = _closure_table(
        ws, head_row, [
            ("Slice", "from the Slice Table"),
            ("α  (as the solver signs it, degrees)", "stated"),
            ("c  (as the solver signs it)", "stated"),
            ("tan φ  (as the solver signs it)", "stated"),
            ("F_h", "stated"),
            ("F_v", "stated"),
            ("m_α  = 1 / (cos(α−θ) + sin(α−θ)·tan φ / F)", "formula"),
            ("Q  (Spencer's equation 23)", "formula"),
            ("Q from the solver", "Slice Table"),
            ("|ΔQ|", "formula"),
            ("Q·(x_b·sin θ − y_Q·cos θ)", "formula"),
            ("|Q|", "formula"),
        ])

    for i in range(n):
        r = body + i
        sr = first + i
        sin_a, cos_a = f"SIN(RADIANS({AL}{r}))", f"COS(RADIANS({AL}{r}))"
        u_dl = f"{st('u', sr)}*{st('dl', sr)}"
        ws[f"{SL}{r}"] = i + 1
        ws[f"{SL}{r}"].number_format = "0"
        for letter, value, fmt in ((AL, state["alpha"][i], "0.0000"),
                                   (CC, state["c"][i], "0.000"),
                                   (TP, state["tan_phi"][i], "0.00000"),
                                   (FH, state["Fh"][i], "0.0"),
                                   (FV, state["Fv"][i], "0.0")):
            ws[f"{letter}{r}"] = _float(value)
            ws[f"{letter}{r}"].number_format = fmt
        for letter, formula, fmt in (
                (MA, f"=1/(COS(RADIANS({AL}{r}-{th_cell}))"
                     f"+SIN(RADIANS({AL}{r}-{th_cell}))*{TP}{r}/{fs_cell})",
                 "0.00000"),
                (QQ, f"=(-{FV}{r}*{sin_a}-{FH}{r}*{cos_a}"
                     f"-{CC}{r}/{fs_cell}*{st('dl', sr)}"
                     f"+({FV}{r}*{cos_a}-{FH}{r}*{sin_a}+{u_dl})"
                     f"*{TP}{r}/{fs_cell})*{MA}{r}", "0.0"),
                (QS, f"={st('q_s', sr)}", "0.0"),
                (DQ, f"=ABS({QQ}{r}-{QS}{r})", RESIDUAL_FORMAT),
                (MO, f"={QQ}{r}*({st('x_c', sr)}*SIN(RADIANS({th_cell}))"
                     f"-{st('y_q', sr)}*COS(RADIANS({th_cell})))", "0.0"),
                (AQ, f"=ABS({QQ}{r})", "0.0")):
            ws[f"{letter}{r}"] = formula
            ws[f"{letter}{r}"].number_format = fmt
    end = body + n - 1

    total = end + 1
    _closure_total(ws, total,
                   [SL, AL, CC, TP, FH, FV, MA, QQ, QS, DQ, MO, AQ], {
                       QQ: (f"=SUM({QQ}{body}:{QQ}{end})", RESIDUAL_FORMAT),
                       DQ: (f"=MAX({DQ}{body}:{DQ}{end})", RESIDUAL_FORMAT),
                       MO: (f"=SUM({MO}{body}:{MO}{end})", RESIDUAL_FORMAT),
                       AQ: (f"=SUM({AQ}{body}:{AQ}{end})", SUM_FORMAT)})
    row = total + 2

    row = _row(ws, row, "Largest difference between the recomputed Q and the "
                        "solver's", f"={DQ}{total}",
               "should be zero — Q really is a function of the stated (F, θ)",
               fmt=RESIDUAL_FORMAT)
    row = _row(ws, row, "Force equilibrium of the sliding mass,  Σ Q",
               f"={QQ}{total}", "should be zero", fmt=RESIDUAL_FORMAT)
    row = _row(ws, row,
               "Moment equilibrium of the sliding mass,  "
               "Σ Q·(x_b·sin θ − y_Q·cos θ)", f"={MO}{total}",
               "should be zero", fmt=RESIDUAL_FORMAT)
    scale_row = row
    row = _row(ws, row, "Σ |Q|, the size the two residuals are judged against",
               f"={AQ}{total}",
               "a residual is small when it is small compared with this",
               fmt=SUM_FORMAT)
    row = _row(ws, row, "Σ Q as a fraction of Σ |Q|",
               f"={QQ}{total}/B{scale_row}", "should be zero",
               fmt=RESIDUAL_FORMAT)
    return row + 1


#: How a method's closure block is built. A method with no entry gets the stated
#: converged unknowns and no per-slice rebuild.
CLOSURES = {"bishop": bishop_closure, "spencer": spencer_closure}


# ---------------------------------------------------------------------------
# The entry point
# ---------------------------------------------------------------------------

def default_filename(model_path, method):
    """``north_levee_calcs_spencer.xlsx`` — what the export is called."""
    stem = (os.path.splitext(os.path.basename(model_path))[0]
            if model_path else "model")
    return f"{stem}_calcs_{str(method or 'lem').lower()}.xlsx"


def export_lem_calcs(slope_data, bundle, path, method=None, model_path=None,
                     style=None):
    """Write the selected method's factor of safety to ``path`` as a workbook.

    Parameters
    ----------
    slope_data : dict
        The model, as :func:`xslope.fileio.load_slope_data` returns it.
    bundle : dict
        A solved LEM bundle — ``{"slice_df", "failure_surface", "results",
        "search", "method"}``, which is what Studio's runner emits.
    path : str
        The ``.xlsx`` to write.
    method : str, optional
        Which method to document. Defaults to the bundle's own.
    model_path : str, optional
        The model file the solution came from, named on the README sheet.
    style : dict, optional
        Studio's plotting style, so the embedded figure matches the screen.

    Returns
    -------
    (bool, dict or str)
        ``(True, {"path", "method", "sheets", "figure"})``, or ``(False,
        message)`` — the package's error convention.
    """
    import tempfile

    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    from .plot import declared_unit_labels
    from .report import bundle_method, calculation

    if not path:
        return False, "No output path was given for the calculation export."
    if not bundle:
        return False, "There is no limit equilibrium solution to export."
    method = str(method or bundle_method(bundle) or "").lower()
    if not method:
        return False, "The solution does not say which method produced it."

    try:
        calc = calculation(slope_data, bundle, method)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return False, f"The calculation could not be worked out: {exc}"
    if calc is None:
        from .report import method_label
        return False, (
            f"{method_label(method)} cannot be written out as a cell-by-cell "
            f"calculation for this model. Either the method's factor of safety "
            f"is not a quotient of two sums over the slices, or the model "
            f"carries passive support, whose capacity divides by the factor of "
            f"safety on the driving side.")

    builder = CLOSURES.get(method)
    closure = None
    if builder is not None:
        try:
            closure = builder(calc)
        except Exception:
            import traceback
            traceback.print_exc()
    calc["_closure"] = closure

    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet(SHEET_README)
    plot_ws = wb.create_sheet(SHEET_PLOT)
    table_ws = wb.create_sheet(SHEET_TABLE)
    calc_ws = wb.create_sheet(SHEET_CALC)

    layout = _slice_sheet(table_ws, calc, declared_unit_labels(slope_data))
    cells = _calc_sheet(calc_ws, calc, layout, closure)

    names = {
        "Sum_resisting": _ref(SHEET_CALC, "B", int(cells["resisting"][3:])),
        "Sum_driving": _ref(SHEET_CALC, "B", int(cells["driving"][3:])),
        "Factor_of_safety": _ref(SHEET_CALC, "B", int(cells["fs"][3:])),
    }
    for name, target in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=target))

    meta = {
        "model": os.path.basename(model_path) if model_path else "",
        "date": datetime.now().strftime("%B %d, %Y"),
        "version": _version(),
    }
    _readme_sheet(readme, calc, slope_data, meta, {"names": sorted(names)})

    figure = ""
    tmp = tempfile.mkdtemp(prefix="xslope_calcs_")
    try:
        png = os.path.join(tmp, "slice_plot.png")
        figure = render_slice_plot(slope_data, bundle, png, style=style)
        if figure:
            _plot_sheet(plot_ws, calc, png, figure)
        else:
            _text(plot_ws, 1, "The slice figure could not be drawn for this "
                              "model.", _NOTE)

        # Every formula in the workbook is written as a string; nothing here
        # caches a result, so a consumer must be told to compute them on open.
        wb.calculation.fullCalcOnLoad = True
        try:
            wb.save(path)
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return False, f"The workbook could not be written: {exc}"
    finally:
        import shutil
        shutil.rmtree(tmp, ignore_errors=True)

    return True, {"path": path, "method": method, "figure": figure,
                  "sheets": [SHEET_README, SHEET_PLOT, SHEET_TABLE, SHEET_CALC]}


def _version():
    try:
        from . import __version__
        return str(__version__)
    except Exception:
        return ""
