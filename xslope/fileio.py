# Copyright 2025 Norman L. Jones
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import io
import math
import os
import pickle
import re
import shutil
import tempfile
import warnings
import zipfile
from xml.sax.saxutils import escape as xml_escape

import numpy as np
import pandas as pd
from lxml import etree
from shapely.geometry import LineString, Point, Polygon
from shapely.ops import unary_union

from .mesh import import_mesh_from_json, build_polygons
from .package import MESH_SIDECAR, SEEP_SIDECARS, is_package, unpack
from .units import (GAMMA_W, infer_system_from_gamma_water, normalize_unit_system,
                    units_check)
# v22 main!D23 vocabulary. It lives in xslope.water, which is where the mode is
# acted on, and is imported rather than restated so the sheet and the engine can
# never offer different words.
from .water import WATER_LOAD_MODES as WATER_LOAD_OPTIONS

def build_ground_surface(profile_lines):
    """
    Constructs the topmost ground surface LineString from a set of profile lines.

    The function finds the highest elevation at each x-coordinate across all profile lines,
    which represents the true ground surface.

    Parameters:
        profile_lines (list of dict): A list of profile lines, each represented
            as a dict with 'coords' key containing a list of (x, y) coordinate tuples.

    Returns:
        shapely.geometry.LineString: A LineString of the top surface, or an empty LineString
        if fewer than two valid points are found.
    """
    
    if not profile_lines:
        return LineString([])
    
    # Extract coordinate lists from profile line dicts
    coord_lists = [line['coords'] for line in profile_lines]
    
    # Step 1: Gather all points from all profile lines
    all_points = []
    for line in coord_lists:
        all_points.extend(line)
    
    # Step 2: Group points by x-coordinate and find the highest y for each x
    x_groups = {}
    for x, y in all_points:
        if x not in x_groups:
            x_groups[x] = y
        else:
            x_groups[x] = max(x_groups[x], y)
    
    # Step 3: For each candidate point, check if any profile line is above it
    ground_surface_points = []
    for x, y in sorted(x_groups.items()):
        # Create a vertical line at this x-coordinate
        vertical_line = LineString([(x, y - 1000), (x, y + 1000)])
        
        # Check intersections with all profile lines
        is_topmost = True
        for profile_line in coord_lists:
            line = LineString(profile_line)
            if line.length == 0:
                continue
            
            # Find intersection with this profile line
            intersection = line.intersection(vertical_line)
            if not intersection.is_empty:
                # Get the y-coordinate of the intersection
                if hasattr(intersection, 'y'):
                    # Single point intersection
                    if intersection.y > y + 1e-6:  # Allow small numerical tolerance
                        is_topmost = False
                        break
                elif hasattr(intersection, 'geoms'):
                    # Multiple points or line intersection
                    for geom in intersection.geoms:
                        if hasattr(geom, 'y') and geom.y > y + 1e-6:
                            is_topmost = False
                            break
                    if not is_topmost:
                        break
        
        if is_topmost:
            ground_surface_points.append((x, y))
    
    # Ensure we have at least 2 points
    if len(ground_surface_points) < 2:
        return LineString([])

    return LineString(ground_surface_points)


def build_ground_surface_from_polygons(polygons):
    """
    Derive the ground surface and domain polygon from material-zone polygons.

    The domain polygon is the union of all input polygons (the total extent of the
    model). The ground surface is the upper boundary of that union, traced from the
    top-left corner to the top-right corner.

    Parameters:
        polygons (list of dict): each dict has a 'polygon' key (shapely Polygon).

    Returns:
        tuple(LineString, shapely.geometry.Polygon): the ground surface (empty
        LineString if it cannot be formed) and the domain polygon.
    """
    domain = unary_union([p['polygon'] for p in polygons])
    if domain.geom_type == 'MultiPolygon':
        # Disjoint zones: use the largest connected component for the surface.
        domain = max(domain.geoms, key=lambda g: g.area)

    ring = list(domain.exterior.coords)[:-1]  # drop the closing duplicate vertex
    n = len(ring)
    if n < 3:
        return LineString([]), domain

    # The leftmost and rightmost vertices (ties broken by highest elevation) split
    # the exterior ring into two arcs; the one with the higher mean elevation is
    # the top boundary.
    imin = min(range(n), key=lambda i: (ring[i][0], -ring[i][1]))
    imax = max(range(n), key=lambda i: (ring[i][0], ring[i][1]))

    def arc(a, b):
        path, i = [], a
        while True:
            path.append(ring[i])
            if i == b:
                break
            i = (i + 1) % n
        return path

    arc1, arc2 = arc(imin, imax), arc(imax, imin)
    mean_y = lambda a: sum(p[1] for p in a) / len(a)
    upper = arc1 if mean_y(arc1) >= mean_y(arc2) else arc2

    # Order the surface left-to-right.
    if upper[0][0] > upper[-1][0]:
        upper = upper[::-1]

    return LineString(upper), domain


# === v20 SSR-zone sentinel Mat IDs (polygon sheet) ===
# A polygon row whose Mat ID is NEGATIVE is not a material zone at all — it is an
# ANALYSIS OVERLAY for the shear-strength-reduction method. Overlay rows are never
# meshed, never material regions and never generate slices; they only tell the SSRM
# which elements to reduce. The three codes (and the display wording, which matches
# the template's row-6 echo formulas exactly) are:
#
#   -1  "SSR reduce"   search area   -- strength reduction applies ONLY inside
#   -2  "SSR hold"     exclusion     -- full strength inside, but can still yield
#   -3  "SSR elastic"  exclusion     -- linear elastic inside, cannot yield at all
#
# Composition: the reduce set is the union of the -1 rows minus the union of the
# -2/-3 rows; with no -1 row present it is the whole domain minus those exclusions.
# So exclusions carve holes out of search areas and out of the model as a whole.
SSR_ZONE_SENTINELS = {
    -1: 'reduce',
    -2: 'hold',
    -3: 'hold_elastic',
}
SSR_ZONE_LABELS = {
    'reduce': 'SSR reduce',
    'hold': 'SSR hold',
    'hold_elastic': 'SSR elastic',
}

# === v21 polygon Type words (polygon sheet, row 5) ===
# v21 retires the sentinel Mat IDs above in favour of an explicit Type dropdown, so
# the overlay kind is stated in words instead of encoded in the field that otherwise
# names a material. Blank means 'material' (that is what makes a v20 file's layout
# and a v21 file with an untouched Type row describe the same model).
#
# 'refine' is new in v21 and is NOT an SSR kind: it is a pure meshing overlay — a
# region that carries no material and no analysis meaning, only a local target
# element size. It is never meshed as a material region, so its Size is REQUIRED
# (a refine polygon with no size would be a no-op the user could not see).
POLYGON_TYPE_WORDS = {
    'material': 'material',
    'ssr reduce': 'reduce',
    'ssr hold': 'hold',
    'ssr elastic': 'hold_elastic',
    'refine': 'refine',
}

# === v12 reinforcement support-type presets (reinforce sheet, Type column) ===
# type -> (dir, appl). The same table the sheet holds in its hidden lookup block
# (reinforce!AB8:AD11), which its Dir and Appl formulas VLOOKUP: picking a Type
# fills both, and typing over either keeps what was typed until the Type is picked
# again. Module level rather than local to the loader so the Studio's reinforcement
# editor fills its Dir/Appl combos from THIS table instead of restating it -- the
# file, the sheet and the editor cannot then disagree about what a Type means.
#
# A blank Type is a generic tensile line: the sheet leaves Dir/Appl blank for it and
# the loader falls back to ('tangent', 'active'), the pre-v12 behavior.
REINFORCE_TYPE_PRESETS = {
    'geosynthetic': ('tangent', 'active'),
    'nail':         ('axial',   'passive'),
    'tieback':      ('axial',   'active'),
    'anchor':       ('axial',   'active'),
}
#: What a line with no Type gets: a generic tensile line.
REINFORCE_TYPE_DEFAULT = ('tangent', 'active')


def _reinf_word(value, default):
    """A reinforcement vocabulary cell's word, or ``default`` when it is blank.

    Blank is None, NaN or an empty string -- never a falsy NUMBER. The obvious
    ``value or default`` reads ``0.0`` as an empty cell, which is how a nonsense
    ``Dir`` got written back to the sheet as the Type's preset instead of as
    itself, hiding from the loader the value that had actually been entered.
    """
    if value is None or (isinstance(value, float) and value != value):
        return default
    text = str(value).strip().lower()
    return text if text else default


def _opt_size_cell(df, row_idx, col_idx, where):
    """Optional local mesh size cell. Blank -> None; a non-numeric or non-positive
    entry is a loud error naming the block, never a silently ignored refinement."""
    try:
        raw = df.iloc[row_idx, col_idx]
    except IndexError:
        return None
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = str(raw).strip()
    if s == '' or s.lower() == 'nan':
        return None
    try:
        val = float(raw)
    except (TypeError, ValueError):
        raise ValueError(
            f"{where} declares a mesh Size of {raw!r}, which is not a number. "
            "Leave it blank to use the global target size.")
    if not (val > 0):
        raise ValueError(
            f"{where} declares a mesh Size of {val}. It must be positive (leave it "
            "blank to use the global target size).")
    return val


def _parse_polygon_sheet(xls, materials, template_version=20):
    """
    Parse the optional 'polygon' sheet into material zones and analysis/mesh overlays.

    Polygon p occupies columns (x_col, y_col) with x_col = 3*p and a "Polygon #N"
    header in row 4. Everything below the header moved in v21:

        v20 and earlier          v21
        row 5  Mat ID            row 5  Type      (blank = material)
        row 6  name echo         row 6  Mat ID
        rows 8+ vertices         row 7  name echo
                                 row 8  Size      (optional, any polygon)
                                 rows 10+ vertices

    v20 encodes an SSR overlay as a NEGATIVE Mat ID (see :data:`SSR_ZONE_SENTINELS`);
    v21 states it in the Type cell (see :data:`POLYGON_TYPE_WORDS`). Both readers stay
    live forever — an existing file is never re-saved just to keep loading — and the
    two produce identical in-memory zones, so a v20 model and its v21 re-save are the
    same analysis.

    mat_id is stored 0-based to match the profile_lines / materials convention.

    Returns:
        tuple(list, list, list):
          - material zones: [{'polygon': shapely Polygon, 'mat_id': int,
                              'size': float|None}, ...]
          - SSR zones:      [{'kind': 'reduce'|'hold'|'hold_elastic',
                              'polygon': [(x, y), ...], 'label': str,
                              'size': float|None}, ...]
          - refine zones:   [{'polygon': [(x, y), ...], 'size': float}, ...]  (v21 only)
        All empty if the sheet is absent or contains no polygons.
    """
    if 'polygon' not in xls.sheet_names:
        return [], [], []

    df = xls.parse('polygon', header=None)
    if template_version >= 21:
        header_row, type_row, mat_id_row = 3, 4, 5
        size_row, coords_start_row = 7, 9            # Excel rows 8 and 10
    else:
        header_row, type_row, mat_id_row = 3, None, 4
        size_row, coords_start_row = None, 7         # Excel row 8
    polygons = []
    ssr_zones = []
    refine_zones = []

    col = 0
    while col + 1 < df.shape[1]:
        x_col, y_col = col, col + 1

        # Stop when the block header is empty (no further polygon blocks).
        try:
            header_val = str(df.iloc[header_row, x_col]).strip()
        except Exception:
            break
        if not header_val or header_val.lower() == 'nan':
            break

        # Read vertices until the first fully empty row.
        coords = []
        row = coords_start_row
        while row < df.shape[0]:
            try:
                x_val, y_val = df.iloc[row, x_col], df.iloc[row, y_col]
            except Exception:
                break
            if pd.isna(x_val) and pd.isna(y_val):
                break
            if pd.notna(x_val) and pd.notna(y_val):
                coords.append((float(x_val), float(y_val)))
            row += 1

        # A header with no vertices is an unused block (template default) — skip.
        if coords:
            if len(coords) >= 2 and coords[0] == coords[-1]:
                coords = coords[:-1]  # drop explicit closing vertex (Shapely closes)
            col_letter = chr(65 + col) if col < 26 else f"col {col + 1}"
            where = f"Polygon in block starting at column {col_letter}"
            if len(coords) < 3:
                raise ValueError(
                    f"Each polygon must have at least 3 vertices. {where} has "
                    f"{len(coords)}.")

            poly = Polygon(coords)
            if not poly.is_valid:
                raise ValueError(
                    f"{where} is not a valid polygon (self-intersecting or "
                    f"degenerate).")

            size = (None if size_row is None
                    else _opt_size_cell(df, size_row, y_col, where))

            # -- v21 Type word. Blank/absent falls through to the Mat ID reading
            # below, which is what makes an untouched Type row mean 'material'.
            kind = 'material'
            if type_row is not None:
                type_raw = df.iloc[type_row, y_col] if type_row < df.shape[0] else None
                type_str = '' if (type_raw is None
                                  or (isinstance(type_raw, float) and pd.isna(type_raw))) \
                    else str(type_raw).strip()
                if type_str and type_str.lower() != 'nan':
                    kind = POLYGON_TYPE_WORDS.get(type_str.lower())
                    if kind is None:
                        # A mis-typed Type must never be read as 'material': an
                        # 'ssr hold' zone that silently became a material region
                        # would mesh a hole into the section AND drop the
                        # constraint, and the run would report a number nobody
                        # asked for. Name the whole vocabulary and stop.
                        raise ValueError(
                            f"{where} declares Type {type_str!r}, which is not a "
                            f"recognized polygon type. Use one of: "
                            f"{', '.join(sorted(POLYGON_TYPE_WORDS))} (or leave the "
                            f"Type cell blank for a material zone).")

            if kind == 'refine':
                # Pure meshing overlay: no material, no analysis meaning, only a
                # local element size — so without a Size it does nothing at all.
                if size is None:
                    raise ValueError(
                        f"{where} declares Type 'refine' but no Size. A refine "
                        "polygon carries no material and no analysis meaning — its "
                        "only effect is the local target element size, so the Size "
                        "cell is required.")
                refine_zones.append({'polygon': list(coords), 'size': size})
                col += 3
                continue

            if kind != 'material':
                ssr_zones.append({'kind': kind,
                                  'polygon': list(coords),
                                  'label': SSR_ZONE_LABELS[kind],
                                  'size': size})
                col += 3
                continue

            mat_id_val = df.iloc[mat_id_row, y_col]
            try:
                raw_id = int(float(mat_id_val))
            except (ValueError, TypeError):
                raw_id = None

            if raw_id is not None and raw_id < 0:
                # v20 SSR overlay row. An unrecognized negative code is a typo, and a
                # typo must never be silently dropped (a mis-typed -4 that vanished
                # would run the SSRM unconstrained and report a number nobody asked
                # for), so name the whole vocabulary and stop.
                zone_kind = SSR_ZONE_SENTINELS.get(raw_id)
                if zone_kind is None:
                    raise ValueError(
                        f"{where} has an "
                        f"unrecognized negative Mat ID ({raw_id}). Negative Mat IDs are "
                        f"SSR zone overlays: -1 = 'SSR reduce' (strength reduction "
                        f"applies only inside), -2 = 'SSR hold' (full strength inside, "
                        f"can still yield), -3 = 'SSR elastic' (linear elastic inside, "
                        f"cannot yield). Use a positive Mat ID (1..{len(materials)}) "
                        f"for a material zone.")
                ssr_zones.append({'kind': zone_kind,
                                  'polygon': list(coords),
                                  'label': SSR_ZONE_LABELS[zone_kind],
                                  'size': size})
                col += 3
                continue

            mat_id = raw_id - 1 if raw_id is not None else None  # 1-based -> 0-based
            if mat_id is None or mat_id < 0 or mat_id >= len(materials):
                _sentinels = ("" if template_version >= 21 else
                              ", or be one of the SSR zone sentinels -1 / -2 / -3")
                raise ValueError(
                    f"{where} has an invalid "
                    f"Mat ID ({mat_id_val!r}); it must reference a material in the "
                    f"'mat' sheet (1..{len(materials)}){_sentinels}.")

            polygons.append({'polygon': poly, 'mat_id': mat_id, 'size': size})

        col += 3  # next block (A->D->G->...)

    return polygons, ssr_zones, refine_zones


def _validate_polygons_no_overlap(polygons):
    """Material zones must tile the section without overlapping — adjacent zones
    share matching edges; a zone is never drawn on top of another. Overlapping
    zones mesh incorrectly (a high-conductivity zone bridges over a low-
    conductivity barrier and can inflate the seepage flowrate several-fold), so
    reject them at load time instead of silently producing wrong results.

    Touching at shared edges/vertices is fine (zero overlap area); only a positive
    intersection area is an error.
    """
    tol = 1e-6
    for i in range(len(polygons)):
        pi = polygons[i]['polygon']
        for j in range(i + 1, len(polygons)):
            area = pi.intersection(polygons[j]['polygon']).area
            if area > tol:
                mi, mj = polygons[i].get('mat_id'), polygons[j].get('mat_id')
                m1 = mi + 1 if mi is not None else '?'
                m2 = mj + 1 if mj is not None else '?'
                raise ValueError(
                    f"Material zones overlap: polygon #{i + 1} (Mat ID {m1}) and "
                    f"polygon #{j + 1} (Mat ID {m2}) overlap by {area:.4g} sq units. "
                    f"Material zones must tile the section without overlapping — adjacent "
                    f"zones share matching edges, and a zone that sits inside or cuts "
                    f"through another (a lens, or a dam core) must be carved out of its "
                    f"neighbor (e.g. a shell around a core is one concave polygon with a "
                    f"notch), not drawn on top of it.")


class PulloutProfile:
    """The pullout resistance of one reinforcement line, integrated along it.

    A pullout law states a resistance per unit length of line, r(s) — the force
    the soil–reinforcement interface can hold back per unit of embedment at arc
    length s from end 1. The capacity the line can develop at s is the resistance
    earned over the embedment between s and whichever end is being counted from,
    i.e. the integral of r. This object stores that integral once, sampled on a
    fine grid, so the envelope can be evaluated at any point by interpolation
    instead of re-integrating per trial surface.

    Build one with :meth:`from_rate` and hand it to
    :func:`reinforce_available_tension` as ``pullout=``.
    """

    __slots__ = ("length", "s", "cum", "total")

    def __init__(self, length, s, cum):
        self.length = float(length)
        self.s = np.asarray(s, dtype=float)
        self.cum = np.asarray(cum, dtype=float)
        self.total = float(self.cum[-1]) if len(self.cum) else 0.0

    @classmethod
    def from_rate(cls, rate, length, n=201):
        """Integrate a rate function r(s), s in [0, length], by the trapezoid
        rule on ``n`` samples.

        The rate is sampled rather than solved because it is not analytic: the
        line crosses material boundaries and the water table, and the effective
        overburden it reads is discontinuous at both. ``n`` = 201 puts a sample
        every half percent of the line, which resolves those steps far finer
        than the geometry that produced them.
        """
        length = float(length)
        if length <= 0 or n < 2:
            return cls(length, np.array([0.0]), np.array([0.0]))
        s = np.linspace(0.0, length, int(n))
        r = np.array([float(rate(float(v))) for v in s], dtype=float)
        r = np.where(np.isfinite(r), r, 0.0)
        cum = np.concatenate(([0.0], np.cumsum(0.5 * (r[1:] + r[:-1]) * np.diff(s))))
        return cls(length, s, cum)

    def _at(self, d):
        return float(np.interp(d, self.s, self.cum, left=0.0, right=self.total))

    def from_end1(self, d1):
        """Resistance developed over the embedment d1 back to end 1."""
        return self._at(d1)

    def from_end2(self, d2):
        """Resistance developed over the embedment d2 back to end 2."""
        return self.total - self._at(self.length - d2)


def reinforce_available_tension(d1, d2, t_max, lp1, lp2, tend1=0.0, tend2=0.0,
                                pullout=None):
    """Available tensile force at a point along a reinforcement line — the
    capacity envelope shared by the LEM point list and the FEM element taper:

        T = min( Tmax,
                 Tend1 + int_0^s r,      (integral from end 1 to the point)
                 Tend2 + int_s^L r )     (integral from the point to end 2)

    d1/d2 are the distances from the point to end 1 / end 2. Tend* are end
    anchorage capacities (plate/connection/anchor); 0 reproduces the classical
    friction-only taper exactly. One implementation for both engines, so the
    two can never drift.

    The pullout rate r comes from one of two laws:

    - **Constant rate** (``pullout=None``, the default): the bond develops the
      full tensile capacity over the pullout length, r = Tmax/Lp, so the
      integrals collapse to the linear ramps Tmax*d1/Lp1 and Tmax*d2/Lp2. An
      Lp of 0 means that end is fully anchored — Tmax is available at the end
      itself.
    - **Overburden-dependent** (``pullout`` = a :class:`PulloutProfile`): r
      varies along the line with the effective overburden, and the integrals
      are read off the profile. Lp1/Lp2 play no part; the profile carries the
      whole law.
    """
    if pullout is not None:
        cap1 = tend1 + pullout.from_end1(d1)
        cap2 = tend2 + pullout.from_end2(d2)
    else:
        cap1 = t_max if lp1 <= 0 else tend1 + t_max * d1 / lp1
        cap2 = t_max if lp2 <= 0 else tend2 + t_max * d2 / lp2
    return max(0.0, min(t_max, cap1, cap2))


def _reinforce_line_points(x1, y1, x2, y2, Tmax, Tres, Lp1, Lp2, E, Area,
                           Tend1=0.0, Tend2=0.0, pullout=None):
    """Build the LEM tension-distribution point list for ONE reinforcement line
    from its raw endpoints, pullout lengths, and end anchorage. Returns [] for a
    zero-length line.

    The available tension is the capacity envelope of
    :func:`reinforce_available_tension`. Under the constant-rate law that
    envelope is piecewise linear and the point list holds its breakpoints, so
    linear interpolation between points reproduces it exactly; under the
    overburden-dependent law (``pullout``) it is a curve and the points are a
    dense sampling of it.

    Extracted from load_slope_data so the same derivation can be reused (e.g. by
    the GUI) to rebuild the display format after editing the raw line data."""
    import math
    line_length = math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
    if line_length == 0:
        return []

    # Unit vector from (x1,y1) to (x2,y2)
    dx = (x2 - x1) / line_length
    dy = (y2 - y1) / line_length

    def T_at(s):
        return reinforce_available_tension(s, line_length - s, Tmax, Lp1, Lp2,
                                           Tend1, Tend2, pullout=pullout)

    if pullout is not None:
        # Overburden-dependent law: the envelope is a curve, not a polyline, so
        # there are no kinks to solve for. Sample it on the profile's own grid,
        # thinned to keep the stored point list a readable size — the LEM reads
        # the envelope function directly at the crossing, so these points are
        # for drawing and for hand-built models, not for the solve.
        step = max(1, (len(pullout.s) - 1) // 40)
        cands = set(float(v) for v in pullout.s[::step])
        cands.update((0.0, line_length))
    else:
        # Candidate breakpoints: the endpoints plus every kink of the envelope —
        # where each end's ramp reaches Tmax, and where the two ramps cross.
        cands = {0.0, line_length}
        if Lp1 > 0 and Tend1 < Tmax:
            cands.add(min(line_length, (Tmax - Tend1) * Lp1 / Tmax))
        if Lp2 > 0 and Tend2 < Tmax:
            cands.add(max(0.0, line_length - (Tmax - Tend2) * Lp2 / Tmax))
        if Lp1 > 0 and Lp2 > 0:
            m1, m2 = Tmax / Lp1, Tmax / Lp2
            s_x = (Tend2 + m2 * line_length - Tend1) / (m1 + m2)
            # only a breakpoint when the ramps cross BELOW the Tmax plateau
            if 0.0 < s_x < line_length and (Tend1 + m1 * s_x) < Tmax:
                cands.add(s_x)

    # Emit ordered, deduplicated breakpoints. The stored Tres is the residual
    # capacity at the breakpoint: the smaller of the material residual and the
    # capacity the envelope develops there. Bond slip is perfectly plastic — an
    # element that slips keeps carrying the ramped capacity — so the only thing
    # that can cut a point's post-peak capacity below the envelope is the
    # material residual itself. An UNSET Tres (NaN) stays unset everywhere along
    # the line — no post-peak drop anywhere — rather than collapsing to a number
    # through min().
    tol = max(1e-9, 1e-9 * line_length)
    line_points = []
    for s in sorted(cands):
        if line_points and abs(s - line_points[-1]["_s"]) < tol:
            continue
        T = T_at(s)
        if Tres != Tres:                      # NaN: unset, no post-peak drop
            tres_pt = float('nan')
        else:
            tres_pt = min(Tres, T)
        line_points.append({"X": x1 + s * dx, "Y": y1 + s * dy, "T": T,
                            "Tres": tres_pt, "E": E, "Area": Area, "_s": s})
    for p in line_points:
        del p["_s"]

    return line_points


def build_reinforce_lines(reinforcement_lines):
    """Map raw FEM-format reinforcement dicts (x1,y1,x2,y2,t_max,t_res,lp1,lp2,E,
    area, and optionally tend1/tend2) to the LEM display/analysis format (a list
    of tension point-lists). Lines resolving to fewer than 2 points are dropped,
    matching load_slope_data."""
    lines = []
    for r in reinforcement_lines:
        pts = _reinforce_line_points(
            r["x1"], r["y1"], r["x2"], r["y2"], r["t_max"], r["t_res"],
            r["lp1"], r["lp2"], r["E"], r["area"],
            r.get("tend1", 0.0), r.get("tend2", 0.0),
            pullout=r.get("_pullout_profile"))
        if len(pts) >= 2:
            lines.append(pts)
    return lines


# ---------------------------------------------------------------------------
# Overburden-dependent pullout
#
# The alternative to a development length: state the interface strength and let
# the resistance follow the depth of burial. Per unit length of a planar
# reinforcement, with soil bearing on both faces,
#
#     r(s) = 2 * ( a + sigma'_v(s) * tan(delta) )
#
# with a the soil-reinforcement adhesion and delta the interface friction angle
# (the 'Adhesion' and 'Delta' columns of the reinforce sheet). sigma'_v is the
# EFFECTIVE overburden at the point: the weight of the soil column standing
# above it, less the pore pressure the model declares there. Both columns blank
# is the default and leaves the constant-rate law untouched.
#
# The FHWA pullout-capacity form, F* and alpha against the same overburden, is
# this law with a = 0 and delta = atan(F* * alpha).
# ---------------------------------------------------------------------------

def _ground_y(ground_surface, x):
    """Ground elevation above x, or None off the ends of the section."""
    if ground_surface is None or ground_surface.is_empty:
        return None
    coords = list(ground_surface.coords)
    xs = [p[0] for p in coords]
    if x < min(xs) or x > max(xs):
        return None
    if xs[0] > xs[-1]:
        coords = coords[::-1]
        xs = xs[::-1]
    for k in range(len(coords) - 1):
        x0, y0 = coords[k]
        x1, y1 = coords[k + 1]
        if x0 <= x <= x1 and x1 > x0:
            return y0 + (x - x0) * (y1 - y0) / (x1 - x0)
    return coords[-1][1] if x >= xs[-1] else coords[0][1]


def _column_total_stress(slope_data, x, y, y_ground, y_water):
    """Total vertical stress at (x, y) from the soil standing above it.

    The same quantity ``generate_slices`` accumulates as ``sum_gam_h`` and
    divides into the slice weight: every material zone the vertical column
    crosses contributes its thickness times its unit weight, saturated below
    the model's water table where the material declares a ``gamma_sat`` and
    moist above it. Splitting the column at the water table before intersecting
    the zones reproduces the slice generator's band-by-band split exactly.

    ``y_water`` is NaN when the model defines no water table, in which case the
    moist unit weight applies throughout — the same fallback the slice
    generator takes.
    """
    if y_ground is None or y_ground <= y:
        return 0.0
    materials = slope_data.get('materials') or []
    polygons = slope_data.get('polygons') or []
    if y_water is None or not np.isfinite(y_water):
        segments = [(y, y_ground, False)]
    else:
        y_split = min(max(float(y_water), y), y_ground)
        segments = [(y, y_split, True), (y_split, y_ground, False)]
    total = 0.0
    for lo, hi, saturated in segments:
        if hi - lo <= 0:
            continue
        column = LineString([(x, lo), (x, hi)])
        # A zone owns the part of the column no zone before it already claimed.
        # A shapely polygon includes its own boundary, so a column standing
        # exactly on the vertical edge two zones share — which is where a
        # reinforcement line ends in every zoned retaining structure, the back
        # of the reinforced fill — intersects BOTH of them over its whole
        # length and would be weighed twice. Claiming each length once makes the
        # zone listed first own an interface, the same convention
        # :func:`_material_at_point` already applies to a point on a boundary.
        claimed = None
        for poly in polygons:
            shape = poly.get('polygon') if isinstance(poly, dict) else poly
            mat_id = poly.get('mat_id') if isinstance(poly, dict) else None
            if shape is None or mat_id is None:
                continue
            try:
                piece = column.intersection(shape)
                if claimed is not None:
                    piece = piece.difference(claimed)
            except Exception:                            # pragma: no cover
                continue
            if piece.is_empty:
                continue
            try:
                mat = materials[int(mat_id)]
            except (IndexError, TypeError, ValueError):
                continue
            claimed = piece if claimed is None else claimed.union(piece)
            g_sat = mat.get('gamma_sat')
            gamma = (g_sat if (saturated and g_sat is not None) else mat['gamma'])
            total += piece.length * float(gamma)
    return total


def _material_at_point(slope_data, x, y):
    """The material zone containing (x, y), or None."""
    materials = slope_data.get('materials') or []
    pt = Point(x, y)
    for poly in slope_data.get('polygons') or []:
        shape = poly.get('polygon') if isinstance(poly, dict) else poly
        mat_id = poly.get('mat_id') if isinstance(poly, dict) else None
        if shape is None or mat_id is None:
            continue
        try:
            if shape.covers(pt):
                return materials[int(mat_id)]
        except (IndexError, TypeError, ValueError):
            continue
    return None


def reinforce_effective_overburden(slope_data, x, y, y_water=None):
    """Effective vertical stress at (x, y): the weight of the soil column above
    it less the pore pressure the model declares there.

    The pore pressure follows the ``u`` option of the material the point lands
    in, exactly as a slice base does — a piezometric head, a pore-pressure ratio
    on the soil column, or an interpolated seepage field — and a model with no
    water reads zero. Suction is not credited: a point above the water table
    contributes its total stress, never more.
    """
    from .water import water_table_y
    if y_water is None:
        y_water = float(water_table_y(slope_data, x))
    y_ground = _ground_y(slope_data.get('ground_surface'), x)
    sigma_v = _column_total_stress(slope_data, x, y, y_ground, y_water)
    if sigma_v <= 0:
        return 0.0
    material = _material_at_point(slope_data, x, y)
    from .generators import _pore_pressure
    u = _pore_pressure(slope_data, material, x, y, sigma_v)
    return max(0.0, sigma_v - u)


def reinforce_pullout_profile(line, slope_data, n=201):
    """The :class:`PulloutProfile` for one raw reinforcement line, or None when
    the line does not use the overburden-dependent law.

    ``line`` is a ``reinforcement_lines`` dict; the law is active when both its
    ``adhesion`` and ``delta`` are present and finite. The rate is divided by
    the line's ``Spacing`` for the same reason every other capacity term is —
    everything downstream of the loader is per unit width of slope — so a
    continuous sheet, whose Spacing is blank, is unaffected.
    """
    adhesion = line.get('adhesion')
    delta = line.get('delta')
    if adhesion is None or delta is None:
        return None
    adhesion, delta = float(adhesion), float(delta)
    if not (np.isfinite(adhesion) and np.isfinite(delta)):
        return None

    x1, y1 = float(line['x1']), float(line['y1'])
    x2, y2 = float(line['x2']), float(line['y2'])
    length = math.hypot(x2 - x1, y2 - y1)
    if length <= 0:
        return None
    # The law reads the soil column above the line. Without material zones there
    # is no column to read and the overburden term would silently vanish,
    # leaving the adhesion alone — a much weaker line, and nothing on screen
    # saying why.
    if not (slope_data.get('polygons') and slope_data.get('materials')):
        warnings.warn(
            f"Reinforcement line {line.get('label') or ''!r} states an "
            f"overburden-dependent pullout law (Adhesion and Delta), but this "
            f"model carries no material zones, so the effective overburden along "
            f"the line reads zero and only the adhesion develops any resistance.")
    dx, dy = (x2 - x1) / length, (y2 - y1) / length
    tan_delta = math.tan(math.radians(delta))
    # The loader has already divided the entered per-element capacities by
    # Spacing; the rate is divided here for the same reason, so a discrete
    # support's envelope stays in one convention end to end.
    spacing = float(line.get('spacing') or 1.0)

    def rate(s):
        x, y = x1 + s * dx, y1 + s * dy
        sigma_v = reinforce_effective_overburden(slope_data, x, y)
        return 2.0 * (adhesion + sigma_v * tan_delta) / spacing

    return PulloutProfile.from_rate(rate, length, n=n)


def attach_reinforce_pullout(slope_data, n=201):
    """Build each reinforcement line's pullout profile and refresh the derived
    LEM point lists.

    Called once by :func:`load_slope_data`, and by any caller that assembles
    ``slope_data`` itself (the Studio document, the GeoStudio importer) after
    the geometry, materials and water are in place — the profile reads all
    three. Lines on the constant-rate law get ``_pullout_profile = None`` and
    behave exactly as before.
    """
    lines = slope_data.get('reinforcement_lines') or []
    if not lines:
        return slope_data
    for r in lines:
        r['_pullout_profile'] = reinforce_pullout_profile(r, slope_data, n=n)
        r['_pullout_key'] = _pullout_key(r)
    if any(r.get('_pullout_profile') is not None for r in lines):
        slope_data['reinforce_lines'] = build_reinforce_lines(lines)
    return slope_data


def _pullout_key(line):
    """The inputs a line's pullout profile was built from.

    Comparing this against the line's current values is how a profile is known
    to be stale: a Studio edit, a sensitivity sweep, or an importer writing new
    values all mutate the line dict in place, and a profile that outlived its
    inputs would answer with the old law and say nothing about it.
    """
    return tuple(line.get(k) for k in
                 ('adhesion', 'delta', 'spacing', 'x1', 'y1', 'x2', 'y2'))


def ensure_reinforce_pullout(slope_data):
    """Resolve any pullout profile that is missing or out of date.

    :func:`load_slope_data` builds them once, but a ``slope_data`` assembled in
    memory — by the Studio, by an importer, by a sweep, by a test — reaches an
    engine with the Adhesion/Delta columns filled and no profile behind them, or
    with a profile built before the last edit. Both engines call this on the way
    in, so the law is honored wherever the model came from. Files on the
    constant-rate law cost one tuple comparison per line.
    """
    lines = slope_data.get('reinforcement_lines') or []
    if not lines and 'reinforcement_lines' in slope_data:
        # The source list is the model's own statement about its reinforcement,
        # the empty list included. 'reinforce_lines' is derived from it, and a
        # derived list that outlives the row it came from is reinforcement the
        # model no longer has -- still drawn by plot.plot_reinforcement_lines,
        # and (before the slicer stopped falling back to it) still carried into
        # the factor of safety.
        if slope_data.get('reinforce_lines'):
            slope_data['reinforce_lines'] = []
        return slope_data
    for r in lines:
        if r.get('_pullout_key') != _pullout_key(r):
            attach_reinforce_pullout(slope_data)
            break
    return slope_data


# Highest input-template version this build can read. Bump together with the
# template (docs/inputs/input_template.xlsx, main!D5) and its reader support.
SUPPORTED_TEMPLATE_VERSION = 25

# The template version that inserted the 1D element size cell at main!D20, pushing
# every main-sheet run option below it down one row. Files at or above it are read
# (and written) with that one-row shift; older ones keep their original layout and
# carry no 1D size at all.
_ELEMENT_SIZE_1D_TEMPLATE_VERSION = 25

# The template version that introduced the water-load mode cell (main!D23). Below it
# a file cannot say who supplies the weight of standing water, so it always means
# 'manual' on read -- and an automatic-water model cannot be written into one at all
# without losing its reservoir (see save_slope_data_to_xlsx).
_WATER_LOADS_TEMPLATE_VERSION = 22

# === v19 run-option vocabularies (main sheet D14/D18) ===
# The template backs both cells with a dropdown, but a hand-edited or
# script-written file can carry anything, so the loader validates and raises.
# Silently ignoring an unknown value is the failure mode that made u='ru' zero
# pore pressure for a year (see the 'u' option check below).
LEM_METHODS = ('oms', 'janbu', 'bishop', 'corps', 'lowe', 'spencer', 'mprice', 'all')
MESH_ELEMENT_TYPES = ('tri3', 'tri6', 'quad4', 'quad8', 'quad9')

# v21 main!D22 -- how the FEM restrains the left/right truncation boundaries.
# 'rollers' is the historical hardwired behavior (u = 0, v free); 'fixed' clamps
# both components, which is what RS2 does and what a vendor-parity comparison needs.
SIDE_BC_OPTIONS = ('rollers', 'fixed')

# v22 amendment, main!D24 -- which surface family a deck that defines BOTH means.
# These are the words the CELL uses (the dropdown's own list); the loader normalizes
# them to the package's internal vocabulary, 'circular' / 'noncircular'.
SURFACE_FAMILY_OPTIONS = ('circular', 'non-circular')

#: The cell spelling of each internal surface-family name, for the writer.
SURFACE_FAMILY_CELL = {'circular': 'circular', 'noncircular': 'non-circular'}

# Optional circles-sheet search window (v19, J8:K17). Each entry maps the label
# in column J to the slope_data['search_window'] key its value in column K feeds.
# Order IS the sheet order: the reader walks rows 8..17 positionally.
SEARCH_WINDOW_KEYS = (
    'entry_x_min', 'entry_x_max', 'exit_x_min', 'exit_x_max',
    'center_box_x_min', 'center_box_x_max', 'center_box_y_min', 'center_box_y_max',
    'max_tangent_depth', 'min_slip_depth',
)


def _read_seep_bc_sheet(seep_df, sheet_name):
    """Parse one ``seep bc`` sheet into specified heads, fluxes, and the exit face.

    Layout (v15): exit face in B/C from row 5 down; then up to 5 BC blocks in
    E/F, H/I, K/L, N/O, Q/R. Row 3 of a block holds the TYPE cell (x column) and
    the VALUE (y column); coordinates start at row 5. A block ends the scan when
    its VALUE cell is empty.

    Backward compatibility: templates v14 and earlier put the literal label
    "Head:" in the type cell, so a block is a flux BC only when the type text
    starts with "flux" — every older file therefore still loads as a head BC.

    Head type (v18): a head block carries a ``kind`` — "reservoir" when the type
    cell text starts with "reservoir" (the submerged-only face), else "head" (a
    plain Dirichlet held at the value at every node of the polyline, at all
    times). A blank or legacy "Head:" type cell reads as "head".

    Value cells (v18): a block's VALUE cell may hold either a number (a constant
    head/flux, as always) or a string naming a ``tseep`` time series (a
    time-varying BC). This reader keeps whichever it finds — a numeric string is
    coerced to float, anything else is kept as the trimmed series-name string —
    and ``load_slope_data`` validates any string against the tseep series headers
    once both sheets are parsed. Files without a tseep sheet therefore still
    reject a non-numeric value there (see the load-time check).
    """
    bc = {"specified_heads": [], "specified_fluxes": [], "exit_face": []}

    def _bc_value(v):
        """Float when the cell is numeric; otherwise the trimmed string (a tseep
        series name, validated later). NaN/None never reach here — the caller
        breaks the block scan on an empty VALUE cell first."""
        try:
            return float(v)
        except (TypeError, ValueError):
            return str(v).strip()

    def _read_coords(x_col, y_col, start_row=4):
        coords = []
        row = start_row
        while row < seep_df.shape[0]:
            try:
                x_val = seep_df.iloc[row, x_col]
                y_val = seep_df.iloc[row, y_col]
                if pd.isna(x_val):
                    break
                if pd.notna(x_val) and pd.notna(y_val):
                    coords.append((float(x_val), float(y_val)))
            except Exception:
                break
            row += 1
        return coords

    bc["exit_face"] = _read_coords(1, 2)

    type_row = 2  # Excel row 3 (0-indexed)
    col = 4       # column E
    while col + 1 < seep_df.shape[1]:
        if seep_df.shape[0] <= type_row:
            break
        value = seep_df.iloc[type_row, col + 1]
        if pd.isna(value):
            break  # empty VALUE cell ends the block scan

        type_cell = seep_df.iloc[type_row, col]
        type_text = str(type_cell).strip().lower() if pd.notna(type_cell) else ""
        is_flux = type_text.startswith("flux")
        # v18: two Dirichlet head types share the head block layout. "reservoir" is
        # the submerged-only face (a node is held at the level only while submerged;
        # nodes above the water line become seepage-exit faces). Anything else that
        # is not a flux — "head", the legacy "Head:" label, or a blank type cell —
        # is a plain Dirichlet head held at the value at every node of the polyline.
        is_reservoir = type_text.startswith("reservoir")

        coords = _read_coords(col, col + 1)
        if is_flux:
            block = (col - 4) // 3 + 1
            if len(coords) < 2:
                raise ValueError(
                    f"Flux BC #{block} on sheet '{sheet_name}' has "
                    f"{len(coords)} coordinate(s). A flux BC is applied over the "
                    "edges of a polyline and needs at least 2 points."
                )
            bc["specified_fluxes"].append({"flux": _bc_value(value), "coords": coords})
        elif coords:
            kind = "reservoir" if is_reservoir else "head"
            bc["specified_heads"].append(
                {"head": _bc_value(value), "coords": coords, "kind": kind})

        col += 3  # E -> H -> K -> ...

    return bc


# tseep sheet cell geometry (0-based row/col into the header-less DataFrame), per the
# v18 template audit (cell_map §3). The time-series table sits top-left: B2 "time"
# header, time anchors down column B; series headers C2:G2..., values down each column.
# The controls live in column I (labels) / J (values); the save_times list is a
# vertical column under its own header in column J.
_TSEEP_TIME_COL = 1          # column B
_TSEEP_SERIES_COL0 = 2       # first series column C
_TSEEP_HEADER_ROW = 1        # Excel row 2 (headers)
_TSEEP_DATA_ROW0 = 2         # Excel row 3 (first data row)
_TSEEP_VAL_COL = 9           # column J (control values, save_times)
# control value cells (0-based row, in column J). These are POSITIONAL: the column-I
# labels are decorative and are never scanned, so a control's row index IS its identity.
_TSEEP_CONTROL_ROWS = {"duration": 2, "save_interval": 4, "stage_1": 6, "stage_2": 7}
# 0-based row 8 (Excel row 9, label "stability_time:") exists only from v22 on, so it is
# kept out of the shared map and added by version. The KEY is always present in the parsed
# dict (None on an older file), which is what keeps a v18 model round-tripping through a
# v22 template without growing or losing a field.
_TSEEP_STABILITY_TIME_ROW0 = 8    # Excel row 9 (v22+)
# save_times header row (0-based). v18-v21 head the list at J10 with values from J11
# down; the reviewed v22 template drops the header one row (J11, values from J12) to
# leave a blank line under stability_time. The controls above it did NOT move, so this
# is the only tseep anchor that is version-dependent.
_TSEEP_SAVE_TIMES_HDR_ROW0 = 9    # Excel row 10 (v18-v21)
_TSEEP_SAVE_TIMES_HDR_ROW0_V22 = 10   # Excel row 11 (v22+)


def _tseep_control_rows(version):
    """The tseep control label -> 0-based row map for a template version.

    ``stability_time`` joins the four original controls at v22; on an older template
    the row does not exist and the control is simply absent from the map."""
    rows = dict(_TSEEP_CONTROL_ROWS)
    if (version or 0) >= 22:
        rows["stability_time"] = _TSEEP_STABILITY_TIME_ROW0
    return rows


def _tseep_save_times_rows(version):
    """``(header_row0, first_value_row0)`` for the save_times column, 0-based."""
    hdr = (_TSEEP_SAVE_TIMES_HDR_ROW0_V22 if (version or 0) >= 22
           else _TSEEP_SAVE_TIMES_HDR_ROW0)
    return hdr, hdr + 1


def _parse_tseep_sheet(xls, template_version=18):
    """Parse the optional ``tseep`` (transient seepage) sheet into a dict, or return
    ``None`` when the sheet is absent OR present-but-empty.

    Sheet presence with actual data = transient enabled; an absent or all-blank sheet
    means steady behavior, so the ``tseep`` key is simply omitted from ``slope_data``
    and the file loads bit-identically to a pre-v18 workbook. (The master template
    ships this sheet with only its labels/headers and every value blank, so it parses
    to ``None`` and adds no key.)

    Returned dict (all times/values in the file's declared ``time`` unit — xslope never
    converts)::

        {
          "times":         [float, ...],          # the shared time axis (column B anchors)
          "series":        {name: [float|None]},  # aligned to times; None = no breakpoint
          "duration":      float | None,
          "save_interval": float | None,
          "save_times":    [float, ...],          # explicit extra save times (column J)
          "stage_1":       float | None,          # rapid-drawdown stage times
          "stage_2":       float | None,
          "stability_time": float | None,         # v22+: which instant a stability run reads
        }

    ``stability_time`` is the single-run extraction time: the instant an LEM or FEM run
    with ``u = seep`` pulls its pore pressures out of a transient solution. It is a
    stability-consumption time, not a control on the seepage march — blank means
    unspecified, and an unspecified time resolves to the LAST saved frame. The key is
    always present; on a pre-v22 template the row does not exist and it reads ``None``.

    A series column is registered only when its header is non-blank AND it carries at
    least one value, so the template's five default (blank) ``t1..t5`` headers do not
    manufacture empty series. Values keep their gaps (``None``): with linear
    interpolation a blank between anchors is identical to filling it in, so each series
    supplies values only at its own breakpoints.
    """
    if 'tseep' not in xls.sheet_names:
        return None

    df = xls.parse('tseep', header=None)
    nrows, ncols = df.shape

    def _num(r, c):
        if r >= nrows or c >= ncols:
            return None
        v = df.iloc[r, c]
        if pd.isna(v):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            raise ValueError(
                f"tseep sheet, cell {cell_ref(r + 1, c + 1)}: expected a number, got "
                f"{v!r}.")

    # --- shared time axis (column B, from the first data row down to the first blank) ---
    times = []
    r = _TSEEP_DATA_ROW0
    while r < nrows:
        t = df.iloc[r, _TSEEP_TIME_COL] if _TSEEP_TIME_COL < ncols else None
        if pd.isna(t):
            break
        try:
            times.append(float(t))
        except (TypeError, ValueError):
            raise ValueError(
                f"tseep sheet, cell {cell_ref(r + 1, _TSEEP_TIME_COL + 1)}: the time "
                f"column must be numeric, got {t!r}.")
        r += 1
    n_times = len(times)

    # --- named series (columns C onward; header text is the series name) ---
    series = {}
    if n_times:
        c = _TSEEP_SERIES_COL0
        while c < ncols:
            hdr = df.iloc[_TSEEP_HEADER_ROW, c] if _TSEEP_HEADER_ROW < nrows else None
            if hdr is None or pd.isna(hdr) or str(hdr).strip() == '':
                c += 1
                continue
            name = str(hdr).strip()
            vals = [_num(_TSEEP_DATA_ROW0 + i, c) for i in range(n_times)]
            if any(v is not None for v in vals):     # skip header-only (blank) columns
                if name in series:
                    raise ValueError(
                        f"tseep sheet: duplicate series name {name!r}. Series headers "
                        f"(row 2) must be unique.")
                series[name] = vals
            c += 1

    # --- controls (column J) ---
    controls = {k: _num(row, _TSEEP_VAL_COL)
                for k, row in _tseep_control_rows(template_version).items()}
    # stability_time has no cell before v22; the key still exists, blank.
    controls.setdefault("stability_time", None)

    # --- explicit save_times (vertical list under the column-J header) ---
    save_times = []
    _, r = _tseep_save_times_rows(template_version)
    while r < nrows:
        v = _num(r, _TSEEP_VAL_COL)
        if v is None:
            break
        save_times.append(v)
        r += 1

    enabled = (bool(times) or bool(series) or bool(save_times)
               or any(v is not None for v in controls.values()))
    if not enabled:
        return None

    return {
        "times": times,
        "series": series,
        "duration": controls["duration"],
        "save_interval": controls["save_interval"],
        "save_times": save_times,
        "stage_1": controls["stage_1"],
        "stage_2": controls["stage_2"],
        "stability_time": controls["stability_time"],
    }


def load_slope_data(filepath, dest=None, overwrite=False, require_analysis_data=True):
    """
    This function reads input data from various Excel sheets and parses it into
    structured components used throughout the slope stability analysis framework.
    It handles circular and non-circular failure surface data, reinforcement, piezometric
    lines, and distributed loads.

    ``require_analysis_data=False`` loads a model that is not yet runnable — no
    failure surfaces, no mesh, and no seepage boundary conditions (e.g. a partially
    built model saved from an editor, or a tutorial starter file). Such a model is
    validated on the seepage-only path (unit weights may be blank); everything else
    is checked as usual. The default (True) keeps the analysis-entry behavior: a
    file with none of the four is rejected.

    ``filepath`` is a workbook (.xlsx) or a project package (.xslz). A package is
    unpacked first and the extracted workbook loaded, because the sidecars this
    function reads — ``{base}_mesh.json``, ``{base}_seep.csv`` — have to exist as
    files beside the workbook, and so do the results the solvers will write next.
    ``dest`` and ``overwrite`` are passed to :func:`xslope.unpack` and are meaningful
    only for a package: by default it extracts to a folder named for the package,
    beside it, and raises rather than write over a folder that is already there.

    Validation is enforced to ensure required geometry and material information is present:
    - Circular failure surface: must contain at least one valid row with Xo and Yo
    - Non-circular failure surface: required if no circular data is provided
    - Profile lines: must contain at least one valid set, and each line must have ≥ 2 points
    - Materials: must match the number of profile lines
    - Piezometric line: only included if it contains ≥ 2 valid rows
    - Distributed loads and reinforcement: each block must contain ≥ 2 valid entries

    Raises:
        ValueError: if required inputs are missing or inconsistent.

    Returns:
        dict: Parsed and validated global data structure for analysis
    """

    if is_package(filepath):
        filepath = unpack(filepath, dest=dest, overwrite=overwrite)
    elif dest is not None or overwrite:
        raise ValueError(
            "dest= and overwrite= apply to unpacking a .xslz project package. "
            f"{os.path.basename(str(filepath))} is a workbook, which is loaded "
            "where it sits.")

    # Parse from an in-memory copy of the workbook, never from an open handle on it.
    #
    # ``pd.ExcelFile(path)`` keeps the .xlsx open for as long as the object lives, and
    # it has no ``__del__`` -- so the handle is released only when the whole object
    # graph behind it happens to be collected. This function then reads ~15 sheets and
    # can raise on any of a dozen validation failures, and an exception keeps the
    # frame (and so the parser) alive with it. The result was a handle on the input
    # file outliving the call.
    #
    # On POSIX that is invisible: an open file can still be renamed, replaced and
    # unlinked. On Windows it is not. A held handle there means the user cannot save
    # over the file they just opened, cannot open it in Excel while Studio is running,
    # and cannot delete it -- "[WinError 32] the process cannot access the file
    # because it is being used by another process".
    #
    # Reading the bytes up front costs one copy of a file that is tens of kilobytes to
    # a few megabytes -- which the archive reader would fault in anyway, since parsing
    # every sheet touches most of it -- and buys a guarantee no amount of ``close()``
    # discipline can: after this line the process holds NO operating-system handle on
    # ``filepath``, on every return path and every raise path, without the rest of the
    # function having to be arranged around a ``finally``.
    with open(filepath, 'rb') as _workbook_fh:
        xls = pd.ExcelFile(io.BytesIO(_workbook_fh.read()))
    globals_data = {}

    # === STATIC GLOBALS ===
    main_df = xls.parse('main', header=None)

    try:
        template_version = main_df.iloc[4, 3]  # Excel row 5, column D
    except Exception as e:
        raise ValueError(f"Error reading the template version from 'main' tab: {e}")

    # Template version gate. Refuse files NEWER than this build understands, so a
    # newer template can never be silently mis-read by an older install (new
    # columns ignored, options like u='ru' silently zeroed). Shipped models carry
    # versions 8-18; all load through the header-name-driven readers below.
    try:
        _tv = int(float(template_version))
    except (TypeError, ValueError):
        raise ValueError(
            f"Unrecognized template version {template_version!r} in cell D5 of the "
            f"'main' sheet. Expected a number (current version is "
            f"{SUPPORTED_TEMPLATE_VERSION}).")
    if _tv > SUPPORTED_TEMPLATE_VERSION:
        raise ValueError(
            f"This input file is template version {_tv}, but this installation of "
            f"xslope supports versions up to {SUPPORTED_TEMPLATE_VERSION}. "
            "Update xslope to read this file.")

    # v25 inserted '1D element size' at main!D20, so every run option below it moved
    # down one row. One offset carries that through the whole main-sheet block --
    # including the cell names the error messages quote, which must name the row the
    # user is actually looking at.
    _shift = 1 if _tv >= _ELEMENT_SIZE_1D_TEMPLATE_VERSION else 0

    # === MAIN-SHEET GLOBALS (version-gated positional map) ===
    # v18 inserted two declaration cells at the top of the block and shifted the three
    # existing globals down two rows:
    #   v18  D8=Units selector  D9=Time unit  D10=gamma_w  D11=tcrack  D12=crackwater  D13=seismic
    #   <=17            (none)        (none)   D8 =gamma_w  D9 =tcrack  D10=crackwater  D11=seismic
    # gamma_w (v18) is autofill-overridable: a set Units selector fixes it to the
    # canonical value when D10 is blank, but a filled D10 always wins (the
    # seawater/brine override) with a >2% divergence warning surfaced by units_check.
    unit_system = None
    time_unit = None

    def _cell_str(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ''
        s = str(v).strip()
        return '' if s.lower() == 'nan' else s

    try:
        if _tv >= 18:
            unit_system = normalize_unit_system(_cell_str(main_df.iloc[7, 3]) or None)
            time_unit = _cell_str(main_df.iloc[8, 3]) or None
            _gamma_raw = main_df.iloc[9, 3]                     # D10
            _gamma_blank = pd.isna(_gamma_raw) or _cell_str(_gamma_raw) == ''
            if _gamma_blank:
                if unit_system is None:
                    raise ValueError(
                        "The 'main' sheet declares no Units selector (D8) and no unit "
                        "weight of water (D10). Set one or the other: with a Units "
                        "selector the canonical gamma_w is filled automatically; "
                        "otherwise enter it in D10.")
                gamma_water = GAMMA_W[unit_system]              # canonical autofill
            else:
                gamma_water = float(_gamma_raw)                 # D10 always wins
                if unit_system is None:
                    unit_system = infer_system_from_gamma_water(gamma_water)
            tcrack_depth = float(main_df.iloc[10, 3])           # D11
            tcrack_water = float(main_df.iloc[11, 3])           # D12
            k_seismic = float(main_df.iloc[12, 3])              # D13
        else:
            gamma_water = float(main_df.iloc[7, 3])             # D8
            tcrack_depth = float(main_df.iloc[8, 3])            # D9
            tcrack_water = float(main_df.iloc[9, 3])            # D10
            k_seismic = float(main_df.iloc[10, 3])              # D11
            unit_system = infer_system_from_gamma_water(gamma_water)
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Error reading static global values from 'main' tab: {e}")

    # === RUN OPTIONS (v19, main D14:D21) ===
    # Eight optional run settings carried BY THE FILE so a model is self-describing:
    # the analysis it was built for travels with it instead of living only in a
    # dialog. Every one of them is optional and BLANK ALWAYS MEANS UNSPECIFIED --
    # the value stays None and the solver/GUI default is used unchanged. That is
    # what makes a v18 file and a v19 file with an empty block behave identically.
    #   D14 LEM method       D15 number of slices   D16 K0 initial stress (FEM)
    #   D17 Tension SRF      D18 mesh element type  D19 mesh target size
    #   D20 SSRM F min       D21 SSRM F max
    # From v25 the block gains D20 '1D element size' and the SSRM bracket moves to
    # D21/D22; the rows below shift with it (see _shift).
    lem_method = None
    num_slices_opt = None
    k0 = None
    tension_srf_opt = None
    element_type = None
    target_size = None
    element_size_1d = None
    ssrm_f_min = None
    ssrm_f_max = None

    def _opt_num(row_idx, label, integer=False):
        """Optional numeric main-sheet cell at (row_idx, D). Blank -> None."""
        try:
            raw = main_df.iloc[row_idx, 3]
        except IndexError:
            return None
        if _cell_str(raw) == '':
            return None
        try:
            return int(float(raw)) if integer else float(raw)
        except (TypeError, ValueError):
            raise ValueError(
                f"The 'main' sheet cell D{row_idx + 1} ({label}) contains "
                f"{raw!r}, which is not a number. Leave it blank to use the "
                "default.")

    if _tv >= 19:
        # -- D14 LEM method. Validated against the canonical solver names (plus
        # 'all', the run-every-method sweep). Case-insensitive, since the cell is
        # free text once someone types over the dropdown; an unrecognized value
        # RAISES rather than falling back, so a typo can never silently run a
        # different method than the file says.
        _lem_raw = _cell_str(main_df.iloc[13, 3]) if main_df.shape[0] > 13 else ''
        if _lem_raw:
            lem_method = _lem_raw.lower()
            if lem_method not in LEM_METHODS:
                raise ValueError(
                    f"The 'main' sheet declares an unrecognized LEM method "
                    f"{_lem_raw!r} in cell D14. Expected one of: "
                    f"{', '.join(LEM_METHODS)} (or leave it blank).")

        num_slices_opt = _opt_num(14, 'number of slices', integer=True)   # D15
        if num_slices_opt is not None and num_slices_opt < 2:
            raise ValueError(
                f"The 'main' sheet declares {num_slices_opt} slices in cell D15. "
                "At least 2 are required (leave it blank for the default).")

        # -- D16 K0. At-rest lateral earth pressure coefficient for the FEM
        # initial stress state. Blank -> None -> gravity turn-on (the historical
        # initialization), so no existing model moves.
        k0 = _opt_num(15, 'K0 initial stress')
        if k0 is not None and k0 <= 0:
            raise ValueError(
                f"The 'main' sheet declares K0 = {k0} in cell D16. The at-rest "
                "coefficient must be positive (leave it blank for the gravity "
                "turn-on initialization).")

        # -- D17 Tension SRF. YES/NO/blank -> True/False/None. None is NOT False:
        # it means unspecified, and the engine default (True) applies.
        _tsrf_raw = _cell_str(main_df.iloc[16, 3]) if main_df.shape[0] > 16 else ''
        if _tsrf_raw:
            _t = _tsrf_raw.strip().lower()
            if _t in ('yes', 'true', 'y'):
                tension_srf_opt = True
            elif _t in ('no', 'false', 'n'):
                tension_srf_opt = False
            else:
                raise ValueError(
                    f"The 'main' sheet has an unrecognized Tension SRF value "
                    f"{_tsrf_raw!r} in cell D17. Expected YES or NO (or blank).")

        # -- D18/D19 mesh defaults.
        _et_raw = _cell_str(main_df.iloc[17, 3]) if main_df.shape[0] > 17 else ''
        if _et_raw:
            element_type = _et_raw.lower()
            if element_type not in MESH_ELEMENT_TYPES:
                raise ValueError(
                    f"The 'main' sheet declares an unrecognized mesh element type "
                    f"{_et_raw!r} in cell D18. Expected one of: "
                    f"{', '.join(MESH_ELEMENT_TYPES)} (or leave it blank).")
        target_size = _opt_num(18, 'mesh target size')                   # D19
        if target_size is not None and target_size <= 0:
            raise ValueError(
                f"The 'main' sheet declares a mesh target size of {target_size} "
                "in cell D19. It must be positive (leave it blank for the "
                "automatic size).")

        # -- D20 1D element size (v25). The target edge length along the 1D
        # members -- piles and reinforcement lines. Blank means unspecified, and
        # the mesher's own subdivision applies, so a v24 file (which has no such
        # cell at all) and a v25 file with the cell empty are the same model.
        if _tv >= _ELEMENT_SIZE_1D_TEMPLATE_VERSION:
            element_size_1d = _opt_num(19, '1D element size')            # D20
            if element_size_1d is not None and element_size_1d <= 0:
                raise ValueError(
                    f"The 'main' sheet declares a 1D element size of "
                    f"{element_size_1d} in cell D20. It must be positive (leave it "
                    "blank to subdivide the 1D members automatically).")

        # -- SSRM bracket (D20/D21 through v24, D21/D22 from v25).
        ssrm_f_min = _opt_num(19 + _shift, 'SSRM F min')
        ssrm_f_max = _opt_num(20 + _shift, 'SSRM F max')
        if (ssrm_f_min is not None and ssrm_f_max is not None
                and ssrm_f_min >= ssrm_f_max):
            raise ValueError(
                f"The 'main' sheet declares SSRM F min = {ssrm_f_min} "
                f"(D{20 + _shift}) >= F max = {ssrm_f_max} (D{21 + _shift}). "
                f"The bracket must be increasing.")

    # === SIDE BOUNDARY CONDITION (v21, main D22) ===
    # How the FEM restrains the left and right truncation boundaries:
    #   'rollers' (default, and every pre-v21 file)  u = 0, v free
    #   'fixed'                                      u = v = 0
    # Rollers are the historical hardwired behavior and stay the default; 'fixed'
    # exists for vendor parity (RS2 fully fixes side boundaries). Blank means
    # unspecified and the engine default (rollers) applies, so a v21 file with an
    # untouched cell is the same model as its v20 original.
    side_bc = None
    if _tv >= 21:
        _sbc_row = 21 + _shift
        _sbc_raw = (_cell_str(main_df.iloc[_sbc_row, 3])
                    if main_df.shape[0] > _sbc_row else '')
        if _sbc_raw:
            side_bc = _sbc_raw.lower()
            if side_bc not in SIDE_BC_OPTIONS:
                raise ValueError(
                    f"The 'main' sheet declares a Side BC of {_sbc_raw!r} in cell "
                    f"D{_sbc_row + 1}. Expected one of: {', '.join(SIDE_BC_OPTIONS)} "
                    "(or leave it blank for the default, rollers).")

    # === WATER LOADS (v22, main D23) ===
    # Who supplies the weight of standing water:
    #   'auto'    the engine derives the ponded-water surface load from the model's
    #             own water definition (piezometric line, or the seepage head
    #             boundaries where a seepage analysis is defined) at solve time, and
    #             the dloads sheets carry NON-WATER loads only
    #   'manual'  the load is whatever the user typed on the dloads sheets
    # Unlike the run options above, blank does NOT stay unspecified: the answer is
    # resolved here, by template version, because the two versions mean opposite
    # things. A v22 file means 'auto' (the template's own default); a v21-or-earlier
    # file means 'manual', and that is a correctness requirement rather than a
    # preference -- those files carry hand-entered water loads, and deriving a load
    # under them would count the reservoir twice.
    water_loads = 'manual'
    if _tv >= 22:
        _wl_row = 22 + _shift
        _wl_raw = (_cell_str(main_df.iloc[_wl_row, 3])
                   if main_df.shape[0] > _wl_row else '')
        if _wl_raw:
            water_loads = _wl_raw.lower()
            if water_loads not in WATER_LOAD_OPTIONS:
                raise ValueError(
                    f"The 'main' sheet declares a Water loads mode of {_wl_raw!r} in "
                    f"cell D{_wl_row + 1}. Expected one of: "
                    f"{', '.join(WATER_LOAD_OPTIONS)} (or leave it blank for the "
                    "default, auto).")
        else:
            water_loads = 'auto'

    # === SURFACE FAMILY (v22 amendment, main D24) ===
    # Which failure-surface family this model means, for the rare deck that defines
    # BOTH a circular surface (circles sheet) and a non-circular one (non-circ
    # sheet). Nothing else in the file can say: with both present the circles simply
    # win, and the non-circular surface is ignored with no message.
    #
    # SHIPS BLANK, and blank is the normal state -- it means "resolve automatically",
    # which is the only family present on a single-family deck and, on a both-family
    # deck, the question the Studio run dialog asks once and then writes here. The
    # cell is normalized to the vocabulary the rest of the package speaks
    # ('circular' / 'noncircular'), so a reader never has to know that the sheet
    # spells the second one with a hyphen.
    surface_family = None
    if _tv >= 22:
        _sf_row = 23 + _shift
        _sf_raw = (_cell_str(main_df.iloc[_sf_row, 3])
                   if main_df.shape[0] > _sf_row else '')
        if _sf_raw:
            _sf = _sf_raw.strip().lower().replace('_', '-').replace(' ', '')
            if _sf not in SURFACE_FAMILY_OPTIONS + ('noncircular',):
                raise ValueError(
                    f"The 'main' sheet declares a surface family of {_sf_raw!r} in "
                    f"cell D{_sf_row + 1}. Expected one of: "
                    f"{', '.join(SURFACE_FAMILY_OPTIONS)} (or leave it blank to use "
                    "whichever family the model defines).")
            surface_family = 'noncircular' if _sf.startswith('non') else 'circular'

    # === PROFILE LINES ===
    profile_df = xls.parse('profile', header=None)

    max_depth = float(profile_df.iloc[1, 1])  # Excel B2 = row 1, column 1

    profile_lines = []
    
    # New format: single data block, profile lines arranged horizontally
    # First profile line: columns A:B, second: D:E, third: G:H, etc.
    # Header row is row 4 (index 3), mat_id is in B5 (row 4, column 1)
    #
    # v21 inserts an optional 'Size:' row (Excel row 7) between the material-name
    # echo and the coordinates, so the vertices start two rows lower. A profile line
    # builds one material zone; its Size is the local target element size inside
    # THAT zone. Pre-v21 files keep the old rows forever.
    header_row = 3  # Excel row 4 (0-indexed)
    mat_id_row = 4  # Excel row 5 (0-indexed)
    if _tv >= 21:
        size_row = 6           # Excel row 7
        coords_start_row = 8   # Excel row 9
    else:
        size_row = None
        coords_start_row = 7   # Excel row 8

    col = 0  # Start with column A (index 0)
    while col < profile_df.shape[1]:
        x_col = col
        y_col = col + 1
        
        # Check if header row is empty (stop reading if empty)
        try:
            header_val = str(profile_df.iloc[header_row, x_col]).strip()
            if not header_val or header_val.lower() == 'nan':
                break  # No more profile lines
        except:
            break  # No more profile lines
        
        # Read mat_id from B5 (row 4, column 1) for this profile line
        # Convert from 1-based to 0-based for internal use
        try:
            mat_id_val = profile_df.iloc[mat_id_row, y_col]
            if pd.isna(mat_id_val):
                mat_id = None
            else:
                # Convert to integer and subtract 1 to make it 0-based.
                # Range validation happens later, once materials are parsed.
                mat_id = int(float(mat_id_val)) - 1
                if mat_id < 0:
                    mat_id = None  # Invalid mat_id
        except (ValueError, TypeError):
            mat_id = None
        
        # Read XY coordinates starting from row 7, stop at first empty row
        coords = []
        row = coords_start_row
        while row < profile_df.shape[0]:
            try:
                x_val = profile_df.iloc[row, x_col]
                y_val = profile_df.iloc[row, y_col]
                
                # Stop at first empty row (both x and y are empty)
                if pd.isna(x_val) and pd.isna(y_val):
                    break
                
                # If at least one coordinate is present, try to convert
                if pd.notna(x_val) and pd.notna(y_val):
                    coords.append((float(x_val), float(y_val)))
            except:
                break
            row += 1
        
        # Validate that we have at least 2 points
        if len(coords) == 1:
            raise ValueError(f"Each profile line must contain at least two points. Profile line starting at column {chr(65 + col)} has only one point.")
        
        if len(coords) >= 2:
            # Store as dict with coords, mat_id and the optional v21 local mesh size
            _size = (None if size_row is None else _opt_size_cell(
                profile_df, size_row, y_col,
                f"Profile line starting at column {chr(65 + col)}"))
            profile_lines.append({
                'coords': coords,
                'mat_id': mat_id,
                'size': _size,
            })

        # Move to next profile line (skip 3 columns: A->D, D->G, etc.)
        col += 3

    # Ground surface, domain polygon, and tensile-crack line are built from the
    # unified polygon representation after materials are parsed — see below.

    # === MATERIALS (Optimized Parsing) ===
    # Locate the header row rather than assuming row 8 -- it has moved before (pre-v8
    # templates put it on row 3) and a legend row above the table moves it again.
    _mat_raw = xls.parse('mat', header=None)
    _mat_hdr = _find_mat_header_row(
        lambda r, c: (_mat_raw.iloc[r - 1, c - 1]
                      if r <= _mat_raw.shape[0] and c <= _mat_raw.shape[1] else None))
    mat_df = _mat_raw.iloc[_mat_hdr:].reset_index(drop=True)
    # Header names are matched underscore-insensitively: the master template's
    # styled headers concatenate to 'pow_a'/'vg_a' while older sheets carry
    # 'powa'/'vga' - both must resolve to the same column.
    mat_df.columns = [('' if pd.isna(v) else str(v).strip().replace('_', ''))
                      for v in _mat_raw.iloc[_mat_hdr - 1]]
    materials = []
    # Names of materials still carrying the RETIRED v19 'ssr_zone' flag, collected
    # in the row loop and reported once, after it (see the v20 warning below).
    retired_ssr_zone_flags = []

    def _num(x):
        v = pd.to_numeric(x, errors="coerce")
        return float(v) if pd.notna(v) else 0.0

    def _pick(row, *names):
        """First of ``names`` that is actually a COLUMN on this sheet.

        v14 renamed the two unsaturated-curve parameters from the van-Genuchten-
        specific 'vga'/'vgn' to the law-agnostic 'a'/'n' (they now serve van
        Genuchten and Gardner alike), and renamed Poisson's ratio from 'n' to
        'nu' — which had collided with the new 'n'. Order matters: a v14 sheet
        has BOTH 'n' (the curve exponent) and 'nu', so nu must prefer 'nu' and
        only fall back to 'n' on a pre-v14 sheet that has no 'nu' column."""
        for nm in names:
            if nm in row.index:
                return row.get(nm, 0)
        return 0

    def _choice(x, default):
        """Normalize a free-text option cell. An empty cell reaches here as the
        float NaN, which ``str()`` renders as 'nan' -- treat both as unset."""
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return default
        s = str(x).strip().lower()
        return default if s in ('', 'nan') else s

    # Read materials row by row until we encounter an empty material name (Column B)
    # Data starts at Excel row 9 (0-indexed row 0 after header=7)
    for i in range(len(mat_df)):
        row = mat_df.iloc[i]
        
        # Check if material name (Column B) is empty - stop reading if empty
        material_name = row.get('name', '')
        if pd.isna(material_name) or str(material_name).strip() == '':
            break  # Stop reading when we encounter an empty material name
        
        # For seep workflows, 'g' (unit weight) and shear strength properties are not required.
        # A material row is considered "missing" only if EVERY property column after
        # 'name' is empty. (Expressed position-free: a column insert can never
        # silently narrow this check the way the old hardcoded C:X window could.)
        props_empty = row.iloc[2:].isna().all() if mat_df.shape[1] > 2 else True
        if props_empty:
            # Excel row number: first data row sits just below the located header
            excel_row = _mat_hdr + 1 + i
            raise ValueError(
                "CRITICAL ERROR: Material row has empty property fields. "
                f"Material '{material_name}' (Excel row {excel_row}) has no property "
                "values after the name column."
            )

        # Unsaturated relative-permeability model (template v11+): 'lf' (linear
        # front, kr0/h0 apply), 'vg' (van Genuchten) or 'gard' (Gardner, v14+).
        # The last two share the a/n parameter pair. Older templates lack the
        # column -> default 'lf' (current behavior). Read by header name, so
        # column position is version-independent.
        unsat_raw = row.get('unsat', 'lf')
        unsat_val = str(unsat_raw).strip().lower() if (
            pd.notna(unsat_raw) and str(unsat_raw).strip().lower() not in ('', 'nan')
        ) else 'lf'
        if unsat_val not in ('lf', 'vg', 'gard'):
            unsat_val = 'lf'

        # Excel row number: first data row sits just below the located header
        excel_row = _mat_hdr + 1 + i

        # Pore pressure option. An unrecognized value used to fall through to
        # u = 0 in slice.py, silently deleting pore pressure and inflating FS.
        u_val = _choice(row.get('u'), 'none')
        if u_val not in ('none', 'piezo', 'seep', 'ru'):
            raise ValueError(
                f"Material '{material_name}' (mat sheet, Excel row {excel_row}) has an "
                f"unrecognized pore pressure option u='{u_val}'. "
                "Expected one of: none, piezo, seep, ru (or leave blank for none)."
            )

        # Strength model. Blank is allowed -- seep-only material rows carry no
        # strength -- but slice.py raises if a blank one reaches a failure surface.
        # 'elastic' (v16): infinite strength / cannot fail -- the FEM holds it out
        # of plasticity entirely and the LEM treats it as impenetrable.
        option_val = _choice(row.get('option'), '')
        if option_val not in ('', 'mc', 'cp', 'pow', 'hb', 'elastic'):
            raise ValueError(
                f"Material '{material_name}' (mat sheet, Excel row {excel_row}) has an "
                f"unrecognized strength option option='{option_val}'. "
                "Expected one of: mc, cp, pow, hb, elastic."
            )

        # v12 columns, read by header name. Older templates lack them entirely:
        # row.get() returns None and the defaults preserve pre-v12 behavior exactly.
        gamma_val = _num(row.get("g", 0))
        _gsat_num = pd.to_numeric(row.get('gsat'), errors='coerce')
        gamma_sat_val = float(_gsat_num) if pd.notna(_gsat_num) else None
        if gamma_sat_val is not None and gamma_sat_val < gamma_val:
            raise ValueError(
                f"Material '{material_name}' (mat sheet, Excel row {excel_row}) has "
                f"gsat = {gamma_sat_val} < g = {gamma_val}. The saturated unit weight "
                "cannot be less than the moist unit weight; leave gsat blank to use "
                "g throughout.")

        ru_val = _num(row.get('ru', 0))
        if u_val == 'ru' and ru_val < 0:
            raise ValueError(
                f"Material '{material_name}' (mat sheet, Excel row {excel_row}) selects "
                f"u='ru' but has a negative pore pressure ratio ru = {ru_val}.")

        pow_a_val = _num(row.get('powa', 0))
        pow_b_val = _num(row.get('powb', 0))
        pow_c_val = _num(row.get('powc', 0))
        pow_d_val = _num(row.get('powd', 0))
        if option_val == 'pow' and (pow_a_val <= 0 or pow_b_val <= 0):
            raise ValueError(
                f"Material '{material_name}' (mat sheet, Excel row {excel_row}) selects "
                f"option='pow' but pow_a ({pow_a_val}) and pow_b ({pow_b_val}) must both "
                "be positive for the envelope tau = pow_a*(sigma_n + pow_d)^pow_b + pow_c.")

        # Generalized Hoek-Brown (v14). mb/s/a are DERIVED from GSI/mi/D at use
        # time (xslope.hoekbrown), so only the four field-observable inputs are
        # entered. GSI is defined on (0, 100] and D on [0, 1]; sigma_ci and mi
        # must be positive or the envelope collapses.
        hb_sci_val = _num(row.get('hbsci', 0))
        hb_gsi_val = _num(row.get('hbgsi', 0))
        hb_mi_val = _num(row.get('hbmi', 0))
        hb_d_val = _num(row.get('hbd', 0))
        if option_val == 'hb':
            if hb_sci_val <= 0 or hb_mi_val <= 0:
                raise ValueError(
                    f"Material '{material_name}' (mat sheet, Excel row {excel_row}) selects "
                    f"option='hb' but hb_sci ({hb_sci_val}) and hb_mi ({hb_mi_val}) must "
                    "both be positive (intact strength and the intact Hoek-Brown constant).")
            if not (0 < hb_gsi_val <= 100):
                raise ValueError(
                    f"Material '{material_name}' (mat sheet, Excel row {excel_row}) selects "
                    f"option='hb' but has hb_gsi = {hb_gsi_val}. The Geological Strength "
                    "Index must lie in (0, 100].")
            if not (0 <= hb_d_val <= 1):
                raise ValueError(
                    f"Material '{material_name}' (mat sheet, Excel row {excel_row}) selects "
                    f"option='hb' but has hb_d = {hb_d_val}. The disturbance factor must "
                    "lie in [0, 1] (0 = undisturbed, 1 = heavily blast-damaged).")

        # Tensile-strength cutoff (v16). Rankine cap on the major principal stress,
        # in stress units. BLANK -> None (no cutoff, unbounded tension -- exactly the
        # pre-v16 behavior); 0 -> the soil carries no tension. FEM only; the LEM
        # ignores it (a tension crack is modeled separately). Read by header name
        # ('t_cut' normalizes to 'tcut'), so pre-v16 sheets without the column load
        # as None.
        _tcut_num = pd.to_numeric(row.get('tcut'), errors='coerce')
        t_cut_val = float(_tcut_num) if pd.notna(_tcut_num) else None

        # Matric-suction strength (v17). Two LEM-only columns for the Fredlund
        # extended Mohr-Coulomb apparent cohesion (xslope.slice generate_slices):
        #   phi_b -- the unsaturated friction angle phi^b (degrees). BLANK -> None:
        #     no suction strength credited, exactly the pre-v17 behavior (the default).
        #   s_cap -- the maximum credited suction (stress units), a cap on the base
        #     suction s before it becomes apparent cohesion. BLANK -> None: uncapped.
        # Read by header name ('phi_b' normalizes to 'phib', 's_cap' to 'scap'), so a
        # pre-v17 sheet without the columns loads as None for both. Caution: with
        # u=piezo the hydrostatic suction above the line is unbounded, so s_cap is
        # essential there; with u=seep the FE field self-bounds and s_cap is a backstop.
        _phib_num = pd.to_numeric(row.get('phib'), errors='coerce')
        phi_b_val = float(_phib_num) if pd.notna(_phib_num) else None
        _scap_num = pd.to_numeric(row.get('scap'), errors='coerce')
        s_cap_val = float(_scap_num) if pd.notna(_scap_num) else None

        # RETIRED mat-sheet 'ssr_zone' flag (v19 column O, deleted in v20). SSR zones
        # are polygon-sheet overlay rows now, not a material property — see
        # SSR_ZONE_SENTINELS. A v19 file that still carries the column is read, but
        # the flag is IGNORED and the file is named loudly below, because silently
        # honoring a per-material search area would give an answer the v20 model does
        # not describe.
        _ssrz_raw = row.get('ssrzone')
        _ssrz = '' if (_ssrz_raw is None or (isinstance(_ssrz_raw, float)
                                             and pd.isna(_ssrz_raw))) else str(_ssrz_raw).strip().lower()
        if _ssrz not in ('', 'nan', 'no', 'false', 'n', '0'):
            retired_ssr_zone_flags.append(str(material_name).strip())

        # Transient storage parameters (v18, seepage block AO/AP). Ss = specific
        # storage [1/len], Sy = specific yield [-]. Both optional here (header-keyed,
        # so a pre-v18 sheet without the columns reads None) and required only when a
        # 'tseep' sheet is present -- validated once, after the sheet is parsed. BLANK
        # stays None (never a silent 0, which would zero the storage term).
        _ss_num = pd.to_numeric(row.get('Ss'), errors='coerce')
        ss_val = float(_ss_num) if pd.notna(_ss_num) else None
        _sy_num = pd.to_numeric(row.get('Sy'), errors='coerce')
        sy_val = float(_sy_num) if pd.notna(_sy_num) else None

        # --- load-time validation warnings (v16) ---
        if option_val == 'elastic':
            # (b) elastic materials ignore every strength input (they cannot fail);
            # (c) t_cut is meaningless on an elastic material.
            _strength_present = any(_num(row.get(h, 0)) for h in (
                'c', 'f', 'c/p', 'powa', 'powb', 'powc', 'powd',
                'hbsci', 'hbgsi', 'hbmi', 'hbd', 'psi'))
            if _strength_present:
                print(f"WARNING: Material '{material_name}' (mat sheet, Excel row "
                      f"{excel_row}) is elastic (cannot fail); its strength values are "
                      "ignored.")
            if t_cut_val is not None:
                print(f"WARNING: Material '{material_name}' (mat sheet, Excel row "
                      f"{excel_row}) is elastic; its t_cut is ignored.")
        elif option_val == 'mc' and t_cut_val is not None:
            # (a) an mc t_cut at/above the cone apex c/tan(phi) never binds -- the
            # shear envelope caps tension there first, so the Rankine cap is inert.
            _c = _num(row.get('c', 0))
            _phi = _num(row.get('f', 0))
            if _phi > 0 and _c > 0:
                apex = _c / np.tan(np.radians(_phi))
                if t_cut_val >= apex - 1e-9:
                    print(f"WARNING: Material '{material_name}' (mat sheet, Excel row "
                          f"{excel_row}) has t_cut = {t_cut_val:g} >= the Mohr-Coulomb "
                          f"apex c/tan(phi) = {apex:g}; the cutoff never binds (inert).")

        materials.append({
            "name": str(material_name).strip(),
            "gamma": gamma_val,
            "gamma_sat": gamma_sat_val,
            "option": option_val,
            "c": _num(row.get('c', 0)),
            "phi": _num(row.get('f', 0)),
            "cp": _num(row.get('c/p', 0)),
            "r_elev": _num(row.get('r-elev', 0)),
            "d": _num(row.get('d', 0)) if pd.notna(row.get('d')) else 0,
            "psi": _num(row.get('psi', 0)) if pd.notna(row.get('psi')) else 0,
            # v16: tensile-strength cutoff. None = blank (no cutoff); FEM-only.
            "t_cut": t_cut_val,
            # v17: matric-suction strength (LEM only, Fredlund extended MC).
            # phi_b None = no suction strength (default); s_cap None = uncapped.
            "phi_b": phi_b_val,
            "s_cap": s_cap_val,
            # v18: transient-seepage storage. None = blank (required only with a tseep
            # sheet); never defaulted to 0 (a silent zero would drop the storage term).
            "Ss": ss_val,
            "Sy": sy_val,
            "pow_a": pow_a_val,
            "pow_b": pow_b_val,
            "pow_c": pow_c_val,
            "pow_d": pow_d_val,
            "u": u_val,
            "ru": ru_val,
            "sigma_gamma": _num(row.get('s(g)', 0)),
            "sigma_c": _num(row.get('s(c)', 0)),
            "sigma_phi": _num(row.get('s(f)', 0)),
            "sigma_cp": _num(row.get('s(c/p)', 0)),
            "sigma_d": _num(row.get('s(d)', 0)),
            "sigma_psi": _num(row.get('s(psi)', 0)),
            "k1": _num(row.get('k1', 0)),
            "k2": _num(row.get('k2', 0)),
            "alpha": _num(row.get('alpha', 0)),
            "unsat": unsat_val,
            "kr0" : _num(row.get('kr0', 0)),
            "h0" : _num(row.get('h0', 0)),
            # v14: 'a'/'n' (law-agnostic); pre-v14: 'vga'/'vgn'. Same two slots
            # feed van Genuchten and Gardner, exactly as SEEP2D's uspar(1..2) do.
            # The OLD names are checked first because they are unambiguous: on a
            # pre-v14 sheet the bare 'n' column is POISSON'S RATIO, not the curve
            # exponent, so preferring 'n' there would silently read nu into vg_n.
            "vg_a": _num(_pick(row, 'vga', 'a')),
            "vg_n": _num(_pick(row, 'vgn', 'n')),
            "E": _num(row.get('E', 0)),
            # v14 renamed this to 'nu'; pre-v14 sheets call it 'n'
            "nu": _num(_pick(row, 'nu', 'n')),
            # Generalized Hoek-Brown (v14). mb/s/a are DERIVED from GSI/mi/D by
            # the Hoek-Brown 2002 relations, not entered — nobody has them to hand.
            "hb_sci": hb_sci_val,
            "hb_gsi": hb_gsi_val,
            "hb_mi": hb_mi_val,
            "hb_d": hb_d_val,
        })

    if retired_ssr_zone_flags:
        print(f"WARNING: the mat sheet flags {retired_ssr_zone_flags} with the RETIRED "
              "'ssr_zone' column (template version 19). That flag is IGNORED. In "
              "template version 20 an SSR zone is a POLYGON drawn on the 'polygon' "
              "sheet with a sentinel Mat ID: -1 = 'SSR reduce' (strength reduction "
              "applies only inside), -2 = 'SSR hold' (full strength inside, can still "
              "yield), -3 = 'SSR elastic' (linear elastic inside, cannot yield). "
              "Redraw the zone there to restore the constraint.")

    # === UNIFIED POLYGON REPRESENTATION ===
    # Geometry is always represented internally as material-zone polygons. If the
    # 'polygon' sheet is populated it is used directly; otherwise the profile lines
    # are converted to polygons via build_polygons(). All downstream code (slicing,
    # search) works from polygons.
    #
    # SSR zone rows come back SEPARATELY (ssr_zones) and never enter `polygons`:
    # they are analysis overlays, so they must not be meshed, must not become
    # material regions, must not generate slices and must not shape the domain.
    # Keeping them out of `polygons` here is what makes that true everywhere at
    # once — every downstream consumer reads `polygons`.
    polygons_from_sheet, ssr_zones, refine_zones = _parse_polygon_sheet(
        xls, materials, template_version=_tv)

    if polygons_from_sheet:
        if profile_lines:
            raise ValueError(
                "Both the 'profile' and 'polygon' sheets contain data. Use one "
                "geometry method, not both.")
        _validate_polygons_no_overlap(polygons_from_sheet)
        polygons = polygons_from_sheet
        # max_depth is a profile-sheet concept; it has no meaning for polygon input.
        max_depth = None
    elif profile_lines:
        # Convert profile lines -> polygons. max_depth is used ONLY here, as the
        # bottom boundary for build_polygons (mat_id is 0-based in both).
        # build_polygons emits one zone per profile line, in line order, so the
        # line's own local mesh Size rides along to the zone it built.
        polygons = [
            {'polygon': Polygon(p['coords']), 'mat_id': p['mat_id'],
             'size': p.get('size')}
            for p in build_polygons(slope_data={'profile_lines': profile_lines,
                                                'max_depth': max_depth})
        ]
    else:
        polygons = []

    # Derive the ground surface and domain polygon from the polygons. The domain
    # polygon (not max_depth) defines the bottom/side boundaries for all downstream
    # use (slice generation, search containment).
    domain_polygon = None
    if polygons:
        ground_surface, domain_polygon = build_ground_surface_from_polygons(polygons)
    else:
        ground_surface = LineString([])

    # === BUILD TENSILE CRACK LINE ===
    tcrack_surface = None
    if tcrack_depth > 0 and not ground_surface.is_empty:
        tcrack_surface = LineString(
            [(x, y - tcrack_depth) for (x, y) in ground_surface.coords])

    # === MESH AND SEEPAGE ANALYSIS FILES ===
    # The suffixes come from xslope.package, which is where the whole companion
    # convention is defined: the same names the packager uses to decide which of two
    # workbooks in one folder a results file belongs to. One definition, so the
    # loader and the packager cannot drift into disagreeing about a project's set.
    base, _ = os.path.splitext(filepath)
    mesh_filename = f"{base}{MESH_SIDECAR}"

    # Load mesh if it exists (used by both seep and fem workflows)
    mesh = None
    if os.path.exists(mesh_filename):
        try:
            mesh = import_mesh_from_json(mesh_filename)
        except Exception as e:
            print(f"WARNING: Error reading mesh file: {e}. Continuing without mesh.")

    # Load seepage solution files if any materials use seep pore pressure
    has_seep_materials = any(material["u"] == "seep" for material in materials)
    seep_u = None
    seep_u2 = None

    if has_seep_materials:
        try:
            solution1_filename = f"{base}{SEEP_SIDECARS[0]}"
            solution2_filename = f"{base}{SEEP_SIDECARS[1]}"

            if mesh is not None and os.path.exists(solution1_filename):
                # comment="#" skips BOTH the trailing "# Total Flowrate:" footer and any
                # leading "# units:" provenance header, leaving exactly the node rows.
                # (Replaces the old read-all-then-drop-last-row trick, which assumed the
                # footer was the only extra line and would mis-drop a real node once a
                # second comment line existed.)
                solution1_df = pd.read_csv(solution1_filename, comment="#")
                seep_u = solution1_df["u"].to_numpy()

                if os.path.exists(solution2_filename):
                    solution2_df = pd.read_csv(solution2_filename, comment="#")
                    seep_u2 = solution2_df["u"].to_numpy()

        except Exception as e:
            print(f"WARNING: Error reading seepage files: {e}. Continuing without seep data.")

    # === PIEZOMETRIC LINE ===
    piezo_df = xls.parse('piezo', header=None)
    piezo_line = []
    piezo_line2 = []

    # Locate the x/y header row by content so both layouts load: v13 adds a
    # 'Type:' row (piezo | phreatic) above the headers, shifting data down one
    # row. The Type value sits one row above the header in the y column.
    _hdr = 2  # v12 default (headers in Excel row 3)
    for _r in range(min(8, piezo_df.shape[0])):
        if str(piezo_df.iloc[_r, 0]).strip().lower() == 'x':
            _hdr = _r
            break
    def _line_type(col):
        if _hdr >= 1:
            v = piezo_df.iloc[_hdr - 1, col]
            if pd.notna(v) and str(v).strip().lower() == 'phreatic':
                return True
            if pd.notna(v) and str(v).strip().lower() not in ('', 'piezo', 'type:'):
                raise ValueError(
                    f"Unrecognized piezometric line Type {v!r} on the 'piezo' "
                    f"sheet. Expected 'piezo' (static head, default) or "
                    f"'phreatic' (cos^2 inclination correction).")
        return False
    piezo_phreatic = _line_type(1)
    piezo_phreatic2 = _line_type(4)

    start_row = _hdr + 1
    x_col = 0  # Column A
    y_col = 1  # Column B
    
    row = start_row
    while row < piezo_df.shape[0]:
        try:
            x_val = piezo_df.iloc[row, x_col]
            y_val = piezo_df.iloc[row, y_col]
            
            # Stop at first empty row (both x and y are empty)
            if pd.isna(x_val) and pd.isna(y_val):
                break
            
            # If at least one coordinate is present, try to convert
            if pd.notna(x_val) and pd.notna(y_val):
                piezo_line.append((float(x_val), float(y_val)))
        except:
            break
        row += 1
    
    # Validate first piezometric line
    if len(piezo_line) == 1:
        raise ValueError("First piezometric line must contain at least two points.")
    
    # Read second piezometric line (columns D:E, starting at row 4, Excel row 4 = index 3)
    # Keep reading until we encounter an empty row
    x_col2 = 3  # Column D
    y_col2 = 4  # Column E
    
    row = start_row
    while row < piezo_df.shape[0]:
        try:
            x_val = piezo_df.iloc[row, x_col2]
            y_val = piezo_df.iloc[row, y_col2]
            
            # Stop at first empty row (both x and y are empty)
            if pd.isna(x_val) and pd.isna(y_val):
                break
            
            # If at least one coordinate is present, try to convert
            if pd.notna(x_val) and pd.notna(y_val):
                piezo_line2.append((float(x_val), float(y_val)))
        except:
            break
        row += 1
    
    # Validate second piezometric line (only if it has data)
    if len(piezo_line2) == 1:
        raise ValueError("Second piezometric line must contain at least two points if provided.")

    # === DISTRIBUTED LOADS ===
    # Both sheets ('dloads' and the rapid-drawdown stage-2 'dloads (2)') have the
    # identical layout and are read by the same routine, so the two can never drift.
    #
    # v21 adds a Direction cell per block (the block's N column):
    # blank/'normal' keeps the historical behavior — the load acts PERPENDICULAR to
    # the loaded surface — and 'vertical' resolves the same resultant straight down
    # (a gravity surcharge: dead weight with no horizontal component). Pre-v21 sheets
    # have no such cell and every block reads 'normal', so an existing file is
    # unchanged.
    #
    # Row anchors by version (Excel rows; the whole block moved, nothing changed shape):
    #   <=v20   block header 2, X/Y/N 3, data 4
    #   v21     block header 2, Direction 3, X/Y/N 4, data 5
    #   v22+    note 2, block header 4, Direction 5, X/Y/N 6, data 7
    _DLOAD_DIRECTIONS = ('normal', 'vertical')
    if _tv >= 22:
        start_row, _dir_row = 6, 4      # Excel data row 7, Direction row 5
    elif _tv >= 21:
        start_row, _dir_row = 4, 2      # Excel data row 5, Direction row 3
    else:
        start_row, _dir_row = 3, None   # Excel data row 4, no Direction cell

    def _parse_dload_sheet(df, sheet_name):
        """(lines, directions) for one dloads sheet. Blocks are 4 columns apart
        starting at column B; each carries X, Y, Normal."""
        lines = []
        directions = []
        col = 1  # Column B
        while col < df.shape[1]:
            x_col, y_col, normal_col = col, col + 1, col + 2

            # Check if dataframe has enough rows before accessing start_row
            if df.shape[0] <= start_row:
                break  # Not enough rows, stop reading

            # Check if this distributed load block is empty (first row's X)
            if pd.isna(df.iloc[start_row, x_col]):
                break  # Stop reading when we encounter an empty distributed load

            # Read points for this distributed load, keep reading down until empty row
            block_points = []
            row = start_row
            while row < df.shape[0]:
                try:
                    x_val = df.iloc[row, x_col]
                    y_val = df.iloc[row, y_col]
                    normal_val = df.iloc[row, normal_col]

                    # Stop at first empty row (all three values are empty)
                    if pd.isna(x_val) and pd.isna(y_val) and pd.isna(normal_val):
                        break

                    # If at least coordinates are present, try to convert
                    if pd.notna(x_val) and pd.notna(y_val):
                        normal = float(normal_val) if pd.notna(normal_val) else 0.0
                        block_points.append({
                            "X": float(x_val),
                            "Y": float(y_val),
                            "Normal": normal
                        })
                except:
                    break
                row += 1

            # Validate that we have at least 2 points
            if len(block_points) == 1:
                raise ValueError(
                    f"Each distributed load must contain at least two points. "
                    f"Distributed load starting at column {chr(65 + col)} of the "
                    f"'{sheet_name}' sheet has only one point.")

            if len(block_points) >= 2:
                lines.append(block_points)
                directions.append(_dload_direction(df, normal_col, col, sheet_name))

            # Next distributed load (3 columns for the dload + 1 empty column)
            col += 4
        return lines, directions

    def _dload_direction(df, dir_col, block_col, sheet_name):
        """The block's Direction word. Blank (and every pre-v21 sheet) -> 'normal'.
        An unrecognized word is a loud error: silently falling back to 'normal' would
        apply a dead-weight surcharge as a surface-normal thrust, which on an inclined
        crest is a horizontal force the user never asked for."""
        if (_dir_row is None or _dir_row >= df.shape[0]
                or dir_col >= df.shape[1]):
            return 'normal'
        raw = df.iloc[_dir_row, dir_col]
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            return 'normal'
        word = str(raw).strip().lower()
        if word in ('', 'nan'):
            return 'normal'
        if word not in _DLOAD_DIRECTIONS:
            raise ValueError(
                f"Distributed load starting at column {chr(65 + block_col)} of the "
                f"'{sheet_name}' sheet declares Direction {str(raw).strip()!r}. "
                f"Expected one of: {', '.join(_DLOAD_DIRECTIONS)} (or leave it blank "
                f"for 'normal', a load perpendicular to the loaded surface).")
        return word

    dloads, dload_dirs = _parse_dload_sheet(xls.parse('dloads', header=None), 'dloads')

    dloads2, dload2_dirs = [], []
    try:
        dloads2, dload2_dirs = _parse_dload_sheet(
            xls.parse('dloads (2)', header=None), 'dloads (2)')
    except (ValueError, KeyError) as _e:
        # A missing "dloads (2)" tab just leaves the second set empty. A Direction /
        # point-count problem ON the sheet is a real error and must not be swallowed.
        if 'dloads (2)' not in xls.sheet_names:
            dloads2, dload2_dirs = [], []
        else:
            raise

    # Orient every distributed-load line left-to-right (increasing X).
    #
    # Nothing stops a user (or an importer transcribing a vendor file) from
    # tracing a load line right-to-left: a downstream reservoir followed from
    # the far boundary back toward the toe reads perfectly well on the sheet.
    # Downstream consumers, however, assume increasing X. slice.py builds the
    # load intensity with np.interp, which requires an ascending x array and
    # silently returns nonsense rather than raising when it is not, so a
    # right-to-left line would price the wrong intensity into every slice.
    # Reversing the point list carries each (X, Y, Normal) triple intact, so
    # the line is the same geometry traced the other way, not a re-shaped load.
    #
    # Only lines that are MONOTONE in X are touched. Equal X on consecutive
    # points is a vertical face and stays monotone (a pool line usually starts
    # with one). A line whose X actually turns around — an overhang, or a face
    # that leans back and then forward — has no increasing-X form at all, so it
    # is left exactly as authored rather than mangled by a sort; the FEM path
    # handles those correctly on its own geometry (fem.build_fem_data orients
    # each loaded edge from the element that owns it), and the LEM path cannot
    # represent them regardless of point order.
    def _orient_dload_lines(lines):
        for line in lines:
            xs = [pt["X"] for pt in line]
            non_increasing = all(b <= a for a, b in zip(xs, xs[1:]))
            if non_increasing and xs[-1] < xs[0]:
                line.reverse()

    _orient_dload_lines(dloads)
    _orient_dload_lines(dloads2)

    # === CIRCLES ===

    # Read the first 3 rows to get the max depth
    raw_df = xls.parse('circles', header=None)  # No header, get full sheet

    # Read the circles data starting from row 2 (index 1)
    circles_df = xls.parse('circles', header=1)
    raw = circles_df.dropna(subset=['Xo', 'Yo'], how='any')
    circles = []
    for _, row in raw.iterrows():
        Xo = row['Xo']
        Yo = row['Yo']
        Option = row.get('Option', None)
        Depth = row.get('Depth', None)
        Xi = row.get('Xi', None)
        Yi = row.get('Yi', None)
        R = row.get('R', None)
        # For each circle, fill in the radius and depth values depending on the circle option
        if Option == 'Depth':
            R = Yo - Depth
        elif Option == 'Intercept':
            R = ((Xi - Xo) ** 2 + (Yi - Yo) ** 2) ** 0.5
            Depth = Yo - R
        elif Option == 'Radius':
            Depth = Yo - R
        else:
            raise ValueError(f"Unknown option '{Option}' for circles.")
        circle = {
            "Xo": Xo,
            "Yo": Yo,
            "Depth": Depth,
            "R": R,
        }
        circles.append(circle)

    # --- optional search window (v19, circles!J7 header, J8:K17 label/value pairs) ---
    # Ten independent limits confining a circular search (Slide2's "search limits").
    # Every one is optional; the dict is present only if at least one value is
    # filled, and carries only the filled keys, so a caller can spread it into
    # circular_search without inventing bounds nobody entered. Read positionally
    # from rows 8-17 of column K (SEARCH_WINDOW_KEYS is in sheet order), gated on
    # the template version so a pre-v19 sheet -- where those rows hold nothing or,
    # on old layouts, unrelated text -- is never scanned.
    search_window = {}
    if _tv >= 19:
        for _i, _key in enumerate(SEARCH_WINDOW_KEYS):
            _r = 7 + _i                      # Excel row 8 -> index 7
            if _r >= raw_df.shape[0] or raw_df.shape[1] < 11:
                break
            _v = raw_df.iloc[_r, 10]         # column K
            if pd.isna(_v) or str(_v).strip() == '':
                continue
            try:
                search_window[_key] = float(_v)
            except (TypeError, ValueError):
                raise ValueError(
                    f"The 'circles' sheet search window cell K{_r + 1} "
                    f"({_key}) contains {_v!r}, which is not a number. Leave it "
                    "blank if the limit does not apply.")
        for _lo, _hi in (('entry_x_min', 'entry_x_max'),
                         ('exit_x_min', 'exit_x_max'),
                         ('center_box_x_min', 'center_box_x_max'),
                         ('center_box_y_min', 'center_box_y_max')):
            if (_lo in search_window and _hi in search_window
                    and search_window[_lo] > search_window[_hi]):
                raise ValueError(
                    f"The 'circles' sheet search window has {_lo} = "
                    f"{search_window[_lo]} > {_hi} = {search_window[_hi]}. "
                    "Ranges must be increasing.")

    # === NON-CIRCULAR SURFACES ===
    noncirc_df = xls.parse('non-circ')
    non_circ = list(noncirc_df.iloc[1:].dropna(subset=['Unnamed: 0']).apply(
        lambda row: {
            "X": float(row['Unnamed: 0']),
            "Y": float(row['Unnamed: 1']),
            "Movement": row['Unnamed: 2']
        }, axis=1))

    # === REINFORCEMENT LINES ===
    # Header-name-driven, so both template layouts load: v11 is
    # (# | x1 y1 x2 y2 Tmax Tres Lp1 Lp2 E Area); v12 inserts Label at column B,
    # regroups by analysis type, and adds Type/Dir/Appl/Tend1/Tend2/Spacing.
    reinforce_df = xls.parse('reinforce', header=1)  # Header in row 2 (0-indexed row 1)
    reinforce_df.columns = [str(c).strip().lower() for c in reinforce_df.columns]
    reinforcement_lines = []    # FEM format: list of dicts with raw line endpoints and properties

    # Type presets fill Dir/Appl when those cells are blank (mirrors the in-sheet
    # default formulas); explicit values win. REINFORCE_TYPE_PRESETS (module level)
    # is the one table -- the sheet's lookup block and the Studio editor's fill.
    _TYPE_PRESETS = REINFORCE_TYPE_PRESETS

    for i, row in reinforce_df.iterrows():
        excel_row = i + 3
        # Stop reading when x1 is empty (NOT column B positionally -- in v12,
        # column B is the optional Label and may be blank on a data row)
        if pd.isna(row.get('x1')):
            break

        label = (str(row['label']).strip()
                 if 'label' in reinforce_df.columns and pd.notna(row.get('label'))
                 else f"Line {i + 1}")

        # Check if other required coordinates are present
        if pd.isna(row.get('y1')) or pd.isna(row.get('x2')) or pd.isna(row.get('y2')):
            continue  # Skip rows with incomplete coordinate data

        # If coordinates are present, check for required parameters (Tmax, Lp1, Lp2)
        if pd.isna(row.get('tmax')) or pd.isna(row.get('lp1')) or pd.isna(row.get('lp2')):
            raise ValueError(
                f"Reinforcement line '{label}' (reinforce sheet, Excel row {excel_row}) has "
                "coordinates but missing required parameters (Tmax, Lp1, Lp2). "
                "All three must be specified.")

        # v12 support-type columns; defaults reproduce pre-v12 behavior exactly
        # (generic tensile line: tangent direction, active application).
        rtype = _choice(row.get('type'), '')
        if rtype not in ('',) + tuple(_TYPE_PRESETS):
            raise ValueError(
                f"Reinforcement line '{label}' (reinforce sheet, Excel row {excel_row}) has "
                f"an unrecognized Type='{rtype}'. Expected one of: "
                f"{', '.join(_TYPE_PRESETS)} (or leave blank for a generic line).")
        _dir_def, _appl_def = _TYPE_PRESETS.get(rtype, REINFORCE_TYPE_DEFAULT)

        direction = _choice(row.get('dir'), _dir_def)
        if direction not in ('tangent', 'axial'):
            raise ValueError(
                f"Reinforcement line '{label}' (reinforce sheet, Excel row {excel_row}) has "
                f"an unrecognized Dir='{direction}'. Expected: tangent or axial.")

        appl = _choice(row.get('appl'), _appl_def)
        if appl not in ('active', 'passive'):
            raise ValueError(
                f"Reinforcement line '{label}' (reinforce sheet, Excel row {excel_row}) has "
                f"an unrecognized Appl='{appl}'. Expected: active or passive.")

        _sp_num = pd.to_numeric(row.get('spacing'), errors='coerce')
        spacing = float(_sp_num) if pd.notna(_sp_num) else 1.0
        if spacing <= 0:
            raise ValueError(
                f"Reinforcement line '{label}' (reinforce sheet, Excel row {excel_row}) has "
                f"Spacing = {spacing}; it must be positive (blank or 1 for geosynthetics).")

        tend1 = _num(row.get('tend1', 0))
        tend2 = _num(row.get('tend2', 0))
        if tend1 < 0 or tend2 < 0:
            raise ValueError(
                f"Reinforcement line '{label}' (reinforce sheet, Excel row {excel_row}) has "
                f"a negative end anchorage capacity (Tend1 = {tend1}, Tend2 = {tend2}).")

        try:
            # Extract coordinates and parameters into the raw (FEM) format. All
            # capacity terms and the axial stiffness are per unit width: discrete
            # supports enter per-element values plus Spacing and are divided here,
            # once, for both engines. (spacing defaults to 1 -> no-op for v11 files
            # and geosynthetics.)
            reinforcement_lines.append({
                "x1": float(row['x1']), "y1": float(row['y1']),
                "x2": float(row['x2']), "y2": float(row['y2']),
                "t_max": float(row['tmax']) / spacing,
                # A BLANK Tres means "no post-peak drop" — the bar is elastic-
                # perfectly-plastic and holds its capacity once it yields. It does
                # NOT mean zero. Zero is a legitimate, and very aggressive, entry:
                # it says the bar ruptures brittly and carries nothing afterwards.
                # Defaulting a blank cell to 0.0 silently made brittle rupture the
                # behavior of every file that never mentions Tres. NaN carries the
                # "unset" sense through to the FEM, which softens only where t_res
                # is finite.
                "t_res": (float(row['tres']) / spacing
                          if pd.notna(row.get('tres')) else float('nan')),
                "lp1": float(row['lp1']) if not pd.isna(row['lp1']) else 0.0,
                "lp2": float(row['lp2']) if not pd.isna(row['lp2']) else 0.0,
                "E": float(row['e']) if pd.notna(row.get('e')) else float('nan'),
                "area": (float(row['area']) if pd.notna(row.get('area')) else float('nan')) / spacing,
                "label": label,
                "type": rtype,
                "dir": direction,
                "appl": appl,
                "tend1": tend1 / spacing,
                "tend2": tend2 / spacing,
                "spacing": spacing,
                # v24 overburden-dependent pullout. NaN = blank = the constant-
                # rate law from Lp1/Lp2, so a file written before the columns
                # existed reads exactly as it always did. These two are entered
                # per unit AREA of interface and stay that way — the rate they
                # produce is divided by Spacing where it is integrated, not
                # here.
                "adhesion": (float(row['adhesion'])
                             if pd.notna(row.get('adhesion')) else float('nan')),
                "delta": (float(row['delta'])
                          if pd.notna(row.get('delta')) else float('nan')),
            })
        except Exception as e:
            raise ValueError(f"Error processing reinforcement line '{label}' in row {excel_row}: {e}")

    # LEM tension-distribution format, derived from the raw endpoints/pullout data.
    reinforce_lines = build_reinforce_lines(reinforcement_lines)


    # === PILE LINES ===
    pile_lines = []
    if 'piles' in xls.sheet_names:
        # Header-name-driven with case-normalized headers, so the v12 column
        # regrouping (and the v11 layout) both load. The force-angle override
        # column was dropped in v23 (never used; the direction auto-derives
        # from the pile endpoints) — pre-v23 files may still carry it as 'qp'
        # ('theta' in some older files) and either is honored when present.
        piles_df = xls.parse('piles', header=1)
        piles_df.columns = [str(c).strip().lower() for c in piles_df.columns]
        _theta_col = 'qp' if 'qp' in piles_df.columns else 'theta'
        # v25 split the single rotation restraint into one per end: the old
        # 'Fixity' column became 'Head' (top node) and a new 'Tip' column
        # (bottom node) was inserted after it. A v24 or older file carries
        # only 'Fixity', which is read as the head, and no Tip at all, which
        # reads as 'free' -- exactly the behavior it had before the column
        # existed.
        _head_col = 'head' if 'head' in piles_df.columns else 'fixity'
        for i, row in piles_df.iterrows():
            # Stop reading when column x1 is empty
            if pd.isna(row.get('x1')):
                break
            # Check required coordinates
            if pd.isna(row.get('y1')) or pd.isna(row.get('x2')) or pd.isna(row.get('y2')):
                continue
            try:
                x1, y1 = float(row['x1']), float(row['y1'])
                x2, y2 = float(row['x2']), float(row['y2'])
                H = float(row['h']) if pd.notna(row.get('h')) else None
                if pd.notna(row.get(_theta_col)):
                    theta_p = float(row[_theta_col])
                else:
                    # Auto-compute: perpendicular to pile axis (0 for vertical)
                    dx = x2 - x1
                    dy = y2 - y1
                    theta_p = np.degrees(np.arctan2(dx, -dy))
                D_pile = float(row['d']) if pd.notna(row.get('d')) else None
                S = float(row['s']) if pd.notna(row.get('s')) else None
                E_pile = float(row['e']) if pd.notna(row.get('e')) else None
                I_pile = float(row['i']) if pd.notna(row.get('i')) else None
                area = float(row['area']) if pd.notna(row.get('area')) else None
                V_cap = float(row['vcap']) if pd.notna(row.get('vcap')) else None
                M_cap = float(row['mcap']) if pd.notna(row.get('mcap')) else None
                head_fixity = (str(row[_head_col]).strip().lower()
                               if pd.notna(row.get(_head_col)) else 'free')
                if head_fixity not in ('free', 'pinned', 'unrotated', 'fixed'):
                    raise ValueError(f"{_head_col.capitalize()} must be 'free', "
                                     f"'pinned', 'unrotated' or 'fixed', got "
                                     f"'{head_fixity}'")
                tip_fixity = (str(row['tip']).strip().lower()
                              if pd.notna(row.get('tip')) else 'free')
                if tip_fixity not in ('free', 'pinned', 'unrotated', 'fixed'):
                    raise ValueError(f"Tip must be 'free', 'pinned', 'unrotated' or "
                                     f"'fixed', got '{tip_fixity}'")
                # Force application (v12, LEM only): Active = allowable force, not
                # divided by FS (default, pre-v12 behavior); Passive = ultimate
                # capacity divided by FS.
                appl_raw = str(row['appl']).strip().lower() if pd.notna(row.get('appl')) else 'active'
                if appl_raw not in ('active', 'passive'):
                    raise ValueError(f"Appl must be 'active' or 'passive', got '{appl_raw}'")
                label = str(row['label']) if pd.notna(row.get('label')) else f"Pile {i+1}"

                # Validate
                line_length = ((x2 - x1)**2 + (y2 - y1)**2)**0.5
                if line_length == 0:
                    continue
                if H is not None and H <= 0:
                    raise ValueError(f"H must be positive, got {H}")
                if V_cap is not None and V_cap <= 0:
                    raise ValueError(f"Vcap must be positive, got {V_cap}")
                if M_cap is not None and M_cap <= 0:
                    raise ValueError(f"Mcap must be positive, got {M_cap}")
                if (V_cap is not None or M_cap is not None) and S is None:
                    raise ValueError(f"S (pile spacing) is required when Vcap or Mcap are specified")

                pile_lines.append({
                    "x1": x1, "y1": y1,
                    "x2": x2, "y2": y2,
                    "H": H,
                    "theta_p": theta_p,
                    "D_pile": D_pile,
                    "S": S,
                    "E": E_pile,
                    "I": I_pile,
                    "area": area,
                    "V_cap": V_cap,
                    "M_cap": M_cap,
                    "head_fixity": head_fixity,
                    "tip_fixity": tip_fixity,
                    "appl": appl_raw,
                    "label": label,
                })
            except Exception as e:
                raise ValueError(f"Error processing pile in row {i + 3}: {e}")

    # === LINE LOADS (v12 'lloads' sheet) ===
    # A concentrated force per unit width applied at a point on the ground
    # surface (e.g. the weight of a shotcrete facing plate). Absent in pre-v12
    # templates -> empty list.
    line_loads = []
    if 'lloads' in xls.sheet_names:
        lloads_df = xls.parse('lloads', header=1)
        lloads_df.columns = [str(c).strip().lower() for c in lloads_df.columns]
        # Snap tolerance for "on the ground surface": 0.5% of the model height,
        # floored to a small absolute value for degenerate geometries.
        if not ground_surface.is_empty:
            _gs_ys = [p[1] for p in ground_surface.coords]
            _ll_tol = max(1e-6, 0.005 * (max(_gs_ys) - min(_gs_ys)))
        for i, row in lloads_df.iterrows():
            excel_row = i + 3
            if pd.isna(row.get('x')):
                break
            ll_label = (str(row['label']).strip()
                        if 'label' in lloads_df.columns and pd.notna(row.get('label'))
                        else f"Load {i + 1}")
            if pd.isna(row.get('y')) or pd.isna(row.get('p')):
                raise ValueError(
                    f"Line load '{ll_label}' (lloads sheet, Excel row {excel_row}) needs "
                    "x, y, and P; one or more are blank.")
            ll_x, ll_y = float(row['x']), float(row['y'])
            ll_p = float(row['p'])
            if ll_p <= 0:
                raise ValueError(
                    f"Line load '{ll_label}' (lloads sheet, Excel row {excel_row}) has "
                    f"P = {ll_p}; the magnitude must be positive (use Angle for direction).")
            ll_angle = float(row['angle']) if pd.notna(row.get('angle')) else -90.0
            # The load must act on the ground surface: snap small mismatches from
            # rounded coordinates, refuse anything farther than the tolerance.
            if not ground_surface.is_empty:
                _pt = Point(ll_x, ll_y)
                _d = ground_surface.distance(_pt)
                if _d > _ll_tol:
                    raise ValueError(
                        f"Line load '{ll_label}' (lloads sheet, Excel row {excel_row}) at "
                        f"({ll_x}, {ll_y}) is {_d:.3g} away from the ground surface "
                        f"(tolerance {_ll_tol:.3g}). Line loads must act on the ground surface.")
                _snapped = ground_surface.interpolate(ground_surface.project(_pt))
                ll_x, ll_y = float(_snapped.x), float(_snapped.y)
            line_loads.append({
                "x": ll_x, "y": ll_y,
                "P": ll_p,
                "angle": ll_angle,
                "label": ll_label,
            })


    # === SEEPAGE ANALYSIS BOUNDARY CONDITIONS ===
    seepage_bc = _read_seep_bc_sheet(xls.parse('seep bc', header=None), 'seep bc')
    try:
        seepage_bc2 = _read_seep_bc_sheet(xls.parse('seep bc (2)', header=None), 'seep bc (2)')
    except (ValueError, KeyError) as e:
        if isinstance(e, ValueError) and 'Flux BC' in str(e):
            raise
        # Sheet absent (older workbook) -> no second BC set.
        seepage_bc2 = {"specified_heads": [], "specified_fluxes": [], "exit_face": []}

    # === TRANSIENT SEEPAGE (v18 'tseep' sheet) ===
    # Absent or all-blank -> None (no key, steady behavior, bit-identical to pre-v18).
    tseep = _parse_tseep_sheet(xls, template_version=_tv)

    # 'seep bc (2)' is the CONSTANT-STEADY rapid-drawdown boundary set (the second of
    # the two steady solves). There is exactly one transient timeline and it belongs
    # to the main 'seep bc' sheet, so set 2 may carry neither a reservoir (submerged-
    # only, inherently time-varying) boundary NOR a time-varying value (a string naming
    # a tseep series). Reject both LOUDLY and specifically here — before the generic
    # series-name check below — so the message names the real rule instead of a vaguer
    # "series not defined". The alternative is a set-2 solve silently ignoring a
    # schedule the user believed was applied.
    for _b in seepage_bc2.get('specified_heads', []):
        if str(_b.get('kind', 'head')).strip().lower() == 'reservoir':
            raise ValueError(
                "The 'seep bc (2)' sheet has a head block with type 'reservoir'. "
                "'seep bc (2)' is the constant-steady rapid-drawdown boundary set and "
                "cannot carry a reservoir (submerged-only, time-varying) boundary. "
                "Reservoir boundaries — and any time-varying boundaries — belong on "
                "the main 'seep bc' sheet.")
    for _kind, _vk in (('specified_heads', 'head'), ('specified_fluxes', 'flux')):
        for _b in seepage_bc2.get(_kind, []):
            if isinstance(_b[_vk], str):
                raise ValueError(
                    f"The 'seep bc (2)' sheet binds a {_vk} to a time series "
                    f"({_b[_vk]!r}). 'seep bc (2)' is the constant-steady rapid-drawdown "
                    f"boundary set and cannot carry a time-varying (tseep series) value. "
                    f"Enter a number here; time-varying boundaries belong on the main "
                    f"'seep bc' sheet.")

    # A seep-BC VALUE cell may name a tseep series (a time-varying head/flux). Resolve
    # every string value now: it must match a tseep series header, else a hard error
    # listing the available names. A string with no tseep sheet is likewise an error
    # (there are no series to bind to) -- pre-v18 files never reach here because their
    # value cells parse as floats.
    _series_names = set(tseep["series"]) if tseep else set()
    for _bcset, _label in ((seepage_bc, 'seep bc'), (seepage_bc2, 'seep bc (2)')):
        for _kind, _vk in (('specified_heads', 'head'), ('specified_fluxes', 'flux')):
            for _b in _bcset.get(_kind, []):
                _v = _b[_vk]
                if not isinstance(_v, str):
                    continue
                if tseep is None:
                    raise ValueError(
                        f"The '{_label}' sheet gives a non-numeric {_vk} value {_v!r}. "
                        f"A time-series name is only valid when a 'tseep' sheet defines "
                        f"series; this file has none. Enter a number, or add a tseep "
                        f"series named {_v!r}.")
                if _v not in _series_names:
                    _avail = ', '.join(sorted(_series_names)) or '(none defined)'
                    raise ValueError(
                        f"The '{_label}' sheet binds a {_vk} to tseep series {_v!r}, "
                        f"which is not defined. Available tseep series: {_avail}.")

    # === VALIDATION ===
 
    circular = len(circles) > 0
    # The file's own surface-family selection (main D24) decides the flag when the
    # deck defines BOTH families -- that flag is what every non-dialog consumer reads
    # (the runners, plot, sensitivity), so a file that says 'non-circular' must not
    # come back as a circular model just because its circles sheet is filled. With
    # one family present the selection changes nothing: presence already answers it,
    # and a stale word left in the cell must never claim a surface the deck does not
    # define.
    if surface_family is not None and circular and len(non_circ) > 0:
        circular = surface_family == 'circular'
    # Check if this is a seep-only analysis (has seep BCs but no slope stability surfaces)
    has_seepage_bc = (len(seepage_bc.get("specified_heads", [])) > 0 or
                     len(seepage_bc.get("specified_fluxes", [])) > 0 or
                     len(seepage_bc.get("exit_face", [])) > 0)
    # An editor load (require_analysis_data=False) of a model with no surfaces takes
    # the same validation path: it is a seepage model in progress, so unit weights
    # may still be blank and no failure surfaces are demanded.
    is_seepage_only = ((has_seepage_bc or not require_analysis_data)
                       and not circular and len(non_circ) == 0)
    # A mesh-based run with no LEM surfaces is a seepage or FEM (SSRM) analysis;
    # neither needs circular/non-circular failure surfaces.
    is_mesh_analysis = mesh is not None and not circular and len(non_circ) == 0

    # Only require circular/non-circular data for a pure LEM run (no seep BCs, no mesh)
    if not is_seepage_only and not is_mesh_analysis and not circular and len(non_circ) == 0:
        raise ValueError(
            "Input must include circular or non-circular surface data (for a "
            "slope-stability run), a mesh, or seepage boundary conditions (for a "
            "seepage run).")
    if not polygons:
        raise ValueError("Geometry is missing: provide either the 'profile' sheet or the 'polygon' sheet.")
    if not materials:
        raise ValueError("Materials sheet is empty.")

    # Every polygon must reference a material that exists in the 'mat' sheet. The
    # polygon-sheet path validates this at parse time; profile-derived polygons are
    # validated here (materials are not yet parsed when profile lines are read).
    for poly in polygons:
        mid = poly.get('mat_id')
        if mid is not None and (mid < 0 or mid >= len(materials)):
            raise ValueError(
                f"A geometry zone references an invalid Mat ID ({mid + 1}); it must "
                f"reference a material in the 'mat' sheet (1..{len(materials)}).")

    # A missing unit weight is NOT refused here. Loading answers "is this workbook
    # structurally readable" -- sheet shapes, required columns, ID references -- and
    # a model being built legitimately has values still to type. Whether gamma is
    # usable is a question about the RUN, and preflight's 'mat.gamma_nonpositive'
    # asks it for every stability analysis (and stays silent for a seepage-only one,
    # which never reads unit weight).

    # === TRANSIENT-SEEPAGE VALIDATION (only when a tseep sheet is in use) ===
    if tseep is not None:
        # A transient run makes the implied TIME unit load-bearing everywhere (k is
        # len/time, storage is 1/len, the tseep times), so it must be declared. Never
        # guessed -- a wrong time label is worse than none.
        if time_unit is None:
            raise ValueError(
                "A 'tseep' (transient seepage) sheet is in use, but no Time unit is "
                "declared in the 'main' sheet (D8/D9 Units & Time selectors). Transient "
                "seepage needs a declared time base; set the Time unit.")
        # Storage parameters: Ss is required for every material; Sy is required only
        # when the model is unconfined (an exit-face BC exists -> desaturation possible;
        # a confined/always-saturated transient needs only Ss).
        _unconfined = bool(seepage_bc.get('exit_face') or seepage_bc2.get('exit_face'))
        for _m in materials:
            if _m.get('Ss') is None:
                raise ValueError(
                    f"Material '{_m['name']}' has no Ss (specific storage), which is "
                    f"required for a transient (tseep) analysis. Set Ss on every "
                    f"material, or remove the tseep sheet for a steady run.")
            if _unconfined and _m.get('Sy') is None:
                raise ValueError(
                    f"Material '{_m['name']}' has no Sy (specific yield). Sy is required "
                    f"for an unconfined transient analysis (this model has an exit-face "
                    f"BC, so desaturation is possible).")
        # Saved-frame schedule = union of the save_interval schedule, save_times, stage
        # times, and series breakpoints. Entries beyond the run duration are never
        # reached -- warn (not an error; the run is still valid).
        _dur = tseep.get('duration')
        if _dur is not None:
            _over = [t for t in tseep.get('save_times', []) if t > _dur]
            for _st in (tseep.get('stage_1'), tseep.get('stage_2'),
                        tseep.get('stability_time')):
                if _st is not None and _st > _dur:
                    _over.append(_st)
            for _name, _vals in tseep.get('series', {}).items():
                for _t, _v in zip(tseep.get('times', []), _vals):
                    if _v is not None and _t > _dur:
                        _over.append(_t)
                        break
            if _over:
                warnings.warn(
                    f"tseep: {len(_over)} scheduled/breakpoint time(s) exceed the run "
                    f"duration ({_dur:g}) and will not be reached: "
                    f"{sorted(set(_over))}.")

    # Add everything to globals_data
    globals_data["template_version"] = template_version
    globals_data["gamma_water"] = gamma_water
    globals_data["tcrack_depth"] = tcrack_depth
    globals_data["tcrack_water"] = tcrack_water
    globals_data["k_seismic"] = k_seismic
    globals_data["max_depth"] = max_depth
    globals_data["profile_lines"] = profile_lines
    globals_data["polygons"] = polygons
    # v20 SSR zone overlays (polygon-sheet sentinel rows). Deliberately a SEPARATE
    # key from 'polygons': these never take part in geometry — only the SSRM reads
    # them (fem.build_fem_data carries them onto fem_data; solve_ssrm composes the
    # element masks) and the input plots draw them as dashed outlines.
    globals_data["ssr_zones"] = ssr_zones
    # v21 mesh-refinement overlays (polygon-sheet Type='refine' rows). Like the SSR
    # zones these are NOT geometry: they carry no material, never become a mesh
    # region and never generate slices. Their only effect is the local target element
    # size inside the ring, applied as a gmsh size field (see mesh.size_regions).
    globals_data["refine_zones"] = refine_zones
    globals_data["domain_polygon"] = domain_polygon
    globals_data["ground_surface"] = ground_surface
    globals_data["tcrack_surface"] = tcrack_surface
    globals_data["materials"] = materials
    globals_data["piezo_line"] = piezo_line
    globals_data["piezo_phreatic"] = piezo_phreatic
    globals_data["piezo_phreatic2"] = piezo_phreatic2
    globals_data["piezo_line2"] = piezo_line2
    globals_data["circular"] = circular # True if circles are present (or selected)
    globals_data["circles"] = circles
    # v22 amendment: the file's surface-family selection (main D24), 'circular' /
    # 'noncircular' / None. None is the normal state and means "resolve
    # automatically" -- from the run's own choice if it makes one, otherwise from
    # which family the model defines.
    globals_data["surface_family"] = surface_family
    # v19 run options. Every one is None when the file does not declare it (and on
    # every pre-v19 file), which every consumer must read as "use your own default".
    globals_data["lem_method"] = lem_method
    globals_data["num_slices"] = num_slices_opt
    globals_data["k0"] = k0
    globals_data["tension_srf"] = tension_srf_opt
    globals_data["element_type"] = element_type
    globals_data["target_size"] = target_size
    globals_data["element_size_1d"] = element_size_1d
    globals_data["ssrm_f_min"] = ssrm_f_min
    globals_data["ssrm_f_max"] = ssrm_f_max
    # v21 side boundary condition (None = unspecified -> the engine default, rollers).
    globals_data["side_bc"] = side_bc
    # v22 water-load mode, already resolved by template version above: this is
    # always 'auto' or 'manual', never None, because every consumer needs a
    # definite answer and the answer for a file that does not say is 'manual'.
    globals_data["water_loads"] = water_loads
    # v19 circles-sheet search window: present only when at least one limit is set.
    if search_window:
        globals_data["search_window"] = search_window
    globals_data["non_circ"] = non_circ
    globals_data["dloads"] = dloads
    globals_data["dloads2"] = dloads2
    # v21 per-block load direction, parallel to the lists above ('normal' | 'vertical').
    # Always the same length as its list, so a consumer can zip the two without
    # guarding; every entry is 'normal' on a pre-v21 file.
    globals_data["dload_dirs"] = dload_dirs
    globals_data["dload2_dirs"] = dload2_dirs
    globals_data["reinforce_lines"] = reinforce_lines
    globals_data["reinforcement_lines"] = reinforcement_lines
    globals_data["pile_lines"] = pile_lines
    globals_data["line_loads"] = line_loads
    globals_data["seepage_bc"] = seepage_bc
    globals_data["seepage_bc2"] = seepage_bc2
    globals_data["has_seepage_bc2"] = bool(seepage_bc2.get("specified_heads")
                                           or seepage_bc2.get("specified_fluxes")
                                           or seepage_bc2.get("exit_face"))
    
    # Add mesh if available (used by both seep and fem workflows)
    globals_data["mesh"] = mesh

    # Add seep solution data if available
    if has_seep_materials:
        globals_data["seep_u"] = seep_u
        if seep_u2 is not None:
            globals_data["seep_u2"] = seep_u2

    # === UNIT DECLARATION (units plan phases 2 & 3) ===
    # unit_system/time_unit were resolved in the version-gated main-globals block above:
    #   - v18: unit_system from the Units selector (D8), or inferred from gamma_w when the
    #     selector is blank; time_unit from the Time selector (D9), else None.
    #   - <=17: no selectors, so unit_system is INFERRED from gamma_water's magnitude
    #     (~9.81/9.807 -> SI, ~62.4 -> Imperial) and time_unit is None.
    # Inference only LABELS the model -- gamma_water is used exactly as entered either
    # way, so physics is unchanged. time_unit is NEVER inferred (a wrong time label is
    # worse than none -- the min-vs-sec mislabel lesson).
    globals_data["unit_system"] = unit_system
    globals_data["time_unit"] = time_unit
    if tseep is not None:
        globals_data["tseep"] = tseep
    if unit_system is None:
        warnings.warn(
            f"This file declares no unit system and its unit weight of water "
            f"({gamma_water:g}) matches neither the SI (~9.81) nor the Imperial "
            f"(~62.4) band, so the model is left unlabeled. Physics is unaffected "
            f"(gamma_water is used as entered); only unit labels are unavailable.")
    # Overburden-dependent pullout reads the geometry, the materials and the
    # water, so it is resolved once the whole model is assembled — the profile
    # then rides on each line for both engines and for the plots.
    attach_reinforce_pullout(globals_data)

    for w in units_check(globals_data):
        warnings.warn(w)

    return globals_data


def default_template_path():
    """Filesystem path to the blank XSLOPE input template bundled with the
    package (``xslope/resources/input_template.xlsx``).

    Pass it as the ``template=`` argument to :func:`save_slope_data_to_xlsx` to
    create a new file from the standard template — useful for library-only
    (no-GUI) users, since the template ships inside the installed package and so
    is available regardless of where pip put it. The docs copy
    (``docs/inputs/input_template.xlsx``) is the editable master; the two are kept
    byte-identical by a check in ``run_tests.py``.
    """
    return os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "resources", "input_template.xlsx")


# in-memory material key -> 'mat' sheet header text. Shared by the writer and by the
# input-building scripts, so a new column is declared in exactly one place.
MAT_NUM_HEADERS = [
    ('gamma', 'g'), ('c', 'c'), ('phi', 'f'), ('cp', 'c/p'), ('r_elev', 'r-elev'),
    ('d', 'd'), ('psi', 'psi'),
    ('pow_a', 'powa'), ('pow_b', 'powb'), ('pow_c', 'powc'), ('pow_d', 'powd'),
    ('ru', 'ru'),
    ('sigma_gamma', 's(g)'), ('sigma_c', 's(c)'), ('sigma_phi', 's(f)'),
    ('sigma_cp', 's(c/p)'), ('sigma_d', 's(d)'), ('sigma_psi', 's(psi)'),
    ('hb_sci', 'hb_sci'), ('hb_gsi', 'hb_gsi'),
    ('hb_mi', 'hb_mi'), ('hb_d', 'hb_d'),
    ('k1', 'k1'), ('k2', 'k2'), ('alpha', 'alpha'),
    # v14: the curve-parameter pair is law-agnostic ('a'/'n'), serving both van
    # Genuchten and Gardner; Poisson's ratio moved off 'n' to 'nu' to break the
    # header collision that created.
    # v14: the curve-parameter pair is law-agnostic ('a'/'n'), serving both van
    # Genuchten and Gardner; Poisson's ratio moved off 'n' to 'nu' to break the
    # header collision that created. A header entry may therefore be a TUPLE of
    # candidate names, tried in order and resolved against the destination sheet
    # -- the writer's mirror of the loader's ``_pick`` (see its docstring). The
    # order is the loader's, deliberately: on a pre-v14 sheet 'n' IS Poisson's
    # ratio, and without the alternatives the writer would skip nu entirely and
    # put vg_n in its cell, silently rewriting Poisson's ratio on every save into
    # an archived template.
    ('kr0', 'kr0'), ('h0', 'h0'),
    ('vg_a', ('vga', 'a')), ('vg_n', ('vgn', 'n')),
    ('E', 'E'), ('nu', ('nu', 'n')),
]
# Optional numerics: written only when set (None must stay a blank cell -- e.g.
# a blank gsat means "fall back to gamma", which 0.0 would not; a blank t_cut
# means "no cutoff", which 0.0 -- "no tension" -- would not).
MAT_OPT_NUM_HEADERS = [('gamma_sat', 'gsat'), ('t_cut', 't_cut'),
                       ('phi_b', 'phi_b'), ('s_cap', 's_cap'),
                       # v18 transient storage: written only when set, so a blank
                       # stays a blank cell (None), never a silent 0.
                       ('Ss', 'Ss'), ('Sy', 'Sy')]
MAT_STR_HEADERS = [('option', 'option'), ('u', 'u'), ('unsat', 'unsat')]

# The 'mat' header row is located by scanning for its sentinel cells rather than
# assumed to be row 8: the header has moved before (pre-v8 templates put it on row 3,
# and none of those load today), and a legend row inserted above the table moves it
# again. Bounded scan -- the legend above the table is short.
_MAT_HEADER_SCAN_ROWS = 20


def _find_mat_header_row(cell):
    """Row index (1-based) of the 'mat' sheet header, found by its sentinel cells:
    column A == 'mat' and column B == 'name'.

    ``cell(row, col) -> value`` is a lookup callable, so this works against either an
    openpyxl worksheet or a header-less DataFrame.

    The row analogue of :data:`MAT_NUM_HEADERS`' by-name column lookup: neither the
    header's row nor its columns may be hardcoded, so inserting a legend row or a new
    property column cannot silently shift what gets read or written.

    Raises:
        ValueError: if no header row is found in the first 20 rows.
    """
    def _s(v):
        return '' if v is None else str(v).strip().lower()

    for r in range(1, _MAT_HEADER_SCAN_ROWS + 1):
        if _s(cell(r, 1)) == 'mat' and _s(cell(r, 2)) == 'name':
            return r
    raise ValueError(
        "The 'mat' sheet has no header row: expected a row whose first two cells are "
        f"'mat' and 'name' within the first {_MAT_HEADER_SCAN_ROWS} rows."
    )


def _read_mat_header_cols(filepath):
    """(header_row, cols) for the 'mat' sheet, read from the destination workbook.

    ``cols`` maps each header's text to its 1-based column index, so
    :func:`save_slope_data_to_xlsx` locates material columns BY NAME rather than by
    hardcoded position and adapts to the template version automatically (v11 inserted
    ``unsat`` and ``vg_a``/``vg_n``, shifting ``kr0``/``h0``/``E``/``nu``). Headers
    absent in an older template are simply skipped by the caller. ``header_row`` is
    found the same way -- see :func:`_find_mat_header_row`.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=False)
    try:
        ws = wb['mat']
        header_row = _find_mat_header_row(lambda r, c: ws.cell(row=r, column=c).value)
        cols = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            if v is not None and str(v).strip() != '':
                # underscore-insensitive, matching load_slope_data's mat parse
                cols[str(v).strip().replace('_', '')] = c
        return header_row, cols
    finally:
        wb.close()


def _read_template_info(filepath):
    """``(version:int, sheet_names:set)`` for a workbook, read via openpyxl.

    The writer targets whatever template it is handed (the current standard template,
    or an explicit ``template=`` for Save As), so the main-sheet globals must be
    written at the positions THAT template uses -- v18 shifted them (see
    :func:`load_slope_data`). Reading the destination's own version keeps the writer
    version-faithful: a v18 template gets v18 cells, a v17 template gets v17 cells,
    regardless of the version the in-memory data was loaded from. ``version`` is 0 when
    the cell is unreadable (older templates still take the <=17 branch)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=False)
    try:
        raw = wb['main'].cell(row=5, column=4).value
        names = set(wb.sheetnames)
    finally:
        wb.close()
    try:
        return int(float(raw)), names
    except (TypeError, ValueError):
        return 0, names


def mat_header_cols(filepath):
    """Public accessor for the 'mat' sheet's ``(header_row, {header: column})`` map.

    Input-building scripts must write material cells through this rather than by
    hardcoded column number, or a future column insert shifts them silently.
    """
    return _read_mat_header_cols(filepath)


def _inplace_save_would_drop(filepath, materials):
    """True if editing ``filepath`` in place would silently drop material data
    because its 'mat' sheet lacks a column the data needs — e.g. a pre-v11 file
    (no ``unsat``/``vg_a``/``vg_n`` columns) being saved with a van Genuchten
    material. The caller then upgrades the file through the current template instead
    of dropping the data. Only flags when a material carries a NON-default value for
    a missing column, so old files that don't use the new feature still save in
    place (preserving their formatting)."""
    if not materials:
        return False
    have = set(_read_mat_header_cols(filepath))
    # (header, key, default) for columns that may be absent in older templates.
    checks = [('unsat', 'unsat', 'lf'), ('vg_a', 'vg_a', 0.0), ('vg_n', 'vg_n', 0.0)]
    for header, key, default in checks:
        if header in have:
            continue
        for m in materials:
            val = m.get(key)
            if val is None:
                continue
            if isinstance(default, str):
                if str(val).strip().lower() not in ('', 'nan', default):
                    return True
            else:
                try:
                    if float(val) != default:
                        return True
                except (TypeError, ValueError):
                    pass
    return False


def save_slope_data_to_xlsx(slope_data, filepath, template=None):
    """
    Write an in-memory ``slope_data`` dict back to an XSLOPE Excel input file.

    This is the inverse of :func:`load_slope_data`: it maps every editable input
    category (global parameters, materials, geometry, piezometric lines, failure
    surfaces, distributed loads, reinforcement, piles, and seepage boundary
    conditions) back into the template's sheet/cell layout. It builds on
    :func:`write_cells_to_xlsx`, which edits the workbook at the XML level and so
    preserves all formatting, formulas, charts, and drawings.

    Round-trip guarantee: for every input category,
    ``load_slope_data(<file written by this function>)`` reproduces the source
    ``slope_data``. Derived geometry (``ground_surface``, ``domain_polygon``,
    ``tcrack_surface``, and — for profile input — ``polygons``) is recomputed by
    the loader and therefore not written here.

    Parameters
    ----------
    slope_data : dict
        A dict in the form returned by :func:`load_slope_data`.
    filepath : str
        Destination ``.xlsx`` path.
    template : str, optional
        Path to a blank XSLOPE template to copy before writing (e.g. "Save As" from
        a chosen file). If ``None`` (the default, and the normal "Save"), the current
        standard template from :func:`default_template_path` is used. Either way the
        destination is written from a fresh, data-free template, so ``filepath`` need
        not exist and no stale rows can survive from a previous version of the file.

    Returns
    -------
    str
        ``filepath``.

    Notes
    -----
    Geometry is written to either the ``profile`` sheet (when ``profile_lines`` is
    present) or the ``polygon`` sheet (otherwise) — never both, matching the
    loader's mutual-exclusivity rule. Formula cells (e.g. the profile/polygon
    row-6 material-name XLOOKUP) are never written; the material IDs that feed
    them are, and :func:`write_cells_to_xlsx` flags the workbook for a full
    recalculation on open so the dependent formulas refresh.

    Circle surfaces are always written with ``Option = "Depth"`` (the loader
    collapses every circle to ``Xo/Yo/Depth/R`` with ``R = Yo - Depth``, so this
    reproduces the radius regardless of how the original was specified). The pile
    force direction is auto-derived by the loader from the pile endpoints (the
    pre-v23 templates' ``qp`` override column, when the target carries one, is
    left blank).
    """
    # Always write into a FRESH copy of the standard template rather than editing the
    # destination in place. The template carries no data — only structural helpers the
    # loader ignores (index numbers in column A, dropdown reference lists) — so every
    # save rebuilds each input category from slope_data on a clean sheet. This is
    # robust by construction: a shortened list (a deleted reinforcement line, profile
    # point, material, …) leaves no stale/orphaned rows behind, and it also auto-
    # upgrades older files to the current template format. The trade-off is that
    # user-added custom formulas/formatting in the destination are not preserved.
    # Callers may pass an explicit `template` (e.g. Save As from a chosen file).
    if template is None:
        template = default_template_path()
    # An automatic-water model states its reservoir as the water definition plus the
    # mode cell, and NOT as a row on the dloads sheet. A pre-v22 template has no mode
    # cell, so writing one there would produce a file that loads back as manual with no
    # water load at all -- the pool silently gone, at a factor of safety that still
    # looks perfectly reasonable. Refused rather than degraded, and refused BEFORE the
    # copy so no half-written file is left behind. Vendor imports reach this first,
    # because a GeoStudio or Slide2 model states its water as a surface.
    if str(slope_data.get('water_loads') or '').strip().lower() == 'auto':
        _tpl_version = _read_template_info(template)[0]
        if _tpl_version < _WATER_LOADS_TEMPLATE_VERSION:
            raise ValueError(
                f"Cannot save this model into a version {_tpl_version} template: it "
                f"takes its water loads automatically (main sheet D23 = auto), and "
                f"cell D23 does not exist before template version "
                f"{_WATER_LOADS_TEMPLATE_VERSION}. The file would load back as manual "
                f"carrying no water load at all, losing the standing water. Save into "
                f"the current template, or set the mode to manual and enter the water "
                f"loads on the dloads sheet.")
    shutil.copy(template, filepath)

    def _f(v):
        return float(v)

    def _isnan(v):
        """True for an unset numeric field (NaN or None) — written as a blank cell."""
        if v is None:
            return True
        try:
            return float(v) != float(v)
        except (TypeError, ValueError):
            return False

    updates = {}

    # Version-faithful to the DESTINATION template (not the loaded data's version):
    # the writer copied `template`, so the main-sheet globals go at the positions that
    # template uses, and tseep is written only if the template actually has the sheet.
    _dest_version, _dest_sheets = _read_template_info(filepath)
    # v25 inserted '1D element size' at main!D20 and pushed the run options below it
    # down one row; the same offset the loader uses, applied to the DESTINATION's
    # layout so an archived template is still written at its own positions.
    _dshift = 1 if _dest_version >= _ELEMENT_SIZE_1D_TEMPLATE_VERSION else 0

    # === main === (version-gated positional map; see load_slope_data)
    if _dest_version >= 18:
        main_u = {
            'D10': _f(slope_data['gamma_water']),
            'D11': _f(slope_data['tcrack_depth']),
            'D12': _f(slope_data['tcrack_water']),
            'D13': _f(slope_data['k_seismic']),
        }
        # Units selector (D8) and Time selector (D9) are written UNCONDITIONALLY --
        # blank (None) when undeclared -- so the template's pre-filled defaults never
        # leak into a saved file (e.g. a model with no time_unit must not inherit the
        # template's stock "day"). The selector persists the declared/inferred system
        # in the template's own vocabulary (SI/Imperial); the time unit is written
        # verbatim, never invented.
        _us = normalize_unit_system(slope_data.get('unit_system'))
        main_u['D8'] = 'SI' if _us == 'si' else 'Imperial' if _us == 'imperial' else None
        _tu = slope_data.get('time_unit')
        main_u['D9'] = str(_tu) if _tu else None
        if _dest_version >= 19:
            # v19 run options (D14:D21). Written UNCONDITIONALLY -- None becomes a
            # blank cell -- for the same reason as the Units/Time selectors above:
            # the blank template ships with D17 pre-filled 'YES', and a model that
            # declares nothing must not silently inherit it on save.
            _srf = slope_data.get('tension_srf')
            main_u['D14'] = str(slope_data['lem_method']) if slope_data.get('lem_method') else None
            main_u['D15'] = (int(slope_data['num_slices'])
                             if slope_data.get('num_slices') is not None else None)
            main_u['D16'] = (_f(slope_data['k0'])
                             if slope_data.get('k0') is not None else None)
            main_u['D17'] = None if _srf is None else ('YES' if _srf else 'NO')
            main_u['D18'] = str(slope_data['element_type']) if slope_data.get('element_type') else None
            main_u['D19'] = (_f(slope_data['target_size'])
                             if slope_data.get('target_size') is not None else None)
            if _dest_version >= _ELEMENT_SIZE_1D_TEMPLATE_VERSION:
                # v25 1D element size (D20). Written unconditionally like the rest
                # of the block: the template ships the cell blank, and a model that
                # declares no 1D size must save as blank rather than inherit one.
                main_u['D20'] = (_f(slope_data['element_size_1d'])
                                 if slope_data.get('element_size_1d') is not None
                                 else None)
            main_u[f'D{20 + _dshift}'] = (
                _f(slope_data['ssrm_f_min'])
                if slope_data.get('ssrm_f_min') is not None else None)
            main_u[f'D{21 + _dshift}'] = (
                _f(slope_data['ssrm_f_max'])
                if slope_data.get('ssrm_f_max') is not None else None)
        if _dest_version >= 21:
            # v21 side BC (D22). Written unconditionally for the same leak reason as
            # the run options above: the blank template ships D22 pre-filled
            # 'rollers', and a model that declares nothing must not inherit it.
            _sbc = slope_data.get('side_bc')
            main_u[f'D{22 + _dshift}'] = str(_sbc).lower() if _sbc else None
        if _dest_version >= 22:
            # v22 water loads (D23). Written EXPLICITLY, never blank, and this one
            # matters more than the leak cases above: the template ships D23
            # pre-filled 'auto', so a v21 model saved into a v22 file with a blank
            # cell would come back as an automatic-water model still carrying the
            # water loads its user typed -- the reservoir counted twice. Writing the
            # mode the model actually ran under keeps a save from changing the
            # analysis.
            _wl = str(slope_data.get('water_loads') or 'manual').strip().lower()
            if _wl not in WATER_LOAD_OPTIONS:
                # Refused rather than coerced. Writing 'manual' for a word we do not
                # recognize would silently answer a question the model got wrong,
                # and the answer decides whether the reservoir is counted once.
                raise ValueError(
                    f"Cannot save this model: its water-load mode is {_wl!r}, which "
                    f"is not a value cell D{23 + _dshift} of the 'main' sheet can "
                    f"hold. Expected one of: {', '.join(WATER_LOAD_OPTIONS)}.")
            main_u[f'D{23 + _dshift}'] = _wl
            # v22 amendment: the surface family (D24). Written unconditionally, so a
            # model that states nothing writes a blank cell -- the normal state, and
            # the template ships blank, so there is nothing here to leak. A model that
            # DOES state one (the both-family deck whose run dialog asked) must carry
            # the answer out with it: without this line the choice survives only until
            # the file is closed, and the next session's circles win again in silence.
            _sf = slope_data.get('surface_family')
            if _sf:
                _sf_key = str(_sf).strip().lower().replace('-', '').replace('_', '')
                _sf_key = 'noncircular' if _sf_key.startswith('non') else _sf_key
                if _sf_key not in SURFACE_FAMILY_CELL:
                    raise ValueError(
                        f"Cannot save this model: its surface family is {_sf!r}, which "
                        f"is not a value cell D{24 + _dshift} of the 'main' sheet "
                        f"can hold. Expected one of: "
                        f"{', '.join(SURFACE_FAMILY_OPTIONS)}.")
                main_u[f'D{24 + _dshift}'] = SURFACE_FAMILY_CELL[_sf_key]
            else:
                main_u[f'D{24 + _dshift}'] = None
        updates['main'] = main_u
    else:
        updates['main'] = {
            'D8': _f(slope_data['gamma_water']),
            'D9': _f(slope_data['tcrack_depth']),
            'D10': _f(slope_data['tcrack_water']),
            'D11': _f(slope_data['k_seismic']),
        }

    # === mat ===  Both the header ROW and its COLUMNS are located by name in the
    # destination file, never hardcoded, so the writer adapts to the template version
    # automatically (v11 inserted 'unsat' and 'vg_a'/'vg_n', shifting kr0/h0/E/nu; a
    # legend row above the table would shift the header row). A header absent in an
    # older template is skipped.
    mat_header_row, mat_cols = _read_mat_header_cols(filepath)
    mat = {}
    # mat_cols is keyed underscore-insensitively (see _read_mat_header_cols), so every
    # lookup must normalize the declared header the same way -- otherwise a header
    # written with an underscore ('hb_sci') misses its column and the writer SILENTLY
    # skips it, zeroing the property on the next load.
    def _col(header):
        """Destination column for a declared header, or None when the sheet has no
        such column (an older template: the property is skipped).

        ``header`` may be a tuple of candidate names, tried in order -- the
        renamed columns of MAT_NUM_HEADERS, resolved exactly as the loader's
        ``_pick`` resolves them on the way in."""
        for name in ((header,) if isinstance(header, str) else tuple(header)):
            col = mat_cols.get(name.replace('_', ''))
            if col is not None:
                return col
        return None

    for idx, material in enumerate(slope_data.get('materials', [])):
        row = mat_header_row + 1 + idx
        if 'mat' in mat_cols:
            mat[cell_ref(row, mat_cols['mat'])] = idx + 1     # 1-based mat number
        if 'name' in mat_cols:
            mat[cell_ref(row, mat_cols['name'])] = str(material.get('name', ''))
        # option / u / unsat are strings; leave the cell blank when unset (the loader
        # reads an empty cell back as a default, so writing literal text is noise).
        for key, header in MAT_STR_HEADERS:
            col = _col(header)
            if col is None:
                continue
            val = material.get(key)
            if val is not None and str(val).strip().lower() not in ('', 'nan'):
                mat[cell_ref(row, col)] = str(val)
        for key, header in MAT_NUM_HEADERS:
            col = _col(header)
            if col is None:
                continue
            # An UNSET property is written as a blank cell, not as 0.0. The editors
            # preserve a blank as None now (a cohesionless material and an unfilled
            # one are not the same model), and writing a zero here would put the
            # invention back on the way out -- the file would come back declaring a
            # unit weight of zero, or a Poisson's ratio of zero, that nobody typed.
            val = material.get(key)
            mat[cell_ref(row, col)] = None if val is None else _f(val or 0)
        for key, header in MAT_OPT_NUM_HEADERS:
            col = _col(header)
            if col is None:
                continue
            val = material.get(key)
            if val is not None:
                mat[cell_ref(row, col)] = _f(val)
    if mat:
        updates['mat'] = mat

    # === geometry: profile OR polygon (mutually exclusive, matching the loader) ===
    # The polygon sheet additionally carries the SSR zone overlays and (v21) the mesh
    # refinement overlays, which are NOT geometry and so are written whichever
    # geometry form the model uses — appended after the material zones (or from
    # block #1 on a profile-geometry model).
    #
    # Row map by destination version (see _parse_polygon_sheet):
    #   <= v20   row 5 Mat ID (negative = SSR sentinel),          vertices from row 8
    #   v21      row 5 Type, row 6 Mat ID, row 8 Size,            vertices from row 10
    _v21_poly = _dest_version >= 21
    _poly_matid_row = 6 if _v21_poly else 5
    _poly_coord_row = 10 if _v21_poly else 8
    _type_word_by_kind = {v: k for k, v in POLYGON_TYPE_WORDS.items()}
    _sentinel_by_kind = {v: k for k, v in SSR_ZONE_SENTINELS.items()}

    poly_u = {}
    n_poly_blocks = 0

    def _write_poly_block(coords, kind='material', mat_id=None, size=None):
        """One polygon-sheet block: header, type/Mat ID, optional Size, vertices.

        ``kind`` is the in-memory word ('material', 'reduce', 'hold', 'hold_elastic',
        'refine'). On a v21 destination it is written as the Type word and the Mat ID
        cell is filled only for a material zone; on v20 and earlier the kind is
        encoded in the Mat ID as its sentinel, and 'refine' has no representation at
        all (v21-only concept) so it is refused rather than written as geometry."""
        nonlocal n_poly_blocks
        x_col = 1 + n_poly_blocks * 3
        y_col = x_col + 1
        poly_u[cell_ref(4, x_col)] = f"Polygon #{n_poly_blocks + 1}"  # block header
        if _v21_poly:
            poly_u[cell_ref(5, y_col)] = _type_word_by_kind[kind]
            # Mat ID written UNCONDITIONALLY — blank for an overlay — so the blank
            # template's pre-filled 1/2/3/... cannot survive underneath a Type the
            # reader ignores it for, leaving a stale material ID in a block that has
            # no material.
            poly_u[cell_ref(_poly_matid_row, y_col)] = (
                int(mat_id) if kind == 'material' else None)
            poly_u[cell_ref(8, y_col)] = _f(size) if size is not None else None
        else:
            if kind == 'refine':
                raise ValueError(
                    "This model carries a mesh refinement polygon (Type 'refine'), "
                    f"which template version {_dest_version} has no way to express. "
                    "Save to a version 21 (or later) template.")
            poly_u[cell_ref(_poly_matid_row, y_col)] = (
                int(mat_id) if kind == 'material' else _sentinel_by_kind[kind])
        pts = list(coords)
        if len(pts) >= 2 and tuple(pts[0]) == tuple(pts[-1]):
            pts = pts[:-1]                                 # loader closes implicitly
        for i, (x, y) in enumerate(pts):
            poly_u[cell_ref(_poly_coord_row + i, x_col)] = _f(x)
            poly_u[cell_ref(_poly_coord_row + i, y_col)] = _f(y)
        n_poly_blocks += 1

    profile_lines = slope_data.get('profile_lines') or []
    if profile_lines:
        # v21 inserts an optional Size row at Excel row 7 and pushes the vertices
        # from row 8 to row 9; earlier templates keep the old map.
        _prof_size_row = 7 if _dest_version >= 21 else None
        _prof_coord_row = 9 if _dest_version >= 21 else 8
        prof = {}
        md = slope_data.get('max_depth')
        prof['B2'] = _f(md) if md is not None else 0.0
        for n, line in enumerate(profile_lines):              # n is 0-based
            x_col = 1 + n * 3                                  # A, D, G, ...
            y_col = x_col + 1
            # Write the row-4 block header. load_slope_data() detects how many
            # profile lines exist by scanning for a non-empty header here, so this
            # makes the file self-describing rather than relying on the template's
            # pre-labeled blocks.
            prof[cell_ref(4, x_col)] = f"Profile Line #{n + 1}"
            mat_id = line.get('mat_id')
            if mat_id is not None:
                prof[cell_ref(5, y_col)] = int(mat_id) + 1    # 0-based -> 1-based
            _size = line.get('size')
            if _size is not None and _prof_size_row is not None:
                prof[cell_ref(_prof_size_row, y_col)] = _f(_size)
            for i, (x, y) in enumerate(line['coords']):
                prof[cell_ref(_prof_coord_row + i, x_col)] = _f(x)
                prof[cell_ref(_prof_coord_row + i, y_col)] = _f(y)
        updates['profile'] = prof
    else:
        for pdict in slope_data.get('polygons') or []:
            mat_id = pdict.get('mat_id')
            if mat_id is None:
                continue
            _write_poly_block(pdict['polygon'].exterior.coords, 'material',
                              mat_id=int(mat_id) + 1, size=pdict.get('size'))

    # SSR zone overlays. The kind maps back to its Type word (v21) or its sentinel Mat
    # ID (v20 and earlier), so a file with zones round-trips as zones (an unknown kind
    # is a programming error, not a user input, and raising here beats writing a row
    # the loader would reject).
    for zone in slope_data.get('ssr_zones') or []:
        kind = str(zone.get('kind', '')).strip()
        if kind not in _sentinel_by_kind:
            raise ValueError(
                f"Unknown SSR zone kind {kind!r}; expected one of "
                f"{sorted(_sentinel_by_kind)}.")
        _write_poly_block(zone['polygon'], kind, size=zone.get('size'))

    # v21 mesh refinement overlays. Size is required by the loader, and a zone that
    # somehow lost it would come back as a hard load error, so refuse it here.
    for zone in slope_data.get('refine_zones') or []:
        if zone.get('size') is None:
            raise ValueError(
                "A mesh refinement polygon (Type 'refine') carries no Size. Its only "
                "effect is the local target element size, so the size is required.")
        _write_poly_block(zone['polygon'], 'refine', size=zone['size'])

    if poly_u:
        updates['polygon'] = poly_u

    # === piezo ===  (v13 layout: Type row at Excel row 3 — 'piezo' static head
    # or 'phreatic' cos^2 correction — x/y headers row 4, data from row 5)
    piezo = {}
    piezo[cell_ref(3, 2)] = 'phreatic' if slope_data.get('piezo_phreatic') else 'piezo'
    piezo[cell_ref(3, 5)] = 'phreatic' if slope_data.get('piezo_phreatic2') else 'piezo'
    for i, (x, y) in enumerate(slope_data.get('piezo_line') or []):
        piezo[cell_ref(5 + i, 1)] = _f(x)
        piezo[cell_ref(5 + i, 2)] = _f(y)
    for i, (x, y) in enumerate(slope_data.get('piezo_line2') or []):
        piezo[cell_ref(5 + i, 4)] = _f(x)
        piezo[cell_ref(5 + i, 5)] = _f(y)
    if piezo:
        updates['piezo'] = piezo

    # === circles ===  (header row 2, data rows 3+; always written as Option="Depth")
    circ = {}
    for n, c in enumerate(slope_data.get('circles') or []):
        row = 3 + n
        circ[cell_ref(row, 1)] = n + 1
        circ[cell_ref(row, 2)] = _f(c['Xo'])
        circ[cell_ref(row, 3)] = _f(c['Yo'])
        circ[cell_ref(row, 4)] = 'Depth'
        circ[cell_ref(row, 5)] = _f(c['Depth'])
    # v19 search window (K8:K17, in SEARCH_WINDOW_KEYS order). Written on a v19
    # destination only; every one of the ten is written unconditionally so an
    # unset limit is a blank cell rather than a leftover.
    if _dest_version >= 19:
        _sw = slope_data.get('search_window') or {}
        for _i, _key in enumerate(SEARCH_WINDOW_KEYS):
            _v = _sw.get(_key)
            circ[cell_ref(8 + _i, 11)] = _f(_v) if _v is not None else None
    if circ:
        updates['circles'] = circ

    # === non-circ ===  (data rows 3+; cols A=X, B=Y, C=Movement)
    nonc = {}
    for i, p in enumerate(slope_data.get('non_circ') or []):
        row = 3 + i
        nonc[cell_ref(row, 1)] = _f(p['X'])
        nonc[cell_ref(row, 2)] = _f(p['Y'])
        mv = p.get('Movement')
        if mv is not None and not (isinstance(mv, float) and pd.isna(mv)):
            nonc[cell_ref(row, 3)] = str(mv)
    if nonc:
        updates['non-circ'] = nonc

    # === dloads / dloads (2) ===  (3-col blocks from col B, +1 gap)
    # v21 adds a Direction cell in each block's N column (Excel row 3) and moves the
    # data down one row (4 -> 5); v22 adds the conditional note at the top of the sheet
    # and pushes the whole block down two more rows (Direction 5, data 7). 'normal' is
    # the default and is left blank so the sheet reads the way it always has;
    # 'vertical' is written explicitly.
    if _dest_version >= 22:
        _dl_data_row, _dl_dir_row = 7, 5
    elif _dest_version >= 21:
        _dl_data_row, _dl_dir_row = 5, 3
    else:
        _dl_data_row, _dl_dir_row = 4, None

    def _dload_updates(blocks, dirs):
        u = {}
        for n, block in enumerate(blocks):
            x_col = 2 + n * 4                                  # B, F, J, ...
            if _dl_dir_row is not None:
                _dir = str((dirs[n] if n < len(dirs) else 'normal') or 'normal').lower()
                if _dir not in ('normal', 'vertical'):
                    raise ValueError(
                        f"Unknown distributed-load direction {_dir!r}; expected "
                        "'normal' or 'vertical'.")
                u[cell_ref(_dl_dir_row, x_col + 2)] = None if _dir == 'normal' else _dir
            for i, pt in enumerate(block):
                u[cell_ref(_dl_data_row + i, x_col)] = _f(pt['X'])
                u[cell_ref(_dl_data_row + i, x_col + 1)] = _f(pt['Y'])
                u[cell_ref(_dl_data_row + i, x_col + 2)] = _f(pt['Normal'])
        return u
    d1 = _dload_updates(slope_data.get('dloads') or [],
                        slope_data.get('dload_dirs') or [])
    if d1:
        updates['dloads'] = d1
    d2 = _dload_updates(slope_data.get('dloads2') or [],
                        slope_data.get('dload2_dirs') or [])
    if d2:
        updates['dloads (2)'] = d2

    # === reinforce ===  (raw endpoint form in 'reinforcement_lines' round-trips)
    # Layout: # | Label | x1 y1 x2 y2 | Type Dir Appl | Tmax Lp1 Lp2 Adhesion
    # Delta Tend1 Tend2 Spacing | Tres E Area. Adhesion/Delta arrived at v24 and
    # shifted everything after Lp2, and this writer also fills ARCHIVED older
    # templates (the legacy round-trip fixtures), so each field's column is read
    # from the target template's own header row rather than hardcoded. A header
    # the target does not carry is skipped -- saving a model that uses the
    # overburden law into a pre-v24 template drops the two columns rather than
    # writing them into whatever sits at that position.
    # Capacity terms were divided by Spacing at load, so they are multiplied back
    # here -- the file carries per-element values. Adhesion and Delta are not:
    # they are interface properties, entered per unit area either way.
    # Dir/Appl carry in-sheet default formulas driven by Type: those cells are
    # written ONLY when the value differs from what the Type preset (or the
    # generic default) would produce, so the formulas survive a round-trip and
    # the preset behavior stays live for hand editing. Same table as the loader,
    # the sheet's lookup block and the Studio editor's fill.
    _REINF_PRESETS = REINFORCE_TYPE_PRESETS
    reinf = {}
    if slope_data.get('reinforcement_lines'):
        _reinf_hdr = pd.read_excel(template, sheet_name='reinforce', header=1,
                                   nrows=0)
        _rcol = {}
        for i, c in enumerate(_reinf_hdr.columns):
            name = str(c).strip().lower()
            if name and not name.startswith('unnamed'):
                _rcol.setdefault(name, i + 1)
    for n, r in enumerate(slope_data.get('reinforcement_lines') or []):
        row = 3 + n
        sp = float(r.get('spacing', 1.0) or 1.0)
        reinf[cell_ref(row, _rcol.get('#', 1))] = n + 1
        reinf[cell_ref(row, _rcol.get('label', 2))] = str(r.get('label', f"Line {n + 1}"))
        for hdr, val in (('x1', _f(r['x1'])), ('y1', _f(r['y1'])),
                         ('x2', _f(r['x2'])), ('y2', _f(r['y2'])),
                         ('tmax', _f(r['t_max']) * sp),
                         ('lp1', _f(r['lp1'])), ('lp2', _f(r['lp2'])),
                         ('tend1', _f(r.get('tend1', 0.0)) * sp),
                         ('tend2', _f(r.get('tend2', 0.0)) * sp),
                         ('spacing', sp),
                         # unset Tres round-trips as a BLANK cell, not a literal NaN
                         ('tres', None if _isnan(r.get('t_res'))
                          else _f(r.get('t_res', 0.0)) * sp),
                         ('e', _f(r['E'])), ('area', _f(r['area']) * sp)):
            col = _rcol.get(hdr)
            if col is not None:
                reinf[cell_ref(row, col)] = val
        # Blank Adhesion/Delta round-trip as blank cells: the constant-rate law
        # is what a blank pair MEANS, and a literal NaN in the sheet would read
        # back as a half-filled pair the preflight then refuses.
        for hdr, key in (('adhesion', 'adhesion'), ('delta', 'delta')):
            col = _rcol.get(hdr)
            if col is not None:
                reinf[cell_ref(row, col)] = (None if _isnan(r.get(key))
                                             else _f(r.get(key)))
        # Type is the INPUT; Dir and Appl are derived from it in the sheet by a
        # VLOOKUP against the type table. A row whose direction and application
        # are exactly what its Type derives is written as the Type alone, so the
        # formula stays in the cell and keeps answering after the next edit —
        # writing the preset back as a literal would replace a live default with
        # a frozen one. A row that OVERRIDES the preset has to state the
        # override, which is what typing over the cell means in the sheet; the
        # cell writer expands the column's shared-formula group first so the
        # rows below keep a formula (_expand_shared_formulas).
        #
        # Each column is written only where the target template actually carries
        # it. v11 has no Type/Dir/Appl at all, and the positional fallbacks these
        # lookups used to carry would have put the words into whatever that
        # template keeps in columns G, H and I.
        rtype = _reinf_word(r.get('type'), '')
        d_def, a_def = _REINF_PRESETS.get(rtype, REINFORCE_TYPE_DEFAULT)
        for hdr, value, derived in (
                ('type', rtype, ''),
                ('dir', _reinf_word(r.get('dir'), d_def), d_def),
                ('appl', _reinf_word(r.get('appl'), a_def), a_def)):
            col = _rcol.get(hdr)
            if col is not None and value and value != derived:
                reinf[cell_ref(row, col)] = value.capitalize()
    if reinf:
        updates['reinforce'] = reinf

    # === piles ===  (header row 2, data rows 3+. The piles layout changed at
    # v23 — the qp force-angle column was dropped, shifting everything after H
    # one column left — and again at v25, where Fixity became Head and a Tip
    # column was inserted after it; this writer also fills ARCHIVED older templates
    # (the legacy round-trip fixtures), so each field's column is read from the
    # target template's own header row rather than hardcoded. A header the
    # target does not carry is skipped.)
    piles_u = {}
    if slope_data.get('pile_lines'):
        _piles_hdr = pd.read_excel(template, sheet_name='piles', header=1,
                                   nrows=0)
        _pcol = {str(c).strip().lower(): i + 1
                 for i, c in enumerate(_piles_hdr.columns)}
        for n, p in enumerate(slope_data['pile_lines']):
            row = 3 + n
            piles_u[cell_ref(row, _pcol.get('#', 1))] = n + 1
            piles_u[cell_ref(row, _pcol.get('label', 2))] = \
                str(p.get('label', f"Pile {n + 1}"))
            for hdr, key in [('x1', 'x1'), ('y1', 'y1'), ('x2', 'x2'),
                             ('y2', 'y2')]:
                piles_u[cell_ref(row, _pcol[hdr])] = _f(p[key])
            for hdr, key in [('h', 'H'), ('d', 'D_pile'), ('s', 'S'),
                             ('vcap', 'V_cap'), ('mcap', 'M_cap'),
                             ('e', 'E'), ('i', 'I'), ('area', 'area')]:
                col = _pcol.get(hdr)
                val = p.get(key)
                if col is not None and val is not None:
                    piles_u[cell_ref(row, col)] = _f(val)
            # Appl is written explicitly both ways (like Head/Tip below): a
            # blank cell still LOADS as active, but a default a reader cannot
            # see in the sheet is a trap, so saved files spell the choice out.
            if 'appl' in _pcol:
                piles_u[cell_ref(row, _pcol['appl'])] = \
                    'Passive' if str(p.get('appl', 'active')) == 'passive' \
                    else 'Active'
            # Head fixity is 'Head' from v25 on and 'Fixity' before it; Tip is
            # v25-only. Each is written only where the target sheet carries the
            # column, so a v24 file round-trips through its own layout.
            _head = str(p.get('head_fixity', p.get('fixity', 'free')) or 'free')
            for hdr in ('head', 'fixity'):
                if hdr in _pcol:
                    piles_u[cell_ref(row, _pcol[hdr])] = _head
            if 'tip' in _pcol:
                piles_u[cell_ref(row, _pcol['tip'])] = \
                    str(p.get('tip_fixity', 'free') or 'free')
    if piles_u:
        updates['piles'] = piles_u

    # === lloads ===  (v12: # | Label | x | y | P | Angle; data rows 3+)
    lloads_u = {}
    for n, ll in enumerate(slope_data.get('line_loads') or []):
        row = 3 + n
        lloads_u[cell_ref(row, 1)] = n + 1
        lloads_u[cell_ref(row, 2)] = str(ll.get('label', f"Load {n + 1}"))
        lloads_u[cell_ref(row, 3)] = _f(ll['x'])
        lloads_u[cell_ref(row, 4)] = _f(ll['y'])
        lloads_u[cell_ref(row, 5)] = _f(ll['P'])
        lloads_u[cell_ref(row, 6)] = _f(ll.get('angle', -90.0))
    if lloads_u:
        updates['lloads'] = lloads_u

    # === seep bc / seep bc (2) ===
    def _seep_updates(bc):
        u = {}
        for i, (x, y) in enumerate(bc.get('exit_face') or []):
            u[cell_ref(5 + i, 2)] = _f(x)                     # B
            u[cell_ref(5 + i, 3)] = _f(y)                     # C
        # A head block's type cell is its kind ("head" or "reservoir"); default
        # "head" so a dict without an explicit kind writes the plain-Dirichlet type.
        blocks = [(b.get('kind', 'head'), b['head'], b['coords'])
                  for b in bc.get('specified_heads') or []]
        blocks += [('flux', b['flux'], b['coords'])
                   for b in bc.get('specified_fluxes') or []]
        for k, (kind, value, coords) in enumerate(blocks):
            x_col = 5 + k * 3                                 # E, H, K, ...
            y_col = x_col + 1
            u[cell_ref(3, x_col)] = kind                      # type cell (head/reservoir/flux)
            # v18: a string value is a tseep series name (time-varying BC) -- written
            # verbatim; a number is a constant head/flux.
            u[cell_ref(3, y_col)] = value if isinstance(value, str) else _f(value)
            for i, (x, y) in enumerate(coords):
                u[cell_ref(5 + i, x_col)] = _f(x)
                u[cell_ref(5 + i, y_col)] = _f(y)
        return u
    s1 = _seep_updates(slope_data.get('seepage_bc') or {})
    if s1:
        updates['seep bc'] = s1
    s2 = _seep_updates(slope_data.get('seepage_bc2') or {})
    if s2:
        updates['seep bc (2)'] = s2

    # === tseep (v18 transient seepage) ===
    # Written only when the model carries transient data AND the destination template
    # actually has the sheet (an older template has no place to put it). The layout
    # mirrors the parser (cell_map §3): time axis down column B, named series across
    # C.., controls in column J, save_times vertical under a column-J header whose row
    # depends on the DESTINATION version (J10 through v21, J11 from v22).
    tseep = slope_data.get('tseep')
    if tseep and 'tseep' in _dest_sheets:
        tu = {'B2': 'time'}
        times = tseep.get('times') or []
        for i, t in enumerate(times):
            tu[cell_ref(_TSEEP_DATA_ROW0 + 1 + i, _TSEEP_TIME_COL + 1)] = _f(t)
        for s, (name, vals) in enumerate(tseep.get('series', {}).items()):
            col = _TSEEP_SERIES_COL0 + 1 + s                  # 1-based C, D, ...
            tu[cell_ref(_TSEEP_HEADER_ROW + 1, col)] = str(name)
            for i, v in enumerate(vals):
                if not _isnan(v):
                    tu[cell_ref(_TSEEP_DATA_ROW0 + 1 + i, col)] = _f(v)
        # stability_time only has a cell from v22 on; writing it into an older
        # destination would land on the save_times header row.
        for key, row0 in _tseep_control_rows(_dest_version).items():
            val = tseep.get(key)
            if val is not None:
                tu[cell_ref(row0 + 1, _TSEEP_VAL_COL + 1)] = _f(val)
        _st_hdr, _st_row0 = _tseep_save_times_rows(_dest_version)
        tu[cell_ref(_st_hdr + 1, _TSEEP_VAL_COL + 1)] = 'save_times'
        for i, t in enumerate(tseep.get('save_times') or []):
            tu[cell_ref(_st_row0 + 1 + i, _TSEEP_VAL_COL + 1)] = _f(t)
        updates['tseep'] = tu

    updates = {k: v for k, v in updates.items() if v}
    write_cells_to_xlsx(filepath, updates)
    return filepath


def save_data_to_pickle(data, filepath):
    """
    Save a data object to a pickle file.
    
    This function serializes the data object and saves it to the specified filepath.
    Useful for saving processed data from Excel templates for later use.
    
    Parameters:
        data: The data object to save (typically a dictionary from load_slope_data)
        filepath (str): The file path where the pickle file should be saved
        
    Returns:
        None
        
    Raises:
        IOError: If the file cannot be written
        PickleError: If the data cannot be serialized
    """
    try:
        with open(filepath, 'wb') as f:
            pickle.dump(data, f)
    except Exception as e:
        raise IOError(f"Failed to save data to pickle file '{filepath}': {e}")


def load_data_from_pickle(filepath):
    """
    Load a data object from a pickle file.
    
    This function deserializes a data object from the specified pickle file.
    Useful for loading previously saved data without re-processing Excel templates.
    
    Parameters:
        filepath (str): The file path of the pickle file to load
        
    Returns:
        The deserialized data object (typically a dictionary)
        
    Raises:
        FileNotFoundError: If the pickle file doesn't exist
        IOError: If the file cannot be read
        PickleError: If the data cannot be deserialized
    """
    try:
        with open(filepath, 'rb') as f:
            data = pickle.load(f)
        return data
    except FileNotFoundError:
        raise FileNotFoundError(f"Pickle file not found: '{filepath}'")
    except Exception as e:
        raise IOError(f"Failed to load data from pickle file '{filepath}': {e}")
    

def print_dictionary(dictionary):
    """
    Print the contents of a dictionary to the console.
    This can be used for slope_data, seep_data, or any other dictionary.
    """
    for key, value in dictionary.items():
        print(f"\n=== {key} ===")
        if isinstance(value, list):
            for item in value:
                print(item)
        else:
            print(value)


# ---------------------------------------------------------------------------
# Surgical xlsx writer
# ---------------------------------------------------------------------------
# Writes individual cell values into an existing xlsx template by editing the
# sheet XML in place, rather than re-saving the whole workbook (which would
# discard formulas, formatting, and named ranges). Used to populate templates
# programmatically — e.g. from the polygon-sheet builders and the planned CAD
# importer.
#
# Two pitfalls this handles (do not "simplify" them away):
#   * Writing a value into a cell that some formula depends on invalidates the
#     cached calcChain. Excel then "recovers" the sheet and silently discards
#     the edits. We force fullCalcOnLoad and delete calcChain.xml so Excel
#     rebuilds it. (See the xlsx calcChain/recalc note in project memory.)
#   * Never write a plain value into a cell that itself holds a formula — that
#     breaks calcChain.xml. Callers must target precedent/value cells only. The
#     one place a value legitimately replaces a formula is a DERIVED column the
#     sheet lets a user type over (reinforce Dir/Appl); those columns are shared
#     formula groups, so _expand_shared_formulas gives the rest of the column its
#     own copy of the formula before the cell is overwritten.
#
# Requires the `zip` CLI (used to replace single entries inside the archive in
# place; Python's zipfile cannot update an existing member without rewriting
# the whole file).

_XLSX_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_XLSX_R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_XLSX_PKG_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'


def _col_num_to_letter(n):
    result = ''
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def cell_ref(row, col):
    """Return an A1-style cell reference for 1-based (row, col)."""
    return f'{_col_num_to_letter(col)}{row}'


def _parse_cell_ref(ref):
    m = re.match(r'^([A-Z]+)(\d+)$', ref)
    col_str, row = m.group(1), int(m.group(2))
    col = 0
    for c in col_str:
        col = col * 26 + (ord(c) - ord('A') + 1)
    return row, col


def _round_cell_float(value):
    """Strip binary-repr noise from a float on its way into a cell.

    The noise to remove is the tail past about ten meaningful digits — what turns
    ``0.1 + 0.2`` into ``0.30000000000000004``. Rounding to ten DECIMAL places
    does that at everyday magnitudes but measures from the decimal point rather
    than from the number, so it eats real digits as the value shrinks: it leaves a
    1e-9 m/s conductivity with two significant figures (8.944271910e-9 written as
    8.9e-9) and sends anything below 5e-11 to exactly zero, so a 1e-13 m/s
    low-permeability lens loads back as a hole in the model with no error anywhere.

    So the rule is ten SIGNIFICANT digits, never coarser than ten decimal places:
    the decimal count is measured from the value's own exponent, and floored at
    ten so nothing that used to survive is now rounded harder. At |v| >= 0.1 that
    is ten decimals exactly — unchanged. Below it the count grows with the
    exponent, which is what keeps a small conductivity intact.
    """
    if value == 0.0 or not math.isfinite(value):
        return value
    dp = 9 - math.floor(math.log10(abs(value)))   # ten significant digits
    return round(value, max(dp, 10))


def _modify_existing_cell(cell_xml, value):
    open_match = re.match(r'(<c\s[^>]*?)(/?>)', cell_xml)
    if not open_match:
        return cell_xml
    open_tag_attrs = re.sub(r'\s+t="[^"]*"', '', open_match.group(1))
    # Blank the cell for None / NaN / inf. A numeric cell holding "nan" in particular
    # makes openpyxl's reader raise int('nan') on the next load (silent corruption).
    if value is None or (isinstance(value, float) and not np.isfinite(value)):
        return f'{open_tag_attrs}/>'
    if isinstance(value, float):
        value = _round_cell_float(value)
    if isinstance(value, str):
        return (f'{open_tag_attrs} t="inlineStr"><is><t>'
                f'{xml_escape(value)}</t></is></c>')
    else:
        return f'{open_tag_attrs}><v>{value}</v></c>'


def _build_new_cell(ref, value):
    if value is None or (isinstance(value, float) and not np.isfinite(value)):
        return f'<c r="{ref}"/>'          # blank cell (None / NaN / inf)
    if isinstance(value, float):
        value = _round_cell_float(value)
    if isinstance(value, str):
        return f'<c r="{ref}" t="inlineStr"><is><t>{xml_escape(value)}</t></is></c>'
    else:
        return f'<c r="{ref}"><v>{value}</v></c>'


_SHARED_CELL_RE = re.compile(
    r'<c\s+r="([A-Z]+\d+)"(?:\s[^>]*)?>(?:(?!</c>).)*?<f\b[^>]*?\bt="shared"'
    r'[^>]*?(?:/>|>.*?</f>).*?</c>', re.DOTALL)
_SHARED_F_RE = re.compile(r'<f\b([^>]*?)(?:/>|>(.*?)</f>)', re.DOTALL)


def _shared_group_of(cell_xml):
    """(si, formula_text_or_None) for a shared-formula cell, else (None, None).

    The master of a group carries the formula text; every other member carries a
    bare ``<f t="shared" si="N"/>`` that points back at it.
    """
    m = _SHARED_F_RE.search(cell_xml)
    if not m or 't="shared"' not in m.group(1):
        return None, None
    si = re.search(r'\bsi="(\d+)"', m.group(1))
    if not si:
        return None, None
    text = m.group(2)
    return si.group(1), (text if text else None)


def _expand_shared_formulas(xml_bytes, refs):
    """Give every member of a touched shared-formula group its own formula.

    A shared group stores its formula ONCE, on the master cell; the other members
    hold a bare back-reference. Overwriting the master with a value therefore
    deletes the only copy of the formula and leaves the rest of the group
    pointing at nothing. Measured on the reinforce sheet, whose Dir and Appl
    columns are one shared group each (master at H3/I3, ``ref="H3:H22"``):
    saving a model whose Dir overrides the Type preset wrote a literal into H3
    and left Dir rows 9-22 reading back as a bare ``'='``. Excel handles an edit
    inside a shared group by expanding it first, and so does this: each member
    gets the master's formula translated to its own address, after which the
    cell about to be written is an ordinary cell.

    Only groups a written cell actually belongs to are expanded; every other
    sheet in the workbook is untouched.
    """
    from openpyxl.formula.translate import Translator

    xml_text = xml_bytes.decode('utf-8')
    if 't="shared"' not in xml_text:
        return xml_bytes
    cells = {}                       # ref -> (si, formula text or None)
    masters = {}                     # si -> (ref, formula text)
    for m in _SHARED_CELL_RE.finditer(xml_text):
        si, text = _shared_group_of(m.group(0))
        if si is None:
            continue
        cells[m.group(1)] = (si, text)
        if text:
            masters[si] = (m.group(1), text)
    touched = {cells[ref][0] for ref in refs if ref in cells}
    touched &= set(masters)
    if not touched:
        return xml_bytes

    def _expand(match):
        ref = match.group(1)
        entry = cells.get(ref)
        if entry is None or entry[0] not in touched:
            return match.group(0)
        origin, text = masters[entry[0]]
        formula = (text if ref == origin else
                   Translator('=' + text, origin=origin)
                   .translate_formula(ref)[1:])
        return _SHARED_F_RE.sub(lambda _m: f'<f>{formula}</f>',
                                match.group(0), count=1)

    return _SHARED_CELL_RE.sub(_expand, xml_text).encode('utf-8')


def _modify_sheet_xml(xml_bytes, cells):
    xml_text = xml_bytes.decode('utf-8')
    rows_data = {}
    for ref, value in cells.items():
        row_num, col_num = _parse_cell_ref(ref)
        rows_data.setdefault(row_num, []).append((ref, col_num, value))
    for row_num in rows_data:
        rows_data[row_num].sort(key=lambda x: x[1])
    for row_num, row_cells in sorted(rows_data.items()):
        row_pattern = re.compile(
            r'(<row\s+r="%d"[^>]*>)(.*?)(</row>)' % row_num, re.DOTALL)
        row_match = row_pattern.search(xml_text)
        if row_match:
            row_open = row_match.group(1)
            row_content = row_match.group(2)
            row_close = row_match.group(3)
            for ref, col_num, value in row_cells:
                cell_pattern = re.compile(
                    r'<c\s+r="%s"(?:\s[^>]*)*/>' % re.escape(ref) +
                    r'|<c\s+r="%s"(?:\s[^>]*)?>.*?</c>' % re.escape(ref), re.DOTALL)
                cell_match = cell_pattern.search(row_content)
                if cell_match:
                    new_cell = _modify_existing_cell(cell_match.group(0), value)
                    row_content = (row_content[:cell_match.start()] +
                                   new_cell + row_content[cell_match.end():])
                else:
                    row_content = row_content + _build_new_cell(ref, value)
            new_row = row_open + row_content + row_close
            xml_text = xml_text[:row_match.start()] + new_row + xml_text[row_match.end():]
        else:
            cells_xml = ''.join(_build_new_cell(ref, value) for ref, _, value in row_cells)
            new_row = f'<row r="{row_num}">{cells_xml}</row>'
            sd_close = xml_text.rfind('</sheetData>')
            xml_text = xml_text[:sd_close] + new_row + xml_text[sd_close:]
    return xml_text.encode('utf-8')


def _reset_view(xml_bytes):
    """Reset a sheet's saved scroll position/selection to A1 so populated cells are
    visible on open. The template ships some sheets (e.g. polygon) scrolled far right
    (topLeftCell="AA1", activeCell="AA8") — Excel then opens to an empty-looking region
    and the populated left-hand columns appear blank until you scroll back."""
    t = xml_bytes.decode('utf-8')
    t = re.sub(r'\s+topLeftCell="[^"]*"', '', t)
    t = re.sub(r'<selection\b[^>]*/>', '<selection activeCell="A1" sqref="A1"/>', t)
    return t.encode('utf-8')


def _force_full_recalc(xml_text):
    m = re.search(r'<calcPr\b[^>]*?/?>', xml_text)
    if m:
        tag = m.group(0)
        if 'fullCalcOnLoad=' in tag:
            new_tag = re.sub(r'fullCalcOnLoad="[^"]*"', 'fullCalcOnLoad="1"', tag)
        elif tag.endswith('/>'):
            new_tag = tag[:-2] + ' fullCalcOnLoad="1"/>'
        else:
            new_tag = tag[:-1] + ' fullCalcOnLoad="1">'
        return xml_text[:m.start()] + new_tag + xml_text[m.end():]
    insert = '<calcPr calcId="191029" fullCalcOnLoad="1"/>'
    idx = xml_text.find('</sheets>')
    pos = (idx + len('</sheets>')) if idx != -1 else xml_text.find('</workbook>')
    return xml_text[:pos] + insert + xml_text[pos:]


def _drop_calcchain(parts, zf_read):
    """Stage edits that remove xl/calcChain.xml and its two references.

    The cached calcChain becomes stale the moment we change a formula's precedent
    cell at the XML level. Excel then "recovers" the affected sheet and discards
    our edits (symptom: a populated sheet opens blank). Deleting calcChain.xml and
    its references makes Excel rebuild it from scratch — safe because we also set
    fullCalcOnLoad="1", so every formula is recomputed on open.

    The rewritten ``[Content_Types].xml`` and ``xl/_rels/workbook.xml.rels`` are
    added to ``parts`` (arcname -> bytes). Returns True if a calcChain part was
    present and must be dropped from the archive."""
    ct = zf_read('[Content_Types].xml').decode('utf-8')
    if 'calcChain' not in ct:
        return False
    ct = re.sub(r'<Override\s+PartName="/xl/calcChain\.xml"[^>]*/>', '', ct)
    parts['[Content_Types].xml'] = ct.encode('utf-8')

    rels = zf_read('xl/_rels/workbook.xml.rels').decode('utf-8')
    rels = re.sub(r'<Relationship\s+[^>]*Target="calcChain\.xml"[^>]*/>', '', rels)
    parts['xl/_rels/workbook.xml.rels'] = rels.encode('utf-8')
    return True


def _rewrite_zip(filepath, parts, drop=()):
    """Replace and/or delete members of the zip at ``filepath``, in place.

    ``parts`` maps arcname -> new bytes (replacing an existing member or adding a
    new one); ``drop`` names members to delete. Every other member is copied
    through byte-for-byte, keeping its stored compression method, timestamp and
    external attributes, so an .xlsx keeps all the formatting, drawings, charts
    and calc metadata this writer never looks at.

    Why in Python and not the ``zip`` command line: there is no ``zip`` executable
    on Windows. Shelling out to one raised ``FileNotFoundError`` — surfaced by the
    frozen app as "[WinError 2] The system cannot find the file specified" — so
    every Save failed there while working perfectly on macOS and Linux, where the
    tool happens to ship with the OS. A save path may not depend on a binary that
    is not part of the runtime we ship.

    The new archive is built beside the destination and swapped in with
    ``os.replace``, which is atomic on every platform: an interrupted save leaves
    the original file intact rather than a truncated one.
    """
    drop = set(drop)
    dest_dir = os.path.dirname(os.path.abspath(filepath)) or '.'
    fd, tmp_path = tempfile.mkstemp(prefix='.xslope-', suffix='.xlsx', dir=dest_dir)
    os.close(fd)
    try:
        with zipfile.ZipFile(filepath) as src:
            with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as out:
                out.comment = src.comment
                for info in src.infolist():
                    if info.filename in drop:
                        continue
                    data = (parts[info.filename] if info.filename in parts
                            else src.read(info.filename))
                    new_info = zipfile.ZipInfo(info.filename, date_time=info.date_time)
                    new_info.compress_type = info.compress_type
                    new_info.external_attr = info.external_attr
                    new_info.internal_attr = info.internal_attr
                    new_info.create_system = info.create_system
                    new_info.comment = info.comment
                    out.writestr(new_info, data)
                seen = {i.filename for i in src.infolist()}
                for name, data in parts.items():
                    if name not in seen:
                        out.writestr(name, data)
        os.replace(tmp_path, filepath)
    except BaseException:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise


def write_cells_to_xlsx(filepath, updates):
    """Write cell values into an existing xlsx template in place.

    Parameters
    ----------
    filepath : str
        Path to the xlsx file to modify (edited in place).
    updates : dict
        Mapping of {sheet_name: {cell_ref: value}}. Use cell_ref(row, col) to
        build the A1-style references. Values may be str, int, or float. Target
        only value/precedent cells — never cells that hold formulas.
    """
    with zipfile.ZipFile(filepath) as zf:
        wb_xml = etree.fromstring(zf.read('xl/workbook.xml'))
        rels_xml = etree.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        rid_map = {r.get('Id'): r.get('Target')
                   for r in rels_xml.iter('{%s}Relationship' % _XLSX_PKG_NS)}
        sheet_paths = {}
        for s in wb_xml.iter('{%s}sheet' % _XLSX_NS):
            rid = s.get('{%s}id' % _XLSX_R_NS)
            if rid and rid in rid_map:
                _t = rid_map[rid]
                sheet_paths[s.get('name')] = _t.lstrip('/') if _t.startswith('/') else f'xl/{_t}'
    parts = {}
    with zipfile.ZipFile(filepath) as zf:
        for sheet_name, cells in updates.items():
            path = sheet_paths[sheet_name]
            parts[path] = _reset_view(_modify_sheet_xml(
                _expand_shared_formulas(zf.read(path), cells), cells))
        wb_text = zf.read('xl/workbook.xml').decode('utf-8')
        parts['xl/workbook.xml'] = _force_full_recalc(wb_text).encode('utf-8')
        drop_cc = _drop_calcchain(parts, zf.read)
    _rewrite_zip(filepath, parts,
                 drop=('xl/calcChain.xml',) if drop_cc else ())